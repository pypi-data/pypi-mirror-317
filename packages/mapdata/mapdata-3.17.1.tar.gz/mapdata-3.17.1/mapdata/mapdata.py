#!/usr/bin/python
#
# mapdata.py
#
# PURPOSE
#	Display a simple interactive map of data points, allowing points to
#	be highlighted by clicking on the map or table or by querying,
#	and allowing some simple data plots.
#
# COPYRIGHT AND LICENSE
#	Copyright (c) 2023-2024, R. Dreas Nielsen
# 	This program is free software: you can redistribute it and/or modify
# 	it under the terms of the GNU General Public License as published by
# 	the Free Software Foundation, either version 3 of the License, or
# 	(at your option) any later version.
# 	This program is distributed in the hope that it will be useful,
# 	but WITHOUT ANY WARRANTY; without even the implied warranty of
# 	MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# 	GNU General Public License for more details.
# 	The GNU General Public License is available at <http://www.gnu.org/licenses/>
#
# AUTHOR
#	Dreas Nielsen (RDN)
#
# ==================================================================

version = "3.17.1"
vdate = "2024-12-28"

copyright = "2023-2024"


import sys
import os.path
import io
import codecs
import argparse
from configparser import ConfigParser
import csv
import re
import datetime
import time
import math
import statistics
import collections
import webbrowser
import threading
import queue
import sqlite3
import tempfile
import random
import uuid
import traceback
import subprocess
import multiprocessing
from functools import reduce
from operator import add
import copy
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.font as tkfont
import tkinter.filedialog as tkfiledialog
import tkintermapview as tkmv
import textwrap
import requests
import PIL
from PIL import Image, ImageGrab, ImageEnhance, ImageTk
import odf.opendocument
import odf.table
import odf.text
import odf.number
import odf.style
import xlrd
import openpyxl
import numpy as np
from pyproj import CRS, Transformer
from numpy.polynomial import Polynomial
import matplotlib
import matplotlib.collections as mpl_coll
import matplotlib.path as mpl_path
matplotlib.use('TkAgg')
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import seaborn as sns
import jenkspy
from loess.loess_1d import loess_1d
import scipy.stats as spstats
import scipy.sparse as sparse
import statsmodels.api as sm
from statsmodels.sandbox.stats.runs import runstest_1samp
from statsmodels.stats.diagnostic import normal_ad, lilliefors, kstest_normal, het_breuschpagan
from sklearn.preprocessing import StandardScaler, Normalizer
from sklearn.metrics import mean_squared_error
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.manifold import TSNE
from sklearn.decomposition import NMF, PCA
try:
	import pymannkendall as mk
	mk_available = True
except:
	mk_available = False


# Default name of configuration file.  Files with other names may be read.
config_file_name = "mapdata.conf"

# Configuration files read on startup
config_files = []
# Configuration file read post-startup
config_files_user = []


# Default options
multiselect = "0"

# User-modifiable map settings
class MapSettings(object):
	def __init__(self):
		#-- Default location marker.  This may be overridden
		self.location_marker = "triangle_open"
		self.location_color = "black"
		self.use_data_marker = True
		self.use_data_color = True
		# Column names in the data table--initially not set
		self.label_col = None		# Column name for location labels
		self.symbol_col = None		# Column with symbol names
		self.color_col = None		# Column with color names
		#-- Selected item marker
		self.select_symbol = "wedge"
		self.select_color = "red"
		#-- Location marker label appearance
		self.label_font = "Liberation Sans"
		self.label_color = "black"
		self.label_bold = False
		self.label_size = 10
		self.label_position = "below"	# above or below
		#-- Count label appearance
		self.count_label_font = "Liberation Sans"
		self.count_label_color = "black"
		self.count_label_size = 10
		self.count_label_bold = True
		self.count_label_offset = 12		# pixels
		self.count_label_show = False
		#-- Basemap color adjustment, triplet of reals, only set by configuration file.
		self.basemap_color_adj = None

map_settings = MapSettings()

# Plot configuration settings
show_regression_stats = False
wrapwidth = 20
wrap_at_underscores = False

# Custom matplotlib color map
c1 = matplotlib.colormaps['tab20'].colors
quant_colors = [c1[0], c1[2], c1[4], c1[8], c1[10], c1[6], c1[14], c1[16], c1[18], c1[12], \
		c1[1], c1[3], c1[5], c1[7], c1[9], c1[11], c1[13], c1[15], c1[17], c1[19]]

# Name of editor, read from environment if it exists, may be set by config file.
editor = os.environ.get("EDITOR")

# Lists of warning and error messages from reading configuration files,
# to be displayed after the MapUI interface creates a Tkinter window.
config_warning_messages = []
config_fatal_messages = []

#-- Operational configuration
# Whether to use a temporary file for Sqlite (as opposed to memory).
temp_dbfile = False

# Translations to SQLite type affinity names
sqlite_type_x = {'int': 'INTEGER', 'float': 'REAL', 'string': 'TEXT', 'timestamptz': 'TEXT',
		         'timestamp': 'TEXT', 'date': 'TEXT', 'boolean': 'INTEGER', None: 'TEXT'}


#=====  Software patches

# Patch the tkintermapview CanvasPositionMarker 'calculate_text_y_offset()' function to
# allow labeling below the icon.  The icon anchor position is always "center" for this app.
def new_calc_text_offset(self):
	if self.icon is not None:
		if map_settings.label_position == "below":
			self.text_y_offset = self.icon.height()/2 + 6 + map_settings.label_size
		else:
			self.text_y_offset = -self.icon.height()/2 - 3
	else:
			self.text_y_offset = -56
tkmv.canvas_position_marker.CanvasPositionMarker.calculate_text_y_offset = new_calc_text_offset


# Patch the tkintermapview CanvasPositionMarker 'draw()' function to create a count
# label if the marker's data is set.
original_marker_draw = tkmv.canvas_position_marker.CanvasPositionMarker.draw
def count_label_draw(self, event=None):
	# Non-null .data means that this is a special marker for each unique X,Y, not directly associated with a data table row.
	if self.data is not None:
		if self.data[0] and self.data[1] > 1:
			lblfont = makefont(map_settings.count_label_font, map_settings.count_label_size, map_settings.count_label_bold)
			canvas_x, canvas_y = self.get_canvas_pos(self.position)
			self.canvas_text = self.map_widget.canvas.create_text(canvas_x + map_settings.count_label_offset, canvas_y, anchor=tk.W,
					text=str(self.data[1]), font=lblfont, fill=map_settings.count_label_color)
		else:
			if self.canvas_text is not None:
				self.map_widget.canvas.delete(self.canvas_text)
		self.map_widget.manage_z_order()
	if self.canvas_icon is not None and self.right_click_function is not None:
		self.map_widget.canvas.tag_bind(self.canvas_icon, "<Button-2>", self.right_click_function)
		self.map_widget.canvas.tag_bind(self.canvas_icon, "<Button-3>", self.right_click_function)
def mod_marker_draw(self, event=None):
	original_marker_draw(self, event)
	count_label_draw(self)
tkmv.canvas_position_marker.CanvasPositionMarker.draw = mod_marker_draw

# Patch function for ImageTk.PhotoImage.__del__ 
def new_img_del(img_self):
	try:
		name = img_self.__photo.name
		img_self.__photo.name = None
		img_self.__photo.tk.call("image", "delete", name)
	except Exception:
		pass


# Patch tkintermapview CanvasPositionMarker to add methods that will be called by 
# a right-click on a marker.
def mkr_right_click_method(self, event):
	# This marker method calls a function in mapdata.
	show_location_table(event, self)

tkmv.canvas_position_marker.CanvasPositionMarker.right_click_function = mkr_right_click_method

def show_location_table(event, marker):
	# This is a stub that is global because it is referenced in the patch above.
	# MapUI.__init__ replaces this with the class method that uses MapUI instance data.
	pass


# Patch tkintermapview request_image to include an image adjustment step.
def alt_request_image(self, zoom: int, x: int, y: int, db_cursor=None) -> ImageTk.PhotoImage:
    # if database is available check first if tile is in database, if not try to use server
	if db_cursor is not None:
		try:
			db_cursor.execute("SELECT t.tile_image FROM tiles t WHERE t.zoom=? AND t.x=? AND t.y=? AND t.server=?;",
								(zoom, x, y, self.tile_server))
			result = db_cursor.fetchone()

			if result is not None:
				image = Image.open(io.BytesIO(result[0]))
				image_tk = ImageTk.PhotoImage(image)
				self.tile_image_cache[f"{zoom}{x}{y}"] = image_tk
				return image_tk
			elif self.use_database_only:
				return self.empty_tile_image
			else:
				pass

		except sqlite3.OperationalError:
			if self.use_database_only:
				return self.empty_tile_image
			else:
				pass

		except Exception:
			return self.empty_tile_image

    # try to get the tile from the server
	try:
		url = self.tile_server.replace("{x}", str(x)).replace("{y}", str(y)).replace("{z}", str(zoom))
		image = Image.open(requests.get(url, stream=True, headers={"User-Agent": "TkinterMapView"}).raw)
		# Adjust color only for newly downloaded images.
		image = adjust_tile_color(image)

		if self.overlay_tile_server is not None:
			url = self.overlay_tile_server.replace("{x}", str(x)).replace("{y}", str(y)).replace("{z}", str(zoom))
			image_overlay = Image.open(requests.get(url, stream=True, headers={"User-Agent": "TkinterMapView"}).raw)
			image = image.convert("RGBA")
			image_overlay = image_overlay.convert("RGBA")

			if image_overlay.size is not (self.tile_size, self.tile_size):
				image_overlay = image_overlay.resize((self.tile_size, self.tile_size), Image.ANTIALIAS)

			image.paste(image_overlay, (0, 0), image_overlay)

		if self.running:
			image_tk = ImageTk.PhotoImage(image)
		else:
			return self.empty_tile_image

		self.tile_image_cache[f"{zoom}{x}{y}"] = image_tk
		return image_tk

	except PIL.UnidentifiedImageError:  # image does not exist for given coordinates
		self.tile_image_cache[f"{zoom}{x}{y}"] = self.empty_tile_image
		return self.empty_tile_image

	except requests.exceptions.ConnectionError:
		return self.empty_tile_image

	except Exception:
		return self.empty_tile_image

tkmv.map_widget.TkinterMapView.request_image = alt_request_image


#=====  Global constants and variables =====

# SQL (SQLite) keywords
sql_kw = ('add', 'all', 'alter', 'analyze', 'and', 'as', 'asc', 'autoincrement', 'before', 'begin',
			'between', 'by', 'cascade', 'case', 'cast', 'check', 'column', 'commit', 'conflict', 'constraint',
			'create', 'cross', 'current', 'current_date', 'current_time', 'current_timestamp', 'database',
			'default', 'delete', 'desc', 'distinct', 'do', 'drop', 'each', 'else', 'end', 'except', 'exclude',
			'exclusive', 'exists', 'explain', 'filter', 'first', 'for', 'foreign', 'from', 'full', 'glob',
			'group', 'groups', 'having', 'if', 'immediate', 'in', 'index', 'indexed', 'initially', 'inner',
			'insert', 'instead', 'intersect', 'into', 'is', 'isnull', 'join', 'key', 'last', 'left', 'like',
			'limit', 'match', 'materialized', 'natural', 'no', 'not', 'nothing', 'notnull', 'null', 'nulls',
			'of', 'offset', 'on', 'or', 'order', 'others', 'outer', 'over', 'partition', 'plan', 'pragma',
			'preceding', 'primary', 'query', 'raise', 'range', 'recursive', 'references', 'reindex', 'release',
			'rename', 'replace', 'restrict', 'returning', 'right', 'rollback', 'row', 'rows', 'savepoint',
			'select', 'set', 'table', 'temp', 'temporary', 'then', 'ties', 'to', 'transaction', 'trigger',
			'unbounded', 'union', 'unique', 'update', 'using', 'vacuum', 'values', 'view', 'virtual',
			'when', 'where', 'window', 'with', 'without')


# Tile servers for map basemap layers
bm_servers = {"OpenStreetMap": "https://a.tile.openstreetmap.org/{z}/{x}/{y}.png",
			"Google streets": "https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga",
			"Google satellite": "https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga",
			"Open topo map": "https://tile.opentopomap.org/{z}/{x}/{y}.png"
			}

map_attributions = {"OpenStreetMap": "Basemap data © OpenStreetMap and contributors <https://www.openstreetmap.org/>",
		"Google streets": "Basemap data © Google",
		"Google satellite": "Basemap data © Google and contributors",
		"Open topo map": "Basemap data © OpenStreetMap and contributors <https://www.openstreetmap.org/>, SRTM | map style: © OpenTopoMap (CC-BY-SA) <https://opentopomap.org/>"
		}

# Triplets for adjustment of saturation, contrast, and brightness.  The dictionary keys
# should match basemap names.  This dictionary is only populated from configuration files.
map_color_adj = {}

# API keys for tile servers that require them.  The dictionary keys should match basemap names.
api_keys = {}

# Initial basemap to use
#initial_basemap = tuple(bm_servers.keys())[0]
initial_basemap = "OpenStreetMap"

# List of initial basemap names, for use when saving config
initial_bms = list(bm_servers.keys())

# Database codes used in command-line arguments and corresponding names.
dbms_name_codes = {"p": "PostgreSQL", "l": "SQLite", "k": "DuckDB", "m": "MariaDB", \
		"s": "SQL Server", "o": "Oracle", "f": "Firebird"}

# X11 bitmaps for map icons
icon_xbm = {
	'anchor': """#define anchor_width 16
#define anchor_height 16
static unsigned char anchor_bits[] = {
   0xc0, 0x03, 0x60, 0x06, 0x60, 0x06, 0xc0, 0x03, 0xfc, 0x3f, 0xfc, 0x3f,
   0x80, 0x01, 0x81, 0x81, 0x83, 0xc1, 0x87, 0xe1, 0x83, 0xc1, 0x87, 0xe1,
   0x8e, 0x71, 0xfc, 0x3f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'ball': """#define ball_width 16
#define ball_height 16
static unsigned char circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0xf8, 0x1f, 0xfc, 0x3f, 0xfe, 0x7f, 0xfe, 0x7f,
   0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xfe, 0x7f, 0xfe, 0x7f,
   0xfc, 0x3f, 0xf8, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

   'binoculars': """#define binoculars_width 16
#define binoculars_height 16
static unsigned char binoculars_bits[] = {
   0x38, 0x1c, 0x38, 0x1c, 0x7c, 0x3e, 0x7c, 0x3e, 0xfc, 0x3f, 0xbc, 0x3d,
   0xbc, 0x3d, 0xfe, 0x7f, 0xfe, 0x7f, 0x7f, 0xfe, 0x7f, 0xfe, 0x3f, 0xfc,
   0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc};""",

	'bird': """#define bird_width 16
#define bird_height 16
static unsigned char bird.xbm_bits[] = {
   0x00, 0x00, 0x00, 0x1c, 0x00, 0x3f, 0x80, 0xef, 0xc0, 0x7f, 0xe0, 0x3f,
   0xf0, 0x3f, 0xf8, 0x1f, 0xff, 0x1f, 0xfc, 0x0f, 0xe0, 0x07, 0x80, 0x01,
   0x00, 0x01, 0x00, 0x01, 0x80, 0x03, 0xe0, 0x0f};""",

	'block': """#define block_width 16
#define block_height 16
static unsigned char square_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfc, 0x3f, 0xfc, 0x3f, 0x00, 0x00, 0x00, 0x00};""",

	'bookmark': """#define bookmark_width 16
#define bookmark_height 16
static unsigned char bookmark_bits[] = {
   0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f,
   0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0x7e, 0x7e, 0x3e, 0x7c,
   0x1e, 0x78, 0x0e, 0x70, 0x06, 0x60, 0x02, 0x40};""",

   'camera': """#define camera.xbm_width 16
#define camera.xbm_height 16
static unsigned char camera.xbm_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0xf0, 0x07, 0xf6, 0x07, 0xfe, 0x7f, 0xff, 0xff,
   0x3f, 0xfe, 0x1f, 0xfc, 0xcf, 0xf9, 0xcf, 0xf9, 0xcf, 0xf9, 0x1f, 0xfc,
   0x3f, 0xfe, 0xff, 0xff, 0xfe, 0x7f, 0x00, 0x00};""",

   'cancel': """#define cancel_width 16
#define cancel_height 16
static unsigned char cancel_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x7c, 0x06, 0x6e,
   0x07, 0xe7, 0x83, 0xc3, 0xc3, 0xc1, 0xe7, 0xe0, 0x76, 0x60, 0x3e, 0x78,
   0x1c, 0x38, 0x78, 0x3e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'car': """#define car_width 16
#define car_height 16
static unsigned char car_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0xf0, 0x00, 0xf8, 0x01, 0x2c, 0x03, 0x2c, 0x06,
   0x26, 0x0c, 0x26, 0x7c, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xc1, 0x83,
   0xdd, 0xbb, 0xdc, 0x3b, 0x1c, 0x38, 0x00, 0x00};""",

	'center8': """#define center8_width 16
#define center8_height 16
static unsigned char center8_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0xf0, 0x0f, 0xf0, 0x0f,
   0xf0, 0x0f, 0xf0, 0x0f, 0xf0, 0x0f, 0xf0, 0x0f, 0xf0, 0x0f, 0xf0, 0x0f,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'checkbox': """#define checkbox_width 16
#define checkbox_height 16
static unsigned char checkbox_bits[] = {
   0xff, 0xff, 0xff, 0xff, 0x03, 0xc0, 0x03, 0xd8, 0x03, 0xd8, 0x03, 0xdc,
   0x03, 0xcc, 0x1b, 0xce, 0x3b, 0xc6, 0x73, 0xc7, 0xe3, 0xc3, 0xc3, 0xc3,
   0x83, 0xc1, 0x03, 0xc0, 0xff, 0xff, 0xff, 0xff};""",

	'circle': """#define circle_width 16
#define circle_height 16
static unsigned char circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60,
   0x07, 0xe0, 0x03, 0xc0, 0x03, 0xc0, 0x07, 0xe0, 0x06, 0x60, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'circle_bar': """#define circle_bar_width 16
#define circle_bar_height 16
static unsigned char circle_bar_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60,
   0x07, 0xe0, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0x06, 0x60, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'circle_plus': """#define circle_plus_width 16
#define circle_plus_height 16
static unsigned char circle_plus_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0xf8, 0x1f, 0x9c, 0x39, 0x8e, 0x71, 0x86, 0x61,
   0x87, 0xe1, 0xff, 0xff, 0xff, 0xff, 0x87, 0xe1, 0x86, 0x61, 0x8e, 0x71,
   0x9c, 0x39, 0xf8, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'circle_x': """#define circle_x_width 16
#define circle_x_height 16
static unsigned char circle_x_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x3e, 0x7c, 0x76, 0x6e,
   0xe7, 0xe7, 0xc3, 0xc3, 0xc3, 0xc3, 0xe7, 0xe7, 0x76, 0x6e, 0x3e, 0x7c,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'clock': """#define clock_width 16
#define clock_height 16
static unsigned char clock_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0xf8, 0x1e, 0xfc, 0x3e, 0xfe, 0x7e, 0xfe, 0x7e,
   0xff, 0xfe, 0xff, 0xfe, 0x07, 0xfe, 0xff, 0xff, 0xfe, 0x7f, 0xfe, 0x7f,
   0xfc, 0x3f, 0xf8, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'darkeye': """#define darkeye_width 16
#define darkeye_height 16
static unsigned char darkeye_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0xe0, 0x07, 0xf8, 0x1f, 0xfc, 0x3f, 0x7e, 0x7e,
   0x3e, 0x7c, 0x1f, 0xf8, 0x1f, 0xf8, 0x3e, 0x7c, 0x7e, 0x7e, 0xfc, 0x3f,
   0xf8, 0x1f, 0xe0, 0x07, 0x00, 0x00, 0x00, 0x00};""",

	'deposition': """
#define deposition_width 16
#define deposition_height 16
static unsigned char deposition_bits[] = {
   0x55, 0xab, 0x00, 0x00, 0x6b, 0xdd, 0x00, 0x00, 0xf7, 0xee, 0x00, 0x00,
   0xbb, 0xbb, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff,
   0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff};""",

	'diag_ll': """#define diag_ll_width 16
#define diag_ll_height 16
static unsigned char diag_ll_bits[] = {
   0x01, 0x00, 0x03, 0x00, 0x07, 0x00, 0x0f, 0x00, 0x1f, 0x00, 0x3f, 0x00,
   0x7f, 0x00, 0xff, 0x00, 0xff, 0x01, 0xff, 0x03, 0xff, 0x07, 0xff, 0x0f,
   0xff, 0x1f, 0xff, 0x3f, 0xff, 0x7f, 0xff, 0xff};""",

	'diag_lr': """#define diag_lr_width 16
#define diag_lr_height 16
static unsigned char diag_lr_bits[] = {
   0x00, 0x80, 0x00, 0xc0, 0x00, 0xe0, 0x00, 0xf0, 0x00, 0xf8, 0x00, 0xfc,
   0x00, 0xfe, 0x00, 0xff, 0x80, 0xff, 0xc0, 0xff, 0xe0, 0xff, 0xf0, 0xff,
   0xf8, 0xff, 0xfc, 0xff, 0xfe, 0xff, 0xff, 0xff};""",

	'diag_ul': """#define diag_ul_width 16
#define diag_ul_height 16
static unsigned char diag_ul_bits[] = {
   0xff, 0x7f, 0xff, 0x3f, 0xff, 0x1f, 0xff, 0x0f, 0xff, 0x07, 0xff, 0x03,
   0xff, 0x01, 0xff, 0x00, 0x7f, 0x00, 0x3f, 0x00, 0x1f, 0x00, 0x0f, 0x00,
   0x07, 0x00, 0x03, 0x00, 0x01, 0x00, 0x00, 0x00};""",

	'diag_ur': """#define diag_ur_width 16
#define diag_ur_height 16
static unsigned char diag_ur_bits[] = {
   0xfe, 0xff, 0xfc, 0xff, 0xf8, 0xff, 0xf0, 0xff, 0xe0, 0xff, 0xc0, 0xff,
   0x80, 0xff, 0x00, 0xff, 0x00, 0xfe, 0x00, 0xfc, 0x00, 0xf8, 0x00, 0xf0,
   0x00, 0xe0, 0x00, 0xc0, 0x00, 0x80, 0x00, 0x00};""",

	'diamond': """#define diamond_width 16
#define diamond_height 16
static unsigned char diamond_bits[] = {
   0x80, 0x01, 0xc0, 0x03, 0xe0, 0x07, 0xf0, 0x0f, 0xf8, 0x1f, 0xfc, 0x3f,
   0xfe, 0x7f, 0xff, 0xff, 0xff, 0xff, 0xfe, 0x7f, 0xfc, 0x3f, 0xf8, 0x1f,
   0xf0, 0x0f, 0xe0, 0x07, 0xc0, 0x03, 0x80, 0x01};""",

	'donkey': """#define donkey.xbm_width 16
#define donkey.xbm_height 16
static unsigned char donkey.xbm_bits[] = {
   0x00, 0x00, 0x0c, 0x00, 0x08, 0x00, 0x1b, 0x00, 0x7e, 0x3c, 0xfe, 0x7f,
   0xfa, 0xff, 0xff, 0xbf, 0xff, 0x9f, 0xef, 0x9f, 0xe6, 0xbf, 0x60, 0x3c,
   0x70, 0x6c, 0x30, 0x68, 0x30, 0x48, 0x30, 0x6c};""",

   'drop': """,
#define drop.xbm_width 16
#define drop.xbm_height 16
static unsigned char drop.xbm_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0xe0, 0x07,
   0xf8, 0x1f, 0xf8, 0x1f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfc, 0x3f, 0xf0, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'elephant': """#define elephant_width 16
#define elephant_height 16
static unsigned char elephant_bits[] = {
   0x00, 0x1c, 0x00, 0x1f, 0xfe, 0x3f, 0xfe, 0x7f, 0xff, 0xff, 0xfd, 0xef,
   0xfd, 0xff, 0xfd, 0xff, 0xfd, 0xff, 0xfd, 0xf7, 0xfc, 0xc7, 0x3c, 0x83,
   0x1c, 0xa3, 0x1c, 0xb3, 0x9e, 0xf3, 0xbe, 0x67};""",

	'eye': """#define eye_width 16
#define eye_height 16
static unsigned char eye_bits[] = {
   0x00, 0x00, 0x40, 0x02, 0x44, 0x22, 0x09, 0x90, 0xe2, 0x47, 0x38, 0x1c,
   0x8c, 0x31, 0xc6, 0x63, 0xc7, 0xe3, 0x86, 0x61, 0x0e, 0x70, 0x3c, 0x3c,
   0xf8, 0x1f, 0xe0, 0x07, 0x00, 0x00, 0x00, 0x00};""",

   'fish': """#define fish_width 16
#define fish_height 16
static unsigned char fish.xbm_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x01, 0x00, 0x03, 0x01, 0x0f,
   0xe3, 0x7f, 0xf7, 0x78, 0x3e, 0xea, 0x9e, 0xfb, 0x73, 0x84, 0xc1, 0x63,
   0x01, 0x3e, 0x00, 0x18, 0x00, 0x18, 0x00, 0x08};""",

	'flag': """#define flag_width 16
#define flag_height 16
static unsigned char flag.xbm_bits[] = {
   0x00, 0x00, 0x0e, 0x00, 0x3e, 0x00, 0xfe, 0x01, 0xfe, 0x1f, 0xfe, 0xff,
   0xfe, 0xff, 0xfe, 0xff, 0xfe, 0xff, 0xfe, 0xff, 0xfe, 0xff, 0xf6, 0xff,
   0x86, 0xff, 0x06, 0xf8, 0x06, 0x00, 0x06, 0x00};""",

	'hand': """#define hand_width 16
#define hand_height 16
static unsigned char hand_bits[] = {
   0xc0, 0x00, 0xd8, 0x06, 0xd8, 0x06, 0xdb, 0x06, 0xdb, 0x06, 0xdb, 0x06,
   0xdb, 0x06, 0xdb, 0xc6, 0xff, 0xe7, 0xff, 0xf7, 0xff, 0x7f, 0xff, 0x3f,
   0xff, 0x3f, 0xff, 0x1f, 0xff, 0x0f, 0xfe, 0x0f};""",

	'heart': """#define heart_width 16
#define heart_height 16
static unsigned char heart_bits[] = {
   0x3c, 0x3c, 0x7e, 0x7e, 0x7f, 0xfe, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff,
   0xff, 0xff, 0xff, 0xff, 0xfe, 0x7f, 0xfe, 0x7f, 0xfc, 0x3f, 0xf8, 0x1f,
   0xf0, 0x0f, 0xe0, 0x07, 0xc0, 0x03, 0x80, 0x01};""",

	'hidden': """#define hidden_width 2
#define hidden_height 2
static unsigned char hidden_bits[] = {
   0x00, 0x00};""",

	'hourglass': """#define hourglass_width 16
#define hourglass_height 16
static unsigned char hourglass_bits[] = {
   0xff, 0xff, 0xff, 0xff, 0x0c, 0x30, 0x0c, 0x30, 0x18, 0x18, 0xf0, 0x0f,
   0xe0, 0x07, 0xc0, 0x03, 0xc0, 0x03, 0x60, 0x06, 0x30, 0x0c, 0x98, 0x19,
   0xcc, 0x33, 0xec, 0x37, 0xff, 0xff, 0xff, 0xff};""",

	'house': """#define house_width 16
#define house_height 16
static unsigned char house_bits[] = {
   0x80, 0x01, 0xc0, 0x33, 0x60, 0x36, 0xb0, 0x3d, 0xd8, 0x3b, 0xec, 0x37,
   0xf6, 0x6f, 0xfb, 0xdf, 0xfd, 0xbf, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f,
   0x7c, 0x3e, 0x7c, 0x3e, 0x7c, 0x3e, 0x7c, 0x3e};""",

	'info': """#define info_width 16
#define info_height 16
static unsigned char info_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x7c, 0x3e, 0xfe, 0x7f, 0xfe, 0x7f,
   0x3f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7f, 0xfe, 0x7e, 0x7e, 0x7e, 0x7e,
   0x3c, 0x3c, 0xf8, 0x1f, 0xf0, 0x0f, 0xc0, 0x03};""",

	'leaf': """#define leaf_width 16
#define leaf_height 16
static unsigned char leaf_bits[] = {
   0x00, 0x00, 0xc0, 0xff, 0xf0, 0xfb, 0xf8, 0xdb, 0xdc, 0xcb, 0xdc, 0xe3,
   0xdc, 0xf9, 0xdc, 0xfc, 0x5c, 0xc0, 0x1c, 0xff, 0x9c, 0x7f, 0xcc, 0x7f,
   0xec, 0x3f, 0xe6, 0x1f, 0x03, 0x00, 0x01, 0x00};""",

	'lightning': """#define lightning_width 16
#define lightning_height 16
static unsigned char Lightning_bits[] = {
   0x00, 0xc0, 0x00, 0x70, 0x00, 0x1c, 0x00, 0x07, 0x80, 0x03, 0xe0, 0x01,
   0xf0, 0x00, 0xf8, 0x03, 0xc0, 0x3f, 0x00, 0x1f, 0x80, 0x07, 0xc0, 0x01,
   0xe0, 0x00, 0x38, 0x00, 0x0e, 0x00, 0x03, 0x00};""",

	'mine': """#define mine_width 16
#define mine_height 16
static unsigned char mine_bits[] = {
   0xe0, 0xf1, 0x70, 0xf8, 0x3c, 0xfc, 0x1c, 0xf8, 0x3e, 0x7c, 0x77, 0x2e,
   0xe3, 0x07, 0xc1, 0x03, 0xc1, 0x03, 0xe0, 0x07, 0x70, 0x0e, 0x38, 0x1c,
   0x1c, 0x38, 0x0e, 0x70, 0x07, 0xe0, 0x03, 0xc0};""",

	'pennant': """#define pennant2_width 16
#define pennant2_height 16
static unsigned char pennant2_bits[] = {
   0x0e, 0x00, 0x3e, 0x00, 0xfe, 0x00, 0xfe, 0x03, 0xfe, 0x0f, 0xfe, 0x3f,
   0xfe, 0xff, 0xfe, 0x3f, 0xfe, 0x07, 0xfe, 0x00, 0x1e, 0x00, 0x06, 0x00,
   0x06, 0x00, 0x06, 0x00, 0x06, 0x00, 0x06, 0x00};""",

	'photo': """#define photo_width 16
#define photo_height 16
static unsigned char photo_bits[] = {
   0xff, 0xff, 0x01, 0x80, 0x01, 0x80, 0x01, 0x98, 0x01, 0x98, 0x01, 0x80,
   0x11, 0x80, 0x39, 0x80, 0x7d, 0x84, 0xfd, 0x8e, 0xfd, 0x9f, 0xfd, 0xbf,
   0xfd, 0xbf, 0xfd, 0xbf, 0x01, 0x80, 0xff, 0xff};""",

	'picnic': """#define picnic_width 16
#define picnic_height 16
static unsigned char picnic_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0xfe, 0x7f, 0xfe, 0x7f,
   0x70, 0x0e, 0x30, 0x0c, 0x38, 0x1c, 0x38, 0x1c, 0xff, 0xff, 0xff, 0xff,
   0x06, 0x60, 0x06, 0x60, 0x03, 0xc0, 0x03, 0xc0};""",

	'plus': """#define plus_width 16
#define plus_height 16
static unsigned char plus_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01,
   0x80, 0x01, 0xff, 0xff, 0xff, 0xff, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01,
   0x80, 0x01, 0x80, 0x01, 0x80, 0x01, 0x80, 0x01};""",

	'q1': """#define q1_width 16
#define q1_height 16
static unsigned char q1_bits[] = {
   0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00,
   0xff, 0x00, 0xff, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'q1_notch': """#define q1_notch_width 16
#define q1_notch_height 16
static unsigned char q1_notch_bits[] = {
   0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0x0f, 0x00, 0x0f, 0x00,
   0x0f, 0x00, 0x0f, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'q1_20': """#define q1_20_width 20
#define q1_20_height 20
static unsigned char q1_20_bits[] = {
   0xff, 0x03, 0x00, 0xff, 0x03, 0x00, 0xff, 0x03, 0x00, 0xff, 0x03, 0x00,
   0xff, 0x03, 0x00, 0xff, 0x03, 0x00, 0xff, 0x03, 0x00, 0xff, 0x03, 0x00,
   0xff, 0x03, 0x00, 0xff, 0x03, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'q2': """#define q2_width 16
#define q2_height 16
static unsigned char q2_bits[] = {
   0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff,
   0x00, 0xff, 0x00, 0xff, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'q2_notch': """#define q2_notch_width 16
#define q2_notch_height 16
static unsigned char q2_notch_bits[] = {
   0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xf0, 0x00, 0xf0,
   0x00, 0xf0, 0x00, 0xf0, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'q2_20': """#define q2_20_width 20
#define q2_20_height 20
static unsigned char q2_20_bits[] = {
   0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f,
   0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f,
   0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'q3': """#define q3_width 16
#define q3_height 16
static unsigned char q3_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff,
   0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff};""",

	'q3_notch': """#define q3_notch_width 16
#define q3_notch_height 16
static unsigned char q3_notch_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0xf0, 0x00, 0xf0, 0x00, 0xf0, 0x00, 0xf0,
   0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff};""",

	'q3_20': """#define q3_20_width 20
#define q3_20_height 20
static unsigned char q3_20_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f,
   0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f,
   0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f, 0x00, 0xfc, 0x0f};""",

	'q4': """#define q4_width 16
#define q4_height 16
static unsigned char q4_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00,
   0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00};""",

	'q4_notch': """#define q4_notch_width 16
#define q4_notch_height 16
static unsigned char q4_notch_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x0f, 0x00, 0x0f, 0x00, 0x0f, 0x00, 0x0f, 0x00,
   0xff, 0x00, 0xff, 0x00, 0xff, 0x00, 0xff, 0x00};""",

	'q4_20': """#define q4_20_width 20
#define q4_20_height 20
static unsigned char q4_20_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0xff, 0x03, 0x00, 0xff, 0x03, 0x00,
   0xff, 0x03, 0x00, 0xff, 0x03, 0x00, 0xff, 0x03, 0x00, 0xff, 0x03, 0x00,
   0xff, 0x03, 0x00, 0xff, 0x03, 0x00, 0xff, 0x03, 0x00, 0xff, 0x03, 0x00};""",

	'qmark_circle': """#define qmark_circle_width 16
#define qmark_circle_height 16
static unsigned char qmark_circle_bits[] = {
   0xc0, 0x03, 0x70, 0x0e, 0x18, 0x18, 0xcc, 0x33, 0xe6, 0x67, 0x72, 0x4e,
   0x33, 0xcc, 0x01, 0x87, 0x81, 0x83, 0x83, 0xc1, 0x02, 0x40, 0x86, 0x61,
   0x8c, 0x31, 0x38, 0x18, 0x70, 0x0e, 0xc0, 0x03};""",

	'raincloud': """#define raincloud_width 16
#define raincloud_height 16
static unsigned char raincloud_bits[] = {
   0x00, 0x00, 0xe0, 0x03, 0xf0, 0x07, 0xf8, 0x0f, 0xf8, 0x1f, 0xfe, 0x1f,
   0xff, 0x7f, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xfe, 0xff, 0xfc, 0x7f,
   0x98, 0x19, 0x88, 0x08, 0xcc, 0x0c, 0x44, 0x04};""",

	'rose': """#define rose_width 16
#define rose_height 16
static unsigned char rose_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07,
   0xfc, 0x3f, 0xff, 0xff, 0xff, 0xff, 0xfc, 0x3f, 0xe0, 0x07, 0xc0, 0x03,
   0xc0, 0x03, 0xc0, 0x03, 0x80, 0x01, 0x80, 0x01};""",

	'square': """#define square_width 16
#define square_height 16
static unsigned char square_bits[] = {
   0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0,
   0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0, 0x07, 0xe0,
   0x07, 0xe0, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff};""",

	'star': """#define star_width 16
#define star_height 16
static unsigned char star_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07,
   0xff, 0xff, 0xff, 0xff, 0xfc, 0x3f, 0xf0, 0x0f, 0xf8, 0x1f, 0xf8, 0x1f,
   0x7c, 0x3e, 0x3c, 0x3c, 0x0e, 0x70, 0x06, 0x60};""",

	'surprise_circle': """#define surprise_circle_width 16
#define surprise_circle_height 16
static unsigned char surprise_circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x9c, 0x39, 0xce, 0x73, 0xc6, 0x63,
   0xc7, 0xe3, 0xc3, 0xc3, 0x83, 0xc1, 0x87, 0xe1, 0x06, 0x60, 0x8e, 0x71,
   0x9c, 0x39, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'target': """#define target_width 16
#define target_height 16
static unsigned char target_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x86, 0x61,
   0xc7, 0xe3, 0xe3, 0xc7, 0xe3, 0xc7, 0xc7, 0xe3, 0x86, 0x61, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'trash': """#define trash_width 16
#define trash_height 16
static unsigned char trash_bits[] = {
   0xf0, 0x0f, 0xff, 0xff, 0xff, 0xff, 0x06, 0x60, 0x66, 0x66, 0x66, 0x66,
   0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66, 0x66,
   0x66, 0x66, 0x06, 0x60, 0xfe, 0x7f, 0xfc, 0x3f};""",

	'tree': """
#define tree_width 16
#define tree_height 16
static unsigned char tree_bits[] = {
   0xf8, 0x00, 0xa8, 0x37, 0x7c, 0x7d, 0xef, 0x4a, 0x37, 0xf5, 0xdf, 0xaf,
   0xbe, 0xdb, 0xfc, 0x7f, 0xb0, 0x77, 0xc0, 0x7b, 0xc0, 0x1f, 0xc0, 0x03,
   0xc0, 0x07, 0xc0, 0x07, 0xe0, 0x0f, 0xfc, 0x0f};""",

	'triangle': """#define triangle_width 16
#define triangle_height 16
static unsigned char triangle_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0xe0, 0x07,
   0xf0, 0x0f, 0xf0, 0x0f, 0xf8, 0x1f, 0xf8, 0x1f, 0xfc, 0x3f, 0xfc, 0x3f,
   0xfe, 0x7f, 0xfe, 0x7f, 0xff, 0xff, 0xff, 0xff};""",

	'triangle_open': """#define triangle_open_width 16
#define triangle_open_height 16
static unsigned char triangle_open_bits[] = {
   0x80, 0x01, 0x80, 0x01, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0xe0, 0x07,
   0x70, 0x0e, 0x70, 0x0e, 0x38, 0x1c, 0x38, 0x1c, 0x1c, 0x38, 0x1c, 0x38,
   0x0e, 0x70, 0xfe, 0x7f, 0xff, 0xff, 0xff, 0xff};""",

	'vapor': """
#define vapor_width 16
#define vapor_height 16
static unsigned char vapor_bits[] = {
   0x20, 0x01, 0x30, 0x03, 0xcc, 0x4c, 0xc4, 0xcc, 0x13, 0x32, 0x31, 0x23,
   0xc8, 0x8c, 0x4c, 0xc4, 0x11, 0x22, 0x33, 0x33, 0xc4, 0xc4, 0x4c, 0x4c,
   0x32, 0x33, 0x20, 0x23, 0xc8, 0x0c, 0x80, 0x04};""",

	'wave': """#define wave_width 16
#define wave_height 16
static unsigned char wave_bits[] = {
   0x00, 0x00, 0x70, 0x00, 0xf8, 0x00, 0xce, 0x00, 0x83, 0x01, 0x00, 0xc3,
   0x00, 0xe6, 0x70, 0x3e, 0xf8, 0x1c, 0xce, 0x00, 0x83, 0x01, 0x00, 0xc3,
   0x00, 0xe6, 0x00, 0x3e, 0x00, 0x1c, 0x00, 0x00};""",

	'wedge': """#define wedge_width 16
#define wedge_height 16
static unsigned char stn_marker_inv_bits[] = {
   0xff, 0xff, 0xff, 0x7f, 0xfe, 0x7f, 0xfe, 0x3f, 0xfc, 0x3f, 0xfc, 0x1f,
   0xf8, 0x1f, 0xf8, 0x0f, 0xf0, 0x0f, 0xf0, 0x07, 0xe0, 0x07, 0xe0, 0x03,
   0xc0, 0x03, 0xc0, 0x01, 0x80, 0x01, 0x80, 0x00};""",

	'wheelchair': """#define wheelchair_width 16
#define wheelchair_height 16
static unsigned char wheelchair_bits[] = {
   0x30, 0x00, 0x78, 0x00, 0x78, 0x00, 0x78, 0x00, 0x30, 0x00, 0x77, 0x1f,
   0x76, 0x1f, 0xf6, 0x06, 0xf3, 0x07, 0xf3, 0x3f, 0x03, 0x3f, 0x03, 0x36,
   0x07, 0x77, 0x06, 0x63, 0xfe, 0xe3, 0xf8, 0xe0};""",

	'x': """#define x_width 16
#define x_height 16
static unsigned char x_bits[] = {
   0x00, 0x00, 0x06, 0x60, 0x0e, 0x70, 0x1c, 0x38, 0x38, 0x1c, 0x70, 0x0e,
   0xe0, 0x07, 0xc0, 0x03, 0xc0, 0x03, 0xe0, 0x07, 0x70, 0x0e, 0x38, 0x1c,
   0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60, 0x00, 0x00};""",

	'0': """#define i0_width 16
#define i0_height 16
static unsigned char i0_bits[] = {
   0xf0, 0x0f, 0xf8, 0x1f, 0xfc, 0x3f, 0x3c, 0x3c, 0x1e, 0x78, 0x0e, 0x70,
   0x8e, 0x71, 0xce, 0x73, 0xce, 0x73, 0x8e, 0x71, 0x0e, 0x70, 0x1e, 0x78,
   0x3c, 0x3c, 0xfc, 0x3f, 0xf8, 0x1f, 0xf0, 0x0f};""",

	'1': """#define i1_width 16
#define i1_height 16
static unsigned char i1_bits[] = {
   0xc0, 0x01, 0xe0, 0x01, 0xf0, 0x01, 0xf8, 0x01, 0xf8, 0x01, 0xc0, 0x01,
   0xc0, 0x01, 0xc0, 0x01, 0xc0, 0x01, 0xc0, 0x01, 0xc0, 0x01, 0xc0, 0x01,
   0xc0, 0x01, 0xc0, 0x01, 0xf8, 0x0f, 0xf8, 0x0f};""",

	'2': """#define i2_width 16
#define i2_height 16
static unsigned char i2_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0xf8, 0x1f, 0x3c, 0x3e, 0x1c, 0x38, 0x00, 0x38,
   0x00, 0x3c, 0x00, 0x1e, 0x80, 0x1f, 0xe0, 0x0f, 0xf8, 0x03, 0xf8, 0x00,
   0x3c, 0x00, 0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f};""",

	'3': """#define i3_width 16
#define i3_height 16
static unsigned char i3_bits[] = {
   0xe0, 0x03, 0xf8, 0x0f, 0xf8, 0x1e, 0x3c, 0x38, 0x1c, 0x38, 0x00, 0x3c,
   0x80, 0x1f, 0xc0, 0x0f, 0x80, 0x1f, 0x00, 0x3c, 0x00, 0x38, 0x1c, 0x38,
   0x3c, 0x3c, 0xf8, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'4': """define i4.xbm_width 16
#define i4.xbm_height 16
static unsigned char i4.xbm_bits[] = {
   0x00, 0x0f, 0x80, 0x0f, 0xc0, 0x0f, 0xc0, 0x0f, 0xe0, 0x0e, 0xf0, 0x0e,
   0x70, 0x0e, 0x78, 0x0e, 0x3c, 0x0e, 0xfc, 0x7f, 0xfc, 0x7f, 0xfc, 0x7f,
   0x00, 0x0e, 0x00, 0x0e, 0x00, 0x0e, 0x00, 0x0e};""",

	'5': """#define i5.xbm_width 16
#define i5.xbm_height 16
static unsigned char i5.xbm_bits[] = {
   0xfc, 0x3f, 0xfc, 0x3f, 0xfc, 0x3f, 0x1c, 0x00, 0x1c, 0x00, 0xfc, 0x07,
   0xfc, 0x1f, 0xe0, 0x3f, 0x00, 0x3e, 0x00, 0x78, 0x00, 0x70, 0x1c, 0x70,
   0x3c, 0x78, 0xf8, 0x3e, 0xf0, 0x1f, 0xc0, 0x07};""",

	'6': """#define i6_width 16
#define i6_height 16
static unsigned char i6_bits[] = {
   0x80, 0x0f, 0xc0, 0x07, 0xe0, 0x03, 0xf0, 0x01, 0xf8, 0x00, 0x78, 0x00,
   0xfc, 0x03, 0xfc, 0x0f, 0xfc, 0x1f, 0x3c, 0x3c, 0x1c, 0x38, 0x1c, 0x38,
   0x3c, 0x38, 0x78, 0x1c, 0xf0, 0x0f, 0xe0, 0x07};""",

	'7': """#define i7_width 16
#define i7_height 16
static unsigned char i7_bits[] = {
   0xfc, 0x7f, 0xfc, 0x7f, 0xfc, 0x7f, 0x00, 0x3c, 0x00, 0x1e, 0x00, 0x1e,
   0x00, 0x0f, 0x80, 0x07, 0x80, 0x07, 0xc0, 0x03, 0xe0, 0x01, 0xe0, 0x01,
   0xf0, 0x00, 0x78, 0x00, 0x78, 0x00, 0x3c, 0x00};""",

	'8': """#define i8_width 16
#define i8_height 16
static unsigned char i8_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x1c, 0x38, 0x3c, 0x3c,
   0xf8, 0x1f, 0xf0, 0x0f, 0xf8, 0x1f, 0x3c, 0x3c, 0x1c, 0x38, 0x1c, 0x38,
   0x3c, 0x3c, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};""",

	'9': """#define i9_width 16
#define i9_height 16
static unsigned char i9_bits[] = {
   0xe0, 0x07, 0xf0, 0x0f, 0x38, 0x1e, 0x1c, 0x3c, 0x1c, 0x38, 0x1c, 0x38,
   0x3c, 0x3c, 0xf8, 0x3f, 0xf0, 0x3f, 0xc0, 0x3f, 0x00, 0x1e, 0x00, 0x1f,
   0x80, 0x0f, 0xc0, 0x07, 0xe0, 0x03, 0xf0, 0x01};""",

	'10': """#define ten_width 16
#define ten_height 16
static unsigned char ten_bits[] = {
   0x1c, 0x0c, 0x1e, 0x1e, 0x1f, 0x3f, 0x9f, 0x73, 0x9c, 0x73, 0xdc, 0xe1,
   0xdc, 0xe1, 0xdc, 0xc0, 0xdc, 0xc0, 0xdc, 0xe1, 0xdc, 0xe1, 0x9c, 0x73,
   0x9c, 0x73, 0x1c, 0x3f, 0x7f, 0x1e, 0x7f, 0x0c};""",


	'ball20': """#define ball20_width 20
#define ball20_height 20
static unsigned char ball20_bits[] = {
   0x00, 0x00, 0xf0, 0x00, 0x00, 0xf0, 0x00, 0x0f, 0xf0, 0xc0, 0x3f, 0xf0,
   0xe0, 0x7f, 0xf0, 0xf0, 0xff, 0xf0, 0xf8, 0xff, 0xf1, 0xf8, 0xff, 0xf1,
   0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3,
   0xf8, 0xff, 0xf1, 0xf8, 0xff, 0xf1, 0xf0, 0xff, 0xf0, 0xe0, 0x7f, 0xf0,
   0xc0, 0x3f, 0xf0, 0x00, 0x0f, 0xf0, 0x00, 0x00, 0xf0, 0x00, 0x00, 0xf0};""",

	'block20': """#define block20_width 20
#define block20_height 20
static unsigned char block20_bits[] = {
   0x00, 0x00, 0xf0, 0x00, 0x00, 0xf0, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3,
   0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3,
   0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3,
   0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3,
   0xfc, 0xff, 0xf3, 0xfc, 0xff, 0xf3, 0x00, 0x00, 0xf0, 0x00, 0x00, 0xf0};""",

	'circle20': """#define circle20_width 20
#define circle20_height 20
static unsigned char circle20_bits[] = {
   0xc0, 0x3f, 0x00, 0xf0, 0xff, 0x00, 0xf8, 0xf0, 0x01, 0x3c, 0xc0, 0x03,
   0x1e, 0x80, 0x07, 0x0e, 0x00, 0x07, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e,
   0x03, 0x00, 0x0c, 0x03, 0x00, 0x0c, 0x03, 0x00, 0x0c, 0x03, 0x00, 0x0c,
   0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x0e, 0x00, 0x07, 0x1e, 0x80, 0x07,
   0x3c, 0xc0, 0x03, 0xf8, 0xf0, 0x01, 0xf0, 0xff, 0x00, 0xc0, 0x3f, 0x00};""",

	'square20': """#define square20_width 20
#define square20_height 20
static unsigned char square20_bits[] = {
   0xff, 0xff, 0x0f, 0xff, 0xff, 0x0f, 0xff, 0xff, 0x0f, 0x07, 0x00, 0x0e,
   0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e,
   0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e,
   0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e, 0x07, 0x00, 0x0e,
   0x07, 0x00, 0x0e, 0xff, 0xff, 0x0f, 0xff, 0xff, 0x0f, 0xff, 0xff, 0x0f};""",


	'ball24': """#define ball24_width 24
#define ball24_height 24
static unsigned char ball24_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x3c, 0x00, 0x80, 0xff, 0x01,
   0xc0, 0xff, 0x03, 0xe0, 0xff, 0x07, 0xf0, 0xff, 0x0f, 0xf8, 0xff, 0x1f,
   0xf8, 0xff, 0x1f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xf8, 0xff, 0x1f, 0xf8, 0xff, 0x1f,
   0xf0, 0xff, 0x1f, 0xf0, 0xff, 0x0f, 0xe0, 0xff, 0x07, 0xc0, 0xff, 0x03,
   0x80, 0xff, 0x01, 0x00, 0x3c, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'block24': """#define block24_width 24
#define block24_height 24
static unsigned char block24_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f,
   0xfc, 0xff, 0x3f, 0xfc, 0xff, 0x3f, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};""",

	'circle24': """#define circle24_width 24
#define circle24_height 24
static unsigned char circle24_bits[] = {
   0x00, 0x7e, 0x00, 0xc0, 0xff, 0x03, 0xe0, 0xc3, 0x07, 0x70, 0x00, 0x0e,
   0x38, 0x00, 0x1c, 0x1c, 0x00, 0x38, 0x0e, 0x00, 0x70, 0x06, 0x00, 0x60,
   0x06, 0x00, 0x60, 0x03, 0x00, 0xc0, 0x03, 0x00, 0xc0, 0x03, 0x00, 0xc0,
   0x03, 0x00, 0xc0, 0x03, 0x00, 0xc0, 0x07, 0x00, 0xe0, 0x06, 0x00, 0x60,
   0x0e, 0x00, 0x60, 0x0e, 0x00, 0x70, 0x1c, 0x00, 0x38, 0x38, 0x00, 0x1c,
   0x70, 0x00, 0x0e, 0xe0, 0xc3, 0x07, 0xc0, 0xff, 0x03, 0x00, 0x7e, 0x00};""",

	'square24': """#define square24_width 24
#define square24_height 24
static unsigned char square24_bits[] = {
   0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0x07, 0x00, 0xe0,
   0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0,
   0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0,
   0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0,
   0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0, 0x07, 0x00, 0xe0,
   0x07, 0x00, 0xe0, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff, 0xff};""",


	'ball28': """#define ball28_width 28
#define ball28_height 28
static unsigned char circle28_bits[] = {
   0x00, 0x00, 0x00, 0xf0, 0x00, 0x00, 0x00, 0xf0, 0x00, 0x00, 0x00, 0xf0,
   0x00, 0xfc, 0x03, 0xf0, 0x80, 0xff, 0x1f, 0xf0, 0xc0, 0xff, 0x3f, 0xf0,
   0xe0, 0xff, 0x7f, 0xf0, 0xf0, 0xff, 0xff, 0xf0, 0xf0, 0xff, 0xff, 0xf0,
   0xf0, 0xff, 0xff, 0xf0, 0xf8, 0xff, 0xff, 0xf1, 0xf8, 0xff, 0xff, 0xf1,
   0xf8, 0xff, 0xff, 0xf1, 0xf8, 0xff, 0xff, 0xf1, 0xf8, 0xff, 0xff, 0xf1,
   0xf8, 0xff, 0xff, 0xf1, 0xf8, 0xff, 0xff, 0xf1, 0xf0, 0xff, 0xff, 0xf0,
   0xf0, 0xff, 0xff, 0xf0, 0xf0, 0xff, 0xff, 0xf0, 0xe0, 0xff, 0x7f, 0xf0,
   0xe0, 0xff, 0x7f, 0xf0, 0xc0, 0xff, 0x3f, 0xf0, 0x80, 0xff, 0x1f, 0xf0,
   0x00, 0xfc, 0x03, 0xf0, 0x00, 0x00, 0x00, 0xf0, 0x00, 0x00, 0x00, 0xf0,
   0x00, 0x00, 0x00, 0xf0};""",

	'block28': """#define block28_width 28
#define block28_height 28
static unsigned char block28_bits[] = {
   0x00, 0x00, 0x00, 0xf0, 0x00, 0x00, 0x00, 0xf0, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3,
   0xfc, 0xff, 0xff, 0xf3, 0xfc, 0xff, 0xff, 0xf3, 0x00, 0x00, 0x00, 0xf0,
   0x00, 0x00, 0x00, 0xf0};""",

	'circle28': """#define circle28_width 28
#define circle28_height 28
static unsigned char circle24_bits[] = {
   0x00, 0xfe, 0x07, 0x00, 0x80, 0xff, 0x1f, 0x00, 0xe0, 0xff, 0x7f, 0x00,
   0xf0, 0x03, 0xfc, 0x00, 0x78, 0x00, 0xe0, 0x01, 0x3c, 0x00, 0xc0, 0x03,
   0x1c, 0x00, 0x80, 0x03, 0x0e, 0x00, 0x00, 0x07, 0x0e, 0x00, 0x00, 0x07,
   0x0f, 0x00, 0x00, 0x0f, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x0f, 0x00, 0x00, 0x0f,
   0x0f, 0x00, 0x00, 0x0f, 0x0e, 0x00, 0x00, 0x07, 0x1e, 0x00, 0x80, 0x07,
   0x1c, 0x00, 0x80, 0x03, 0x3c, 0x00, 0xc0, 0x03, 0x78, 0x00, 0xe0, 0x01,
   0xf0, 0x03, 0xfc, 0x00, 0xe0, 0xff, 0x7f, 0x00, 0x80, 0xff, 0x1f, 0x00,
   0x00, 0xfe, 0x07, 0x00};""",

	'square28': """#define square28_width 28
#define square28_height 28
static unsigned char square28_bits[] = {
   0xff, 0xff, 0xff, 0x0f, 0xff, 0xff, 0xff, 0x0f, 0xff, 0xff, 0xff, 0x0f,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e, 0x07, 0x00, 0x00, 0x0e,
   0x07, 0x00, 0x00, 0x0e, 0xff, 0xff, 0xff, 0x0f, 0xff, 0xff, 0xff, 0x0f,
   0xff, 0xff, 0xff, 0x0f};"""

	}

# X11 bitmaps for map button bar icons
expand_xbm = """#define expand_width 16
#define expand_height 16
static unsigned char expand_bits[] = {
   0x3f, 0xfc, 0x07, 0xe0, 0x0f, 0xf0, 0x1d, 0xb8, 0x39, 0x9c, 0x71, 0x8e,
   0x60, 0x06, 0x00, 0x00, 0x00, 0x00, 0x61, 0x06, 0x71, 0x8e, 0x39, 0x9c,
   0x1d, 0xb8, 0x0f, 0xf0, 0x07, 0xe0, 0x3f, 0xfc};"""

wedges_3_xbm = """#define wedges_3_width 16
#define wedges_3_height 16
static unsigned char wedges_3_bits[] = {
   0xff, 0x01, 0xfe, 0x00, 0x7c, 0x00, 0x38, 0x00, 0x10, 0x00, 0x00, 0x00,
   0x80, 0xff, 0x00, 0x7f, 0x00, 0x3e, 0x00, 0x1c, 0x00, 0x08, 0xff, 0x01,
   0xfe, 0x00, 0x7c, 0x00, 0x38, 0x00, 0x10, 0x00};"""

wedge_sm_xbm = """#define wedge_sm_width 16
#define wedge_sm_height 16
static unsigned char wedge_sm_bits[] = {
   0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0xf8, 0x1f, 0xf8, 0x1f, 0xf0, 0x0f,
   0xf0, 0x0f, 0xe0, 0x07, 0xe0, 0x07, 0xc0, 0x03, 0xc0, 0x03, 0x80, 0x01,
   0x80, 0x01, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00};"""

circle_xbm = """#define circle_width 16
#define circle_height 16
static unsigned char circle_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x70, 0x06, 0x60,
   0x07, 0xe0, 0x03, 0xc0, 0x03, 0xc0, 0x07, 0xe0, 0x06, 0x60, 0x0e, 0x70,
   0x1c, 0x38, 0x78, 0x1e, 0xf0, 0x0f, 0xc0, 0x03};"""

cancel_xbm = """#define cancel_width 16
#define cancel_height 16
static unsigned char cancel_bits[] = {
   0xc0, 0x03, 0xf0, 0x0f, 0x78, 0x1e, 0x1c, 0x38, 0x0e, 0x7c, 0x06, 0x6e,
   0x07, 0xe7, 0x83, 0xc3, 0xc3, 0xc1, 0xe7, 0xe0, 0x76, 0x60, 0x3e, 0x78,
   0x1c, 0x38, 0x78, 0x3e, 0xf0, 0x0f, 0xc0, 0x03};"""

# X11 bitmap for (invisible) marking of locations for count labeling
null_xbm = """
#define null_width 1
#define null_height 1
static unsigned char null_bits[] = {
   0x00};"""


# Color names for map symbols.  See https://www.w3schools.com/colors/colors_names.asp.
color_names = ("aliceblue", "antiquewhite", "aqua", "aquamarine", "azure", "beige", "bisque", "black", "blanchedalmond",
		"blue", "blueviolet", "brown", "burlywood", "cadetblue", "chartreuse", "chocolate", "coral", "cornflowerblue",
		"cornsilk", "crimson", "cyan", "darkblue", "darkcyan", "darkgoldenrod", "darkgray", "darkgrey", "darkgreen",
		"darkkhaki", "darkmagenta", "darkolivegreen", "darkorange", "darkorchid", "darkred", "darksalmon", "darkseagreen",
		"darkslateblue", "darkslategray", "darkslategrey", "darkturquose", "darkviolet", "deeppink", "deepskyblue",
		"dimgray", "dimgrey", "dodgerblue", "firebrick", "floralwhite", "forestgreen", "fuschia", "gainsboro", "ghostwhite",
		"gold", "goldenrod", "gray", "grey", "green", "greenyellow", "honeydew", "hotpink", "indianred", "indigo", "ivory",
		"khaki", "lavender", "lavenderblush", "lawngreen", "lemonchiffon", "lightblue", "lightcoral", "lightcyan",
		"lightgoldenrodyellow", "lightgray", "lightgrey", "lightgreen", "lightpink", "lightsalmon", "lightseagreen",
		"lightskyblue", "lightslategray", "lightslategrey", "lightsteelblue", "lightyellow", "lime", "limegreen", "linen",
		"magenta", "maroon", "mediumaquamarine", "mediumblue", "mediumorchid", "mediumpurple", "mediumseagreen",
		"mediumslateblue", "mediumspringgreen", "mediumturquose", "mediumvioletred", "midnightblue", "mintcream", "mistyrose",
		"moccasin", "navajowhite", "navy", "oldlace", "olive", "olivedrab", "orange", "orangered", "orchid", "palegoldenrod",
		"palegreen", "paleturquose", "palevioletred", "papayawhip", "peachpuff", "peru", "pink", "plum", "powderblue",
		"purple", "rebeccapurple", "red", "rosybrown", "royalblue", "saddlebrown", "salmon", "sandybrown", "seagreen",
		"seashell", "sienna", "silver", "skyblue", "slateblue", "slategray", "slategrey", "snow", "springgreen",
		"steelblue", "tan", "teal", "thistle", "tomato", "turquoise", "violet", "wheat", "white", "whitesmoke", "yellow",
		"yellowgreen")

# A shorter list for interactive selection of the marker color
select_colors = ('aqua', 'black', 'blue', 'blueviolet', 'brown', 'chartreuse', 'cornflowerblue', 'crimson',
		'cyan', 'darkblue', 'darkgreen', 'darkmagenta', 'darkorange', 'darkred', 'darkslategray', 'deeppink',
		'forestgreen', 'fuschia', 'green', 'greenyellow','lightblue', 'lightgray', 'lightgreen', 'lightyellow',
		'magenta', 'maroon', 'navy', 'orange', 'orangered', 'purple', 'red', 'violet', 'white', 'yellow', 'yellowgreen')


# List of imported symbol names and paths
imported_symbols = []

# Keys for custom symbols are made up of the color name and the icon name, separated with a space.
custom_icons = {}


# X11 bitmap for the application window icon
win_icon_xbm = """#define window_icon_width 16
#define window_icon_height 16
static unsigned char window_icon_bits[] = {
   0xff, 0xff, 0x01, 0x80, 0x01, 0x84, 0x01, 0x8e, 0x01, 0x9f, 0x81, 0xbf,
   0x21, 0x80, 0x71, 0x80, 0xf9, 0x80, 0xfd, 0x81, 0x01, 0x84, 0x01, 0x8e,
   0x01, 0x9f, 0x81, 0xbf, 0x01, 0x80, 0xff, 0xff};"""


# SQLite3 database in which the data table is stored for querying.
dbtmpdir = None
temp_dbfile = None
data_db = None

# Dictionary keyed by lat,lon tuple containing a list of location IDs for that location.
location_ids = {} 


# Continuity storage for random selection dialog
random_last_n = 1
random_use_selected = False


#=====  Functions and classes  =====

def warning(message, kwargs):
	dlg = MsgDialog("MapData Warning", message, parent=kwargs.get('parent'), bgcolor="Gold")
	dlg.show()

def fatal_error(message, kwargs):
	dlg = MsgDialog("MapData Fatal Error", message, parent=kwargs.get('parent'), bgcolor="Red")
	dlg.show()
	sys.exit()

def warning_nolog(msg="Data can't be log-transformed.", parent=None):
	if parent is not None:
		warning(msg, {"parent": parent})
	else:
		warning(msg, {})

def make_data_db():
	global dbtmpdir, temp_dbfile, data_db
	if data_db is not None:
		data_db.close()
	if dbtmpdir is not None:
		dbtmpdir.cleanup(ignore_cleanup_errors = True)
	if temp_dbfile:
		dbtmpdir = tempfile.TemporaryDirectory()
		dbname = os.path.join(self.dbtmpdir.name, "mapdata.db")
		data_db = sqlite3.connect(dbname)
	else:
		tmpdir = None
		dbname = None
		data_db = sqlite3.connect(":memory:")

def makefont(fontname, fontsize, fontbold):
	fams = tkfont.families()
	if not fontname in fams:
		alt_fonts = ["Liberation Sans", "Arial", "Helvetica", "Nimbus Sans", "Liberation Sans", "Trebuchet MS", "Tahoma", "DejaVu Sans", "Bitstream Vera Sans", "Open Sans"]
		font_found = False
		for f in alt_fonts:
			if f in fams:
				fontname = f
				font_found = True
				break
		if not font_found:
			label_font = tkfont.nametofont("TkDefaultFont").actual()["family"]
	boldstr = "normal" if not fontbold else "bold"
	return tkfont.Font(family=fontname, size=fontsize, weight=boldstr)

def map_colors(lbl_list, color_list=quant_colors):
	lbllen = len(lbl_list)
	collen = len(color_list)
	if lbllen > collen:
		ext_colors = color_list * math.ceil(lbllen/collen)
		return dict(zip(lbl_list, ext_colors[:len(lbl_list)]))
	else:
		return dict(zip(lbl_list, color_list[:len(lbl_list)]))

def adjust_tile_color(tile_image):
	if map_settings.basemap_color_adj is None:
		return tile_image
	else:
		sat, con, bri = map_settings.basemap_color_adj
		if sat != 0:
			tile_image = ImageEnhance.Color(tile_image).enhance(sat)
		if con != 0:
			tile_image = ImageEnhance.Contrast(tile_image).enhance(con)
		if bri != 0:
			tile_image = ImageEnhance.Brightness(tile_image).enhance(bri)
		return tile_image

def doane_bins(d):
	# The Doane algorithm for histogram bins.
	dset = set(d)
	if len(dset) == 1:
		return 1
	else:
		g1 = spstats.skew(d)
		if math.isnan(g1):
			return 1
		else:
			nd = len(d)
			g1_sd = math.sqrt((6*(nd-2))/((nd+1)*(nd+3)))
			if g1_sd == 0.0:
				return 1
			else:
				return math.ceil(1 + math.log(nd,2) + math.log(1+abs(g1)/g1_sd,2))

#def scott_bins(d):
	# The Scott algorithm for histogram bins.
#	dset = set(d)
#	if len(dset) == 1:
#		return 1
#	else:
#		d_sd = statistics.stdev(d)
#		if d_sd == 0.0:
#			return 1
#		else:
#			nd = len(d)
#			return math.ceil((3.49 * d_sd)/(nd**(1.0/3)))

def clear_dlg_hotkeys(dlg):
	# Not Alt-c or Alt-h
	dlg.unbind("<Alt-a>")
	dlg.unbind("<Alt-b>")
	dlg.unbind("<Alt-f>")
	dlg.unbind("<Alt-g>")
	dlg.unbind("<Alt-l>")
	dlg.unbind("<Alt-o>")
	dlg.unbind("<Alt-p>")
	dlg.unbind("<Alt-r>")
	dlg.unbind("<Alt-s>")
	dlg.unbind("<Alt-t>")
	dlg.unbind("<Alt-x>")
	dlg.unbind("<Alt-y>")
	dlg.unbind("<Control-s>")
	dlg.unbind("<Control-z>")

def xicor(X, Y, ties=True):
	# xi correlation coefficient.  Adapted from https://towardsdatascience.com/a-new-coefficient-of-correlation-64ae4f260310
    n = len(X)
    order = np.array([i[0] for i in sorted(enumerate(X), key=lambda x: x[1])])
    if ties:
        l = np.array([sum(y >= Y[order]) for y in Y[order]])
        r = l.copy()
        for j in range(n):
            if sum([r[j] == r[i] for i in range(n)]) > 1:
                tie_index = np.array([r[j] == r[i] for i in range(n)])
                r[tie_index] = np.random.choice(r[tie_index] - np.arange(0, sum([r[j] == r[i] for i in range(n)])), sum(tie_index), replace=False)
        return 1 - n*sum( abs(r[1:] - r[:n-1]) ) / (2*sum(l*(n - l)))
    else:
        r = np.array([sum(y >= Y[order]) for y in Y[order]])
        return 1 - 3 * sum( abs(r[1:] - r[:n-1]) ) / (n**2 - 1)

def pearsonstat(x, y):
	return spstats.pearsonr(x, y).statistic
def spearmanstat(x, y):
	return spstats.spearmanr(x, y).statistic
def kendallstat(x, y):
	return spstats.kendalltau(x, y).statistic

def array_rss(m1, m2):
	# Arguments are numpy 2D arrays of the same dimensions.
	r, c = m1.shape
	rss = 0.0
	for i in range(r):
		for j in range(c):
			rss += (m1[i][j] - m2[i][j])**2
	return rss

def rowsums(a):
    return np.sum(a, axis=1)

def colsums(a):
    return rowsums(a.transpose())

def arraysum(a):
    return sum(rowsums(a))

# Utility functions

def enable_if(widget, tf_val):
	widget["state"] = tk.NORMAL if tf_val else tk.DISABLED

def chosen_dataset(maingui, column_list, selected_only):
	if selected_only:
		return maingui.get_sel_data(column_list)
	else:
		return maingui.get_all_data(column_list)

def chosen_dataset_and_ids(maingui, column_list, selected_only):
	if selected_only:
		return maingui.get_sel_data(column_list), maingui.get_sel_rowids()
	else:
		return maingui.get_all_data(column_list), maingui.get_all_rowids()

def new_button(container, label, row, column, action, sticky=tk.W, padx=(0,0), pady=(0,0), underline=None, state=tk.NORMAL):
	btn = ttk.Button(container, text=label, command=action, state=state, underline=underline)
	btn.grid(row=row, column=column, sticky=sticky, padx=padx, pady=pady)
	return btn

def new_help_button(dlg, container, action):
	btn = new_button(container, "Help", 0, 0, action, padx=(6,3), underline=0)
	dlg.bind("<Alt-h>", action)
	return btn

def new_data_button(dlg, container, column, action):
	return new_button(container, "Source Data", 0, column, action, padx=(3,3), state=tk.DISABLED)

def new_plot_button(dlg, container, column, action):
	return new_button(container, "Plot Data", 0, column, action, padx=(3,3), state=tk.DISABLED)

def new_addcol_button(dlg, container, column, action):
	return new_button(container, "Add Column", 0, column, action, padx=(3,3), state=tk.DISABLED)

def new_ok_button(dlg, container, column, action, ok_enabled=False):
	btn = new_button(container, "OK", 0, column, action, tk.E, padx=(3,3), underline=0, state=tk.NORMAL if ok_enabled else tk.DISABLED)
	dlg.bind("<Return>", action)
	dlg.bind("<Alt-o>", action)
	return btn

def new_cancel_button(dlg, container, column, action):
	btn = new_button(container, "Cancel", 0, column, action, tk.E, padx=(6,6), underline=0)
	dlg.bind("<Alt-c>", action)
	dlg.bind("<Escape>", action)
	return btn

def new_close_button(dlg, container, column, action):
	btn = new_button(container, "Close", 0, column, action, tk.E, padx=(6,6), underline=0)
	dlg.bind("<Alt-c>", action)
	dlg.bind("<Escape>", action)
	return btn

def add_help_close_buttons(dlg, container, help_action, close_action):
	new_help_button(dlg, container, help_action)
	new_close_button(dlg, container, 1, close_action)
	container.columnconfigure(1, weight=1)

def add_help_ok_cancel_buttons(dlg, container, help_action, ok_action, cancel_action, ok_enabled=False):
	container.columnconfigure(0, weight=0)
	container.columnconfigure(1, weight=1)
	new_help_button(dlg, container, help_action)
	ok_btn = new_ok_button(dlg, container, 1, ok_action, ok_enabled)
	new_cancel_button(dlg, container, 2, cancel_action)
	return ok_btn

def add_help_src_plot_close_btns(dlg, container, help_action, data_action, plot_action, close_action):
	container.columnconfigure(0, weight=0)
	container.columnconfigure(1, weight=0)
	container.columnconfigure(2, weight=1)
	new_help_button(dlg, container, help_action)
	data_btn = new_data_button(dlg, container, 1, data_action)
	plot_data_btn = new_plot_button(dlg, container, 2, plot_action)
	new_close_button(dlg, container, 3, close_action)
	return data_btn, plot_data_btn

def add_help_src_close_btns(dlg, container, help_action, data_action, close_action):
	container.columnconfigure(0, weight=0)
	container.columnconfigure(1, weight=1)
	new_help_button(dlg, container, help_action)
	data_btn = new_data_button(dlg, container, 1, data_action)
	new_close_button(dlg, container, 2, close_action)
	return data_btn

def add_sel_only(container, row, col, change_action, colspan=None):
	sel_var = tk.StringVar(container, "0")
	sel_ck = ttk.Checkbutton(container, text="Selected data only", command=change_action, variable=sel_var,
			onvalue="1", offvalue="0")
	sel_ck.grid(row=row, column=col, sticky=tk.W, padx=(6,3), pady=(3,3), columnspan=colspan)
	return sel_var, sel_ck

def add_autoupdate(container, row, col, change_action, colspan=None, padx=(3,3)):
	au_var = tk.StringVar(container, "1")
	au_ck = ttk.Checkbutton(container, text="Auto-update", command=change_action, variable=au_var,
			onvalue="1", offvalue="0")
	au_ck.grid(row=row, column=col, sticky=tk.W, padx=padx, pady=(3,3), columnspan=colspan)
	return au_var


class VScrollFrame(ttk.Frame):
	# A custom frame with a vertical scrollbar.
	# Widgets should be added to the embedded 'content_frame'.
	def __init__(self, parent, height=150, *args, **kwargs):
		ttk.Frame.__init__(self, parent, *args, **kwargs)
		self.canvas = tk.Canvas(self, bd=0, highlightthickness=0, height=height)
		self.canvas.grid(row=0, column=0, sticky=tk.NSEW)
		self.canvas.bind('<Configure>', self.upd_canvas)
		self.rowconfigure(0, weight=1)
		self.columnconfigure(0, weight=1)
		vsb = ttk.Scrollbar(self, orient='vertical')
		vsb.grid(row=0, column=1, sticky=tk.NS)
		vsb.configure(command=self.canvas.yview)
		self.canvas.configure(yscrollcommand=vsb.set)
		self.content_frame = ttk.Frame(self.canvas)
		self.content_frame.bind('<Configure>', self.upd_content_frame)
		self.content_win = self.canvas.create_window(0, 0, window=self.content_frame, anchor=tk.NW)
	def upd_content_frame(self, event):
		reqwid = self.content_frame.winfo_reqwidth()
		self.canvas.config(scrollregion=(0, 0, reqwid, self.content_frame.winfo_reqheight()))
		if self.content_frame.winfo_reqwidth() != self.canvas.winfo_width():
			self.canvas.config(width = reqwid)
	def upd_canvas(self, event):
		if self.content_frame.winfo_reqwidth() != self.canvas.winfo_width():
			self.canvas.itemconfigure(self.content_win, width=self.canvas.winfo_width())
	def scrolltop(self):
		self.canvas.xview_moveto(0)
		self.canvas.yview_moveto(0)
    	

class CsvFile(object):
	def __init__(self, csvfname, junk_header_lines=0, dialect=None):
		self.csvfname = csvfname
		self.junk_header_lines = junk_header_lines
		self.lineformat_set = False		# Indicates whether delimiter, quotechar, and escapechar have been set
		self.delimiter = None
		self.quotechar = None
		self.escapechar = None
		self.inf = None
		self.colnames = None
		self.rows_read = 0
		# Detect encoding of the file has a BOM
		encoding = 'utf-8-sig'	# Default
		with open(csvfname, 'rb') as f:
			raw = f.read(4)
		for enc, boms in (
					('utf-8-sig', (codecs.BOM_UTF8,)),
					('utf_16', (codecs.BOM_UTF16_LE, codecs.BOM_UTF16_BE)),
					('utf_32', (codecs.BOM_UTF32_LE, codecs.BOM_UTF32_BE))):
			if any(raw.startswith(bom) for bom in boms):
				encoding = enc
		# Python 3 only
		self.reader = csv.reader(open(csvfname, mode="rt", encoding=encoding, newline=''), dialect=dialect)
	def __next__(self):
		row = next(self.reader)
		self.rows_read = self.rows_read + 1
		return row
	def next(self):
		row = next(self.reader)
		self.rows_read = self.rows_read + 1
		if self.rows_read == 1:
			self.colnames = row
		return row
	def __iter__(self):
		return self

def columns_to_rows(column_table):
	# Columns may have unequal lengths
	maxcollen = max([len(c) for c in column_table])
	ct = [c + [None]*(maxcollen - len(c)) for c in column_table]
	return list(map(list, zip(*ct)))

def rows_to_columns(row_table):
	# Every row must be filled equally
	return list(map(list, zip(*row_table)))

# Sorting strings respecting numeric part
rx_numstr = re.compile(r'(\d*\.?\d+)|(.)')
def key_numstr(ns):
	try:
		return [(sp, float(np)) if np != '' else (sp, np) for np, sp in re.findall(r'(\d+)|(.)', ns)]
	except:
		return [str(v) for v in ns]

def sorted_numstrs(v):
	if all([type(s) == str for s in v]):
		return sorted(v, key=key_numstr)
	else:
		return sorted(v)

def sort_columns(columns, sortby=0):
	# Sorts a list of sublists, where each sublist is a column.
	# Rows are sorted by the 'sortby' column, which is zero-based.
	# The returned value is also column-wise, but sorted by rows.
	ncols = len(columns)
	if ncols == 1:
		try:
			return [sorted_numstrs(columns[0])]
		except:
			try:
				return [sorted(columns[0])]
			except:
				return [sorted([str(v) for v in columns[0]])]
	nrows = len(columns[0])
	sortcol_is_str = any([type(s) == str for s in columns[sortby]])
	rowdata = [[columns[c][r] for c in range(ncols)] for r in range(nrows)]
	try:
		if sortcol_is_str:
			rowdata.sort(key = lambda c: (c[sortby] is None, key_numstr(c[sortby])))
		else:
			rowdata.sort(key = lambda c: (c[sortby] is None, c[sortby]))
	except:
		for r in range(nrows):
			rowdata[r][sortby] = str(rowdata[r][sortby])
		rowdata.sort(key = lambda c: (c[sortby] == 'None', c[sortby]))
	return [[rowdata[r][c] for r in range(nrows)] for c in range(ncols)]

def revsort_columns(columns, sortby=0):
	sc = sort_columns(columns, sortby)
	return [list(reversed(c)) for c in sc]

def check_rowtable(rowtable, headers, parent, tablename=None):
	# Checks that every row of the table has the same number of elements as the headers.
	# Halts with a fatal error if not.
	n_headers = len(headers)
	for rowno in range(len(rowtable)):
		if len(rowtable[rowno]) != n_headers:
			msg = tablename + ": " if tablename is not None else ""
			msg += f"Row {rowno+1} has {len(rowtable[rowno])} columns but should have {n_headers}.\nBad row: {rowtable[rowno]}"
			fatal_error(msg, kwargs={'parent': parent})

def clean_missing(dataset, column_indexes):
	# Removes rows for a table as a list of columns if there is a missing value in any
	# of the columns specified by the indexes
	clean_data = [[] for _ in dataset]
	for i in range(len(dataset[0])):
		ok = True
		for col in column_indexes:
			if dataset[col][i] is None or dataset[col][i] == '':
				ok = False
		if ok:
			for col in range(len(dataset)):
				clean_data[col].append(dataset[col][i])
	return clean_data

def clean_missing_columns(dataset, column_names, column_indexes):
	# Removes entire columns that have any missing values.
	# Returns the cleaned dataset, the revised column names, and the number of columns removed.
	badcols = []
	for col in column_indexes:
		if not all(dataset[col]):
			badcols.append(col)
	if len(badcols) > 0:
		badcols.reverse()
		for bc in badcols:
			del dataset[bc]
			del column_names[bc]
	return dataset, column_names, len(badcols)

def clean_missing_bycol(dataset, column_indexes):
	# Removes missing values from each column of a column-ordered data set.  This does
	# not preserve rows across columns.
	clean_data = [[] for _ in dataset]
	for col in column_indexes:
		for i in range(len(dataset[col])):
			if dataset[col][i] is not None and dataset[col][i] != '':
				clean_data[col].append(dataset[col][i])
	return clean_data

def logvector(v):
	# Return a vector of log10 values, or None.
	try:
		return [math.log10(x) for x in v]
	except:
		return None

def logdataset(dataset, column_indexes):
	# Log10 transforms specified columns of a column-oriented dataset.
	log_data = [[] for _ in dataset]
	for colidx in range(len(dataset)):
		if colidx in column_indexes:
			try:
				log_data[colidx] = [math.log10(x) for x in dataset[colidx]]
			except:
				return None
		else:
			log_data[colidx] = dataset[colidx]
	return log_data


def no_none(datalist):
	return [d for d in datalist if d is not None]

def spread_by_groups(grouplist, valuelist):
	# Split 'valuelist' into several lists, one for each group in grouplist.
	# The lists must have the same length.
	# Returns a tuple of group names (as strings) and a list of lists.  Missing values are eliminated.
	grpstrs = [str(g) for g in grouplist]
	grp_vals = sorted(list(set(grpstrs)))
	ds = list(zip(grpstrs, valuelist))
	dspread = []
	for g in grp_vals:
		dspread.append([d[1] for d in ds if d[0] == g and d[1] is not None])
	return grp_vals, dspread

def subset_by_groups(dataset, grouplist):
	# Returns a dictionary with keys of unique elements of 'grouplist' and values
	# corresponding to elements of 'dataset'.  'dataset' is a column-oriented list of lists.
	# The length of 'grouplist' must be the same as the lengths of the lists in 'dataset'--this is not checked.
	groups = sorted_numstrs(list(set(grouplist)))
	ncols = len(dataset)
	grp_datasets = {}
	for g in groups:
		grp_datasets[g] = [[] for _ in range(ncols)]
	for r in range(len(dataset[0])):
		g = grouplist[r]
		for col in range(ncols):
			grp_datasets[g][col].append(dataset[col][r])
	return grp_datasets

def aggregate_groups(subset_dict, agg_by):
	# Takes a dictionary like that created by 'subset_by_groups()' and returns a column-major list
	# of lists with data for each dictionary element aggregated according to the specified function.
	# All values in 'subset_dict' must have the same number of columns.
	# Side effect: may display a warning message.
	agg_data = {}
	agg_error = False
	groups = list(subset_dict.keys())
	datacols = len(subset_dict[groups[0]])
	for g in groups:
		agg_data[g] = [math.nan for _ in range(datacols)]
	if agg_by == "Count":
		for g in groups:
			grp_subset = subset_dict[g]
			for colidx in range(len(grp_subset)):
				vals = no_none(grp_subset[colidx])
				agg_data[g][colidx] = len(vals)
	elif agg_by == "Minimum":
		for g in groups:
			grp_subset = subset_dict[g]
			for colidx in range(len(grp_subset)):
				vals = no_none(grp_subset[colidx])
				agg_data[g][colidx] = min(vals) if len(vals) > 0 else math.nan
	elif agg_by == "Maximum":
		for g in groups:
			grp_subset = subset_dict[g]
			for colidx in range(len(grp_subset)):
				vals = no_none(grp_subset[colidx])
				agg_data[g][colidx] = max(vals) if len(vals) > 0 else math.nan
	elif agg_by == "Median":
		for g in groups:
			grp_subset = subset_dict[g]
			for colidx in range(len(grp_subset)):
				vals = no_none(grp_subset[colidx])
				agg_data[g][colidx] = statistics.median(vals) if len(vals) > 0 else math.nan
	elif agg_by == "Arithmetic mean":
		for g in groups:
			grp_subset = subset_dict[g]
			for colidx in range(len(grp_subset)):
				vals = no_none(grp_subset[colidx])
				agg_data[g][colidx] = statistics.mean(vals) if len(vals) > 0 else math.nan
	elif agg_by == "Geometric mean":
		for g in groups:
			grp_subset = subset_dict[g]
			for colidx in range(len(grp_subset)):
				vals = no_none(grp_subset[colidx])
				try:
					agg_data[g][colidx] = spstats.gmean(vals) if len(vals) > 0 else math.nan
				except:
					agg_data[g][colidx] = math.nan
					agg_error = True
					warning("Cannot calculate a geometric mean with negative values", {"parent": self.dlg})
	elif agg_by == "Harmonic mean":
		for g in groups:
			grp_subset = subset_dict[g]
			for colidx in range(len(grp_subset)):
				vals = no_none(grp_subset[colidx])
				try:
					agg_data[g][colidx] = spstats.hmean(vals) if len(vals) > 0 else math.nan
				except:
					agg_data[g][colidx] = math.nan
					agg_error = True
					warning("Cannot calculate a harmonic mean with negative values", {"parent": self.dlg})
	elif agg_by == "Sum":
		for g in groups:
			grp_subset = subset_dict[g]
			for colidx in range(len(grp_subset)):
				vals = no_none(grp_subset[colidx])
				agg_data[g][colidx] = sum(vals) if len(vals) > 0 else np.nan
	else:
		agg_error = True
		warning(f"Unrecognized aggregation method: {agg_by}.", {"parent": self.dlg})
	return agg_error, list(map(list, zip(*[([k]+agg_data[k]) for k in agg_data.keys()])))

def dataset_contains(ds, test_fn):
	# Returns True or False to indicate whether the 'test_fn' is true for any member of 'ds'.
	# 'ds' is a list of lists.
	contains = False
	for x in ds:
		for y in x:
			if test_fn(y):
				contains = True
				break
	return contains

def treeview_sort_column(tv, col, reverse):
    colvals = [(tv.set(k, col), k) for k in tv.get_children()]
    try:
    	colvals.sort(key=lambda v: float(v[0]), reverse=reverse)
    except ValueError:
    	try:
    		colvals.sort(key=key_numstr, reverse=reverse)
    	except TypeError:
    		colvals.sort(reverse=reverse)
    # Rearrange items in sorted positions
    for index, (val, k) in enumerate(colvals):
        tv.move(k, '', index)
    # Reverse sort next time
    tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))

def set_tv_headers(tvtable, column_headers, colwidths, charpixels):
	pixwidths = [charpixels * col for col in colwidths]
	for i in range(len(column_headers)):
		hdr = column_headers[i]
		tvtable.column(hdr, width=pixwidths[i])
		tvtable.heading(hdr, text=hdr, command=lambda _col=hdr: treeview_sort_column(tvtable, _col, False))

def fill_tv_table(tvtable, rowset, status_label=None):
	for i, row in enumerate(rowset):
		enc_row = [c if c is not None else '' for c in row]
		tvtable.insert(parent='', index='end', iid=str(i), values=enc_row)
	if status_label is not None:
		status_label.config(text = "    %d rows" % len(rowset))

def treeview_table(parent, rowset, column_headers, select_mode="none", nrows=None):
	# Creates a TreeView table containing the specified data, with scrollbars and status bar
	# in an enclosing frame.
	# This does not grid the table frame in its parent widget.
	# Returns a tuple of 0: the frame containing the table,  and 1: the table widget itself.
	nrows = range(len(rowset))
	ncols = range(len(column_headers))
	hdrwidths = [len(column_headers[j]) for j in ncols]
	if len(rowset) > 0:
		datawidthtbl = [[len(rowset[i][j] if isinstance(rowset[i][j], str) else str(rowset[i][j])) for i in nrows] for j in ncols]
		#datawidthtbl = [[len(str(rowset[i][j])) for i in nrows] for j in ncols]
		datawidths = [max(cwidths) for cwidths in datawidthtbl]
		colwidths = [max(hdrwidths[i], datawidths[i]) for i in ncols]
	else:
		#datawidths = hdrwidths
		colwidths = hdrwidths
	#colwidths = [max(hdrwidths[i], datawidths[i]) for i in ncols]
	# Set the font.
	ff = tkfont.nametofont("TkFixedFont")
	tblstyle = ttk.Style()
	tblstyle.configure('tblstyle', font=ff)
	charpixels = int(1.3 * ff.measure(u"0"))
	tableframe = ttk.Frame(master=parent, padding="3 3 3 3")
	statusframe = ttk.Frame(master=tableframe)
	# Create and configure the Treeview table widget
	tv_widget = ttk.Treeview(tableframe, columns=column_headers, selectmode=select_mode, show="headings")
	tv_widget.configure()["style"] = tblstyle
	if nrows is not None:
		tv_widget.configure()["height"] = nrows
	ysb = ttk.Scrollbar(tableframe, orient='vertical', command=tv_widget.yview)
	xsb = ttk.Scrollbar(tableframe, orient='horizontal', command=tv_widget.xview)
	tv_widget.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
	# Status bar
	parent.statusbar = ttk.Label(statusframe, text="    %d rows" % len(rowset), relief=tk.RIDGE, anchor=tk.W)
	tableframe.statuslabel = parent.statusbar
	# Fill the Treeview table widget with data
	set_tv_headers(tv_widget, column_headers, colwidths, charpixels)
	fill_tv_table(tv_widget, rowset, parent.statusbar)
	# Place the table
	tv_widget.grid(column=0, row=0, sticky=tk.NSEW)
	ysb.grid(column=1, row=0, sticky=tk.NS)
	xsb.grid(column=0, row=1, sticky=tk.EW)
	statusframe.grid(column=0, row=3, sticky=tk.EW)
	tableframe.columnconfigure(0, weight=1)
	tableframe.rowconfigure(0, weight=1)
	# Place the status bar
	parent.statusbar.pack(side=tk.BOTTOM, fill=tk.X)
	#
	return tableframe, tv_widget

def add_tv_column(tv_table, column_name, item_ids, data_values):
	# tv_table: the widget identifier; column_name: text of the new column name;
	# item_ids: list of Treeview row identifiers; data_values: data for the new
	# column, in the same order as the row identifiers.
	data_strs = [str(d) if d is not None else '' for d in data_values]
	datawidths = [len(v) for v in data_strs]
	datawidths.append(len(column_name))
	colwidth = max(datawidths)
	ff = tkfont.nametofont("TkFixedFont")
	tblstyle = ttk.Style()
	tblstyle.configure('tblstyle', font=ff)
	charpixels = int(1.3 * ff.measure(u"0"))
	colhdrs = list(tv_table["columns"])
	colwidths = [tv_table.column(col, 'width') for col in colhdrs]
	hdr_attrs = {col: tv_table.heading(col) for col in colhdrs}
	for col in hdr_attrs:
		del hdr_attrs[col]["state"]
	hdr_attrs[column_name] = {"text": column_name, "command": lambda _col=column_name: treeview_sort_column(tv_table, _col, False)}
	colhdrs.append(column_name)
	colwidths.append(colwidth*charpixels)
	tv_table["columns"] = tuple(colhdrs)
	for i, col in enumerate(colhdrs):
		tv_table.column(col, width = colwidths[i], stretch=tk.NO)
	for col in hdr_attrs:
		tv_table.heading(col, **hdr_attrs[col])
	allrowids = tv_table.get_children()
	if len(item_ids) < len(allrowids):
		for i in range(len(allrowids)):
			tv_table.set(allrowids[i], column=column_name, value='')
	for i in range(len(data_strs)):
		tv_table.set(item_ids[i], column=column_name, value=data_strs[i])



def export_data_table(headers, rows, sheetname="mapdata_export"):
	# Exports the selected data to a CSV or ODS file.  Returns the file name or None if canceled.
	outfile = tkfiledialog.asksaveasfilename(title="File name for saved data rows",
		filetypes=[('CSV files', '.csv'), ('ODS files', '.ods'), ('TSV files', '.tsv'), ('Plain text', '.txt'), ('LaTeX', '.tex')])
	if outfile:
		if outfile[-3:].lower() == 'csv':
			write_delimited_file(outfile, "csv", headers, rows)
		elif outfile[-3:].lower() == 'tsv':
			write_delimited_file(outfile, "tsv", headers, rows)
		elif outfile[-3:].lower() == 'txt':
			write_delimited_file(outfile, "plain", headers, rows)
		elif outfile[-3:].lower() == 'tex':
			write_delimited_file(outfile, "tex", headers, rows)
		elif outfile[-3:].lower() == 'ods':
			export_ods(outfile, headers, rows, append=True, sheetname=sheetname)
		else:
			# Force write as CSV.
			outfile = outfile + ".csv"
			write_delimited_file(outfile, "csv", headers, rows)
	return outfile

def dquote(v):
	# Returns a double-quoted value if it is not an identifier.
	if not v.isidentifier():
		return '"%s"' % v
	else:
		return v

def squotelist(vv):
	return ["'" + v.replace("'", "''") + "'" for v in vv]

def db_colnames(tbl_hdrs):
	# Takes a list of table headers and returns a list of database column names,
	# with double-quoting of any name that is not all alphanumeric and starts with
	# an alphabetic.
	colnames = []
	for hdr in tbl_hdrs:
		h2 = re.subn(r'\W', '_', hdr)[0]
		if not h2[0].isalpha() and h2[0] != '_':
			h2 = '_' + h2
		if h2.lower() in sql_kw:
			h2 = h2 + '_'
		colnames.append(h2)
	return colnames

def db_colnamestr(tbl_hdrs):
	return ",".join([c for c in db_colnames(tbl_hdrs)])

def isint(v):
	# Missing values match and will be handled by 'conv_int()'
	if v is None or (type(v) is str and v.strip() == ''):
		return True
	if type(v) == int:
		return True
	if type(v) == float:
		return False
	try:
		int(v)
		return True
	except ValueError:
		return False
	except TypeError:
		return False

def conv_int(v):
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	try:
		rv = int(v)
		return rv
	except:
		return None

def isfloat(v):
	# Missing values match and will be handled by 'conv_float()'
	if v is None or (type(v) is str and v.strip() == ''):
		return True
	try:
		float(v)
		return True
	except ValueError:
		return False
	except TypeError:
		return False

def conv_float(v):
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	try:
		rv = float(v)
		return rv
	except:
		return None

def isboolean(v):
	return parse_boolean(v) is not None

def conv_datetime(v):
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	try:
		rv = parse_datetime(v)
	except:
		try:
			d = parse_date(v)
			rv = datetime.datetime.combine(d, datetime.datetime.min.time())
		except:
			return None
	return rv

def conv_none(v):
    return None


def dt_type(v):
	# Type of date/time: timestamp, date, or None
	if type(v) is str and v.strip() == '':
		v = None
	if parse_date(v):
		return "date"
	if parse_datetime(v):
		return "timestamp"
	if parse_datetimetz(v):
		return "timestamptz"
	return None

def data_type(v):
	# Characterizes the value v as one of a simple set of data types.
	# Returns "timestamp", "date", "timestamptz", "int", "float", "boolean", or "string"
	if v is None or (type(v) is str and v.strip() == ''):
		return None
	if isint(v):
		return "int"
	if isfloat(v):
		return "float"
	dt = dt_type(v)
	if dt is not None:
		return dt
	if isboolean(v):
		return "boolean"
	return "string"

def conv_vect_dt(vect, v_type):
	# Convert all values in the vector to date, timestamp, timestamptz, or strings as specified.
	if v_type == "date":
		try:
			return [parse_date(v) for v in vect]
		except:
			return [str(v) for v in vect]
	elif v_type == "timestamp":
		try:
			return [parse_datetime(v) for v in vect]
		except:
			return [str(v) for v in vect]
	elif v_type == "timestamptz":
		try:
			return [parse_datetimetz(v) for v in vect]
		except:
			return [str(v) for v in vect]
	else:
		return [str(v) for v in vect]



# Lookup table for priorities among data types
data_type_pair_priorities = {
		"int" : {"int":"int", "float":"float", "date":"string", "timestamp":"string", "timestamptz":"string", "boolean":"string", "string":"string", None:"int"},
		"float" : {"int":"float", "float":"float", "date":"string", "timestamp":"string", "timestamptz":"string", "boolean":"string", "string":"string", None:"float"},
		"date" : {"int":"string", "float":"string", "date":"date", "timestamp":"timestamp", "timestamptz":"string", "boolean":"string", "string":"string", None:"date"},
		"timestamp" : {"int":"string", "float":"string", "date":"timestamp", "timestamp":"timestamp", "datetimetz":"string", "boolean":"string", "string":"string", None:"timestamp"},
		"timestamptz" : {"int":"string", "float":"string", "date":"string", "timestamp":"string", "timestamptz":"timestamptz", "boolean":"string", "string":"string", None:"timestamptz"},
		"string" : {"int":"string", "float":"string", "date":"string", "timestamp":"string", "timestamptz":"string", "boolean":"string", "string":"string", None:"string"},
		"boolean" : {"int":"string", "float":"string", "date":"string", "timestamp":"string", "timestamptz":"string", "boolean":"boolean", "string":"string", None:"boolean"},
		None : {"int":"int", "float":"float", "date":"date", "timestamp":"timestamp", "timestamptz":"timestamptz", "boolean":"boolean", "string":"string", None:None}
		}
def priority_data_type(dt1, dt2):
	return data_type_pair_priorities[dt1][dt2]


def data_type_cast_fn(data_type_str):
	# Return functions to convert text representations into Python data types.
	# Text representations may come from input data or SQLite.
	if data_type_str is None or data_type_str == "None":
		return conv_none
	if data_type_str == "string":
		return str
	elif data_type_str == "date":
		return parse_date
	elif data_type_str == "timestamp":
		return conv_datetime
	elif data_type_str == "timestamptz":
		return parse_datetimetz
	elif data_type_str == "int":
		return conv_int
	elif data_type_str == "float":
		return conv_float
	elif data_type_str == "boolean":
		return parse_boolean

def datetime_to_iso(dt):
	# An adapter to convert date and datetime values to strings for storage in SQLite.
	if dt is None:
		return None
	else:
		return dt.isoformat()

def datestr_to_sqlite(val):
	return datetime_to_iso(parse_date(val))

def timestamp_to_sqlite(val):
	return datetime_to_iso(conv_datetime(val))

def timestamptz_to_sqlite(val):
	return datetime_to_iso(parse_datetimetz(val))

def data_type_cast_sqlite_fn(data_type_str):
	if data_type_str is None or data_type_str == "None":
		return conv_none
	elif data_type_str == "date":
		return datestr_to_sqlite
	elif data_type_str == "timestamp":
		return timestamp_to_sqlite
	elif data_type_str == "timestamptz":
		return datetimetz_to_sqlite
	else:
		return str

def set_data_types_core(headers, rows):
	# Column-by-column processing is slightly faster than row-by-row processing.
	coltypes = []
	for i, colname in enumerate(headers):
		datavals = [row[i] for row in rows]
		dt = None
		for d in datavals:
			ndt = data_type(d)
			if ndt != dt:
				dt = priority_data_type(ndt, dt)
				if dt == "string":
					break
		non_null = [d for d in datavals if d is not None and not (type(d) is str and d.strip() == '')]
		nullcount = len(datavals) - len(non_null)
		uniquevals = len(set(non_null))
		if dt is None:
			dt = "string"
		coltypes.append((colname, dt, nullcount, uniquevals))
	return coltypes

def set_data_types(headers, rows, q):
	q.put(set_data_types_core(headers, rows))


def round_figs(x, figs=3):
	if x == 0.0 or x == -0.0 or x == float('inf') or x == float('-inf'):
		return 0.0
	elif math.isnan(x):
		return x
	else:
		return round(x, figs - int(math.floor(math.log10(abs(x)))) - 1)

def fp_display(x, figs=3):
	if x is None or math.isnan(x):
		return "NC"
	if type(x) == str:
		return x
	if x < 0.0005:
		return "%.2E" % x
	else:
		return str(round_figs(x, figs))


def center_window(win, x_offset=0, y_offset=0):
	win.update_idletasks()
	m = re.match(r"(\d+)x(\d+)\+(-?\d+)\+(-?\d+)", win.geometry())
	if m is not None:
		wwd = int(m.group(1))
		wht = int(m.group(2))
		swd = win.winfo_screenwidth()
		sht = win.winfo_screenheight()
		xpos = (swd/2) - (wwd/2) + x_offset
		ypos = (sht/2) - (wht/2) + y_offset
		win.geometry("%dx%d+%d+%d" % (wwd, wht, xpos, ypos))

def raise_window(win):
	win.attributes('-topmost', 1)
	win.attributes('-topmost', 0)

def shift_window(win, x_offset=0, y_offset=0):
	win.update_idletasks()
	m = re.match(r"(\d+)x(\d+)\+(-?\d+)\+(-?\d+)", win.geometry())
	if m is not None:
		xpos = int(m.group(1)) + x_offset
		ypos = int(m.group(2)) + y_offset
		win.geometry("+%d+%d" % (xpos, ypos))

# Inverse normal CDF function (qnorm) by Acklam's algorithm
A = (-3.969683028665376e+01, 2.209460984245205e+02, -2.759285104469687e+02,
		1.383577518672690e+02, -3.066479806614716e+01, 2.506628277459239e+00)
B = (-5.447609879822406e+01, 1.615858368580409e+02, -1.556989798598866e+02,
		6.680131188771972e+01, -1.328068155288572e+01)
C = (-7.784894002430293e-03, -3.223964580411365e-01, -2.400758277161838e+00,
		-2.549732539343734e+00, 4.374664141464968e+00, 2.938163982698783e+00)
D = (7.784695709041462e-03, 3.224671290700398e-01, 2.445134137142996e+00,
		3.754408661907416e+00)
P_LOW = 0.02425
P_HIGH = 1.0 - P_LOW
def qnorm(p):
	if p <= 0 or p >= 1.0:
		raise ValueError("Invalid input to qnorm()")
	if p >= P_LOW and p <= P_HIGH:
		q = p - 0.5
		r = q*q
		return (((((A[0]*r+A[1])*r+A[2])*r+A[3])*r+A[4])*r+A[5])*q / \
        (((((B[0]*r+B[1])*r+B[2])*r+B[3])*r+B[4])*r+1)
	elif p < P_LOW:
		q = math.sqrt(-2 * math.log(p))
		return (((((C[0]*q+C[1])*q+C[2])*q+C[3])*q+C[4])*q+C[5]) / \
         ((((D[0]*q+D[1])*q+D[2])*q+D[3])*q+1)
	else:
		q = math.sqrt(-2 * math.log(1.0 - p))
		return -(((((C[0]*q+C[1])*q+C[2])*q+C[3])*q+C[4])*q+C[5]) / \
          ((((D[0]*q+D[1])*q+D[2])*q+D[3])*q+1)


# Functions for calculation of Fisher-Jenks breaks
def ssd(values):
    m = statistics.mean(values)
    return sum([(x-m)**2 for x in values])

def all_jenks_breaks(d, max_groups):
    if max_groups > len(set(d)) - 2:
    	max_groups = len(set(d)) - 2
    d_ssd = ssd(d)
    groups = [1]
    gvf = [0]
    for i in range(2, max_groups+2):
        jnb = jenkspy.JenksNaturalBreaks(i)
        jnb.fit(d)
        grps = jnb.groups_
        g_ssd = [ssd(g) for g in grps]
        groups.append(i)
        gvf.append((d_ssd - sum(g_ssd)) / d_ssd)
    return groups, gvf

def slopes(xs, ys):
    return [(ys[i+1]-ys[i])/(xs[i+1]-xs[i]) for i in range(len(xs)-1)]

def reldiffs(v):
    return [(v[i] - v[i+1]) / ((v[i] + v[i+1])/2) for i in range(len(v)-1)]

def optimum_jenks(d, max_groups):
    # Returns the optimum number of Fisher-Jenks groups, >= 2.
	if len(set(d)) < 3:
		return [1]
	else:
		df = reldiffs(slopes(*all_jenks_breaks(d, max_groups)))
		imd = df.index(max(df))
		return list(range(2, max_groups+2))[imd]

# Functions for calculating categorical similarity
def cat_freqs(d):
	# 'd' is a list of values for a categorical variable, possibly including missing values.
	# Returns a dictionary where the keys are categorical values and the value is a tuple
	# containing the number of instances of that categorical value and the frequency.
	N = len(d)
	nnd = [v for v in d if v is not None]
	catvals = list(set(nnd))
	freqdict = {}
	for v in catvals:
		n = nnd.count(v)
		freqdict[v] = (n, n/N)
	return freqdict

def cat_sim_Lin(d, ids):
	# 'd' is a column-oriented table of categorical variables with the same number of rows
	# in every column and no missing values.
	# 'ids' is a list of unique row identifiers, with length equal to the number of rows in d.
	# Per Boriah et al. 2008
	pdict = [cat_freqs(col) for col in d]
	nrows = len(ids)
	similarities = []	# A row-oriented table with row values of row_id_1, row_id_2, and similarity.
	for r1 in range(nrows):
		for r2 in range(r1, nrows):
			sim_numerator = 0
			sim_denom = 0
			for catvar in range(len(d)):
				if d[catvar][r1] == d[catvar][r2]:
					sim_numerator += 2 * math.log(pdict[catvar][d[catvar][r1]][1])
				else:
					sim_numerator += 2 * math.log(pdict[catvar][d[catvar][r1]][1] + pdict[catvar][d[catvar][r2]][1])
				sim_denom += math.log(pdict[catvar][d[catvar][r1]][1]) + math.log(pdict[catvar][d[catvar][r2]][1])
			if sim_denom == 0.0:
				similarities.append([ids[r1], ids[r2], "NC"])
			else:
				similarities.append([ids[r1], ids[r2], abs(sim_numerator/sim_denom)])
	return similarities

def cat_sim_Goodall3(d, ids):
	# 'd' is a column-oriented table of categorical variables with the same number of rows
	# in every column and no missing values.
	# 'ids' is a list of unique row identifiers, with length equal to the number of rows in d.
	# Per Boriah et al. 2008
	pdict = [cat_freqs(col) for col in d]
	nrows = len(ids)
	similarities = []	# A row-oriented table with row values of row_id_1, row_id_2, and similarity.
	for r1 in range(nrows):
		for r2 in range(r1, nrows):
			sim_numerator = 0
			for catvar in range(len(d)):
				if d[catvar][r1] == d[catvar][r2]:
					sim_numerator += 1 - pdict[catvar][d[catvar][r1]][1]
			similarities.append([ids[r1], ids[r2], sim_numerator/len(d)])
	return similarities

def cat_sim_IOF(d, ids):
	# 'd' is a column-oriented table of categorical variables with the same number of rows
	# in every column and no missing values.
	# 'ids' is a list of unique row identifiers, with length equal to the number of rows in d.
	# Per Boriah et al. 2008
	pdict = [cat_freqs(col) for col in d]
	nrows = len(ids)
	similarities = []	# A row-oriented table with row values of row_id_1, row_id_2, and similarity.
	for r1 in range(nrows):
		for r2 in range(r1, nrows):
			sim_numerator = 0
			for varno in range(len(d)):
				if d[varno][r1] == d[varno][r2]:
					sim_numerator += 1
				else:
					sim_numerator += 1 / (1 + math.log(pdict[varno][d[varno][r1]][0]) * math.log(pdict[varno][d[varno][r2]][0]))
			similarities.append([ids[r1], ids[r2], sim_numerator/len(d)])
	return similarities

def cat_sim_OF(d, ids):
	# 'd' is a column-oriented table of categorical variables with the same number of rows
	# in every column and no missing values.
	# 'ids' is a list of unique row identifiers, with length equal to the number of rows in d.
	# Per Boriah et al. 2008
	pdict = [cat_freqs(col) for col in d]
	nrows = len(ids)
	similarities = []	# A row-oriented table with row values of row_id_1, row_id_2, and similarity.
	for r1 in range(nrows):
		for r2 in range(r1, nrows):
			sim_numerator = 0
			for catvar in range(len(d)):
				if d[catvar][r1] == d[catvar][r2]:
					sim_numerator += 1
				else:
					sim_numerator += 1 / (1 + math.log(nrows/pdict[catvar][d[catvar][r1]][0]) * math.log(nrows/pdict[catvar][d[catvar][r2]][0]))
			similarities.append([ids[r1], ids[r2], sim_numerator/len(d)])
	return similarities

def cat_sim_Overlap(d, ids):
	# 'd' is a column-oriented table of categorical variables with the same number of rows
	# in every column and no missing values.
	# 'ids' is a list of unique row identifiers, with length equal to the number of rows in d.
	# Per Boriah et al. 2008
	pdict = [cat_freqs(col) for col in d]
	nrows = len(ids)
	similarities = []	# A row-oriented table with row values of row_id_1, row_id_2, and similarity.
	for r1 in range(nrows):
		for r2 in range(r1, nrows):
			sim_numerator = 0
			for catvar in range(len(d)):
				if d[catvar][r1] == d[catvar][r2]:
					sim_numerator += 1
			similarities.append([ids[r1], ids[r2], sim_numerator/len(d)])
	return similarities

# Functions for NMF unmixing
def rowsums(a):
	# 'a' is a numpy array.
    return np.sum(a, axis=1)

def colsums(a):
	# 'a' is a numpy array.
    return rowsums(a.transpose())

def array_rss(m1, m2):
	# 'm1' and 'm2' are numpy arrays.
    r, c = m1.shape
    rss = 0.0
    for i in range(r):
        for j in range(c):
            rss += (m1[i][j] - m2[i][j])**2
    return rss

def array_rmse(m1, m2):
    return math.sqrt(mean_squared_error(m1, m2))

def W_rowsum_normalize(aW, aH):
    # Return modified versions of numpy arrays aW and aH so that all the row sums
    # of the modified aW are 1.0 and aH is adjusted so the product aW*aH is the same.
    rowsW = aW.shape[0]
    rowsums_W = rowsums(aW)
    revW = aW / rowsums_W[:, np.newaxis]
    scalefacts = colsums(aW) / colsums(revW)
    revH = aH * scalefacts[:, np.newaxis]
    return revW, revH

def sample_EMs(W, H):
    # Return a list of lists (not a numpy array) representing
    # a matrix of samples by EMs, where each cell contains the
    # effective concentration of an EM in a sample.
    # Both W and H should be numpy arrays produced by NMF.
    #return [np.matmul(s, H).tolist() for s in W]
    em_chemsum = rowsums(H)
    res = []
    for row in W:
    	res.append([row[i] * em_chemsum[i] for i in range(len(row))])
    return res

def list_mse(y):
    # This computes the MSE for differences between elements of the list and 1.0
    return sum([(s-1)**2 for s in y])/len(y)

def iterate_nmf(D, n_em, max_iterations=200, mse_goal=0.00001):
	# Performs NMF on D until the row sums of the W matrix are all (approximately) 1.0.
    nmf_model1 = NMF(n_components=n_em, solver='mu', max_iter=500)
    W = nmf_model1.fit_transform(D)
    H = nmf_model1.components_
    nmf_model = NMF(n_components=n_em, solver='mu', init='custom', max_iter=2000)
    last_mse = -1.0
    for _ in range(max_iterations):
        W = nmf_model.fit_transform(D, W=W, H=H)
        H = nmf_model.components_
        rsw = list(rowsums(W))
        mse = list_mse(rsw)
        if (mse < mse_goal or abs(mse - last_mse) < 0.0001) and max(rsw) <= 1.0:
            break
        else:
            last_mse = mse
            W, H = W_rowsum_normalize(W, H)
    return W, H, nmf_model.reconstruction_err_


# Functions for map markers

def set_sel_marker(symbol, color):
	select_marker = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
	mkr_key = "%s %s" % (color, symbol)
	if mkr_key not in custom_icons:
		custom_icons[mkr_key] = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
	select_marker = custom_icons[mkr_key]
	return select_marker

def set_get_loc_marker(location_marker, location_color):
	global map_settings
	mkr_key = "%s %s" % (location_color, location_marker)
	if mkr_key not in custom_icons:
		custom_icons[mkr_key] = tk.BitmapImage(data=icon_xbm[location_marker], foreground=location_color)
	return custom_icons[mkr_key]


# ==================================================================================
#		class MapUI
# The main UI element
# ==================================================================================
class MapUI(object):
	def __init__(self, args):
		# 'args' is an object with the following attributes:
		#	color			- String; the data table column with color names.
		#	database		- String: the name of the client-server database.
		#	db_type			- String: code to identify DBMS.
		#	file			- String: name of CSV or spreadsheet file to read.
		#	identifier		- String: the data table column with location identifiers.
		#	image			- String: the name of an image file to write (and then exit).
		#	imagewait		- Integer: the number of seconds to wait before writing an image file.
		#	lat				- String: the data table column with latitude values.
		#	lon				- String: the data table column with longitude values.
		#	message			- String: title for the MapUI window.
		#	no_passwd		- Boolean: whether to not prompt for a password for client-server databases.
		#	port			- Integer: port for client-server databases.
		#	projection		- Integer: the CRS for the lat/lon values in the data table.
		#	script			- String: the name of a SQL script file to run before importing a database table.
		#	server			- String: the name of the database server
		#	sheet			- String: the spreadsheet sheet to import.
		#	symbol			- String: the data table column with symbol names.
		#	table			- String: the database table to import.
		#	user			- String: the user name for client-server databases.
		#
		# Valid combinations/interpretations of arguments are:
		# 1. No arguments
		#    mapdata starts by presenting a prompt for the data source.
		# 2. File import (CSV or spreadsheet)
		#    file, lon, and lat are required; sheet may be required
		#    symbol, color, and identifier are optional
		#        sheet must be specified if the file is a spreadsheet
		#        If image and imagewait are specified, the map will be drawn and exported, and mapdata will exit..
		# 3. Server-based database import
		#    db_type, server, db_name, user, no_passwd, table, lon, and lat are required
		#    port, script, symbol, color, and identifier are optional.
		#        If port is missing, the default port for the specified db_type will be used.
		#        If script is specified, it will be run before table is imported.
		#        If no_passwd is False (the default), the user will be prompted for a password.
		# 4. File-based database import
		#    file, table, lon, and lat are required.
		#    script, symbol, color, and identifier are optional.
		#        If script is specified, it will be run before table is imported.
		# 5. Map image creation and export -- in addition to one of the previous sets of requirement.
		#    image and imagewait are required
		#
		# If file and db_type are both missing, the GUI prompt will be launched regardless of
		# any other arguments that have been provided.
		#
		self.args = args
		self.win = tk.Tk()
		self.win.withdraw()
		self.win.protocol("WM_DELETE_WINDOW", self.cancel)

		#=====  Display any warning and fatal error messages.
		if len(config_warning_messages) > 0:
			warning("\n".join(config_warning_messages), kwargs={'parent': self.win})
		if len(config_fatal_messages) > 0:
			fatal("\n".join(config_fatal_messages), kwargs={'parent': self.win})

		#=====  Tkinter style modifications
		# Set entry of readonly comboboxes to white instead of default grey
		self.ttkstyle = ttk.Style(self.win)
		self.ttkstyle.map('TCombobox', fieldbackground=[('disabled', 'lightgrey'), ('readonly', 'white')])

		#==== Check for inconsistent or invalid arguments
		if args.file is not None and not os.path.exists(args.file):
			fatal_error("File %s does not exist" % args.file, kwargs={'parent': self.win})
		# CSV, spreadsheet, or database file but also client-server database specifications
		if (args.file and any((args.port, args.server, args.database, args.user))) \
				or (args.file and args.db_type and args.sheet):
					fatal_error("Command-line arguments include specifications for both a file and a database.", kwargs={'parent': self.win})
		# Incomplete client-server database specification:
		if any((args.server, args.database)) and not all((args.server, args.database)):
			fatal_error("Incomplete client-server data source specifications on the command line.", kwargs={'parent': self.win})

		self.loading_dlg = LoadingDialog(self.win)
		self.loading_dlg.display("Preparing map")

		# Size and position window.
		self.win.geometry("1200x1000")
		center_window(self.win)

		# Establish default projection if none is specified.
		if args.projection is None:
			args.projection = 4326

		# References to data table, populated by add_data()
		self.lat_col = None
		self.lon_col = None
		self.src_lat_col = None
		self.src_lon_col = None
		self.lat_4326_col = None
		self.lon_4326_col = None
		self.lat_index = None
		self.lon_index = None
		self.src_lat_index = None
		self.src_lon_index = None
		self.label_col = None
		self.label_index = None
		self.symbol_index = None
		self.color_index = None

		# Table of data column names, types, and other information.  This is populated
		# by a separate process after the data table is loaded.
		self.data_types = None

		if args.file is not None:
			src_name = os.path.basename(args.file)
		else:
			src_name = args.table
		if args.message is not None:
			self.win.title(args.message)
		else:
			self.win.title("Map of %s" % src_name)
		self.data_src_name = src_name

		#====  Patch ImageTk.PhotoImage.__del__ 
		ImageTk.PhotoImage.__del__ = new_img_del

		#====  Map configuration
		self.mapfont = makefont(map_settings.label_font, map_settings.label_size, map_settings.label_bold)
		self.attribfont = makefont(map_settings.label_font, 7, False)
		# Set the application window icon
		#win_icon = tk.BitmapImage(data=win_icon_xbm, foreground="black", background="tan")
		#self.win.iconbitmap(win_icon)
		# The markers for all the locations in the data table
		self.loc_map_markers = []
		# The markers for the selected location(s)
		self.sel_map_markers = []
		# The markers used to label unique coordinates with the count of data rows
		self.count_label_markers = []
		map_settings.count_label_show = False
		# The number of table rows without coordinates
		self.missing_latlon = 0
		# Map bounds
		self.min_lat = None
		self.max_lat = None
		self.min_lon = None
		self.max_lon = None
		# List of FindCandKeysDialog objects, so they can be told to update themselves
		self.candcol_list = []
		# List of FindDupRowsDialog objects, so they can be told to update themselves
		self.duprows_list = []
		# List of CardinalityTestDialog objects, so they can be told to update themselves
		self.cardinality_list = []
		# List of AggregateDialog objects, so they can be told to update themselves
		self.aggdlg_list = []
		# List of DistanceDialog objects, so they can be told to update themselves
		self.distdlg_list = []
		# List of PlotDialog objects, so they can be told to update themselves, or be deleted.
		self.plot_list = []
		# List of UnivarStatsDialog objects, so they can be updated.
		self.univar_list = []
		# List of BivarStatsDialog objects, so they can be updated.
		self.bivar_list = []
		# List of ANOVADialog objects, so they can be updated.
		self.anova_list = []
		# List of FitDistDialog objects, so they can be updated.
		self.fitdist_list = []
		# List of TSNEDialog objects, so they can be updated.
		self.tsne_list = []
		# List of UMAPDialog objects, so they can be updated.
		self.umap_list = []
		# List of CorrMatrixDialog objects, so they can be updated.
		self.corrmat_list = []
		# List of CosineSimilarityDialog objects, so they can be updated.
		self.cosmat_list = []
		# List of CategCorrespDialog objects, so they can be updated.
		self.catcorresp_list = []
		# List of ContTableDialog objects, so they can be updated.
		self.conttable_list = []
		# List of CategSimilarityDialog objects, so they can be updated.
		self.simmat_list = []
		# List of PCADialog objects, so they can be updated.
		self.pca_list = []
		# List of NMFUnmixingDialog objects, so they can be updated.
		self.unmixing_list = []
		# List of ROCCurveDialog objects, so they can be updated.
		self.roccurve_list = []
		# Create default markers for the map
		self.loc_marker_icon = set_get_loc_marker(map_settings.location_marker, map_settings.location_color)
		# Initializes selection marker to the global settings
		self.sel_marker_icon = set_sel_marker(map_settings.select_symbol, map_settings.select_color)
		# Fixed 'null' icon for count labels
		self.count_label_icon = tk.BitmapImage(data=null_xbm, foreground="white")
		# Create icons for the buttonbar
		expand_icon = tk.BitmapImage(data=expand_xbm, foreground="black")
		focus_icon = tk.BitmapImage(data=wedge_sm_xbm, foreground="red")
		zoom_sel_icon = tk.BitmapImage(data=wedges_3_xbm, foreground="red")
		unselect_icon = tk.BitmapImage(data=cancel_xbm, foreground="black")
		# Use stacked frames for the main application window components.  Map and table in a PanedWindow.
		msgframe = ttk.Frame(self.win, padding="3 2")
		ctrlframe = ttk.Frame(self.win, padding="3 2")
		datapanes = ttk.PanedWindow(self.win, orient=tk.VERTICAL)
		mapframe = ttk.Frame(datapanes, borderwidth=2, relief=tk.RIDGE)
		self.tblframe = ttk.Frame(datapanes, padding="3 2")
		datapanes.add(mapframe, weight=1)
		datapanes.add(self.tblframe, weight=1)
		# Allow vertical resizing of map and table frames, not of message and control frames
		self.win.columnconfigure(0, weight=1)
		self.win.rowconfigure(0, weight=0)		# msgframe
		self.win.rowconfigure(1, weight=0)		# ctrlframe
		self.win.rowconfigure(2, weight=1)		# datapanes
		# Grid all the main frames
		msgframe.grid(row=0, column=0, sticky=tk.NSEW)
		ctrlframe.grid(row=1, column=0, sticky=tk.W)
		datapanes.grid(row=2, column=0, sticky=tk.NSEW)
		# Populate the message frame
		self.msg_label = ttk.Label(msgframe, text=args.message or "Map display")
		def wrap_msg(event):
			self.msg_label.configure(wraplength=event.width - 5)
		self.msg_label.bind("<Configure>", wrap_msg)
		self.msg_label.grid(column=0, row=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		msgframe.columnconfigure(0, weight=1)
		msgframe.rowconfigure(0, weight=1)
		# Populate the map control frame
		ctrlframe.rowconfigure(0, weight=0)
		ctrlframe.columnconfigure(0, weight=0)
		# Basemap controls and buttons
		self.basemap_label = ttk.Label(ctrlframe, text="Basemap:", anchor="w")
		self.basemap_label.grid(row=0, column=0, padx=(5, 5), pady=(2, 5), sticky=tk.W)
		global initial_basemap
		bm_name = initial_basemap
		if bm_name not in bm_servers:
			bm_name = tuple(bm_servers.keys())[0]
			initial_basemap = bm_name
		self.basemap_var = tk.StringVar(self.win, bm_name)
		self.map_option_menu = ttk.Combobox(ctrlframe, state="readonly", textvariable=self.basemap_var,
				values=self.available_tile_servers(), width=22)
		self.map_option_menu["state"] = tk.NORMAL
		self.map_option_menu.bind('<<ComboboxSelected>>', self.change_basemap)
		self.map_option_menu.grid(row=0, column=1, padx=(5, 30), pady=(2, 5), sticky=tk.W)
		# Multi-select option
		def ck_changed():
			ck = self.multiselect_var.get()
			if ck == '0':
				self.unselect_map()
				self.tbl.configure(selectmode = tk.BROWSE)
			else:
				self.tbl.configure(selectmode = tk.EXTENDED)
			self.set_status()
		# The multilselect setting
		self.multiselect_var = tk.StringVar(self.win, multiselect)
		ck_multiselect = ttk.Checkbutton(ctrlframe, text="Multi-select", variable=self.multiselect_var, command=ck_changed)
		ck_multiselect.grid(row=0, column=2, sticky=tk.W, padx=(0, 20))
		# Map control buttons
		zoomsel_btn = ttk.Button(ctrlframe, text="Zoom selected", image=zoom_sel_icon, compound=tk.LEFT, command=self.zoom_selected)
		zoomsel_btn.grid(row=0, column=3, sticky=tk.W)
		zoomsel_btn.image = zoom_sel_icon
		expand_btn = ttk.Button(ctrlframe, text="Zoom full", image=expand_icon, compound=tk.LEFT, command=self.zoom_full)
		expand_btn.image = expand_icon
		expand_btn.grid(row=0, column=4, sticky=tk.W)
		focus_btn = ttk.Button(ctrlframe, text="Center", image=focus_icon, compound=tk.LEFT, command=self.focus_map)
		focus_btn.image = focus_icon
		focus_btn.grid(row=0, column=5, sticky=tk.W)
		unselect_btn = ttk.Button(ctrlframe, text="Un-select", image=unselect_icon, compound=tk.LEFT, command=self.unselect_map)
		unselect_btn.image = unselect_icon
		unselect_btn.grid(row=0, column=6, sticky=tk.W)
		# Map widget
		mapframe.rowconfigure(0, weight=1)
		mapframe.columnconfigure(0, weight=1)
		self.map_widget = tkmv.TkinterMapView(mapframe, height=600, width=600, corner_radius=0)
		# Set the map center to the Gulf of Guinea instead of Brandenburg Tor while data are loading
		self.map_widget.set_position(0.0, 0.0)

		# Patch the tkintermapview right-click action to do nothing--it will be replaced.
		# Just a 'bind' command with no command argument does not un-bind the click.
		def do_nothing(args=None):
			pass
		if sys.platform == "darwin":
			self.map_widget.canvas.bind("<Button-2>", do_nothing)
		else:
			self.map_widget.canvas.bind("<Button-3>", do_nothing)

		# Patch the tkintermapview position marker to show location IDs on a right-click.
		global show_location_table
		show_location_table = self.show_location_ids

		# Set basemap
		bm_name = initial_basemap or "OpenMapServer"
		tileserver = self.tile_url(bm_name)
		self.map_widget.set_tile_server(tileserver, max_zoom=20)
		self.map_widget.grid(row=0, column=0, sticky=tk.NSEW)
		self.attrib_frame = tk.Frame(mapframe)
		self.attrib_frame.grid(row=0, column=0, sticky=tk.S+tk.E)
		self.set_attrib_labels(map_attributions[bm_name])

		# Remove the splash screen message and show the main UI.
		self.loading_dlg.hide()
		# Re-setting the size is necessary on Windows in at least some environments.
		self.win.geometry("1200x1000")
		self.win.deiconify()

		#==== Import data
		# Prompt if command-line arguments are missing
		if not (args.file or args.table):
			sdsd = SelDataSrcDialog(parent=self.win, mapui=self)
			src_name, label_col, lat_col, lon_col, crs, symbol_col, color_col, message, headers, rows = sdsd.select()
			if src_name is None:
				self.cancel()
		else:
			if args.lat is None or args.lon is None:
				warning("Latitude and longitude columns must be specified with other data source arguments.", kwargs={'parent': self.win})
				sdsd = SelDataSrcDialog(parent=self.win, mapui=self)
				src_name, label_col, lat_col, lon_col, crs, symbol_col, color_col, message, headers, rows = sdsd.select()
				if src_name is None:
					self.cancel()
			else:
				self.loading_dlg.display("Loading data")
				# src_name is a filename (either CSV or spreadsheet) or a database table name
				if args.file is not None:
					src_name = os.path.basename(args.file)
				else:
					src_name = args.table
				label_col, lat_col, lon_col, crs = args.id, args.lat, args.lon, args.projection
				symbol_col, color_col = args.symbol, args.color
				message  = args.message
				if args.file is not None and args.db_type is None:
					# CSV file or spreadsheet file
					fn, ext = os.path.splitext(args.file)
					if ext.lower() == ".csv":
						try:
							headers, rows = file_data(args.file, 0)
						except:
							self.loading_dlg.hide_all()
							fatal_error("Could not read data from %s" % src_data, kwargs={'parent': self.win})
					else:
						if args.sheet is None:
							self.loading_dlg.hide_all()
							fatal_error("A sheet name must be specified for spreadsheets", kwargs={'parent': self.win})
						try:
							if ext.lower() == '.ods':
								headers, rows = ods_data(src_name, sheet, 0)
							else:
								headers, rows = xls_data(src_name, sheet, 0)
						except:
							self.loading_dlg.hide_all()
							fatal_error("Could not read table from %s, sheet %s" % (args.file, args.sheet), kwargs={'parent': self.win})
				else:
					# Client-server or file-based database
					# Prompt for password only for client-server databases
					if args.script and not os.path.exists(args.script):
						fatal_error(f"The SQL script file {args.script} does not exist.", kwargs={'parent': self.win})
					if not args.no_passwd and args.user and not args.file:
						dlg = OneEntryDialog(self.win, "Password", f"Enter the password for {args.user}, database {args.database}:", obscure=True)
						self.loading_dlg.hide()
						pw = dlg.show()
						if pw is None:
							self.loading_dlg.hide_all()
							fatal_error(f"Password entry failed for user {args.user}, database {args.database}.", kwargs={'parent': self.win})
						self.loading_dlg.display("Loading data")
					else:
						pw = None
					db = open_dbms(dbms_name_codes[args.db_type], args.server, args.database, args.user, not args.no_passwd, pw, args.port, args.file)
					if db is None:
						self.loading_dlg.hide_all()
						fatal_error(f"Could not open database {args.database}.", kwargs={'parent': self.win})
					if args.script is not None:
						try:
							runscriptfile(db, args.script)
						except ErrInfo as e:
							self.loading_dlg.hide()
							fatal_error(e.eval_err(), kwargs={'parent':self.win})
						except:
							self.loading_dlg.hide_all()
							fatal_error(f"Could not run script file {args.script}", kwargs={'parent': self.win})
					try:
						headers, rows = db.select_data(f"select * from {args.table};")
					except ErrInfo as e:
						self.loading_dlg.hide_all()
						fatal_error(e.eval_err(), kwargs={'parent': self.win})
					except:
						self.loading_dlg.hide_all()
						fatal_error(f"Could not import data from table {args.table}", kwargs={'parent': self.win})
					db.close()
					if lat_col not in headers or lon_col not in headers:
						self.loading_dlg.hide_all()
						fatal_error("The specified latitude and longitude columns are not in the table", kwargs={'parent': self.win})
					if label_col is not None and label_col not in headers:
						self.loading_dlg.hide_all()
						fatal_error("The specified labeling column is not in the table", kwargs={'parent': self.win})
					if symbol_col is not None and symbol_col not in headers:
						self.loading_dlg.hide_all()
						fatal_error("The specified symbol column is not in the table", kwargs={'parent': self.win})
					if color_col is not None and color_col not in headers:
						self.loading_dlg.hide_all()
						fatal_error("The specified color column is not in the table", kwargs={'parent': self.win})
				self.loading_dlg.hide()

		self.msg_label.config(text=message)

		# Re-set the window title if a custom message is not specified
		if args.message is None:
			self.win.title("Map of %s" % src_name)

		# Source and possibly un-projected crs
		self.src_crs = crs
		self.crs = crs

		# Check that the specified column names are valid.
		if not (lat_col in headers and lon_col in headers):
			fatal_error("The specified coordinate columns are not in the data table.", kwargs={"parent": self.win})
		if label_col is not None and label_col != '' and not (label_col in headers):
			fatal_error("The label column is not in the data table.", kwargs={"parent": self.win})
		if symbol_col is not None and symbol_col != '' and not (symbol_col in headers):
			fatal_error("The symbol type column is not in the data table.", kwargs={"parent": self.win})
		if color_col is not None and color_col != '' and not (color_col in headers):
			fatal_error("The symbol color column is not in the data table.", kwargs={"parent": self.win})

		# Populate the table frame
		self.loading_dlg.display("Populating data table")
		self.tblframe.rowconfigure(0, weight=1)
		self.tblframe.columnconfigure(0, weight=1)
		try:
			self.tableframe, self.tbl = self.add_data(rows, headers, lat_col, lon_col, label_col, symbol_col, color_col)
		except:
			self.loading_dlg.hide_all()
			fatal_error("Cannot load data.  Check latitude, longitude, and CRS values; column counts; and headers.", kwargs={'parent': self.win})
		self.tableframe.grid(column=0, row=0, sticky=tk.NSEW)
		self.set_tbl_selectmode()
		self.set_status()
		# Add menu
		self.add_menu()
		self.map_widget.canvas.bind("<Control-ButtonRelease-1>", self.delete_selected)
		self.tbl.bind('<ButtonRelease-1>', self.mark_map)
		# Other key bindings
		self.win.protocol("WM_DELETE_WINDOW", self.cancel)
		self.win.bind("<Alt-q>", self.run_query)
		self.win.bind("<Alt-c>", self.showhide_count_labels)
		self.win.bind("<Alt-g>", self.new_plot)
		# Limit resizing
		self.win.minsize(width=860, height=640)
		# Set table status message
		self.set_status()
		# Just export the map and quit?
		if args.imagefile is not None:
			self.imageoutputfile = args.imagefile
			self.win.after(args.imagewait * 1000, self.export_map_and_quit)
		self.loading_dlg.hide()
	def available_tile_servers(self):
		# Return a list of those with attributions and either without API keys or for which API keys are provided
		avail = []
		for k in bm_servers:
			if k in map_attributions:
				if self.tile_url(k) is not None:
					avail.append(k)
		return avail
	def tile_url(self, source_name):
		# Return the URL with the API key replaced, unless it is not available.
		source_url = bm_servers[source_name]
		if "<api_key>" in source_url.lower():
			if source_name in api_keys:
				api_key = api_keys[source_name]
				for matched in re.findall("<api_key>", source_url, re.IGNORECASE):
					source_url = source_url.replace(matched, api_key)
				return source_url
			else:
				return None
		else:
			return source_url
	def mark_map(self, event):
		# Highlight the selected row(s) in the table and get the coordinates to map it
		if self.tbl.selection():
			new_markers = []
			for sel_row in self.tbl.selection():
				rowdata = self.tbl.item(sel_row)["values"]
				try:
					lat_val = float(rowdata[self.lat_index])
				except:
					lat_val = None
				try:
					lon_val = float(rowdata[self.lon_index])
				except:
					lon_val = None
				if lon_val is not None and lat_val is not None:
					new_marker = self.map_widget.set_marker(lat_val, lon_val, icon=self.sel_marker_icon)
					new_markers.append(new_marker)
			for m in self.sel_map_markers:
				m.delete()
			self.sel_map_markers = new_markers
		else:
			for m in self.sel_map_markers:
				m.delete()
		self.update_plot_data()
		self.set_status()
	def redraw_sel_markers(self):
		new_markers = []
		for mkr in self.sel_map_markers:
			mposition = mkr.position
			micon = mkr.icon
			mkr.delete()
			new_marker = self.map_widget.set_marker(mposition[0], mposition[1], icon=micon)
			new_markers.append(new_marker)
		self.sel_map_markers = new_markers
	def draw_sel_markers(self):
		for mkr in self.sel_map_markers:
			mkr.draw()
	def redraw_count_labels(self):
		if map_settings.count_label_show:
			for mkr in self.count_label_markers:
				mkr.draw()
		else:
			for mkr in self.count_label_markers:
				if mkr.canvas_text is not None:
					self.map_widget.canvas.delete(mkr.canvas_text)
	def redraw_loc_markers(self, tdata):
		# tdata is the treeview control containing the data table.
		self.loading_dlg.display("Redrawing markers")
		while len(self.loc_map_markers) > 0:
			self.loc_map_markers.pop().delete()
		self.draw_loc_markers(tdata)
		self.loading_dlg.hide()
	def draw_loc_markers(self, tdata):
		# tdata is the treeview control containing the data table.
		# Also set the number of rows missing coordinates and the bounding box.
		self.loading_dlg.display("Preparing map")
		self.missing_latlon = 0
		for row_id in tdata.get_children():
			rowdata = tdata.item(row_id)["values"]
			try:
				lat_val = float(rowdata[self.lat_index])
			except:
				lat_val = None
			try:
				lon_val = float(rowdata[self.lon_index])
			except:
				lon_val = None
			if lon_val is not None and lat_val is not None:
				if self.min_lat is None or lat_val < self.min_lat:
					self.min_lat = lat_val
				if self.max_lat is None or lat_val > self.max_lat:
					self.max_lat = lat_val
				if self.min_lon is None or lon_val < self.min_lon:
					self.min_lon = lon_val
				if self.max_lon is None or lon_val > self.max_lon:
					self.max_lon = lon_val
				if self.color_index is None and self.symbol_index is None:
					marker_icon = self.loc_marker_icon
				else:
					if self.color_index is None or not map_settings.use_data_color:
						color = map_settings.location_color
					else:
						color = rowdata[self.color_index].lower()
						if color not in color_names:
							color = map_settings.location_color
					if self.symbol_index is None or not map_settings.use_data_marker:
						symbol = map_settings.location_marker
					else:
						symbol = rowdata[self.symbol_index].lower()
						if symbol not in icon_xbm:
							symbol = map_settings.location_marker
					mkr_key = "%s %s" % (color, symbol)
					if mkr_key not in custom_icons:
						custom_icons[mkr_key] = tk.BitmapImage(data=icon_xbm[symbol], foreground=color)
					marker_icon = custom_icons[mkr_key]
				if self.label_index is not None:
					lbl = rowdata[self.label_index]
					try:
						mkr = self.map_widget.set_marker(lat_val, lon_val, icon=marker_icon,
								text=lbl, font=self.mapfont, text_color=map_settings.label_color,
								command=self.map_sel_table)
					except:
						pass
					else:
						self.loc_map_markers.append(mkr)
				else:
					try:
						mkr = self.map_widget.set_marker(lat_val, lon_val, icon=marker_icon, command=self.map_sel_table)
					except:
						pass
					else:
						self.loc_map_markers.append(mkr)
			else:
				self.missing_latlon += 1
		self.update_plot_data()
		self.loading_dlg.hide()
	def reset_count_labels_flag(self):
		for mkr in self.count_label_markers:
			if mkr.data is not None and type(mkr.data) == list:
				mkr.data[0] = map_settings.count_label_show
	def showhide_count_labels(self, event=None):
		map_settings.count_label_show = not map_settings.count_label_show
		self.reset_count_labels_flag()
		self.redraw_count_labels()
	def delete_selected(self, event):
		# Remove selection markers from the map.
		ix_to_del = []
		for i, mkr in enumerate(self.sel_map_markers):
			mkrx, mkry = mkr.get_canvas_pos(mkr.position)
			diffx = abs(event.x - mkrx)
			diffy = abs(event.y - mkry)
			if diffx <= mkr.icon.width()/2 and diffy <= mkr.icon.height()/2:
				mkr_lat, mkr_lon = mkr.position
				# Delete marker from map
				mkr.delete()
				# Record index of sel_map_markers items to delete
				ix_to_del.append(i)
				# Un-select in table
				for sel_row in self.tbl.selection():
					rowdata = self.tbl.item(sel_row)["values"]
					try:
						lat_val = float(rowdata[self.lat_index])
					except:
						lat_val = None
					try:
						lon_val = float(rowdata[self.lon_index])
					except:
						lon_val = None
					if lon_val is not None and lat_val is not None:
						if abs(lon_val - mkr_lon) < 0.000001 and abs(lat_val - mkr_lat) < 0.000001:
							self.tbl.selection_remove(sel_row)
		# Remove item(s) from self.sel_map_markers
		if len(ix_to_del) > 0:
			for i in reversed(sorted(ix_to_del)):
				del self.sel_map_markers[i]
		self.update_plot_data()
		self.set_status()
	def launch_set_data_types(self, headers, rows):
		# Tabulate data types for table columns, using a separate process for performance.
		if os.name == 'posix':
			self.data_types_queue = multiprocessing.Queue()
			self.data_types_process = multiprocessing.Process(target=set_data_types, args=(headers, rows, self.data_types_queue))
			self.data_types_process.start()
		else:
			self.data_types = set_data_types_core(headers, rows)

	def build_database(self, tv_table):
		self.loading_dlg.display("Building database")
		cur = data_db.cursor()
		cur.execute("drop table if exists mapdata;")
		colnames = db_colnames([r[0] for r in self.data_types])
		colnames.append("treeviewid")
		coltypes = [r[1] for r in self.data_types]
		coltypes.append("int")
		db_coltypes = [sqlite_type_x[t] for t in coltypes]
		coldecs = ", ".join([dquote(c[0])+' '+c[1] for c in list(zip(colnames, db_coltypes))])
		colrange = range(len(coltypes))
		cur.execute("create table mapdata (%s);" % coldecs)
		tbldata = []
		#castfuncs = [data_type_cast_fn(t) for t in coltypes]
		castfuncs = [data_type_cast_sqlite_fn(t) for t in coltypes]
		for row_id in tv_table.get_children():
			row = tv_table.item(row_id)["values"]
			row.append(row_id)
			row_vals = [castfuncs[i](row[i]) for i in colrange]
			tbldata.append(row_vals)
		params = ",".join(['?'] * len(colnames))
		cur.executemany("insert into mapdata values (%s)" % params, tbldata)
		cur.close()
		self.loading_dlg.hide()
		return True

	def show_location_ids(self, event, marker):
		global location_ids, data_db
		if self.label_index is not None:
			position = marker.position
			if position not in location_ids:
				cur = data_db.cursor()
				lat, lon, label = db_colnames([self.lat_col, self.lon_col, map_settings.label_col])
				latval = float(position[0])
				lonval = float(position[1])
				sqlcmd = f"select {label}, count(*) as data_rows from mapdata where abs({latval} - cast({lat} as real)) < 0.00001 and abs({lonval} - cast({lon} as real)) < 0.00001 group by {label} order by {label}"
				lbllist = cur.execute(sqlcmd).fetchall()
				rows = [ [r[0], r[1]] for r in lbllist ]
				headers = [d[0] for d in cur.description]
				location_ids[position] = (headers, rows)
				cur.close()
			else:
				headers, rows = location_ids[position]
			show_table(self.win, "Location Identifiers", f"Latitude: {position[0]}\nLongitude: {position[1]}", rows, headers, "Identifiers")

	def add_data(self, rows, headers, lat_col, lon_col, label_col, symbol_col, color_col):
		# Re-set all data-specific variables and widgets
		self.headers = headers
		self.rows = rows
		self.lat_col = lat_col
		self.lon_col = lon_col
		self.label_col = label_col
		self.src_lat_col = lat_col
		self.src_lon_col = lon_col
		self.lat_4326_col = None
		self.lon_4326_col = None
		map_settings.label_col = label_col
		map_settings.symbol_col = symbol_col
		map_settings.color_col = color_col
		self.lat_index = headers.index(lat_col)
		self.lon_index = headers.index(lon_col)
		self.src_lat_index = headers.index(lat_col)
		self.src_lon_index = headers.index(lon_col)
		self.label_index = headers.index(label_col) if label_col is not None and label_col != '' else None
		self.symbol_index = headers.index(symbol_col) if symbol_col is not None and symbol_col != '' else None
		self.color_index = headers.index(color_col) if color_col is not None and color_col != '' else None
		global location_ids
		location_ids = {}

		if self.crs != 4326:
			try:
				from pyproj import CRS, Transformer
			except:
				self.loading_dlg.hide_all()
				fatal_error("The pyproj library is required to re-project spatial coordinates", kwargs={})
			try:
				crs_proj = CRS(self.crs)
			except:
				self.loading_dlg.hide_all()
				fatal_error("Invalid CRS (%s)" % self.crs, kwargs={})
			if self.lat_4326_col is None:
				for colname in ('lat_4326', 'latitude_4326', 'y_4326', 'unprojected_lat'):
					if colname not in headers:
						self.lat_4326_col = colname
						headers.append(colname)
						break
			if self.lon_4326_col is None:
				for colname in ('lon_4326', 'longitude_4326', 'x_4326', 'unprojected_lon'):
					if colname not in headers:
						self.lon_4326_col = colname
						headers.append(colname)
						break
			self.lat_col = self.lat_4326_col
			self.lon_col = self.lon_4326_col
			self.lat_index = headers.index(self.lat_col)
			self.lon_index = headers.index(self.lon_col)
			crs_4326 = CRS(4326)
			reproj = Transformer.from_crs(crs_proj, crs_4326, always_xy=True)
			for i, r in enumerate(rows):
				rows[i] = list(r)
				r = rows[i]
				y = r[self.src_lat_index]
				x = r[self.src_lon_index]
				if y is not None and y != 0 and x is not None and x != 0:
					try:
						newx, newy = reproj.transform(x, y)
					except:
						newx = None
						newy = None
				else:
					newx = None
					newy = None
				if len(r) < len(headers):
					r.extend([newy, newx])
				else:
					r[self.lat_index] = newy
					r[self.lon_index] = newx

		# Check that not all rows are missing lat/lon.
		has_latlon = False
		for rowdata in rows:
			try:
				lat_val = float(rowdata[self.lat_index])
			except:
				lat_val = None
			try:
				lon_val = float(rowdata[self.lon_index])
			except:
				lon_val = None
			if lon_val is not None and lat_val is not None:
				has_latlon = True
				break
		if not has_latlon:
			fatal_error("All data rows are missing latitude and longitude coordinates.", {'parent': self.win})

		# Determine data types for use in table statistics display and in column selection for plotting.
		# This is done in a separate process for performance (on Linux).  After the data types are set,
		# values in numeric columns in SQLite and the Treeview table are cast to the appropriate numeric type.
		# This is done as early as possible in 'add_data()' to give its process as much time as possible
		# before it is needed by the 'build_database' process.
		self.launch_set_data_types(headers, rows)

		# Populate the treeview
		tframe, tdata = treeview_table(self.tblframe, rows, headers, "browse")
		self.table_row_count = len(tdata.get_children())

		# Scan the table, put points on the map, and find the map extent.
		self.min_lat = self.max_lat = self.min_lon = self.max_lon = None
		self.sel_map_markers = []
		self.missing_latlon = 0
		self.draw_loc_markers(tdata)

		# Copy data from the treeview table to the database.  This includes the treeview IDs.
		self.wait_for_data_types()
		self.build_database(tdata)

		# Set the map extent based on coordinates.
		# The box must not be a point.
		if self.min_lat == self.max_lat:
			self.min_lat = self.min_lat - 0.005
			self.max_lat = self.max_lat + 0.005
		if self.min_lon == self.max_lon:
			self.min_lon = self.min_lon - 0.005
			self.max_lon = self.max_lon + 0.005
		self.map_widget.fit_bounding_box((self.max_lat, self.min_lon), (self.min_lat, self.max_lon))

		# Get all unique coordinates and count of data rows, then add null markers to the map
		lat, lon = db_colnames([self.lat_col, self.lon_col])
		sqlcmd = f"select {lat}, {lon}, count(*) as nrows from mapdata where {lat} is not null and {lon} is not null group by {lat}, {lon}"
		cur = data_db.cursor()
		result = cur.execute(sqlcmd).fetchall()
		new_count_markers = []
		for row in result:
			try:
				lat_val, lon_val = float(row[0]), float(row[1])
			except:
				pass
			else:
				# Marker data is a tuple of T/F and the count.  T/F is the flag controlling display
				dataval = [False, row[2]]
				try:
					new_marker = self.map_widget.set_marker(lat_val, lon_val, icon=self.count_label_icon, data=dataval)
				except:
					pass
				else:
					new_count_markers.append(new_marker)
		for m in self.count_label_markers:
			m.delete()
		self.count_label_markers = new_count_markers
		cur.close()

		# Initial value for data-recoding expression
		self.recode_expr = ""

		# Initial value for user-entered WHERE clause
		self.whereclause = ""

		# Return frame and data table
		return tframe, tdata

	def remove_data(self):
		while len(self.sel_map_markers) > 0:
			self.sel_map_markers.pop().delete()
		while len(self.loc_map_markers) > 0:
			self.loc_map_markers.pop().delete()
		self.map_widget.delete_all_marker()
		self.close_all_plots()
		self.data_types = None
		self.tableframe.destroy()
		self.tbl.destroy()
	def set_tbl_selectmode(self):
		ck = self.multiselect_var.get()
		if ck == '0':
			self.tbl.configure(selectmode = tk.BROWSE)
		else:
			self.tbl.configure(selectmode = tk.EXTENDED)
		self.tbl.bind('<ButtonRelease-1>', self.mark_map)
	def replace_data(self, rows, headers, lat_col, lon_col, label_col, symbol_col, color_col):
		self.remove_data()
		try:
			self.tableframe, self.tbl = self.add_data(rows, headers, lat_col, lon_col, label_col, symbol_col, color_col)
		except:
			self.loading_dlg.hide_all()
			fatal_error("Cannot load data.  Check latitude, longitude, and CRS values.", kwargs={'parent': self.win})
		self.tableframe.grid(column=0, row=0, sticky=tk.NSEW)
		self.set_tbl_selectmode()
		self.set_status()
	def new_data_file(self):
		dfd = DataFileDialog()
		fn, id_col, lat_col, lon_col, crs, sym_col, col_col, msg, headers, rows = dfd.get_datafile()
		if fn is not None and fn != '':
			self.crs = crs
			self.data_src_name = os.path.abspath(fn)
			base_fn = os.path.basename(fn)
			self.win.title("Map of %s" % base_fn)
			self.replace_data(rows, headers, lat_col, lon_col, id_col, sym_col, col_col)
			if msg is not None and msg != '':
				self.msg_label['text'] = msg
	def new_spreadsheet_file(self):
		dfd = ImportSpreadsheetDialog(self.win, self)
		fn, id_col, lat_col, lon_col, crs, sym_col, col_col, msg, headers, rows = dfd.get_datafile()
		if fn is not None and fn != '':
			self.crs = crs
			self.data_src_name = os.path.abspath(fn)
			base_fn = os.path.basename(fn)
			self.win.title("Map of %s" % base_fn)
			self.replace_data(rows, headers, lat_col, lon_col, id_col, sym_col, col_col)
			if msg is not None and msg != '':
				self.msg_label['text'] = msg
	def new_db_table(self):
		dbd = DbConnectDialog(self.win, self)
		tablename, id_col, lat_col, lon_col, crs, sym_col, col_col, desc, headers, rows = dbd.get_data()
		if tablename is not None and tablename != '':
			self.crs = crs
			self.win.title("Map of %s" % tablename)
			self.replace_data(rows, headers, lat_col, lon_col, id_col, sym_col, col_col)
			if desc is not None and desc != '':
				self.msg_label['text'] = desc

	def zoom_full(self):
		self.map_widget.fit_bounding_box((self.max_lat, self.min_lon), (self.min_lat, self.max_lon))
	def zoom_selected(self):
		if len(self.sel_map_markers) > 0:
			if len(self.sel_map_markers) == 1:
				self.focus_map()
			else:
				min_lat = max_lat = min_lon = max_lon = None
				for m in self.sel_map_markers:
					lat, lon = m.position
					if min_lat is None or lat < min_lat:
						min_lat = lat
					if max_lat is None or lat > max_lat:
						max_lat = lat
					if min_lon is None or lon < min_lon:
						min_lon = lon
					if max_lon is None or lon > max_lon:
						max_lon = lon
				if max_lat == min_lat:
					min_lat = min_lat - 0.00001
					max_lat = max_lat + 0.00001
				if max_lon == min_lon:
					min_lon = min_lon - 0.00001
					max_lon = max_lon + 0.00001
				self.map_widget.fit_bounding_box((max_lat, min_lon), (min_lat, max_lon))
	def focus_map(self):
		# Center the map on the last marker
		if len(self.sel_map_markers) > 0:
			m = self.sel_map_markers[-1]
			self.map_widget.set_position(m.position[0], m.position[1])
	def unselect_map(self):
		for m in self.sel_map_markers:
			self.map_widget.delete(m)
		self.tbl.selection_remove(*self.tbl.selection())
		self.sel_map_markers = []
		self.update_plot_data()
		self.set_status()
	def clean_attrib(self, raw_attrib):
		# Removes URLs from attribution string and closes up spaces before commas.
		return re.sub(r"\s+,", ",", re.sub(r"<[^>]+>", "", raw_attrib)).strip()
	def attrib_split(self, raw_attrib):
		# Returns a list of tuple, each tuple consisting of the text and the URL (if any).
		attriblist = [a.strip() for a in raw_attrib.split(",")]
		rv = []
		for attr in attriblist:
			m = re.findall(r"<([^>]+)>", attr)
			rv.append((self.clean_attrib(attr), m[0] if len(m) > 0 else None))
		return rv
	def set_attrib_labels(self, raw_attrib):
		attrs = self.attrib_split(raw_attrib)
		for lbl in self.attrib_frame.winfo_children():
			lbl.destroy()
		for att in attrs:
			lbl = ttk.Label(self.attrib_frame, font=self.attribfont, text=att[0])
			if att[1] is not None:
				lbl["cursor"] = "hand2"
				lbl.bind("<Button-1>", lambda x: webbrowser.open(att[1]))
			lbl.pack(side=tk.LEFT, padx=(3,3))
	def change_basemap(self, *args):
		new_map = self.basemap_var.get()
		tileserver = self.tile_url(new_map)
		if new_map in map_color_adj:
			map_settings.basemap_color_adj = [10**x for x in map_color_adj[new_map]]
		else:
			map_settings.basemap_color_adj = None
		self.map_widget.set_tile_server(tileserver, max_zoom=20)
		self.set_attrib_labels(map_attributions[new_map])
	def map_sel_table(self, marker):
		# Highlight the table row for the clicked map marker
		lat, lon = marker.position
		if self.multiselect_var.get() == '0':
			for mkr in self.sel_map_markers:
				self.map_widget.delete(mkr)
			self.sel_map_markers = []
			self.tbl.selection_remove(*self.tbl.selection())
		for row_id in self.tbl.get_children():
			rowdata = self.tbl.item(row_id)["values"]
			try:
				lat_val = float(rowdata[self.lat_index])
			except:
				lat_val = None
			try:
				lon_val = float(rowdata[self.lon_index])
			except:
				lon_val = None
			if lon_val is not None and lat_val is not None:
				if lat_val == lat and lon_val == lon:
					self.tbl.selection_add(row_id)
					self.tbl.see(row_id)
					new_marker = self.map_widget.set_marker(lat, lon, icon=self.sel_marker_icon)
					if not new_marker in self.sel_map_markers:
						self.sel_map_markers.append(new_marker)
		self.update_plot_data()
		self.set_status()
	def set_status(self):
		statusmsg = "    %d rows" % self.table_row_count
		if self.missing_latlon > 0:
			statusmsg = statusmsg + " (%d without lat/lon)" % self.missing_latlon
		if len(self.tbl.selection()) > 0:
			statusmsg = statusmsg + "  |  %s selected" % len(self.tbl.selection())
		if self.multiselect_var.get() == "1":
			statusmsg = statusmsg + "  |  Ctrl-click to select multiple rows"
		self.tblframe.statusbar.config(text = statusmsg)

	def get_all_data(self, column_list):
		# Plotting and statistics support.  Return all data for the specified columns as a list of column-oriented lists.
		res = [[] for _ in column_list]
		indices = [self.headers.index(c) for c in column_list]
		for sel_row in self.tbl.get_children():
			datarow = self.tbl.item(sel_row)["values"]
			for i, index in enumerate(indices):
				res[i].append(datarow[index])
		return res
	def get_all_rowids(self):
		return self.tbl.get_children()
	def get_sel_data(self, column_list):
		# Plotting and statistics support.  Return data from selected rows for the specified columns, as a list of lists.
		res = [[] for _ in column_list]
		indices = [self.headers.index(c) for c in column_list]
		for sel_row in self.tbl.selection():
			datarow = self.tbl.item(sel_row)["values"]
			for i, index in enumerate(indices):
				res[i].append(datarow[index])
		return res
	def get_sel_rowids(self):
		return self.tbl.selection()
	def update_plot_data(self, unconditional=False):
		# Pushes updates to all plots and other dialogs that may use only selected data.
		# The 'push' is done by calling each dialog's refresh method.
		for dlg in self.candcol_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.q_recalc(get_data=True)
		for dlg in self.duprows_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.recalc(None)
		for dlg in self.cardinality_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.q_recalc(None)
		for dlg in self.aggdlg_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.q_recalc(get_data=True)
		for dlg in self.distdlg_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.reset_locid_list()
		for plot in self.plot_list:
			if unconditional or (plot.sel_only_var.get() == "1" and plot.auto_update):
				plot.q_redraw()
		for dlg in self.univar_list:
			if unconditional or (dlg.sel_only_var.get() == "1" and dlg.auto_update):
				dlg.q_recalc()
		for dlg in self.bivar_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.q_recalc()
		for dlg in self.anova_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.q_recalc()
		for dlg in self.fitdist_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.q_recalc()
		for dlg in self.tsne_list:
			if unconditional or (dlg.sel_only_var.get() == "1"):
				dlg.q_recalc()
		for dlg in self.umap_list:
			if unconditional or (dlg.sel_only_var.get() == "1"):
				dlg.q_recalc()
		for dlg in self.corrmat_list:
			if unconditional or (dlg.sel_only_var.get() == "1" and dlg.auto_update):
				dlg.q_redraw()
		for dlg in self.cosmat_list:
			if unconditional or (dlg.sel_only_var.get() == "1" and dlg.auto_update):
				dlg.q_redraw()
		for dlg in self.catcorresp_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.q_recalc()
		for dlg in self.simmat_list:
			if unconditional or (dlg.sel_only_var.get() == "1" and dlg.auto_update):
				dlg.q_redraw()
		for dlg in self.pca_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.q_recalc()
		for dlg in self.unmixing_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.q_recalc()
		for dlg in self.conttable_list:
			if unconditional or dlg.sel_only_var.get() == "1":
				dlg.q_recalc()
	def clone_plot(self, plot_obj):
		self.wait_for_data_types()
		clone = PlotDialog(self, self.data_types)
		self.plot_list.append(clone)
		gm = re.match(r'(\d+)x(\d+)\+(\d+)\+(\d+)', plot_obj.dlg.geometry())
		clone.geom = f"{gm.group(1)}x{gm.group(2)}+{str(int(gm.group(3))+30)}+{str(int(gm.group(4))+30)}"
		shift_window(clone.dlg, x_offset=10, y_offset=10)
		clone.type_var.set(plot_obj.type_var.get())
		clone.sel_only_var.set(plot_obj.sel_only_var.get())
		clone.autoupdate_var.set(plot_obj.autoupdate_var.get())
		clone.x_var.set(plot_obj.x_var.get())
		clone.y_var.set(plot_obj.y_var.get())
		clone.z_var.set(plot_obj.z_var.get())
		clone.z_width_var.set(plot_obj.z_width_var.get())
		clone.z_sel.bind("<<ComboboxSelected>>", clone.z_changed)
		clone.xlog_var.set(plot_obj.xlog_var.get())
		clone.ylog_var.set(plot_obj.ylog_var.get())
		clone.groupby_var.set(plot_obj.groupby_var.get())
		clone.x_sel["values"] = copy.copy(plot_obj.x_sel["values"])
		clone.y_sel["values"] = copy.copy(plot_obj.y_sel["values"])
		clone.z_sel["values"] = copy.copy(plot_obj.z_sel["values"])
		if plot_obj.type_var.get() == "Bubble plot":
			clone.z_frame.grid(row=2, column=3, columnspan=2)
		clone.groupby_sel["values"] = copy.copy(plot_obj.groupby_sel["values"])
		clone.x_sel["state"] = plot_obj.x_sel["state"]
		clone.y_sel["state"] = plot_obj.y_sel["state"]
		clone.groupby_sel["state"] = plot_obj.groupby_sel["state"]
		clone.xlog_ck["state"] = plot_obj.xlog_ck["state"]
		clone.ylog_ck["state"] = plot_obj.ylog_ck["state"]
		clone.data_btn["state"] = plot_obj.data_btn["state"]
		clone.plot_data_btn["state"] = plot_obj.plot_data_btn["state"]
		clone.dlg.bind("<Alt-h>", clone.do_help)
		clone.dlg.bind("<Alt-n>", clone.clone_plot)
		clone.dlg.bind("<Alt-c>", clone.do_close)
		clone.dlg.bind("<Escape>", clone.do_close)
		clone.dlg.bind("<Alt-t>", clone.set_title)
		clone.dlg.bind("<Alt-x>", clone.set_xlabel)
		clone.dlg.bind("<Alt-y>", clone.set_ylabel)
		if clone.type_var.get() == "Normal Q-Q plot":
			clone.dlg.bind("<Alt-g>", clone.show_groups)
		else:
			clone.dlg.bind("Alt-g>")
		if clone.type_var.get() in ("Line plot", "Scatter plot"):
			clone.dlg.bind("<Alt-l>", clone.set_loess)
			clone.dlg.bind("<Alt-r>", clone.set_linreg)
			clone.dlg.bind("<Alt-s>", clone.set_theilsen)
		else:
			clone.dlg.bind("<Alt-l>")
			clone.dlg.bind("<Alt-r>")
			clone.dlg.bind("<Alt-s>")
		if clone.type_var.get() == "Histogram":
			clone.dlg.bind("<Alt-b>", clone.set_bins)
		elif clone.type_var.get() == "Scatter plot":
			clone.dlg.bind("<Alt-b>", clone.set_scatter_breaks)
		elif clone.type_var.get() == "Line plot":
			clone.dlg.bind("<Alt-b>", clone.set_lineplot_breaks)
		else:
			clone.dlg.bind("Alt-b")
		if clone.type_var.get() in ("Count by category", "Min-max by bin", "Min-max by category", "Box plot", "Mean by category", "CV by category", "Stripchart", "Total by category", "Violin plot"):
			clone.dlg.bind("<Alt-r>", clone.set_rotated)
		if clone.type_var.get() in ("Box plot", "Scatter plot", "Stripchart" "Kernel density (KD) plot", "Violin plot"):
			clone.dlg.bind("<Alt-a>", clone.set_alpha)
		else:
			clone.dlg.bind("<Alt-a>")
		if clone.type_var.get() in ("Count by category", "Mean by category", "CV by category", "Min-max by bin", "Min-max by category", "Total by category", "Y range plot"):
			clone.dlg.bind("<Alt-f>", clone.set_flip_y)
		else:
			clone.dlg.bind("<Alt-f>")
		clone.plot_title = plot_obj.plot_title
		clone.xlabel = plot_obj.xlabel
		clone.ylabel = plot_obj.ylabel
		clone.alpha = plot_obj.alpha
		clone.rotated = plot_obj.rotated
		clone.flip_y = plot_obj.flip_y
		clone.qq_groups = plot_obj.qq_groups
		clone.scatter_breaks = plot_obj.scatter_breaks
		clone.scatter_x_breaks = plot_obj.scatter_x_breaks
		clone.scatter_y_breaks = plot_obj.scatter_y_breaks
		clone.lineplot_breaks = plot_obj.lineplot_breaks
		clone.lineplot_x_breaks = plot_obj.lineplot_x_breaks
		clone.loess = plot_obj.loess
		clone.linreg = plot_obj.linreg
		clone.theilsen = plot_obj.theilsen
		clone.anot = plot_obj.annot
		clone.numeric_columns = copy.copy(plot_obj.numeric_columns)
		clone.pos_numeric_columns = copy.copy(plot_obj.pos_numeric_columns)
		clone.categ_columns = copy.copy(plot_obj.categ_columns)
		clone.categ_columns2 = copy.copy(plot_obj.categ_columns2)
		clone.quant_columns = copy.copy(plot_obj.quant_columns)
		clone.date_columns = copy.copy(plot_obj.date_columns)
		clone.dataset = copy.copy(plot_obj.dataset)
		clone.n_dataset_columns = copy.copy(plot_obj.n_dataset_columns)
		clone.plot_data = copy.copy(plot_obj.plot_data)
		clone.data_labels = copy.copy(plot_obj.data_labels)
		clone.plot_data_labels = copy.copy(plot_obj.plot_data_labels)
		clone.q_redraw(get_data=False)
		#raise_window(clone.dlg)
		clone.dlg.focus()
		clone.show()
	def remove_candcol(self, candcol_dlg):
		try:
			self.candcol_list.remove(candcol_dlg)
		except:
			pass
	def remove_duprows(self, duprow_dlg):
		try:
			self.duprows_list.remove(duprow_dlg)
		except:
			pass
	def remove_cardinality(self, cardinality_dlg):
		try:
			self.cardinality_list.remove(cardinality_dlg)
		except:
			pass
	def remove_aggdlg(self, aggregate_dlg):
		try:
			self.aggdlg_list.remove(aggregate_dlg)
		except:
			pass
	def remove_distdlg(self, distance_dlg):
		try:
			self.distdlg_list.remove(distance_dlg)
		except:
			pass
	def remove_plot(self, plot_obj):
		# For use by the plot 'do_close()' method.
		try:
			self.plot_list.remove(plot_obj)
		except:
			pass
	def close_plot(self, plot_obj):
		try:
			plot_obj.do_close()
			self.remove_plot()
		except:
			pass
	def close_all_plots(self):
		while len(self.plot_list) > 0:
			self.plot_list[0].do_close()
			# The callback will remove the plot.
		self.plot_list = []
	def remove_univar(self, univar_dlg):
		try:
			self.univar_list.remove(univar_dlg)
		except:
			pass
	def remove_bivar(self, bivar_dlg):
		try:
			self.bivar_list.remove(bivar_dlg)
		except:
			pass
	def remove_anova(self, anova_dlg):
		try:
			self.anova_list.remove(anova_dlg)
		except:
			pass
	def remove_fitdist(self, fitdist_dlg):
		try:
			self.fitdist_list.remove(fitdist_dlg)
		except:
			pass
	def remove_corrmat(self, corrmat_dlg):
		try:
			self.corrmat_list.remove(corrmat_dlg)
		except:
			pass
	def remove_cosmat(self, cosmat_dlg):
		try:
			self.cosmat_list.remove(cosmat_dlg)
		except:
			pass
	def remove_pca(self, pca_dlg):
		try:
			self.pca_list.remove(pca_dlg)
		except:
			pass
	def remove_tsne(self, tsne_dlg):
		try:
			self.tsne_list.remove(tsne_dlg)
		except:
			pass
	def remove_umap(self, umap_dlg):
		try:
			self.umap_list.remove(umap_dlg)
		except:
			pass
	def remove_categcorresp(self, categcorresp_dlg):
		try:
			self.catcorresp_list.remove(categcorresp_dlg)
		except:
			pass
	def remove_simmat(self, simmat_dlg):
		try:
			self.simmat_list.remove(simmat_dlg)
		except:
			pass
	def remove_unmixing(self, unmixing_dlg):
		try:
			self.unmixing_list.remove(unmixing_dlg)
		except:
			pass
	def remove_conttable(self, conttable_dlg):
		try:
			self.conttable_list.remove(conttable_dlg)
		except:
			pass
	def remove_roccurve(self, roc_dlg):
		try:
			self.rocccurve_list.remove(roc_dlg)
		except:
			pass

	def change_crs(self):
		crsdlg = NewCrsDialog(self.crs)
		new_crs = crsdlg.show()
		if new_crs is not None:
			if new_crs != self.crs:
				try:
					crs_proj = CRS(new_crs)
				except:
					warning("Invalid CRS (%s)" % new_crs, kwargs={})
				else:
					if self.lat_4326_col is None:
						for colname in ('lat_4326', 'latitude_4326', 'y_4326', 'unprojected_lat'):
							if colname not in self.headers:
								self.lat_4326_col = colname
								self.headers.append(colname)
								for r in self.rows:
									r.append(None)
								break
					if self.lon_4326_col is None:
						for colname in ('lon_4326', 'longitude_4326', 'x_4326', 'unprojected_lon'):
							if colname not in self.headers:
								self.lon_4326_col = colname
								self.headers.append(colname)
								for r in self.rows:
									r.append(None)
								break
					self.lat_col = self.lat_4326_col
					self.lon_col = self.lon_4326_col
					self.lat_index = self.headers.index(self.lat_4326_col)
					self.lon_index = self.headers.index(self.lon_4326_col)
					crs_4326 = CRS(4326)
					self.crs = new_crs
					try:
						reproj = Transformer.from_crs(crs_proj, crs_4326, always_xy=True)
					except:
						warning("Invalid or unrecognized CRS.")
					else:
						for r in self.rows:
							y = r[self.src_lat_index]
							x = r[self.src_lon_index]
							if y is not None and y != 0 and x is not None and x != 0:
								try:
									newx, newy = reproj.transform(x, y)
									r[self.lat_index] = newy
									r[self.lon_index] = newx
								except:
									r[self.lat_index] = None
									r[self.lon_index] = None
							else:
								r[self.lat_index] = None
								r[self.lon_index] = None
						selected = self.tbl.selection()
						self.replace_data(self.rows, self.headers, self.src_lat_col, self.src_lon_col, self.label_col, self.symbol_col, self.color_col)
						self.tbl.selection_set(tuple(selected))
						self.mark_map({})
	def cancel(self):
		try:
			self.win.destroy()
		except:
			pass
		sys.exit()
	def export_map_and_quit(self):
		fn, ext = os.path.splitext(self.imageoutputfile)
		if ext.lower() == ".ps":
			self.export_map_to_ps(self.imageoutputfile)
		else:
			self.map_widget.update_idletasks()
			#self.win.after(200, self.save_imageoutputfile)
			self.save_imageoutputfile()
		self.win.destroy()
	def export_map_to_ps(self, outfile):
		self.map_widget.canvas.postscript(file=outfile, colormode='color')
	def save_imageoutputfile(self):
		obj = self.map_widget.canvas
		bounds = (obj.winfo_rootx(), obj.winfo_rooty(), 
				obj.winfo_rootx() + obj.winfo_width(), obj.winfo_rooty() + obj.winfo_height())
		ImageGrab.grab(bbox=bounds).save(self.imageoutputfile)
	def export_map_to_img(self, outfile):
		# Allow map to recover from blocking by the file dialog box before grabbing and exporting the canvas
		self.map_widget.update_idletasks()
		self.imageoutputfile = outfile
		self.win.after(1000, self.save_imageoutputfile)

	def coord_candkeys(self):
		# Returns a (possibly empty) list of Treeview column names that are candidate keys for coordinate pairs.
		coordkey_ids = []
		curs = data_db.cursor()
		lat, lon = db_colnames([self.src_lat_col, self.src_lon_col])
		sqlcmd = f"""select count(*) from (select distinct {lat}, {lon} from mapdata
		where {lat} is not null and {lon} is not null);"""
		result = curs.execute(sqlcmd).fetchall()
		if len(result) == 0:
			warning("There are no coordinates", kwargs={})
		else:
			unique_coords = result[0][0]
			self.wait_for_data_types()
			str_columns = [c[0] for c in self.data_types if c[1] == "string" and c[0] != lat and c[0] != lon]
			db_columns = db_colnames(str_columns)
			db_to_tv = dict(zip(db_columns, str_columns))
			for colname in db_columns:
				sqlcmd = f"select count(*) from (select distinct {colname} from mapdata where {colname} is not null)"
				result = curs.execute(sqlcmd).fetchall()
				if result[0][0] == unique_coords:
					sqlcmd = f"""select count(*), max(nrows) from (select {lat},{lon},count(*) as nrows from  (
					select distinct {lat}, {lon}, {colname} from mapdata
					where {lat} is not null and {lon} is not null and {colname} is not null)
					group by {lat},{lon})"""
					result = curs.execute(sqlcmd).fetchall()
					if result[0][0] == unique_coords and result[0][1] == 1:
						coordkey_ids.append(db_to_tv[colname])
		curs.close()
		return coordkey_ids

	def counts_by_loc(self):
		# Display a table of data row counts for each unique lat/lon pair, with
		# any other columns that are 1:1 with lat/lon.
		cur = data_db.cursor()
		lat, lon = db_colnames([self.src_lat_col, self.src_lon_col])
		sqlcmd = f"""select count(*) from (select distinct {lat}, {lon} from mapdata
		where {lat} is not null and {lon} is not null);"""
		result = cur.execute(sqlcmd).fetchall()
		if len(result) == 0:
			cur.close()
			warning("There are no coordinates", kwargs={})
		else:
			unique_coords = result[0][0]
			ok_ids = []
			self.wait_for_data_types()
			self.loading_dlg.display("Finding data row\ncounts by location")
			ok_ids = self.coord_candkeys()
			db_ok_ids = db_colnames(ok_ids)
			sqlcmd = f"select distinct {lat}, {lon}, count(*) as data_rows"
			sqlcmdp2 = f" from mapdata group by {lat}, {lon}"
			if len(ok_ids) > 0:
				sqlcmd = sqlcmd + ", " + ", ".join(db_ok_ids)
				sqlcmdp2 = sqlcmdp2 + ", " + ", ".join(db_ok_ids)
			sqlcmd = sqlcmd + sqlcmdp2 + " order by count(*);"
			result = cur.execute(sqlcmd).fetchall()
			colhdrs = [c[0] for c in cur.description]
			cur.close()
			self.loading_dlg.hide()
			show_table(self.win, "Row Counts By Location", "Number of data rows for each unique location", result, colhdrs, "Rows per location")

	def select_by_sqlcmd(self, sqlcmd, action, errmsg):
		# The SQL command must select treeviewid values from the main data table.
		# Return T/F to indicate success or error.
		cur = data_db.cursor()
		try:
			result = cur.execute(sqlcmd)
			id_list = [str(r[0]) for r in result.fetchall()]
		except:
			cur.close()
			warning(errmsg, kwargs={})
			return False
		cur.close()
		# Enable multiselect
		global multiselect
		multiselect = "1"
		self.multiselect_var.set("1")
		self.tbl.configure(selectmode = tk.EXTENDED)
		if action == "Replace":
			self.unselect_map()
			self.tbl.selection_set(list(id_list))
		elif action == "Union":
			all_selections = tuple(set(self.tbl.selection()) | set(id_list))
			self.tbl.selection_set(all_selections)
		elif action == "Intersection":
			int_selections = tuple(set(self.tbl.selection()) & set(id_list))
			self.tbl.selection_set(int_selections)
		elif action == "Difference O-N":
			# Old - New
			diff_selections = tuple(set(self.tbl.selection()) - set(id_list))
			self.tbl.selection_set(diff_selections)
		else:
			# New - Old
			diff_selections = tuple(set(id_list) - set(self.tbl.selection()))
			self.tbl.selection_set(diff_selections)
		self.mark_map(None)
		self.set_status()
		return True

	def get_plot_config(self):
		global show_regression_stats, wrapwidth, wrap_at_underscores
		dlg = PlotConfigDialog(show_regression_stats, wrapwidth, wrap_at_underscores)
		plotconfig = dlg.show()
		show_regression_stats = plotconfig["show_regr_stats"]
		wrapwidth = plotconfig["wrapwidth"]
		wrap_at_underscores = plotconfig["wrap_underscores"]
		self.update_plot_data(unconditional=True)

	def wait_for_data_types(self, msg=True):
		if self.data_types is None:
			if msg:
				self.loading_dlg.display("Evaluating data types")
			self.data_types = self.data_types_queue.get()
			self.data_types_process.join()
			self.data_types_process.close()
			if msg:
				self.loading_dlg.hide()

	def show_data_types(self):
		self.wait_for_data_types()
		show_table(self.win, "Data Types", \
				"Data types, data completeness, and number of unique non-missing values for columns of the data table:", \
				self.data_types, ["Column", "Type", "Missing", "Unique"], "Data Types", \
				dlg_args={"can_resize": True}, tv_args={"select_mode": "browse"})

	def hide_columns(self):
		allcols = self.tbl["columns"]
		dispcols = self.tbl["displaycolumns"]
		if len(dispcols)==1 and dispcols[0]=='#all':
			dispcols = allcols
		dlg = TableHideColsDialog(self.win, allcols, dispcols)
		result = dlg.show()
		if result is not None:
			self.tbl["displaycolumns"] = result

	def find_candkeys(self, args=None):
		dlg = FindCandKeysDialog(self, self.table_row_count, self.data_types)
		self.candcol_list.append(dlg)
		dlg.show()

	def find_duprows(self, args=None):
		bad_cols = [self.lat_col, self.lon_col]
		if self.src_lat_col is not None:
			bad_cols.extend([self.src_lat_col, self.src_lon_col])
		if self.lat_4326_col is not None:
			bad_cols.extend([self.lat_4326_col, self.lon_4326_col])
		dlg = FindDupRowsDialog(self, self.table_row_count, self.data_types, bad_cols)
		self.duprows_list.append(dlg)
		dlg.show()

	def cardinality_test(self, args=None):
		dlg = CardinalityTestDialog(self, self.data_types)
		self.cardinality_list.append(dlg)
		dlg.show()

	def recode_data(self, args=None):
		bad_cols = [self.lat_col, self.lon_col]
		if self.src_lat_col is not None:
			bad_cols.extend([self.src_lat_col, self.src_lon_col])
		if self.lat_4326_col is not None:
			bad_cols.extend([self.lat_4326_col, self.lon_4326_col])
		dlg = RecodeDialog(self, self.data_types, bad_cols, self.recode_expr)
		recode_expr, recode_col, col_type, selected_selection, null_selection = dlg.show()
		if recode_expr is not None:
			self.loading_dlg.display("Re-coding data")
			# Try to evaluate the expression to see if it is valid.
			db_col_name = db_colnames([recode_col])[0]
			sql = f"select treeviewid, cast(({recode_expr}) as {col_type}) as {db_col_name} from mapdata"
			sel_sql = None
			if selected_selection != 'all':
				selected = self.tbl.selection()
				if len(selected) > 0:
					selected = ["'" + s + "'" for s in selected]
					sel_sql = f"treeviewid in ({','.join(selected)})"
					if selected_selection == "unselected":
						sel_sql = f"not ({sel_sql})"
			null_sql = None
			if null_selection != 'all' and recode_col in self.headers:
				if null_selection == "empty":
					null_sql = f"({db_col_name} is null or {db_col_name} = '')"
				else:
					null_sql = f"({db_col_name} is not NULL and {db_col_name} <> '')"
			if sel_sql is not None:
				sql = sql + " where " + sel_sql
				if null_sql is not None:
					sql = sql + " and " + null_sql
			else:
				if null_sql is not None:
					sql = sql + " where " + null_sql
			cur = data_db.cursor()
			try:
				expr_values = cur.execute(sql).fetchall()
			except:
				cur.close()
				self.loading_dlg.hide()
				warning("The recoding expression is invalid.", {"parent": self.win})
			else:
				# Update the database
				new_column = recode_col not in self.headers
				if new_column:
					cur.execute(f"alter table mapdata add column {db_col_name} {col_type};")
				for v in expr_values:
					if v[0] is not None and v[1] is not None:
						if col_type == "text":
							newval = f"'{v[1]}'"
						else:
							newval = v[1]
						cur.execute(f"update mapdata set {db_col_name} = {newval} where treeviewid = {v[0]};")
				# Update the Treeview
				if new_column:
					# Add a column.
					tv_keys = [r[0] for r in expr_values]
					new_values = [r[1] for r in expr_values]
					add_tv_column(self.tbl, recode_col, tv_keys, new_values)
					self.headers = self.headers + [recode_col]
					viscols = list(self.tbl["displaycolumns"])
					if viscols[0] != '#all':
						viscols.append(recode_col)
						self.tbl["displaycolumns"] = viscols
				else:
					for row in expr_values:
						self.tbl.set(row[0], column=recode_col, value=str(row[1]))
				# Update the table specs
				if col_type == "text":
					dt = "string"
				elif col_type == "integer":
					dt = "int"
				elif col_type == "real":
					dt = "float"
				else:
					dt = col_type
				coldata = self.get_all_data([recode_col])[0]
				missing = len([v for v in coldata if v is None or v == ''])
				unique = len(set([v for v in coldata if v is not None and v != '']))
				if new_column:
					self.data_types.append([recode_col, dt, missing, unique])
				else:
					col_ix = self.headers.index(recode_col)
					self.data_types[col_ix] = [recode_col, dt, missing, unique]
			self.loading_dlg.hide()

	def aggregate_data(self, args=None):
		dlg = AggregateDialog(self, self.data_types)
		self.aggdlg_list.append(dlg)
		dlg.show()

	def unique_data_values(self, args=None):
		dlg = UniqueValuesDialog(self, self.data_types)
		dlg.show()

	def add_rowid(self, args=None):
		dlg = AddRowIDDialog(self, self.data_types)
		colname, prefix = dlg.show()
		if colname is not None and colname.strip() != '':
			self.loading_dlg.display("Adding unique row ID")
			current_rowids = list(self.tbl.get_children())
			# Get the maximum width of the new IDs, to set padding
			loglen = math.log10(len(current_rowids))
			ceil_len = math.ceil(loglen)
			if int(loglen) == ceil_len:
				val_width = ceil_len + 1
			else:
				val_width = ceil_len
			if prefix != '':
				prefix = prefix + "_"
			new_ids = [prefix + str(int(rowid)+1).rjust(val_width, '0') for rowid in current_rowids]
			# Update SQLite
			db_col_name = db_colnames([colname])[0]
			cur = data_db.cursor()
			cur.execute(f"alter table mapdata add column {db_col_name} text;")
			for rn in range(len(current_rowids)):
				cur.execute(f"update mapdata set {db_col_name} = '{new_ids[rn]}' where treeviewid = {current_rowids[rn]};")

			# Update the Treeview
			add_tv_column(self.tbl, colname, current_rowids, new_ids)
			viscols = list(self.tbl["displaycolumns"])
			if viscols[0] != '#all':
				viscols.append(colname)
				self.tbl["displaycolumns"] = viscols

			# Update the header names and table specs
			self.headers = self.headers + [colname]
			self.data_types.append([colname, "string", 0, len(new_ids)])
			#
			self.loading_dlg.hide()

	def add_coordinate_key(self, args=None):
		dlg = AddCoordKeyDialog(self, self.data_types)
		colname, prefix = dlg.show()
		if colname is not None:
			self.loading_dlg.display("Adding coordinate keys")
			# Get all unique coordinates with a row number
			cur = data_db.cursor()
			lat, lon = db_colnames([self.src_lat_col, self.src_lon_col])
			sqlcmd = f"""select {lat}, {lon}, row_number() over (order by {lat}, {lon})
			from (select distinct {lat}, {lon} from mapdata
			where {lat} is not null and {lon} is not null) order by 1,2;"""
			result = cur.execute(sqlcmd).fetchall()
			if len(result) == 0:
				cur.close()
				self.loading_dlg.hide()
				warning("There are no coordinates", kwargs={})
			else:
				# Eliminate rows that aren't floats.
				for i in reversed(list(range(len(result)))):
					if not isfloat(result[i][0]) or not isfloat(result[i][1]) \
							or result[i][0] is None or result[i][1] is None \
							or (type(result[i][0]) == str and len(result[i][0].strip()) == 0) \
							or (type(result[i][1]) == str and len(result[i][1].strip()) == 0):
								del result[i]
				if len(result) == 0:
					cur.close()
					self.loading_dlg.hide()
					warning("There are no coordinates", kwargs={})
				else:
					# Get the maximum width of the new IDs, to set padding
					loglen = math.log10(len(result))
					ceil_len = math.ceil(loglen)
					if int(loglen) == ceil_len:
						val_width = ceil_len + 1
					else:
						val_width = ceil_len
					if prefix != '':
						prefix = prefix + "_"
					# Assign new IDs
					coord_ids = [r[2] for r in result]
					new_ids = [prefix + str(int(rowid)+1).rjust(val_width, '0') for rowid in coord_ids]
					id_tbl = [ [result[i][0], result[i][1], new_ids[i]] for i in range(len(result))]
					# Update SQLite
					db_col_name = db_colnames([colname])[0]
					cur.execute(f"alter table mapdata add column {db_col_name} text;")
					for rn in range(len(id_tbl)):
						cur.execute(f"""update mapdata set {db_col_name} = '{id_tbl[rn][2]}' where {lat} = {id_tbl[rn][0]}
								and {lon} = {id_tbl[rn][1]};""")
					# Update the Treeview
					coord_tvids = cur.execute(f"""select treeviewid, {db_col_name} as coord_key
						from mapdata where {db_col_name} is not null;""").fetchall()
					add_tv_column(self.tbl, colname, [r[0] for r in coord_tvids], [r[1] for r in coord_tvids])
					cur.close()
					# Update the header names and table specs (name, type, #missing, #unique)
					self.headers = self.headers + [colname]
					viscols = list(self.tbl["displaycolumns"])
					if viscols[0] != '#all':
						viscols.append(colname)
						self.tbl["displaycolumns"] = viscols
					self.data_types.append([colname, "string", self.table_row_count - len(coord_tvids), len(new_ids)])
					#
					self.loading_dlg.hide()
		
	def calc_distances(self, args=None):
		coordkeys = self.coord_candkeys()
		if len(coordkeys) == 0:
			warning("There must be at least one candidate key column for location coordinates.  The 'Table/Add coordinate key' menu item can be used to create a candidate key.", {"parent": self.win})
		else:
			dlg = DistanceDialog(self, coordkeys)
			self.distdlg_list.append(dlg)
			dlg.show()
	def run_query(self, args=None):
		dlg = QueryDialog(self, self.headers, data_db, self.whereclause)
		any_changed = dlg.show()
	def new_plot(self, args=None):
		self.wait_for_data_types()
		dlg = PlotDialog(self, self.data_types)
		self.plot_list.append(dlg)
		dlg.show()
	def pairplot(self, args=None):
		dlg = PairPlotDialog(self, self.data_types)
		self.plot_list.append(dlg)
		dlg.show()
	def multivarbarplot(self, args=None):
		dlg = MultivarBarPlotDialog(self, self.data_types)
		self.plot_list.append(dlg)
		dlg.show()
	def variabledistplot(self, args=None):
		dlg = VariableDistDialog(self, self.data_types)
		self.plot_list.append(dlg)
		dlg.show()
	def pctstackedareaplot(self, args=None):
		dlg = PctStackedAreaDialog(self, self.data_types)
		self.plot_list.append(dlg)
		dlg.show()
	def multivarlineplot(self, args=None):
		dlg = MultiVariableLinePlotDialog(self, self.data_types)
		self.plot_list.append(dlg)
		dlg.show()
	def multivargrppctchart(self, args=None):
		dlg = MultivarGrpPctBarDialog(self, self.data_types)
		self.plot_list.append(dlg)
		dlg.show()
	def rankabundplot(self, args=None):
		bad_cols = [self.lat_col, self.lon_col]
		if self.src_lat_col is not None:
			bad_cols.extend([self.src_lat_col, self.src_lon_col])
		if self.lat_4326_col is not None:
			bad_cols.extend([self.lat_4326_col, self.lon_4326_col])
		dlg = RankAbundDialog(self, self.data_types, bad_cols)
		self.plot_list.append(dlg)
		dlg.show()
	def select_colocated(self, args=None):
		dlg = ColocatedDialog()
		matchstr, num, merge = dlg.show()
		if matchstr is not None:
			if matchstr == "more than":
				comp = ">"
			elif matchstr == "less than":
				comp = "<"
			else:
				comp = "="
			lat = self.src_lat_col
			lon = self.src_lon_col
			sqlcmd = f'SELECT treeviewid FROM mapdata WHERE ({lat}, {lon}) in (select {lat}, {lon} from mapdata group by {lat}, {lon} having count(*) {comp} {num})'
			self.select_by_sqlcmd(sqlcmd, merge, "Invalid data selection: %s" % sqlcmd)
	def select_nonmissing(self, args=None):
		dlg = SelNonMissingDialog(self.data_types)
		sel_columns, action = dlg.show()
		if len(sel_columns) > 0:
			dbcols = db_colnames(sel_columns)
			whereclause = " and ".join(['"'+col+'" is not null and length(trim(cast("' + col + '" as text))) > 0' for col in dbcols])
			sqlcmd = "SELECT treeviewid FROM mapdata WHERE %s" % whereclause
			self.select_by_sqlcmd(sqlcmd, action, "Invalid data selection: %s" % whereclause)
	def select_random(self, args=None):
		global random_last_n, random_use_selected
		data_rows = self.table_row_count
		sel_rows = len(self.tbl.selection())
		dlg = RandomSelDialog(data_rows, sel_rows, random_last_n, random_use_selected)
		rand_n, from_selected = dlg.show()
		if rand_n is not None:
			global multiselect
			multiselect = "1"
			self.multiselect_var.set("1")
			self.tbl.configure(selectmode = tk.EXTENDED)
			random_last_n = rand_n
			random_use_selected = from_selected
			if from_selected:
				tbl_selection = random.sample(self.tbl.selection(), min(rand_n, sel_rows))
			else:
				tbl_selection = random.sample(self.tbl.get_children(), min(rand_n, data_rows))
			self.unselect_map()
			self.tbl.selection_set(tuple(tbl_selection))
			self.mark_map(None)
			self.set_status()

	def add_menu(self):
		mnu = tk.Menu(self.win)
		self.win.config(menu=mnu)
		file_menu = tk.Menu(mnu, tearoff=0)
		tbl_menu = tk.Menu(mnu, tearoff=0)
		map_menu = tk.Menu(mnu, tearoff=0)
		sel_menu = tk.Menu(mnu, tearoff=0)
		plot_menu = tk.Menu(mnu, tearoff=0)
		stats_menu = tk.Menu(mnu, tearoff=0)
		help_menu = tk.Menu(mnu, tearoff=0)
		mnu.add_cascade(label="File", menu=file_menu, underline=0)
		mnu.add_cascade(label="Table", menu=tbl_menu, underline=0)
		mnu.add_cascade(label="Map", menu=map_menu, underline=0)
		mnu.add_cascade(label="Selections", menu=sel_menu, underline=0)
		mnu.add_cascade(label="Plot", menu=plot_menu, underline=0)
		mnu.add_cascade(label="Statistics", menu=stats_menu, underline=2)
		mnu.add_cascade(label="Help", menu=help_menu, underline=0)
		def show_selected():
			if self.tbl.selection():
				rowset = []
				for sel_row in self.tbl.selection():
					rowset.append(self.tbl.item(sel_row)["values"])
				show_table(self.win, "Selected Rows", "Selected rows from the main data table.", rowset, self.headers, "Selected data", dlg_args={}, tv_args={})
		def save_table():
			if self.tbl.selection():
				rowset = []
				for sel_row in self.tbl.selection():
					rowset.append(self.tbl.item(sel_row)["values"])
				export_data_table(self.headers, rowset, sheetname="Selected map items")
		def save_entire_table():
			export_data_table(self.headers, [self.tbl.item(r)["values"] for r in self.tbl.get_children()],
					sheetname="Mapdata table")
		def save_map():
			outfile = tkfiledialog.asksaveasfilename(title="File to save map",
				filetypes=[('Postscript files', '.ps'), ('JPEG files', '.jpg'), ('PNG files', '.png')],
				defaultextension=".png")
			fn, ext = os.path.splitext(outfile)
			if len(ext) > 1 and outfile[-2:].lower() == 'ps':
				self.export_map_to_ps(outfile)
			else:
				if ext == '':
					outfile = outfile + ".jpg"
				self.export_map_to_img(outfile)
		def change_map_settings():
			dlg = MapConfigDialog(self, self.win, self.headers)
			dlg.show()
			self.update_plot_data(unconditional=True)
		def import_symbol_file():
			sd = ImportSymbolDialog()
			name, fn = sd.show()
			if name is not None and fn is not None:
				import_symbol(name, fn)
				fqfn = os.path.abspath(fn)
				symb_spec = (name, fqfn)
				if not symb_spec in imported_symbols:
					imported_symbols.append(symb_spec)
		def read_config_file():
			fn = tkfiledialog.askopenfilename(filetypes=([('Config files', '.conf')]))
			if fn != '' and fn is not None and fn != ():
				global multiselect
				pre_select = multiselect
				pre_basemap = self.basemap_var.get()
				pre_symbol = map_settings.select_symbol
				pre_color = map_settings.select_color
				pre_loc_symbol = map_settings.location_marker
				pre_loc_color = map_settings.location_color
				pre_label_color = map_settings.label_color
				pre_label_font = map_settings.label_font
				pre_label_size = map_settings.label_size
				pre_label_bold = map_settings.label_bold
				pre_label_position = map_settings.label_position
				read_config(fn, self.args)
				# (Re)set configuration options to global defaults
				self.map_option_menu['values'] = self.available_tile_servers()
				if multiselect != pre_select:
					self.multiselect_var.set(multiselect)
				if initial_basemap != pre_basemap:
					self.basemap_var.set(initial_basemap)
					tileserver = self.tile_url(initial_basemap)
					self.map_widget.set_tile_server(tileserver, max_zoom=20)
				if map_settings.select_symbol != pre_symbol or map_settings.select_color != pre_color:
					set_sel_marker(map_settings.select_symbol, map_settings.select_color)
				# Redraw markers if any setting has changed
				if map_settings.location_marker != pre_loc_symbol or map_settings.location_color != pre_loc_color or \
						map_settings.label_color != pre_label_color or map_settings.label_font != pre_label_font or \
						map_settings.label_size != pre_label_size or map_settings.label_bold != pre_label_bold or \
						map_settings.label_position != pre_label_position:
							if map_settings.label_font != pre_label_font or map_settings.label_size != pre_label_size or \
									map_settings.label_bold != pre_label_bold:
								self.mapfont = makefont(map_settings.label_font, map_settings.label_size, map_settings.label_bold)
							self.loc_marker_icon = set_get_loc_marker(map_settings.location_marker, map_settings.location_color)
							self.redraw_loc_markers(self.tbl)
							self.redraw_sel_markers()
				global config_files_user
				config_files_user.append(os.path.abspath(fn))
		def save_config():
			fn = tkfiledialog.asksaveasfilename(filetypes=([('Config files', '.conf')]))
			if fn != '':
				global map_settings
				f = open(fn, "w")
				f.write("# Configuration file for mapdata.py\n# Created by export from mapdata.py at %s\n" % datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
				f.write("\n[basemap_tile_servers]\n")
				added_bms = [k for k in bm_servers if not k in initial_bms]
				for k in added_bms:
					f.write("%s=%s\n" % (k, bm_servers[k]))
				f.write("\n[api_keys]\n")
				for k in api_keys:
					f.write("%s=%s\n" % (k, api_keys[k]))
				f.write("\n[map_attributions]\n")
				for tileserver in map_attributions:
					if tileserver not in ("OpenStreetMap", "Google streets", "Google satellite", "Open topo map"):
						f.write("%s=%s\n" % (tileserver, map_attributions[tileserver]))
				f.write("\n[map_color_adj]\n")
				for tileserver in map_color_adj:
					f.write("%s=%s\n" % (tileserver, ",".join([str(x) for x in map_color_adj[tileserver]])))
				f.write("\n[symbols]\n")
				for s in imported_symbols:
					f.write("%s=%s\n" % (s[0], s[1]))
				f.write("\n[defaults]\n")
				f.write("basemap=%s\n" % self.basemap_var.get())
				f.write("location_color=%s\n" % map_settings.location_color)
				f.write("location_marker=%s\n" % map_settings.location_marker)
				f.write("label_bold=%s\n" % ('No' if not map_settings.label_bold else 'Yes'))
				f.write("label_color=%s\n" % map_settings.label_color)
				f.write("label_font=%s\n" % map_settings.label_font)
				f.write("label_position=%s\n" % map_settings.label_position)
				f.write("label_size=%s\n" % map_settings.label_size)
				f.write("multiselect=%s\n" % ('Yes' if self.multiselect_var.get() == '1' else 'No'))
				f.write("use_data_color=%s\n" % map_settings.use_data_color)
				f.write("use_data_marker=%s\n" % map_settings.use_data_marker)
				f.write("select_color=%s\n" % map_settings.select_color)
				f.write("select_symbol=%s\n" % map_settings.select_symbol)
				f.write("show_regression_stats=%s\n" % show_regression_stats)
				f.write("wrapwidth=%s\n" % wrapwidth)
				f.write("wrap_at_underscores=%s\n" % wrap_at_underscores)
				f.write("\n[misc]\n")
				if editor is not None:
					f.write("editor=%s\n" % editor)
				f.write("temp_dbfile=%s\n" % temp_dbfile)

		def set_editor():
			global editor
			dlg = GetEditorDialog(self.win, editor)
			new_editor = dlg.show()
			if new_editor is not None:
				editor = new_editor
		def invert_selections():
			selected = self.tbl.selection()
			new_selections = []
			for iid in self.tbl.get_children():
				if not iid in selected:
					new_selections.append(iid)
			self.tbl.selection_set(tuple(new_selections))
			self.mark_map(None)
			self.set_status()

		def univar_stats():
			self.wait_for_data_types()
			dlg = UnivarStatsDialog(self, self.data_types)
			self.univar_list.append(dlg)
			dlg.show()

		def bivar_stats():
			self.wait_for_data_types()
			dlg = BivarStatsDialog(self, self.data_types)
			self.bivar_list.append(dlg)
			dlg.show()

		def anova_stats():
			self.wait_for_data_types()
			dlg = ANOVADialog(self, self.data_types)
			self.anova_list.append(dlg)
			dlg.show()

		def fitdist_stats():
			self.wait_for_data_types()
			dlg = FitDistDialog(self, self.data_types)
			self.fitdist_list.append(dlg)
			dlg.show()

		def corr_matrix():
			self.wait_for_data_types()
			dlg = CorrMatrixDialog(self, self.data_types)
			self.corrmat_list.append(dlg)
			dlg.show()

		def cos_matrix():
			self.wait_for_data_types()
			dlg = CosineSimilarityDialog(self, self.data_types)
			self.cosmat_list.append(dlg)
			dlg.show()

		def pca():
			self.wait_for_data_types()
			row_ids = sorted([c[0] for c in self.data_types if c[1] == "string" and c[2] == 0 and c[3] == self.table_row_count])
			if len(row_ids) == 0:
				warning("The data table must have a column with unique row IDs.  These can be added with the 'Table/Add unique row ID' menu item.", {"parent":self.win})
			else:
				dlg = PCADialog(self, [r for r in self.data_types if r[0] not in row_ids], row_ids)
				self.pca_list.append(dlg)
				dlg.show()

		def tsne_anal():
			bad_cols = [self.lat_col, self.lon_col]
			if self.src_lat_col is not None:
				bad_cols.extend([self.src_lat_col, self.src_lon_col])
			if self.lat_4326_col is not None:
				bad_cols.extend([self.lat_4326_col, self.lon_4326_col])
			self.loading_dlg.hide()
			self.wait_for_data_types()
			dlg = TSNEDialog(self, self.data_types, bad_cols)
			self.tsne_list.append(dlg)
			dlg.show()

		def umap_anal():
			self.loading_dlg.display("Initializing UMAP")
			try:
				import umap
			except:
				self.loading_dlg.hide()
				warning("Libraries required for UMAP analysis are missing.", {'parent': self.win})
			else:
				bad_cols = [self.lat_col, self.lon_col]
				if self.src_lat_col is not None:
					bad_cols.extend([self.src_lat_col, self.src_lon_col])
				if self.lat_4326_col is not None:
					bad_cols.extend([self.lat_4326_col, self.lon_4326_col])
				self.loading_dlg.hide()
				self.wait_for_data_types()
				dlg = UMAPDialog(self, self.data_types, bad_cols)
				self.umap_list.append(dlg)
				dlg.show()

		def categ_corresp():
			self.wait_for_data_types()
			dlg = CategCorrespDialog(self, self.data_types)
			self.catcorresp_list.append(dlg)
			dlg.show()

		def sim_matrix():
			self.wait_for_data_types()
			row_ids = sorted([c[0] for c in self.data_types if c[1] == "string" and c[2] == 0 and c[3] == self.table_row_count])
			if len(row_ids) == 0:
				warning("The data table must have a column with unique row IDs.  These can be added with the 'Table/Add unique row ID' menu item.", {"parent":self.win})
			else:
				dlg = CategSimDialog(self, [r for r in self.data_types if r[0] not in row_ids], row_ids)
				self.simmat_list.append(dlg)
				dlg.show()

		def unmixing():
			self.wait_for_data_types()
			row_ids = sorted([c[0] for c in self.data_types if c[1] == "string" and c[2] == 0 and c[3] == self.table_row_count])
			if len(row_ids) == 0:
				warning("The data table must have a column with unique row IDs.  These can be added with the 'Table/Add unique row ID' menu item.", {"parent":self.win})
			else:
				dlg = NMFUnmixingDialog(self, [r for r in self.data_types if r[0] not in row_ids], row_ids)
				self.unmixing_list.append(dlg)
				dlg.show()

		def cont_table():
			self.wait_for_data_types()
			dlg = ContTableDialog(self, self.data_types)
			self.conttable_list.append(dlg)
			dlg.show()

		def roc_curve():
			self.wait_for_data_types()
			dlg = ROCCurveDialog(self, self.data_types)
			self.roccurve_list.append(dlg)
			dlg.show()

		def online_help():
			webbrowser.open("https://mapdata.readthedocs.io/en/latest/", new=2, autoraise=True)
		def show_config_files():
			if len(config_files) == 0 and len(config_files_user) == 0:
				msg = "No configuration files have been read."
			else:
				if len(config_files) > 0:
					msg = "Configuration files read on startup:\n   %s" % "\n   ".join(config_files)
					if len(config_files_user) > 0:
						msg = msg + "\n\n"
				else:
					msg = ""
				if len(config_files_user) > 0:
					msg = msg + "Configuration files read after startup, in sequence:\n   %s" % "\n   ".join(config_files_user)
			dlg = MsgDialog("Config files", msg)
			dlg.show()
		def show_hotkeys():
			dlg = HelpHotkeysDialog()
			dlg.show()
		def show_about():
			message="""
                     mapdata.py

           version: %s, %s
      Copyright %s, R Dreas Nielsen
               License: GNU GPL3""" % (version, vdate, copyright)
			dlg = MsgDialog("About", message, can_resize=False)
			dlg.show()

		file_menu.add_command(label="Open CSV", command = self.new_data_file, underline=5)
		file_menu.add_command(label="Open spreadsheet", command = self.new_spreadsheet_file, underline=5)
		file_menu.add_command(label="Open database", command = self.new_db_table, underline=5)
		file_menu.add_command(label="Import symbol", command = import_symbol_file, underline=0)
		file_menu.add_command(label="Set editor", command = set_editor, underline=4)
		file_menu.add_command(label="Read config", command = read_config_file, underline=0)
		file_menu.add_command(label="Save config", command = save_config, underline=0)
		file_menu.add_command(label="Quit", command = self.cancel, underline=0)
		tbl_menu.add_command(label="Un-select all", command = self.unselect_map, underline=0)
		tbl_menu.add_command(label="Show selected", command = show_selected, underline=0)
		tbl_menu.add_command(label="Export selected", command = save_table, underline=1)
		tbl_menu.add_command(label="Export all", command = save_entire_table, underline=7)
		tbl_menu.add_command(label="Data types", command = self.show_data_types, underline=5)
		tbl_menu.add_command(label="Hide/show columns", command = self.hide_columns, underline=0)
		tbl_menu.add_command(label="Counts by location", command=self.counts_by_loc, underline=0)
		tbl_menu.add_command(label="Unique values", command=self.unique_data_values, underline=3)
		tbl_menu.add_command(label="Find candidate keys", command=self.find_candkeys, underline=0)
		tbl_menu.add_command(label="Find duplicate rows", command=self.find_duprows, underline=3)
		tbl_menu.add_command(label="Test cardinality", command=self.cardinality_test, underline=10)
		tbl_menu.add_command(label="Recode data", command=self.recode_data, underline=0)
		tbl_menu.add_command(label="Aggregate data", command=self.aggregate_data, underline=1)
		tbl_menu.add_command(label="Add unique row ID", command=self.add_rowid, underline=7)
		tbl_menu.add_command(label="Add coordinate key", command=self.add_coordinate_key, underline=5)
		tbl_menu.add_command(label="Calculate distances", command=self.calc_distances, underline=5)
		map_menu.add_command(label="Zoom selected", command = self.zoom_selected, underline=5)
		map_menu.add_command(label="Zoom full", command = self.zoom_full, underline=5)
		map_menu.add_command(label="Center on selection", command = self.focus_map, underline=0)
		map_menu.add_command(label="Un-select all", command = self.unselect_map, underline=0)
		map_menu.add_command(label="Change CRS", command = self.change_crs, underline=1)
		map_menu.add_command(label="Export", command = save_map, underline=1)
		map_menu.add_command(label="Settings", command = change_map_settings, underline=0)
		sel_menu.add_command(label="Invert", command = invert_selections, underline=0)
		sel_menu.add_command(label="Un-select all", command = self.unselect_map, underline=0)
		sel_menu.add_command(label="Data query", command = self.run_query, underline=5)
		sel_menu.add_command(label="Co-located data", command = self.select_colocated, underline=0)
		sel_menu.add_command(label="Non-missing values", command = self.select_nonmissing, underline=0)
		sel_menu.add_command(label="Random", command = self.select_random, underline=0)
		plot_menu.add_command(label="General", command = self.new_plot, underline=2)
		plot_menu.add_command(label="Pair plot", command = self.pairplot, underline=0)
		plot_menu.add_command(label="Variable distribution plot", command = self.variabledistplot, underline=0)
		plot_menu.add_command(label="Multi-variable line/scatter plot", command = self.multivarlineplot, underline=2)
		plot_menu.add_command(label="Multi-variable grouped bar chart", command = self.multivarbarplot, underline=15)
		plot_menu.add_command(label="Grouped percentage bar chart", command = self.multivargrppctchart, underline=5)
		plot_menu.add_command(label="Percent stacked area plot", command = self.pctstackedareaplot, underline=10)
		plot_menu.add_command(label="Rank-abundance plot", command = self.rankabundplot, underline=0)
		plot_menu.add_command(label="Close all", command = self.close_all_plots, underline=0)
		plot_menu.add_command(label="Settings", command = self.get_plot_config, underline=0)
		stats_menu.add_command(label="Univariate statistics", command = univar_stats, underline=0)
		stats_menu.add_command(label="Fit univariate distribution", command = fitdist_stats, underline=0)
		stats_menu.add_command(label="Bivariate statistics", command = bivar_stats, underline=0)
		stats_menu.add_command(label="Correlation matrix", command = corr_matrix, underline=0)
		stats_menu.add_command(label="Cosine similarity matrix", command = cos_matrix, underline=3)
		stats_menu.add_command(label="One-way ANOVA", command = anova_stats, underline=8)
		stats_menu.add_command(label="Contingency table", command = cont_table, underline=2)
		stats_menu.add_command(label="ROC curve", command = roc_curve, underline=0)
		stats_menu.add_command(label="PCA", command = pca, underline=0)
		stats_menu.add_command(label="t-SNE analysis", command = tsne_anal, underline=0)
		stats_menu.add_command(label="UMAP analysis", command = umap_anal, underline=1)
		stats_menu.add_command(label="NMF unmixing", command = unmixing, underline=0)
		stats_menu.add_command(label="Categorical correspondence", command = categ_corresp, underline=1)
		stats_menu.add_command(label="Categorical similarity matrix", command = sim_matrix, underline=12)
		help_menu.add_command(label="Online help", command = online_help, underline=7)
		help_menu.add_command(label="Config files", command = show_config_files, underline=0)
		help_menu.add_command(label="Hotkeys", command = show_hotkeys, underline=0)
		help_menu.add_command(label="About", command = show_about, underline=0)



class LoadingDialog(object):
	def __init__(self, parent):
		self.parent = parent
		self.dlg = tk.Toplevel(parent)
		self.dlg.title("MapData")
		#self.dlg.geometry("150x50")
		center_window(self.dlg)
		self.dlg.update_idletasks()
		self.dlg.withdraw()
		self.dlg.wm_overrideredirect(True)
		self.dlg.configure(bg="Gold")
		self.messages = []
		self.var_lbl = tk.StringVar(self.dlg, "")
		self.lbl_loading = tk.Label(self.dlg, bg="Gold", textvariable=self.var_lbl)
		self.lbl_loading.place(relx=0.5, rely=0.5, anchor="center")
		self.dlg.update()
		self.dots = 3
	def update_lbl(self):
		if len(self.messages) > 0:
			self.dots = self.dots % 3 + 1
			lbl = self.messages[0] + '.' * self.dots
			self.var_lbl.set(lbl)
			self.dlg.update()
			self.after_id = self.dlg.after(250, self.update_lbl)
	def display(self, message):
		self.messages.append(message)
		self.var_lbl.set(message)
		self.dlg.deiconify()
		raise_window(self.dlg)
		self.dlg.config(cursor="watch")
		self.dlg.update()
		#self.dlg.focus_force()
		#self.after_id = self.dlg.after(250, self.update_lbl)
	def hide(self):
		self.var_lbl.set("")
		#self.dlg.after_cancel(self.after_id)
		do_withdraw = True
		self.dlg.config(cursor="arrow")
		if len(self.messages) > 0:
			self.messages.pop(-1)
			if len(self.messages) > 0:
				self.var_lbl.set(self.messages[0])
				do_withdraw = False
		if do_withdraw:
			self.dlg.withdraw()
	def hide_all(self):
		self.var_lbl.set("")
		self.messages = []
		self.dlg.config(cursor="arrow")
		self.dlg.withdraw()


class Dialog(object):
	# A dialog template containing a prompt frame that is automataically
	# populated, a to ctrl_frame, a content_frame, and a btn_frame.
	# The dialog and the content_frame are expandable by default.
	def __init__(self, title, message, parent=None, msgwraplength=80, help_url=None, modal=False):
		# parent is the parent widget
		if parent is not None:
			self.dlg = tk.Toplevel(parent)
		else:
			self.dlg = tk.Toplevel()
		self.dlg.title(title)
		self.help_url = help_url
		self.modal = modal
		self.rv = None
		self.dlg.columnconfigure(0, weight=1)
		if message is not None:
			prompt_frame = tk.Frame(self.dlg)
			prompt_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(6,6))
			prompt_frame.columnconfigure(0, weight=1)
			msg_lbl = ttk.Label(prompt_frame, wraplength=msgwraplength, text=message)
			msg_lbl.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
			def wrap_msg(event):
				msg_lbl.configure(wraplength=event.width - 5)
			msg_lbl.bind("<Configure>", wrap_msg)
		self.ctrl_frame = tk.Frame(self.dlg)
		self.ctrl_frame.grid(row=1, column=0, sticky=tk.N+tk.EW)
		self.ctrl_frame.columnconfigure(0, weight=1)
		self.content_frame = tk.Frame(self.dlg)
		self.content_frame.grid(row=2, column=0, sticky=tk.NSEW)
		self.dlg.rowconfigure(2, weight=1)
		self.content_frame.rowconfigure(0, weight=1)
		self.content_frame.columnconfigure(0, weight=1)
		self.btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		self.btn_frame.columnconfigure(0, weight=1)
		self.btn_frame.grid(row=3, column=0, sticky=tk.EW)
		self.canceled = False
	def do_help(self, *args):
		if self.help_url is not None:
			webbrowser.open(self.help_url, new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		# Subclass this if self.rv needs to be set.
		self.canceled = False
		self.dlg.destroy()
	def show(self):
		if self.modal:
			self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.wait_window(self.dlg)
		if self.dlg is not None:
			try:
				self.dlg.destroy()
			except:
				pass
		return self.rv


class TableHideColsDialog(Dialog):
	def __init__(self, parent, column_list, displayed_columns):
		self.parent = parent
		self.column_list = column_list
		self.displayed_columns = displayed_columns
		super().__init__("Hide/Show Columns", "Choose the columns to show.\nAt least one column must be selected.",
				parent=parent, help_url="https://mapdata.readthedocs.io/en/latest/hide_cols.html")
		self.dlg.rowconfigure(1, weight=1)
		self.check_vars = [tk.StringVar(self.dlg, "1" if col in self.displayed_columns else "0") for col in self.column_list]

		# Controls
		ctlbtn_frame = tk.Frame(self.ctrl_frame)
		ctlbtn_frame.grid(row=0, column=0, sticky=tk.NW)
		ttk.Button(ctlbtn_frame, text="Hide", command=self.do_hide, underline=0).grid(row=0, column=0, sticky=tk.W, padx=(3,3), pady=(3,3))
		ttk.Button(ctlbtn_frame, text="Show", command=self.do_show, underline=0).grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))
		ck_frame = tk.Frame(self.ctrl_frame, borderwidth=3, relief=tk.RIDGE)
		ck_frame.grid(row=1, column=0, sticky=tk.NSEW)
		ck_frame.rowconfigure(0, weight=1)
		ck_frame.columnconfigure(0, weight=1)
		self.ctrl_frame.rowconfigure(1, weight=1)
		self.ctrl_frame.columnconfigure(0, weight=1)
		self.vs_frame = VScrollFrame(ck_frame, height=300)
		self.vs_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.content_frame = self.vs_frame.content_frame

		# Add checkboxes
		for i, colname in enumerate(self.column_list):
			ttk.Checkbutton(self.content_frame, text=colname, variable=self.check_vars[i], command=self.check_sel,
				onvalue="1", offvalue="0").grid(row=i, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))

		# Bottom buttons
		self.ok_btn = add_help_ok_cancel_buttons(self.dlg, self.btn_frame, self.do_help, self.do_select, self.do_cancel, ok_enabled=True)

	def do_hide(self):
		# Hide all but the first
		for col in self.check_vars[1:]:
			col.set("0")
	def do_show(self):
		for col in self.check_vars:
			col.set("1")
	def check_sel(self):
		# Ensure that at least one column is selected.
		shown = [cv.get() == "1" for cv in self.check_vars]
		if not any(shown):
			self.check_vars[0].set("1")
			self.vs_frame.scrolltop()
			warning("At least one column must be shown.", {"parent": self.dlg})
	def do_select(self, *args):
		self.rv = [self.column_list[i] for i in range(len(self.column_list)) if self.check_vars[i].get() == "1"]
		super().do_select(args)


	

class MapConfigDialog(Dialog):
	def __init__(self, parent, parent_widget, column_list):
		self.parent = parent
		self.column_list = column_list
		self.old_set = copy.copy(map_settings)
		fonts = list(set(list(tkfont.families())))
		fonts.sort()
		symbol_vals = list(icon_xbm.keys())
		symbol_vals.sort()
		color_vals = list(select_colors)
		columns = ['']
		columns.extend(column_list)
		super().__init__("Map Settings", message=None, parent=parent_widget,
				help_url= "https://mapdata.readthedocs.io/en/latest/map_settings.html")

		# Use a Notebook with left-side tabs.
		nbstyle = ttk.Style(self.dlg)
		nbstyle.configure("lefttab.TNotebook", tabposition="wn", tabplacement=tk.N+tk.EW)
		self.nb = ttk.Notebook(self.content_frame, style="lefttab.TNotebook")
		# Separate notebook pages for location symbols, selected location markers, and count labels
		pg_loc = tk.Frame(self.nb)
		pg_sel = tk.Frame(self.nb)
		pg_cnt = tk.Frame(self.nb)
		self.nb.add(pg_loc, text="Locations")
		self.nb.add(pg_sel, text="Selections")
		self.nb.add(pg_cnt, text="Counts")
		self.nb.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))

		# Populate the location page
		loc_instr_frame = tk.Frame(pg_loc)
		loc_instr_frame.grid(row=0, column=0, sticky=tk.NSEW)
		loc_prompt_frame = tk.Frame(pg_loc)
		loc_prompt_frame.grid(row=1, column=0, sticky=tk.NSEW)
		instr_lbl = ttk.Label(loc_instr_frame, text="Settings for location markers.")
		instr_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(12,12))
		loc_symbol_lbl = ttk.Label(loc_prompt_frame, text="Location symbol:")
		loc_symbol_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3))
		self.loc_symbol_var = tk.StringVar(self.dlg, map_settings.location_marker)
		self.loc_symbol_opts = ttk.Combobox(loc_prompt_frame, state="readonly", textvariable=self.loc_symbol_var,
				values=symbol_vals, width=15)
		self.loc_symbol_opts.grid(row=1, column=1, columnspan=3, sticky=tk.W, padx=(6,3))
		#
		loc_color_lbl = ttk.Label(loc_prompt_frame, text="Symbol color:")
		loc_color_lbl.grid(row=2, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.loc_color_var = tk.StringVar(self.dlg, map_settings.location_color)
		loc_color_opts = ttk.Combobox(loc_prompt_frame, state="readonly", textvariable=self.loc_color_var,
				values=list(select_colors), width=15)
		loc_color_opts.grid(row=2, column=1, columnspan=3, sticky=tk.W, padx=(6,3), pady=(3,3))
		#
		self.use_data_marker_var = tk.StringVar(self.dlg, map_settings.use_data_marker)
		ck_use_data_marker = ttk.Checkbutton(loc_prompt_frame, text="Use data symbol", variable=self.use_data_marker_var)
		ck_use_data_marker.grid(row=1, column=3, columnspan=2, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.use_data_color_var = tk.StringVar(self.dlg, map_settings.use_data_color)
		ck_use_color_marker = ttk.Checkbutton(loc_prompt_frame, text="Use data color", variable=self.use_data_color_var)
		ck_use_color_marker.grid(row=2, column=3, columnspan=2, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		self.loc_col_var = tk.StringVar(self.dlg, map_settings.label_col or '')
		loc_col_lbl = ttk.Label(loc_prompt_frame, text = "Label column:")
		loc_col_lbl.grid(row=3, column=0, sticky=tk.E, padx=(6,3), pady=(6,3))
		loc_col_opts = ttk.Combobox(loc_prompt_frame, state="readonly", textvariable=self.loc_col_var,
				values=columns, width=40)
		loc_col_opts.grid(row=3, column=1, columnspan=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		loc_lblcolor_lbl = ttk.Label(loc_prompt_frame, text="Label color:")
		loc_lblcolor_lbl.grid(row=4, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.loc_lblcolor_var = tk.StringVar(self.dlg, map_settings.label_color)
		loc_lblcolor_opts = ttk.Combobox(loc_prompt_frame, state="readonly", textvariable=self.loc_lblcolor_var,
				values=list(select_colors), width=15)
		loc_lblcolor_opts.grid(row=4, column=1, columnspan=3, sticky=tk.W, padx=(6,3), pady=(3,3))
		#
		self.loc_symbcol_var = tk.StringVar(self.dlg, map_settings.symbol_col or '')
		loc_symbcol_lbl = ttk.Label(loc_prompt_frame, text="Symbol column:")
		loc_symbcol_lbl.grid(row=5, column=0, sticky=tk.E, padx=(6,3), pady=(6,3))
		loc_symbcol_opts = ttk.Combobox(loc_prompt_frame, state="readonly", textvariable=self.loc_symbcol_var,
				values=columns, width=40)
		loc_symbcol_opts.grid(row=5, column=1, columnspan=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		self.loc_symbcolor_var = tk.StringVar(self.dlg, map_settings.color_col or '')
		loc_symbcolor_lbl = ttk.Label(loc_prompt_frame, text="Color column:")
		loc_symbcolor_lbl.grid(row=6, column=0, sticky=tk.E, padx=(6,3), pady=(6,3))
		loc_symbcolor_opts = ttk.Combobox(loc_prompt_frame, state="readonly", textvariable=self.loc_symbcolor_var,
				values=columns, width=40)
		loc_symbcolor_opts.grid(row=6, column=1, columnspan=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		self.loc_font_var = tk.StringVar(self.dlg, map_settings.label_font)
		loc_font_lbl = ttk.Label(loc_prompt_frame, text="Font:")
		loc_font_lbl.grid(row=7, column=0, sticky=tk.E, padx=(3,3), pady=(6,3))
		loc_font_opts = ttk.Combobox(loc_prompt_frame, state="readonly", textvariable=self.loc_font_var,
				values=fonts, width=25)
		loc_font_opts.grid(row=7, column=1, columnspan=3, sticky=tk.W, padx=(3,3), pady=(3,3))
		self.loc_bold_var = tk.StringVar(self.dlg, "0" if not map_settings.label_bold else "1")
		loc_ck_bold = ttk.Checkbutton(loc_prompt_frame, text="Bold", variable=self.loc_bold_var, onvalue="1", offvalue="0")
		loc_ck_bold.grid(row=7, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		self.loc_size_var = tk.IntVar(self.dlg, map_settings.label_size)
		loc_size_lbl = ttk.Label(loc_prompt_frame, text="Size:")
		loc_size_lbl.grid(row=8, column=0, sticky=tk.E, padx=(6,3), pady=(3,12))
		loc_size_opt = ttk.Combobox(loc_prompt_frame, state=tk.NORMAL, textvariable=self.loc_size_var,
				values=[8, 10, 12, 14, 16, 20, 24], width=3)
		loc_size_opt.grid(row=8, column=1, sticky=tk.W, padx=(6,3), pady=(3,12))
		self.loc_position_var = tk.StringVar(self.dlg, map_settings.label_position)
		loc_position_lbl = ttk.Label(loc_prompt_frame, text="Position:")
		loc_position_lbl.grid(row=8, column=2, sticky=tk.E, padx=(3,3), pady=(3,12))
		loc_position_sel = ttk.Combobox(loc_prompt_frame, state="readonly", textvariable=self.loc_position_var,
				values=["above", "below"], width=6)
		loc_position_sel.grid(row=8, column=3, sticky=tk.W, padx=(3,6), pady=(3,12))

		# Populate the selection marker page
		sel_instr_frame = tk.Frame(pg_sel)
		sel_instr_frame.grid(row=0, column=0, sticky=tk.NSEW)
		sel_prompt_frame = tk.Frame(pg_sel)
		sel_prompt_frame.grid(row=1, column=0, sticky=tk.NSEW)
		sel_instr_lbl = ttk.Label(sel_instr_frame, text="Settings for selection markers.")
		sel_instr_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(12,12))
		sel_symbol_lbl = ttk.Label(sel_prompt_frame, text="Marker symbol:")
		sel_symbol_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.sel_symbol_var = tk.StringVar(self.dlg, map_settings.select_symbol)
		self.sel_symbol_opts = ttk.Combobox(sel_prompt_frame, state="readonly", textvar=self.sel_symbol_var,
				values=symbol_vals, width=15)
		self.sel_symbol_opts.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		sel_color_lbl = ttk.Label(sel_prompt_frame, text="Color:")
		sel_color_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,12))
		self.sel_color_var = tk.StringVar(self.dlg, map_settings.select_color)
		sel_color_opts = ttk.Combobox(sel_prompt_frame, state="readonly", textvar=self.sel_color_var,
				values=color_vals, width=15)
		sel_color_opts.grid(row=1, column=1, sticky=tk.W, padx=(3,6), pady=(3,12))

		# Populate the count marker page
		cnt_instr_frame = tk.Frame(pg_cnt)
		cnt_instr_frame.grid(row=0, column=0, sticky=tk.NSEW)
		cnt_prompt_frame = tk.Frame(pg_cnt)
		cnt_prompt_frame.grid(row=1, column=0, sticky=tk.NSEW)
		cnt_instr_lbl = ttk.Label(cnt_instr_frame, text="Settings for rowcount labels.")
		cnt_instr_lbl.grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(12,12))
		#
		self.cnt_font_var = tk.StringVar(self.dlg, map_settings.count_label_font)
		cnt_font_lbl = ttk.Label(cnt_prompt_frame, text="Font:")
		cnt_font_lbl.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(6,3))
		cnt_font_opts = ttk.Combobox(cnt_prompt_frame, state="readonly", textvariable=self.cnt_font_var,
				values=fonts, width=25)
		cnt_font_opts.grid(row=0, column=1, columnspan=3, sticky=tk.W, padx=(3,3), pady=(3,3))
		#
		self.cnt_bold_var = tk.StringVar(self.dlg, "0" if not map_settings.count_label_bold else "1")
		cnt_ck_bold = ttk.Checkbutton(cnt_prompt_frame, text="Bold", variable=self.cnt_bold_var, onvalue="1", offvalue="0")
		cnt_ck_bold.grid(row=0, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		self.cnt_size_var = tk.IntVar(cnt_prompt_frame, map_settings.count_label_size)
		cnt_size_lbl = ttk.Label(cnt_prompt_frame, text="Size:")
		cnt_size_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		cnt_size_opt = ttk.Combobox(cnt_prompt_frame, state=tk.NORMAL, textvariable=self.cnt_size_var,
				values=[8, 10, 12, 14, 16, 20, 24], width=3)
		cnt_size_opt.grid(row=1, column=1, sticky=tk.W, padx=(6,3), pady=(3,3))
		#
		self.cnt_color_var = tk.StringVar(self.dlg, map_settings.count_label_color)
		cnt_color_lbl = ttk.Label(cnt_prompt_frame, text="Color:")
		cnt_color_lbl.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		cnt_color_opts = ttk.Combobox(cnt_prompt_frame, state="readonly", textvar=self.cnt_color_var,
				values=color_vals, width=15)
		cnt_color_opts.grid(row=2, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		self.cnt_offset_var = tk.IntVar(self.dlg, map_settings.count_label_offset)
		cnt_offset_lbl = ttk.Label(cnt_prompt_frame, text="Offset (px):")
		cnt_offset_lbl.grid(row=3, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		cnt_offset_entry = ttk.Spinbox(cnt_prompt_frame, textvariable=self.cnt_offset_var, from_=10, to=64, width=4)
		cnt_offset_entry.grid(row=3, column=1, sticky=tk.W, columnspan=2, padx=(6,3), pady=(3,3))
		#
		self.cnt_show_var = tk.StringVar(self.dlg, "0" if not map_settings.count_label_show else "1")
		cnt_ck_show = ttk.Checkbutton(cnt_prompt_frame, text="Show counts", variable=self.cnt_show_var, onvalue="1", offvalue="0")
		cnt_ck_show.grid(row=4, column=1, sticky=tk.W, padx=(3,6), pady=(3,12))

		# Buttons
		self.canceled = False
		self.ok_btn = add_help_ok_cancel_buttons(self.dlg, self.btn_frame, self.do_help, self.do_select, self.do_cancel, ok_enabled=True)
	def show(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.loc_symbol_opts.focus()
		self.dlg.wait_window(self.dlg)
		if not self.canceled:
			global map_settings
			map_settings.location_marker = self.loc_symbol_var.get()
			map_settings.location_color = self.loc_color_var.get()
			map_settings.use_data_marker = self.use_data_marker_var.get() == '1'
			map_settings.use_data_color = self.use_data_color_var.get() == '1'
			map_settings.label_col = self.loc_col_var.get() if self.loc_col_var.get() != '' else None
			map_settings.symbol_col = self.loc_symbcol_var.get() if self.loc_symbcol_var.get() != '' else None
			map_settings.color_col = self.loc_symbcolor_var.get() if self.loc_symbcolor_var.get() != '' else None
			map_settings.label_font = self.loc_font_var.get()
			map_settings.label_color = self.loc_lblcolor_var.get()
			map_settings.label_bold = self.loc_bold_var.get() == '1'
			map_settings.label_size = int(self.loc_size_var.get())
			map_settings.label_position = self.loc_position_var.get()
			if self.sel_symbol_var.get() != '':
				map_settings.select_symbol = self.sel_symbol_var.get()
			if self.sel_color_var.get() != '':
				map_settings.select_color = self.sel_color_var.get()
			map_settings.count_label_font = self.cnt_font_var.get()
			map_settings.count_label_bold = self.cnt_bold_var.get() == '1'
			map_settings.count_label_size = int(self.cnt_size_var.get())
			map_settings.count_label_color = self.cnt_color_var.get()
			map_settings.count_label_offset = int(self.cnt_offset_var.get())
			map_settings.count_label_show = self.cnt_show_var.get() == '1'

			if map_settings.label_col != self.old_set.label_col:
				global location_ids
				location_ids = {}

			redraw_loc_markers = map_settings.location_marker != self.old_set.location_marker \
					or map_settings.location_color != self.old_set.location_color \
					or map_settings.use_data_marker != self.old_set.use_data_marker \
					or map_settings.use_data_color != self.old_set.use_data_color 
			if not redraw_loc_markers and (map_settings.use_data_marker or map_settings.use_data_color) \
					and (map_settings.label_col != self.old_set.label_col \
					or map_settings.symbol_col != self.old_set.symbol_col \
					or map_settings.color_col != self.old_set.color_col):
						redraw_loc_markers = True
			if not redraw_loc_markers and (map_settings.label_font != self.old_set.label_font \
					or map_settings.label_color != self.old_set.label_color \
					or map_settings.label_bold != self.old_set.label_bold \
					or map_settings.label_size != self.old_set.label_size \
					or map_settings.label_position != self.old_set.label_position):
				redraw_loc_markers = True

			redraw_sel_markers = redraw_loc_markers or map_settings.select_symbol != self.old_set.select_symbol \
					or map_settings.select_color != self.old_set.select_color

			if redraw_loc_markers:
				self.parent.label_index = self.parent.headers.index(map_settings.label_col) if map_settings.label_col is not None and map_settings.label_col != '' else None
				self.parent.symbol_index = self.parent.headers.index(map_settings.symbol_col) if map_settings.symbol_col is not None and map_settings.symbol_col != '' else None
				self.parent.color_index = self.parent.headers.index(map_settings.color_col) if map_settings.color_col is not None and map_settings.color_col != '' else None
				self.parent.mapfont = makefont(map_settings.label_font, map_settings.label_size, map_settings.label_bold)
				self.parent.loc_marker_icon = set_get_loc_marker(map_settings.location_marker, map_settings.location_color)
				self.label_index = self.parent.headers.index(map_settings.label_col) if map_settings.label_col is not None else None
				self.parent.redraw_loc_markers(self.parent.tbl)

			if redraw_sel_markers:
				symb_name = "%s %s" % (map_settings.select_color, map_settings.select_symbol)
				if symb_name not in custom_icons:
					custom_icons[symb_name] = tk.BitmapImage(data=icon_xbm[map_settings.select_symbol], foreground=map_settings.select_color)
				self.parent.sel_marker_icon = custom_icons[symb_name]
				self.parent.redraw_sel_markers()

			if redraw_loc_markers or redraw_sel_markers or map_settings.count_label_show != self.old_set.count_label_show \
					or map_settings.count_label_font != self.old_set.count_label_font \
					or map_settings.count_label_bold != self.old_set.count_label_bold \
					or map_settings.count_label_size != self.old_set.count_label_size \
					or map_settings.count_label_color != self.old_set.count_label_color \
					or map_settings.count_label_offset != self.old_set.count_label_offset:
				self.parent.reset_count_labels_flag()
				self.parent.redraw_count_labels()


class ImportSymbolDialog(Dialog):
	def __init__(self):
		def get_fn(*args):
			fn = tkfiledialog.askopenfilename(filetypes=([('X11 bitmaps', '.xbm')]))
			if fn != '' and fn is not None and fn != ():
				self.fn_var.set(fn)
			self.dlg.lift()
		def check_enable(*args):
			enable_if(self.ok_btn, self.fn_var.get() != '' and self.symbol_var.get() != '')
		super().__init__("Import X11 Bitmap Symbol", message=None, help_url="https://mapdata.readthedocs.io/en/latest/import_symbol.html")
		self.rv = (None, None)
		# Prompts
		symbol_lbl = ttk.Label(self.content_frame, text="Symbol name:")
		symbol_lbl.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.symbol_var = tk.StringVar(self.dlg, "")
		self.symbol_var.trace('w', check_enable)
		self.symbol_entry = ttk.Entry(self.content_frame, textvariable=self.symbol_var, width=12)
		self.symbol_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		#
		fn_label = ttk.Label(self.content_frame, text="File:")
		fn_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3))
		self.fn_var = tk.StringVar(self.content_frame, '')
		self.fn_var.trace('w', check_enable)
		fn_entry = ttk.Entry(self.content_frame, textvariable=self.fn_var)
		fn_entry.configure(width=64)
		fn_entry.grid(row=1, column=1, sticky=tk.EW, padx=(3,3))
		fn_btn = new_button(self.content_frame, "Browse", 1, 2, get_fn)
		# Buttons
		self.canceled = False
		self.ok_btn = add_help_ok_cancel_buttons(self.dlg, self.btn_frame, self.do_help, self.do_select, self.do_cancel)
		self.symbol_entry.focus()
		self.dlg.resizable(False, False)
	def do_select(self, *args):
		if self.ok_btn["state"] != tk.DISABLED:
			self.canceled = False
			self.rv = (self.symbol_var.get(), self.fn_var.get())
			self.dlg.destroy()


class DataFileDialog(object):
	def __init__(self):
		def get_fn():
			fn = tkfiledialog.askopenfilename(filetypes=([('CSV files', '.csv')]), parent=self.dlg)
			if fn != '' and fn != () and fn is not None:
				self.fn_var.set(fn)
				csvreader = CsvFile(fn)
				self.header_list = csvreader.next()
				self.id_sel["values"] = self.header_list
				self.lat_sel["values"] = self.header_list
				self.lon_sel["values"] = self.header_list
				self.sym_sel["values"] = self.header_list
				self.col_sel["values"] = self.header_list
			self.dlg.lift()
		def check_enable(*args):
			enable_if(ok_btn, self.fn_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '')
		def new_fn(*args):
			check_enable()
		self.header_list = []
		self.dlg = tk.Toplevel()
		self.dlg.title("Open CSV Data File for Map Display")
		# Main frames
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		dir_frame = tk.Frame(prompt_frame)
		dir_frame.grid(row=0, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		dir_frame.rowconfigure(0, weight=1)
		dir_frame.columnconfigure(0, weight=1)
		req_frame = ttk.LabelFrame(prompt_frame, text="Required")
		req_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,3), pady=(3,3))
		req_frame.columnconfigure(0, weight=1)
		opt_frame = ttk.LabelFrame(prompt_frame, text="Optional")
		opt_frame.grid(row=2, column=0, sticky=tk.EW, padx=(6,3), pady=(9,3))
		opt_frame.columnconfigure(0, weight=1)
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		# Prompts
		#-- Directions
		dir_lbl = ttk.Label(dir_frame,
				text="Select a CSV file with columns containing latitude and longitude values, and optionally other information.",
				width=80, justify=tk.LEFT, wraplength=600)
		dir_lbl.grid(row=0, column=0, padx=(3,3), pady=(3,3))
		def wrap_msg(event):
			dir_lbl.configure(wraplength=event.width - 5)
		dir_lbl.bind("<Configure>", wrap_msg)
		#-- Filename
		fn_frame = tk.Frame(req_frame)
		fn_frame.grid(row=0, column=0, sticky=tk.EW, pady=(5,5))
		fn_label = ttk.Label(fn_frame, text="File:")
		fn_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.fn_var = tk.StringVar(fn_frame, '')
		self.fn_var.trace('w', new_fn)
		fn_entry = ttk.Entry(fn_frame, textvariable=self.fn_var)
		fn_entry.configure(width=64)
		fn_entry.grid(row=0, column=1, sticky=tk.EW, padx=(3,3))
		fn_btn = ttk.Button(fn_frame, text="Browse", command=get_fn)
		fn_btn.grid(row=0, column=2, sticky=tk.W, padx=(3,3))
		#-- Required columns
		column_choices = list(self.header_list)
		#
		req_col_frame = tk.Frame(req_frame)
		req_col_frame.grid(row=1, column=0, sticky=tk.EW, pady=(3,3))
		lat_label = ttk.Label(req_col_frame, text="Latitude column:")
		lat_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lat_var = tk.StringVar(req_col_frame, '')
		self.lat_var.trace('w', check_enable)
		self.lat_sel = ttk.Combobox(req_col_frame, state="readonly", textvariable=self.lat_var, values=self.header_list, width=24)
		self.lat_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))
		#
		lon_label = ttk.Label(req_col_frame, text="Longitude column:")
		lon_label.grid(row=0, column=2, sticky=tk.E, padx=(20,3), pady=(3,3))
		self.lon_var = tk.StringVar(req_frame, '')
		self.lon_var.trace('w', check_enable)
		self.lon_sel = ttk.Combobox(req_col_frame, state="readonly", textvariable=self.lon_var, values=self.header_list, width=24)
		self.lon_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,3), pady=(3,3))
		#-- Optional columns
		opt_col_frame = tk.Frame(opt_frame)
		opt_col_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		id_label = ttk.Label(opt_col_frame, text="Label column:")
		id_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.id_var = tk.StringVar(opt_col_frame, '')
		self.id_sel = ttk.Combobox(opt_col_frame, state="readonly", textvariable=self.id_var, values=self.header_list, width=24)
		self.id_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,20), pady=(3,3))
		#
		crs_label = ttk.Label(opt_col_frame, text="CRS:")
		crs_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.crs_var = tk.IntVar(opt_col_frame, 4326)
		self.crs_var.trace('w', check_enable)
		self.crs_sel = ttk.Entry(opt_col_frame, width=8, textvariable=self.crs_var)
		self.crs_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,20), pady=(3,3))
		#
		sym_label = ttk.Label(opt_col_frame, text="Symbol column:")
		sym_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.sym_var = tk.StringVar(opt_col_frame, '')
		self.sym_sel = ttk.Combobox(opt_col_frame, state="readonly", textvariable=self.sym_var, values=self.header_list, width=24)
		self.sym_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,20), pady=(3,3))
		#
		col_label = ttk.Label(opt_col_frame, text="Color column:")
		col_label.grid(row=1, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.col_var = tk.StringVar(opt_col_frame, '')
		self.col_sel = ttk.Combobox(opt_col_frame, state="readonly", textvariable=self.col_var, values=self.header_list, width=24)
		self.col_sel.grid(row=1, column=3, sticky=tk.W, padx=(3,20), pady=(3,3))
		#-- Description
		title_label = ttk.Label(opt_col_frame, text="Description:")
		title_label.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.title_var = tk.StringVar(opt_col_frame, '')
		title_entry = ttk.Entry(opt_col_frame, width=60, textvariable=self.title_var)
		title_entry.grid(row=2, column=1, columnspan=3, sticky=tk.EW, padx=(3,6), pady=(3,3))
		# Buttons
		self.canceled = False
		ok_btn = add_help_ok_cancel_buttons(self.dlg, btn_frame, self.do_help, self.do_select, self.do_cancel)
		self.dlg.resizable(False, False)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/importingdata.html#importing-from-a-csv-file", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.fn_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def get_datafile(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.dlg.wait_window(self.dlg)
		self.dlg = None
		if not self.canceled:
			headers, rows = file_data(self.fn_var.get())
			return (self.fn_var.get(), self.id_var.get(), self.lat_var.get(), self.lon_var.get(),
					self.crs_var.get(), self.sym_var.get(), self.col_var.get(), self.title_var.get(),
					headers, rows)
		else:
			return (None, None, None, None, None, None, None, None, None, None)



class ImportSpreadsheetDialog(object):
	def __init__(self, parent, mapui):
		self.parent = parent
		self.mapui = mapui
		def get_fn(*args):
			fn = tkfiledialog.askopenfilename(filetypes=([('Spreadsheets', '.ods .xlsx .xls')]), parent=self.dlg)
			if fn != '' and fn != () and fn is not None:
				self.fn_var.set(fn)
		def check_w1enable(*args):
			if self.fn_var.get() != '':
				enable_if(w1next_btn, os.path.isfile(self.fn_var.get()))
			else:
				w1next_btn["state"] = tk.DISABLED
		def check_w2enable(*args):
			enable_if(w2next_btn, self.fn_var.get() != '' and self.sheet_var.get() != '')
		def check_w3enable(*args):
			enable_if(w3ok_btn, self.fn_var.get() != '' and self.sheet_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '')
		def new_fn(*args):
			check_w1enable()
		self.sheet_list = []
		self.header_list = []
		self.dlg = tk.Toplevel(parent)
		self.dlg.title("Open Spreadsheet File for Map Display")
		# Main frames
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		# Wizard frames 1, 2, and 3 are gridded in the same cell to make a wizard.
		self.dlg.rowconfigure(0, weight=0)
		wiz1_frame = tk.Frame(self.dlg)		# For description, filename, and sheet name
		wiz1_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		wiz1_frame.rowconfigure(0, weight=1)
		wiz2_frame = tk.Frame(self.dlg)		# For sheet selections
		wiz2_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		wiz2_frame.rowconfigure(0, weight=1)
		wiz2_frame.columnconfigure(0, weight=1)
		wiz3_frame = tk.Frame(self.dlg)		# For column selections
		wiz3_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		wiz3_frame.columnconfigure(0, weight=1)
		self.dlg.rowconfigure(1, weight=0)
		self.dlg.resizable(False, False)
		wiz1_frame.lift()

		# Populate directions frame
		dir_lbl = ttk.Label(prompt_frame,
				text="Select a spreadsheet file with columns containing latitude and longitude values, and optionally other information.",
				width=80, justify=tk.LEFT, wraplength=600)
		dir_lbl.grid(row=0, column=0, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			dir_lbl.configure(wraplength=event.width - 5)
		dir_lbl.bind("<Configure>", wrap_msg)

		# Populate wiz1_frame
		w1req_frame = ttk.LabelFrame(wiz1_frame, text="Required")
		w1req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w1req_frame.columnconfigure(0, weight=1)
		fn_frame = tk.Frame(w1req_frame)
		fn_frame.grid(row=0, column=0, sticky=tk.EW, pady=(3,3))
		fn_label = ttk.Label(fn_frame, text="File:")
		fn_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.fn_var = tk.StringVar(fn_frame, '')
		self.fn_var.trace('w', new_fn)
		fn_entry = ttk.Entry(fn_frame, textvariable=self.fn_var)
		fn_entry.configure(width=64)
		fn_entry.grid(row=0, column=1, sticky=tk.EW, padx=(3,3))
		fn_btn = ttk.Button(fn_frame, text="Browse", command=get_fn, underline=0)
		fn_btn.grid(row=0, column=2, sticky=tk.W, padx=(3,3))
		self.dlg.bind("<Alt-b>", get_fn)

		w1opt_frame = ttk.LabelFrame(wiz1_frame, text="Optional")
		w1opt_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,3), pady=(9,3))
		w1opt_frame.columnconfigure(0, weight=1)
		desc_label = ttk.Label(w1opt_frame, text="Description:")
		desc_label.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.desc_var = tk.StringVar(w1opt_frame, '')
		desc_entry = ttk.Entry(w1opt_frame, width=60, textvariable=self.desc_var)
		desc_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))

		def w1_next(*args):
			if self.fn_var.get() != '':
				# Open spreadsheet, get sheet names
				self.mapui.loading_dlg.display("Importing spreadsheet")
				fn, ext = os.path.splitext(self.fn_var.get())
				ext = ext.lower()
				try:
					if ext == '.ods':
						sso = OdsFile()
					elif ext == '.xlsx':
						sso = XlsxFile()
					else:
						sso = XlsFile()
				except:
					warning("Could not open %s" % self.fn_var.get(), kwargs={'parent': self.dlg})
				else:
					sso.open(self.fn_var.get())
					self.sheet_list = sso.sheetnames()
					self.sheet_sel["values"] = self.sheet_list
					if ext in ('.ods', '.xlsx'):
						try:
							sso.close()
						except:
							pass
					else:
						try:
							sso.release_resources()
							del sso
						except:
							pass
					self.dlg.bind("<Alt-b>")
					self.dlg.bind("<Alt-n>")
					wiz2_frame.lift()
					self.dlg.bind("<Alt-b>", w2_back)
					self.dlg.bind("<Alt-n>", w2_next)
				self.mapui.loading_dlg.hide()

		w1btn_frame = tk.Frame(wiz1_frame, borderwidth=3, relief=tk.RIDGE)
		w1btn_frame.grid(row=2, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		w1btn_frame.columnconfigure(0, weight=1)
		self.canceled = False
		#
		w1help_btn = new_button(w1btn_frame, "Help", 0, 0, self.do_help, tk.W, (6,3), underline=0)
		self.dlg.bind("<Alt-h>", self.do_help)
		w1next_btn = new_button(w1btn_frame, "Next", 0, 1, w1_next, tk.E, (3,3), underline=0, state=tk.DISABLED)
		self.dlg.bind("<Alt-n>", w1_next)
		w1cancel_btn = new_button(w1btn_frame, "Cancel", 0, 2, self.do_cancel, tk.E, (3,6), underline=0)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Escape>", self.do_cancel)


		# Populate wiz2_frame
		w2req_frame = ttk.LabelFrame(wiz2_frame, text="Required")
		w2req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w2req_frame.columnconfigure(0, weight=1)
		w2inner_frame = tk.Frame(w2req_frame)
		w2inner_frame.grid(row=0, column=0, sticky=tk.W)
		#
		sheet_label = ttk.Label(w2inner_frame, text="Sheet:")
		sheet_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.sheet_var = tk.StringVar(w2req_frame, '')
		self.sheet_var.trace('w', check_w2enable)
		self.sheet_sel = ttk.Combobox(w2inner_frame, state="readonly", textvariable=self.sheet_var, values=self.sheet_list, width=24)
		self.sheet_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		#
		xrows_label = ttk.Label(w2inner_frame, text="Initial rows to skip:")
		xrows_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3))
		self.xrows_var = tk.IntVar(w2req_frame, 0)
		self.xrows_var.trace('w', check_w2enable)
		xrows_entry = ttk.Entry(w2inner_frame, textvariable=self.xrows_var, width=6)
		xrows_entry.grid(row=1, column=1, sticky=tk.W, padx=(3,3))

		def w2_back(*args):
			self.dlg.bind("<Alt-b>")
			self.dlg.bind("<Alt-n>")
			wiz1_frame.lift()
			self.dlg.bind("<Alt-n>", w1_next)
			self.dlg.bind("<Alt-b>", get_fn)

		def w2_next(*args):
			# Open spreadsheet, get column names
			if self.fn_var.get() != '' and self.sheet_var.get() != '':
				fn, ext = os.path.splitext(self.fn_var.get())
				try:
					if ext.lower() == '.ods':
						hdrs, data = ods_data(self.fn_var.get(), self.sheet_var.get(), junk_header_rows=self.xrows_var.get())
					else:
						hdrs, data = xls_data(self.fn_var.get(), self.sheet_var.get(), junk_header_rows=self.xrows_var.get())
				except:
					warning("Could not read table from %s, sheet %s" % (self.fn_var.get(), self.sheet_var.get()), 
							kwargs={'parent': self.dlg})
				else:
					self.headers = hdrs
					self.header_list = list(hdrs)
					self.rows = data
					# Set list box values
					self.id_sel["values"] = self.header_list
					self.lat_sel["values"] = self.header_list
					self.lon_sel["values"] = self.header_list
					self.sym_sel["values"] = self.header_list
					self.col_sel["values"] = self.header_list
					self.dlg.bind("<Alt-b>")
					self.dlg.bind("<Alt-n>")
					wiz3_frame.lift()
					self.dlg.bind("<Alt-b>", w3_back)

		w2btn_frame = tk.Frame(wiz2_frame, borderwidth=3, relief=tk.RIDGE)
		w2btn_frame.grid(row=2, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		w2btn_frame.columnconfigure(0, weight=1)
		#
		w2help_btn = new_button(w2btn_frame, "Help", 0, 0, self.do_help, tk.W, (6,3), underline=0)
		self.dlg.bind("<Alt-h>", self.do_help)
		w2prev_btn = new_button(w2btn_frame, "Back", 0, 1, w2_back, tk.E, (3,3), underline=0)
		w2next_btn = new_button(w2btn_frame, "Next", 0, 2, w2_next, tk.E, (3,3), underline=0, state=tk.DISABLED)
		w2cancel_btn = new_button(w2btn_frame, "Cancel", 0, 3, self.do_cancel, tk.E, (3,6), underline=0)
	
		# Populate wiz3_frame
		w3req_frame = ttk.LabelFrame(wiz3_frame, text="Required")
		w3req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w3req_frame.columnconfigure(0, weight=1)
		#
		lat_label = ttk.Label(w3req_frame, text="Latitude column:")
		lat_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lat_var = tk.StringVar(w3req_frame, '')
		self.lat_var.trace('w', check_w3enable)
		self.lat_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lat_var, values=self.header_list, width=24)
		self.lat_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		lon_label = ttk.Label(w3req_frame, text="Longitude column:")
		lon_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lon_var = tk.StringVar(w3req_frame, '')
		self.lon_var.trace('w', check_w3enable)
		self.lon_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lon_var, values=self.header_list, width=24)
		self.lon_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))


		w3opt_frame = ttk.LabelFrame(wiz3_frame, text="Optional")
		w3opt_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(9,3))
		w3opt_frame.columnconfigure(0, weight=1)
		#
		id_label = ttk.Label(w3opt_frame, text="Label column:")
		id_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.id_var = tk.StringVar(w3opt_frame, '')
		self.id_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.id_var, values=self.header_list, width=24)
		self.id_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		crs_label = ttk.Label(w3opt_frame, text="CRS:")
		crs_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.crs_var = tk.IntVar(w3opt_frame, 4326)
		self.crs_var.trace('w', check_w2enable)
		self.crs_sel = ttk.Entry(w3opt_frame, width=8, textvariable=self.crs_var)
		self.crs_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		sym_label = ttk.Label(w3opt_frame, text="Symbol column:")
		sym_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.sym_var = tk.StringVar(w3opt_frame, '')
		self.sym_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.sym_var, values=self.header_list, width=24)
		self.sym_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		col_label = ttk.Label(w3opt_frame, text="Color column:")
		col_label.grid(row=1, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.col_var = tk.StringVar(w3opt_frame, '')
		self.col_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.col_var, values=self.header_list, width=24)
		self.col_sel.grid(row=1, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))

		def w3_back(*args):
			self.dlg.bind("<Alt-b>")
			wiz2_frame.lift()
			self.dlg.bind("<Alt-b>", w2_back)
			self.dlg.bind("<Alt-n>", w2_next)

		w3btn_frame = tk.Frame(wiz3_frame, borderwidth=3, relief=tk.RIDGE)
		w3btn_frame.grid(row=2, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		w3btn_frame.columnconfigure(0, weight=1)
		#
		w3help_btn = new_button(w3btn_frame, "Help", 0, 0, self.do_help, tk.W, (6,3), underline=0)
		self.dlg.bind("<Alt-h>", self.do_help)
		w2prev_btn = new_button(w3btn_frame, "Back", 0, 1, w3_back, tk.E, (3,3), underline=0)
		w3ok_btn = new_button(w3btn_frame, "OK", 0, 2, self.do_select, tk.E, (3,3), underline=0, state=tk.DISABLED)
		self.dlg.bind('<Alt-o>', self.do_select)
		w3cancel_btn = new_button(w3btn_frame, "Cancel", 0, 3, self.do_cancel, tk.E, (3,6), underline=0)
	
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/importingdata.html#importing-from-a-spreadsheet-file", new=2, autoraise=True)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		if self.fn_var.get() != '' and self.sheet_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def get_datafile(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		self.dlg.resizable(False, False)
		self.dlg.wait_window(self.dlg)
		self.dlg = None
		if not self.canceled:
			return (self.fn_var.get(), self.id_var.get(), self.lat_var.get(), self.lon_var.get(),
					self.crs_var.get(), self.sym_var.get(), self.col_var.get(), self.desc_var.get(),
					self.headers, self.rows)
		else:
			return (None, None, None, None, None, None, None, None, None, None)


class NewCrsDialog(Dialog):
	def __init__(self, current_crs):
		super().__init__("Change CRS", message=None, help_url= "https://mapdata.readthedocs.io/en/latest/change_crs.html")
		crs_lbl = ttk.Label(self.content_frame, text="New CRS:")
		crs_lbl.grid(row=0, column=0, sticky=tk.E, padx=(3,3))
		self.crs_var = tk.IntVar(self.dlg, current_crs)
		self.crs_entry = ttk.Entry(self.content_frame, width=12, textvariable=self.crs_var)
		self.crs_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		# Buttons
		ok_btn = add_help_ok_cancel_buttons(self.dlg, self.btn_frame, self.do_help, self.do_select, self.do_cancel, True)
		self.dlg.resizable(False, False)
		self.crs_entry.focus()
	def do_select(self, *args):
		self.rv = self.crs_var.get()
		super().do_select(args)


class QueryDialog(Dialog):
	def __init__(self, parent, column_headers, db_conn, init_sql=""):
		# parent is the class object, not the parent widget
		self.parent = parent
		super().__init__("Query Data",
				"Enter an expression below to identify the data rows that you want to select.  The syntax of this expression should correspond to a SQL 'WHERE' clause.  Column names with non-alphanumeric characters should be double-quoted.  String literals should be single-quoted.  The '%' character is a wildcard.",
				msgwraplength=300, help_url="https://mapdata.readthedocs.io/en/latest/sql_select.html")
		self.scriptfilepath = None
		self.dlg.columnconfigure(0, weight=1, minsize=620)
		#self.dlg.rowconfigure(0, weight=1, minsize=180)
		self.applied = False
		self.ever_applied = False
		self.sql = init_sql
		# Frames
		query_frame = tk.Frame(self.content_frame)
		query_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		query_frame.rowconfigure(1, weight=1)
		query_frame.columnconfigure(0, weight=0)
		query_frame.columnconfigure(1, weight=3)
		col_frame = tk.Frame(self.content_frame)
		col_frame.grid(row=0, column=1, rowspan=2, sticky=tk.NS, padx=(3,3), pady=(3,3))
		col_frame.rowconfigure(0, weight=1)
		chars_frame = tk.Frame(query_frame)
		chars_frame.grid(row=0, column=1, columnspan=2, sticky = tk.EW, padx=(3,3))
		edit_frame = tk.Frame(query_frame)
		edit_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,0), pady=(3,3))
		sql_frame = tk.Frame(query_frame)
		sql_frame.grid(row=1, column=1, sticky=tk.NSEW, padx=(1,3), pady=(3,1))
		sql_frame.rowconfigure(0, weight=1, minsize=75)
		# Setting the minsize should prevent resizing that crushes the buttons, but it does not.
		sql_frame.columnconfigure(0, weight=1, minsize=200)
		act_frame = tk.Frame(query_frame)
		act_frame.grid(row=2, column=0, columnspan=2, sticky=tk.EW, padx=(3,3), pady=(1,3))
		# SQL text-insertion buttons
		ttk.Button(chars_frame, text="=", width=1, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, "="))).grid(row=0, column=0, sticky=tk.W)
		ttk.Button(chars_frame, text="<>", width=3, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, "<>"))).grid(row=0, column=1, sticky=tk.W)
		ttk.Button(chars_frame, text="<", width=1, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, "<"))).grid(row=0, column=2, sticky=tk.W)
		ttk.Button(chars_frame, text=">", width=1, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, ">"))).grid(row=0, column=3, sticky=tk.W)
		ttk.Button(chars_frame, text="not", width=3, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, " not "))).grid(row=0, column=4, sticky=tk.W)
		ttk.Button(chars_frame, text="in", width=2, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, " in "))).grid(row=0, column=5, sticky=tk.W)
		ttk.Button(chars_frame, text="(", width=1, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, " ("))).grid(row=0, column=6, sticky=tk.W)
		ttk.Button(chars_frame, text=",", width=1, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, ","))).grid(row=0, column=7, sticky=tk.W)
		ttk.Button(chars_frame, text=")", width=1, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, ")"))).grid(row=0, column=8, sticky=tk.W)
		ttk.Button(chars_frame, text="'", width=1, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, "'"))).grid(row=0, column=9, sticky=tk.W)
		ttk.Button(chars_frame, text="and", width=3, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, " and "))).grid(row=0, column=10, sticky=tk.W)
		ttk.Button(chars_frame, text="or", width=2, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, " or "))).grid(row=0, column=11, sticky=tk.W)
		ttk.Button(chars_frame, text="like", width=4, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, " like "))).grid(row=0, column=12, sticky=tk.W)
		ttk.Button(chars_frame, text="%", width=2, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, "%"))).grid(row=0, column=13, sticky=tk.W)
		ttk.Button(chars_frame, text="_", width=1, command=lambda:self.exec(lambda:self.sql_text.insert(tk.END, "_"))).grid(row=0, column=14, sticky=tk.W)
		# Clear, Load, Edit, and Save buttons
		clear_btn = new_button(edit_frame, "Clear", 1, 0, self.clear_sql, tk.E, (0,1), underline=1)
		load_btn = new_button(edit_frame, "Open", 2, 0, self.load_script, tk.E, (0,1), underline=1)
		save_btn = new_button(edit_frame, "Save", 3, 0, self.save_script, tk.E, (0,1), underline=0)
		edit_btn = new_button(edit_frame, "Edit", 4, 0, self.edit_sql, tk.E, (0,1), (0,2), underline=0)
		enable_if(edit_btn, editor is not None)

		# SQL text entry
		self.sql_text = tk.Text(sql_frame, width=60, height=10)
		if init_sql is not None and init_sql != "":
			self.sql_text.insert(tk.END, init_sql)
		self.sql_text.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,0), pady=(3,3))
		self.sql_text.bind("<KeyRelease>", self.check_enable)
		sbar = tk.Scrollbar(sql_frame)
		sbar.grid(row=0, column=1, sticky=tk.NS, padx=(0,3), pady=(3,3))
		sbar.config(command=self.sql_text.yview)
		self.sql_text.config(yscrollcommand = sbar.set)
		# Column values
		col_lbl = ttk.Label(col_frame, text="Column values:")
		col_lbl.grid(row=0, column=0, sticky=tk.NW, padx=(3,3), pady=(3,3))
		col_var = tk.StringVar(col_frame, "")
		colsel = ttk.Combobox(col_frame, state="readonly", textvariable=col_var, values=db_colnames(column_headers), width=20)
		colsel.grid(row=1, column=0, sticky=tk.NW, padx=(3,3), pady=(3,3))
		def copycol():
			cval = col_var.get()
			if cval != '':
				self.sql_text.insert(tk.END, cval)
				self.ok_btn["state"] = tk.NORMAL
				self.apply_btn["state"] = tk.NORMAL
		self.colcopy = ttk.Button(col_frame, state=tk.DISABLED, text="\u2B05", width=2, command=copycol)
		self.colcopy.grid(row=1, column=1, sticky=tk.W, pady=(3,3))
		tv_frame = tk.Frame(col_frame)
		tv_frame.grid(row=2, column=0, sticky=tk.NS, padx=(3,3), pady=(3,3))
		col_frame.rowconfigure(0, weight=0)
		col_frame.rowconfigure(1, weight=0)
		col_frame.rowconfigure(2, weight=1)
		col_frame.columnconfigure(2, weight=1)
		def colval_to_sql(event):
			item_iid = self.tv_tbl.identify('item', event.x, event.y)
			item_val = self.tv_tbl.item(item_iid, "values")[0]
			if not isfloat(item_val):
				item_val = "'%s'" % item_val
			self.sql_text.insert(tk.END, " "+item_val)
			self.ok_btn["state"] = tk.NORMAL
			self.sql_text.focus()
		def list_col_vals(event):
			curs = db_conn.cursor()
			colname = dquote(col_var.get())
			res = curs.execute('SELECT DISTINCT %s FROM mapdata ORDER BY %s' % (colname, colname))
			rowset = res.fetchall()
			for widget in tv_frame.winfo_children():
				widget.destroy()
			tblframe, self.tv_tbl = treeview_table(tv_frame, rowset, [col_var.get()])
			tblframe.grid(column=0, row=0, sticky=tk.NSEW)
			tv_frame.rowconfigure(0, weight=1)
			tv_frame.columnconfigure(0, weight=1)
			curs.close()
			self.tv_tbl.bind("<Double-1>", colval_to_sql)
			self.colcopy["state"] = tk.NORMAL
		colsel.bind("<<ComboboxSelected>>", list_col_vals)
		# Action selection
		self.act_var = tk.StringVar(act_frame, "Replace")
		act_lbl = ttk.Label(act_frame, text="Action:")
		act_lbl.grid(row=0, column=0, sticky=tk.E, padx=(35,3))
		act_sel = ttk.Combobox(act_frame, state="readonly", textvariable=self.act_var, values=["Replace", "Union", "Intersection", "Difference O-N", "Difference N-O"], width=15)
		act_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6))
		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.apply_btn = new_button(self.btn_frame, "Apply", 0, 1, self.do_apply, tk.E, (3,3), (0,0), 0)
		if init_sql is None or len(init_sql) == 0:
			self.apply_btn["state"] = tk.DISABLED
		self.dlg.bind('<Alt-a>', self.do_apply)
		self.ok_btn = new_ok_button(self.dlg, self.btn_frame, 2, self.do_select)
		if init_sql is None or len(init_sql) == 0:
			self.ok_btn["state"] = tk.DISABLED
		self.cancel_btn = new_cancel_button(self.dlg, self.btn_frame, 3, self.do_cancel)
		self.dlg.unbind("<Return>")
		self.dlg.bind("<Control-Return>", self.do_select)
		self.sql_text.focus()
	def check_enable(self, args=None):
		curr_sql = self.sql_text.get("1.0", "end-1c")
		enable_if(self.ok_btn, curr_sql != '')
		enable_if(self.apply_btn, curr_sql != '')
		if curr_sql != self.sql:
			self.applied = False
	def exec(self, command):
		command()
		self.check_enable()
	def clear_sql(self):
		self.sql_text.delete(1.0, tk.END)
		self.ok_btn["state"] = tk.DISABLED
		self.apply_btn["state"] = tk.DISABLED
		self.applied = False
		self.sql_text.focus()
	def load_script(self):
		fn = tkfiledialog.askopenfilename(parent=self.dlg, title="SQL script file to open", filetypes=([('SQL script files', '.sql')]))
		if not (fn is None or fn == '' or fn == ()):
			path, filename = os.path.split(os.path.abspath(fn))
			self.scriptfilepath = path
			with open(fn, "r") as f:
				sql = f.read()
			self.sql_text.insert("end", sql)
			self.check_enable()
		self.sql_text.focus()
	def save_script(self):
		outfile = tkfiledialog.asksaveasfilename(initialdir=self.scriptfilepath, parent=self.dlg, title="SQL script file to save", filetypes=[('SQL script files', '.sql')])
		if not (outfile is None or outfile == ''):
			sql = self.sql_text.get("1.0", "end")
			with open(outfile, "w") as f:
				f.write(sql)
		self.sql_text.focus()
	def edit_sql(self):
		td = tempfile.TemporaryDirectory()
		edit_fn = os.path.join(td.name, "mapfile_temp.sql")
		with open(edit_fn, "w") as f:
			f.write(self.sql_text.get("1.0", "end"))
		returncode = subprocess.call([editor, edit_fn])
		if returncode == 0:
			with open(edit_fn, "r") as f:
				sql = f.read()
			self.sql_text.delete("1.0", "end")
			self.sql_text.insert("end", sql)
			self.check_enable()
		else:
			warning("Failure attempting to edit the SQL with %s" % editor, kwargs={'parent':self.dlg})
		self.sql_text.focus()
	def do_apply(self, *args):
		success = False
		whereclause = self.sql_text.get("1.0", "end-1c")
		action = self.act_var.get()
		if whereclause is not None and whereclause != '':
			sqlcmd = "SELECT treeviewid FROM mapdata WHERE %s" % whereclause
			success = self.parent.select_by_sqlcmd(sqlcmd, action, "Invalid data selection: %s" % whereclause)
			if success:
				self.sql = whereclause
				self.parent.whereclause = whereclause
				self.applied = True
				self.ever_applied = True
		self.sql_text.focus()
		return success
	def do_select(self, *args):
		if not self.applied:
			self.do_apply(None)
		self.rv = self.ever_applied
		self.dlg.destroy()



class RecodeDialog(Dialog):
	def __init__(self, parent, column_specs, prohibited_columns, init_sql=None):
		# parent is the class object, not widget
		self.parent = parent
		super().__init__("Recode Data",
				message="Select or enter the type and name of a column to alter, then enter an expression for the new or replacement values.  The expression should be valid SQL.",
				msgwraplength=100, help_url="https://mapdata.readthedocs.io/en/latest/recode.html")
		self.rv = (None, None, None, None, None)
		self.scriptfilepath = None
		self.dlg.columnconfigure(0, weight=1, minsize=620)
		#self.dlg.rowconfigure(1, weight=1, minsize=180)
		# Valid column selections, by type
		self.init_sql = init_sql
		self.sql = init_sql
		self.column_specs = column_specs
		self.prohibited_columns = prohibited_columns
		self.int_columns = [c[0] for c in self.column_specs if c[1] == "int" and not (c[1] in prohibited_columns)]
		self.int_columns.sort()
		self.float_columns = [c[0] for c in self.column_specs if c[1] == "float" and not (c[1] in prohibited_columns)]
		self.float_columns.sort()
		self.string_columns = [c[0] for c in self.column_specs if c[1] == "string" and not (c[1] in prohibited_columns)]
		self.string_columns.sort()
		# Frames
		column_frame = tk.Frame(self.content_frame)
		column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		column_frame.columnconfigure(3, weight=1)
		act_frame = tk.Frame(self.content_frame)
		act_frame.grid(row=1, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		query_frame = tk.Frame(self.content_frame)
		query_frame.grid(row=2, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		query_frame.rowconfigure(1, weight=1)
		query_frame.columnconfigure(0, weight=0)
		query_frame.columnconfigure(1, weight=3)
		# Row 0 of the query_frame is used for a label
		edit_frame = tk.Frame(query_frame)
		edit_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		sql_frame = tk.Frame(query_frame)
		sql_frame.grid(row=1, column=1, sticky=tk.NSEW, padx=(3,3), pady=(3,3))
		sql_frame.rowconfigure(0, weight=1, minsize=75)
		sql_frame.columnconfigure(0, weight=1)
		# Column selection
		coltype_lbl = ttk.Label(column_frame, justify=tk.LEFT, text="Data type:")
		coltype_lbl.grid(row=0, column=0, sticky=tk.NW, padx=(6,3), pady=(3,3))
		self.coltype_var = tk.StringVar(column_frame, "text")
		self.coltype_var.trace_add("write", self.ck_coltype)
		# The button values match SQLite data types.
		coltype_btn1 = ttk.Radiobutton(column_frame, variable=self.coltype_var, text="Text", value="text")
		coltype_btn2 = ttk.Radiobutton(column_frame, variable=self.coltype_var, text="Real number", value="real")
		coltype_btn3 = ttk.Radiobutton(column_frame, variable=self.coltype_var, text="Integer", value="integer")
		coltype_btn1.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,1))
		coltype_btn2.grid(row=1, column=1, sticky=tk.W, padx=(3,3), pady=(1,1))
		coltype_btn3.grid(row=2, column=1, sticky=tk.W, padx=(3,3), pady=(1,3))
		self.col_var = tk.StringVar(column_frame, "")
		self.col_var.trace_add("write", self.ck_col_name)
		col_lbl = ttk.Label(column_frame, justify=tk.RIGHT, text="New or existing column name:")
		col_lbl.grid(row=0, column=2, sticky=tk.W, padx=(25,3), pady=(3,1))
		self.col_sel = ttk.Combobox(column_frame, state=tk.NORMAL, textvariable=self.col_var, values=self.string_columns, width=24)
		self.col_sel.grid(row=1, column=2, sticky=tk.E, padx=(50,3), pady=(1,3))

		# Actions -- which values to replace
		ttk.Label(act_frame, justify=tk.LEFT, text="Values to replace:").grid(row=0, column=0, columnspan=2, sticky=tk.EW, padx=(6,3), pady=(6,3))
		self.repl_sel_var = tk.StringVar(act_frame, "all")
		repl_sel_btn1 = ttk.Radiobutton(act_frame, variable=self.repl_sel_var, text="Both selected and un-selected", value="all")
		repl_sel_btn2 = ttk.Radiobutton(act_frame, variable=self.repl_sel_var, text="Selected only", value="selected")
		repl_sel_btn3 = ttk.Radiobutton(act_frame, variable=self.repl_sel_var, text="Un-selected only", value="unselected")
		repl_sel_btn1.grid(row=3, column=0, sticky=tk.W, padx=(30,6), pady=(3,3))
		repl_sel_btn2.grid(row=1, column=0, sticky=tk.W, padx=(30,6), pady=(1,1))
		repl_sel_btn3.grid(row=2, column=0, sticky=tk.W, padx=(30,6), pady=(3,3))
		self.repl_null_var = tk.StringVar(act_frame, "all")
		repl_null_btn1 = ttk.Radiobutton(act_frame, variable=self.repl_null_var, text="Both empty and non-empty", value="all")
		repl_null_btn2 = ttk.Radiobutton(act_frame, variable=self.repl_null_var, text="Empty only", value="empty")
		repl_null_btn3 = ttk.Radiobutton(act_frame, variable=self.repl_null_var, text="Non-empty only", value="nonempty")
		repl_null_btn1.grid(row=3, column=1, sticky=tk.W, padx=(20,3), pady=(3,3))
		repl_null_btn2.grid(row=1, column=1, sticky=tk.W, padx=(20,3), pady=(1,1))
		repl_null_btn3.grid(row=2, column=1, sticky=tk.W, padx=(20,3), pady=(3,3))

		ttk.Label(query_frame, justify=tk.LEFT, text="Expression:").grid(row=0, column=0, columnspan=2, sticky=tk.EW, padx=(6,3), pady=(6,3))

		# Clear, Load, Edit, and Save buttons
		clear_btn = new_button(edit_frame, "Clear", 1, 0, self.clear_sql, tk.E, (0,1), underline=1)
		load_btn = new_button(edit_frame, "Open", 2, 0, self.load_script, tk.E, (0,1), underline=1)
		save_btn = new_button(edit_frame, "Save", 3, 0, self.save_script, tk.E, (0,1), underline=0)
		edit_btn = new_button(edit_frame, "Edit", 4, 0, self.edit_sql, tk.E, (0,1), (0,2), underline=0)
		enable_if(edit_btn, editor is not None)

		# SQL expression text entry
		self.sql_text = tk.Text(sql_frame, width=60, height=6)
		if self.init_sql is not None and self.init_sql != "":
			self.sql_text.insert(tk.END, self.init_sql)
		self.sql_text.grid(row=0, column=0, sticky=tk.NSEW, padx=(3,0), pady=(3,3))
		self.sql_text.bind("<KeyRelease>", self.check_enable)
		sbar = tk.Scrollbar(sql_frame)
		sbar.grid(row=0, column=1, sticky=tk.NS, padx=(0,3), pady=(3,3))
		sbar.config(command=self.sql_text.yview)
		self.sql_text.config(yscrollcommand = sbar.set)

		# Buttons
		self.ok_btn = add_help_ok_cancel_buttons(self.dlg, self.btn_frame, self.do_help, self.do_select, self.do_cancel, not (init_sql is None or len(init_sql) == 0))
		self.dlg.unbind("<Return>")
		self.dlg.bind("<Control-Return>", self.do_select)
		self.sql_text.focus()

	def check_enable(self, args=None):
		enable_if(self.ok_btn, self.col_var.get() != '' and self.sql_text.get("1.0", "end-1c") != '')
	def ck_coltype(self, varname, ix, mode):
		ct = self.coltype_var.get()
		cv = self.col_var.get()
		if ct == "text":
			self.col_sel["values"] = self.string_columns
			if cv in self.float_columns or cv in self.int_columns:
				self.col_var.set('')
				self.ok_btn["state"] = tk.DISABLED
		elif ct == "real":
			self.col_sel["values"] = self.float_columns
			if cv in self.string_columns or cv in self.int_columns:
				self.col_var.set('')
				self.ok_btn["state"] = tk.DISABLED
		else:
			self.col_sel["values"] = self.int_columns
			if cv in self.string_columns or cv in self.float_columns:
				self.col_var.set('')
				self.ok_btn["state"] = tk.DISABLED
	def ck_col_name(self, varname, ix, mode):
		if self.col_var.get() in self.prohibited_columns:
			self.col_var.set('')
		self.check_enable()
	def clear_sql(self):
		self.sql_text.delete(1.0, tk.END)
		self.ok_btn["state"] = tk.DISABLED
		self.sql_text.focus()
	def load_script(self):
		fn = tkfiledialog.askopenfilename(parent=self.dlg, title="SQL script file to open", filetypes=([('SQL script files', '.sql')]))
		if not (fn is None or fn == '' or fn == ()):
			path, filename = os.path.split(os.path.abspath(fn))
			self.scriptfilepath = path
			with open(fn, "r") as f:
				sql = f.read()
			self.sql_text.insert("end", sql)
		self.check_enable()
		self.sql_text.focus()
	def save_script(self):
		outfile = tkfiledialog.asksaveasfilename(initialdir=self.scriptfilepath, parent=self.dlg, title="SQL script file to save", filetypes=[('SQL script files', '.sql')])
		if not (outfile is None or outfile == ''):
			sql = self.sql_text.get("1.0", "end")
			with open(outfile, "w") as f:
				f.write(sql)
		self.check_enable()
		self.sql_text.focus()
	def edit_sql(self):
		td = tempfile.TemporaryDirectory()
		edit_fn = os.path.join(td.name, "mapfile_temp.sql")
		with open(edit_fn, "w") as f:
			f.write(self.sql_text.get("1.0", "end"))
		returncode = subprocess.call([editor, edit_fn])
		if returncode == 0:
			with open(edit_fn, "r") as f:
				sql = f.read()
			self.sql_text.delete("1.0", "end")
			self.sql_text.insert("end", sql)
			self.check_enable()
		else:
			warning("Failure attempting to edit the SQL with %s" % editor, kwargs={'parent':self.dlg})
		self.check_enable()
		self.sql_text.focus()
	def do_select(self, *args):
		self.sql = self.sql_text.get("1.0", "end-1c")
		if len(self.sql) > 0:
			self.rv = (self.sql, self.col_var.get(), self.coltype_var.get(), self.repl_sel_var.get(), self.repl_null_var.get())
		self.dlg.destroy()


class Plot(object):
	def __init__(self, frame, height=2, width=2, layout='none'):
		frame.rowconfigure(0, weight=1)
		frame.columnconfigure(0, weight=1)
		self.figure = Figure(figsize=(width, height), dpi=100, layout=layout)
		self.canvas = FigureCanvasTkAgg(self.figure, frame)
		self.plot_nav = NavigationToolbar2Tk(self.canvas, frame)
		self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot_nav.update()
		self.clear()
	def set_axis_labels(self, xlabel=None, ylabel=None):
		if xlabel is not None:
			self.axes.set_xlabel(xlabel)
		if ylabel is not None:
			self.axes.set_ylabel(ylabel)
	def draw(self):
		self.canvas.draw()
		self.plot_nav.update()
	def redraw(self):
		self.draw()
	def clear(self):
		self.figure.clear()
		self.axes = self.figure.add_subplot(111)
		self.canvas.draw()


class Hoverer(object):
	def __init__(self, plot_obj, point_labels, plot_pathcoll=None):
		self.plot = plot_obj
		self.splot = plot_pathcoll
		self.labels = point_labels
		self.annot = self.plot.axes.annotate("", xy=(0,0), xytext=(8,12), textcoords="offset points", \
				bbox={"boxstyle": "round", "fc": "w"}, arrowprops={"arrowstyle": "->"})
		self.annot.set_visible(False)
	def set_plot_pathcoll(self, plot_pathcoll):
		self.splot = plot_pathcoll
	def hover(self, event):
		if len(self.labels) > 0 and self.splot is not None:
			vis = self.annot.get_visible()
			if event.inaxes == self.plot.axes:
				cont, ind = self.splot.contains(event)
				if cont:
					pos = self.splot.get_offsets()[ind["ind"][0]]
					self.annot.xy = pos
					text = ", ".join([str(self.labels[n]) for n in ind["ind"]])
					self.annot.set_text(text)
					self.annot.get_bbox_patch().set_facecolor("wheat")
					self.annot.get_bbox_patch().set_alpha(0.8)
					self.annot.set_visible(True)
				else:
					if vis:
						self.annot.set_visible(False)
				self.plot.canvas.draw_idle()


class FindCandKeysDialog(Dialog):
	def __init__(self, parent, table_rows, column_specs):
		# parent is the class object, not the widget
		self.parent = parent
		self.column_specs = column_specs
		super().__init__("Find Candidate Keys",
				"Select one or more variables from the list below to evaluate whether they form a candidate key.  Use Ctrl-click or Shift-click to select multiple rows.",
				msgwraplength=50, help_url="https://mapdata.readthedocs.io/en/latest/candkeys.html")
		# Data
		self.dataset = None
		self.data_labels = None
		self.duplicates = None
		self.select_dups_sql = None
		# Columns that are not themselves singly a candidate column.
		self.cand_cols = [c[0] for c in self.column_specs if not (c[2] == 0 and c[3] == table_rows)]
		self.cand_cols.sort()

		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_recalc)

		# Add multi-select list of variables
		self.column_frame, self.column_table = treeview_table(self.content_frame, rowset=[[v] for v in self.cand_cols], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.cand_cols)))
		self.column_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW, padx=(6,6), pady=(3,3))
		self.column_table.bind('<ButtonRelease-1>', self.q_recalc)

		# Textarea to report the results of the check
		ttk.Label(self.content_frame, text="Results of check:").grid(row=1, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		self.results_text = tk.Text(self.content_frame, width=50, height=5, state=tk.DISABLED, wrap=tk.WORD)
		self.results_text.bind("<1>", lambda event: self.results_text.focus_set())
		self.results_text.grid(row=2, column=0, sticky=tk.NSEW, padx=(6,6), pady=(3,3))
		sbar = tk.Scrollbar(self.content_frame)
		sbar.grid(row=2, column=1, sticky=tk.NS, padx=(0,3), pady=(3,3))
		sbar.config(command=self.results_text.yview)
		self.results_text.config(yscrollcommand = sbar.set)

		# Buttons
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(4, weight=1)
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.data_btn = new_data_button(self.dlg, self.btn_frame, 1, self.show_data)
		self.dups_btn = new_button(self.btn_frame, "Duplicates", 0, 2, self.show_dups, tk.W, (3,3), state=tk.DISABLED)
		self.seldups_btn = new_button(self.btn_frame, "Select", 0, 3, self.select_dups, tk.W, (3,3), state=tk.DISABLED)
		new_close_button(self.dlg, self.btn_frame, 4, self.do_close)

	def q_recalc(self, *args, get_data=True):
		if self.dataset is None or get_data:
			self.clear_output()
			self.get_data()
		self.dups_btn["state"] = tk.DISABLED
		self.seldups_btn["state"] = tk.DISABLED
		if self.dataset is not None and len(self.dataset[0]) > 1:
			self.data_btn["state"] = tk.NORMAL
			self.check_cols()
		else:
			self.data_btn["state"] = tk.DISABLED
			self.clear_output()

	def get_data(self):
		# Get the selected data into 'dataset'
		self.dataset = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		if len(column_list) > 0:
			# Get either only the selected data or all data.
			dataset, rowids = chosen_dataset_and_ids(self.parent, column_list, self.sel_only_var.get()=='1')
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
				self.rowids = None
			else:
				self.dataset = dataset
				self.data_labels = column_list
				self.rowids = rowids

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for candidate key check")

	def show_dups(self, *args):
		if self.duplicates is not None:
			show_table(self.dlg, "Duplicates", "Selected columns with more than one instance:", self.duplicates, self.data_labels[0:len(self.dataset)]+["Duplicates"], \
					"Duplicates")

	def select_dups(self, *args):
		if self.select_dups_sql is not None:
			self.parent.select_by_sqlcmd(self.select_dups_sql, "Replace", "Error when selecting duplicates" )

	def check_cols(self):
		# Check for nulls in any column
		has_nulls = []
		for i, c in enumerate(self.dataset):
			if any([True for d in c if d is None or (type(d) is str and d.strip() == '')]):
				has_nulls.append(i)
		# SQL to find duplicates
		colstr = db_colnamestr(self.data_labels)
		sqlcmd = f"select {colstr}, count(*) as duplicates from mapdata where treeviewid in ({','.join(self.rowids)}) group by {colstr} having count(*) > 1;"
		cur = data_db.cursor()
		duprows = cur.execute(sqlcmd).fetchall()
		n_dups = len(duprows)
		if n_dups > 0:
			self.duplicates = duprows
			self.dups_btn["state"] = tk.NORMAL
			self.seldups_btn["state"] = tk.NORMAL
			# Get TreeView row IDs for all duplicates
			dbcols = db_colnames(self.data_labels)
			dupsel = []
			for duprow in self.duplicates:
				rowexprs = []
				for colno in range(len(self.data_labels)):
					if duprow[colno] is None:
						rowexprs.append(f"{dbcols[colno]} is NULL")
					else:
						v = duprow[colno]
						colexpr = dbcols[colno]
						if type(v) == str:
							v = "'"+v+"'"
							colexpr = f"cast({colexpr} as text)"
						rowexprs.append(f"{colexpr} = {v}")
				where_row = "(" + " and ".join(rowexprs) + ")"
				dupsel.append(where_row)
			where_tbl = " or ".join(dupsel)
			sqlcmd = f"select treeviewid from mapdata where {where_tbl};"
			self.select_dups_sql = sqlcmd
		#
		self.results_text["state"] = tk.NORMAL
		if len(has_nulls) > 0 or n_dups > 0:
			self.results_text.insert(tk.END, "NO -- not a candidate key.\n")
			if len(has_nulls) > 0:
				badcols = ", ".join([self.data_labels[i] for i in has_nulls])
				self.results_text.insert(tk.END, f"There are nulls in column{'s' if len(has_nulls) > 1 else ''}: {badcols}.\n")
			if n_dups > 0:
				if n_dups == 1:
					self.results_text.insert(tk.END, "There is one case with more than one row\nfor the tested key.\n")
				else:
					self.results_text.insert(tk.END, f"There are {n_dups} cases with more than one row\nfor the tested key.\n")
			self.results_text["state"] = tk.DISABLED
		else:
			self.results_text.insert(tk.END, f"YES -- a candidate key.")
		self.results_text["state"] = tk.DISABLED

	def clear_output(self):
		self.results_text["state"] = tk.NORMAL
		self.results_text.delete('1.0', tk.END)
		self.results_text["state"] = tk.DISABLED
		self.duplicates = None
		self.select_dups_sql = None
	def do_close(self, *args):
		self.parent.remove_candcol(self)
		super().do_cancel(args)



class FindDupRowsDialog(Dialog):
	def __init__(self, parent, table_rows, column_specs, prohibited_columns):
		self.parent = parent
		self.column_specs = column_specs
		self.prohibited_columns = prohibited_columns
		super().__init__("Find Duplicate Rows", "The table below lists all duplicated rows, for all columns that are not candidate keys.",
				help_url= "https://mapdata.readthedocs.io/en/latest/findduplicates.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.duplicates = None
		self.duplicates_labels = None
		# Columns that are not themselves singly a candidate column.
		self.noncand_cols = [c[0] for c in self.column_specs if not (c[1] == "string" and c[2] == 0 and c[3] == table_rows)]
		self.string_columns = sorted([c[0] for c in self.column_specs if c[1] == "string" and not c[1] in prohibited_columns])

		# Top controls
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.recalc)

		# The content_frame encompasses the two panes of the variable frame and the content frame
		frame_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)

		# Variable frame for list of variables
		var_frame = tk.Frame(frame_panes, width=250, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		varvals = [[v] for v in self.noncand_cols]
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=varvals, \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.noncand_cols)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.recalc)
		# Select all
		self.column_table.selection_set(self.column_table.get_children())

		# Dup frame for table of duplicates.
		self.dup_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		self.dup_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.dup_frame.rowconfigure(0, weight=1)
		self.dup_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.dup_frame, weight=12)
		# Initially add output table with default message
		self.out_tbl_frame, self.output_table = treeview_table(self.dup_frame, [["No duplicates found"]], column_headers=["Result"], select_mode="none", nrows=10)
		self.out_tbl_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW, padx=(6,6), pady=(3,3))

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.add_columns_btn = new_addcol_button(self.dlg, self.btn_frame, 1, self.add_column)
		new_close_button(self.dlg, self.btn_frame, 2, self.do_close)
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(2, weight=1)

		self.dlg.bind("<Control-s>")

	def recalc(self, *args):
		self.add_columns_btn["state"] = tk.DISABLED
		self.dlg.bind("<Control-s>")
		sel_columns = [str(self.column_table.item(sel_row)["values"][0]) for sel_row in self.column_table.selection()]
		if len(sel_columns) > 0:
			self.db_cols = db_colnames(sel_columns)
			col_expr = ",".join(self.db_cols)
			if self.sel_only_var.get() == '1':
				sel_tvids = self.parent.get_sel_rowids()
				whereclause = f"where treeviewid in ({','.join(sel_tvids)})"
			else:
				whereclause = ''
			cur = data_db.cursor()
			cur.execute("drop view if exists dups;")
			fd_sql = f"""create temporary view dups as
				SELECT
					'Duplicate set ' || row_number() over(order by {col_expr}) as duplicate_set,
					dd.*
				from (
					select 
						count(*) as duplicate_count, {col_expr}
					from mapdata
					{whereclause}
					group by {col_expr}
					having count(*) > 1
					) as dd
					;"""
			cur.execute(fd_sql)
			dupdata = cur.execute("select * from dups;").fetchall()
			cur.close()
			if len(dupdata) > 0:
				hdrs = ["Duplicate set", "Rows"]+sel_columns
				self.out_tbl_frame, self.output_table = treeview_table(self.dup_frame, dupdata, column_headers=hdrs, select_mode="none", nrows=10)
				self.out_tbl_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW, padx=(6,6), pady=(3,3))
				self.add_columns_btn["state"] = tk.NORMAL
				self.sel_only_ck["state"] = tk.NORMAL
				self.duplicates = dupdata
				self.duplicates_labels = hdrs
				self.dlg.bind("<Control-s>", self.save_data)
			else:
				self.out_tbl_frame, self.output_table = treeview_table(self.dup_frame, [["No duplicates found"]], column_headers=["Result"], select_mode="none", nrows=10)
				self.out_tbl_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW, padx=(6,6), pady=(3,3))

	def add_column(self, *args):
		dlg = CustomContentDialog(parent=self.dlg, title="Save Duplicate Labels", prompt="Save labels for duplicate groups in a table column.")
		def ck_col_name(varname, ix, mode):
			if self.col_var.get() in self.prohibited_columns:
				self.col_var.set('')
				dlg.ok_btn["state"] = tk.DISABLED
			else:
				enable_if(dlg.ok_btn, self.col_var.get() != '')
		self.col_var = tk.StringVar(self.dup_frame, "")
		self.col_var.trace_add("write", ck_col_name)
		col_lbl = ttk.Label(dlg.content_frame, justify=tk.RIGHT, text="New or existing column name:")
		col_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,6))
		self.col_sel = ttk.Combobox(dlg.content_frame, state=tk.NORMAL, textvariable=self.col_var, values=self.string_columns, width=24)
		self.col_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,6))
		self.col_sel.focus()
		ok = dlg.show()
		if ok:
			self.loading_dlg.display("Saving labels\nfor duplicates")
			# Get the treeview IDs for the duplicated rows.
			cur = data_db.cursor()
			joinexpr = " and ".join([f"(dups.{c}=md.{c} or coalesce(dups.{c}, md.{c}) is null)" for c in self.db_cols])
			sql = f"select dups.duplicate_set, md.treeviewid from dups inner join mapdata as md on {joinexpr};"
			dupkeylabels = cur.execute(sql).fetchall()
			# Update the database and data table
			target_col = self.col_var.get()
			db_col_name = db_colnames([target_col])[0]
			# Update the database
			new_column = target_col not in self.parent.headers
			if new_column:
				cur.execute(f"alter table mapdata add column {db_col_name} TEXT;")
			for i in range(len(dupkeylabels)):
				cur.execute(f"update mapdata set {db_col_name} = '{dupkeylabels[i][0]}' where treeviewid = '{dupkeylabels[i][1]}';")
			# Update the Treeview
			if new_column:
				dups_by_col = rows_to_columns(dupkeylabels)
				add_tv_column(self.parent.tbl, target_col, dups_by_col[1], dups_by_col[0])
				self.parent.headers = self.parent.headers + [target_col]
				viscols = list(self.parent.tbl["displaycolumns"])
				if viscols[0] != '#all':
					viscols.append(target_col)
					self.parent.tbl["displaycolumns"] = viscols
			else:
				for rowid in self.parent.tbl.get_children():
					for dup in dupkeylabels:
						if dup[1] == rowid:
							self.parent.tbl.set(rowid, column=target_col, value=dup[0])
					else:
						self.parent.tbl.set(rowid, column=target_col, value='')
			# Update the table specs
			dt = "string"
			coldata = self.parent.get_all_data([target_col])[0]
			missing = len([v for v in coldata if v is None or v == ''])
			unique = len(set([v for v in coldata if v is not None and v != '']))
			if new_column:
				self.parent.data_types.append([target_col, dt, missing, unique])
				self.string_columns.append(target_col)
			else:
				col_ix = self.parent.headers.index(target_col)
				self.parent.data_types[col_ix] = [target_col, dt, missing, unique]
			self.loading_dlg.hide()

	def save_data(self, *args):
		export_data_table(self.duplicates_labels, self.duplicates, sheetname="Duplicate row sets")
		self.dlg.lift()

	def show(self):
		if len(self.noncand_cols) == 0:
			warning("There are no columns that are not candidate keys.")
			parent.remove_duprows(self)
			self.dlg.destroy()
		else:
			center_window(self.dlg)
			raise_window(self.dlg)
			self.recalc(None)
			self.dlg.wait_window(self.dlg)
	def do_close(self, *args):
		self.parent.remove_duprows(self)
		super().do_cancel(args)



class CardinalityTestDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		super().__init__("Test Cardinality", "Choose one or more candidate key columns and one or more attribute columns to test the cardinality of the relationship between key and attribute.",
				help_url= "https://mapdata.readthedocs.io/en/latest/testcardinality.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.results = []
		self.results_labels = ["Selection", "Key column(s)", "Attribute column(s)", "Cardinality"]
		self.colnames = [c[0] for c in self.column_specs]
		self.dbcols = db_colnames(self.colnames)

		# Top controls
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_recalc)

		# The content_frame is split into upper and lower input and output frames
		frame_panes = ttk.PanedWindow(self.content_frame, orient=tk.VERTICAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)
		inp_frame = tk.Frame(frame_panes)
		inp_frame.grid(row=0, column=0, sticky=tk.NSEW)
		frame_panes.add(inp_frame, weight=1)
		self.outp_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		self.outp_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(6,6), pady=(3,6))
		frame_panes.add(self.outp_frame, weight=1)
		self.content_frame.rowconfigure(1, weight=1)
		# The input frame is split left and right into key and attribute column lists
		key_frame = tk.Frame(inp_frame, borderwidth=2, relief=tk.RIDGE)
		key_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(6,3), pady=(3,3))
		inp_frame.rowconfigure(0, weight=1)
		inp_frame.columnconfigure(0, weight=1)
		key_frame.columnconfigure(0, weight=1)
		attr_frame = tk.Frame(inp_frame, borderwidth=2, relief=tk.RIDGE)
		attr_frame.grid(row=0, column=1, sticky=tk.NSEW, padx=(3,6), pady=(3,3))
		inp_frame.columnconfigure(1, weight=1)
		attr_frame.columnconfigure(0, weight=1)
		# Input - Key columns
		ttk.Label(key_frame, text="Key column(s):").grid(row=0, column=0, padx=(6,6), pady=(3,3), sticky=tk.W)
		self.keylist_frame, self.keylist_table = treeview_table(key_frame, rowset=[[c] for c in self.colnames], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=10)
		self.keylist_table.bind('<ButtonRelease-1>', self.q_recalc)
		self.keylist_frame.grid(row=1, column=0, sticky=tk.NSEW)
		key_frame.rowconfigure(1, weight=1)
		# Input - Attribute column
		ttk.Label(attr_frame, text="Attribute column(s):").grid(row=0, column=0, padx=(6,6), pady=(3,3), sticky=tk.W)
		self.attrlist_frame, self.attrlist_table = treeview_table(attr_frame, rowset=[[c] for c in self.colnames], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=8)
		self.attrlist_table.bind('<ButtonRelease-1>', self.q_recalc)
		self.attrlist_frame.grid(row=1, column=0, sticky=tk.NSEW)
		attr_frame.rowconfigure(1, weight=1)
		# Output - Test results
		testbtn_frame = tk.Frame(self.outp_frame)
		testbtn_frame.grid(row=0, column=0, sticky=tk.EW)
		self.test_btn = new_button(testbtn_frame, "Test", 0, 0, self.recalc, sticky=tk.W, padx=(6,3), pady=(3,3), underline=None, state=tk.DISABLED)
		self.clear_btn = new_button(testbtn_frame, "Clear", 0, 1, self.do_clear_output, sticky=tk.W, padx=(3,6), pady=(3,3), underline=None, state=tk.DISABLED)
		testbtn_frame.columnconfigure(1, weight=1)
		#ttk.Label(self.outp_frame, text="Test results:").grid(row=0, column=0, padx=(6,6), pady=(3,3), sticky=tk.W)
		self.outp_frame.rowconfigure(1, weight=1)
		self.outp_frame.columnconfigure(0, weight=1)
		self.do_clear_output([])
		
		# Buttons
		add_help_close_buttons(self.dlg, self.btn_frame, self.do_help, self.do_close)

		self.dlg.bind("<Control-s>", self.save_data)

	def q_recalc(self, *args):
		input_ok = False
		key_columns = [str(self.keylist_table.item(sel_row)["values"][0]) for sel_row in self.keylist_table.selection()]
		if len(key_columns) > 0:
			attr_columns = [str(self.attrlist_table.item(sel_row)["values"][0]) for sel_row in self.attrlist_table.selection()]
			attr_columns = [c for c in attr_columns if not (c in key_columns)]
			input_ok = len(attr_columns) > 0
		self.test_btn["state"] = tk.NORMAL if input_ok else tk.DISABLED

	def recalc(self, *args):
		selected = "All"
		whereclause = ""
		if self.sel_only_var.get() == "1":
			sel_tvids = [f"'{id}'" for id in self.parent.get_sel_rowids()]
			selected = f"{len(sel_tvids)} rows"
			whereclause = f" where treeviewid in ({','.join(sel_tvids)})"
		if selected[0] == "0":
			self.test_btn["state"] = tk.DISABLED
			warning("No rows selected.", {"parent": self.dlg})
		else:
			key_columns = [str(self.keylist_table.item(sel_row)["values"][0]) for sel_row in self.keylist_table.selection()]
			keycolstr = ", ".join(key_columns)
			keydbcols = db_colnames(key_columns)
			keydbcolstr = ",".join(keydbcols)
			attr_columns = [str(self.attrlist_table.item(sel_row)["values"][0]) for sel_row in self.attrlist_table.selection()]
			attr_columns = [c for c in attr_columns if not (c in key_columns)]
			attrcolstr = ",".join(attr_columns)
			attrdbcolstr = ",".join(db_colnames(attr_columns))
			cur = data_db.cursor()
			sqlcmd = f"select {keydbcolstr} from (select distinct {keydbcolstr},{attrdbcolstr} from mapdata"
			sqlcmd = sqlcmd + whereclause + f") as d group by {keydbcolstr} having count(*) > 1;"
			resultlist = cur.execute(sqlcmd).fetchall()
			resultdesc = "One to many" if len(resultlist) > 0 else "One to one"
			self.results.append([selected, keycolstr, attrcolstr, resultdesc])
			self.update_output()
			cur.close()
			self.clear_btn["state"] = tk.NORMAL
			self.test_btn["state"] = tk.DISABLED

	def update_output(self):
		self.results_frame, self.results_table = treeview_table(self.outp_frame, rowset=self.results, \
				column_headers=self.results_labels, select_mode="none", nrows=5)
		self.results_frame.grid(row=1, column=0, sticky=tk.NSEW)

	def do_clear_output(self, *args):
		self.results = []
		self.update_output()
		self.clear_btn["state"] = tk.DISABLED

	def save_data(self, *args):
		if len(self.results) > 0:
			export_data_table(self.results_labels, self.results, sheetname="Cardinality Test Results")
			self.dlg.lift()

	def do_close(self, *args):
		self.parent.remove_cardinality(self)
		super().do_cancel(args)



class AddRowIDDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.curr_colnames = [c[0].lower() for c in column_specs]
		self.db_names = db_colnames(self.curr_colnames)
		super().__init__("Add Unique Row ID",
				"This will create a new column with a unique identifier.  Enter a name, and optionally a prefix, for the new column.  The new column must have a name different from any existing column.")
		self.rv = (None, None)
		self.new_col_var = tk.StringVar(self.content_frame, '')
		self.new_col_var.trace_add("write", self.ck_colname)
		ttk.Label(self.content_frame, text="New column name:").grid(row=0, column=0, sticky=tk.E, padx=(12,3), pady=(3,3))
		self.new_col_entry = ttk.Entry(self.content_frame, width=24, textvariable=self.new_col_var)
		self.new_col_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,12), pady=(3,3))
		self.prefix_var = tk.StringVar(self.content_frame, '')
		ttk.Label(self.content_frame, text="Optional prefix:").grid(row=1, column=0, sticky=tk.E, padx=(12,3), pady=(3,6))
		self.prefix_entry = ttk.Entry(self.content_frame, width=24, textvariable=self.prefix_var)
		self.prefix_entry.grid(row=1, column=1, sticky=tk.W, padx=(3,12), pady=(3,6))
		# Buttons
		self.ok_btn = new_ok_button(self.dlg, self.btn_frame, 0, self.do_select)
		self.cancel_btn = new_cancel_button(self.dlg, self.btn_frame, 1, self.do_cancel)
		self.btn_frame.columnconfigure(0, weight=1)
		self.new_col_entry.focus()
	def ck_colname(self, varname, ix, mode):
		cname = self.new_col_var.get().lower()
		enable_if(self.ok_btn, not (cname.strip() == '' or cname in self.curr_colnames or db_colnames([cname])[0] in self.db_names))
	def do_select(self, *args):
		if str(self.ok_btn["state"]) == tk.NORMAL:
			self.canceled = False
			self.rv = self.new_col_var.get(), self.prefix_var.get()
			self.dlg.destroy()


class AddCoordKeyDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.curr_colnames = [c[0].lower() for c in column_specs]
		self.db_names = db_colnames(self.curr_colnames)
		super().__init__("Add Candidate Key for Coordinates",
				"This will create a new column with a unique identifier for every unique combination of coordinates.  Enter a name, and optionally a prefix, for the new column.  The new column must have a name different from any existing column.")
		self.rv = (None, None)
		self.new_col_var = tk.StringVar(self.content_frame, '')
		self.new_col_var.trace_add("write", self.ck_colname)
		ttk.Label(self.content_frame, text="New column name:").grid(row=0, column=0, sticky=tk.E, padx=(12,3), pady=(3,3))
		self.new_col_entry = ttk.Entry(self.content_frame, width=24, textvariable=self.new_col_var)
		self.new_col_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,12), pady=(3,3))
		self.prefix_var = tk.StringVar(self.content_frame, '')
		ttk.Label(self.content_frame, text="Optional prefix:").grid(row=1, column=0, sticky=tk.E, padx=(12,3), pady=(3,6))
		self.prefix_entry = ttk.Entry(self.content_frame, width=24, textvariable=self.prefix_var)
		self.prefix_entry.grid(row=1, column=1, sticky=tk.W, padx=(3,12), pady=(3,6))
		# Buttons
		self.ok_btn = new_ok_button(self.dlg, self.btn_frame, 0, self.do_select)
		self.cancel_btn = new_cancel_button(self.dlg, self.btn_frame, 1, self.do_cancel)
	def ck_colname(self, varname, ix, mode):
		cname = self.new_col_var.get().lower()
		enable_if(self.ok_btn, not (cname == '' or cname in self.curr_colnames or db_colnames([cname])[0] in self.db_names))
	def do_select(self, *args):
		if str(self.ok_btn["state"]) == tk.NORMAL:
			self.canceled = False
			self.rv = self.new_col_var.get(), self.prefix_var.get()
			self.dlg.destroy()


class DistanceDialog(Dialog):
	def __init__(self, parent, coord_keys):
		# 'coord_keys' is list of column names that are candidate keys for coordinate pairs.
		self.parent = parent
		self.coord_keys = coord_keys
		msg = "Select two or more location IDs from the list at the left to see all pairwise distances."
		if len(coord_keys) > 2:
			msg += "  Alternate sets of location IDs can be selected from the dropdown above the list."
		super().__init__("Distances Between Locations", msg, help_url="https://mapdata.readthedocs.io/en/latest/distances.html")
		self.column_frame = None
		self.column_table = None
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.dataset = None
		self.data_labels = None
		self.calc_data = None	# A subset of self.dataset used for the calculation
		self.dlg.bind("<Control-s>")

		# Controls
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(2, weight=1)
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.reset_locid_list)

		# Distance units
		self.dist_var = tk.StringVar(self.ctrl_frame, "meters")
		ttk.Label(self.ctrl_frame, text="Units:").grid(row=0, column=1, sticky=tk.W, padx=(12,3), pady=(3,3))
		dist_sel = ttk.Combobox(self.ctrl_frame, state="readonly", textvariable=self.dist_var, width=11)
		dist_sel["values"] = ["meters", "kilometers", "feet", "miles"]
		dist_sel.grid(row=0, column=2, sticky=tk.W, padx=(3,15), pady=(3,3))
		dist_sel.bind('<<ComboboxSelected>>', self.q_recalc)

		# The content_frame encompasses the two panes of the variable frame and the content frame
		frame_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)

		# Variable frame for list of quantitative columns/variables
		self.var_frame = tk.Frame(frame_panes, width=250, borderwidth=2, relief=tk.RIDGE)
		self.var_frame.grid(row=2, column=0, sticky=tk.NSEW)
		self.var_frame.rowconfigure(2, weight=1)
		self.var_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.var_frame, weight=1)

		# Frame for table of distances
		self.dist_frame = tk.Frame(frame_panes, width=450, borderwidth=2, relief=tk.RIDGE)
		self.dist_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.dist_frame.rowconfigure(0, weight=1)
		self.dist_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.dist_frame, weight=12)

		# Coordinate candidate key selection - only if there are multiple
		self.coordkey_var = tk.StringVar(self.var_frame, self.coord_keys[0])
		if len(self.coord_keys) > 1:
			ttk.Label(self.var_frame, text="Identifiers:").grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
			coordkey_sel = ttk.Combobox(self.var_frame, state="readonly", textvariable=self.coordkey_var, width=24, values=self.coord_keys)
			coordkey_sel.grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
			coordkey_sel.bind('<<ComboboxSelected>>', self.reset_locid_list)

		# Buttons
		self.data_btn = add_help_src_close_btns(self.dlg, self.btn_frame, self.do_help, self.show_data, self.do_close)

		# Get the dataset of selected location identifiers and coordinates and set the selection list.
		# Also clears the output frame.
		self.reset_locid_list()

	def reset_locid_list(self, *args):
		self.clear_output()
		self.get_data()
		if self.dataset is not None:
			locids = columns_to_rows([sorted(self.dataset[0])])
			self.column_frame, self.column_table = treeview_table(self.var_frame, rowset=locids, \
					column_headers=[self.coordkey_var.get()], select_mode=tk.EXTENDED, nrows=min(20, len(locids)))
			self.column_frame.grid(row=2, column=0, sticky=tk.NSEW)
			self.column_table.bind('<ButtonRelease-1>', self.q_recalc)
		else:
			self.column_frame, self.column_table = treeview_table(self.var_frame, rowset=[], \
					column_headers=[self.coordkey_var.get()], select_mode=tk.EXTENDED, nrows=15)
			self.column_frame.grid(row=2, column=0, sticky=tk.NSEW)

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.calc_data, self.data_labels, \
					"Data for distance calculations")

	def clear_output(self):
		for ctl in self.dist_frame.winfo_children():
			ctl.destroy()
		tframe, tdata = treeview_table(self.dist_frame, [], ["Location 1", "Location 2", "Distance"])
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.calc_data = None
		self.data_btn["state"] = tk.DISABLED
		self.dlg.bind("<Control-s>")

	def q_recalc(self, *event):
		if len(self.column_table.selection()) > 1:
			self.recalc()
		else:
			self.clear_output()

	def get_data(self):
		# Get the selected data into 'dataset'.  This is used to set up the input selections as
		# well as to get data used to calcualte the output
		column_list = [self.coordkey_var.get(), self.parent.lat_col, self.parent.lon_col]
		self.n_dataset_columns = len(column_list)
		# Get either only the selected data or all data.
		dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is None or len(dataset[0]) == 0:
			self.dataset = None
			self.data_labels = None
		else:
			dataset = clean_missing(dataset, [0,1,2])
			if len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				# Convert all location ids to strings
				dataset[0] = [str(v) for v in dataset[0]]
				# Convert all lat/lon to floats
				for i in [1,2]:
					dataset[i] = [conv_float(v) for v in dataset[i]]
				#
				# Reduce to only unique rows
				dataset = rows_to_columns([list(d) for d in list(set(map(tuple, columns_to_rows(dataset))))])
				self.dataset = dataset
				self.data_labels = column_list

	def recalc(self):
		# Calculate and display the distances
		self.loading_dlg.display("Calculating distances")
		# Get the rows of self.dataset to be used for the calculation into self.calc_data
		id_list = [str(self.column_table.item(sel_row)["values"][0]) for sel_row in self.column_table.selection()]
		ds = columns_to_rows(self.dataset)
		sel_ds = [r for r in ds if r[0] in id_list]
		self.calc_data = rows_to_columns(sel_ds)
		from pyproj import Proj, transform, Geod
		geod = Geod(ellps='WGS84')
		self.calc_result = []
		for i in range(len(sel_ds)-1):
			for j in range(i+1, len(sel_ds)):
				a1, a2, dist = geod.inv(sel_ds[i][2], sel_ds[i][1], sel_ds[j][2], sel_ds[j][1])
				if self.dist_var.get() == 'kilometers':
					dist = dist/1000
				elif self.dist_var.get() == 'feet':
					dist = dist * 3.28084
				elif self.dist_var.get() ==  'miles':
					dist = dist / 1609.34708789
				self.calc_result.append([sel_ds[i][0], sel_ds[j][0], dist])
		self.calc_result_labels = [self.coordkey_var.get()+" 1", self.coordkey_var.get()+" 2", self.dist_var.get()]
		tframe, tdata = treeview_table(self.dist_frame, self.calc_result, self.calc_result_labels)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.data_btn["state"] = tk.NORMAL
		self.dlg.bind("<Control-s>", self.save_data)
		self.loading_dlg.hide()

	def save_data(self, *args):
		export_data_table(self.calc_result_labels, self.calc_result, sheetname="Distances between locations")
		self.dlg.lift()
	def do_close(self, *args):
		self.parent.remove_distdlg(self)
		super().do_cancel(args)


class AggregateDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		super().__init__("Numerical Aggregates",
				"Select one or more variables from the left, and a grouping variable, to see the selected aggregates.  Use Ctrl-click or Shift-click to select multiple rows.",
				help_url="https://mapdata.readthedocs.io/en/latest/aggregates.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.dataset = None
		self.data_labels = None
		self.agg_data = None
		self.agg_data_labels = None
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.categ_columns = sorted([c[0] for c in self.column_specs if c[1] in ("string", "boolean", "date", "timestamp")])
		self.show_labels = True
		self.dlg.bind("<Control-s>")

		# Controls
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(2, weight=1)
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_recalc)

		# Aggregation method for rows within a group (required)
		self.aggreg_var = tk.StringVar(self.ctrl_frame, "Arithmetic mean")
		ttk.Label(self.ctrl_frame, text="Aggregate rows by:").grid(row=0, column=1, sticky=tk.W, padx=(12,3), pady=(3,3))
		aggreg_sel = ttk.Combobox(self.ctrl_frame, state="readonly", textvariable=self.aggreg_var, width=24)
		aggreg_sel["values"] = ["Count", "Minimum", "Maximum", "Median", "Arithmetic mean", "Geometric mean", "Harmonic mean", "Sum"]
		aggreg_sel.grid(row=0, column=2, sticky=tk.W, padx=(3,12), pady=(3,3))
		aggreg_sel.bind('<<ComboboxSelected>>', self.q_recalc)

		# The content_frame encompasses the two panes of the variable frame and the content frame
		frame_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)

		# Variable frame for list of quantitative columns/variables
		var_frame = tk.Frame(frame_panes, width=250, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_recalc)

		# Grouping variable (required)
		self.groupby_var = tk.StringVar(var_frame, "")
		ttk.Label(var_frame, text="Group by:").grid(row=1, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		groupby_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.groupby_var, width=24, values=self.categ_columns)
		groupby_sel.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		groupby_sel.bind('<<ComboboxSelected>>', self.q_recalc)

		# agg_frame frame for table of aggregates
		self.agg_frame = tk.Frame(frame_panes, width=450, borderwidth=2, relief=tk.RIDGE)
		self.agg_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.agg_frame.rowconfigure(0, weight=1)
		self.agg_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.agg_frame, weight=12)

		# initialize content frame with an empty table
		self.clear_output()

		# Buttons
		self.data_btn = add_help_src_close_btns(self.dlg, self.btn_frame, self.do_help, self.show_data, self.do_close)

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for calculation of aggregates")
	def clear_output(self):
		for ctl in self.agg_frame.winfo_children():
			ctl.destroy()
		tframe, tdata = treeview_table(self.agg_frame, [], ["Group"])
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.dataset = None
		self.agg_data = None
		clear_dlg_hotkeys(self.dlg)

	def q_recalc(self, *event, get_data=True):
		self.clear_output()
		if (len(self.column_table.selection()) > 0 and self.groupby_var.get() != '') and get_data:
			self.get_data()
			if self.dataset is not None and len(self.dataset) > 0 and len(self.dataset[0]) > 0:
				self.recalc()
			else:
				self.data_btn["state"] = tk.DISABLED

	def get_data(self):
		# Get the selected data into 'dataset'
		self.clear_output()
		self.dataset = None
		self.agg_data = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		self.n_dataset_columns = len(column_list)
		column_list.append(self.groupby_var.get())
		if self.n_dataset_columns > 0:
			# Get either only the selected data or all data.
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				# Convert all group ids to strings
				dataset[self.n_dataset_columns] = [str(v) for v in dataset[self.n_dataset_columns]]
				groups = sorted_numstrs(list(set(dataset[self.n_dataset_columns])))
				# Convert all numerics to floats
				for i in range(self.n_dataset_columns):
					dataset[i] = [conv_float(v) for v in dataset[i]]
				#
				self.dataset = dataset
				self.data_labels = column_list
				self.data_btn["state"] = tk.NORMAL
				# Aggregate rows by group
				subsets = subset_by_groups(dataset[0:self.n_dataset_columns], dataset[self.n_dataset_columns])
				agg_error, agg_data = aggregate_groups(subsets, self.aggreg_var.get())
				if not agg_error:
					self.agg_data = sort_columns(agg_data)
					self.agg_data_labels = [self.data_labels[self.n_dataset_columns]] + self.data_labels[0:self.n_dataset_columns]

	def recalc(self):
		# Display the aggregates.  These were calculated in get_data()
		if self.agg_data is not None and len(self.agg_data) > 0 and len(self.agg_data[0]) > 0:
			self.dlg.bind("<Control-s>", self.save_aggregates)
			self.loading_dlg.display("Calculating aggregates")
			tframe, tdata = treeview_table(self.agg_frame, columns_to_rows(self.agg_data), self.agg_data_labels)
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			self.loading_dlg.hide()

	def save_aggregates(self, *args):
		if self.agg_data is not None:
			export_data_table(self.agg_data_labels, columns_to_rows(self.agg_data), sheetname=f"{self.aggreg_var.get()}")
			self.dlg.lift()
	def do_close(self, *args):
		self.parent.remove_aggdlg(self)
		super().do_cancel(args)



class UniqueValuesDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.variables = [c[0] for c in column_specs]
		super().__init__("Unique Values", "Select one or more variables from the left to see all unique values.",
				help_url="https://mapdata.readthedocs.io/en/latest/unique_values.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.dataset = None
		self.data_labels = None
		self.dlg.bind("<Control-s>")

		# The content_frame encompasses the two panes of the variable frame and the content frame
		frame_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL, width=600)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)

		# Variable frame for list of variables
		var_frame = tk.Frame(frame_panes, width=250, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		varvals = [[v] for v in self.variables]
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=varvals, \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.variables)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_recalc)

		# values frame for table of unique values
		self.values_frame = tk.Frame(frame_panes, width=400, borderwidth=3, relief=tk.RIDGE)
		self.values_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.values_frame.rowconfigure(0, weight=1)
		self.values_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.values_frame, weight=12)

		# initialize content frame with an empty table
		self.clear_output()

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		new_close_button(self.dlg, self.btn_frame, 2, self.do_cancel)

	def clear_output(self):
		for ctl in self.values_frame.winfo_children():
			ctl.destroy()
		tframe, tdata = treeview_table(self.values_frame, [], ["Variable", "Count"])
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.dataset = None
		self.agg_data = None
		clear_dlg_hotkeys(self.dlg)

	def q_recalc(self, *event, get_data=True):
		self.clear_output()
		if len(self.column_table.selection()) > 0 and get_data:
			self.recalc()

	def recalc(self):
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		dbcols = db_colnames(column_list)
		colstr = ",".join(dbcols)
		sqlcmd = f"select distinct {colstr}, count(*) as nrows from mapdata group by {colstr} order by {colstr};"
		cur = data_db.cursor()
		self.dataset = cur.execute(sqlcmd).fetchall()
		self.data_labels = column_list + ["Count"]
		tframe, tdata = treeview_table(self.values_frame, self.dataset, self.data_labels)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.dlg.bind("<Control-s>", self.save_data)

	def save_data(self, *args):
		export_data_table(self.data_labels, self.dataset, sheetname="Unique Values")
		self.dlg.wait_window(self.dlg)


class SelNonMissingDialog(Dialog):
	def __init__(self, column_specs):
		super().__init__("Select Non-Missing Values",
				"Select one or more variables from the list below to select data rows where they are non-missing.",
				help_url="https://mapdata.readthedocs.io/en/latest/select_nonmissing.html")
		self.column_list = [c[0] for c in column_specs]
		self.sel_columns = []

		self.ctrl_frame.destroy()

		# Add multi-select list of variables
		self.column_frame, self.column_table = treeview_table(self.content_frame, rowset=[[v] for v in self.column_list], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.column_list)))
		self.column_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW, padx=(6,6), pady=(3,3))
		self.column_table.bind('<ButtonRelease-1>', self.check_ok)

		# Action selection
		act_frame = tk.Frame(self.content_frame)
		act_frame.grid(row=1, column=0, sticky=tk.NSEW)
		self.act_var = tk.StringVar(act_frame, "Replace")
		act_lbl = ttk.Label(act_frame, text="Action:")
		act_lbl.grid(row=0, column=0, sticky=tk.E, padx=(12,3))
		act_sel = ttk.Combobox(act_frame, state="readonly", textvariable=self.act_var, values=["Replace", "Union", "Intersection", "Difference O-N", "Difference N-O"], width=15)
		act_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6))

		# Buttons
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(1, weight=1)
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.select_btn = new_button(self.btn_frame, "Select", 0, 1, self.do_select, tk.W, (3,3), state=tk.DISABLED)
		new_cancel_button(self.dlg, self.btn_frame, 2, self.do_close)
	
	def check_ok(self, *args):
		self.sel_columns = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			self.sel_columns.append(datarow[0])
		enable_if(self.select_btn, len(self.sel_columns) > 0)

	def do_select(self, *args):
		self.rv = (self.sel_columns, self.act_var.get())
		self.canceled = False
		self.dlg.destroy()
	
	def do_close(self, *args):
		self.rv = ([], self.act_var.get())
		self.canceled = True
		self.dlg.destroy()


class ColocatedDialog(Dialog):
	def __init__(self):
		super().__init__("Co-located Data", "The options below will select data based on the number of co-located data values.",
				help_url= "https://mapdata.readthedocs.io/en/latest/select_colocated.html")
		self.rv = (None, None, None)
		# Data selection controls
		self.match_var = tk.StringVar(self.ctrl_frame, "more than")
		match_lbl = ttk.Label(self.ctrl_frame, text="Select data at locations with")
		match_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3))
		self.match_sel = ttk.Combobox(self.ctrl_frame, state="readonly", textvariable=self.match_var, values=["more than", "less than", "exactly"], width=10)
		self.match_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3))
		self.num_var = tk.IntVar(self.ctrl_frame, 1)
		self.num_entry = ttk.Spinbox(self.ctrl_frame, textvariable=self.num_var, from_=1, to=250, width=4)
		self.num_entry.grid(row=0, column=2, sticky=tk.EW, padx=(6,6), pady=(3,3))
		ttk.Label(self.ctrl_frame, text="data values").grid(row=0, column=3, sticky=tk.W, padx=(3, 12), pady=(3,3))
		# Action selection
		act_frame = tk.Frame(self.ctrl_frame)
		act_frame.grid(row=1, column=0, sticky=tk.NSEW)
		self.act_var = tk.StringVar(act_frame, "Replace")
		act_lbl = ttk.Label(act_frame, text="Action:")
		act_lbl.grid(row=0, column=0, sticky=tk.E, padx=(12,3))
		act_sel = ttk.Combobox(act_frame, state="readonly", textvariable=self.act_var, values=["Replace", "Union", "Intersection", "Difference O-N", "Difference N-O"], width=15)
		act_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6))
		# Buttons
		self.ok_btn = add_help_ok_cancel_buttons(self.dlg, self.btn_frame, self.do_help, self.do_select, self.do_cancel, ok_enabled=True)
		self.dlg.resizable(False, False)
		self.match_sel.focus()
	def do_select(self, *args):
		if self.match_var.get() != '' and self.num_var.get() > 0:
			self.canceled = False
			self.rv = (self.match_var.get(), self.num_var.get(), self.act_var.get())
			self.dlg.destroy()


	
class PlotDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		# Numeric columns
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.pos_numeric_columns = None		# Set only if a bubble plot is selected
		# Category columns.  Does not include date columns for most uses, but may include dates for some.
		self.categ_columns = sorted([c[0] for c in self.column_specs if c[1] in ("string", "boolean")])
		self.categ_columns2 = sorted([c[0] for c in self.column_specs if c[1] in ("string", "boolean", "date")])
		# quant_columns includes date and timestamp columns
		self.quant_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float", "date", "timestamp", "timestamptz")])
		self.date_columns = sorted([c[0] for c in self.column_specs if c[1] in ("date", "timestamp", "timestamptz")])
		self.dataset = None
		self.data_labels = None
		self.n_dataset_columns = None
		self.plot_data = None
		self.plot_data_labels = None
		super().__init__("Plot", "Select the type of plot, data columns to use, and whether to show only selected data.",
				help_url="https://mapdata.readthedocs.io/en/latest/plot_general.html")
		self.auto_update = True
		self.plot_title = None
		self.xlabel = None
		self.ylabel = None
		# For transparency on some plots; also initial value set by plot type in 'set_xy()'
		self.alpha = 0.45
		# For rotation of some plot types
		self.rotated = False
		# For histogram; will be initially overridden based on data set
		self.bins = None
		# For display of groups at Jenks breaks on Q-Q plots
		self.qq_groups = False
		# For display of lines at X and Y Jenks breaks on scatter plots
		self.scatter_breaks = False
		self.scatter_x_breaks = None
		self.scatter_y_breaks = None
		self.lineplot_breaks = False
		self.lineplot_x_breaks = None
		# For flipped Y axis
		self.flip_y = False
		# For display of semi-fitted lines on scatter plots
		self.loess = False
		self.linreg = False
		self.theilsen = False
		# Hover annotation for scatter plots
		self.annot = None
		self.canvas_conn_id = None

		def set_autoupdate():
			if self.autoupdate_var.get() == "1":
				self.auto_update = True
				self.q_redraw()
			else:
				self.auto_update = False

		# Controls
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.type_var = tk.StringVar(self.ctrl_frame, "")
		type_lbl = ttk.Label(self.ctrl_frame, text="Plot type:")
		type_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		plot_types = ["Box plot", "Breaks groups", "Breaks optimum", "Bubble plot", "Count by category", "CV by category"]
		if len(self.date_columns) > 1:
			plot_types.append("Date range by category")
		plot_types.extend(["Empirical CDF", "Histogram", "Kernel density (KD) plot", "Line plot", "Mean by category",
					"Min-max by bin", "Min-max by category", "Normal Q-Q plot", "Pareto chart", "Scatter plot", "Stripchart",
					"Total by category", "Violin plot", "Y range plot"])
		self.type_sel = ttk.Combobox(self.ctrl_frame, state="readonly", textvariable=self.type_var, width=20, height=21, values=plot_types)
		self.type_sel.grid(row=0, column=1, columnspan=2, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.type_sel["state"] = tk.NORMAL
		self.type_sel.bind("<<ComboboxSelected>>", self.set_xy)

		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 1, 0, self.q_redraw, colspan=2)

		self.autoupdate_var = add_autoupdate(self.ctrl_frame, 1, 2, set_autoupdate)

		self.x_var = tk.StringVar(self.ctrl_frame, "")
		self.x_lbl = ttk.Label(self.ctrl_frame, text="X column:")
		self.x_lbl.grid(row=0, column=3, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.x_sel = ttk.Combobox(self.ctrl_frame, state=tk.DISABLED, textvariable=self.x_var, width=24)
		self.x_sel.grid(row=0, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.x_sel.bind("<<ComboboxSelected>>", self.x_changed)

		self.y_var = tk.StringVar(self.ctrl_frame, "")
		self.y_lbl = ttk.Label(self.ctrl_frame, text="Y column:")
		self.y_lbl.grid(row=1, column=3, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.y_sel = ttk.Combobox(self.ctrl_frame, state=tk.DISABLED, textvariable=self.y_var, width=24)
		self.y_sel.grid(row=1, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.y_sel.bind("<<ComboboxSelected>>", self.y_changed)

		self.xlog_var = tk.StringVar(self.ctrl_frame, "0")
		self.xlog_ck = ttk.Checkbutton(self.ctrl_frame, text="Log X", state=tk.DISABLED, command=self.q_redraw, variable=self.xlog_var,
				onvalue="1", offvalue="0")
		self.xlog_ck.grid(row=0, column=5, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.ylog_var = tk.StringVar(self.ctrl_frame, "0")
		self.ylog_ck = ttk.Checkbutton(self.ctrl_frame, text="Log Y", state=tk.DISABLED, command=self.q_redraw, variable=self.ylog_var,
				onvalue="1", offvalue="0")
		self.ylog_ck.grid(row=1, column=5, sticky=tk.W, padx=(6,6), pady=(3,3))

		# Frame for Z variable--not gridded except for a bubble plot.
		# When needed it should be gridded at row=2, column=3, columnspan=2.
		self.z_frame = tk.Frame(self.ctrl_frame)
		self.z_var = tk.StringVar(self.z_frame, "")
		z_lbl = ttk.Label(self.z_frame, text="Z column:")
		z_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.z_sel = ttk.Combobox(self.z_frame, state="readonly", textvariable=self.z_var, width=24)
		self.z_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.z_sel.bind("<<ComboboxSelected>>", self.z_changed)
		self.z_width_var = tk.IntVar(self.z_frame, 300)
		self.z_width_var.trace_add("write", self.z_width_changed)
		ttk.Label(self.z_frame, text="Max. size:").grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.z_width_sel = ttk.Spinbox(self.z_frame, textvariable=self.z_width_var, from_=10, to=5000, increment=10, width=5)
		self.z_width_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))

		self.groupby_var = tk.StringVar(self.ctrl_frame, "")
		self.groupby_lbl = ttk.Label(self.ctrl_frame, text="Group by:")
		self.groupby_lbl.grid(row=3, column=3, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.groupby_sel = ttk.Combobox(self.ctrl_frame, state=tk.DISABLED, textvariable=self.groupby_var, width=24)
		self.groupby_sel.grid(row=3, column=4, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.groupby_sel.bind("<<ComboboxSelected>>", self.groupby_changed)

		# Plot
		self.plot = Plot(self.content_frame, 5, 5)

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.data_btn = new_data_button(self.dlg, self.btn_frame, 1, self.show_data)
		self.plot_data_btn = new_plot_button(self.dlg, self.btn_frame, 2, self.show_plot_data)
		self.clone_btn = new_button(self.btn_frame, "Clone", 0, 3, self.clone_plot, tk.W, (3,6), (0,0), 3)
		self.dlg.bind("<Alt-n>", self.clone_plot)
		new_close_button(self.dlg, self.btn_frame, 4, self.do_close)
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(4, weight=1)

		self.dlg.bind("<Alt-t>", self.set_title)
		self.dlg.bind("<Alt-x>", self.set_xlabel)
		self.dlg.bind("<Alt-y>", self.set_ylabel)

		self.geom = '650x700'

	def show_data(self, *args):
		# Show data that have been collected for plotting, but not summarized as needed for a particular plot type.
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Original data:", self.dataset, \
					self.data_labels[0:len(self.dataset)], "Source data")
			self.dlg.lift()

	def show_plot_data(self, *args):
		# Show data as summarized for a particular plot type.
		if self.plot_data is not None:
			show_columnar_table(self.dlg, "Data for Plotting", "Data to be plotted:", self.plot_data, \
					self.plot_data_labels, "Plot data")
			self.dlg.lift()
	
	def clone_plot(self, *args):
		self.parent.clone_plot(self)

	def x_changed(self, *args):
		self.xlabel = None
		plot_type = self.type_var.get()
		if plot_type in ("Count by category", "Date range by category", "Pareto chart"):
			self.xlog_ck["state"] = tk.DISABLED
		else:
			self.xlog_ck["state"] = "readonly"
		xval = self.x_var.get()
		yval = self.y_var.get()
		if xval != '' and yval == xval:
			self.y_var.set('')
		self.bins = None
		self.q_redraw(args)

	def y_changed(self, *args):
		self.ylabel = None
		plot_type = self.type_var.get()
		if plot_type in ("Count by category", "Date range by category", "Histogram", "Empirical CDF", "Min-max by category"):
			self.ylog_ck["state"] = tk.DISABLED
		else:
			self.ylog_ck["state"] = "readonly"
		xval = self.x_var.get()
		yval = self.y_var.get()
		if yval != '' and yval == xval:
			self.x_var.set('')
		self.bins = None
		self.q_redraw(args)

	def z_changed(self, *args):
		self.q_redraw(args)

	def z_width_changed(self, *args):
		try:
			sz = self.z_width_var.get()
		except:
			pass
		else:
			self.q_redraw(args)

	def groupby_changed(self, *args):
		self.q_redraw(args)

	def set_xy(self, *args):
		# Enable X and Y value selection, and set Combobox values based on plot type and column types.
		# Also sets the groupby Combobox values if applicable.
		self.x_lbl["text"] = "X column:"
		self.y_lbl["text"] = "Y column:"
		self.plot.clear()
		self.plot_title = None
		self.xlabel = None
		self.ylabel = None
		self.dlg.bind("<Alt-a>")
		self.dlg.bind("<Alt-b>")
		self.dlg.bind("<Alt-f>")
		self.dlg.bind("<Alt-g>")
		self.dlg.bind("<Alt-l>")
		self.dlg.bind("<Alt-r>")
		self.dlg.bind("<Alt-s>")
		self.bins = None
		self.flip_y = False
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED
		self.x_var.set('')
		self.y_var.set('')
		self.xlog_var.set('0')
		self.ylog_var.set('0')
		self.groupby_var.set('')
		self.xlog_ck["state"] = tk.DISABLED
		self.ylog_ck["state"] = tk.DISABLED
		self.x_sel["state"] = tk.DISABLED
		self.y_sel["state"] = tk.DISABLED
		self.groupby_sel["state"] = tk.DISABLED
		plot_type = self.type_var.get()
		if plot_type == "Bubble plot":
			if self.pos_numeric_columns is None:
				self.pos_numeric_columns = []
				nd = self.parent.get_all_data(self.numeric_columns)
				for i, col in enumerate(nd):
					if min([float(c) for c in col if c != '']) >= 0:
						self.pos_numeric_columns.append(self.numeric_columns[i])
				del nd
			if self.pos_numeric_columns == []:
				warning("There are no strictly positive numeric columns to use for the bubble size.", {"parent": self.dlg})
				self.z_var.set('')
			else:
				self.z_frame.grid(row=2, column=3, columnspan=2)
				self.x_sel["values"] = self.numeric_columns
				self.y_sel["values"] = self.numeric_columns
				self.z_sel["values"] = self.pos_numeric_columns
				self.groupby_sel["state"] = "readonly"
				self.groupby_sel["values"] = [''] + self.categ_columns2
		else:
			self.z_frame.grid_forget()
			self.z_var.set('')
		if plot_type in ("Count by category", "Pareto chart"):
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = self.categ_columns2
		elif plot_type in ("Empirical CDF", "Normal Q-Q plot", "Breaks groups", "Breaks optimum"):
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = self.numeric_columns
			self.xlog_ck["state"] = tk.NORMAL
			if plot_type == "Normal Q-Q plot":
				self.dlg.bind("<Alt-g>", self.show_groups)
		elif plot_type == "Date range by category":
			self.x_lbl["text"] = "First date:"
			self.y_lbl["text"] = "Last date:"
			self.x_sel["state"] = "readonly"
			self.y_sel["state"] = "readonly"
			self.groupby_sel["state"] = "readonly"
			self.x_sel["values"] = self.date_columns
			self.y_sel["values"] = self.date_columns
			self.groupby_sel["values"] = self.categ_columns
		elif plot_type in ("Box plot", "Histogram", "Kernel density (KD) plot", "Mean by category", "CV by category", "Stripchart", "Total by category", "Violin plot"):
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = self.numeric_columns
			self.xlog_ck["state"] = tk.NORMAL
			self.groupby_sel["state"] = "readonly"
			self.groupby_sel["values"] = [''] + self.categ_columns2
			self.alpha = 0.45
			if plot_type == "Histogram":
				self.dlg.bind("<Alt-b>", self.set_bins)
		elif plot_type == "Min-max by bin":
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = self.numeric_columns
			self.xlog_ck["state"] = tk.NORMAL
			self.y_sel["state"] = "readonly"
			self.y_sel["values"] = self.numeric_columns
		elif plot_type == "Min-max by category":
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = self.quant_columns
			self.xlog_ck["state"] = tk.NORMAL
			self.y_sel["state"] = "readonly"
			self.y_sel["values"] = self.categ_columns2
		else:
			self.x_sel["state"] = "readonly"
			self.x_sel["values"] = self.quant_columns
			self.xlog_ck["state"] = tk.NORMAL
			self.y_sel["state"] = "readonly"
			self.y_sel["values"] = self.quant_columns
			self.ylog_ck["state"] = tk.NORMAL
			if plot_type in ("Scatter plot", "Line plot"):
				self.groupby_sel["state"] = "readonly"
				if plot_type == "Scatter plot":
					self.groupby_sel["values"] = [''] + self.categ_columns2 + ['* Breaks in X', '* Breaks in Y']
					self.alpha = 0.45
				else:
					self.groupby_sel["values"] = [''] + self.categ_columns2 + ['* Breaks in X']
					self.alpha = 1.0

	def q_redraw(self, get_data=True, *args):
		# Conditionally (re)draw the plot.
		plot_type = self.type_var.get()
		can_redraw = (plot_type in ("Count by category", "Empirical CDF", "Normal Q-Q plot", "Breaks groups", "Breaks optimum", \
				"Histogram", "Pareto chart") and self.x_var.get() != '') \
				or (plot_type in ("Scatter plot", "Line plot", "Min-max by bin", "Min-max by category", "Y range plot") and \
				self.x_var.get() != '' and self.y_var.get() != '') \
				or (plot_type in ("Box plot", "Kernel density (KD) plot", "Stripchart", "Violin plot") and self.x_var.get() != '') \
				or (plot_type in ("Mean by category", "CV by category", "Total by category") and self.x_var.get() != '' and self.groupby_var.get() != '') \
				or (plot_type == "Bubble plot" and self.x_var.get() != '' and self.y_var.get() != '' and self.z_var.get() != '') \
				or (plot_type == "Date range by category" and self.x_var.get() != '' and self.y_var.get() != '' and self.groupby_var.get() != '')
		if can_redraw:
			self.plot.clear()
			if get_data or self.dataset is None or self.plot_data is None:
				self.get_data()
			if self.dataset is not None:
				self.redraw()

	def get_data(self):
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED
		self.dataset = None
		plot_type = self.type_var.get()
		column_list = []
		if self.x_var.get() != '':
			column_list = [self.x_var.get()]
		if self.y_var.get() != '':
			column_list.append(self.y_var.get())
		if self.z_var.get() != '':
			column_list.append(self.z_var.get())
		grpbyvar = self.groupby_var.get()
		if grpbyvar != '' and grpbyvar not in ('* Breaks in X', '* Breaks in Y'):
			column_list.append(self.groupby_var.get())
			grpby_col_index = len(column_list) - 1
		else:
			grpby_col_index = None
		# Record the number of data columns without labels.
		self.n_dataset_columns = len(column_list)
		if map_settings.label_col is not None:
			column_list.append(map_settings.label_col)
		# Get either only the selected data or all data.
		dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is None or len(dataset[0]) == 0:
			self.dataset = None
			self.data_labels = None
			self.plot_data = None
			self.plot_data_labels = None
			self.data_btn["state"] = tk.DISABLED
			self.plot_data_btn["state"] = tk.DISABLED
		else:
			# Remove missing data
			clean_data = clean_missing(dataset, list(range(len(dataset))))
			# Ensure all values of any grouping variable are strings
			if grpby_col_index is not None:
				clean_data[grpby_col_index] = [str(grp) for grp in clean_data[grpby_col_index]]
			# Convert quantitative data types
			if plot_type not in ("Count by category", "Pareto chart"):
				x_data_type = [cs[1] for cs in self.column_specs if cs[0] == self.x_var.get()][0]
				cast_fn = data_type_cast_fn(x_data_type)
				for i in range(len(clean_data[0])):
					clean_data[0][i] = cast_fn(clean_data[0][i])
			if self.y_sel["state"] != tk.DISABLED and self.y_var.get() != "" and len(clean_data) > 1:
				y_data_type = [cs[1] for cs in self.column_specs if cs[0] == self.y_var.get()][0]
				cast_fn = data_type_cast_fn(y_data_type)
				for i in range(len(clean_data[1])):
					clean_data[1][i] = cast_fn(clean_data[1][i])
			# Sort the dataset by X values
			clean_data = sort_columns(clean_data)
			# Set data labels
			self.data_labels = column_list
			# Log-transform data if specified.
			if self.xlog_ck["state"] != tk.DISABLED and self.xlog_var.get() == "1":
				logged_x = logvector(clean_data[0])
				if logged_x is not None:
					clean_data[0] = logged_x
					self.data_labels[0] = "Log10 of " + self.x_var.get()
				else:
					self.xlog_var.set("0")
					self.xlog_ck["state"] = tk.DISABLED
			if self.ylog_ck["state"] != tk.DISABLED and self.ylog_var.get() == "1":
				logged_y = logvector(clean_data[1])
				if logged_y is not None:
					clean_data[1] = logged_y
					self.data_labels[1] = "Log10 of " + self.y_var.get()
				else:
					self.ylog_var.set("0")
					self.ylog_ck["state"] = tk.DISABLED
			#
			self.dataset = clean_data
			self.data_btn["state"] = tk.NORMAL
			# Summarize and sort the data as needed for each type of plot.
			# The label column, if present, is not retained when data are summarized.
			if plot_type == "Count by category":
				# Count of values for each X, ordered by X
				counter = collections.Counter(self.dataset[0])
				x_vals = list(counter.keys())
				x_vals.sort()
				x_counts = [counter[k] for k in x_vals]
				self.plot_data = [x_vals, x_counts]
				self.plot_data_labels = [self.data_labels[0], "Count"]
			elif plot_type in ("Box plot", "Histogram", "Kernel density (KD) plot", "Stripchart", "Violin plot"):
				if self.groupby_var.get() != '':
					self.plot_data_labels, self.plot_data = spread_by_groups(self.dataset[1], self.dataset[0])
				else:
					self.plot_data = [self.dataset[0]]
					self.plot_data_labels = [self.data_labels[0]]
			elif plot_type == "Breaks groups":
				x_vals = copy.copy(self.dataset[0])
				x_vals.sort()
				oj = optimum_jenks(x_vals, 8)
				jnb = jenkspy.JenksNaturalBreaks(oj)
				jnb.fit(x_vals)
				self.plot_data = [list(x) for x in jnb.groups_]
				self.plot_data_labels = [str(i+1) for i in range(len(self.plot_data))]
			elif plot_type == "Breaks optimum":
				nrows = len(self.dataset[0])
				if nrows < 4:
					warning("The data set must have at least 4 values.", {})
				else:
					x_vals = copy.copy(self.dataset[0])
					x_vals.sort()
					x, y = all_jenks_breaks(x_vals, min(len(self.dataset[0])-1, 8))
					oj = optimum_jenks(x_vals, min(nrows-1, 8))
					self.plot_data = [x, y, [x[oj-1]], [y[oj-1]]]
					self.plot_data_labels = ["Groups", "Goodness of Variance Fit", "Optimum group", "Optimum GVF"]
			elif plot_type == "Bubble plot":
					self.plot_data = self.dataset
					self.plot_data_labels = self.data_labels
			elif plot_type == "Date range by category":
				d1_type = [cs[1] for cs in self.column_specs if cs[0] == self.x_var.get()][0]
				d2_type = [cs[1] for cs in self.column_specs if cs[0] == self.y_var.get()][0]
				if d1_type != d2_type:
					if d1_type == "date":
						d1 = [datetime.datetime.combine(dt, datetime.datetime.min.time()) for dt in self.dataset[0]]
						d2 = self.dataset[1]
					else:
						d1 = self.dataset[0]
						d2 = [datetime.datetime.combine(dt, datetime.datetime.min.time()) for dt in self.dataset[1]]
				else:
					d1 = self.dataset[0]
					d2 = self.dataset[1]
				grplists = subset_by_groups([d1, d2], self.dataset[2])
				grps = list(grplists.keys())
				mins = []
				maxs = []
				order_error = False
				for g in grps:
					newmin = min(grplists[g][0])
					newmax = max(grplists[g][1])
					if newmax < newmin:
						order_error = True
					mins.append(newmin)
					maxs.append(newmax)
				self.plot_data = revsort_columns([grps, mins, maxs], sortby=1)
				self.plot_data_labels = [self.groupby_var.get(), self.x_var.get(), self.y_var.get()]
				if order_error:
					warning("First and last dates are out of order.", {"parent": self.dlg})
			elif plot_type == "Empirical CDF":
				# Y is the fraction of data points below each X value
				x_counts = np.unique(self.dataset[0], return_counts=True)
				y_vals = list(np.cumsum(x_counts[1]/np.sum(x_counts[1])))
				self.plot_data = [list(x_counts[0]), y_vals]
				self.plot_data_labels = [self.data_labels[0], "Cumulative frequency"]
			elif plot_type == "Min-max by bin":
				if self.bins is None:
					self.bins = doane_bins(self.dataset[1])
				if len(self.dataset[0]) > 1:
					min_vals, min_edges, min_no = spstats.binned_statistic(self.dataset[1], self.dataset[0], statistic='min', bins=self.bins)
					max_vals, max_edges, max_no = spstats.binned_statistic(self.dataset[1], self.dataset[0], statistic='max', bins=self.bins)
					bin_width = min_edges[1] - min_edges[0]
					bin_centers = min_edges[1:] - bin_width/2
					self.plot_data = [bin_centers, min_vals, max_vals]
				else:
					self.plot_data = [[],[],[]]
				xpfx = "" if self.xlog_var.get() == "0" else "Log10 "
				ypfx = "" if self.ylog_var.get() == "0" else "Log10 "
				self.plot_data_labels = [ypfx+self.y_var.get(), xpfx+self.x_var.get()+", min", xpfx+self.x_var.get()+", max"]
			elif plot_type == "Min-max by category":
				# Min and max X for each Y
				y_vals = list(set(self.dataset[1]))
				y_vals.sort()
				plotdata = dict(zip(y_vals, [[None, None] for _ in y_vals]))
				for i in range(len(self.dataset[1])):
					x = self.dataset[0][i]
					y = self.dataset[1][i]
					x_vals = plotdata[y]
					if x_vals[0] is None or x < x_vals[0]:
						plotdata[y][0] = x
					if x_vals[1] is None or x > x_vals[1]:
						plotdata[y][1] = x
				x1 = [plotdata[y][0] for y in y_vals]
				x2 = [plotdata[y][1] for y in y_vals]
				self.plot_data = [y_vals, x1, x2]
				self.plot_data_labels = [self.data_labels[1], self.data_labels[0] + ", min", self.data_labels[0] + ", max"]
			elif plot_type == "Normal Q-Q plot":
				x_vals = copy.copy(self.dataset[0])
				x_vals.sort()
				x_mean = statistics.mean(x_vals)
				x_sd = statistics.stdev(x_vals)
				if x_sd == 0:
					warning("Normal quantiles cannot be computed because there is no variance.", {"parent": self.dlg})
					self.x_var.set('')
					self.dataset = None
					self.data_labels = None
					self.plot_data = None
					self.plot_data_labels = None
				else:
					x_quantiles = [(x - x_mean)/x_sd for x in x_vals]
					x_len = len(x_vals)
					q = [(i + 0.5)/x_len for i in range(x_len)]
					norm_quantiles = [qnorm(p) for p in q]
					if x_len < 4:
						oj = 2
					else:
						oj = optimum_jenks(x_vals, min(x_len-1, 8))
					jnb = jenkspy.JenksNaturalBreaks(oj)
					jnb.fit(x_vals)
					self.plot_data = [x_vals, x_quantiles, norm_quantiles, jnb.labels_]
					self.plot_data_labels = [self.data_labels[0], "Sample quantiles", "Theoretical quantiles", "Group"]
			elif plot_type == "Pareto chart":
				# Count of values for each X, ordered by Y desc.
				counter = collections.Counter(self.dataset[0])
				x_vals = list(counter.keys())
				x_counts = [counter[k] for k in x_vals]
				pd = sort_columns([x_vals, x_counts], sortby=1)
				self.plot_data = [list(reversed(pd[0])), list(reversed(pd[1]))]
				total_cases = sum(self.plot_data[1])
				cumpct = list(np.cumsum([100*x/total_cases for x in self.plot_data[1]]))
				self.plot_data.append(cumpct)
				self.plot_data_labels = [self.data_labels[0], "Number of cases", "Cumulative percent"]
			elif plot_type == "Y range plot":
				# Min and max Y for each X
				x_vals = list(set(self.dataset[0]))
				x_vals.sort()
				plotdata = dict(zip(x_vals, [[None, None] for i in x_vals]))
				for i in range(len(self.dataset[0])):
					x = self.dataset[0][i]
					y = self.dataset[1][i]
					y_vals = plotdata[x]
					if y_vals[0] is None or y < y_vals[0]:
						plotdata[x][0] = y
					if y_vals[1] is None or y > y_vals[1]:
						plotdata[x][1] = y
				y1 = [plotdata[x][0] for x in x_vals]
				y2 = [plotdata[x][1] for x in x_vals]
				self.plot_data = [x_vals, y1, y2]
				self.plot_data_labels = [self.data_labels[0], self.data_labels[1] + " min", self.data_labels[1] + " max"]
			elif plot_type == "Line plot":
				ds = sort_columns(self.dataset)
				if self.groupby_var.get() == "* Breaks in X":
					if self.x_var.get() in self.numeric_columns:
						oj = optimum_jenks(ds[0], 8)
						jnb = jenkspy.JenksNaturalBreaks(oj)
						jnb.fit(ds[0])
						self.plot_data = [ds[0], ds[1], ["Group "+str(lbl+1) for lbl in jnb.labels_]]
						self.plot_data_labels = self.data_labels[0:2] + ['Breaks in X']
						if len(self.data_labels) == 3:
							# Include the label column
							self.plot_data.append(ds[2])
							self.plot_data_labels.append(self.data_labels[2])
					else:
						# Can't compute breaks for X
						self.groupby_var.set('')
						self.plot_data = self.dataset
						self.plot_data_labels = self.data_labels
				else:
					self.plot_data = ds
					self.plot_data_labels = self.data_labels
			elif plot_type == "Scatter plot":
				if self.groupby_var.get() == "* Breaks in X":
					if self.x_var.get() in self.numeric_columns:
						ds = sort_columns(self.dataset)
						oj = optimum_jenks(ds[0], 8)
						jnb = jenkspy.JenksNaturalBreaks(oj)
						jnb.fit(ds[0])
						self.plot_data = [ds[0], ds[1], ["Group "+str(lbl+1) for lbl in jnb.labels_]]
						self.plot_data_labels = self.data_labels[0:2] + ['Breaks in X']
						if len(self.data_labels) == 3:
							# Include the label column
							self.plot_data.append(ds[2])
							self.plot_data_labels.append(self.data_labels[2])
					else:
						# Can't compute breaks for X
						self.groupby_var.set('')
						self.plot_data = self.dataset
						self.plot_data_labels = self.data_labels
				elif self.groupby_var.get() == "* Breaks in Y":
					if self.y_var.get() in self.numeric_columns:
						ds = sort_columns(self.dataset, sortby=1)
						oj = optimum_jenks(ds[1], 8)
						jnb = jenkspy.JenksNaturalBreaks(oj)
						jnb.fit(ds[1])
						self.plot_data = [ds[0], ds[1], ["Group "+str(lbl+1) for lbl in jnb.labels_]]
						self.plot_data_labels = self.data_labels[0:2] + ['Breaks in Y']
						if len(self.data_labels) == 3:
							# Include the label column
							self.plot_data.append(ds[2])
							self.plot_data_labels.append(self.data_labels[2])
					else:
						# Can't compute breaks for Y
						self.groupby_var.set('')
						self.plot_data = self.dataset
						self.plot_data_labels = self.data_labels
				else:
					self.plot_data = self.dataset
					self.plot_data_labels = self.data_labels
			elif plot_type in ("Mean by category", "CV by category", "Total by category"):
				# Both X and groupby variables must be set
				grp_vals = list(set(self.dataset[1]))
				grp_vals.sort()
				ds = list(zip(self.dataset[1], self.dataset[0]))
				plot_data = []
				for g in grp_vals:
					plot_data.append([d[1] for d in ds if d[0] == g])
				self.plot_data = plot_data
				self.plot_data_labels = grp_vals
			self.plot_data_btn["state"] = tk.NORMAL

	def redraw(self):
		plot_type = self.type_var.get()
		if self.plot_data is not None and len(self.plot_data) > 0 and len(self.plot_data[0]) > 0:
			if plot_type == "Count by category":
				self.dlg.bind("<Alt-r>", self.set_rotated)
				self.dlg.bind("<Alt-f>", self.set_flip_y)
				groups = self.plot_data[0]
				grp_lbls = []
				for g in groups:
					if wrap_at_underscores:
						g = g.replace("_", " ")
					grp_lbls.append("\n".join(textwrap.wrap(g, width=wrapwidth)))
				if not self.rotated:
					self.plot.axes.bar(grp_lbls, self.plot_data[1])
					self.plot.set_axis_labels(self.xlabel or self.plot_data_labels[0], self.ylabel or self.plot_data_labels[1])
				else:
					self.plot.axes.barh(grp_lbls, self.plot_data[1])
					self.plot.set_axis_labels(self.ylabel or self.plot_data_labels[1], self.xlabel or self.plot_data_labels[0])
					if self.flip_y:
						self.plot.axes.invert_yaxis()
			elif plot_type == "Histogram":
				if self.bins is None:
					self.bins = doane_bins(self.plot_data[0])
				if self.groupby_var.get() == '':
					self.plot.axes.hist(self.plot_data[0], bins=self.bins)
				else:
					self.plot.axes.hist(self.plot_data, bins=self.bins, stacked=True, label=self.plot_data_labels)
					self.plot.axes.legend()
				self.plot.set_axis_labels(self.xlabel or self.x_var.get(), "Counts")
			elif plot_type == "Scatter plot":
				self.dlg.bind("<Alt-a>", self.set_alpha)
				self.dlg.bind("<Alt-b>", self.set_scatter_breaks)
				self.dlg.bind("<Alt-l>", self.set_loess)
				self.dlg.bind("<Alt-r>", self.set_linreg)
				self.dlg.bind("<Alt-s>", self.set_theilsen)
				# Remover orange from the palette because it is used for the linear fit line.
				scatter_palette = [quant_colors[0]] + quant_colors[2:]
				if self.scatter_breaks:
					if self.x_var.get() in self.numeric_columns:
						x_vals = copy.copy(self.plot_data[0])
						nx = len(x_vals)
						if nx < 4:
							warning("There must be at least 4 X values to calculate natural breaks.", {})
						else:
							x_vals.sort()
							oj = optimum_jenks(x_vals, min(nx-1, 8))
							jnb = jenkspy.JenksNaturalBreaks(oj)
							jnb.fit(x_vals)
							self.scatter_x_breaks = jnb.inner_breaks_
							for b in self.scatter_x_breaks:
								self.plot.axes.axvline(b, color='0.8', alpha=0.5)
					if self.y_var.get() in self.numeric_columns:
						y_vals = copy.copy(self.plot_data[1])
						ny = len(y_vals)
						if ny < 4:
							warning("There must be at least 4 Y values to calculate natural breaks.", {})
						else:
							y_vals.sort()
							oj = optimum_jenks(y_vals, min(ny-1, 8))
							jnb = jenkspy.JenksNaturalBreaks(oj)
							jnb.fit(y_vals)
							self.scatter_y_breaks = jnb.inner_breaks_
							for b in self.scatter_y_breaks:
								self.plot.axes.axhline(b, color='0.8', alpha=0.5)
				# Flip Y?
				if self.flip_y:
					self.plot.axes.invert_yaxis()
				# Single or grouped scatter plot
				if self.groupby_var.get() == '':
					splot = self.plot.axes.scatter(self.plot_data[0], self.plot_data[1], alpha=self.alpha)
					handle_list = []
					label_list = []
				else:
					groups = sorted_numstrs(list(set(self.plot_data[2])))
					cmap = map_colors(groups, scatter_palette)
					colors = [cmap[g] for g in self.plot_data[2]]
					splot = self.plot.axes.scatter(self.plot_data[0], self.plot_data[1], c=colors, alpha=self.alpha)
					# Custom legend handle for grouped scatter plot
					lbls = list(cmap.keys())
					if wrap_at_underscores:
						lbls = [l.replace("_", " ") for l in lbls]
					lbls = ["\n".join(textwrap.wrap(l, width=wrapwidth)) for l in lbls]
					colkey = list(cmap.values())
					symbs = [matplotlib.lines.Line2D([], [], marker='o', color=colkey[i], label=lbls[i], linestyle='None') for i in range(len(lbls))]
					handle_list = symbs
					label_list = lbls
				bbox = self.plot.axes.get_clip_box()
				# Hover annotation
				if len(self.plot_data) > self.n_dataset_columns:
					hoverer = Hoverer(self.plot, self.plot_data[self.n_dataset_columns], splot)
					self.canvas_conn_id = self.plot.canvas.mpl_connect("motion_notify_event", lambda ev: hoverer.hover(ev))
				else:
					if self.canvas_conn_id is not None:
						self.plot.canvas.mpl_disconnect(self.canvas_conn_id)
				# LOESS line
				if self.loess and self.x_var.get() in self.numeric_columns:
					loess_x, loess_y, wts = loess_1d(np.array(self.plot_data[0]), np.array(self.plot_data[1]))
					self.plot.axes.plot(loess_x, loess_y, label="LOESS", color="black", linestyle="dashed", alpha=0.65)
					# Use a proxy artist for the handle
					handle_list.append(matplotlib.lines.Line2D([], [], color="black", linestyle="dashed"))
					label_list.append("LOESS")
				# Regression line
				if self.linreg and self.x_var.get() in self.numeric_columns:
					ns, stats = Polynomial.fit(np.array(self.plot_data[0]), np.array(self.plot_data[1]), 1, full=True)
					intercept, slope = ns.convert().coef
					# Use the mean point to plot instead of the intercept because the intercept may be out of the bounding box, especially for log-transformed data
					xmean = statistics.fmean(self.plot_data[0])
					ymean = statistics.fmean(self.plot_data[1])
					handle_list.append(self.plot.axes.axline((xmean, ymean), slope=slope, clip_box=bbox, clip_on=True, label="Linear fit", color="darkorange", linestyle="dashdot", linewidth=2, alpha=0.65))
					label_list.append("Linear fit")
					if show_regression_stats:
						N = len(self.plot_data[1])
						total_ss = sum([(self.plot_data[1][i] - ymean)**2 for i in range(N)])
						resid_ss = list(stats[0])[0]
						exp_ss = total_ss - resid_ss
						if total_ss == 0.0:
							r_square = "NC"
						else:
							r_square = fp_display(exp_ss/total_ss, 4)
						statdlg = MsgDialog(title="Linear Regression", message = "Slope: %s\nIntercept: %s\nR-square: %s\nN: %s" % \
								(fp_display(slope, 4), fp_display(intercept, 4), r_square, N))
						statdlg.show()
				# Theil-Sen line
				if self.theilsen and self.x_var.get() in self.numeric_columns:
					ts_slope, ts_intercept, ts_high, ts_low = spstats.theilslopes(np.array(self.plot_data[1]), np.array(self.plot_data[0]))
					handle_list.append(self.plot.axes.axline((statistics.median(self.plot_data[0]), statistics.median(self.plot_data[1])), slope=ts_slope, \
							clip_box=bbox, clip_on=True, label="Theil-Sen line", color="darkgreen", linestyle="dotted", linewidth=2, alpha=0.65))
					label_list.append("Theil-Sen line")
				# Legend
				if len(handle_list) > 0:
					self.plot.axes.legend(handles=handle_list, labels=label_list)
				# Axis labels
				self.plot.set_axis_labels(self.xlabel or self.plot_data_labels[0], self.ylabel or self.plot_data_labels[1])
			elif plot_type == "Bubble plot":
				self.dlg.bind("<Alt-a>", self.set_alpha)
				self.dlg.bind("<Alt-f>", self.set_flip_y)
				sz = self.z_width_var.get()
				z_data = [float(v) for v in self.dataset[1]]
				max_v = max(z_data)
				sizes = [v * sz/max_v for v in z_data]
				# Flip Y?
				if self.flip_y:
					self.plot.axes.invert_yaxis()
				# Single or grouped
				if self.groupby_var.get() == '':
					self.plot.axes.scatter(self.plot_data[0], self.plot_data[1], s=sizes, alpha=self.alpha)
				else:
					grpvals = sorted(list(set(self.plot_data[3])))
					cmap = map_colors(grpvals)
					colors = [cmap[g] for g in self.plot_data[3]]
					self.plot.axes.scatter(self.plot_data[0], self.plot_data[1], s=sizes, c=colors, alpha=self.alpha)
					# Custom legend
					lbls = list(cmap.keys())
					if wrap_at_underscores:
						lbls = [l.replace("_", " ") for l in lbls]
					lbls = ["\n".join(textwrap.wrap(l, width=wrapwidth)) for l in lbls]
					colkey = list(cmap.values())
					symbs = [matplotlib.lines.Line2D([], [], marker='o', color=colkey[i], label=lbls[i], linestyle='None') for i in range(len(lbls))]
					self.plot.axes.legend(handles=symbs)
				# Axis labels
				self.plot.set_axis_labels(self.xlabel or self.plot_data_labels[0], self.ylabel or self.plot_data_labels[1])
			elif plot_type in ("CV by category"):
				self.dlg.bind("<Alt-r>", self.set_rotated)
				self.dlg.bind("<Alt-f>", self.set_flip_y)
				values = np.array([statistics.stdev(x)/statistics.fmean(x) for x in self.plot_data if len(x) > 1])
				if len(values) > 0:
					lbl = "Coefficient of variation of "
					groups = [self.plot_data_labels[i] for i in range(len(self.plot_data)) if len(self.plot_data[i]) > 1]
					grp_lbls = []
					for g in groups:
						if wrap_at_underscores:
							g = g.replace("_", " ")
						grp_lbls.append("\n".join(textwrap.wrap(g, width=wrapwidth)))
					loglbl = "" if self.xlog_var.get() == "0" else "Log10 of "
					if not self.rotated:
						self.plot.axes.bar(grp_lbls, values)
						self.plot.set_axis_labels(self.xlabel or self.groupby_var.get(), self.ylabel or lbl + loglbl + self.x_var.get())
					else:
						self.plot.axes.barh(grp_lbls, means)
						self.plot.set_axis_labels(self.ylabel or lbl + loglbl + self.x_var.get(), self.xlabel or self.groupby_var.get())
						if self.flip_y:
							self.plot.axes.invert_yaxis()
			elif plot_type == "Line plot":
				self.dlg.bind("<Alt-a>", self.set_alpha)
				self.dlg.bind("<Alt-b>", self.set_lineplot_breaks)
				self.dlg.bind("<Alt-l>", self.set_loess)
				self.dlg.bind("<Alt-r>", self.set_linreg)
				self.dlg.bind("<Alt-s>", self.set_theilsen)
				if self.lineplot_breaks:
					if self.x_var.get() in self.numeric_columns:
						x_vals = copy.copy(self.plot_data[0])
						#x_vals.sort()
						nx = len(x_vals)
						if nx < 4:
							warning("There must be at least 4 X values to calculate natural breaks.", {})
						else:
							oj = optimum_jenks(x_vals, min(nx-1, 8))
							jnb = jenkspy.JenksNaturalBreaks(oj)
							jnb.fit(x_vals)
							self.lineplot_x_breaks = jnb.inner_breaks_
							for b in self.lineplot_x_breaks:
								self.plot.axes.axvline(b, color='0.8', alpha=0.5)
				if self.groupby_var.get() == '':
					self.plot.axes.plot(self.plot_data[0], self.plot_data[1], alpha=self.alpha)
				else:
					groups = list(set(self.plot_data[2]))
					groups.sort()
					datarows = len(self.plot_data[0])
					for g in groups:
						pdx = [self.plot_data[0][i] for i in range(datarows) if self.plot_data[2][i] == g]
						pdy = [self.plot_data[1][i] for i in range(datarows) if self.plot_data[2][i] == g]
						self.plot.axes.plot(pdx, pdy, label=g, alpha=self.alpha)
				bbox = self.plot.axes.get_clip_box()
				if self.loess and self.x_var.get() in self.numeric_columns:
					loess_x, loess_y, wts = loess_1d(np.array(self.plot_data[0]), np.array(self.plot_data[1]))
					self.plot.axes.plot(loess_x, loess_y, label="LOESS", color="black", linestyle="dashed", alpha=0.65)
				if self.linreg and self.x_var.get() in self.numeric_columns:
					ns, stats = Polynomial.fit(np.array(self.plot_data[0]), np.array(self.plot_data[1]), 1, full=True)
					intercept, slope = ns.convert().coef
					# Use the mean point to plot instead of the intercept because the intercept may be out of the bounding box, especially for log-transformed data
					xmean = statistics.fmean(self.plot_data[0])
					ymean = statistics.fmean(self.plot_data[1])
					self.plot.axes.axline((xmean, ymean), slope=slope, clip_box=bbox, clip_on=True, label="Linear fit", color="darkorange", linestyle="dashdot", linewidth=2, alpha=0.65)
					if show_regression_stats:
						N = len(self.plot_data[1])
						total_ss = sum([(self.plot_data[1][i] - ymean)**2 for i in range(N)])
						resid_ss = list(stats[0])[0]
						exp_ss = total_ss - resid_ss
						if total_ss == 0.0:
							r_square = "NC"
						else:
							r_square = fp_display(exp_ss/total_ss, 4)
						statdlg = MsgDialog(title="Linear Regression", message = "Slope: %s\nIntercept: %s\nR-square: %s\nN: %s" % \
								(fp_display(slope, 4), fp_display(intercept, 4), r_square, N))
						statdlg.show()
				if self.theilsen and self.x_var.get() in self.numeric_columns:
					ts_slope, ts_intercept, ts_high, ts_low = spstats.theilslopes(np.array(self.plot_data[1]), np.array(self.plot_data[0]))
					self.plot.axes.axline((statistics.median(self.plot_data[0]), statistics.median(self.plot_data[1])), slope=ts_slope, \
							clip_box=bbox, clip_on=True, label="Theil-Sen line", color="darkgreen", linestyle="dotted", linewidth=2, alpha=0.65)
				if self.groupby_var.get() != '' or ((self.loess or self.linreg or self.theilsen) and self.x_var.get() in self.numeric_columns):
					self.plot.axes.legend()
				self.plot.set_axis_labels(self.xlabel or self.plot_data_labels[0], self.ylabel or self.plot_data_labels[1])
			elif plot_type == "Breaks groups":
				for i in range(len(self.plot_data)):
					xs = [i+1] * len(self.plot_data[i])
					self.plot.axes.scatter(xs, self.plot_data[i])
				ticks = [int(l) for l in self.plot_data_labels]
				ticks.insert(0,0)
				ticks.append(ticks[-1]+1)
				self.plot.axes.set_xticks(ticks)
				lbls = copy.copy(self.plot_data_labels)
				lbls.insert(0, "")
				lbls.append("")
				self.plot.axes.set_xticklabels(lbls)
				self.plot.set_axis_labels("Group", self.ylabel or self.data_labels[0])
			elif plot_type == "Breaks optimum":
				self.plot.axes.plot(self.plot_data[0], self.plot_data[1])
				self.plot.axes.scatter(self.plot_data[2], self.plot_data[3])
				self.plot.set_axis_labels(self.xlabel or self.plot_data_labels[0], self.ylabel or self.plot_data_labels[1])
			elif plot_type == "Date range by category":
				self.dlg.bind("<Alt-f>", self.set_flip_y)
				self.plot.axes.hlines(self.plot_data[0], self.plot_data[1], self.plot_data[2], linewidths=5.0)
				self.plot.set_axis_labels(self.xlabel or self.x_var.get()+" to "+self.y_var.get(), self.ylabel or self.plot_data_labels[0])
				if self.flip_y:
					self.plot.axes.invert_yaxis()
			elif plot_type == "Empirical CDF":
				self.plot.axes.stackplot(self.plot_data[0], self.plot_data[1])
				self.plot.set_axis_labels(self.xlabel or self.plot_data_labels[0], self.ylabel or self.plot_data_labels[1])
			elif plot_type == "Kernel density (KD) plot":
				self.dlg.bind("<Alt-a>", self.set_alpha)
				if self.groupby_var.get() == '':
					sns.kdeplot({self.x_var.get(): self.dataset[0]}, x=self.x_var.get(), fill=True, ax=self.plot.axes)
					self.plot.axes.set_xlabel(self.xlabel or self.x_var.get())
				else:
					grplbls = copy.copy(self.dataset[1])
					if wrap_at_underscores:
						grplbls = [lbl.replace("_", " ") for lbl in grplbls]
					grplbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in grplbls]
					grplbls = ["\n".join(lbl) for lbl in grplbls]
					sns.kdeplot({self.groupby_var.get(): grplbls, self.x_var.get(): self.dataset[0]},
							x=self.x_var.get(), hue=self.groupby_var.get(), multiple="layer", fill=True, alpha=self.alpha, ax=self.plot.axes,
							warn_singular=False)
					self.plot.axes.set_xlabel(self.xlabel or self.x_var.get())
				self.plot.axes.set_ylabel("Density")
			elif plot_type in ("Mean by category"):
				self.dlg.bind("<Alt-r>", self.set_rotated)
				self.dlg.bind("<Alt-f>", self.set_flip_y)
				values = np.array([statistics.fmean(x) for x in self.plot_data])
				lbl = "Mean "
				groups = self.plot_data_labels
				grp_lbls = []
				for g in groups:
					if wrap_at_underscores:
						g = g.replace("_", " ")
					grp_lbls.append("\n".join(textwrap.wrap(g, width=wrapwidth)))
				loglbl = "" if self.xlog_var.get() == "0" else "Log10 of "
				if not self.rotated:
					self.plot.axes.bar(grp_lbls, values)
					self.plot.set_axis_labels(self.xlabel or self.groupby_var.get(), self.ylabel or lbl + loglbl + self.x_var.get())
				else:
					self.plot.axes.barh(grp_lbls, values)
					self.plot.set_axis_labels(self.ylabel or lbl + loglbl + self.x_var.get(), self.xlabel or self.groupby_var.get())
					if self.flip_y:
						self.plot.axes.invert_yaxis()
			elif plot_type == "Min-max by bin":
				if len(self.plot_data[0]) > 1:
					self.dlg.bind("<Alt-r>", self.set_rotated)
					self.dlg.bind("<Alt-f>", self.set_flip_y)
					self.dlg.bind("<Alt-b>", self.set_bins)
					# Data with min = max will be plotted as points because hlines and vlines do not show them.
					point_data = [[],[]]
					for i in range(len(self.plot_data[0])):
						if self.plot_data[1][i] == self.plot_data[2][i]:
							point_data[0].append(self.plot_data[0][i])
							point_data[1].append(self.plot_data[1][i])
					if not self.rotated:
						self.plot.axes.hlines(self.plot_data[0], self.plot_data[1], self.plot_data[2], linewidths=5.0)
						self.plot.axes.set_ylabel(self.plot_data_labels[0])
						self.plot.axes.set_xlabel(("" if self.xlog_var.get() == "0" else "Log10 ") + self.x_var.get())
						if len(point_data[0]) > 0:
							self.plot.axes.scatter(point_data[0], point_data[1])
					else:
						self.plot.axes.vlines(self.plot_data[0], self.plot_data[1], self.plot_data[2], linewidths=5.0)
						self.plot.set_axis_labels(self.plot_data_labels[0], ("" if self.xlog_var.get() == "0" else "Log10 ") + self.x_var.get())
						if len(point_data[0]) > 0:
							self.plot.axes.scatter(point_data[1], point_data[0])
					if self.flip_y:
						self.plot.axes.invert_yaxis()
			elif plot_type == "Min-max by category":
				self.dlg.bind("<Alt-r>", self.set_rotated)
				self.dlg.bind("<Alt-f>", self.set_flip_y)
				if not self.rotated:
					self.plot.axes.hlines(self.plot_data[0], self.plot_data[1], self.plot_data[2], linewidths=5.0)
					if self.xlog_ck["state"] != tk.DISABLED and self.xlog_var.get() == "1":
						self.plot.axes.set_xlabel(self.xlabel or "Log10 of " + self.x_var.get())
					else:
						self.plot.axes.set_xlabel(self.xlabel or self.x_var.get())
					self.plot.axes.set_ylabel(self.ylabel or self.plot_data_labels[0])
				else:
					self.plot.axes.vlines(self.plot_data[0], self.plot_data[1], self.plot_data[2], linewidths=5.0)
					if self.xlog_ck["state"] != tk.DISABLED and self.xlog_var.get() == "1":
						self.plot.axes.set_ylabel(self.xlabel or "Log10 of " + self.x_var.get())
					else:
						self.plot.axes.set_ylabel(self.xlabel or self.x_var.get())
					self.plot.axes.set_xlabel(self.ylabel or self.plot_data_labels[0])
				if self.flip_y:
					self.plot.axes.invert_yaxis()
			elif plot_type == "Normal Q-Q plot":
				self.dlg.bind("<Alt-a>", self.set_alpha)
				if self.qq_groups:
					self.plot.axes.scatter(self.plot_data[2], self.plot_data[1], c=self.plot_data[3], cmap="tab10", alpha=self.alpha)
				else:
					self.plot.axes.scatter(self.plot_data[2], self.plot_data[1], alpha=self.alpha)
				pmin = min(self.plot_data[2][0], self.plot_data[1][0])
				pmax = max(self.plot_data[2][-1], self.plot_data[1][-1])
				self.plot.axes.plot([pmin, pmax], [pmin, pmax])
				self.plot.set_axis_labels(self.xlabel or self.plot_data_labels[2], self.ylabel or self.plot_data_labels[1])
			elif plot_type == "Pareto chart":
				groups = self.plot_data[0]
				grp_lbls = []
				for g in groups:
					if wrap_at_underscores:
						g = g.replace("_", " ")
					grp_lbls.append("\n".join(textwrap.wrap(g, width=wrapwidth)))
				# Plot bars
				self.plot.axes.bar(grp_lbls, self.plot_data[1])
				self.plot.set_axis_labels(self.xlabel or self.plot_data_labels[0], self.ylabel or self.plot_data_labels[1])
				# Plot cumulative percentage
				ax2 = self.plot.axes.twinx()
				ax2.set_ylim(0, 105)
				ax2.plot(self.plot_data[0], self.plot_data[2], color=quant_colors[1], marker=".")
				ax2.set_ylabel("Cumulative percent")
			elif plot_type == "Y range plot":
				self.dlg.bind("<Alt-f>", self.set_flip_y)
				self.plot.axes.fill_between(self.plot_data[0], self.plot_data[1], self.plot_data[2])
				if self.xlog_ck["state"] != tk.DISABLED and self.xlog_var.get() == "1":
					self.plot.axes.set_xlabel(self.xlabel or "Log10 of " + self.x_var.get())
				else:
					self.plot.axes.set_xlabel(self.xlabel or self.x_var.get())
				if self.ylog_ck["state"] != tk.DISABLED and self.ylog_var.get() == "1":
					self.plot.axes.set_ylabel(self.ylabel or "Log10 of " + self.y_var.get())
				else:
					self.plot.axes.set_ylabel(self.ylabel or self.y_var.get())
				if self.flip_y:
					self.plot.axes.invert_yaxis()
			elif plot_type == "Box plot":
				self.dlg.bind("<Alt-r>", self.set_rotated)
				orient = not self.rotated
				grplbls = copy.copy(self.plot_data_labels)
				if wrap_at_underscores:
					grplbls = [lbl.replace("_", " ") for lbl in grplbls]
				grplbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in grplbls]
				grplbls = ["\n".join(lbl) for lbl in grplbls]
				self.plot.axes.boxplot(self.plot_data, labels=grplbls, vert=orient)
				if not self.rotated:
					self.plot.set_axis_labels(self.xlabel or self.groupby_var.get(), self.ylabel or self.data_labels[0])
				else:
					self.plot.set_axis_labels(self.ylabel or self.data_labels[0], self.xlabel or self.groupby_var.get())
			elif plot_type == "Stripchart":
				self.dlg.bind("<Alt-a>", self.set_alpha)
				self.dlg.bind("<Alt-r>", self.set_rotated)
				orientation = "v" if not self.rotated else "h"
				if self.groupby_var.get() != '':
					grplbls = copy.copy(self.dataset[1])
					if wrap_at_underscores:
						grplbls = [lbl.replace("_", " ") for lbl in grplbls]
					grplbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in grplbls]
					grplbls = ["\n".join(lbl) for lbl in grplbls]
					if not self.rotated:
						sns.stripplot({self.groupby_var.get(): grplbls, \
								self.x_var.get(): self.dataset[0]}, \
								x=self.groupby_var.get(), y=self.x_var.get(), alpha=self.alpha, \
								ax=self.plot.axes, orient=orientation)
						self.plot.set_axis_labels(self.xlabel or self.groupby_var.get(), self.ylabel or self.data_labels[0])
					else:
						sns.stripplot({self.groupby_var.get(): grplbls, \
								self.x_var.get(): self.dataset[0]}, \
								y=self.groupby_var.get(), x=self.x_var.get(), alpha=self.alpha, \
								ax=self.plot.axes, orient=orientation)
						self.plot.set_axis_labels(self.ylabel or self.groupby_var.get(), self.xlabel or self.data_labels[0])
				else:
					if not self.rotated:
						sns.stripplot({self.x_var.get(): self.dataset[0]}, y=self.x_var.get(), \
								alpha=self.alpha, ax=self.plot.axes, orient=orientation)
						self.plot.set_axis_labels(self.xlabel or self.x_var.get(), self.ylabel or self.data_labels[0])
					else:
						sns.stripplot({self.x_var.get(): self.dataset[0]}, x=self.x_var.get(), \
								alpha=self.alpha, ax=self.plot.axes, orient=orientation)
						self.plot.set_axis_labels(self.ylabel or self.data_labels[0], self.xlabel or self.x_var.get())
			elif plot_type in ("Total by category"):
				self.dlg.bind("<Alt-r>", self.set_rotated)
				self.dlg.bind("<Alt-f>", self.set_flip_y)
				totals = np.array([sum(x) for x in self.plot_data])
				groups = self.plot_data_labels
				grp_lbls = []
				for g in groups:
					if wrap_at_underscores:
						g = g.replace("_", " ")
					grp_lbls.append("\n".join(textwrap.wrap(g, width=wrapwidth)))
				loglbl = "" if self.xlog_var.get() == "0" else "Log10 of "
				if not self.rotated:
					self.plot.axes.bar(grp_lbls, totals)
					self.plot.set_axis_labels(self.xlabel or self.groupby_var.get(), self.ylabel or "Total " + loglbl + self.x_var.get())
				else:
					self.plot.axes.barh(grp_lbls, totals)
					self.plot.set_axis_labels(self.ylabel or "Total " + loglbl + self.x_var.get(), self.xlabel or self.groupby_var.get())
					if self.flip_y:
						self.plot.axes.invert_yaxis()
			elif plot_type == "Violin plot":
				self.dlg.bind("<Alt-r>", self.set_rotated)
				orientation = "v" if not self.rotated else "h"
				if self.groupby_var.get() != '':
					grplbls = copy.copy(self.dataset[1])
					if wrap_at_underscores:
						grplbls = [lbl.replace("_", " ") for lbl in grplbls]
					grplbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in grplbls]
					grplbls = ["\n".join(lbl) for lbl in grplbls]
					if not self.rotated:
						sns.violinplot({self.groupby_var.get(): grplbls, self.x_var.get(): self.dataset[0]},
								x=self.groupby_var.get(), y=self.x_var.get(), ax=self.plot.axes)
						self.plot.set_axis_labels(self.xlabel or self.groupby_var.get(), self.ylabel or self.data_labels[0])
					else:
						sns.violinplot({self.groupby_var.get(): grplbls, self.x_var.get(): self.dataset[0]},
								y=self.groupby_var.get(), x=self.x_var.get(), ax=self.plot.axes)
						self.plot.set_axis_labels(self.ylabel or self.data_labels[0], self.xlabel or self.groupby_var.get())
				else:
					if not self.rotated:
						sns.violinplot({self.x_var.get(): self.dataset[0]}, y=self.x_var.get(), alpha=self.alpha, ax=self.plot.axes)
						self.plot.set_axis_labels(self.xlabel or self.x_var.get(), self.ylabel or self.data_labels[0])
					else:
						sns.violinplot({self.x_var.get(): self.dataset[0]}, x=self.x_var.get(), alpha=self.alpha, ax=self.plot_axes)
						self.plot.set_axis_labels(self.ylabel or self.data_labels[0], self.xlabel or self.x_var.get())
			if self.plot_title is not None:
				self.plot.axes.set_title(self.plot_title)
				self.plot.axes.title.set_visible(True)
			else:
				self.plot.axes.title.set_visible(False)
			self.plot.draw()

	def set_title(self, *args):
		dlg = OneEntryDialog(self.dlg, "Plot Title", "Enter a title for the plot:", init_value=self.plot_title, nullable=True)
		self.plot_title = dlg.show()
		self.plot.axes.set_title(self.plot_title)
		self.plot.axes.title.set_visible(self.plot_title is not None)
		self.plot.draw()
	def set_xlabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "X-Axis Label", "Enter a title for the un-rotated X-axis label:", init_value=self.xlabel, nullable=True)
		self.xlabel = dlg.show()
		if not self.rotated:
			self.plot.axes.set_xlabel(self.xlabel or self.x_var.get())
		else:
			self.plot.axes.set_ylabel(self.xlabel or self.x_var.get())
		self.plot.draw()
	def set_ylabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "Y-Axis Label", "Enter a title for the un-rotated Y-axis label:", init_value=self.ylabel, nullable=True)
		self.ylabel = dlg.show()
		if not self.rotated:
			self.plot.axes.set_ylabel(self.ylabel or self.y_var.get())
		else:
			self.plot.axes.set_xlabel(self.ylabel or self.y_var.get())
		self.plot.draw()
	def set_bins(self, *args):
		if self.type_var.get() in ("Min-max by bin", "Histogram"):
			dlg = OneIntDialog(self.dlg, "Data Bins", "Enter the number of bins to be used for histograms and min-max plots", min_value=2, max_value=50, initial=self.bins)
			num_bins = dlg.show()
			if num_bins is not None:
				self.bins = num_bins
				self.q_redraw()
	def set_alpha(self, *args):
		if self.type_var.get() in ("Box plot", "Bubble plot", "Scatter plot", "Kernel density (KD) plot", "Line plot", "Normal Q-Q plot", "Stripchart", "Violin plot"):
			dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the transparency (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
			new_alpha = dlg.show()
			if new_alpha is not None:
				self.alpha = min(1.0, max(new_alpha, 0.0))
				self.q_redraw()
	def set_scatter_breaks(self, *args):
		# Toggle display of vertical and horizontal lines at natural breaks values on a scatter plot.
		if self.type_var.get() == "Scatter plot":
			self.scatter_breaks = not self.scatter_breaks
			self.q_redraw()
	def set_lineplot_breaks(self, *args):
		# Toggle display of vertical lines at natural breaks values on a scatter plot.
		if self.type_var.get() == "Line plot":
			self.lineplot_breaks = not self.lineplot_breaks
			self.q_redraw()
	def set_loess(self, *args):
		# Toggle display of loess line on a scatter plot.
		if self.type_var.get() in ("Line plot", "Scatter plot"):
			self.loess = not self.loess
			self.q_redraw()
	def set_linreg(self, *args):
		# Toggle display of linear regression line on a scatter plot.
		if self.type_var.get() in ("Line plot", "Scatter plot"):
			self.linreg = not self.linreg
			self.q_redraw()
	def set_theilsen(self, *args):
		if self.type_var.get() in ("Line plot", "Scatter plot"):
			self.theilsen = not self.theilsen
			self.q_redraw()
	def set_rotated(self, *args):
		if self.type_var.get() in ("Count by category", "Mean by category", "Min-max by bin", "Min-max by category", "Box plot", "Stripchart", "Total by category", "Violin plot"):
			self.rotated = not self.rotated
			self.q_redraw()
	def set_flip_y(self, *args):
		if self.type_var.get() in ("Bubble plot", "Count by category", "Date range by category", "Mean by category", "Min-max by bin",
				"Min-max by category", "Scatter plot", "Total by category", "Y range plot"):
			self.flip_y = not self.flip_y
			self.q_redraw()
	def show_groups(self, *args):
		if self.type_var.get() == "Normal Q-Q plot":
			self.qq_groups = not self.qq_groups
			self.q_redraw()
	def do_close(self, *args):
		self.parent.remove_plot(self)
		super().do_cancel(args)
	def show(self):
		center_window(self.dlg, x_offset=-225, y_offset=-250)
		self.dlg.deiconify()
		self.dlg.geometry(self.geom)
		raise_window(self.dlg)
		self.dlg.wait_window(self.dlg)


def rosners_stat(x):
	# Calculates the statistic for Rosner's test.
	# Returns the standard score (Z score) for the value with the maximum absolute
	# deviation from the mean, and the index into the data list for that value.
	# From https://www.itl.nist.gov/div898/handbook/eda/section3/eda35h3.htm
	mn, sd = statistics.fmean(x), statistics.stdev(x)
	absdevs = [abs(v - mn) for v in x]
	maxdev = max(absdevs)
	maxidx = absdevs.index(maxdev)
	return maxdev/sd, maxidx

def rosners_critical(n, alpha, outlier_no):
	# From https://www.itl.nist.gov/div898/handbook/eda/section3/eda35h3.htm
	tval = spstats.t.ppf(1 - alpha / (2*(n-outlier_no+1)), n-outlier_no-1)
	return ((n-1) * tval) / (math.sqrt((n - outlier_no - 1 + tval**2) * (n - outlier_no + 1)))

def rosners_test(x, alpha, max_outliers):
	# Carries out Rosner's test for up to max_outliers.
	# Returns the number of outliers, from 0 up to max_outliers.
	d = copy.copy(x)
	n_outliers = 0
	for i in range(max_outliers):
		rs, ix = rosners_stat(d)
		cv = rosners_critical(len(d), alpha, i+1)
		if rs > cv:
			n_outliers = i+1
		del d[i]
	# No non-outlier was found
	return n_outliers

def tukey_outliers(x):
	d = np.array(x)
	q25, q75 = np.percentile(d, 25), np.percentile(d, 75)
	bound = 1.5 * (q75 - q25)
	lbound, ubound = q25 - bound, q75 + bound
	return len([v for v in x if v < lbound or v > ubound])



class PairPlotDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		super().__init__("Pair Plot",
				"Select two or more variables from the left to see the pair plot.  Use Ctrl-click or Shift-click to select multiple rows.",
				help_url="https://mapdata.readthedocs.io/en/latest/plot_pairplot.html")
		self.auto_update = True
		self.alpha = 0.45
		self.yrot = 90
		self.ypad = 4.0
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.categ_columns = sorted([c[0] for c in self.column_specs if c[1] in ("string")])
		self.dlg.bind("<Control-s>")
		self.dlg.bind("<Alt-y>")

		# Controls
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_redraw)
		self.autoupdate_var = add_autoupdate(self.ctrl_frame, 0, 1, self.set_autoupdate)
		self.log10_var = tk.StringVar(self.ctrl_frame, "0")
		self.log10_ck = ttk.Checkbutton(self.ctrl_frame, text="Log10 transform data", command=self.q_redraw, variable=self.log10_var,
				onvalue="1", offvalue="0")
		self.log10_ck.grid(row=0, column=2, sticky=tk.W, padx=(12,3), pady=(3,3))
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(2, weight=1)

		# The content_frame encompasses the two panes of the variable frame and the pairplot frame
		frame_panes = ttk.PanedWindow(self.content_frame, width=750, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW, padx=3, pady=3)

		# Variable frame for list of quantitative columns/variables and the 'Group by' dropdown.
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_redraw)
		# Add 'Group by' dropdown.
		self.groupby_var = tk.StringVar(var_frame, "")
		ttk.Label(var_frame, text="Group by:").grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		groupby_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.groupby_var, width=24)
		groupby_sel["values"] = [''] + self.categ_columns
		groupby_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		groupby_sel.bind('<<ComboboxSelected>>', self.q_redraw)

		# pplot frame for pair plot figure
		self.pplot_frame = tk.Frame(frame_panes, width=450, borderwidth=3, relief=tk.RIDGE)
		self.pplot_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.pplot_frame.rowconfigure(0, weight=1)
		self.pplot_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.pplot_frame, weight=12)
		self.pairplot = Plot(self.pplot_frame, 4.25, 4.25)
		# initialize content frame with an empty plot
		self.pairplot.clear()

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.data_btn = new_data_button(self.dlg, self.btn_frame, 1, self.show_data)
		new_close_button(self.dlg, self.btn_frame, 2, self.do_close)
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(1, weight=1)

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for pair plot")

	def clear_output(self):
		self.pairplot.clear()

	def q_redraw(self, get_data=True, *args):
		if self.dataset is None or get_data:
			self.clear_output()
			self.get_data()
		if self.dataset is not None and len(self.dataset) > 1 and len(self.dataset[0]) > 1:
			self.redraw()
		else:
			self.clear_output()
			self.data_btn["state"] = tk.DISABLED

	def get_data(self):
		# Get the selected data into 'dataset'
		self.dataset = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		self.n_data_columns = len(column_list)
		groupby = self.groupby_var.get()
		if groupby != '':
			column_list.append(groupby)
		if len(column_list) > 0:
			# Get either only the selected data or all data.
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				self.dataset = dataset
				self.data_labels = column_list
			if self.dataset is not None:
				clean_labels = self.data_labels
				# Remove missing data
				clean_data = clean_missing(dataset, list(range(len(dataset))))
				# If the last column is a label, ensure all values are strings.
				if self.groupby_var.get() != '':
					clean_data[self.n_data_columns] = [str(lbl) for lbl in clean_data[self.n_data_columns]]
				# Convert to floats
				for i in range(self.n_data_columns):
					clean_data[i] = [conv_float(v) for v in clean_data[i]]
				# Log-transform data if specified.
				if self.log10_var.get() == '1':
					logged_data = logdataset(clean_data, list(range(self.n_data_columns)))
					if logged_data is not None:
						clean_data = logged_data
						clean_labels = ["Log10 of " + v for v in clean_labels]
					else:
						warning_nolog(parent=self.dlg)
						self.log10_var.set("0")
				#
				self.dataset = clean_data
				self.data_labels = clean_labels
				self.data_btn["state"] = tk.NORMAL

	def redraw(self):
		# (Re)draw the pair plot.  The Seaborn pairplot function produces a figure that
		# can't be used with FigureCanvasTkAgg, so this is a custom pairplot routine.
		self.loading_dlg.display("Drawing plots")
		self.dlg.bind("<Alt-a>", self.set_alpha)
		self.dlg.bind("<Alt-y>", self.set_yrot)
		if self.data_labels is not None and self.n_data_columns > 1:
			groupby = self.groupby_var.get()
			if groupby != '':
				grplbls = [lbl for lbl in self.dataset[self.n_data_columns]]
				if wrap_at_underscores:
					grplbls = [lbl.replace("_", " ") for lbl in grplbls]
				grplbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in grplbls]
				grplbls = ["\n".join(lbl) for lbl in grplbls]
				groups = sorted_numstrs(list(set(self.dataset[self.n_data_columns])))
				cmap = map_colors(groups)
				colors = [cmap[g] for g in self.dataset[self.n_data_columns]]
			self.pairplot.figure.clear()
			dim = self.n_data_columns
			orig_xsize = matplotlib.rcParams['xtick.labelsize']
			orig_ysize = matplotlib.rcParams['ytick.labelsize']
			matplotlib.rcParams['xtick.labelsize'] = 6
			matplotlib.rcParams['ytick.labelsize'] = 6
			for rowindex in range(dim):
				for colindex in range(dim):
					subplot_no = (rowindex * dim) + (colindex+1)
					ax = self.pairplot.figure.add_subplot(dim, dim, subplot_no)
					if colindex == rowindex:
						if groupby == '':
							kdp = sns.kdeplot({self.data_labels[rowindex]: self.dataset[rowindex]}, fill=True, ax=ax)
						else:
							kdp = sns.kdeplot({self.data_labels[rowindex]: self.dataset[rowindex], 'colors': colors},
									x=self.data_labels[rowindex], hue=colors, multiple="layer", alpha=self.alpha, fill=True, ax=ax,
									warn_singular=False)
						kdp.set(xlabel=None)
						kdp.set(ylabel=None)
					else:
						if groupby == '':
							ax.scatter(self.dataset[colindex], self.dataset[rowindex], alpha=self.alpha)
						else:
							ax.scatter(self.dataset[colindex], self.dataset[rowindex], c=colors, alpha=self.alpha)
					ax.legend([]).set_visible(False)
					if colindex == 0:
						ax.set_ylabel(self.data_labels[rowindex], rotation=self.yrot, labelpad=self.ypad)
					if rowindex == dim-1:
						ax.set_xlabel(self.data_labels[colindex])
			if groupby != '':
				handle_list = []
				lbls = list(cmap.keys())
				colkey = list(cmap.values())
				symbs = [matplotlib.lines.Line2D([], [], marker='o', color=colkey[i], label=lbls[i], linestyle='None') for i in range(len(lbls))]
				self.pairplot.figure.subplots_adjust(right=0.75)
				self.pairplot.figure.legend(handles=symbs, labels=lbls, loc="outside right upper")
			self.pairplot.draw()
			matplotlib.rcParams['xtick.labelsize'] = orig_xsize
			matplotlib.rcParams['ytick.labelsize'] = orig_ysize
		self.loading_dlg.hide()

	def set_autoupdate(self):
		if self.autoupdate_var.get() == "1":
			self.auto_update = True
		else:
			self.auto_update = False

	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the opacity (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = min(1.0, max(new_alpha, 0.0))
			self.q_redraw()
	def set_yrot(self, *args):
		if self.yrot == 0:
			self.yrot = 90
			self.ypad = 4.0
		else:
			self.yrot = 0
			if self.log10_var.get() == "1":
				self.ypad = 50.0
			else:
				self.ypad = 30.0
		self.q_redraw()
	def do_close(self, *args):
		self.parent.remove_plot(self)
		super().do_cancel(args)



class VariableDistDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.auto_update = True
		self.alpha = 0.45
		self.plot_title = None
		self.xlabel = None
		self.ylabel = None
		super().__init__("Variable Distribution Plot", "Select one or more variables from the list below.",
			help_url="https://mapdata.readthedocs.io/en/latest/plot_vardist.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		self.alpha = 0.45
		self.rotated = False
		self.xlabel = None
		self.ylabel = None
		self.plot_title = None
		# Data
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None

		# Controls
		ckb_frame = tk.Frame(self.ctrl_frame)
		ckb_frame.grid(row=0, column=0, sticky = tk.NSEW)
		self.sel_only_var, self.sel_only_ck = add_sel_only(ckb_frame, 0, 0, self.q_redraw)
		self.autoupdate_var = add_autoupdate(ckb_frame, 0, 1, self.set_autoupdate)
		self.ylog_var = tk.StringVar(ckb_frame, "0")
		self.ylog_ck = ttk.Checkbutton(ckb_frame, text="Log values", state=tk.NORMAL, command=self.q_redraw, variable=self.ylog_var,
				onvalue="1", offvalue="0")
		self.ylog_ck.grid(row=0, column=3, sticky=tk.W, padx=(6,6), pady=(3,3))

		pt_frame = tk.Frame(self.ctrl_frame)
		pt_frame.grid(row=1, column=0, sticky=tk.NSEW)
		self.type_var = tk.StringVar(pt_frame, "Box plot")
		type_lbl = ttk.Label(pt_frame, text="Plot type:")
		type_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.type_sel = ttk.Combobox(pt_frame, state="readonly", textvariable=self.type_var, width=25, height=15,
				values=["Box plot", "Kernel density (KD) plot", "Stripchart", "Violin plot"])
		self.type_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.type_sel.bind("<<ComboboxSelected>>", self.pt_changed)

		# PanedWindow within content_frame: left is input listbox and other data selection controls, right is plot
		io_panes = ttk.PanedWindow(self.content_frame, width=700, orient=tk.HORIZONTAL)
		io_panes.grid(row=0, column=0, sticky=tk.NSEW, padx=3, pady=3)

		self.inp_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE)
		self.inp_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.inp_frame.rowconfigure(0, weight=1)
		self.inp_frame.columnconfigure(0, weight=1)
		io_panes.add(self.inp_frame, weight=1)

		self.output_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE)
		self.output_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.output_frame.rowconfigure(0, weight=1)
		self.output_frame.columnconfigure(0, weight=1)
		io_panes.add(self.output_frame, weight=12)

		# Variable selection
		# Add multi-select list of numeric variables to the leftmost pane
		var_frame = tk.Frame(self.inp_frame, width=250)
		var_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_redraw)

		# Output plot
		plot_frame = tk.Frame(self.output_frame)
		plot_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.varplot = Plot(plot_frame, 3.5, 3.5)

		# Buttons
		self.data_btn, self.plot_data_btn = add_help_src_plot_close_btns(self.dlg, self.btn_frame,
				self.do_help, self.show_data, self.show_plot_data, self.do_close)

		# Initialize output frames
		self.clear_output()

	def set_autoupdate(self):
		if self.autoupdate_var.get() == "1":
			self.auto_update = True
		else:
			self.auto_update = False

	def pt_changed(self, *args):
		self.xlabel = None
		self.ylabel = None
		self.q_redraw([])

	def q_redraw(self, *args):
		# At least two variables must be selected
		self.clear_output()
		curr_selections = self.column_table.selection()
		if len(curr_selections) > 0:
			self.get_data()
			if self.dataset is not None:
				self.redraw()

	def get_data(self):
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED
		self.ylog_ck["state"] = tk.NORMAL
		self.dataset = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		# Record the number of data columns without labels.
		self.n_dataset_columns = len(column_list)
		# Get the selected data.
		dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is None:
			self.dataset = None
			self.data_labels = None
		else:
			# Remove missing values by column
			dataset = clean_missing_bycol(dataset, list(range(self.n_dataset_columns)))
			# Convert to floats for numeric data columns only
			for i in range(self.n_dataset_columns):
				if column_list[i] in self.numeric_columns:
					dataset[i] = [conv_float(v) for v in dataset[i]]
			self.dataset = dataset
			self.data_labels = column_list
			self.data_btn["state"] = tk.NORMAL
			self.plot_data = self.dataset
			self.plot_data_labels = self.data_labels
			self.plot_data_btn["state"] = tk.NORMAL
			# Maybe log-transform every column
			if self.ylog_var.get() == '1' and self.ylog_ck["state"] != tk.DISABLED:
				logged_data = logdataset(dataset, list(range(len(dataset))))
				if logged_data is not None:
					self.plot_data = logged_data
					self.plot_data_labels = ["Log10 of " + c for c in self.data_labels]
				else:
					self.ylog_var.set("0")
					self.ylog_ck["state"] = tk.DISABLED
					warning_nolog(parent=self.dlg)
			self.dlg.bind("<Alt-x>", self.set_xlabel)
			self.dlg.bind("<Alt-y>", self.set_ylabel)

	def clear_output(self):
		clear_dlg_hotkeys(self.dlg)
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED
		self.varplot.clear()

	def redraw(self):
		plot_type = self.type_var.get()
		grplbls = copy.copy(self.data_labels)
		if wrap_at_underscores:
			grplbls = [lbl.replace("_", " ") for lbl in grplbls]
		grplbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in grplbls]
		grplbls = ["\n".join(lbl) for lbl in grplbls]
		if self.ylog_var.get() == '1' and self.ylog_ck["state"] != tk.DISABLED:
			value_lbl = "Log10 of value"
		else:
			value_lbl = "Value"
		if plot_type != "Box_plot":
			plt_data = {}
			for i in range(len(grplbls)):
				plt_data[grplbls[i]] = self.plot_data[i] 
		if plot_type == "Box plot":
			self.dlg.bind("<Alt-r>", self.set_rotated)
			self.varplot.axes.boxplot(self.plot_data, labels=grplbls, vert=not self.rotated)
			if not self.rotated:
				self.varplot.axes.set_ylabel(self.ylabel or value_lbl)
			else:
				self.varplot.axes.set_xlabel(self.ylabel or value_lbl)
		elif plot_type == "Kernel density (KD) plot":
			self.dlg.bind("<Alt-a>", self.set_alpha)
			sns.kdeplot(data=plt_data, multiple="layer", fill=True, alpha=self.alpha, ax=self.varplot.axes,
					warn_singular=False)
			self.varplot.axes.set_ylabel(self.ylabel or "Density")
		elif plot_type == "Stripchart":
			self.dlg.bind("<Alt-a>", self.set_alpha)
			self.dlg.bind("<Alt-r>", self.set_rotated)
			orientation = "v" if not self.rotated else "h"
			sns.stripplot(data=plt_data, alpha=self.alpha, ax=self.varplot.axes, orient=orientation)
			if not self.rotated:
				self.varplot.axes.set_ylabel(self.ylabel or value_lbl)
			else:
				self.varplot.axes.set_xlabel(self.ylabel or value_lbl)
		else:
			# Violin plot
			self.dlg.bind("<Alt-r>", self.set_rotated)
			orientation = True if not self.rotated else False
			try:
				sns.violinplot(data=plt_data, vert=orientation, ax=self.varplot.axes)
				if not self.rotated:
					self.varplot.axes.set_ylabel(self.ylabel or value_lbl)
				else:
					self.varplot.axes.set_xlabel(self.ylabel or value_lbl)
			except:
				sns.violinplot(data=plt_data, ax=self.varplot.axes)
				self.varplot.axes.set_ylabel(self.ylabel or value_lbl)
		if self.xlabel is not None:
			if self.rotated:
				self.varplot.axes.set_ylabel(self.xlabel)
			else:
				self.varplot.axes.set_xlabel(self.xlabel)
		if self.plot_title is not None:
			self.varplot.axes.set_title(self.plot_title)
			self.varplot.axes.title.set_visible(True)
		else:
			self.varplot.axes.title.set_visible(False)
		self.dlg.bind("<Alt-t>", self.set_title)
		self.varplot.draw()

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels, "Data for Variable Distribution Plot")
			self.dlg.lift()

	def show_plot_data(self, *args):
		# Show data as summarized for a particular plot type.
		if self.plot_data is not None:
			show_columnar_table(self.dlg, "Data for Plotting", "Data to be plotted:", self.plot_data, \
					self.plot_data_labels, "Plot data")
			self.dlg.lift()
	
	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the transparency (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = min(1.0, max(new_alpha, 0.0))
			self.q_redraw()
	def set_title(self, *args):
		dlg = OneEntryDialog(self.dlg, "Plot Title", "Enter a title for the plot:", init_value=self.plot_title, nullable=True)
		self.plot_title = dlg.show()
		self.varplot.axes.set_title(self.plot_title)
		self.varplot.axes.title.set_visible(self.plot_title is not None)
		self.varplot.draw()
	def set_xlabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "X-Axis Label", "Enter a label for the X-axis:", init_value=self.xlabel, nullable=True)
		self.xlabel = dlg.show()
		self.varplot.axes.set_xlabel(self.xlabel)
		self.varplot.canvas.draw()
	def set_ylabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "Y-Axis Label", "Enter a label for the Y-axis label:", init_value=self.ylabel, nullable=True)
		self.ylabel = dlg.show() or "Value"
		self.varplot.axes.set_ylabel(self.ylabel)
		self.varplot.draw()
	def set_rotated(self, *args):
		self.rotated = not self.rotated
		self.q_redraw()

	def do_close(self, *args):
		self.parent.remove_plot(self)
		super().do_cancel(args)



class MultivarBarPlotDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		super().__init__("Muli-variable Grouped Bar Chart",
				"Select variables from the list at the left, and then a grouping variable and aggregation method.  Use Ctrl-click or Shift-click to select multiple rows.",
				help_url="https://mapdata.readthedocs.io/en/latest/multivar_bar_plot.html")
		self.auto_update = True
		self.plot_by_variable = True	# A separate bar plot for each variable; alternative: a separate bar plot for each group.
		self.plot_title = None
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.categ_columns = sorted([c[0] for c in self.column_specs if c[1] in ("string")])
		self.dlg.bind("<Control-s>")
		self.dlg.bind("<Alt-r>")
		self.dlg.bind("<Alt-t>")

		# Controls
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_redraw)
		self.autoupdate_var = add_autoupdate(self.ctrl_frame, 0, 1, self.set_autoupdate)
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(1, weight=1)

		# The content_frame encompasses the two panes of the variable frame and the plot frame
		frame_panes = ttk.PanedWindow(self.content_frame, width=750, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW, padx=3, pady=3)

		# Variable frame for list of quantitative columns/variables and the 'Group by' and 'Aggregate' dropdowns.
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_redraw)
		# Add 'Group by' dropdown.
		self.groupby_var = tk.StringVar(var_frame, "")
		ttk.Label(var_frame, text="Group by:").grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		groupby_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.groupby_var, width=24)
		groupby_sel["values"] = [''] + self.categ_columns
		groupby_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		groupby_sel.bind('<<ComboboxSelected>>', self.q_redraw)
		# Add aggregation method
		self.aggreg_var = tk.StringVar(var_frame, "Arithmetic mean")
		ttk.Label(var_frame, text="Aggregate:").grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		aggreg_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.aggreg_var, width=24)
		aggreg_sel["values"] = ["Arithmetic mean", "Geometric mean", "Harmonic mean", "Minimum", "Maximum", "Median", "Count", "Sum"]
		aggreg_sel.grid(row=2, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		aggreg_sel.bind('<<ComboboxSelected>>', self.q_redraw)

		# bplot frame for plot figure
		self.bplot_frame = tk.Frame(frame_panes, width=450, borderwidth=3, relief=tk.RIDGE)
		self.bplot_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.bplot_frame.rowconfigure(0, weight=1)
		self.bplot_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.bplot_frame, weight=12)
		self.barplot = Plot(self.bplot_frame, 4.25, 4.25)
		# initialize content frame with an empty plot
		self.barplot.clear()

		# Buttons
		self.data_btn, self.plot_data_btn = add_help_src_plot_close_btns(self.dlg, self.btn_frame,
				self.do_help, self.show_data, self.show_plot_data, self.do_close)

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for multi-variable bar plot")

	def clear_output(self):
		self.barplot.clear()
		self.dlg.bind("<Alt-r>")
		self.dlg.bind("<Alt-t>")
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED

	def q_redraw(self, get_data=True, *args):
		if len(self.column_table.selection()) > 0 and self.groupby_var.get() != '':
			if self.dataset is None or self.agg_data is None or get_data:
				self.clear_output()
				self.get_data()
			if self.dataset is not None and len(self.dataset) > 1 and len(self.dataset[0]) > 1 and self.agg_data is not None:
				self.redraw()
			else:
				self.clear_output()
		else:
			self.clear_output()

	def get_data(self):
		self.dataset = None
		self.agg_data = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		self.n_data_columns = len(column_list)
		groupby = self.groupby_var.get()
		column_list.append(groupby)
		if len(column_list) > 0:
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				self.dataset = dataset
				self.data_labels = column_list
			if self.dataset is not None:
				self.data_btn["state"] = tk.NORMAL
				# Convert the grouping variable to strings
				self.dataset[self.n_data_columns] = [str(lbl) for lbl in self.dataset[self.n_data_columns]]
				# Convert variable values to floats
				for i in range(self.n_data_columns):
					self.dataset[i] = [conv_float(v) for v in self.dataset[i]]
				# Aggregate rows by group for each variable
				subsets = subset_by_groups(self.dataset[0:self.n_data_columns], self.dataset[self.n_data_columns])
				agg_error, agg_data = aggregate_groups(subsets, self.aggreg_var.get())
				if not agg_error:
					self.agg_data = agg_data
					self.agg_data_labels = [self.data_labels[self.n_data_columns]] + self.data_labels[0:self.n_data_columns]
					self.plot_data_btn["state"] = tk.NORMAL

	def redraw(self):
		# (Re)draw the multi-variable stacked bar plot.
		self.loading_dlg.display("Drawing plots")
		#
		self.barplot.figure.clear()
		if self.plot_by_variable:
			nplots = len(self.agg_data)-1
			bpy = self.data_labels[:self.n_data_columns]
			bpx = self.agg_data[0]
			if nplots > 1:
				self.barplot.axes = axs = self.barplot.figure.subplots(nplots, sharex=True, sharey=True)
				for bp_idx in range(nplots):
					axs[bp_idx].bar(bpx, self.agg_data[bp_idx+1])
					axs[bp_idx].set_ylabel(bpy[bp_idx])
					axs[bp_idx].label_outer()
			else:
				self.barplot.clear()
				self.barplot.axes.bar(self.agg_data[0], self.agg_data[1])
				self.barplot.axes.set_ylabel(bpy[0])
		else:
			pdata = columns_to_rows(self.agg_data)
			nplots = len(pdata)
			if nplots > 1:
				self.barplot.axes = axs = self.barplot.figure.subplots(nplots, sharex=True, sharey=True)
				bpy = [r[0] for r in pdata]
				bpx = self.data_labels[:self.n_data_columns]
				for bp_idx in range(nplots):
					axs[bp_idx].bar(bpx, pdata[bp_idx][1:])
					axs[bp_idx].set_ylabel(bpy[bp_idx])
					axs[bp_idx].label_outer()
			else:
				self.barplot.clear()
				self.barplot.axes.bar(self.data_labels[:self.n_data_columns], pdata[0][1:])
				self.barplot.axes.set_ylabel(pdata[0][0])
		self.barplot.draw()
		self.dlg.bind("<Alt-r>", self.set_rotated)
		self.dlg.bind("<Alt-t>", self.set_title)
		#
		self.loading_dlg.hide()

	def set_autoupdate(self):
		if self.autoupdate_var.get() == "1":
			self.auto_update = True
		else:
			self.auto_update = False

	def show_plot_data(self, *args):
		# Show data as summarized for a particular plot type.
		if self.agg_data is not None:
			show_columnar_table(self.dlg, "Data for Bar Charts", "Data to be plotted:", self.agg_data, \
					self.agg_data_labels, "Plot data")
			self.dlg.lift()

	def set_title(self, *args):
		dlg = OneEntryDialog(self.dlg, "Plot Title", "Enter a title for the plot:", init_value=self.plot_title, nullable=True)
		self.plot_title = dlg.show()
		self.barplot.axes[0].set_title(self.plot_title)
		self.barplot.axes[0].title.set_visible(self.plot_title is not None)
		self.barplot.draw()

	def set_rotated(self, *args):
		self.plot_by_variable = not self.plot_by_variable
		self.q_redraw(False, [])

	def do_close(self, *args):
		self.parent.remove_plot(self)
		super().do_cancel(args)



class MultivarGrpPctBarDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		super().__init__("Muli-variable Grouped Percentage Bar Chart",
				"Select variables from the list at the left, and then a grouping variable and aggregation method.  Use Ctrl-click or Shift-click to select multiple rows.",
				help_url="https://mapdata.readthedocs.io/en/latest/grp_pct_bar_chart.html")
		self.plot_by_variable = True	# A separate bar plot for each variable; alternative: a separate bar plot for each group.
		self.auto_update = True
		self.xlabel = None
		self.plot_title = None
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.categ_columns = sorted([c[0] for c in self.column_specs if c[1] in ("string")])
		self.flip_y = False
		self.dlg.bind("<Control-s>")
		self.dlg.bind("<Alt-r>")
		self.dlg.bind("<Alt-t>")

		# Controls
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_redraw)
		self.autoupdate_var = add_autoupdate(self.ctrl_frame, 0, 1, self.set_autoupdate)
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(1, weight=1)

		# The content_frame encompasses the two panes of the variable frame and the plot frame
		frame_panes = ttk.PanedWindow(self.content_frame, width=750, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW, padx=3, pady=3)

		# Variable frame for list of quantitative columns/variables and the 'Group by' and 'Aggregate' dropdowns.
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_redraw)
		# Add 'Group by' dropdown.
		self.groupby_var = tk.StringVar(var_frame, "")
		ttk.Label(var_frame, text="Group by:").grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		groupby_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.groupby_var, width=24)
		groupby_sel["values"] = [''] + self.categ_columns
		groupby_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		groupby_sel.bind('<<ComboboxSelected>>', self.q_redraw)
		# Add aggregation method
		self.aggreg_var = tk.StringVar(var_frame, "Arithmetic mean")
		ttk.Label(var_frame, text="Aggregate:").grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		aggreg_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.aggreg_var, width=24)
		aggreg_sel["values"] = ["Arithmetic mean", "Geometric mean", "Harmonic mean", "Minimum", "Maximum", "Median", "Count", "Sum"]
		aggreg_sel.grid(row=2, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		aggreg_sel.bind('<<ComboboxSelected>>', self.q_redraw)

		# bplot frame for plot figure
		self.bplot_frame = tk.Frame(frame_panes, width=450, borderwidth=3, relief=tk.RIDGE)
		self.bplot_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.bplot_frame.rowconfigure(0, weight=1)
		self.bplot_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.bplot_frame, weight=12)
		self.barplot = Plot(self.bplot_frame, 4.25, 6.25)
		# initialize content frame with an empty plot
		self.barplot.clear()

		# Buttons
		self.data_btn, self.plot_data_btn = add_help_src_plot_close_btns(self.dlg, self.btn_frame,
				self.do_help, self.show_data, self.show_plot_data, self.do_close)

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for multi-variable grouped percentage chart")

	def clear_output(self):
		self.barplot.clear()
		clear_dlg_hotkeys(self.dlg)
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED

	def q_redraw(self, get_data=True, *args):
		if len(self.column_table.selection()) > 0 and self.groupby_var.get() != '':
			if self.dataset is None or self.agg_data is None or get_data:
				self.clear_output()
				self.get_data()
			if self.dataset is not None and len(self.dataset) > 1 and len(self.dataset[0]) > 1 and self.agg_data is not None:
				self.redraw()
			else:
				self.clear_output()
		else:
			self.clear_output()

	def get_data(self):
		self.dataset = None
		self.agg_data = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		self.n_data_columns = len(column_list)
		groupby = self.groupby_var.get()
		column_list.append(groupby)
		if len(column_list) > 0:
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				self.dataset = dataset
				self.data_labels = column_list
			if self.dataset is not None:
				self.data_btn["state"] = tk.NORMAL
				# Convert the grouping variable to strings
				self.dataset[self.n_data_columns] = [str(lbl) for lbl in self.dataset[self.n_data_columns]]
				# Convert variable values to floats
				for i in range(self.n_data_columns):
					self.dataset[i] = [conv_float(v) for v in self.dataset[i]]
				# Aggregate rows by group for each variable
				subsets = subset_by_groups(self.dataset[0:self.n_data_columns], self.dataset[self.n_data_columns])
				agg_error, agg_data = aggregate_groups(subsets, self.aggreg_var.get())
				if not agg_error:
					# Change nan to 0.0 in aggregated data
					for varno in range(1, len(agg_data)):
						agg_data[varno] = [v if not math.isnan(v) else 0.0 for v in agg_data[varno]]
					grp_agg = columns_to_rows(agg_data)
					# Convert aggregates to percentages by variable
					for varno in range(1, len(agg_data)):
						tot = sum(agg_data[varno])
						if tot > 0:
							agg_data[varno] = [100 * v /tot for v in agg_data[varno]]
						else:
							agg_data[varno] = [0.0 for _ in agg_data[varno]]
					self.agg_data = agg_data
					self.agg_data_labels = [self.data_labels[self.n_data_columns]] + self.data_labels[0:self.n_data_columns]
					self.plot_data_btn["state"] = tk.NORMAL
					# Convert aggregates to percentages by group, for rotated plot
					for grpno in range(len(grp_agg)):
						tot = sum(grp_agg[grpno][1:])
						grp_agg[grpno] = [grp_agg[grpno][0]] + [100 * v/tot for v in grp_agg[grpno][1:]]
					self.grp_agg_data = columns_to_rows(grp_agg)

	def redraw(self):
		# (Re)draw the multi-variable stacked bar plot.
		self.loading_dlg.display("Drawing plots")
		#
		self.barplot.clear()
		if self.plot_by_variable:
			plot_data = columns_to_rows(self.agg_data[1:])
			labels = self.agg_data[0]
			variable_names = self.agg_data_labels[1:]
			barbase = [0]*len(variable_names)
			ext_colors = quant_colors * math.ceil(len(labels)/len(quant_colors))
			for i in range(len(plot_data)):
				self.barplot.axes.barh(variable_names, plot_data[i], left=barbase, color=ext_colors[i], label=labels[i], height=0.9)
				barbase = list(map(add, barbase, plot_data[i]))
		else:
			plot_data = self.grp_agg_data[1:]
			labels = self.agg_data_labels[1:]
			grp_names = self.grp_agg_data[0]
			barbase = [0]*len(grp_names)
			ext_colors = quant_colors * math.ceil(len(labels)/len(quant_colors))
			for i in range(len(plot_data)):
				self.barplot.axes.barh(grp_names, plot_data[i], left=barbase, color=ext_colors[i], label=labels[i], height=0.9)
				barbase = list(map(add, barbase, plot_data[i]))
		if self.flip_y:
			self.barplot.axes.invert_yaxis()
		self.barplot.axes.set_xlabel(self.xlabel or "Percent")
		self.barplot.axes.legend()
		self.barplot.figure.subplots_adjust(right=0.65)
		self.barplot.axes.legend(loc="upper left", bbox_to_anchor=(1.05, 1.0))
		self.barplot.draw()
		self.dlg.bind("<Alt-x>", self.set_xlabel)
		self.dlg.bind("<Alt-t>", self.set_title)
		self.dlg.bind("<Alt-r>", self.set_rotated)
		self.dlg.bind("<Alt-f>", self.set_flip_y)
		#
		self.loading_dlg.hide()

	def set_autoupdate(self):
		if self.autoupdate_var.get() == "1":
			self.auto_update = True
		else:
			self.auto_update = False

	def show_plot_data(self, *args):
		# Show data as summarized for a particular plot type.
		if self.agg_data is not None:
			show_columnar_table(self.dlg, "Data for Bar Charts", "Data to be plotted:", self.agg_data, \
					self.agg_data_labels, "Plot data")
			self.dlg.lift()

	def set_xlabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "X-Axis Label", "Enter a label for the X-axis:", init_value=self.xlabel, nullable=True)
		self.xlabel = dlg.show()
		self.barplot.axes.set_xlabel(self.xlabel)
		self.barplot.canvas.draw()
	def set_title(self, *args):
		dlg = OneEntryDialog(self.dlg, "Plot Title", "Enter a title for the plot:", init_value=self.plot_title, nullable=True)
		self.plot_title = dlg.show()
		self.barplot.axes.set_title(self.plot_title)
		self.barplot.axes.title.set_visible(self.plot_title is not None)
		self.barplot.draw()

	def set_flip_y(self, *args):
			self.flip_y = not self.flip_y
			self.q_redraw()

	def set_rotated(self, *args):
		self.plot_by_variable = not self.plot_by_variable
		self.q_redraw(False, [])

	def do_close(self, *args):
		self.parent.remove_plot(self)
		super().do_cancel(args)



class PctStackedAreaDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.all_columns = sorted([c[0] for c in self.column_specs])
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.categ_columns = sorted([c[0] for c in self.column_specs if c[1] in ("string", "date", "timestamp", "timestamptz")])
		self.auto_update = True
		self.alpha = 0.45
		self.plot_title = None
		self.xlabel = None
		self.ylabel = None
		super().__init__("Percentage Stacked Area Plot", "Select an X variable and two or more Y variables from the list below.",
				help_url="https://mapdata.readthedocs.io/en/latest/pctstackedarea.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		self.alpha = 0.45
		self.rotated = False
		self.xlabel = None
		self.ylabel = None
		self.plot_title = None
		# Data
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None

		# Controls
		ckb_frame = tk.Frame(self.ctrl_frame)
		ckb_frame.grid(row=0, column=0, sticky = tk.NSEW)
		self.sel_only_var, self.sel_only_ck = add_sel_only(ckb_frame, 0, 0, self.q_redraw)
		self.autoupdate_var = add_autoupdate(ckb_frame, 0, 1, self.set_autoupdate)

		pt_frame = tk.Frame(self.ctrl_frame)
		pt_frame.grid(row=1, column=0, sticky=tk.NSEW)
		self.x_var = tk.StringVar(pt_frame, "")
		x_lbl = ttk.Label(pt_frame, text="X variable:")
		x_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.x_sel = ttk.Combobox(pt_frame, state="readonly", textvariable=self.x_var, width=25, height=15, values=self.all_columns)
		self.x_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.x_sel.bind("<<ComboboxSelected>>", self.q_redraw)

		# PanedWindow within content_frame: left is input listbox and other data selection controls, right is plot
		io_panes = ttk.PanedWindow(self.content_frame, width=700, orient=tk.HORIZONTAL)
		io_panes.grid(row=0, column=0, sticky=tk.NSEW, padx=3, pady=3)

		self.inp_frame = tk.Frame(io_panes, width=250, borderwidth=3, relief=tk.RIDGE)
		self.inp_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.inp_frame.rowconfigure(0, weight=1)
		self.inp_frame.columnconfigure(0, weight=1)
		io_panes.add(self.inp_frame, weight=1)

		self.output_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE)
		self.output_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.output_frame.rowconfigure(0, weight=1)
		self.output_frame.columnconfigure(0, weight=1)
		io_panes.add(self.output_frame, weight=12)

		# Y variable selection
		# Add multi-select list of numeric variables to the leftmost pane
		var_frame = tk.Frame(self.inp_frame)
		var_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_redraw)

		# Aggregation method for Y values with the same X value
		self.aggreg_var = tk.StringVar(var_frame, "Arithmetic mean")
		ttk.Label(var_frame, text="Aggregate Y values by:").grid(row=1, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		aggreg_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.aggreg_var, width=24)
		aggreg_sel["values"] = ["Arithmetic mean", "Geometric mean", "Harmonic mean", "Median", "Sum"]
		aggreg_sel.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		aggreg_sel.bind('<<ComboboxSelected>>', self.q_redraw)

		# Output plot
		plot_frame = tk.Frame(self.output_frame)
		plot_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot = Plot(plot_frame, 3.5, 5.8)

		# Buttons
		self.data_btn, self.plot_data_btn = add_help_src_plot_close_btns(self.dlg, self.btn_frame,
				self.do_help, self.show_data, self.show_plot_data, self.do_close)

		# Initialize output frames
		self.clear_output()

	def set_autoupdate(self):
		if self.autoupdate_var.get() == "1":
			self.auto_update = True
		else:
			self.auto_update = False

	def q_redraw(self, *args):
		# At least two variables must be selected
		self.clear_output()
		self.dlg.bind("<Alt-x>")
		self.dlg.bind("<Alt-y>")
		self.dlg.bind("<Alt-t>")
		curr_selections = self.column_table.selection()
		if len(curr_selections) > 1 and self.x_var.get() != '' and self.aggreg_var.get() != '':
			self.get_data()
			if self.plot_data is not None:
				self.redraw()

	def get_data(self):
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None
		column_list = [self.x_var.get()]
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		# Record the number of data columns without labels.
		self.n_dataset_columns = len(column_list)
		# Get the selected data.
		dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is not None:
			# Remove missing values by case
			dataset = clean_missing(dataset, list(range(self.n_dataset_columns)))
			if len(dataset[0]) > 1:
				# Convert the X variable to strings if it is not a date/time, else to date or timestamp if possible
				x_type = [c[1] for c in self.column_specs if c[0] == self.x_var.get()][0]
				dataset[0] = conv_vect_dt(dataset[0], x_type)
				# Convert to floats for numeric data columns only
				for i in range(1,self.n_dataset_columns):
					dataset[i] = [conv_float(v) for v in dataset[i]]
				self.dataset = dataset
				self.data_labels = column_list
				self.data_btn["state"] = tk.NORMAL
				# Aggregate rows by X variable
				subsets = subset_by_groups(dataset[1:self.n_dataset_columns], dataset[0])
				agg_error, agg_data = aggregate_groups(subsets, self.aggreg_var.get())
				if not agg_error:
					# Check for NaN
					if dataset_contains(agg_data, lambda x: type(x) == float and math.isnan(x)):
						warning("Aggregated data has NaN values; plot cannot be produced.", {"parent": self.dlg})
					else:
						# Reject negative values
						if dataset_contains(agg_data, lambda x: type(x) == float and x < 0.0):
							warning("Aggregated data has negative values; plot cannot be produced.", {"parent": self.dlg})
						else:
							self.plot_data = sort_columns(agg_data)
							self.plot_data_labels = self.data_labels
							self.plot_data_btn["state"] = tk.NORMAL

	def clear_output(self):
		clear_dlg_hotkeys(self.dlg)
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED
		self.plot.clear()
		self.dlg.bind("<Alt-x>")
		self.dlg.bind("<Alt-y>")
		self.dlg.bind("<Alt-t>")

	def redraw(self):
		y_lbls = copy.copy(self.data_labels[1:])
		if wrap_at_underscores:
			y_lbls = [lbl.replace("_", " ") for lbl in y_lbls]
		y_lbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in y_lbls]
		y_lbls = ["\n".join(lbl) for lbl in y_lbls]
		# Row-sum standardize all the Y values
		yrows = columns_to_rows(self.plot_data[1:])
		ysums = [sum(row) for row in yrows]
		n_yvars = self.n_dataset_columns - 1
		yprop = []
		for i in range(len(yrows)):
			yprop.append([100.0*yrows[i][v]/ysums[i] for v in range(n_yvars)])
		custpal = [quant_colors[0]] + quant_colors[2:]
		# Draw the stacked area chart
		self.plot.axes.stackplot(np.array(self.plot_data[0]), np.array(rows_to_columns(yprop)), labels=y_lbls, colors=custpal)
		# Add the line for totals.
		ax2 = self.plot.axes.twinx()
		max_y = max(ysums)
		min_y = min(ysums)
		diff_y = max_y - min_y
		if diff_y < 0.2 * max_y:
			min_y = min_y - 0.1 * diff_y
			max_y = max_y + 0.05 * diff_y
		else:
			max_y = 1.05 * max_y
		ax2.set_ylim(min_y, max_y)
		#ax2.set_ylim(0, 1.05*max_y)
		ax2.plot(self.plot_data[0], ysums, color=quant_colors[1], marker=".", label="Total")
		ax2.set_ylabel("Total")

		self.plot.figure.subplots_adjust(right=0.65)
		self.plot.axes.legend(loc="center left", bbox_to_anchor=(1.2, 0.5))

		self.plot.set_axis_labels(self.xlabel or self.x_var.get(), self.ylabel or "Percent of total")
		if self.plot_title is not None:
			self.plot.axes.set_title(self.plot_title)
			self.plot.axes.title.set_visible(True)
		else:
			self.plot.axes.title.set_visible(False)
		self.dlg.bind("<Alt-x>", self.set_xlabel)
		self.dlg.bind("<Alt-y>", self.set_ylabel)
		self.dlg.bind("<Alt-t>", self.set_title)
		self.plot.draw()

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels, "Data for Percent Stacked Plot")
			self.dlg.lift()

	def show_plot_data(self, *args):
		# Show data as summarized for a particular plot type.
		if self.plot_data is not None:
			show_columnar_table(self.dlg, "Data for Plotting", "Data to be plotted:", self.plot_data, \
					self.plot_data_labels, "Plot data")
			self.dlg.lift()
	
	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the transparency (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = min(1.0, max(new_alpha, 0.0))
			self.q_redraw()
	def set_title(self, *args):
		dlg = OneEntryDialog(self.dlg, "Plot Title", "Enter a title for the plot:", init_value=self.plot_title, nullable=True)
		self.plot_title = dlg.show()
		self.plot.axes.set_title(self.plot_title)
		self.plot.axes.title.set_visible(self.plot_title is not None)
		self.plot.draw()
	def set_xlabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "X-Axis Label", "Enter a label for the X-axis:", init_value=self.xlabel, nullable=True)
		self.xlabel = dlg.show()
		self.plot.axes.set_xlabel(self.xlabel or self.x_var.get())
		self.plot.draw()
	def set_ylabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "Y-Axis Label", "Enter a label for the Y-axis label:")
		self.ylabel = dlg.show()
		self.plot.axes.set_ylabel(self.ylabel or "Percent of total")
		self.plot.draw()

	def do_close(self, *args):
		self.parent.remove_plot(self)
		super().do_cancel(args)



class MultiVariableLinePlotDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.all_columns = sorted([c[0] for c in self.column_specs])
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.yvar_list = []			# Updated after an X variable is selected: numeric but not X.
		self.x_is_numeric = True	# True to force initial Y list population.
		self.auto_update = True
		self.alpha = 0.45
		self.plot_title = None
		self.xlabel = None
		self.ylabel = None
		super().__init__("Multi-Variable Line or Scatter Plot", "Select an X variable and then one or more Y variables from the list below.",
				help_url="https://mapdata.readthedocs.io/en/latest/multivarlineplot.html")
		# Data
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None

		# Controls
		ckb_frame = tk.Frame(self.ctrl_frame)
		ckb_frame.grid(row=0, column=0, sticky = tk.NSEW)
		self.sel_only_var, self.sel_only_ck = add_sel_only(ckb_frame, 0, 0, self.q_redraw)
		self.autoupdate_var = add_autoupdate(ckb_frame, 0, 1, self.set_autoupdate)
		pt_frame = tk.Frame(self.ctrl_frame)
		pt_frame.grid(row=1, column=0, sticky=tk.NSEW)

		# X variable
		self.x_var = tk.StringVar(pt_frame, "")
		x_lbl = ttk.Label(pt_frame, text="X column:")
		x_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.x_sel = ttk.Combobox(pt_frame, state="readonly", textvariable=self.x_var, width=25, height=15, values=self.all_columns)
		self.x_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.x_sel.bind("<<ComboboxSelected>>", self.x_selected)

		# Scatter plot?
		self.scatter_var = tk.StringVar(pt_frame, "0")
		self.scatter_ck = ttk.Checkbutton(pt_frame, text="Scatter plot", variable=self.scatter_var, command=self.q_redraw,
				onvalue="1", offvalue="0")
		self.scatter_ck.grid(row=0, column=2, sticky=tk.W, padx=(12,3), pady=(3,3))
		pt_frame.columnconfigure(2, weight=1)

		# PanedWindow within content_frame: left is input listbox and other data selection controls, right is plot
		io_panes = ttk.PanedWindow(self.content_frame, width=700, orient=tk.HORIZONTAL)
		io_panes.grid(row=0, column=0, sticky=tk.NSEW, padx=3, pady=3)

		self.inp_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE)
		self.inp_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.inp_frame.rowconfigure(0, weight=1)
		self.inp_frame.columnconfigure(0, weight=1)
		io_panes.add(self.inp_frame, weight=1)

		self.output_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE)
		self.output_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.output_frame.rowconfigure(0, weight=1)
		self.output_frame.columnconfigure(0, weight=1)
		io_panes.add(self.output_frame, weight=12)

		# Y variable selection
		# Add multi-select list of numeric variables to the leftmost pane
		self.var_frame = tk.Frame(self.inp_frame)
		self.var_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW)
		self.var_frame.rowconfigure(0, weight=1)
		self.var_frame.columnconfigure(0, weight=1)

		# Aggregation method for Y values with the same X value
		self.aggreg_var = tk.StringVar(self.var_frame, "Arithmetic mean")
		ttk.Label(self.var_frame, text="Aggregate Y values by:").grid(row=1, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		aggreg_sel = ttk.Combobox(self.var_frame, state="readonly", textvariable=self.aggreg_var, width=24)
		aggreg_sel["values"] = ["Arithmetic mean", "Geometric mean", "Harmonic mean", "Minimum", "Maximum", "Sum"]
		aggreg_sel.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		aggreg_sel.bind('<<ComboboxSelected>>', self.q_redraw)

		# Output plot
		plot_frame = tk.Frame(self.output_frame)
		plot_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot = Plot(plot_frame, 3.5, 3.5)

		# Buttons
		self.data_btn, self.plot_data_btn = add_help_src_plot_close_btns(self.dlg, self.btn_frame,
				self.do_help, self.show_data, self.show_plot_data, self.do_close)

		# Set initial Y variables, and also clear output
		self.x_selected(None)

	def set_autoupdate(self):
		if self.autoupdate_var.get() == "1":
			self.auto_update = True
		else:
			self.auto_update = False

	def x_selected(self, *args):
		self.clear_output()
		new_x_is_numeric = self.x_var.get() in self.numeric_columns
		if self.x_is_numeric or new_x_is_numeric:
			self.yvar_list = [v for v in self.numeric_columns if v != self.x_var.get()]
			self.column_frame, self.column_table = treeview_table(self.var_frame, rowset=[[v] for v in self.yvar_list], \
					column_headers=['Y Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
			self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
			self.column_table.bind('<ButtonRelease-1>', self.q_redraw)
		else:
			self.q_redraw(None)
		self.x_is_numeric = new_x_is_numeric

	def q_redraw(self, *args):
		# At least two variables must be selected
		self.clear_output()
		self.dlg.bind("<Alt-x>")
		self.dlg.bind("<Alt-y>")
		self.dlg.bind("<Alt-t>")
		curr_selections = self.column_table.selection()
		if len(curr_selections) > 0 and self.x_var.get() != '' and self.aggreg_var.get() != '':
			self.get_data()
			if self.plot_data is not None:
				self.redraw()

	def get_data(self):
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None
		column_list = [self.x_var.get()]
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		# Record the number of data columns without labels.
		self.n_dataset_columns = len(column_list)
		# Get the selected data.
		dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is not None:
			# Remove missing values by case
			dataset = clean_missing(dataset, list(range(self.n_dataset_columns)))
			if len(dataset[0]) > 1:
				# Convert the X variable to strings if it is not a date/time, else to date or timestamp if possible
				x_type = [c[1] for c in self.column_specs if c[0] == self.x_var.get()][0]
				dataset[0] = conv_vect_dt(dataset[0], x_type)
				# Convert to floats for numeric data columns only
				for i in range(1,self.n_dataset_columns):
					dataset[i] = [conv_float(v) for v in dataset[i]]
				self.dataset = dataset
				self.data_labels = column_list
				self.data_btn["state"] = tk.NORMAL
				# Aggregate rows by X variable
				subsets = subset_by_groups(dataset[1:self.n_dataset_columns], dataset[0])
				agg_error, agg_data = aggregate_groups(subsets, self.aggreg_var.get())
				if not agg_error:
					# Check for NaN
					if dataset_contains(agg_data, lambda x: type(x) == float and math.isnan(x)):
						warning("Aggregated data has NaN values; plot cannot be produced.", {"parent": self.dlg})
					else:
						self.plot_data = sort_columns(agg_data)
						self.plot_data_labels = self.data_labels
						self.plot_data_btn["state"] = tk.NORMAL

	def clear_output(self):
		clear_dlg_hotkeys(self.dlg)
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED
		self.plot.clear()
		self.dlg.bind("<Alt-x>")
		self.dlg.bind("<Alt-y>")
		self.dlg.bind("<Alt-t>")

	def redraw(self):
		y_lbls = copy.copy(self.data_labels[1:])
		if wrap_at_underscores:
			y_lbls = [lbl.replace("_", " ") for lbl in y_lbls]
		y_lbls = [textwrap.wrap(lbl, width=wrapwidth) for lbl in y_lbls]
		y_lbls = ["\n".join(lbl) for lbl in y_lbls]
		# Draw a line plot for each Y variable
		plot_fn = self.plot.axes.plot if self.scatter_var.get() == '0' else self.plot.axes.scatter
		for yv in range(1, self.n_dataset_columns):
			clean_data = clean_missing([self.plot_data[0], self.plot_data[yv]], [0,1])
			#self.plot_axes.plot(clean_data[0], clean_data[1], label=self.plot_data_labels[yv])
			plot_fn(clean_data[0], clean_data[1], label=self.plot_data_labels[yv])

		self.plot.axes.legend()

		self.plot.set_axis_labels(self.xlabel or self.x_var.get(), self.ylabel or "Value")
		if self.plot_title is not None:
			self.plot.axes.set_title(self.plot_title)
			self.plot.axes.title.set_visible(True)
		else:
			self.plot.axes.title.set_visible(False)
		self.dlg.bind("<Alt-x>", self.set_xlabel)
		self.dlg.bind("<Alt-y>", self.set_ylabel)
		self.dlg.bind("<Alt-t>", self.set_title)
		self.plot.draw()

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels, "Data for Multi-Variable Line Plot")
			self.dlg.lift()

	def show_plot_data(self, *args):
		# Show data as summarized for a particular plot type.
		if self.plot_data is not None:
			show_columnar_table(self.dlg, "Data for Plotting", "Data to be plotted:", self.plot_data, \
					self.plot_data_labels, "Plot data")
			self.dlg.lift()
	
	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the transparency (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = min(1.0, max(new_alpha, 0.0))
			self.q_redraw()
	def set_title(self, *args):
		dlg = OneEntryDialog(self.dlg, "Plot Title", "Enter a title for the plot:", init_value=self.plot_title, nullable=True)
		self.plot_title = dlg.show()
		self.plot.axes.set_title(self.plot_title)
		self.plot.axes.title.set_visible(self.plot_title is not None)
		self.plot.canvas.draw()
	def set_xlabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "X-Axis Label", "Enter a label for the X-axis:", init_value=self.xlabel, nullable=True)
		self.xlabel = dlg.show()
		self.plot.axes.set_xlabel(self.xlabel or self.x_var.get())
		self.plot.canvas.draw()
	def set_ylabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "Y-Axis Label", "Enter a label for the Y-axis label:", init_value=self.ylabel, nullable=True)
		self.ylabel = dlg.show()
		self.plot.axes.set_ylabel(self.ylabel or "Value")
		self.plot.canvas.draw()
	def do_close(self, *args):
		self.parent.remove_plot(self)
		super().do_cancel(args)



class RankAbundDialog(Dialog):
	def __init__(self, parent, column_specs, prohibited_columns):
		self.parent = parent
		self.column_specs = column_specs
		self.prohibited_columns = prohibited_columns
		self.numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float") and not c[1] in prohibited_columns]
		self.numeric_columns.sort()
		self.string_columns = sorted([c[0] for c in self.column_specs if c[1] == "string" and not c[1] in prohibited_columns])
		self.auto_update = True
		self.alpha = 0.45
		self.plot_title = None
		self.xlabel = None
		self.ylabel = None
		super().__init__("Rank Abundance Plot", "Select two or more variables from the list below.",
				help_url="https://mapdata.readthedocs.io/en/latest/plot_rankabund.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		self.alpha = 0.45
		# Data
		self.dataset = None
		self.data_labels = None
		self.plot_data = None
		self.plot_data_labels = None

		# Controls
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_redraw)
		self.autoupdate_var = add_autoupdate(self.ctrl_frame, 0, 1, self.set_autoupdate)
		self.ylog_var = tk.StringVar(self.ctrl_frame, "0")
		self.ylog_ck = ttk.Checkbutton(self.ctrl_frame, text="Log Y axis", state=tk.NORMAL, command=self.q_redraw, variable=self.ylog_var,
				onvalue="1", offvalue="0")
		self.ylog_ck.grid(row=0, column=3, sticky=tk.W, padx=(6,6), pady=(3,3))
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(3, weight=1)

		# PanedWindow within content_frame: left is input listbox and other data selection controls, right is plot
		io_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		io_panes.grid(row=0, column=0, sticky=tk.NSEW, padx=3, pady=3)

		self.inp_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE)
		self.inp_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.inp_frame.rowconfigure(0, weight=1)
		self.inp_frame.columnconfigure(0, weight=1)
		io_panes.add(self.inp_frame, weight=1)

		self.output_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE)
		self.output_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.output_frame.rowconfigure(0, weight=1)
		self.output_frame.columnconfigure(0, weight=1)
		io_panes.add(self.output_frame, weight=12)

		# Variable selection
		# Add multi-select list of numeric variables to the leftmost pane
		var_frame = tk.Frame(self.inp_frame)
		var_frame.grid(row=0, column=0, columnspan=2, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_redraw)

		# Grouping selection
		self.groupby_var = tk.StringVar(self.inp_frame, "")
		groupby_lbl = ttk.Label(self.inp_frame, text="Group by:")
		groupby_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.groupby_sel = ttk.Combobox(self.inp_frame, state=tk.NORMAL, textvariable=self.groupby_var, values=[""]+self.string_columns, width=24)
		self.groupby_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.groupby_sel.bind("<<ComboboxSelected>>", self.q_redraw)

		# Output plot
		plot_frame = tk.Frame(self.output_frame)
		plot_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot = Plot(plot_frame, 3.5, 3.5)

		# Buttons
		self.data_btn, self.plot_data_btn = add_help_src_plot_close_btns(self.dlg, self.btn_frame,
				self.do_help, self.show_data, self.show_plot_data, self.do_close)

		# Initialize output frames
		self.clear_output()

	def set_autoupdate(self):
		if self.autoupdate_var.get() == "1":
			self.auto_update = True
		else:
			self.auto_update = False

	def q_redraw(self, *args):
		# At least two variables must be selected
		self.clear_output()
		curr_selections = self.column_table.selection()
		if len(curr_selections) > 1:
			self.get_data()
			if self.dataset is not None:
				self.redraw()

	def get_data(self):
		self.data_btn["state"] = tk.DISABLED
		self.plot_data_btn["state"] = tk.DISABLED
		self.dataset = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		# Record the number of data columns without labels.
		self.n_dataset_columns = len(column_list)
		# Add the grouping column -- ALWAYS immediately after the variable columns for indexing consistency
		if self.groupby_var.get() != "":
			column_list.append(self.groupby_var.get())
		# Get additional map labeling column if specified
		if map_settings.label_col is not None:
			column_list.append(map_settings.label_col)
		# Get the selected data.
		dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is None:
			self.dataset = None
			self.data_labels = None
		else:
			# Convert to floats for numeric data columns only
			for i in range(self.n_dataset_columns):
				if column_list[i] in self.numeric_columns:
					dataset[i] = [conv_float(v) for v in dataset[i]]
			self.dataset = dataset
			self.data_labels = column_list
			self.data_btn["state"] = tk.NORMAL
			self.dlg.bind("<Alt-x>", self.set_xlabel)
			self.dlg.bind("<Alt-y>", self.set_ylabel)

	def clear_output(self):
		clear_dlg_hotkeys(self.dlg)
		self.plot_data_btn["state"] = tk.DISABLED
		self.plot.clear()

	def redraw(self):
		self.clear_output()
		ranks = list(range(1, self.n_dataset_columns+1))
		# Sum abundances for each selected variable column, across all rows or across groups
		ylbl = "Relative Abundance (Value)"
		if self.ylog_var.get() == '1':
			ylbl = "Log10 of Relative Abundance (Value)"
		groupvar = self.groupby_var.get()
		if groupvar != "":
			# Sum each column by the grouping variable, producing a dictionary of lists
			grpstrs = [str(g) for g in self.dataset[self.n_dataset_columns]]
			grps = sorted(list(set(grpstrs)))
			grpsums = {}
			for g in grps:
				grpsums[g] = [0] * self.n_dataset_columns
			for i in range(len(self.dataset[0])):
				grp = grpstrs[i]
				for j in range(self.n_dataset_columns):
					if self.dataset[j][i] is not None:
						grpsums[grp][j] = grpsums[grp][j] + self.dataset[j][i]
			# For each group, calculate the total for the list, standardize each value, and sort from largest to smallest
			dictstd = {}
			for k in grpsums.keys():
				dsums = grpsums[k]
				dsumsum = sum(dsums)
				if dsumsum > 0:
					dstd = [d/dsumsum for d in dsums]
					dstd.sort(reverse=True)
					if self.ylog_var.get() == '1':
						dstd = [math.log10(d) if d > 0 else None for d in dstd]
					dictstd[k] = dstd
			if len(dictstd) > 0:
				# Create the plotdata set
				self.plot_data_labels = ["Rank"] + list(dictstd.keys())
				self.plot_data = [ranks] + [dictstd[k] for k in dictstd.keys()]
				self.plot_data_btn["state"] = tk.NORMAL
				# Legend
				# Create a line plot for each row
				for i in range(len(self.plot_data_labels)-1):
					self.plot.axes.plot(self.plot_data[0], self.plot_data[i+1], label=self.plot_data_labels[i+1], alpha=self.alpha)
				if len(self.plot_data_labels)-1 > 1:
					self.plot.axes.legend()
		else:
			# Sum each column across all rows, producing a list
			dsums = [0] * self.n_dataset_columns
			for i in range(len(self.dataset[0])):
				for j in range(self.n_dataset_columns):
					if self.dataset[j][i] is not None:
						dsums[j] = dsums[j] + self.dataset[j][i]
			# Calculate the total for the list, standardize each value, and sort from largest to smallest
			dsumsum = sum(dsums)
			if dsumsum > 0:
				dstd = [d/dsumsum for d in dsums]
				dstd.sort(reverse=True)
				if self.ylog_var.get() == '1':
					dstd = [math.log10(d) if d > 0 else None for d in dstd]
				# Create the plotdata set
				self.plot_data = [ranks, dstd]
				self.plot_data_labels = ["Rank", ylbl]
				self.plot_data_btn["state"] = tk.NORMAL
				# Create a line plot
				self.plot.axes.plot(self.plot_data[0], self.plot_data[1], alpha=self.alpha)
		self.plot.set_axis_labels(self.xlabel or self.plot_data_labels[0], self.ylabel or ylbl)
		if self.plot_title is not None:
			self.plot.axes.set_title(self.plot_title)
		self.plot.draw()
		self.dlg.bind("<Alt-a>", self.set_alpha)
		self.dlg.bind("<Alt-x>", self.set_xlabel)
		self.dlg.bind("<Alt-y>", self.set_ylabel)
		self.dlg.bind("<Alt-t>", self.set_title)

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels, "Data for Rank-Abundance Plot")
			self.dlg.lift()

	def show_plot_data(self, *args):
		# Show data as summarized for a particular plot type.
		if self.plot_data is not None:
			show_columnar_table(self.dlg, "Data for Plotting", "Data to be plotted:", self.plot_data, \
					self.plot_data_labels, "Plot data")
			self.dlg.lift()
	
	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the transparency (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = min(1.0, max(new_alpha, 0.0))
			self.q_redraw()
	def set_title(self, *args):
		dlg = OneEntryDialog(self.dlg, "Plot Title", "Enter a title for the plot:", init_value=self.plot_title, nullable=True)
		self.plot_title = dlg.show()
		self.plot.axes.set_title(self.plot_title)
		self.plot.axes.title.set_visible(self.plot_title is not None)
		self.plot.canvas.draw()
	def set_xlabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "X-Axis Label", "Enter a label for the X-axis:", init_value=self.xlabel, nullable=True)
		self.xlabel = dlg.show()
		self.plot.axes.set_xlabel(self.xlabel or "Rank")
		self.plot.canvas.draw()
	def set_ylabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "Y-Axis Label", "Enter a label for the Y-axis label:", init_value=self.ylabel, nullable=True)
		self.ylabel = dlg.show()
		self.plot.axes.set_ylabel(self.ylabel or "Relative Abundance (Value)")
		self.plot.canvas.draw()

	def do_close(self, *args):
		self.parent.remove_plot(self)
		super().do_cancel(args)



class UnivarStatsDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.auto_update = True
		super().__init__("Univariate Statistics",
				"Select one or more variables from the left to see univariate statistics.  Use Ctrl-click or Shift-click to select multiple rows.",
				help_url="https://mapdata.readthedocs.io/en/latest/univarstats.html")
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.categ_columns = sorted([c[0] for c in self.column_specs if c[1] in ("string", "boolean")])
		self.dnames = ["Variable", " N ", "Min.", "Max.", "Mean", "Median", "Mode", \
				"Geo. mean", "Std.Dev.", "C.V.", "Sum", "5th %ile", "95th %ile", "Anderson-Darling p", \
				"Lilliefors p", "Omnibus normality p", "Rosner's outliers", "Tukey outliers"]
		self.logdnames = ["Variable", " N ", "Min.", "Max.", "Mean", "Median", "Mode", \
				"Std.Dev.", "C.V.", "Sum", "5th %ile", "95th %ile", "Anderson-Darling p", "Lilliefors p", \
				"Omnibus normality p", "Rosner's outliers", "Tukey outliers"]
		self.statdata = []
		self.dlg.bind("<Control-s>")
		self.dlg.bind("<Control-z>")

		# Top controls
		# Only the 'Selected only' checkbox
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_recalc)
		self.autoupdate_var = add_autoupdate(self.ctrl_frame, 0, 1, self.set_autoupdate)
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(1, weight=1)

		# Two panes for the variable frame and the content frame
		frame_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW, padx=3, pady=3)

		# Variable frame for list of quantitative columns/variables
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_recalc)

		# Add an optional 'group by' variable selection
		self.groupby_var = tk.StringVar(var_frame, "")
		groupby_lbl = ttk.Label(var_frame, text="Group by:")
		groupby_lbl.grid(row=1, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		self.groupby_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.groupby_var, 
				values= [''] + self.categ_columns, width=24)
		self.groupby_sel.grid(row=2, column=0, sticky=tk.W, padx=(12,3), pady=(3,3))
		self.groupby_sel.bind("<<ComboboxSelected>>", self.q_recalc)

		# output frame.  This contains a tabbed Notebook widget,
		# with separate pages for statistics of untransformed and log-transformed data.
		self.output_frame = tk.Frame(frame_panes, borderwidth=3, relief=tk.RIDGE)
		self.output_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.output_frame.rowconfigure(0, weight=1)
		self.output_frame.columnconfigure(0, weight=1)
		frame_panes.add(self.output_frame, weight=12)
		output_pages = ttk.Notebook(self.output_frame)
		output_pages.grid(row=0, column=0, sticky=tk.NSEW)
		self.data_page = tk.Frame(output_pages)
		self.log_page = tk.Frame(output_pages)
		self.data_page.name = "untransformed"
		self.log_page.name = "logtransformed"
		self.data_page.rowconfigure(0, weight=1)
		self.data_page.columnconfigure(0, weight=1)
		self.log_page.rowconfigure(0, weight=1)
		self.log_page.columnconfigure(0, weight=1)
		self.data_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.log_page.grid(row=0, column=0, sticky=tk.NSEW)
		output_pages.add(self.data_page, text="Untransformed")
		output_pages.add(self.log_page, text="Log-transformed")

		# initialize content frame with headings, no data
		self.clear_output()

		# Buttons
		self.data_btn = add_help_src_close_btns(self.dlg, self.btn_frame, self.do_help, self.show_data, self.do_close)

	def q_recalc(self, *args):
		if len(self.column_table.selection()) > 0:
			self.clear_output()
			self.get_data()
		if self.dataset is not None and len(self.dataset[0]) > 0:
			self.data_btn["state"] = tk.NORMAL
			self.recalc()
		else:
			self.data_btn["state"] = tk.DISABLED

	def get_data(self):
		# Get the selected data into 'dataset'
		self.dataset = None
		self.n_dataset_columns = 0
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		if len(column_list) > 0:
			self.n_dataset_columns = len(column_list)
			if self.groupby_var.get() != "":
				column_list.append(self.groupby_var.get())
			# Get either only the selected data or all data.
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				self.dataset = dataset
				self.data_labels = column_list

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for univariate stats")

	def univar_stats(self, dlist):
		# Return the univariate statistics for the data values in 'dlist' and for log-transformed versions of those.
		# The geometric mean is returned only for the untransformed data.
		#
		def stats(d, meanlog, do_geomean=False):
			mean_d = statistics.fmean(d)
			if len(d) > 1:
				stdev_d = statistics.stdev(d)
			else:
				stdev_d = 0.0
			dd = [len(d), min(d), max(d), fp_display(mean_d, 4)]
			if len(d) > 1:
				dd.extend([fp_display(statistics.median(d), 4), statistics.mode(d)])
			else:
				dd.extend([None, None])
			# Geo. mean
			if do_geomean:
				if meanlog is not None:
					dd.append(fp_display(10**meanlog, 4))
				else:
					dd.append("NC")
			if len(d) > 1:
				# Std. dev.
				dd.append(fp_display(stdev_d, 4))
				if mean_d != 0:
					# Coefficient of variation
					dd.append(fp_display(stdev_d/mean_d, 4))
				else:
					dd.append("NC")
			else:
				dd.extend([None, None])
			# Sum
			dd.append(sum(d))
			# Percentiles
			if len(d) > 4:
				dd.extend([fp_display(np.percentile(d, 5)), fp_display(np.percentile(d, 95))])
			else:
				dd.extend(["NC", "NC"])
			#=== Normality tests
			if len(d) > 4:
				da = np.array(d)
				normstats = []
				if stdev_d != 0.0:
					# Anderson-Darling
					try:
						adval, adpval = normal_ad(da)
						normstats.append("%.2E" % adpval)
					except:
						normstats.append("NC")
					# Lilliefors
					try:
						lfval, lfpval = lilliefors(da)
						normstats.append("%.2E" % lfpval)
					except:
						normstats.append("NC")
					# Omnibus
					if len(da) > 19:
						try:
							stat, p = spstats.normaltest(da)
							normstats.append("%.2E" % p)
						except:
							normstats.append("NC")
					else:
						normstats.append("NC")
				else:
					normstats = ["NC", "NC", "NC"]
				dd.extend(normstats)
			else:
				dd.extend(["NC", "NC", "NC"])
			#=== Outlier evaluation
			if len(d) > 14:
				if stdev_d > 0.0:
					max_outliers = 5 if len(d) < 100 else 10
					try:
						dd.append(rosners_test(d, 0.05, max_outliers))
					except:
						dd.append("NC")
				else:
					dd.append("NC")
			else:
				dd.append("NC")
			if len(d) > 5:
				dd.append(tukey_outliers(d))
			else:
				dd.append("NC")
			return dd
		#
		ds = [float(dv) for dv in dlist if dv is not None and not (type(dv) == type('') and dv.strip()=='')]
		if len(ds) == 0:
			return [0], [0]
		logd = logvector(ds)
		if logd is not None:
			meanlog = statistics.fmean(logd)
		else:
			meanlog = None
		normstats = stats(ds, meanlog, do_geomean=True)
		if logd is not None:
			logstats = stats(logd, meanlog, do_geomean=False)
			return normstats, logstats
		else:
			return normstats, None
	
	def recalc(self):
		# Calculate and display statistics for each selected column.
		for ctl in self.data_page.winfo_children():
			ctl.destroy()
		for ctl in self.log_page.winfo_children():
			ctl.destroy()
		self.statdata = []
		self.logstatdata = []
		if self.data_labels is not None:
			if self.groupby_var.get() == "":
				for var in range(self.n_dataset_columns):
					vardata = self.dataset[var]
					if len(vardata) > 0 and not (len(vardata) == 1 and (vardata[0] is None or vardata[0] == '')):
						dd, logdd = self.univar_stats(self.dataset[var])
						self.statdata.append([self.data_labels[var]] + dd)
						if logdd is not None:
							self.logstatdata.append([self.data_labels[var]] + logdd)
				dnames = self.dnames
				logdnames = self.logdnames
			else:
				datasetdict = subset_by_groups(self.dataset[0:self.n_dataset_columns], self.dataset[self.n_dataset_columns])
				for grp in datasetdict.keys():
					ds = datasetdict[grp]
					for var in range(self.n_dataset_columns):
						vardata = ds[var]
						if len(vardata) > 0 and not (len(vardata) == 1 and (vardata[0] is None or vardata[0] == '')):
							dd, logdd = self.univar_stats(vardata)
							self.statdata.append([grp, self.data_labels[var]] + dd)
							if logdd is not None:
								self.logstatdata.append([grp, self.data_labels[var]] + logdd)
				dnames = ["Group"] + self.dnames
				logdnames = ["Group"] + self.logdnames
		tframe, tdata = treeview_table(self.data_page, self.statdata, dnames)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		ltframe, ltdata = treeview_table(self.log_page, self.logstatdata, logdnames)
		ltframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.dlg.bind("<Control-s>", self.save_data)
		self.dlg.bind("<Control-z>", self.save_logdata)

	def clear_output(self):
		for ctl in self.data_page.winfo_children():
			ctl.destroy()
		for ctl in self.log_page.winfo_children():
			ctl.destroy()
		clear_dlg_hotkeys(self.dlg)
		self.dataset = None
		tframe, tdata = treeview_table(self.data_page, [], self.dnames)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		ltframe, ltdata = treeview_table(self.log_page, [], self.logdnames)
		ltframe.grid(row=0, column=0, sticky=tk.NSEW)

	def set_autoupdate(self):
		self.auto_update = self.autoupdate_var.get() == "1"
		if self.auto_update:
			self.q_recalc(None)
	def save_data(self, *args):
		export_data_table(self.dnames, self.statdata, sheetname="Stats for untransformed data")
	def save_logdata(self, *args):
		export_data_table(self.logdnames, self.logstatdata, sheetname="Stats for log-transformed data")
	def do_close(self, *args):
		self.parent.remove_univar(self)
		super().do_cancel(args)




class BivarStatsDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.alpha = 0.45
		self.theilsen = False
		self.plot_title = None
		self.xlabel = None
		self.ylabel = None
		super().__init__("Bivariate Statistics",
				"Select two variables.  If the X variable is a date/time, only the results of a Runs Test will be shown.",
				help_url="https://mapdata.readthedocs.io/en/latest/bivariatestats.html")
		clear_dlg_hotkeys(self.dlg)
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float")]
		self.numeric_columns.sort()
		self.quant_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float", "date", "timestamp", "timestamptz")]
		self.quant_columns.sort()
		self.output_columns = ["Statistic", "Value"]
		self.statdata = []

		# Controls
		self.x_var = tk.StringVar(self.ctrl_frame, "")
		x_lbl = ttk.Label(self.ctrl_frame, text="X column:")
		x_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.x_sel = ttk.Combobox(self.ctrl_frame, state=tk.NORMAL, textvariable=self.x_var, values=self.quant_columns, width=24)
		self.x_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.x_sel.bind("<<ComboboxSelected>>", self.q_recalc)

		self.y_var = tk.StringVar(self.ctrl_frame, "")
		y_lbl = ttk.Label(self.ctrl_frame, text="Y column:")
		y_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.y_sel = ttk.Combobox(self.ctrl_frame, state=tk.NORMAL, textvariable=self.y_var, values=self.numeric_columns, width=24)
		self.y_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.y_sel.bind("<<ComboboxSelected>>", self.q_recalc)

		self.xlog_var = tk.StringVar(self.ctrl_frame, "0")
		self.xlog_ck = ttk.Checkbutton(self.ctrl_frame, text="Log X", state=tk.NORMAL, command=self.q_recalc, variable=self.xlog_var,
				onvalue="1", offvalue="0")
		self.xlog_ck.grid(row=0, column=2, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.ylog_var = tk.StringVar(self.ctrl_frame, "0")
		self.ylog_ck = ttk.Checkbutton(self.ctrl_frame, text="Log Y", state=tk.NORMAL, command=self.q_recalc, variable=self.ylog_var,
				onvalue="1", offvalue="0")
		self.ylog_ck.grid(row=1, column=2, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 2, 0, self.q_recalc, colspan=2)

		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(2, weight=1)

		# PanedWindow within content_frame: left is table, right is plot
		output_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		output_panes.grid(row=0, column=0, sticky=tk.NSEW)

		self.out_tbl_frm = tk.Frame(output_panes, borderwidth=3, relief=tk.RIDGE)
		self.out_tbl_frm.grid(row=0, column=0, sticky=tk.NSEW)
		self.out_tbl_frm.rowconfigure(0, weight=1)
		self.out_tbl_frm.columnconfigure(0, weight=1, minsize=400)
		output_panes.add(self.out_tbl_frm, weight=1)

		self.out_plt_frm = tk.Frame(output_panes, borderwidth=3, relief=tk.RIDGE)
		self.out_plt_frm.grid(row=0, column=1, sticky=tk.NSEW)
		self.out_plt_frm.rowconfigure(0, weight=1)
		self.out_plt_frm.columnconfigure(0, weight=1)
		output_panes.add(self.out_plt_frm, weight=1)
		# Add a notebook to the plot frame with pages for the data plot, residual plot, and Sen trend.
		#   Must pack, not grid, sub-frames because NavigationToolbar2Tk uses pack internally.
		plot_pages = ttk.Notebook(self.out_plt_frm)
		plot_pages.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		# Data plot
		plot_frame = tk.Frame(plot_pages)
		plot_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.dataplot = Plot(plot_frame, 3.5, 3.5)
		plot_pages.add(plot_frame, text="Data")
		# Residual plot
		resid_frame = tk.Frame(plot_pages)
		resid_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.residplot = Plot(resid_frame, 3.5, 3.5)
		plot_pages.add(resid_frame, text="Residuals")
		# Şen 2012 trend plot
		sen_frame = tk.Frame(plot_pages)
		sen_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.senplot = Plot(sen_frame, 3.5, 3.5)
		plot_pages.add(sen_frame, text="Şen trend")

		# Initialize output frames
		self.clear_output()

		# Buttons
		self.data_btn = add_help_src_close_btns(self.dlg, self.btn_frame, self.do_help, self.show_data, self.do_close)

	def get_data(self):
		self.data_btn["state"] = tk.DISABLED
		self.xlog_ck["state"] = tk.NORMAL
		self.ylog_ck["state"] = tk.NORMAL
		self.dataset = None
		column_list = [self.x_var.get(), self.y_var.get()]
		# Get either only the selected data or all data.
		dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is None or len(dataset[0]) == 0:
			self.dataset = None
			self.data_labels = None
			self.data_btn["state"] = tk.DISABLED
		else:
			# Remove missing data
			clean_data = clean_missing(dataset, list(range(len(dataset))))
			# Convert to floats for numeric data only
			for i in range(len(clean_data)):
				if column_list[i] in self.numeric_columns:
					clean_data[i] = [conv_float(v) for v in clean_data[i]]
			# Convert X to date or datetime values if necessary
			xvar_type = [self.column_specs[i][1] for i in range(len(self.column_specs)) if self.column_specs[i][0] == self.x_var.get()][0]
			if xvar_type == "date":
				clean_data[0] = [parse_date(x) for x in clean_data[0]]
			elif xvar_type == "timestamp":
				clean_data[0] = [parse_datetime(x) for x in clean_data[0]]
			elif xvar_type == "timestamptz":
				clean_data[0] = [parse_datetimetz(x) for x in clean_data[0]]
			# Log-transform data if specified.
			self.data_labels = column_list
			if self.xlog_ck["state"] != tk.DISABLED and self.xlog_var.get() == "1":
				log_err = False
				if self.x_var.get() in self.numeric_columns:
					logged_x = logvector(clean_data[0])
					if logged_x is not None:
						clean_data[0] = logged_x
						self.data_labels[0] = "Log10 of " + self.x_var.get()
					else:
						log_err = True
				else:
					log_err = True
				if log_err:
					self.xlog_var.set("0")
					self.xlog_ck["state"] = tk.DISABLED
					warning_nolog(msg="X axis data can't be log-transformed.", parent=self.dlg)
			if self.ylog_ck["state"] != tk.DISABLED and self.ylog_var.get() == "1" and len(clean_data) > 1 \
					and self.y_var.get() in self.numeric_columns:
				logged_y = logvector(clean_data[1])
				if logged_y is not None:
					clean_data[1] = logged_y
					self.data_labels[1] = "Log10 of " + self.y_var.get()
				else:
					self.ylog_var.set("0")
					self.ylog_ck["state"] = tk.DISABLED
			#
			self.dataset = sort_columns(clean_data)
			self.data_btn["state"] = tk.NORMAL

	def clear_output(self):
		for ctl in self.out_tbl_frm.winfo_children():
			ctl.destroy()
		clear_dlg_hotkeys(self.dlg)
		tframe, tdata = treeview_table(self.out_tbl_frm, [], ["Statistic", "Value"])
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.dataplot.clear()
		self.residplot.clear()
		self.senplot.clear()

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for bivariate stats")

	def q_recalc(self, get_data=True, *args):
		if self.x_var.get() != '' and self.y_var.get() != '':
			if get_data or self.dataset is None:
				self.get_data()
			if self.dataset is not None and len(self.dataset[0]) > 2:
				self.recalc()
			else:
				self.clear_output()

	def recalc(self):
		self.clear_output()
		for ctl in self.out_tbl_frm.winfo_children():
			ctl.destroy()
		self.statdata = []
		if self.dataset is not None:
			regr_run = False
			total_ss = 0.0
			x_is_numeric = self.x_var.get() in self.numeric_columns
			N = len(self.dataset[0])
			xpfx = "" if self.xlog_var.get() == '0' else "Log10 of "
			ypfx = "" if self.ylog_var.get() == '0' else "Log10 of "
			self.statdata.append(["X variable", xpfx + self.x_var.get()])
			self.statdata.append(["Y variable", ypfx + self.y_var.get()])
			self.statdata.append(["N", N])
			if N > 1:
				if x_is_numeric:
					xmean = statistics.fmean(self.dataset[0])
					xmid = xmean
				else:
					xmean = None
					xmid = (max(self.dataset[0]) - min(self.dataset[0]))/2
				ymean = statistics.fmean(self.dataset[1])
				xarray = np.array(self.dataset[0])
				yarray = np.array(self.dataset[1])
				if x_is_numeric:
					try:
						self.statdata.append(["Covariance", fp_display(statistics.covariance(self.dataset[0], self.dataset[1]), 3)])
					except:
						pass
					try:
						pearsonr = spstats.pearsonr(xarray, yarray)
						self.statdata.append(["Pearson's r", fp_display(pearsonr.statistic)])
						rp = pearsonr.pvalue
						rp = fp_display(rp, 3)
						self.statdata.append(["p value for Pearson's r", rp])
					except:
						pass
					try:
						spearmanr = spstats.spearmanr(xarray, yarray)
						self.statdata.append(["Spearman's rho", fp_display(spearmanr.statistic)])
						sp = spearmanr.pvalue
						sp = fp_display(sp, 3)
						self.statdata.append(["p value for Spearman's rho", sp])
					except TypeError:
						pass
					try:
						kendalltau = spstats.kendalltau(np.array(self.dataset[0]), np.array(self.dataset[1]))
						self.statdata.append(["Kendall's tau", fp_display(kendalltau.statistic)])
						ktau = kendalltau.pvalue
						ktau = fp_display(ktau, 3)
						self.statdata.append(["p value for Kendall's tau", ktau])
					except:
						pass
					try:
						xi = xicor(np.array(self.dataset[0]), np.array(self.dataset[1]))
						self.statdata.append(["Chatterjee's xi", fp_display(xi)])
					except:
						pass
					# Linear regression
					ols_model = sm.OLS(np.array(self.dataset[1]), sm.add_constant(np.array(self.dataset[0])))
					ols_fit = ols_model.fit()
					slope = ols_fit.params[1]
					regr_run = True
					self.statdata.append(["OLS Regression slope", fp_display(slope)])
					self.statdata.append(["OLS Regression intercept", fp_display(ols_fit.params[0])])
					r_square = None
					try:
						r_square = ols_fit.rsquared
						self.statdata.append(["Regression R squared", fp_display(r_square)])
					except:
						pass
					try:
						self.statdata.append(["Regression adj. R squared", fp_display(ols_fit.rsquared_adj)])
					except:
						pass
					self.statdata.append(["Regression total SS", fp_display(ols_fit.centered_tss, 4)])
					total_ss = sum([(self.dataset[1][i] - ymean)**2 for i in range(N)])
					try:
						self.statdata.append(["Regression explained SS", fp_display(ols_fit.ess, 4)])
					except:
						pass
					try:
						self.statdata.append(["Regression residual SS", fp_display(ols_fit.ssr, 4)])
					except:
						pass
					if r_square is not None:
						r_sd = math.sqrt(r_square)
						self.statdata.append(["Coef. Det. in std. dev. (CoD_SD)", fp_display(r_sd / (r_sd + math.sqrt(1 - r_square)))])
					self.statdata.append(["Regression p for slope=0", "%.2E" % ols_fit.pvalues[1]])
					self.statdata.append(["Regression p for intercept=0", "%.2E" % ols_fit.pvalues[0]])
					self.statdata.append(["Regression AIC", fp_display(ols_fit.aic, 3)])
					self.statdata.append(["Regression BIC", fp_display(ols_fit.bic, 3)])
					bp_test = het_breuschpagan(ols_fit.resid, ols_fit.model.exog)
					self.statdata.append(["Regression Breusch-Pagan p value", "%.2E" % bp_test[1]])

				# Mann-Kendall Trend test
				if mk_available and len(self.dataset[0]) > 3:
					mk_out = mk.original_test(self.dataset[1])
					mk_p = mk_out.p
					if mk_p < 0.001:
						mk_p = "%.2E" % mk_p
					else:
						mk_p = fp_display(mk_p, 3)
					self.statdata.append(["Mann-Kendall trend test; p value", mk_p])

				# Runs test
				#if not regr_run or total_ss == 0.0:
				#	runs_p = "NC"
				#else:
				rz, rp = runstest_1samp(self.dataset[1], cutoff="median")
				if rp < 0.001:
					runs_p = "%.2E" % rp
				else:
					runs_p = fp_display(rp, 3)
				self.statdata.append(["Runs test; p value", runs_p])

				if x_is_numeric:
					# Theil-Sen estimators
					ts_slope, ts_intercept, ts_high, ts_low = spstats.theilslopes(yarray, xarray)
					self.statdata.append(["Theil-Sen slope", fp_display(ts_slope)])
					self.statdata.append(["Theil-Sen 95% CI on slope", "(%s, %s)" % (fp_display(ts_high), fp_display(ts_low))])
					self.statdata.append(["Theil-Sen intercept", fp_display(ts_intercept)])
					# Robust R-square per Kvalseth 1985 on the Theil-Sen fit.
					y_est = [ts_intercept + ts_slope * x for x in xarray]
					med_ydiffs = statistics.median([abs(yarray[i] - y_est[i]) for i in range(len(yarray))])
					med_meandiffs = statistics.median([abs(yarray[i] - ymean) for i in range(len(yarray))])
					robust_rsquare = 1 - (med_ydiffs/med_meandiffs)**2
					self.statdata.append(["Robust R-square", fp_display(robust_rsquare)])
				# Update plot
				self.dataplot.set_axis_labels(self.data_labels[0], self.data_labels[1])
				if regr_run:
					ols_ci = ols_fit.get_prediction().summary_frame()
					self.dataplot.axes.fill_between(self.dataset[0], list(ols_ci["mean_ci_lower"]), list(ols_ci["mean_ci_upper"]), color="antiquewhite", edgecolor="goldenrod", label="95% CI")
					self.dataplot.axes.axline((xmean, ymean), slope=slope, clip_on=True, label="Linear fit", color="darkorange", linestyle="dotted", linewidth=2)
					# Plot of residuals
					ols_resid = list(ols_fit.resid)
					self.residplot.axes.axhline(0.0, linestyle="dotted")
					self.residplot.axes.scatter(self.dataset[0], ols_resid, alpha=self.alpha)
					self.residplot.axes.set_xlabel(self.data_labels[0])
					self.residplot.axes.set_ylabel("Residuals for " + self.data_labels[1])
				if self.theilsen and x_is_numeric:
					self.dataplot.axes.axline((statistics.median(self.dataset[0]), statistics.median(self.dataset[1])), slope=ts_slope, \
							label="Theil-Sen line", color="darkgreen", linestyle="dotted", linewidth=2, alpha=0.65)
				if regr_run or (self.theilsen and x_is_numeric):
					self.dataplot.axes.legend()
				self.dataplot.axes.scatter(self.dataset[0], self.dataset[1], alpha=self.alpha)
				self.dataplot.draw()
				self.residplot.draw()
				# Plot of Şen trend
				ylen = len(self.dataset[1])
				if ylen > 3:
					if ylen % 2:
						ydata = self.dataset[1][1:]
						ylen = ylen - 1
					else:
						ydata = self.dataset[1]
					ymin = min(ydata)
					ymax = max(ydata)
					yofs = 0.05 * (ymax - ymin)
					ylims = (ymin - yofs, ymax + yofs)
					ysplit = int(ylen/2)
					firsthalf = ydata[:ysplit]
					secondhalf = ydata[ysplit:]
					self.senplot.axes.scatter(firsthalf, secondhalf)
					self.senplot.axes.set_xlim(ylims)
					self.senplot.axes.set_ylim(ylims)
					mid = statistics.median(ydata)
					self.senplot.axes.axline((mid,mid), slope=1, color="grey", alpha=0.65)
					self.senplot.set_axis_labels("First half of "+self.data_labels[1], "Second half of "+self.data_labels[1])
					self.senplot.draw()
				#
				self.dlg.bind("<Alt-a>", self.set_alpha)
				self.dlg.bind("<Alt-t>", self.set_title)
				self.dlg.bind("<Alt-x>", self.set_xlabel)
				self.dlg.bind("<Alt-y>", self.set_ylabel)
				self.dlg.bind("<Alt-s>", self.set_theilsen)
				if self.plot_title is not None:
					self.dataplot.axes.set_title(self.plot_title)
		if len(self.statdata) > 0:
			tframe, tdata = treeview_table(self.out_tbl_frm, self.statdata, self.output_columns)
			tframe.grid(row=0, column=0, stick=tk.NSEW)
			self.dlg.bind("<Control-s>", self.save_table)
			self.dlg.minsize(width=900, height=500)

	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the opacity (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = min(1.0, max(new_alpha, 0.0))
			self.q_recalc()
	def set_title(self, *args):
		dlg = OneEntryDialog(self.dlg, "Plot Title", "Enter a title for the plot:", init_value=self.plot_title, nullable=True)
		self.plot_title = dlg.show()
		self.dataplot.axes.set_title(self.plot_title)
		self.dataplot.draw()
	def set_xlabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "X-Axis Label", "Enter the X-axis label:", init_value=self.xlabel, nullable=True)
		self.xlabel = dlg.show()
		self.dataplot.axes.set_xlabel(self.xlabel or self.x_var.get())
		self.dataplot.draw()
	def set_ylabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "Y-Axis Label", "Enter the Y-axis label:", init_value=self.ylabel, nullable=True)
		self.ylabel = dlg.show()
		self.dataplot.axes.set_ylabel(self.ylabel or self.y_var.get())
		self.dataplot.draw()
	def set_theilsen(self, *args):
		self.theilsen = not self.theilsen
		self.q_recalc()
	def save_table(self, *args):
		export_data_table(self.output_columns, self.statdata, sheetname="Bivariate statistsics")
	def do_close(self, *args):
		self.parent.remove_bivar(self)
		super().do_cancel(args)


class CorrMatrixDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		super().__init__("Correlation Matrix",
				"Select two or more variables from the left to see the correlation matrix.  Use Ctrl-click or Shift-click to select multiple rows.",
				help_url="https://mapdata.readthedocs.io/en/latest/corrmatrix.html")
		# Data
		self.auto_update = True
		self.dataset = None
		self.data_labels = None
		self.corr_table = None
		self.corr_table_labels = None
		self.numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float")]
		self.numeric_columns.sort()
		self.show_labels = True
		self.dlg.bind("<Alt-l>")
		self.dlg.bind("<Control-s>")
		# Controls
		# Top controls are only the 'Selected only' checkbox
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_redraw)
		self.autoupdate_var = add_autoupdate(self.ctrl_frame, 0, 1, self.set_autoupdate)
		self.log10_var = tk.StringVar(self.ctrl_frame, "0")
		self.log10_ck = ttk.Checkbutton(self.ctrl_frame, text="Log10 transform data", command=self.q_redraw, variable=self.log10_var,
				onvalue="1", offvalue="0")
		self.log10_ck.grid(row=0, column=2, sticky=tk.W, padx=(12,3), pady=(3,3))
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(2, weight=1)

		# The content_frame encompasses the two panes of the variable frame and the content frame
		frame_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)

		# Variable frame for list of quantitative columns/variables
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1, minsize=200)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_redraw)
		self.corr_type_var = tk.StringVar(var_frame, "Pearson r")
		ttk.Label(var_frame, text="Correlation type:").grid(row=1, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		corr_type_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.corr_type_var, width=24)
		corr_type_sel["values"] = ["Pearson r", "Spearman rho", "Kendall tau", "Chatterjee xi"]
		corr_type_sel.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		corr_type_sel.bind('<<ComboboxSelected>>', self.q_redraw)

		# Content frame for correlation matrix figure
		self.corrfig_frame = tk.Frame(frame_panes, borderwidth=3, relief=tk.RIDGE)
		self.corrfig_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.plot = Plot(self.corrfig_frame, 3.5, 3.5)
		frame_panes.add(self.corrfig_frame, weight=12)

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.data_btn = new_data_button(self.dlg, self.btn_frame, 1, self.show_data)
		self.corr_data_btn = new_button(self.btn_frame, "Correlations", 0, 2, self.show_corr_data, tk.W, (3,3), (0,0), None, tk.DISABLED)
		new_close_button(self.dlg, self.btn_frame, 3, self.do_close)
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(3, weight=1)

		# initialize content frame with an empty plot
		self.clear_output()

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for correlation matrix")

	def show_corr_data(self, *args):
		if self.corr_table is not None:
			show_columnar_table(self.dlg, "Correlations", "Correlation coefficients:", self.corr_table, self.corr_table_labels, \
					"Correlations")

	def clear_output(self):
		self.plot.clear()
		self.data_btn["state"] = tk.DISABLED
		self.corr_data_btn["state"] = tk.DISABLED

	def q_redraw(self, *args, get_data=True):
		self.clear_output()
		if self.dataset is None or get_data:
			self.get_data()
		if self.dataset is not None and len(self.dataset) > 1 and len(self.dataset[0]) > 1:
			self.redraw()
			self.data_btn["state"] = tk.NORMAL
			self.corr_data_btn["state"] = tk.NORMAL

	def get_data(self):
		# Get the selected data into 'dataset'
		self.clear_output()
		self.dataset = None
		self.log10_ck["state"] = tk.NORMAL
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		if len(column_list) > 0:
			# Get either only the selected data or all data.
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				self.dataset = dataset
				self.data_labels = column_list
			if self.dataset is not None:
				# Remove missing data
				clean_data = clean_missing(dataset, list(range(len(dataset))))
				# Convert to floats
				for i in range(len(clean_data)):
					clean_data[i] = [conv_float(v) for v in clean_data[i]]
				# Log-transform data if specified.
				clean_labels = self.data_labels
				if self.log10_var.get() == '1':
					log_data = logdataset(clean_data, list(range(len(clean_data))))
					if log_data is None:
						warning_nolog(parent=self.dlg)
						self.log10_var.set("0")
						self.log10_ck["state"] = tk.DISABLED
					else:
						clean_data = log_data
						clean_labels = ["Log10 of " + v for v in clean_labels]
				#
				self.dataset = clean_data
				self.data_labels = clean_labels
				self.data_btn["state"] = tk.NORMAL

	def redraw(self):
		# (Re)draw the correlation matrix
		if self.data_labels is not None and len(self.data_labels) > 1:
			self.dlg.bind("<Alt-l>", self.set_labeling)
			nvar = len(self.data_labels)
			corrmat = np.eye(nvar)
			corrtype = self.corr_type_var.get()
			if corrtype == "Pearson r":
				corrfn = pearsonstat
			elif corrtype == "Spearman rho":
				corrfn = spearmanstat
			elif corrtype == "Kendall tau":
				corrfn = kendallstat
			else:
				corrfn = xicor
			dmat = np.asarray(self.dataset)
			for i in range(nvar):
				for j in range(i, nvar):
					if i != j:
						corrcoeff = corrfn(dmat[i], dmat[j])
						corrmat[i][j] = corrcoeff
						corrmat[j][i] = corrcoeff
			caxes = self.plot.axes.matshow(corrmat, cmap="BrBG", vmin=-1.0, vmax=1.0)
			self.plot.figure.colorbar(caxes)
			self.plot.axes.set_xticks(range(nvar), labels=self.data_labels, rotation=25)
			self.plot.axes.set_yticks(range(nvar), labels=self.data_labels)
			if self.show_labels:
				for i in range(nvar):
					for j in range(nvar):
						v = corrmat[i,j]
						c = "white" if abs(v) > 0.40 else "black"
						self.plot.axes.text(j, i, f"{corrmat[i,j]:.2f}", ha="center", va="center", color=c)
			self.plot.axes.set_title(corrtype)
			self.plot.draw()
			# Convert the upper right corner of the matrix into a table for display.
			self.corr_table_labels = ["Variable 1", "Variable 2", f"{corrtype}"]
			self.corr_table = [[], [], []]
			for i in range(len(self.data_labels)-1):
				for j in range(i+1,len(self.data_labels)):
					self.corr_table[0].append(self.data_labels[i])
					self.corr_table[1].append(self.data_labels[j])
					self.corr_table[2].append(corrmat[i][j])
			self.corr_data_btn["state"] = tk.NORMAL

	def set_autoupdate(self, *args):
		self.auto_update = self.autoupdate_var.get() == "1"
		if self.auto_update:
			self.q_redraw(None, get_data=True)
	def set_labeling(self, *args):
		self.show_labels = not self.show_labels
		self.q_redraw(get_data=False)
	def do_close(self, *args):
		self.parent.remove_corrmat(self)
		super().do_cancel(args)



class CosineSimilarityDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		super().__init__("Cosine Similarity",
				"Select two or more variables from the left, and a grouping variable, to see the cosine similarity matrix.  Use Ctrl-click or Shift-click to select multiple rows.",
				help_url="https://mapdata.readthedocs.io/en/latest/cossimmatrix.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.auto_update = True
		self.dataset = None
		self.data_labels = None
		self.agg_data = None		# Has same column headers as self.data_labels
		self.cos_table = None
		self.cos_table_labels = None
		self.agg_data_groups = None	# List of group IDs in order of the aggregated data
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		# Only allow group variables that have at least two values
		self.categ_columns = sorted([c[0] for c in self.column_specs if c[1] in ("string", "boolean") and c[3] > 1])
		self.show_labels = True
		self.dlg.bind("<Alt-l>")
		self.dlg.bind("<Control-s>")

		# Controls
		# Top controls are only the 'Selected only' checkbox
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_redraw)
		self.autoupdate_var = add_autoupdate(self.ctrl_frame, 0, 1, self.set_autoupdate)

		self.cleantype_var = tk.StringVar(self.ctrl_frame, "Drop cases")
		ttk.Label(self.ctrl_frame, text="Remove missing values by:").grid(row=0, column=2, sticky=tk.W, padx=(9,3), pady=(3,3))
		self.cleantype_sel = ttk.Combobox(self.ctrl_frame, state=tk.NORMAL, textvariable=self.cleantype_var, values=["Drop cases","Drop variables"], width=15)
		self.cleantype_sel.bind('<<ComboboxSelected>>', self.q_redraw)
		self.cleantype_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(3, weight=1)

		# The content_frame encompasses the two panes of the variable frame and the content frame
		frame_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)

		# Variable frame for list of quantitative columns/variables
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1, minsize=250)
		frame_panes.add(var_frame, weight=1)
		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_redraw)

		# Grouping variable (required)
		self.groupby_var = tk.StringVar(var_frame, "")
		ttk.Label(var_frame, text="Group by:").grid(row=1, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		groupby_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.groupby_var, width=24, values=self.categ_columns)
		groupby_sel.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		groupby_sel.bind('<<ComboboxSelected>>', self.q_redraw)

		# Aggregation method for rows within a group (required)
		self.aggreg_var = tk.StringVar(var_frame, "Arithmetic mean")
		ttk.Label(var_frame, text="Aggregate rows by:").grid(row=3, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		aggreg_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.aggreg_var, width=24)
		aggreg_sel["values"] = ["Arithmetic mean", "Geometric mean", "Harmonic mean", "Sum"]
		aggreg_sel.grid(row=4, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		aggreg_sel.bind('<<ComboboxSelected>>', self.q_redraw)

		# Plot frame for cosine similarity matrix figure
		self.simplot_frame = tk.Frame(frame_panes, borderwidth=3, relief=tk.RIDGE)
		self.simplot_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.plot = Plot(self.simplot_frame, 3.5, 3.5)
		frame_panes.add(self.simplot_frame, weight=12)
		# initialize content frame with an empty plot
		self.clear_output()

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.data_btn = new_data_button(self.dlg, self.btn_frame, 1, self.show_data)
		self.agg_data_btn = new_button(self.btn_frame, "Aggregates", 0, 2, self.show_agg, padx=(3,3), state=tk.DISABLED)
		self.sim_data_btn = new_button(self.btn_frame, "Similarities", 0, 3, self.show_sim, padx=(3,3), state=tk.DISABLED)
		new_close_button(self.dlg, self.btn_frame, 4, self.do_close)
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(4, weight=1)

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for covariance matrix")
	def show_agg(self, *args):
		if self.agg_data is not None:
			show_columnar_table(self.dlg, "Aggregated Data", "Selected data, aggregated for analysis:", self.agg_data, self.data_labels[0:len(self.dataset)], \
					"Aggregated data for cosine similarity matrix")

	def show_sim(self, *args):
		if self.cos_table is not None:
			show_columnar_table(self.dlg, "Cosine Similarities", "Pairwise cosine similarities:", self.cos_table, self.cos_table_labels, \
					"Cosine similarities")

	def clear_output(self):
		self.plot.clear()
		self.agg_data = None

	def q_redraw(self, *event):
		self.clear_output()
		if (len(self.column_table.selection()) > 1 and self.groupby_var.get() != ''):
			self.get_data()
			if self.dataset is not None and len(self.dataset) > 1 and len(self.dataset[0]) > 1 and self.agg_data is not None:
				self.redraw()
			else:
				self.data_btn["state"] = tk.DISABLED
				self.agg_data_btn["state"] = tk.DISABLED
				self.sim_data_btn["state"] = tk.DISABLED

	def get_data(self):
		# Get the selected data into 'dataset'
		self.clear_output()
		self.dataset = None
		self.agg_data = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		self.n_dataset_columns = len(column_list)
		column_list.append(self.groupby_var.get())
		if self.n_dataset_columns > 0:
			# Get either only the selected data or all data.
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				self.dataset = dataset
				self.data_labels = column_list
			# Remove missing data
			if self.cleantype_var.get() == "Drop variables":
				# Remove columns with missing data.  This may remove all columns.
				clean_data, column_list, n_removed = clean_missing_columns(dataset,
						column_list, list(range(self.n_dataset_columns)))
				self.n_dataset_columns = self.n_dataset_columns - n_removed
			else:
				# Remove rows with missing data.  This may remove all rows.
				clean_data = clean_missing(dataset, list(range(self.n_dataset_columns+1)))
			dataset = None
			# Check that there are at least two groups and two variables
			groups = list(set(clean_data[self.n_dataset_columns]))
			n_groups = len(groups)
			if n_groups < 2:
				self.dataset = None
				self.data_labels = None
				self.data_btn["state"] = tk.DISABLED
				warning("There are too few groups.", {"parent": self.dlg})
			elif self.n_dataset_columns < 2:
				self.dataset = None
				self.data_labels = None
				self.data_btn["state"] = tk.DISABLED
				warning("There are too few variables.", {"parent": self.dlg})
			else:
				# Convert all group ids to strings
				clean_data[self.n_dataset_columns] = [str(v) for v in clean_data[self.n_dataset_columns]]
				groups = sorted_numstrs(list(set(clean_data[self.n_dataset_columns])))
				# Convert all numerics to floats
				for i in range(self.n_dataset_columns):
					clean_data[i] = [conv_float(v) for v in clean_data[i]]
				#
				self.dataset = clean_data
				self.data_labels = column_list
				self.data_btn["state"] = tk.NORMAL
				# Aggregate rows by group
				subsets = subset_by_groups(clean_data[0:self.n_dataset_columns], clean_data[self.n_dataset_columns])
				agg_error, agg_data = aggregate_groups(subsets, self.aggreg_var.get())
				if not agg_error:
					# Check for NaN
					if dataset_contains(agg_data, lambda x: type(x) == float and math.isnan(x)):
						warning("Aggregated data has NaN values; matrix cannot be produced.", {"parent": self.dlg})
					else:
						self.agg_data = sort_columns(agg_data)
						self.agg_data_groups = self.agg_data[0]
						self.agg_data_btn["state"] = tk.NORMAL

	def redraw(self):
		# (Re)draw the cosine similarity matrix
		if self.agg_data is not None and len(self.agg_data) > 1 and len(self.agg_data[0]) > 1:
			self.loading_dlg.display("Calculating similarity matrix")
			self.dlg.bind("<Alt-l>", self.set_labeling)
			#ncases = len(self.agg_data[self.n_dataset_columns])
			#inp_data = columns_to_rows(self.agg_data[0:self.n_dataset_columns])
			ncases = len(self.agg_data_groups)
			inp_data = columns_to_rows(self.agg_data[1:])
			cosmat = cosine_similarity(np.asarray(inp_data, float))
			caxes = self.plot.axes.matshow(cosmat, cmap="BrBG", vmin=-1.0, vmax=1.0)
			self.plot.figure.colorbar(caxes)
			self.plot.axes.set_xticks(range(ncases), labels=self.agg_data_groups, rotation=25)
			self.plot.axes.set_yticks(range(ncases), labels=self.agg_data_groups)
			if self.show_labels:
				for i in range(ncases):
					for j in range(ncases):
						v = cosmat[i,j]
						c = "white" if abs(v) > 0.40 else "black"
						self.plot.axes.text(j, i, f"{cosmat[i,j]:.2f}", ha="center", va="center", color=c)
			self.plot.draw()
			# Convert the upper right corner of the matrix into a table for display.
			self.cos_table_labels = [f"{self.groupby_var.get()} 1", f"{self.groupby_var.get()} 2", "Cosine similarity"]
			self.cos_table = [[], [], []]
			groups = sorted_numstrs(list(set(self.dataset[self.n_dataset_columns])))
			for i in range(len(groups)-1):
				for j in range(i+1,len(groups)):
					self.cos_table[0].append(groups[i])
					self.cos_table[1].append(groups[j])
					self.cos_table[2].append(cosmat[i][j])
			self.sim_data_btn["state"] = tk.NORMAL
			self.loading_dlg.hide()

	def set_autoupdate(self, *args):
		self.auto_update = self.autoupdate_var.get() == "1"
		if self.auto_update:
			self.q_redraw(())
	def set_labeling(self, *args):
		self.show_labels = not self.show_labels
		self.q_redraw()
	def do_close(self, *args):
		self.parent.remove_cosmat(self)
		super().do_cancel(args)



class ANOVADialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.alpha = 0.45
		self.theilsen = False
		self.plot_title = None
		super().__init__("Analysis of Variance (ANOVA)", "Select numeric and grouping variables",
				help_url="https://mapdata.readthedocs.io/en/latest/anova.html")
		# Data
		self.dataset = None
		self.data_labels = None
		self.group_data = None
		self.group_labels = None
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.categ_columns = sorted([c[0] for c in self.column_specs if c[1] in ("string", "boolean", "date")])
		self.output_colhdrs = ["Test", "Statistic", "Value", "p value"]
		self.dist_colhdrs = ["Group", "N", "Mean", "Median", "Variance", "Skewness", "Kurtosis", "Anderson-Darling p", "Lilliefors p", "Omnibus normality p"]
		self.statdata = []

		# Controls
		self.x_var = tk.StringVar(self.ctrl_frame, "")
		x_lbl = ttk.Label(self.ctrl_frame, text="X column:")
		x_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.x_sel = ttk.Combobox(self.ctrl_frame, state=tk.NORMAL, textvariable=self.x_var, values=self.numeric_columns, width=24)
		self.x_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.x_sel.bind("<<ComboboxSelected>>", self.q_recalc)

		self.xlog_var = tk.StringVar(self.ctrl_frame, "0")
		self.xlog_ck = ttk.Checkbutton(self.ctrl_frame, text="Log X", state=tk.NORMAL, command=self.q_recalc, variable=self.xlog_var,
				onvalue="1", offvalue="0")
		self.xlog_ck.grid(row=0, column=2, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.groupby_var = tk.StringVar(self.ctrl_frame, "")
		groupby_lbl = ttk.Label(self.ctrl_frame, text="Group by:")
		groupby_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.groupby_sel = ttk.Combobox(self.ctrl_frame, state=tk.NORMAL, textvariable=self.groupby_var, values=self.categ_columns, width=24)
		self.groupby_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.groupby_sel.bind("<<ComboboxSelected>>", self.q_recalc)

		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 2, 0, self.q_recalc, colspan=2)

		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(2, weight=1)

		# Frames for output tables
		# with tabbed Notebook for two different tables
		table_pages = ttk.Notebook(self.content_frame)
		table_pages.grid(row=0, column=0, sticky=tk.NSEW)
		table_pages.rowconfigure(0, weight=1)
		table_pages.columnconfigure(0, weight=1)
		# Frame for ANOVA statistics
		self.stat_frame = tk.Frame(table_pages)
		self.stat_frame.rowconfigure(0, weight=1)
		self.stat_frame.columnconfigure(0, weight=1)
		table_pages.add(self.stat_frame, text="ANOVA")
		# Frame for subgroup distribution statistics
		self.dist_frame = tk.Frame(table_pages)
		self.dist_frame.rowconfigure(0, weight=1)
		self.dist_frame.columnconfigure(0, weight=1)
		table_pages.add(self.dist_frame, text="Distributions")

		# Initialize output frames
		self.clear_output()

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.data_btn = new_data_button(self.dlg, self.btn_frame, 1, self.show_data)
		self.group_btn = new_button(self.btn_frame, "Group Data", 0, 2, self.show_group_data, padx=(3,3), state=tk.DISABLED)
		new_close_button(self.dlg, self.btn_frame, 3, self.do_close)
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(3, weight=1)

	def get_data(self):
		self.data_btn["state"] = tk.DISABLED
		self.xlog_ck["state"] = tk.NORMAL
		self.dataset = None
		self.data_labels = None
		self.group_data = None
		self.group_labels = None
		self.dlg.bind("<Alt-s>")
		self.dlg.bind("<Alt-g>")
		column_list = [self.x_var.get(), self.groupby_var.get()]
		# Get either only the selected data or all data.
		dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is None or len(dataset[0]) == 0:
			self.data_btn["state"] = tk.DISABLED
			self.group_btn["state"] = tk.DISABLED
		else:
			# Drop missing data
			dataset = clean_missing(dataset, [0,1])
			# Convert nominally numeric values to floats
			float_vals = [conv_float(v) for v in dataset[0]]
			# Log-transform data if specified.
			if self.xlog_ck["state"] != tk.DISABLED and self.xlog_var.get() == "1":
				log_data = logvector(float_vals)
				if log_data is None:
					warning_nolog(parent=self.dlg)
					self.xlog_var.set("0")
					self.xlog_ck["state"] = tk.DISABLED
				else:
					float_vals = log_data
			self.group_labels, self.group_data = spread_by_groups(dataset[1], float_vals)
			self.group_btn["state"] = tk.NORMAL
			self.dataset = [float_vals, dataset[1]]
			if self.xlog_ck["state"] != tk.DISABLED and self.xlog_var.get() == "1":
				self.data_labels = ["Log10 of %s" % column_list[0]]
			else:
				self.data_labels = [column_list[0]]
			self.data_labels.append(column_list[1])
			self.data_btn["state"] = tk.NORMAL
			self.dlg.bind("<Alt-s>", self.show_data)
			self.dlg.bind("<Alt-g>", self.show_group_data)

	def clear_output(self):
		for ctl in self.stat_frame.winfo_children():
			ctl.destroy()
		for ctl in self.dist_frame.winfo_children():
			ctl.destroy()
		clear_dlg_hotkeys(self.dlg)
		tframe, tdata = treeview_table(self.stat_frame, [], self.output_colhdrs)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		tframe, tdata = treeview_table(self.dist_frame, [], self.dist_colhdrs)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Source Data for ANOVA")

	def show_group_data(self, *args):
		if self.group_data is not None:
			show_columnar_table(self.dlg, "Grouped Data", "Data for each group:", self.group_data, self.group_labels, "Data for ANOVA")

	def q_recalc(self, get_data=True, *args):
		if self.x_var.get() != '' and self.groupby_var.get() != '':
			if get_data or self.dataset is None:
				self.get_data()
			if self.group_data is not None and len(self.group_data) > 1 and max([len(d) for d in self.group_data]) > 1:
				self.recalc()
			else:
				self.clear_output()

	def recalc(self):
		self.clear_output()
		self.statdata = []
		self.distdata = []
		if self.group_data is not None:
			group_lengths = [len(g) for g in self.group_data]
			# One-way ANOVA
			F, p = spstats.f_oneway(*self.group_data)
			self.statdata.append(["One-way ANOVA", "F", fp_display(F), fp_display(p)])
			# Kruskal-Wallace test
			H, p = spstats.kruskal(*self.group_data)
			self.statdata.append(["Kruskal-Wallis", "H", fp_display(H), fp_display(p)])
			# Alexander-Govern test
			if min(group_lengths) > 1:
				ag = spstats.alexandergovern(*self.group_data)
				A = ag.statistic
				p = ag.pvalue
				self.statdata.append(["Alexander-Govern", "A", fp_display(A), fp_display(p)])
			# Levene's test
			W, p = spstats.levene(*self.group_data, center="mean")
			self.statdata.append(["Levene", "W", fp_display(W), fp_display(p)])
			# Brown-Forsythe test
			F, p = spstats.levene(*self.group_data, center="median")
			self.statdata.append(["Brown-Forsythe", "F", fp_display(F), fp_display(p)])
			# Bartlett's test
			if min(group_lengths) > 1:
				try:
					X2, p = spstats.bartlett(*self.group_data)
				except:
					pass
				else:
					if not math.isnan(X2) and not math.isnan(p):
						self.statdata.append(["Bartlett", "X2", fp_display(X2), fp_display(p)])
			# Distribution statistics for each group
			for i in range(len(self.group_labels)):
				# Group, N, Mean, Median, Variance, Skewness, Kurtosis, Normality p]
				lbl = self.group_labels[i]
				d = self.group_data[i]
				dd = [lbl, group_lengths[i], fp_display(statistics.fmean(d))]
				dd.append(fp_display(statistics.median(d)))
				var = None
				try:
					var = statistics.variance(d)
					dd.append(fp_display(var))
				except:
					var = None
					dd.append("NC")
				if group_lengths[i] > 1:
					try:
						dd.extend([fp_display(spstats.skew(d)), fp_display(spstats.kurtosis(d))])
					except:
						dd.extend(["NC","NC"])
				else:
					dd.extend(["NC","NC"])
				# Normality tests
				if var is not None:
					da = np.array(d)
					try:
						adval, adpval = normal_ad(da)
						dd.append("%.2E" % adpval)
					except:
						dd.append("NC")
					try:
						lfval, lfpval = lilliefors(da, dist='norm')
						dd.append("%.2E" % lfpval)
					except:
						dd.append("NC")
					if len(self.group_data[i]) > 19:
						try:
							stat, p = spstats.normaltest(d)
						except:
							dd.append("NC")
						else:
							dd.append("%.2E" % p)
					else:
						dd.append("NC")
				else:
					dd.extend(["NC", "NC", "NC"])

				self.distdata.append(dd)
		if len(self.statdata) > 0:
			tframe, tdata = treeview_table(self.stat_frame, self.statdata, self.output_colhdrs)
			tframe.grid(row=0, column=0, stick=tk.NSEW)
			tframe, tdata = treeview_table(self.dist_frame, self.distdata, self.dist_colhdrs)
			tframe.grid(row=0, column=0, stick=tk.NSEW)
			self.dlg.bind("<Control-s>", self.save_table)
			self.dlg.minsize(width=700, height=400)
			self.dlg.bind("<Control-s>", self.save_table)
			self.dlg.bind("<Control-z>", self.save_norm_table)

	def save_table(self, *args):
		export_data_table(self.output_colhdrs, self.statdata, sheetname="ANOVA statistsics")
	def save_norm_table(self, *args):
		export_data_table(self.dist_colhdrs, self.distdata, sheetname="Normality statistsics")
	def do_close(self, *args):
		self.parent.remove_anova(self)
		super().do_cancel(args)



class TSNEDialog(Dialog):
	def __init__(self, parent, column_specs, prohibited_columns):
		self.parent = parent
		self.column_specs = column_specs
		self.prohibited_columns = prohibited_columns
		self.string_columns = [c[0] for c in self.column_specs if c[1] == "string" and not (c[1] in prohibited_columns)]
		self.auto_update = True
		self.alpha = 0.45
		self.plot_title = None
		super().__init__("t-SNE Analysis",
				"Select three or more variables from the list below, and then use the 'Calculate' button.",
				help_url="https://mapdata.readthedocs.io/en/latest/tsne.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		self.alpha = 0.45
		self.tsne_coords = None
		self.tsne_labels = None
		# Hover annotation for scatter plots
		self.annot = None
		self.canvas_conn_id = None
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float")]
		self.numeric_columns.sort()
		self.categ_columns = [c[0] for c in self.column_specs if c[1] == "string"]
		self.categ_columns.sort()

		# Controls
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_recalc)

		# Dropdown for missing data treatment
		cleantype_frame = tk.Frame(self.ctrl_frame)
		cleantype_frame.grid(row=1, column=0, columnspan=2, sticky=tk.NSEW)
		self.cleantype_var = tk.StringVar(cleantype_frame, "Drop variables")
		ttk.Label(cleantype_frame, text="Remove missing values by:").grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		self.cleantype_sel = ttk.Combobox(cleantype_frame, state=tk.NORMAL, textvariable=self.cleantype_var, values=["Drop cases","Drop variables", "Set to zero"], width=14)
		self.cleantype_sel.bind('<<ComboboxSelected>>', self.q_recalc)
		self.cleantype_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		cleantype_frame.columnconfigure(1, weight=1)

		# Dropdown for standardization
		stdize_frame = tk.Frame(self.ctrl_frame)
		stdize_frame.grid(row=1, column=2, sticky=tk.NSEW)
		stdize_opts = ["", "Z score by variable", "L1 norm by case"]
		self.stdize_var = tk.StringVar(stdize_frame, stdize_opts[1])
		ttk.Label(stdize_frame, text="Standardization:").grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		self.stdize_sel = ttk.Combobox(stdize_frame, state=tk.NORMAL, textvariable=self.stdize_var, values=stdize_opts, width=20)
		self.stdize_sel.bind('<<ComboboxSelected>>', self.q_recalc)
		self.stdize_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		stdize_frame.columnconfigure(1, weight=1)

		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(1, weight=0)
		self.ctrl_frame.columnconfigure(2, weight=1)

		# PanedWindow within content_frame: left is input listbox and 'Calculate' button, right is plot and 'Data Table' button
		io_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		io_panes.grid(row=0, column=0, sticky=tk.NSEW)

		self.inp_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE, width=350)
		self.inp_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.inp_frame.rowconfigure(0, weight=1)
		self.inp_frame.columnconfigure(0, weight=1)
		io_panes.add(self.inp_frame, weight=1)

		self.output_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE, width=350)
		self.output_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.output_frame.rowconfigure(0, weight=1)
		self.output_frame.columnconfigure(0, weight=1)
		io_panes.add(self.output_frame, weight=4)

		# Variable selection
		# Add multi-select list of variables to the leftmost pane
		var_frame = tk.Frame(self.inp_frame)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		self.column_frame, self.column_table = treeview_table(self.inp_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_recalc)

		# Perplexity_selection
		inpbtn_frame = tk.Frame(self.inp_frame)
		inpbtn_frame.grid(row=1, column=0, sticky=tk.NSEW)
		inpbtn_frame.columnconfigure(1, weight=1)
		self.perp_var = tk.IntVar(inpbtn_frame, min(50, math.floor(self.parent.table_row_count/3)))
		self.perp_lbl = ttk.Label(inpbtn_frame, text="Perplexity:")
		self.perp_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.perp_sel = ttk.Combobox(inpbtn_frame, state=tk.NORMAL, textvariable=self.perp_var, values=[2,3,4,5,10,15,20,25,30,40,50], width=5)
		self.perp_sel.bind('<<ComboboxSelected>>', self.q_recalc)
		self.perp_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))

		# Distance metric
		ttk.Label(inpbtn_frame, text="Distance metric:").grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.metric_var = tk.StringVar(inpbtn_frame, 'euclidean')
		metric_sel = ttk.Combobox(inpbtn_frame, state="readonly", textvariable=self.metric_var, width=15,
				values=['braycurtis', 'canberra', 'chebyshev', 'cosine', 'correlation', 'euclidean', 'manhattan', 'minkowski'])
		metric_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))
		metric_sel.bind('<<ComboboxSelected>>', self.q_recalc)

		# Optional grouping
		self.groupby_var = tk.StringVar(self.inp_frame, "")
		groupby_lbl = ttk.Label(inpbtn_frame, text="Group by:")
		groupby_lbl.grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		groupby_sel = ttk.Combobox(inpbtn_frame, state="readonly", textvariable=self.groupby_var, width=24)
		groupby_sel["values"] = [''] + self.categ_columns
		groupby_sel.bind('<<ComboboxSelected>>', self.q_recalc)
		groupby_sel.grid(row=2, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))

		# 'Calculate' button
		self.calc_btn = new_button(inpbtn_frame, "Calculate", 3, 1, self.recalculate, tk.E, (3,6), (3,3), state=tk.DISABLED)

		# t-SNE output plot
		plot_frame = tk.Frame(self.output_frame)
		plot_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot = Plot(plot_frame, 3.5, 3.5)

		# t-SNE Data Table button frame
		dtbtn_frame = tk.Frame(self.output_frame)
		dtbtn_frame.pack(side=tk.BOTTOM, fill=tk.X, expand=0)
		self.dtable_btn = new_button(dtbtn_frame, "Data Table", 0, 0, self.tsne_table, tk.W, (6,3), (21,3), state=tk.DISABLED)

		# k-Means clustering options
		kmeans_lf = tk.LabelFrame(dtbtn_frame, text="k-Means clusters")
		kmeans_lf.grid(row=0, column=1, padx=(3,3), pady=(3,3))
		self.cluster_btn = new_button(kmeans_lf, "Create", 0, 0, self.kclusters, tk.W, (3,3), (2,2), state=tk.DISABLED)
		self.clust_save_btn = new_button(kmeans_lf, "Add Column", 0, 1, self.kcolumn, tk.W, (3,6), (2,2), state=tk.DISABLED)

		# Buttons
		self.data_btn = add_help_src_close_btns(self.dlg, self.btn_frame, self.do_help, self.show_data, self.do_close)

		# Initialize output frames
		self.clear_output()

	def q_recalc(self, *args):
		# At least three variables must be selected
		self.clear_output()
		curr_selections = self.column_table.selection()
		enable_if(self.calc_btn, len(curr_selections) > 2)

	def get_data(self):
		self.data_btn["state"] = tk.DISABLED
		self.dataset = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		# Record the number of data columns without labels.
		self.n_dataset_columns = len(column_list)
		# Get additional columns if specified
		if map_settings.label_col is not None:
			column_list.append(map_settings.label_col)
			self.label_col = len(column_list)
		else:
			self.label_col = None
		grpbyvar = self.groupby_var.get()
		if grpbyvar != '':
			column_list.append(grpbyvar)
			self.groupby_col = len(column_list)
		else:
			self.groupby_col = None
		# Get either only the selected data or all data.  Get row IDs in case k-means clustering results
		# are to be assigned to a column.
		dataset, self.data_rowids = chosen_dataset_and_ids(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is None or len(dataset[0]) < 6:
			self.dataset = None
			self.data_labels = None
			self.data_btn["state"] = tk.DISABLED
			self.calc_btn["state"] = tk.DISABLED
		else:
			# t-SNE with some metrics can operate on sparse arrays, so missing values may be converted to zero or removed.
			column_indexes = range(self.n_dataset_columns)
			if self.cleantype_var.get() == "Set to zero":
				for col in column_indexes:
					for row in range(len(dataset[0])):
						if dataset[col][row] is None:
							dataset[col][row] = 0
			elif self.cleantype_var.get() == "Drop variables":
				# Remove columns with missing data.  This may remove all columns.
				dataset, column_list, n_removed = clean_missing_columns(dataset, column_list, list(column_indexes))
				if n_removed > 0:
					self.n_dataset_columns = self.n_dataset_columns - n_removed
					if self.label_col is not None:
						self.label_col = self.label_col - n_removed
					if self.groupby_col is not None:
						self.groupby_col = len(column_list)
			elif self.cleantype_var.get() == "Drop rows" and self.n_dataset_columns > 0:
				# Remove rows with missing data.  This may remove all rows.
				dataset = clean_missing(dataset, column_indexes)

			if dataset is not None and len(dataset) > 0 and self.n_dataset_columns > 0:
				# Convert to floats for numeric data only
				for i in range(self.n_dataset_columns):
					if column_list[i] in self.numeric_columns:
						dataset[i] = [conv_float(v) for v in dataset[i]]

			if dataset is None or self.n_dataset_columns < 3 or len(dataset[0]) < 6:
				self.dataset = None
				self.data_labels = None
				self.data_btn["state"] = tk.DISABLED
				self.calc_btn["state"] = tk.DISABLED
				self.loading_dlg.hide()
				warning("The dataset must have at least 3 variables and 6 cases for t-SNE analysis", {'parent':self.dlg})
			else:
				self.dataset = sort_columns(dataset)
				self.data_labels = column_list
				self.data_btn["state"] = tk.NORMAL
				self.calc_btn["state"] = tk.NORMAL
				self.dtable_btn["state"] = tk.DISABLED

	def clear_output(self):
		clear_dlg_hotkeys(self.dlg)
		self.data_btn["state"] = tk.DISABLED
		self.dtable_btn["state"] = tk.DISABLED
		self.plot.clear()

	def recalculate(self):
		self.loading_dlg.display("Calculating t-SNE")
		self.clear_output()
		self.get_data()
		if self.dataset is not None and len(self.dataset[0]) > 5:
			# Check to see if this is a sparse matrix
			ds = copy.copy(self.dataset[0:self.n_dataset_columns])
			is_sparse = False
			for col in range(len(ds)):
				if any([x is None for x in ds[col]]):
					is_sparse = True
					break
			if is_sparse:
				# Replace all None with 0
				for col in range(len(ds)):
					ds[col] = [0.0 if x is None else x for x in ds[col]]
			dsa = np.array(columns_to_rows(ds), dtype=float)
			#if is_sparse:
			#	dsa = sparse.csr_matrix(dsa)
			#	tsne_init = 'random'
			#else:
			#	tsne_init = 'pca'
			tsne_init = 'pca'
			# Maybe transform
			if self.stdize_var.get() == "L1 norm by case":
				inp_data = Normalizer(norm='l1').fit_transform(dsa)
			elif self.stdize_var.get() == "Z score by variable":
				#if is_sparse:
				#	inp_data = StandardScaler(with_mean=False).fit_transform(dsa)
				#else:
				#	inp_data = StandardScaler().fit_transform(dsa)
				inp_data = StandardScaler().fit_transform(dsa)
			else:
				inp_data = dsa
			# Calculate t-SNE
			perp = self.perp_var.get()
			if perp > len(self.dataset[0]) - 1:
				perp = min(50, math.floor(len(self.dataset[0])/2))
				self.perp_var.set(perp)
			tsne = TSNE(n_components=2, metric=self.metric_var.get(), perplexity=perp, init=tsne_init)
			tsne_result = tsne.fit_transform(inp_data)
			# Get data for output table
			self.tsne_coords = [list(tsne_result[:,0]), list(tsne_result[:,1])]
			self.tsne_labels = ["Dimension 1", "Dimension 2"]
			self.dtable_btn["state"] = tk.NORMAL
			# Create a single or grouped scatter plot
			if self.groupby_col is None:
				splot = self.plot.axes.scatter(self.tsne_coords[0], self.tsne_coords[1], alpha=self.alpha)
				self.plot_handle_list = []
				self.plot_label_list = []
			else:
				groups = sorted_numstrs([str(x) for x in list(set(self.dataset[self.groupby_col-1]))])
				cmap = map_colors(groups)
				colors = [cmap[str(g)] for g in self.dataset[self.groupby_col-1]]
				splot = self.plot.axes.scatter(self.tsne_coords[0], self.tsne_coords[1], c=colors, alpha=self.alpha)
				# Custom legend handle for grouped scatter plot
				lbls = list(cmap.keys())
				if wrap_at_underscores:
					lbls = [l.replace("_", " ") for l in lbls]
				lbls = ["\n".join(textwrap.wrap(l, width=wrapwidth)) for l in lbls]
				colkey = list(cmap.values())
				symbs = [matplotlib.lines.Line2D([], [], marker='o', color=colkey[i], label=lbls[i], linestyle='None') for i in range(len(lbls))]
				self.plot_handle_list = symbs
				self.plot_label_list = lbls
			# Hover annotation
			if self.label_col is not None:
				hoverer = Hoverer(self.plot, self.dataset[self.label_col-1], splot)
				self.canvas_conn_id = self.plot.canvas.mpl_connect("motion_notify_event", lambda ev: hoverer.hover(ev))
			else:
				if self.canvas_conn_id is not None:
					self.plot.canvas.mpl_disconnect(self.canvas_conn_id)
			# Legend
			if len(self.plot_handle_list) > 0:
				self.plot.axes.legend(handles=self.plot_handle_list, labels=self.plot_label_list)

			self.plot.set_axis_labels(self.tsne_labels[0], self.tsne_labels[1])
			self.plot.draw()
			self.dlg.bind("<Alt-a>", self.set_alpha)
			self.cluster_btn["state"] = tk.NORMAL
			self.clust_save_btn["state"] = tk.DISABLED
			self.loading_dlg.hide()

	def kclusters(self):
		# Compute and display k-means clusters
		dlg = OneIntDialog(self.dlg, "k-Means Clusters", "Enter the number of clusters to identify", min_value=2, max_value=min(30, len(self.dataset[0])-1), initial=2)
		num_clusters = dlg.show()
		if num_clusters is not None:
			self.loading_dlg.display("Clustering t-SNE output")
			from sklearn.cluster import KMeans
			km = KMeans(n_clusters=num_clusters)
			km.fit(np.array(columns_to_rows([self.tsne_coords[0], self.tsne_coords[1]])))
			self.kmlabels = ["Cluster "+str(int(lbl)+1) for lbl in km.labels_]
			self.tsne_coords.append(self.kmlabels)
			self.tsne_labels.append("k-Means cluster")
			kgroups = sorted_numstrs([str(x) for x in list(set(self.kmlabels))])
			kcmap = map_colors(kgroups)
			kcolors = [kcmap[str(g)] for g in self.kmlabels]
			self.plot.axes.scatter(self.tsne_coords[0], self.tsne_coords[1], marker="v", s=15, c=kcolors)
			klbls = [str(k) for k in kcmap.keys()]
			kcolkey = list(kcmap.values())
			ksymbs = [matplotlib.lines.Line2D([], [], marker='v', color=kcolkey[i], label=klbls[i], linestyle='None') for i in range(len(klbls))]
			self.plot_handle_list.extend(ksymbs)
			self.plot_label_list.extend(klbls)
			try:
				self.plot.axes.get_legend().remove()
			except:
				pass
			self.plot.axes.legend(handles=self.plot_handle_list, labels=self.plot_label_list)
			self.plot.draw()
			# Only allow creation of one set of clusters for each t-SNE output.
			self.cluster_btn["state"] = tk.DISABLED
			self.clust_save_btn["state"] = tk.NORMAL
			self.loading_dlg.hide()

	def kcolumn(self):
		# Save k-means cluster output as a data table column.
		# Column selection
		dlg = CustomContentDialog(parent=self.dlg, title="Save Clusters", prompt="Save k-means clusters in a table column.")
		def ck_col_name(varname, ix, mode):
			if self.col_var.get() in self.prohibited_columns:
				self.col_var.set('')
				dlg.ok_btn["state"] = tk.DISABLED
			else:
				enable_if(dlg.ok_btn, self.col_var.get() != '')
		self.col_var = tk.StringVar(dlg.content_frame, "")
		self.col_var.trace_add("write", ck_col_name)
		col_lbl = ttk.Label(dlg.content_frame, justify=tk.RIGHT, text="New or existing column name:")
		col_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,6))
		self.col_sel = ttk.Combobox(dlg.content_frame, state=tk.NORMAL, textvariable=self.col_var, values=self.string_columns, width=24)
		self.col_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,6))
		self.col_sel.focus()
		ok = dlg.show()
		if ok:
			self.loading_dlg.display("Saving clusters")
			target_col = self.col_var.get()
			db_col_name = db_colnames([target_col])[0]
			# Update the database
			new_column = target_col not in self.parent.headers
			cur = data_db.cursor()
			if new_column:
				cur.execute(f"alter table mapdata add column {db_col_name} TEXT;")
			for i in range(len(self.kmlabels)):
				cur.execute(f"update mapdata set {db_col_name} = '{self.kmlabels[i]}' where treeviewid = '{self.data_rowids[i]}';")
			# Update the Treeview
			if new_column:
				add_tv_column(self.parent.tbl, target_col, self.data_rowids, self.kmlabels)
				self.parent.headers = self.parent.headers + [target_col]
				viscols = list(self.parent.tbl["displaycolumns"])
				if viscols[0] != '#all':
					viscols.append(target_col)
					self.parent.tbl["displaycolumns"] = viscols
			else:
				for rowid in self.parent.tbl.get_children():
					if rowid in self.data_rowids:
						clust_index = self.data_rowids.index(rowid)
						self.parent.tbl.set(rowid, column=target_col, value=self.kmlabels[clust_index])
					else:
						self.parent.tbl.set(rowid, column=target_col, value='')
			# Update the table specs
			dt = "string"
			coldata = self.parent.get_all_data([target_col])[0]
			missing = len([v for v in coldata if v is None or v == ''])
			unique = len(set([v for v in coldata if v is not None and v != '']))
			if new_column:
				self.parent.data_types.append([target_col, dt, missing, unique])
				self.string_columns.append(target_col)
			else:
				col_ix = self.parent.headers.index(target_col)
				self.parent.data_types[col_ix] = [target_col, dt, missing, unique]
			self.loading_dlg.hide()

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels, "Data for t-SNE Analysis")

	def tsne_table(self, *args):
		if self.tsne_coords is not None:
			dt = copy.copy(self.tsne_coords)
			dl = copy.copy(self.tsne_labels)
			if self.label_col is not None:
				dt.append(self.dataset[self.label_col-1])
				dl.append(self.data_labels[self.label_col-1])
			if self.groupby_col is not None:
				dt.append(self.dataset[self.groupby_col-1])
				dl.append(self.data_labels[self.groupby_col-1])
			show_columnar_table(self.dlg, "t-SNE Results", "Coordinates for the two t-SNE dimensions:", dt, dl, "t-SNE coordinates")

	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the transparency (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = min(1.0, max(new_alpha, 0.0))
			self.q_recalc()

	def do_close(self, *args):
		self.parent.remove_tsne(self)
		super().do_cancel(self)



class UMAPDialog(Dialog):
	def __init__(self, parent, column_specs, prohibited_columns):
		self.parent = parent
		self.column_specs = column_specs
		self.prohibited_columns = prohibited_columns
		self.string_columns = [c[0] for c in self.column_specs if c[1] == "string" and not (c[1] in prohibited_columns)]
		self.auto_update = True
		self.alpha = 0.45
		self.plot_title = None
		super().__init__("UMAP Analysis",
				"Select three or more variables from the list below, and then use the 'Calculate' button.",
				help_url="https://mapdata.readthedocs.io/en/latest/umap.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		self.alpha = 0.45
		self.umap_coords = None
		self.umap_labels = None
		# Hover annotation for scatter plots
		self.annot = None
		self.canvas_conn_id = None
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float")]
		self.numeric_columns.sort()
		self.categ_columns = [c[0] for c in self.column_specs if c[1] == "string"]
		self.categ_columns.sort()

		# Controls
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_recalc)

		# Dropdown for missing data treatment
		cleantype_frame = tk.Frame(self.ctrl_frame)
		cleantype_frame.grid(row=1, column=0, columnspan=2, sticky=tk.NSEW)
		self.cleantype_var = tk.StringVar(cleantype_frame, "Drop variables")
		ttk.Label(cleantype_frame, text="Remove missing values by:").grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		self.cleantype_sel = ttk.Combobox(cleantype_frame, state=tk.NORMAL, textvariable=self.cleantype_var, values=["Drop cases","Drop variables", "Set to zero"], width=14)
		self.cleantype_sel.bind('<<ComboboxSelected>>', self.q_recalc)
		self.cleantype_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		cleantype_frame.columnconfigure(1, weight=1)

		# Dropdown for standardization
		stdize_frame = tk.Frame(self.ctrl_frame)
		stdize_frame.grid(row=1, column=2, columnspan=2, sticky=tk.NSEW)
		stdize_opts = ["", "Z score by variable", "L1 norm by case"]
		self.stdize_var = tk.StringVar(stdize_frame, stdize_opts[1])
		ttk.Label(stdize_frame, text="Standardization:").grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		self.stdize_sel = ttk.Combobox(stdize_frame, state=tk.NORMAL, textvariable=self.stdize_var, values=stdize_opts, width=20)
		self.stdize_sel.bind('<<ComboboxSelected>>', self.q_recalc)
		self.stdize_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(2,2))
		stdize_frame.columnconfigure(1, weight=1)

		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(2, weight=1)

		# PanedWindow within content_frame: left is input listbox and 'Calculate' button, right is plot and 'Data Table' button
		io_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		io_panes.grid(row=0, column=0, sticky=tk.NSEW)

		self.inp_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE, width=350)
		self.inp_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.inp_frame.rowconfigure(0, weight=1)
		self.inp_frame.columnconfigure(0, weight=1)
		io_panes.add(self.inp_frame, weight=1)

		self.output_frame = tk.Frame(io_panes, borderwidth=3, relief=tk.RIDGE, width=350)
		self.output_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.output_frame.rowconfigure(0, weight=1)
		self.output_frame.columnconfigure(0, weight=1)
		io_panes.add(self.output_frame, weight=4)

		# Variable selection
		# Add multi-select list of variables to the leftmost pane
		var_frame = tk.Frame(self.inp_frame)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1)
		self.column_frame, self.column_table = treeview_table(self.inp_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=0, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_recalc)

		# UMAP parameter selection
		inpopt_frame = tk.Frame(self.inp_frame)
		inpopt_frame.grid(row=1, column=0, sticky=tk.NSEW)
		inpopt_frame.columnconfigure(1, weight=1)
		self.neighbors_var = tk.IntVar(inpopt_frame, min(15, math.floor(self.parent.table_row_count/4)))
		ttk.Label(inpopt_frame, text="Neighbors:").grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		neighbor_sel = ttk.Spinbox(inpopt_frame, textvariable=self.neighbors_var, from_=2, to=math.floor(self.parent.table_row_count/4), width=5,
				command=self.q_recalc)
		neighbor_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))
		ttk.Label(inpopt_frame, text="Minimum distance:").grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.mindist_var = tk.DoubleVar(inpopt_frame, 0.1)
		mindist_sel = ttk.Spinbox(inpopt_frame, textvariable=self.mindist_var, from_=0.0, to=0.99, increment=0.01, width=4, command=self.q_recalc)
		mindist_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))
		# n_components is fixed at 2, with no user input allowed.
		ttk.Label(inpopt_frame, text="Distance metric:").grid(row=2, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.metric_var = tk.StringVar(inpopt_frame, 'euclidean')
		metric_sel = ttk.Combobox(inpopt_frame, state="readonly", textvariable=self.metric_var, width=15,
				values=['braycurtis', 'canberra', 'chebyshev', 'cosine', 'correlation', 'euclidean', 'manhattan', 'minkowski'])
		metric_sel.grid(row=2, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))
		metric_sel.bind('<<ComboboxSelected>>', self.q_recalc)

		# Optional grouping
		self.groupby_var = tk.StringVar(self.inp_frame, "")
		ttk.Label(inpopt_frame, text="Group by:").grid(row=3, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		groupby_sel = ttk.Combobox(inpopt_frame, state="readonly", textvariable=self.groupby_var, width=24)
		groupby_sel["values"] = [''] + self.categ_columns
		groupby_sel.grid(row=3, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		groupby_sel.bind('<<ComboboxSelected>>', self.q_recalc)

		# 'Calculate' button
		self.calc_btn = new_button(inpopt_frame, "Calculate", 4, 1, self.recalculate, tk.E, (3,6), (3,3), state=tk.DISABLED)

		# UMAP output plot
		plot_frame = tk.Frame(self.output_frame)
		plot_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot = Plot(plot_frame, 3.5, 3.5)

		# UMAP Data Table button frame
		dtbtn_frame = tk.Frame(self.output_frame)
		dtbtn_frame.pack(side=tk.BOTTOM, fill=tk.X, expand=0)
		self.dtable_btn = new_button(dtbtn_frame, "Data Table", 0, 0, self.umap_table, tk.W, (6,3), (21,3), state=tk.DISABLED)

		# k-Means clustering options
		kmeans_lf = tk.LabelFrame(dtbtn_frame, text="k-Means clusters")
		kmeans_lf.grid(row=0, column=1, padx=(3,3), pady=(3,3))
		self.cluster_btn = new_button(kmeans_lf, "Create", 0, 0, self.kclusters, tk.W, (3,3), (2,2), state=tk.DISABLED)
		self.clust_save_btn = new_button(kmeans_lf, "Add Column", 0, 1, self.kcolumn, tk.W, (3,6), (2,2), state=tk.DISABLED)

		# Buttons
		self.data_btn = add_help_src_close_btns(self.dlg, self.btn_frame, self.do_help, self.show_data, self.do_close)

		# Initialize output frames
		self.clear_output()

	def q_recalc(self, *args):
		# At least three variables must be selected
		self.clear_output()
		curr_selections = self.column_table.selection()
		enable_if(self.calc_btn, len(curr_selections) > 2)

	def get_data(self):
		self.data_btn["state"] = tk.DISABLED
		self.dataset = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		# Record the number of data columns without labels.
		self.n_dataset_columns = len(column_list)
		# Get additional columns if specified
		if map_settings.label_col is not None:
			column_list.append(map_settings.label_col)
			self.label_col = len(column_list)
		else:
			self.label_col = None
		grpbyvar = self.groupby_var.get()
		if grpbyvar != '':
			column_list.append(grpbyvar)
			self.groupby_col = len(column_list)
		else:
			self.groupby_col = None
		# Get either only the selected data or all data.  Get row IDs in case k-means clustering results
		# are to be assigned to a column.
		dataset, self.data_rowids = chosen_dataset_and_ids(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is not None:
			# UMAP can operate on sparse arrays, so removal of missing data is optional
			column_indexes = range(self.n_dataset_columns)
			if self.cleantype_var.get() == "Set to zero":
				for col in column_indexes:
					for row in range(len(dataset[0])):
						if dataset[col][row] is None:
							dataset[col][row] = 0
			elif self.cleantype_var.get() == "Drop variables":
				# Remove columns with missing data.  This may remove all columns.
				dataset, column_list, n_removed = clean_missing_columns(dataset, column_list, list(column_indexes))
				if n_removed > 0:
					self.n_dataset_columns = self.n_dataset_columns - n_removed
					if self.label_col is not None:
						self.label_col = self.label_col - n_removed
					if self.groupby_col is not None:
						self.groupby_col = len(column_list)
			elif self.cleantype_var.get() == "Drop rows" and self.n_dataset_columns > 0:
				# Remove rows with missing data.  This may remove all rows.
				dataset = clean_missing(dataset, column_indexes)

		if dataset is not None and len(dataset) > 0 and self.n_dataset_columns > 0:
			# Convert to floats for numeric data only
			for i in range(self.n_dataset_columns):
				if column_list[i] in self.numeric_columns:
					dataset[i] = [conv_float(v) for v in dataset[i]]

		if dataset is None or self.n_dataset_columns < 3 or len(dataset[0]) < 6:
			self.dataset = None
			self.data_labels = None
			self.data_btn["state"] = tk.DISABLED
			self.calc_btn["state"] = tk.DISABLED
			self.loading_dlg.hide()
			warning("The dataset must have at least 3 variables and 6 cases for UMAP analysis", {'parent':self.dlg})
		else:
			self.dataset = sort_columns(dataset)
			self.data_labels = column_list
			self.data_btn["state"] = tk.NORMAL
			self.calc_btn["state"] = tk.NORMAL
			self.dtable_btn["state"] = tk.DISABLED

	def clear_output(self):
		clear_dlg_hotkeys(self.dlg)
		self.data_btn["state"] = tk.DISABLED
		self.dtable_btn["state"] = tk.DISABLED
		self.plot.clear()

	def recalculate(self):
		self.loading_dlg.display("Calculating UMAP")
		import umap
		from sklearn.preprocessing import Normalizer, StandardScaler
		self.clear_output()
		self.get_data()
		if self.dataset is not None and len(self.dataset[0]) > 5:
			# Check to see if this is a sparse matrix
			ds = copy.copy(self.dataset[0:self.n_dataset_columns])
			is_sparse = False
			for col in range(len(ds)):
				if any([x is None for x in ds[col]]):
					is_sparse = True
					break
			if is_sparse:
				# Replace all None with 0
				for col in range(len(ds)):
					ds[col] = [0.0 if x is None else x for x in ds[col]]
			dsa = np.array(columns_to_rows(ds), dtype=float)
			if is_sparse:
				dsa = sparse.csr_matrix(dsa)
			# Maybe transform
			if self.stdize_var.get() == "L1 norm by case":
				inp_data = Normalizer(norm='l1').fit_transform(dsa)
			elif self.stdize_var.get() == "Z score by variable":
				if is_sparse:
					inp_data = StandardScaler(with_mean=False).fit_transform(dsa)
				else:
					inp_data = StandardScaler().fit_transform(dsa)
			else:
				inp_data = dsa
			reducer = umap.UMAP(n_neighbors=min(self.neighbors_var.get(), len(self.dataset[0])-1),
					min_dist=self.mindist_var.get(), n_components=2, metric=self.metric_var.get())
			umap_result = reducer.fit_transform(inp_data)
			self.umap_coords = [list(umap_result[:,0]), list(umap_result[:,1])]
			self.umap_labels = ["Dimension 1", "Dimension 2"]
			self.dtable_btn["state"] = tk.NORMAL
			# Single or grouped scatter plot
			if self.groupby_col is None:
				splot = self.plot.axes.scatter(self.umap_coords[0], self.umap_coords[1], alpha=self.alpha)
				self.plot_handle_list = []
				self.plot_label_list = []
			else:
				groups = sorted_numstrs([str(x) for x in list(set(self.dataset[self.groupby_col-1]))])
				cmap = map_colors(groups)
				colors = [cmap[str(g)] for g in self.dataset[self.groupby_col-1]]
				splot = self.plot.axes.scatter(self.umap_coords[0], self.umap_coords[1], c=colors, alpha=self.alpha)
				# Custom legend handle for grouped scatter plot
				lbls = list(cmap.keys())
				if wrap_at_underscores:
					lbls = [l.replace("_", " ") for l in lbls]
				lbls = ["\n".join(textwrap.wrap(l, width=wrapwidth)) for l in lbls]
				colkey = list(cmap.values())
				symbs = [matplotlib.lines.Line2D([], [], marker='o', color=colkey[i], label=lbls[i], linestyle='None') for i in range(len(lbls))]
				self.plot_handle_list = symbs
				self.plot_label_list = lbls
			# Hover annotation
			if self.label_col is not None:
				hoverer = Hoverer(self.plot, self.dataset[self.label_col-1], splot)
				self.canvas_conn_id = self.plot.canvas.mpl_connect("motion_notify_event", lambda ev: hoverer.hover(ev))
			else:
				if self.canvas_conn_id is not None:
					self.plot.canvas.mpl_disconnect(self.canvas_conn_id)
			# Legend
			if len(self.plot_handle_list) > 0:
				self.plot.axes.legend(handles=self.plot_handle_list, labels=self.plot_label_list)

			self.plot.set_axis_labels(self.umap_labels[0], self.umap_labels[1])
			self.plot.draw()
			self.dlg.bind("<Alt-a>", self.set_alpha)
			self.cluster_btn["state"] = tk.NORMAL
			self.clust_save_btn["state"] = tk.DISABLED
			self.loading_dlg.hide()

	def kclusters(self):
		# Compute and display k-means clusters
		dlg = OneIntDialog(self.dlg, "k-Means Clusters", "Enter the number of clusters to identify", min_value=2, max_value=min(30, len(self.dataset[0])-1), initial=2)
		num_clusters = dlg.show()
		if num_clusters is not None:
			self.loading_dlg.display("Clustering UMAP output")
			from sklearn.cluster import KMeans
			km = KMeans(n_clusters=num_clusters)
			km.fit(np.array(columns_to_rows([self.umap_coords[0], self.umap_coords[1]])))
			self.kmlabels = ["Cluster "+str(int(lbl)+1) for lbl in km.labels_]
			self.umap_coords.append(self.kmlabels)
			self.umap_labels.append("k-Means cluster")
			kgroups = sorted_numstrs([str(x) for x in list(set(self.kmlabels))])
			kcmap = map_colors(kgroups)
			kcolors = [kcmap[str(g)] for g in self.kmlabels]
			self.plot.axes.scatter(self.umap_coords[0], self.umap_coords[1], marker="v", s=15, c=kcolors)
			klbls = [str(k) for k in kcmap.keys()]
			kcolkey = list(kcmap.values())
			ksymbs = [matplotlib.lines.Line2D([], [], marker='v', color=kcolkey[i], label=klbls[i], linestyle='None') for i in range(len(klbls))]
			self.plot_handle_list.extend(ksymbs)
			self.plot_label_list.extend(klbls)
			try:
				self.plot.axes.get_legend().remove()
			except:
				pass
			self.plot.axes.legend(handles=self.plot_handle_list, labels=self.plot_label_list)
			self.plot.draw()
			# Only allow creation of one set of clusters for each UMAP output.
			self.cluster_btn["state"] = tk.DISABLED
			self.clust_save_btn["state"] = tk.NORMAL
			self.loading_dlg.hide()

	def kcolumn(self):
		# Save k-means cluster output as a data table column.
		# Column selection
		dlg = CustomContentDialog(parent=self.dlg, title="Save Clusters", prompt="Save k-means clusters in a table column.")
		def ck_col_name(varname, ix, mode):
			if self.col_var.get() in self.prohibited_columns:
				self.col_var.set('')
				dlg.ok_btn["state"] = tk.DISABLED
			else:
				if self.col_var.get() == '':
					dlg.ok_btn["state"] = tk.DISABLED
				else:
					dlg.ok_btn["state"] = tk.NORMAL
		self.col_var = tk.StringVar(dlg.content_frame, "")
		self.col_var.trace_add("write", ck_col_name)
		col_lbl = ttk.Label(dlg.content_frame, justify=tk.RIGHT, text="New or existing column name:")
		col_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,6))
		self.col_sel = ttk.Combobox(dlg.content_frame, state=tk.NORMAL, textvariable=self.col_var, values=self.string_columns, width=24)
		self.col_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,6))
		self.col_sel.focus()
		ok = dlg.show()
		if ok:
			self.loading_dlg.display("Saving clusters")
			target_col = self.col_var.get()
			db_col_name = db_colnames([target_col])[0]
			# Update the database
			new_column = target_col not in self.parent.headers
			cur = data_db.cursor()
			if new_column:
				cur.execute(f"alter table mapdata add column {db_col_name} TEXT;")
			for i in range(len(self.data_rowids)):
				cur.execute(f"update mapdata set {db_col_name} = '{self.kmlabels[i]}' where treeviewid = '{self.data_rowids[i]}';")
			# Update the Treeview
			if new_column:
				add_tv_column(self.parent.tbl, target_col, self.data_rowids, self.kmlabels)
				self.parent.headers = self.parent.headers + [target_col]
				viscols = list(self.parent.tbl["displaycolumns"])
				if viscols[0] != '#all':
					viscols.append(target_col)
					self.parent.tbl["displaycolumns"] = viscols
			else:
				for rowid in self.parent.tbl.get_children():
					if rowid in self.data_rowids:
						clust_index = self.data_rowids.index(rowid)
						self.parent.tbl.set(rowid, column=target_col, value=self.kmlabels[clust_index])
					else:
						self.parent.tbl.set(rowid, column=target_col, value='')
			# Update the table specs
			dt = "string"
			coldata = self.parent.get_all_data([target_col])[0]
			missing = len([v for v in coldata if v is None or v == ''])
			unique = len(set([v for v in coldata if v is not None and v != '']))
			if new_column:
				self.parent.data_types.append([target_col, dt, missing, unique])
				self.string_columns.append(target_col)
			else:
				col_ix = self.parent.headers.index(target_col)
				self.parent.data_types[col_ix] = [target_col, dt, missing, unique]
			self.loading_dlg.hide()

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels, "Data for UMAP Analysis")

	def umap_table(self, *args):
		if self.umap_coords is not None:
			dt = copy.copy(self.umap_coords)
			dl = copy.copy(self.umap_labels)
			if self.label_col is not None:
				dt.append(self.dataset[self.label_col-1])
				dl.append(self.data_labels[self.label_col-1])
			if self.groupby_col is not None:
				dt.append(self.dataset[self.groupby_col-1])
				dl.append(self.data_labels[self.groupby_col-1])
			show_columnar_table(self.dlg, "UMAP Results", "Coordinates for the two UMAP dimensions:", dt, dl, "UMAP coordinates")

	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the transparency (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = min(1.0, max(new_alpha, 0.0))
			self.q_recalc()

	def do_close(self, *args):
		self.parent.remove_umap(self)
		super().do_cancel(args)



class FitDistDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.plot_title = None
		super().__init__("Distribution Fitting", "Select a variable and a distribution type to fit.",
				help_url="https://mapdata.readthedocs.io/en/latest/fitunivardist.html")
		self.bins = None
		# scipy.stats distributions, by support and type.
		# The dictionary itemss are a tuple consisting of:
		#	* the scipy distribution object
		#	* a dictionary of arguments to be passed to the 'fit()' function.
		#	* the user names of the return values of the 'fit()' function.
		#	* the 'goodness_of_fit()' argument names of the return values of the 'fit()' function.
		#-- Continuous, -inf to inf:
		self.dist_cont = {
				'Laplace': (spstats.laplace, {}, ('Location', 'Scale'), ('loc', 'scale')),
				'Logistic': (spstats.logistic, {}, ('Location', 'Scale'), ('loc', 'scale')),
				'Normal': (spstats.norm, {}, ('Mean', 'Std. deviation'), ('loc', 'scale')),
				# Shifted log-logistic fitting triggers an error in scipy.
				#'Shifted log-logistic': (spstats.fisk, {}, ('Shape', 'Location', 'Scale'), ('shape', 'loc', 'scale')),
				'von Mises': (spstats.vonmises, {}, ('Kappa', 'Location', 'Scale'), ('kappa', 'loc', 'scale'))}
		#-- Continuous, 0 to inf only:
		self.dist_cont_pos = {
				'Exponential': (spstats.expon, {}, ('Location', 'Scale'), ('loc', 'scale')),
				'Gamma': (spstats.gamma, {'floc':0}, ('Shape', None, 'Scale'), ('a', 'loc', 'scale')),
				'Gompertz': (spstats.gompertz, {'floc':0}, ('Shape', None, 'Scale'), ('c', 'loc', 'scale')),
				# Log-logistic fitting triggers an error in scipy.
				#'Log-logistic': (spstats.fisk, {'floc':0}, ('Shape', None, 'Scale'), ('shape', 'loc', 'scale')),
				'Lognormal': (spstats.lognorm, {'floc':0}, ('Log std. dev.', None, 'Log mean'), ('s', 'loc', 'scale')),
				'Pareto': (spstats.pareto, {'floc':0}, ('Shape', None, 'Scale'), ('b', 'loc', 'scale')),
				'Rayleigh': (spstats.rayleigh, {}, ('Location', 'Scale'), ('loc', 'scale'))}
		#-- Continuous, 0 to 1 only:
		self.dist_cont_01 = {
				'Beta': (spstats.beta, {}, ('Shape 1', 'Shape 2', 'Location', 'Scale'), ('a', 'b', 'loc', 'scale')),
				'Uniform': (spstats.uniform, {}, ('Location', 'Scale'), ('loc', 'scale'))}
		self.all_dist = {**self.dist_cont, **self.dist_cont_pos, **self.dist_cont_01}
		#
		clear_dlg_hotkeys(self.dlg)
		self.xlabel = None
		self.ylabel = None
		# Data
		self.dataset = None
		self.data_labels = None
		self.numeric_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float")]
		self.numeric_columns.sort()
		self.output_columns = ["Statistic", "Value"]
		self.statdata = []

		# Controls
		self.x_var = tk.StringVar(self.ctrl_frame, "")
		x_lbl = ttk.Label(self.ctrl_frame, text="X column:")
		x_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.x_sel = ttk.Combobox(self.ctrl_frame, state=tk.NORMAL, textvariable=self.x_var, values=self.numeric_columns, width=24)
		self.x_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.x_sel.bind("<<ComboboxSelected>>", self.set_dist)

		self.xlog_var = tk.StringVar(self.ctrl_frame, "0")
		self.xlog_ck = ttk.Checkbutton(self.ctrl_frame, text="Log X", state=tk.NORMAL, command=self.set_dist, variable=self.xlog_var,
				onvalue="1", offvalue="0")
		self.xlog_ck.grid(row=0, column=2, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 3, self.set_dist)

		self.dist_var = tk.StringVar(self.ctrl_frame, "")
		dist_lbl = ttk.Label(self.ctrl_frame, text="Distribution:")
		dist_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.dist_sel = ttk.Combobox(self.ctrl_frame, state=tk.DISABLED, textvariable=self.dist_var, values=[], width=22)
		self.dist_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))
		self.dist_sel.bind("<<ComboboxSelected>>", self.refit)

		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(3, weight=1)

		# PanedWindow within content_frame: left is plot, right is table
		output_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		output_panes.grid(row=0, column=0, sticky=tk.NSEW)

		self.out_plt_frm = tk.Frame(output_panes, borderwidth=3, relief=tk.RIDGE, width=350)
		self.out_plt_frm.grid(row=0, column=0, sticky=tk.NSEW)
		self.out_plt_frm.rowconfigure(0, weight=1)
		self.out_plt_frm.columnconfigure(0, weight=1)
		output_panes.add(self.out_plt_frm, weight=1)

		self.out_tbl_frm = tk.Frame(output_panes, borderwidth=3, relief=tk.RIDGE, width=350)
		self.out_tbl_frm.grid(row=0, column=1, sticky=tk.NSEW)
		self.out_tbl_frm.rowconfigure(0, weight=1)
		self.out_tbl_frm.columnconfigure(0, weight=1, minsize=400)
		output_panes.add(self.out_tbl_frm, weight=1)

		# Data plot
		plot_frame = tk.Frame(self.out_plt_frm)
		plot_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot = Plot(plot_frame, 3.5, 3.5, layout="tight")

		# Initialize output frames
		self.clear_output()

		# Buttons
		self.data_btn = add_help_src_close_btns(self.dlg, self.btn_frame, self.do_help, self.show_data, self.do_close)

	def q_recalc(self):
		self.get_data()
		if self.dataset is not None:
			self.set_dist()

	def get_data(self):
		self.dataset = None
		self.data_labels = None
		self.dist_sel["state"] = tk.DISABLED
		self.data_btn["state"] = tk.DISABLED
		self.xlog_ck["state"] = tk.NORMAL
		self.dist_var.set('')
		self.statdata = []
		if self.x_var.get() != '':
			column_list = [self.x_var.get()]
			# Get either only the selected data or all data.
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
				self.dist_sel["state"] = tk.DISABLED
				self.data_btn["state"] = tk.DISABLED
			else:
				# Remove missing data
				clean_data = clean_missing(dataset, list(range(len(dataset))))
				# Convert to floats for numeric data only
				for i in range(len(clean_data)):
					if column_list[i] in self.numeric_columns:
						clean_data[i] = [conv_float(v) for v in clean_data[i]]
				self.data_labels = [column_list[0]]
				# Log-transform data if specified.
				if self.xlog_ck["state"] != tk.DISABLED and self.xlog_var.get() == "1" and self.x_var.get() in self.numeric_columns:
					log_x = logvector(clean_data[0])
					if log_x is not None:
						clean_data[0] = log_x
						self.data_labels = ["Log10 of %s" % column_list[0]]
					else:
						self.xlog_var.set("0")
						self.xlog_ck["state"] = tk.DISABLED
						warning_nolog(parent=self.dlg)
				#
				self.dataset = sort_columns(clean_data)
				self.dist_sel["state"] = "readonly"
				self.data_btn["state"] = tk.NORMAL
				#
				xpfx = "" if self.xlog_var.get() == '0' else "Log10 of "
				self.statdata.append(["X variable", self.data_labels[0]])
				self.statdata.append(["N", len(self.dataset[0])])

	def clear_output(self):
		clear_dlg_hotkeys(self.dlg)
		for ctl in self.out_tbl_frm.winfo_children():
			ctl.destroy()
		tframe, tdata = treeview_table(self.out_tbl_frm, [], ["Statistic", "Value"])
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.plot.clear()

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for distribution fitting")

	def set_dist(self, *args):
		# Set the allowable values for the distribution type based on the data range.
		self.clear_output()
		if self.x_var.get() != '':
			self.get_data()

			# Plot the histogram
			if self.bins is None:
				self.bins = doane_bins(self.dataset[0])
			self.hist_val, hist_bins, patches = self.plot.axes.hist(self.dataset[0], bins=self.bins, density=True)
			hist_ofs = (self.dataset[0][-1] - self.dataset[0][0])/self.bins/2
			self.hist_centers = [hist_bins[i] + hist_ofs for i in range(len(hist_bins)-1)]
			self.plot.set_axis_labels(self.xlabel or self.data_labels[0], "Probability Density")
			self.dlg.bind("<Alt-b>", self.set_bins)
			self.plot.draw()

			self.dlg.bind("<Alt-t>", self.set_title)
			self.dlg.bind("<Alt-x>", self.set_xlabel)
			self.dlg.bind("<Alt-y>", self.set_ylabel)

			# Set the allowable distributions
			if len(self.dataset[0]) < 5:
				self.dist_var.set("")
				self.dist_sel["state"] = tk.DISABLED
			else:
				self.dist_sel["state"] = "readonly"
				if self.dataset is not None:
					xmin = self.dataset[0][0]
					xmax = self.dataset[0][-1]
					if xmin < 0:
						dist = self.dist_cont
					elif xmax <= 1.0:
						dist = {**self.dist_cont_pos, **self.dist_cont, **self.dist_cont_01}
					else:
						dist = {**self.dist_cont_pos, **self.dist_cont}
				if self.dist_var.get() not in dist:
					self.dist_var.set("")
				self.dist_sel["values"] = list(dist.keys())

				# Maybe fit the distribution also
				if self.dist_var.get() != '':
					self.refit()
		else:
			self.clear_output()

	def refit(self, *args):
		running_dlg = LoadingDialog(self.dlg)
		running_dlg.display("Fitting data")
		fitter, fitterargs, usernames, argnames = self.all_dist[self.dist_sel.get()]
		try:
			fitres = fitter.fit(self.dataset[0], **fitterargs)
		except:
			running_dlg.hide()
			warning("Data fitting unsuccessful", {"parent": self.dlg})
		else:
			N = len(self.dataset[0])
			self.statdata.append(["\u2015"*15, ""])
			self.statdata.append(["Distribution", self.dist_var.get()])
			for i, name in enumerate(usernames):
				if name is not None:
					self.statdata.append([name, fp_display(fitres[i])])
			# Add goodness-of-fit statistics to output table
			fit_pdf = fitter.pdf(self.dataset[0], *fitres) * N
			sse = sum([(fit_pdf[i] - self.dataset[0][i])**2 for i in range(N)])
			self.statdata.append(["SSE", fp_display(sse, figs=5)])
			self.statdata.append(["RMSE", fp_display(math.sqrt(sse/N))])
			gf_args = dict(zip(argnames, fitres))
			gf_succeeded = False
			gf_statname = None
			for stat in (("ad", "Anderson-Darling"), ("ks", "Kolmogorov-Smirnov"), ("cvm", "Cramer-von Mises")):
				try:
					gf_res = spstats.goodness_of_fit(fitter, self.dataset[0], fit_params=gf_args, statistic=stat[0], n_mc_samples=1000)
				except:
					pass
				else:
					if not math.isnan(gf_res.statistic):
						gf_succeeded = True
						gf_statname = stat[1]
						break
			if gf_succeeded:
				self.statdata.append([gf_statname + " statistic", fp_display(gf_res.statistic)])
				self.statdata.append([gf_statname + " p value", fp_display(gf_res.pvalue)])
				loglik = fitter.logpdf(self.dataset[0], *fitres).sum()
				aic = 2*len(fitres) - 2*loglik
				self.statdata.append(["AIC", fp_display(aic)])
			# Display table of statistics
			tframe, tdata = treeview_table(self.out_tbl_frm, self.statdata, self.output_columns)
			tframe.grid(row=0, column=0, stick=tk.NSEW)
			self.dlg.bind("<Control-s>", self.save_table)

			# Display PDF of fitted distribution on the plot
			fit_x = np.linspace(self.dataset[0][0], self.dataset[0][-1], 100)
			fit_pdf = fitter.pdf(fit_x, *fitres)
			self.plot.axes.plot(fit_x, fit_pdf, lw=2, label=self.dist_var.get())
			self.plot.axes.legend()
			self.plot.draw()

			self.dlg.bind("<Alt-t>", self.set_title)
			self.dlg.bind("<Alt-x>", self.set_xlabel)
			self.dlg.bind("<Alt-y>", self.set_ylabel)
			running_dlg.hide()

	def set_title(self, *args):
		dlg = OneEntryDialog(self.dlg, "Plot Title", "Enter a title for the plot:", init_value=self.plot_title, nullable=True)
		self.plot_title = dlg.show()
		self.plot.axes.set_title(self.plot_title)
		self.plot.draw()
	def set_xlabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "X-Axis Label", "Enter the X-axis label:", init_value=self.xlabel, nullable=True)
		self.xlabel = dlg.show()
		self.plot.axes.set_xlabel(self.xlabel or self.x_var.get())
		self.plot.draw()
	def set_ylabel(self, *args):
		dlg = OneEntryDialog(self.dlg, "Y-Axis Label", "Enter the Y-axis label:", init_value=self.ylabel, nullable=True)
		self.ylabel = dlg.show()
		self.plot.axes.set_ylabel(self.ylabel or "Probability Density")
		self.plot.draw()
	def set_bins(self, *args):
		dlg = OneIntDialog(self.dlg, "Data Bins", "Enter the number of bins to be used for the histogram", min_value=2, max_value=100, initial=self.bins)
		num_bins = dlg.show()
		if num_bins is not None:
			self.bins = num_bins
			self.q_recalc()
	def save_table(self, *args):
		export_data_table(self.output_columns, self.statdata, sheetname="Bivariate statistsics")
	def do_close(self, *args):
		self.parent.remove_fitdist(self)
		super().do_cancel(args)



class ContTableDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		self.plot_title = None
		super().__init__("Contingency Table",
				"Select row and column variables. Conventionally, the row represents the explanatory or treatment variable and the column represents the response, condition, or effect variable.  For each, the range of positive responses must be specified.",
				help_url="https://mapdata.readthedocs.io/en/latest/conttable.html")
		# Data
		self.dataset = None
		self.data_labels = None
		self.selectable_columns = [c[0] for c in self.column_specs if c[1] in ("int", "float", "string", "date", "boolean")]
		self.selectable_columns.sort()
		self.categ_columns = [c[0] for c in self.column_specs if c[1] in ("string", "date", "boolean")]
		self.output_columns = ["Statistic", "Value"]
		self.statdata = []

		# Remove the ctrl_frame and content_frame widgets because this dialog is highly customized.
		self.ctrl_frame.destroy()
		self.content_frame.destroy()
		self.dlg.rowconfigure(2, weight=0)

		# Two horizontal frames, 'input' for the selected and freeze checkboxes and the row var and column var
		# specifications, and 'content' for the stats and table graphic.
		# These two frames are embedded in io_frame.
		io_frame = tk.Frame(self.dlg)
		io_frame.grid(row=1, column=0, sticky=tk.NSEW)
		self.dlg.rowconfigure(1, weight=1)
		io_frame.rowconfigure(0, weight=1)
		# Only the second column can grow horizontally
		io_frame.columnconfigure(1, weight=1)

		input_frame = tk.Frame(io_frame)
		input_frame.grid(row=0, column=0, sticky=tk.NSEW)
		input_frame.rowconfigure(1, weight=1)
		input_frame.columnconfigure(0, weight=1)
		select_frame = tk.Frame(input_frame)
		select_frame.grid(row=0, column=0, columnspan=2, sticky=tk.N+tk.EW)
		rowvar_frame = tk.Frame(input_frame, borderwidth=1)
		rowvar_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(6,6), pady=(3,6))
		colvar_frame = tk.Frame(input_frame, borderwidth=1)
		colvar_frame.grid(row=1, column=1, sticky=tk.NSEW, padx=(6,6), pady=(3,6))

		self.content_frame = tk.Frame(io_frame, borderwidth=1)
		self.content_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.content_frame.rowconfigure(0, weight=1)
		self.content_frame.columnconfigure(0, weight=1)

		# The rowvar and colvar frames have a label on row 0, a variable name selector on row 1,
		# another label on row 2, and a threshold selector on row 3.  Because the threshold selector
		# may be a list box, row3, and only row3, is expandable.
		rowvar_frame.rowconfigure(3, weight=1)
		colvar_frame.rowconfigure(3, weight=1)

		# Selected and auto-update checkboxes
		self.sel_only_var, self.sel_only_ck = add_sel_only(select_frame, 0, 0, self.q_recalc)

		self.autoupdate_var = add_autoupdate(select_frame, 0, 1, self.set_autoupdate)

		# Row and column variables
		num_validator = self.dlg.register(self.validate_num_entry)
		ttk.Label(rowvar_frame, text="Row variable (R)").grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,1))
		ttk.Label(colvar_frame, text="Column variable (C)").grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,1))
		self.row_var = tk.StringVar(rowvar_frame)
		self.row_sel = ttk.Combobox(rowvar_frame, state="readonly", textvariable=self.row_var, values=self.selectable_columns, width=24)
		self.row_sel.grid(row=1, column=0, columnspan=3, sticky=tk.N+tk.EW, padx=(3,3), pady=(3,3))
		self.row_sel.bind("<<ComboboxSelected>>", self.ck_rowvar)
		self.col_var = tk.StringVar(colvar_frame)
		self.col_sel = ttk.Combobox(colvar_frame, state="readonly", textvariable=self.col_var, values=self.selectable_columns, width=24)
		self.col_sel.grid(row=1, column=0, columnspan=3, sticky=tk.N+tk.EW, padx=(3,3), pady=(3,3))
		self.col_sel.bind("<<ComboboxSelected>>", self.ck_colvar)

		self.row_thresh_lbl = ttk.Label(rowvar_frame, text="Threshold for\npositive response")
		self.row_thresh_lbl.grid(row=2, column=0, sticky=tk.N+tk.EW)
		self.col_thresh_lbl = ttk.Label(colvar_frame, text="Threshold for\npositive response")
		self.col_thresh_lbl.grid(row=2, column=0, sticky=tk.N+tk.EW)

		# The threshold for positive responses is either a numeric variable specified in an Entry widget
		# or a selected list of categorical values specified in a Listbox widget.  The Entry widget is
		# initially displayed, but disabled.  The 'ck_rowvar' and 'ck_colvar' methods hide and re-grid
		# these widgets as necessary.
		gt_lt_list = ['>', '<']
		self.row_thresh_var = tk.DoubleVar(rowvar_frame, 0.0)
		self.row_thresh_listvar = tk.Variable(rowvar_frame)
		self.row_thresh_gtlt_var = tk.StringVar(rowvar_frame, '>')
		self.row_thresh_gtlt = ttk.Combobox(rowvar_frame, width=2, state="disabled", textvariable=self.row_thresh_gtlt_var, values=gt_lt_list)
		self.row_thresh_gtlt.bind("<<ComboboxSelected>>", self.q_recalc)
		self.row_thresh_gtlt.grid(row=2, column=1, sticky=tk.N+tk.E, padx=(3,0), pady=(3,3))
		self.row_thresh_entry = ttk.Entry(rowvar_frame, width=6, textvariable=self.row_thresh_var, state="disabled", validate='key', validatecommand=(num_validator, '%P'))
		self.row_thresh_entry.bind("<KeyRelease>", self.ck_row_thresh_num)
		self.row_thresh_entry.grid(row=2, column=2, sticky=tk.N+tk.EW, padx=(3,3), pady=(3,3))
		self.col_thresh_var = tk.DoubleVar(colvar_frame, 0.0)
		self.col_thresh_gtlt_var = tk.StringVar(colvar_frame, '>')
		self.col_thresh_gtlt = ttk.Combobox(colvar_frame, width=2, state="disabled", textvariable=self.col_thresh_gtlt_var, values=gt_lt_list)
		self.col_thresh_gtlt.bind("<<ComboboxSelected>>", self.q_recalc)
		self.col_thresh_gtlt.grid(row=2, column=1, sticky=tk.N+tk.E, padx=(3,0), pady=(3,3))
		self.col_thresh_listvar = tk.Variable(colvar_frame)
		self.col_thresh_entry = ttk.Entry(colvar_frame, width=6, textvariable=self.col_thresh_var, state="disabled", validate='key', validatecommand=(num_validator, '%P'))
		self.col_thresh_entry.bind("<KeyRelease>", self.ck_col_thresh_num)
		self.col_thresh_entry.grid(row=2, column=2, sticky=tk.N+tk.EW, padx=(3,3), pady=(3,3))
		self.row_thresh_list = tk.Listbox(rowvar_frame, listvariable=self.row_thresh_listvar, selectmode=tk.MULTIPLE, width=24, height=24)
		self.row_thresh_list.configure(exportselection=False)
		self.row_thresh_list.bind("<<ListboxSelect>>", self.ck_row_thresh_list)
		self.rsb = ttk.Scrollbar(rowvar_frame, orient='vertical', command=self.row_thresh_list.yview)
		self.row_thresh_list.configure(yscrollcommand=self.rsb.set)
		self.col_thresh_list = tk.Listbox(colvar_frame, listvariable=self.col_thresh_listvar, selectmode=tk.MULTIPLE, width=24, height=24)
		self.col_thresh_list.configure(exportselection=False)
		self.col_thresh_list.bind("<<ListboxSelect>>", self.ck_col_thresh_list)
		self.csb = ttk.Scrollbar(colvar_frame, orient='vertical', command=self.col_thresh_list.yview)
		self.col_thresh_list.configure(yscrollcommand=self.csb.set)

		# PanedWindow within content_frame: left is table, right is plot
		output_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		output_panes.grid(row=0, column=0, sticky=tk.NSEW)

		self.out_tbl_frm = tk.Frame(output_panes, borderwidth=3, relief=tk.RIDGE)
		self.out_tbl_frm.grid(row=0, column=0, sticky=tk.NSEW)
		self.out_tbl_frm.rowconfigure(0, weight=1)
		self.out_tbl_frm.columnconfigure(0, weight=1, minsize=400)
		output_panes.add(self.out_tbl_frm, weight=1)

		self.out_plt_frm = tk.Frame(output_panes, borderwidth=3, relief=tk.RIDGE)
		self.out_plt_frm.grid(row=0, column=1, sticky=tk.NSEW, padx=(1,6))
		self.out_plt_frm.rowconfigure(0, weight=1)
		self.out_plt_frm.columnconfigure(0, weight=1)
		output_panes.add(self.out_plt_frm, weight=1)
		# Contingency table (a matplotlib plot)
		self.plotfig = Figure(dpi=100, layout="tight")
		self.plotfig_canvas = FigureCanvasTkAgg(self.plotfig, self.out_plt_frm)
		self.plot_nav = NavigationToolbar2Tk(self.plotfig_canvas, self.out_plt_frm)
		self.plotfig_canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
		self.plot_nav.update()

		# Initialize output frames
		self.clear_output()

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		new_close_button(self.dlg, self.btn_frame, 1, self.do_close)

	def validate_num_entry(self, potential_new_value):
		return not (potential_new_value is None or len(str(potential_new_value)) == 0)

	def clear_output(self):
		for ctl in self.out_tbl_frm.winfo_children():
			ctl.destroy()
		tframe, tdata = treeview_table(self.out_tbl_frm, [], self.output_columns)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.plotfig.clear()
		rc_labels = ['Positive', 'Negative', 'Sums']
		wh = "white"
		lg = "lightgrey"
		ccolors = [[wh, wh, lg], [wh, wh, lg], [lg, lg, lg]]
		self.plot_axes = self.plotfig.add_subplot(111)
		self.plotfig.patch.set_visible(False)
		self.plot_axes.axis("off")
		self.plot_axes.axis("tight")
		self.plotfig.tight_layout()
		self.cont_tbl = self.plot_axes.table(cellText=[[0,0,0],[0,0,0],[0,0,0]], cellColours=ccolors, rowLabels=rc_labels, colLabels=rc_labels, cellLoc="center", loc="center")
		self.cont_tbl.set_fontsize(10)
		self.cont_tbl.scale(1,4)
		self.plot_axes.set_aspect(1)
		self.plotfig_canvas.draw()
		self.dataset = None
		self.dlg.bind("<Control-s>")

	def ck_rowvar(self, event):
		# Enable and disable controls depending on whether the variable is numeric or categorical.
		rowvar = self.row_var.get()
		if rowvar in self.categ_columns:
			self.row_thresh_lbl["text"] = "Select categories for\npositive response"
			self.row_thresh_entry.grid_forget()
			self.row_thresh_gtlt.grid_forget()
			self.row_thresh_list.grid(row=3, column=0, sticky=tk.N+tk.EW, padx=(3,0), pady=(3,3))
			self.rsb.grid(row=3, column=1, sticky=tk.NS, pady=(3,3))
			# Populate the list box
			curs = data_db.cursor()
			rowdata = curs.execute(f"select distinct {db_colnames([rowvar])[0]} from mapdata order by 1;").fetchall()
			curs.close()
			rowvals = [r[0] for r in rowdata]
			self.row_thresh_listvar.set(rowvals)
		else:
			self.row_thresh_lbl["text"] = "Threshold for\npositive response"
			self.row_thresh_list.grid_forget()
			self.rsb.grid_forget()
			self.row_thresh_gtlt.grid(row=2, column=1, sticky=tk.N+tk.E, padx=(3,0), pady=(3,3))
			self.row_thresh_entry.grid(row=2, column=2, sticky=tk.N+tk.EW, padx=(3,3), pady=(3,3))
			self.row_thresh_gtlt["state"] = "readonly"
			self.row_thresh_entry["state"] = tk.NORMAL
		self.q_recalc()

	def ck_colvar(self, event):
		# Enable and disable controls depending on whether the variable is numeric or categorical.
		colvar = self.col_var.get()
		if colvar in self.categ_columns:
			self.col_thresh_lbl["text"] = "Select categories for\npositive response"
			self.col_thresh_entry.grid_forget()
			self.col_thresh_gtlt.grid_forget()
			self.col_thresh_list.grid(row=3, column=0, sticky=tk.N+tk.EW, padx=(3,0), pady=(3,3))
			self.csb.grid(row=3, column=1, sticky=tk.NS, pady=(3,3))
			# Populate the list box
			curs = data_db.cursor()
			coldata = curs.execute(f"select distinct {db_colnames([colvar])[0]} from mapdata order by 1;").fetchall()
			curs.close()
			colvals = [c[0] for c in coldata]
			self.col_thresh_listvar.set(colvals)
		else:
			self.col_thresh_lbl["text"] = "Threshold for\npositive response"
			self.col_thresh_list.grid_forget()
			self.csb.grid_forget()
			self.row_thresh_gtlt.grid(row=2, column=1, sticky=tk.N+tk.E, padx=(3,0), pady=(3,3))
			self.col_thresh_entry.grid(row=2, column=2, sticky=tk.N+tk.EW, padx=(3,3), pady=(3,3))
			self.col_thresh_gtlt["state"] = "readonly"
			self.col_thresh_entry["state"] = tk.NORMAL
		self.q_recalc()

	def ck_row_thresh_num(self, event=None):
		self.q_recalc()

	def ck_col_thresh_num(self, event=None):
		self.q_recalc()

	def ck_col_thresh_list(self, event=None):
		self.q_recalc()

	def ck_row_thresh_list(self, event=None):
		self.q_recalc()

	def q_recalc(self, event=None):
		# Recalculation can be done if both row and column variables are specifiied and have
		# thresholds selected.
		self.parent.loading_dlg.display("Checking data")
		can_recalc = 0
		self.dlg.bind("<Control-s>")
		rv, cv = self.row_var.get(), self.col_var.get()
		if rv != '' and cv != '':
			if rv in self.categ_columns:
				# Check that at least one value has been selected for the threshold set.
				rowthreshvals = self.row_thresh_list.curselection()
				if rowthreshvals is not None:
					can_recalc += 1
			else:
				# There is always a valid numeric threshold, defaulting to zero.
				can_recalc += 1
			if can_recalc == 1:
				if cv in self.categ_columns:
					# Check that at least one value has been selected for the threshold set.
					colthreshvals = self.col_thresh_list.curselection()
					if colthreshvals is not None:
						can_recalc += 1
				else:
					# There is always a valid numeric threshold, defaulting to zero.
					can_recalc += 1
		self.parent.loading_dlg.hide()
		if can_recalc == 2:
			self.recalc()
		else:
			self.clear_output()

	def get_data(self):
		# self.dataset will contain the contingency table values in row-major order.
		self.dataset = None
		self.dlg.bind("<Control-s>")
		rv, cv = db_colnames([self.row_var.get(), self.col_var.get()])
		if rv in self.categ_columns:
			list_items = self.row_thresh_listvar.get()
			curr_selections = self.row_thresh_list.curselection()
			rowthreshvals = [str(list_items[i]) for i in curr_selections]
			rowvals = ",".join(squotelist(rowthreshvals))
			row_sql = f"{rv} in ({rowvals})"
		else:
			rtv = self.row_thresh_var.get()
			if type(rtv) not in (int, float):
				rtv = 0
			if self.row_thresh_gtlt_var.get() == '>':
				row_sql = f"{rv} > {rtv}"
			else:
				row_sql = f"{rv} < {rtv}"
		if cv in self.categ_columns:
			list_items = self.col_thresh_listvar.get()
			curr_selections = self.col_thresh_list.curselection()
			colthreshvals = [list_items[i] for i in curr_selections]
			colvals = ",".join(squotelist(colthreshvals))
			col_sql = f"{cv} in ({colvals})"
		else:
			ctv = self.col_thresh_var.get()
			if type(ctv) not in (int, float):
				ctv = 0
			if self.col_thresh_gtlt_var.get() == '>':
				col_sql = f"{cv} > {ctv}"
			else:
				col_sql = f"{cv} < {ctv}"
		if self.sel_only_var.get() == '1':
			tv_ids = ",".join(self.parent.tbl.selection())
			sel_sql = f"and treeviewid in ({tv_ids})"
		else:
			sel_sql = ""
		sqlcmd = f"""select c.rowvar, c.colvar, coalesce(num, 0)
				from (
				select 1 as rowvar, 1 as colvar
				union select 1 as rowvar, 0 as colvar
				union select 0 as rowvar, 1 as colvar
				union select 0 as rowvar, 0 as colvar
				) as c
  				left join (select rowvar, colvar, count(*) as num
  				from (select ({row_sql}) as rowvar, ({col_sql}) as colvar
  				from mapdata
  				where {rv} is not null and {cv} is not null {sel_sql}
  				)
				group by rowvar, colvar
  				) as d on coalesce(d.rowvar, 0) = c.rowvar and coalesce(d.colvar, 0) = c.colvar
				order by c.rowvar desc, c.colvar desc;"""
		curs = data_db.cursor()
		data = curs.execute(sqlcmd).fetchall()
		if len(data) > 0:
			self.dataset = [r[2] for r in data]
		curs.close()

	def recalc(self):
		self.parent.loading_dlg.display("Recalculating")
		self.dlg.bind("<Control-s>")
		self.get_data()
		if self.dataset is not None:
			rc_labels = ['Positive', 'Negative', 'Sums']
			wh = "white"
			lg = "lightgrey"
			ccolors = [[wh, wh, lg], [wh, wh, lg], [lg, lg, lg]]
			ct_data = np.array([[self.dataset[0], self.dataset[1], self.dataset[0]+self.dataset[1]],
					[self.dataset[2], self.dataset[3], self.dataset[2]+self.dataset[3]],
					[self.dataset[0]+self.dataset[2], self.dataset[1]+self.dataset[3], sum(self.dataset)]])
			self.cont_tbl = self.plot_axes.table(cellText=ct_data, cellColours=ccolors, rowLabels=rc_labels, colLabels=rc_labels, cellLoc="center", loc="center")
			self.cont_tbl.set_fontsize(10)
			self.cont_tbl.scale(1,4)
			self.plotfig_canvas.draw()

			self.statdata = []
			dt = np.array([[self.dataset[0], self.dataset[1]],[self.dataset[2], self.dataset[3]]])
			if sum(self.dataset) >= 20:
				try:
					chi2 = spstats.chi2_contingency(dt)
				except:
					self.statdata.append(["Chi-square statistic", "NC"])
					self.statdata.append(["Chi-square p value", "NC"])
					self.statdata.append(["Degrees of freedom", "NC"])
				else:
					self.statdata.append(["Chi-square statistic", fp_display(chi2.statistic)])
					self.statdata.append(["Chi-square p value", fp_display(chi2.pvalue)])
					self.statdata.append(["Degrees of freedom", chi2.dof])
				fe = spstats.fisher_exact(dt)
				self.statdata.append(["Fisher exact test statistic", fp_display(fe.statistic)])
				self.statdata.append(["Fisher exact test p value", fp_display(fe.pvalue)])
				be = spstats.barnard_exact(dt)
				self.statdata.append(["Barnard exact test statistic", fp_display(be.statistic)])
				self.statdata.append(["Barnard exact test p value", fp_display(be.pvalue)])
				be = spstats.boschloo_exact(dt)
				self.statdata.append(["Boschloo exact test statistic", fp_display(be.statistic)])
				self.statdata.append(["Boschloo exact test p value", fp_display(be.pvalue)])
			if ct_data[0][1] != 0 and not (ct_data[0][0] == 0 and ct_data[1][0] == 0):
				self.statdata.append(["Risk ratio", fp_display((ct_data[0][0] * (ct_data[0][1] + ct_data[1][1])) / \
						(ct_data[0][1] * (ct_data[0][0] + ct_data[1][0])))])
			if ct_data[0][1] != 0 and ct_data[1][0] != 0:
				oddsr = ((ct_data[0][0] * ct_data[1][1]) / (ct_data[0][1]) * ct_data[1][0])
				self.statdata.append(["Odds ratio", fp_display(oddsr)])
				if oddsr != 0.0:
					log_or = math.log(oddsr)
					self.statdata.append(["Log odds ratio", fp_display(log_or)])
					se_logor = math.sqrt(1/ct_data[0][0] + 1/ct_data[0][1] + 1/ct_data[1][0] + 1/ct_data[1][1])
					self.statdata.append(["SE of log odds ratio", fp_display(se_logor)])
					halfint = 1.96 * se_logor
					self.statdata.append(["95% CI for log odds ratio", fp_display(log_or - halfint) + " - " + fp_display(log_or + halfint)])
				self.statdata.append(["Yule's association (Q)", fp_display((oddsr - 1)/(oddsr + 1))])
				self.statdata.append(["Yule's colligation (Y)", fp_display((math.sqrt(oddsr) - 1)/(math.sqrt(oddsr) + 1))])
				n = ct_data[2][2]
				phi = (((ct_data[0][0]/n) * (ct_data[1][1]/n)) - ((ct_data[0][1]/n) * (ct_data[1][0]/n))) / \
						math.sqrt((ct_data[0][2]/n) * (ct_data[1][2]/n) * (ct_data[2][0]/n) * (ct_data[2][1]/n))
				self.statdata.append(["Yule's phi", fp_display(phi)])
			if ct_data[2][2] != 0:
				self.statdata.append(["P(R)", fp_display(ct_data[0][2]/ct_data[2][2])])
				if ct_data[2][0] != 0:
					self.statdata.append(["P(R | C)", fp_display(ct_data[0][0]/ct_data[2][0])])
				if ct_data[2][1] != 0:
					self.statdata.append(["P(R | ¬C)", fp_display(ct_data[0][1]/ct_data[2][1])])
				self.statdata.append(["P(C)", fp_display(ct_data[2][0]/ct_data[2][2])])
				if ct_data[0][2] != 0:
					self.statdata.append(["P(C | R)", fp_display(ct_data[0][0]/ct_data[0][2])])
				if ct_data[1][2] != 0:
					self.statdata.append(["P(C | ¬R)", fp_display(ct_data[1][0]/ct_data[1][2])])
				self.statdata.append(["P(R ∧ C)", fp_display(ct_data[0][0]/ct_data[2][2])])
				self.statdata.append(["P(R ∨ C)", fp_display((ct_data[0][2] + ct_data[1][0])/ct_data[2][2])])

			for ctl in self.out_tbl_frm.winfo_children():
				ctl.destroy()
			tframe, tdata = treeview_table(self.out_tbl_frm, self.statdata, self.output_columns)
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			self.dlg.bind("<Control-s>", self.save_table)
		self.parent.loading_dlg.hide()

	def set_autoupdate(self):
		if self.autoupdate_var.get() == "1":
			self.auto_update = True
			self.q_recalc()
		else:
			self.auto_update = False

	def save_table(self, *args):
		export_data_table(self.output_columns, self.statdata, sheetname="Contingency table statistics")
	def do_close(self, *args):
		self.parent.remove_conttable(self)
		super().do_cancel(args)



class ROCCurveDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		super().__init__("Receiver Operating Characteristics",
				"Select condition and predictor variables. For the condition variable, the positive responses must be defined.",
				help_url="https://mapdata.readthedocs.io/en/latest/roccurve.html")
		clear_dlg_hotkeys(self.dlg)
		# Data
		self.output_columns = ["Statistic", "Value"]
		self.dataset = None
		self.data_labels = None
		self.condition_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float", "string", "date")])
		self.categ_columns = [c[0] for c in self.column_specs if c[1] in ["string", "date"]]
		self.pred_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.numeric_columns = self.pred_columns
		self.statdata = []

		# Selected and auto-update checkboxes
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_recalc)
		self.autoupdate_var = add_autoupdate(self.ctrl_frame, 0, 1, self.set_autoupdate)
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(1, weight=1)

		# All other input and output frames are inside content_frame
		# Only the second column, containing the output frame, can grow horizontally
		self.content_frame.columnconfigure(0, weight=0)
		self.content_frame.columnconfigure(1, weight=1)

		condition_frame = tk.Frame(self.content_frame)
		condition_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(6,3))
		# The condition frame has a label on row 0, a variable name selector on row 1,
		# another label on row 2, and a threshold selector on row 3.  Because the threshold selector
		# may be a list box, row3, and only row3, is expandable.
		condition_frame.rowconfigure(3, weight=1)

		# Condition variable
		ttk.Label(condition_frame, text="Condition variable").grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,1))
		self.cond_var = tk.StringVar(condition_frame)
		self.cond_sel = ttk.Combobox(condition_frame, state="readonly", textvariable=self.cond_var, values=self.condition_columns, width=24)
		self.cond_sel.grid(row=1, column=0, columnspan=3, sticky=tk.N+tk.EW, padx=(3,3), pady=(3,3))
		self.cond_sel.bind("<<ComboboxSelected>>", self.ck_condvar)

		self.cond_thresh_lbl = ttk.Label(condition_frame, text="Threshold for\npositive response")
		self.cond_thresh_lbl.grid(row=2, column=0, sticky=tk.N+tk.EW)

		# The threshold for positive responses is either a numeric variable specified in an Entry widget
		# or a selected list of categorical values specified in a Listbox widget.  The Entry widget is
		# initially displayed, but disabled.  The 'ck_rowvar' and 'ck_colvar' methods hide and re-grid
		# these widgets as necessary.
		gt_lt_list = ['>', '<']
		num_validator = self.dlg.register(self.validate_num_entry)
		self.cond_thresh_var = tk.DoubleVar(condition_frame, 0.0)
		self.cond_thresh_listvar = tk.Variable(condition_frame)
		self.cond_thresh_gtlt_var = tk.StringVar(condition_frame, '>')
		self.cond_thresh_gtlt = ttk.Combobox(condition_frame, width=2, state="disabled", textvariable=self.cond_thresh_gtlt_var, values=gt_lt_list)
		self.cond_thresh_gtlt.bind("<<ComboboxSelected>>", self.q_recalc)
		self.cond_thresh_gtlt.grid(row=2, column=1, sticky=tk.N+tk.E, padx=(3,0), pady=(3,3))
		self.cond_thresh_entry = ttk.Entry(condition_frame, width=6, textvariable=self.cond_thresh_var, validate='key', validatecommand=(num_validator, '%P'))
		self.cond_thresh_entry.bind("<KeyRelease>", self.ck_cond_thresh_num)
		self.cond_thresh_entry.grid(row=2, column=2, sticky=tk.N+tk.EW, padx=(3,3), pady=(3,3))
		self.cond_thresh_entry["state"] = tk.DISABLED
		self.cond_thresh_list = tk.Listbox(condition_frame, listvariable=self.cond_thresh_listvar, selectmode=tk.MULTIPLE, width=24, height=15)
		self.cond_thresh_list.configure(exportselection=False)
		self.cond_thresh_list.bind("<<ListboxSelect>>", self.ck_cond_thresh_list)
		self.condsb = ttk.Scrollbar(condition_frame, orient='vertical', command=self.cond_thresh_list.yview)
		self.cond_thresh_list.configure(yscrollcommand=self.condsb.set)

		# PanedWindow within self.content_frame: left is plot, right is table
		output_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		output_panes.grid(row=0, column=1, sticky=tk.NSEW, padx=(0,6))

		# The two panes are roc_frame at left, stat_frame at right.
		# The first two rows of roc_frame contain a combobox for the predictor variable.
		# The first two rows of stat_frame contain entry for the predictor threshold.
		roc_frame = tk.Frame(output_panes)
		roc_frame.columnconfigure(0, weight=1)
		roc_frame.rowconfigure(2, weight=1)
		roc_frame.grid(row=2, column=0)
		output_panes.add(roc_frame, weight=1)

		stat_frame = tk.Frame(output_panes)
		stat_frame.columnconfigure(0, weight=1)
		stat_frame.rowconfigure(2, weight=1)
		stat_frame.grid(row=2, column=1)
		output_panes.add(stat_frame, weight=1)

		# Populate roc_frame
		ttk.Label(roc_frame, text="Predictor variable").grid(row=0, column=0, sticky=tk.W, padx=(3,3), pady=(3,1))
		self.pred_var = tk.StringVar(roc_frame)
		self.pred_sel = ttk.Combobox(roc_frame, state="readonly", textvariable=self.pred_var, values=self.pred_columns, width=24)
		self.pred_sel.grid(row=1, column=0, sticky=tk.N+tk.W, padx=(3,3), pady=(3,3))
		self.pred_sel.bind("<<ComboboxSelected>>", self.ck_predvar)

		self.out_plt_frm = tk.Frame(roc_frame, borderwidth=3, relief=tk.RIDGE)
		self.out_plt_frm.grid(row=2, column=0, sticky=tk.NSEW, padx=(2,2))
		self.plot = Plot(self.out_plt_frm, 3.5, 3.5, layout="tight")

		# Populate stat_frame
		ttk.Label(stat_frame, text="Prediction threshold").grid(row=0, column=0, sticky=tk.W, padx=(3,3), pady=(3,3))
		self.pred_thresh_var = tk.DoubleVar(stat_frame, 0.0)
		self.pred_thresh_entry = ttk.Entry(stat_frame, width=6, textvariable=self.pred_thresh_var, validate='key', validatecommand=(num_validator, '%P'))
		self.pred_thresh_entry.bind("<KeyRelease>", self.ck_pred_thresh_num)
		self.pred_thresh_entry.grid(row=1, column=0, sticky=tk.N+tk.W, padx=(3,3), pady=(3,3))

		self.out_tbl_frm = tk.Frame(stat_frame, borderwidth=3, relief=tk.RIDGE)
		self.out_tbl_frm.grid(row=2, column=0, sticky=tk.NSEW)
		self.out_tbl_frm.rowconfigure(0, weight=1)
		self.out_tbl_frm.columnconfigure(0, weight=1, minsize=450)

		# Buttons
		self.data_btn = add_help_src_close_btns(self.dlg, self.btn_frame, self.do_help, self.show_data, self.do_close)

		# Initialize output frames
		self.clear_output()

	def validate_num_entry(self, potential_new_value):
		return not (potential_new_value is None or len(str(potential_new_value)) == 0)

	def clear_output(self):
		for ctl in self.out_tbl_frm.winfo_children():
			ctl.destroy()
		tframe, tdata = treeview_table(self.out_tbl_frm, [], self.output_columns)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.plot.clear()
		self.plot.set_axis_labels("False positive rate (1 - Specificity)", "True positive rate (Sensitivity)")
		self.plot.draw()
		self.dataset = None
		clear_dlg_hotkeys(self.dlg)
		self.data_btn["state"] = tk.DISABLED

	def ck_condvar(self, event):
		# Enable and disable controls depending on whether the variable is numeric or categorical.
		condvar = self.cond_var.get()
		if condvar in self.categ_columns:
			self.cond_thresh_lbl["text"] = "Select categories for\npositive response"
			self.cond_thresh_entry.grid_forget()
			self.cond_thresh_gtlt.grid_forget()
			self.cond_thresh_list.grid(row=3, column=0, sticky=tk.N+tk.EW, padx=(3,0), pady=(3,3))
			self.condsb.grid(row=3, column=1, sticky=tk.NS, pady=(3,3))
			# Populate the list box
			curs = data_db.cursor()
			conddata = curs.execute(f"select distinct {db_colnames([condvar])[0]} from mapdata order by 1;").fetchall()
			curs.close()
			condvals = [r[0] for r in conddata]
			self.cond_thresh_listvar.set(condvals)
		else:
			self.cond_thresh_lbl["text"] = "Threshold for\npositive response"
			self.cond_thresh_list.grid_forget()
			self.condsb.grid_forget()
			self.cond_thresh_gtlt.grid(row=2, column=1, sticky=tk.NE, padx=(3,0), pady=(3,3))
			self.cond_thresh_entry.grid(row=2, column=2, sticky=tk.N+tk.EW, padx=(3,3), pady=(3,3))
			self.cond_thresh_gtlt["state"] = "readonly"
			self.cond_thresh_entry["state"] = tk.NORMAL
		self.q_recalc()

	def ck_predvar(self, event):
		# Enable and disable controls depending on whether the variable is numeric or categorical.
		self.q_recalc()

	def ck_cond_thresh_num(self, event=None):
		self.q_recalc()

	def ck_pred_thresh_num(self, event=None):
		self.q_recalc()

	def ck_cond_thresh_list(self, event=None):
		self.q_recalc()

	def q_recalc(self, event=None):
		# Recalculation can be done if the condition variable and positive values have been
		# specified, and a prediction variable specified.  The prediction threshold is always
		# specified, defaulting to zero, so does not affect whether recalculation can be done.
		self.parent.loading_dlg.display("Checking data")
		can_recalc = False
		self.clear_output()
		self.dlg.bind("<Control-s>")
		self.data_btn["state"] = tk.DISABLED
		condvar, predvar = self.cond_var.get(), self.pred_var.get()
		if condvar != '' and predvar != '':
			if condvar in self.condition_columns:
				# Check that at least one value has been selected for the threshold set.
				condthreshvals = self.cond_thresh_list.curselection()
				if condthreshvals is not None:
					can_recalc = True
		self.parent.loading_dlg.hide()
		if can_recalc:
			self.recalc()

	def get_data(self):
		self.dataset = None
		self.dlg.bind("<Control-s>")
		self.data_btn["state"] = tk.DISABLED
		column_list = [self.cond_var.get(), self.pred_var.get()]
		# Get either only the selected data or all data.
		dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is not None and len(dataset[0]) > 0:
			# Remove missing data
			clean_data = clean_missing(dataset, list(range(len(dataset))))
			# Convert to floats for numeric data only
			for i in range(len(clean_data)):
				if column_list[i] in self.numeric_columns:
					clean_data[i] = [conv_float(v) for v in clean_data[i]]
			# Assign Boolean flags to identify positive conditions
			condvar = self.cond_var.get()
			if condvar in self.categ_columns:
				list_items = self.cond_thresh_listvar.get()
				curr_selections = self.cond_thresh_list.curselection()
				condthreshvals = [list_items[i] for i in curr_selections]
				positives = [clean_data[0][i] in condthreshvals for i in range(len(clean_data[0]))]
			else:
				threshval = self.cond_thresh_var.get()
				if self.cond_thresh_gtlt_var.get() == '>':
					positives = [clean_data[0][i] > threshval for i in range(len(clean_data[0]))]
				else:
					positives = [clean_data[0][i] < threshval for i in range(len(clean_data[0]))]
			clean_data.append(positives)
			self.dataset = sort_columns(clean_data)
			self.data_labels = [self.cond_var.get(), self.pred_var.get(), "Positive"]
			self.data_btn["state"] = tk.NORMAL

	def true_positives(self, threshold):
		# Returns the number of correctly predicted positive values by threshold.
		return sum([(self.dataset[1][i] > threshold) and self.dataset[2][i] for i in range(len(self.dataset[1]))])

	def true_negatives(self, threshold):
		return sum([(self.dataset[1][i] <= threshold) and not self.dataset[2][i] for i in range(len(self.dataset[1]))])

	def sens_spec(self):
		# Return vectors of sensitivity and 1-specificity for the dataset
		totobs = len(self.dataset[1])
		p = sum(self.dataset[2])
		n = totobs - p
		if n > 0 and p > 0:
			tp = [self.true_positives(self.dataset[1][i]) for i in range(totobs)]
			tpr = [tp[i]/p for i in range(len(tp))]
			tn = [self.true_negatives(self.dataset[1][i]) for i in range(totobs)]
			tnr_inv = [1.0 - (tn[i]/n) for i in range(len(tn))]
			return tpr, tnr_inv
		else:
			return None, None

	def recalc(self):
		self.parent.loading_dlg.display("Recalculating")
		self.dlg.bind("<Control-s>")
		self.get_data()
		if self.dataset is not None:
			# Create the ROC plot.
			sensitivity, spec_inv = self.sens_spec()
			if sensitivity is not None and spec_inv is not None:
				self.plot.clear()
				self.plot.axes.scatter(spec_inv, sensitivity)
				self.plot.set_axis_labels("False positive rate (1 - Specificity)", "True positive rate (Sensitivity)")
				self.plot.draw()

				# Calculate maximum LR+, J, and minimum Euclidean distance
				lr_pos_max = max([sensitivity[i]/spec_inv[i] for i in range(len(sensitivity)) if spec_inv[i] != 0])
				j_max = max([sensitivity[i] + (1-spec_inv[i]) - 1 for i in range(len(sensitivity))])
				ed_min = min([math.sqrt((1 - sensitivity[i])**2 + (spec_inv[i])**2) for i in range(len(sensitivity))])

			# Show the ROC statistics for the selected predictor threshold
			thresh = self.pred_thresh_var.get()
			self.statdata = []
			obs = len(self.dataset[2])
			p = sum(self.dataset[2])
			n = obs - p
			self.statdata.append(["Observations", obs])
			self.statdata.append(["Actual positives (P)", p])
			self.statdata.append(["Actual negatives (N)", n])
			self.statdata.append(["Prediction variable", self.pred_var.get()])
			self.statdata.append(["Minimum of pred. var.", min(self.dataset[1])])
			self.statdata.append(["Maximum of pred. var.", max(self.dataset[1])])
			self.statdata.append(["Prediction threshold", thresh])
			pp = sum([self.dataset[1][i] > thresh for i in range(obs)])
			pn = sum([self.dataset[1][i] <= thresh for i in range(obs)])
			self.statdata.append(["Predicted positives", pp])
			self.statdata.append(["Predicted negatives", pn])
			tp = self.true_positives(thresh)
			tn = self.true_negatives(thresh)
			self.statdata.append(["Correctly predicted positives (TP)", tp])
			self.statdata.append(["Correctly predicted negatives (TN)", tn])
			fp = sum([not self.dataset[2][i] and (self.dataset[1][i] > thresh) for i in range(obs)])
			fn = sum([self.dataset[2][i] and (self.dataset[1][i] <= thresh) for i in range(obs)])
			self.statdata.append(["False positives (FP)", fp])
			self.statdata.append(["False negatives (FP)", fn])
			if p > 0:
				self.statdata.append(["Sensitivity (TP/P)", fp_display(tp/p)])
			if n > 0:
				self.statdata.append(["Specificity (TN/N)", fp_display(tn/n)])
				if p > 0:
					if tn != 1 and n != 0 and tn != n:
						self.statdata.append(["Positive likelihood ratio (LR+)", fp_display((tp/p)/(1-tn/n))])
					if tn > 0 and n != 0:
						self.statdata.append(["Negative likelihood ratio", fp_display((1-tp/p)/(tn/n))])
			if sensitivity is not None and spec_inv is not None:
				self.statdata.append(["Maximum LR+ for all data", fp_display(lr_pos_max)])
			if tp+fp > 0:
				self.statdata.append(["Precision (TP/(TP + FP))", fp_display(tp/(tp+fp))])
			if n > 0:
				self.statdata.append(["False positive rate (FP/N)", fp_display(fp/n)])
			if p > 0:
				self.statdata.append(["False negative rate (FN/P)", fp_display(fn/p)])
			if tp+fn+fp > 0:
				self.statdata.append(["Critical success index", fp_display(tp/(tp+fn+fp))])
			if obs > 0:
				self.statdata.append(["Accuracy ((TP+TN)/(P+N)", fp_display((tp+tn)/(obs))])
			if p > 0 and n > 0:
				self.statdata.append(["Youden's J statistic", fp_display(tp/p + tn/n - 1)])
				self.statdata.append(["Maximum J for all data", fp_display(j_max)])
				self.statdata.append(["Euclidean distance (ED)", fp_display(math.sqrt((1 - tp/p)**2 + (1 - tn/n)**2))])
				self.statdata.append(["Minimum ED for all data", fp_display(ed_min)])
			for ctl in self.out_tbl_frm.winfo_children():
				ctl.destroy()
			tframe, tdata = treeview_table(self.out_tbl_frm, self.statdata, self.output_columns)
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			self.dlg.bind("<Control-s>", self.save_table)
		self.parent.loading_dlg.hide()

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for ROC curve")

	def set_autoupdate(self):
		if self.autoupdate_var.get() == "1":
			self.auto_update = True
			self.q_recalc()
		else:
			self.auto_update = False

	def save_table(self, *args):
		export_data_table(self.output_columns, self.statdata, sheetname="Contingency table statistics")
	def do_close(self, *args):
		self.parent.remove_roccurve(self)
		super().do_cancel(args)



class PCADialog(Dialog):
	def __init__(self, parent, column_specs, id_columns):
		self.parent = parent
		self.column_specs = column_specs
		self.id_columns = id_columns
		super().__init__("Principal Component Analysis",
				"Select three or more variables from the left for PCA analysis.  Use Ctrl-click or Shift-click to select multiple rows.",
				help_url="https://mapdata.readthedocs.io/en/latest/pca.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		self.alpha = 0.45
		# Hover annotation for PCxPC scatter plots
		self.annot = None
		self.canvas_conn_id = None
		# Data
		self.dataset = None
		self.data_labels = None
		self.scores = None
		self.scores_labels = None
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.null_colnames = []
		self.tab_no = 0				# The tab of the output table
		self.out_tables = {}		# keys: output tab numbers; values: 3-tuple of column headers, column data, and table name
		self.dlg.bind("<Control-s>")

		# Top controls
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_recalc)

		self.cleantype_var = tk.StringVar(self.ctrl_frame, "Drop cases")
		ttk.Label(self.ctrl_frame, text="Remove missing values by:").grid(row=0, column=1, sticky=tk.W, padx=(12,3), pady=(3,3))
		self.cleantype_sel = ttk.Combobox(self.ctrl_frame, state=tk.NORMAL, textvariable=self.cleantype_var, values=["Drop cases","Drop variables"], width=15)
		self.cleantype_sel.bind('<<ComboboxSelected>>', self.q_recalc)
		self.cleantype_sel.grid(row=0, column=2, sticky=tk.W, padx=(3,6), pady=(3,3))

		self.zscore_var = tk.StringVar(self.ctrl_frame, "1")
		self.zscore_ck = ttk.Checkbutton(self.ctrl_frame, text="Use Z scores", command=self.q_recalc, variable=self.zscore_var,
				onvalue="1", offvalue="0")
		self.zscore_ck.grid(row=0, column=3, sticky=tk.W, padx=(12,3), pady=(3,3))
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(3, weight=1)

		# The data_frame encompasses the two panes of the variable frame and the content frame
		frame_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW, padx=(6,6))

		# Variable frame for list of quantitative columns/variables
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE, width=300)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.columnconfigure(0, weight=1, minsize=200)
		frame_panes.add(var_frame, weight=1)

		# Maybe add a dropdown to select the row ID variable if there are multiple.
		self.id_col_var = tk.StringVar(var_frame, self.id_columns[0])
		if len(self.id_columns) > 1:
			ttk.Label(var_frame, text="Row ID:").grid(row=0, column=0, sticky=tk.NW, padx=(6,3), pady=(3,3))
			var_frame.rowconfigure(0, weight=0)
			id_col_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.id_col_var,
					values=self.id_columns, width=24)
			id_col_sel.grid(row=1, column=0, sticky=tk.NE, padx=(6,3), pady=(3,3))
			var_frame.rowconfigure(1, weight=0)
			id_col_sel.bind('<<ComboboxSelected>>', self.q_recalc)

		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(15, len(self.numeric_columns)))
		self.column_frame.grid(row=2, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(2, weight=5)
		self.column_table.bind('<ButtonRelease-1>', self.q_recalc)

		# Frame for output.  This contains a tabbed Notebook widget,
		# with separate pages for scores (samples by PCs), loadings (PCs by variables), explained variance plot, PCxPC plots
		self.output_frame = tk.Frame(frame_panes, borderwidth=3, relief=tk.RIDGE)
		self.output_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.output_frame.rowconfigure(0, weight=1)
		self.output_frame.columnconfigure(0, weight=1, minsize=600)
		frame_panes.add(self.output_frame, weight=12)
		self.output_pages = ttk.Notebook(self.output_frame)
		self.output_pages.bind("<<NotebookTabChanged>>", self.tab_changed)
		self.output_pages.grid(row=0, column=0, sticky=tk.NSEW)
		# Page for PCA scores
		self.scores_page = tk.Frame(self.output_pages)
		self.scores_page.name = "scores"
		self.scores_page.rowconfigure(0, weight=1)
		self.scores_page.columnconfigure(0, weight=1)
		self.scores_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.output_pages.add(self.scores_page, text="Scores")
		# Page for PCA loadings
		self.loadings_page = tk.Frame(self.output_pages)
		self.loadings_page.name = "loadings"
		self.loadings_page.rowconfigure(0, weight=1)
		self.loadings_page.columnconfigure(0, weight=1)
		self.loadings_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.output_pages.add(self.loadings_page, text="Loadings")
		# Page for explained variance plot
		self.expvar_page = tk.Frame(self.output_pages)
		self.expvar_page.name = "variance"
		self.expvar_page.rowconfigure(0, weight=1)
		self.expvar_page.columnconfigure(0, weight=1)
		self.expvar_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.output_pages.add(self.expvar_page, text="Explained variance")
		# Page for the scree plot
		self.screeplot_page = tk.Frame(self.output_pages)
		self.screeplot_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.screeplot_page.name = "screeplot"
		self.screeplot = Plot(self.screeplot_page)
		self.output_pages.add(self.screeplot_page, text="Scree plot")
		# Page for the PC x PC plots
		self.pcplot_page = tk.Frame(self.output_pages)
		self.pcplot_page.name = "pcplots"
		self.pcplot_page.columnconfigure(0, weight=1)
		self.pcplot_page.grid(row=0, column=0, sticky=tk.NSEW)
		pcphdr = tk.Frame(self.pcplot_page)
		pcphdr.grid(row=0, column=0, sticky=tk.NSEW)
		pcpfig = tk.Frame(self.pcplot_page)
		pcpfig.grid(row=1, column=0, sticky=tk.NSEW)
		self.pcplot_page.rowconfigure(1, weight=1)
		ttk.Label(pcphdr, text="X:").grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(2,2))
		self.pcp_x_var = tk.StringVar(pcphdr, "")
		self.pcp_x_sel = ttk.Combobox(pcphdr, state="readonly", textvariable=self.pcp_x_var, values=[], width=5, height=5)
		self.pcp_x_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(2,2))
		self.pcp_x_sel.bind('<<ComboboxSelected>>', self.pcplot_redraw)
		ttk.Label(pcphdr, text="Y:").grid(row=0, column=2, sticky=tk.E, padx=(6,3), pady=(2,2))
		self.pcp_y_var = tk.StringVar(pcphdr, "")
		self.pcp_y_sel = ttk.Combobox(pcphdr, state="readonly", textvariable=self.pcp_y_var, values=[], width=5, height=5)
		self.pcp_y_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(2,2))
		self.pcp_y_sel.bind('<<ComboboxSelected>>', self.pcplot_redraw)
		self.pcplot = Plot(pcpfig)

		self.output_pages.add(self.pcplot_page, text="PC x PC plots")

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.data_btn = new_data_button(self.dlg, self.btn_frame, 1, self.show_data)
		self.addcols_btn = new_addcol_button(self.dlg, self.btn_frame, 2, self.add_columns)
		self.addcols_btn["text"]="Add Columns"
		new_close_button(self.dlg, self.btn_frame, 3, self.do_close)
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(3, weight=1)

		# initialize content frame with headings, no data
		self.clear_output()

	def q_recalc(self, *args):
		self.clear_output()
		self.addcols_btn["state"] = tk.DISABLED
		self.data_btn["state"] = tk.DISABLED
		if len(self.column_table.selection()) > 2:
			self.get_data()
			if self.dataset is not None and len(self.dataset) > 2 and len(self.dataset[0]) > 2:
				self.recalc()

	def get_data(self):
		# Get the selected data into 'dataset'
		self.dataset = None
		self.n_dataset_columns = 0
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		if len(column_list) > 2:
			self.n_dataset_columns = len(column_list)
			column_list.append(self.id_col_var.get())
			# Get either only the selected data or all data.
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) < 3:
				self.dataset = None
				self.data_labels = None
			else:
				if self.cleantype_var.get() == "Drop cases":
					# Remove rows with missing values
					dataset = clean_missing(dataset, list(range(self.n_dataset_columns)))
				else:
					# Remove columns with missing data.  This may remove all columns.
					dataset, column_list, n_removed = clean_missing_columns(dataset,
							column_list, list(range(self.n_dataset_columns)))
					self.n_dataset_columns = self.n_dataset_columns - n_removed
				if self.n_dataset_columns < 3 or len(dataset[0]) < 3:
					warning("Not enough data after removing missing values.", {"parent": self.dlg})
					self.dataset = None
					self.data_labels = None
				else:
					# Convert to floats
					for i in range(self.n_dataset_columns):
						dataset[i] = [conv_float(v) for v in dataset[i]]
					self.dataset = dataset
					self.data_labels = column_list
					self.data_btn["state"] = tk.NORMAL

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels, "Data for Principal Component Analysis")

	def recalc(self):
		self.dlg.bind("<Alt-a>")
		self.loading_dlg.display("Calculating PCA")
		if self.dataset is not None and len(self.dataset[0]) > 2:
			ds = np.array(columns_to_rows(self.dataset[:self.n_dataset_columns]))
			if self.zscore_var.get() == "1":
				ds = StandardScaler().fit_transform(ds)
			pca_model = PCA()
			pca_result = pca_model.fit_transform(ds)
			expvar = pca_model.explained_variance_
			expvar_ratio = pca_model.explained_variance_ratio_
			expvar_total = [sum(expvar_ratio[:i+1]) for i in range(len(expvar_ratio))]
			loadings = pca_model.components_.T * np.sqrt(expvar)
			pc_labels = [f"PC{i+1}" for i in range(len(pca_result[0]))]
		
			# Display the scores table
			# This is pca_result with row and column headers added
			scoretbl = pca_result.tolist()
			hdrs = [self.id_col_var.get()] + pc_labels
			for caseid_i, caseid in enumerate(self.dataset[self.n_dataset_columns]):
				scoretbl[caseid_i].insert(0, caseid)
			self.out_tables[0] = (hdrs, scoretbl, "PCA scores")
			sc_frame, scores_table = treeview_table(self.scores_page, rowset=scoretbl, column_headers=hdrs, \
					select_mode="none", nrows=min(10, len(scoretbl)))
			sc_frame.grid(row=0, column=0, sticky=tk.NSEW)
			# Save as column-oriented table to simplify PCxPC plots
			self.scores = rows_to_columns(scoretbl)
			self.scores_labels = hdrs
		
			# Display the loadings table
			loadtbl = loadings.T.tolist()
			for pc_i, pc in enumerate(loadtbl):
				pc.insert(0, f"PC{pc_i+1}")
			hdrs = ["PC"] + self.data_labels[:self.n_dataset_columns]
			self.out_tables[1] = (hdrs, loadtbl, "PCA loadings")
			ld_frame, loadings_table = treeview_table(self.loadings_page, rowset=loadtbl, column_headers=hdrs, \
					select_mode="none", nrows=min(10, len(scoretbl)))
			ld_frame.grid(row=0, column=0, sticky=tk.NSEW)

			# Display the table of PC by explained and total variance by PC.
			vartbl = columns_to_rows([expvar.tolist(), expvar_ratio.tolist(), expvar_total])
			for pc_i, pc in enumerate(vartbl):
				pc.insert(0, f"PC{pc_i+1}")
			hdrs = ["PC", "Variance", "Variance fraction", "Total"]
			self.out_tables[2] = (hdrs, vartbl, "PCA explained variance")
			var_frame, variance_table = treeview_table(self.expvar_page, rowset=vartbl, column_headers=hdrs, \
					select_mode="none", nrows=min(10, len(vartbl)))
			var_frame.grid(row=0, column=0, sticky=tk.NSEW)

			# Display the scree plot
			vartbl_v = rows_to_columns(vartbl)
			self.screeplot.axes.plot(vartbl_v[0], vartbl_v[3], marker=".")
			self.screeplot.set_axis_labels("Principal Component", "Total Variance")
			self.screeplot.redraw()

			# Populate controls for PC x PC plots
			self.pcp_x_sel["values"] = pc_labels
			self.pcp_y_sel["values"] = pc_labels
			self.dlg.bind("<Alt-a>", self.set_alpha)

		self.dlg.bind("<Control-s>", self.save_data)
		self.tab_changed(None)
		self.loading_dlg.hide()

	def pcplot_redraw(self, *args):
		xvar = self.pcp_x_sel.get()
		yvar = self.pcp_y_sel.get()
		if self.scores is not None and xvar != '' and yvar != '':
			self.pcplot.clear()
			xvals = self.scores[self.scores_labels.index(xvar)]
			yvals = self.scores[self.scores_labels.index(yvar)]
			splot = self.pcplot.axes.scatter(xvals, yvals, alpha=self.alpha)
			self.pcplot.set_axis_labels(xvar + " score", yvar + " score")
			self.pcplot.redraw()
			# Hover annotation
			hoverer = Hoverer(self.pcplot, self.scores[0], splot)
			self.canvas_conn_id = self.pcplot.canvas.mpl_connect("motion_notify_event", lambda ev: hoverer.hover(ev))

	def clear_output(self):
		for ctl in self.scores_page.winfo_children():
			ctl.destroy()
		tframe, tdata = treeview_table(self.scores_page, [], [])
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		for ctl in self.loadings_page.winfo_children():
			ctl.destroy()
		ltframe, ltdata = treeview_table(self.loadings_page, [], [])
		ltframe.grid(row=0, column=0, sticky=tk.NSEW)
		for ctl in self.expvar_page.winfo_children():
			ctl.destroy()
		evframe, evdata = treeview_table(self.expvar_page, [], [])
		evframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.screeplot.clear()
		self.pcp_x_var.set("")
		self.pcp_y_var.set("")
		self.pcplot.clear()
		clear_dlg_hotkeys(self.dlg)
		self.dataset = None
		self.scores = None
		self.data_btn["state"] = tk.DISABLED
		self.addcols_btn["state"] = tk.DISABLED

	def add_columns(self, *args):
		dlg = CustomContentDialog(parent=self.dlg, title="Save PCA Scores", prompt="Save PCA scores in table columns.")

		self.n_pc_var = tk.IntVar(dlg.content_frame, min(5, len(self.scores)-1))
		n_pc_lbl = ttk.Label(dlg.content_frame, justify=tk.RIGHT, text="Number of PCs:")
		n_pc_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		n_pc_entry = ttk.Spinbox(dlg.content_frame, textvariable=self.n_pc_var, from_=1, to=len(self.scores)-1, width=4)
		n_pc_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))

		self.prefix_var = tk.StringVar(dlg.content_frame, "")
		col_lbl = ttk.Label(dlg.content_frame, justify=tk.RIGHT, text="Prefix for new column names:")
		col_lbl.grid(row=1, column=0, sticky=tk.E, padx=(6,3), pady=(3,6))
		self.prefix_entry = ttk.Entry(dlg.content_frame, textvariable=self.prefix_var, width=6)
		self.prefix_entry.grid(row=1, column=1, sticky=tk.W, padx=(3,12), pady=(3,6))

		def ck_entry(*args):
			enable_if(dlg.ok_btn, len(self.prefix_var.get()) > 0)

		self.prefix_var.trace_add("write", ck_entry)
		self.prefix_entry.focus()
		ok = dlg.show()
		self.dlg.lift()
		if ok:
			self.loading_dlg.display("Saving columns of\nPCA scores")
			col_prefix = self.prefix_var.get()
			n_pcs = self.n_pc_var.get()
			key_col_id = self.id_col_var.get()
			db_key_col_id = db_colnames([key_col_id])[0]
			key_col_vals = self.scores[0]
			col_names = self.scores_labels[1:n_pcs+1]
			col_table = self.scores[1:n_pcs+1]
			cur = data_db.cursor()
			kv_list = ",".join(["'"+kv+"'" for kv in key_col_vals])
			data_rowids = cur.execute(f"select treeviewid from mapdata where {db_colnames([key_col_id])[0]} in ({kv_list});").fetchall()
			data_rowids = [str(item[0]) for item in data_rowids]
			for col_idx in range(len(col_names)):
				col_name = col_prefix + "_" + col_names[col_idx]
				db_col_name = db_colnames([col_name])[0]
				# Update the database
				new_column = col_name not in self.parent.headers
				if new_column:
					cur.execute(f"alter table mapdata add column {db_col_name} REAL;")
				for i in range(len(key_col_vals)):
					cur.execute(f"update mapdata set {db_col_name} = '{col_table[col_idx][i]}' where {db_key_col_id} = '{key_col_vals[i]}';")
				# Update the Treeview
				if new_column:
					add_tv_column(self.parent.tbl, col_name, data_rowids, col_table[col_idx])
					self.parent.headers = self.parent.headers + [col_name]
					viscols = list(self.parent.tbl["displaycolumns"])
					if viscols[0] != '#all':
						viscols.append(col_name)
						self.parent.tbl["displaycolumns"] = viscols
				else:
					for rowid in self.parent.tbl.get_children():
						if rowid in data_rowids:
							col_index = data_rowids.index(rowid)
							self.parent.tbl.set(rowid, column=col_name, value=col_table[col_idx][col_index])
						else:
							self.parent.tbl.set(rowid, column=col_name, value='')
				# Update the table specs
				dt = "float"
				coldata = self.parent.get_all_data([col_name])[0]
				missing = len([v for v in coldata if v is None or v == ''])
				unique = len(set([v for v in coldata if v is not None and v != '']))
				if new_column:
					self.parent.data_types.append([col_name, dt, missing, unique])
				else:
					col_ix = self.parent.headers.index(col_name)
					self.parent.data_types[col_ix] = [col_name, dt, missing, unique]
			self.loading_dlg.hide()

	def tab_changed(self, *args):
		self.tab_no = self.output_pages.index(self.output_pages.select())
		enable_if(self.addcols_btn, self.tab_no == 0 and self.scores is not None)
	def set_alpha(self, *args):
		dlg = OneFloatDialog(self.dlg, "Transparency", "Enter the opacity (alpha) value", min_value=0.0, max_value=1.0, initial=self.alpha)
		new_alpha = dlg.show()
		if new_alpha is not None:
			self.alpha = min(1.0, max(new_alpha, 0.0))
			self.pcplot_redraw()
	def save_data(self, *args):
		if self.tab_no in (0,1,2):
			hdrs, tbl, name = self.out_tables[self.tab_no]
			export_data_table(hdrs, tbl, sheetname=name)
			self.dlg.lift()
	def do_close(self, *args):
		self.parent.remove_unmixing(self)
		super().do_cancel(args)



class NMFUnmixingDialog(Dialog):
	def __init__(self, parent, column_specs, id_columns):
		self.parent = parent
		self.column_specs = column_specs
		self.id_columns = id_columns
		super().__init__("NMF Unmixing",
				"Select three or more variables from the left for unmixing to find end member patterns.  Use Ctrl-click or Shift-click to select multiple rows.  Use the 'Unmix' button to initiate the calculation.",
				help_url="https://mapdata.readthedocs.io/en/latest/unmixing.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.dataset = None
		self.data_labels = None
		self.em_diagnostics_cols = None
		self.em_data = None			# EM composition (concentration) for each variable (chemical)
		self.comp_data = None		# EM fractions in each case (sample)
		self.numeric_columns = sorted([c[0] for c in self.column_specs if c[1] in ("int", "float")])
		self.null_colnames = []
		self.tab_no = 0				# The tab of the output table
		self.out_tables = {}		# keys: output tab numbers; values: 3-tuple of column headers, column data, and table name
		self.dlg.bind("<Control-s>")

		# Top controls
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_recalc)

		self.cleantype_var = tk.StringVar(self.ctrl_frame, "Drop cases")
		ttk.Label(self.ctrl_frame, text="Remove missing values by:").grid(row=0, column=1, sticky=tk.W, padx=(12,3), pady=(3,3))
		self.cleantype_sel = ttk.Combobox(self.ctrl_frame, state=tk.NORMAL, textvariable=self.cleantype_var, values=["Drop cases","Drop variables"], width=15)
		self.cleantype_sel.bind('<<ComboboxSelected>>', self.q_recalc)
		self.cleantype_sel.grid(row=0, column=2, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(2, weight=1)

		# The content_frame encompasses the two panes of the variable frame and the content frame
		frame_panes = ttk.PanedWindow(self.content_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)

		# Variable frame for list of quantitative columns/variables
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(0, weight=1)
		var_frame.columnconfigure(0, weight=1, minsize=200)
		frame_panes.add(var_frame, weight=1)

		# Maybe add a dropdown to select the row ID variable if there are multiple.
		self.id_col_var = tk.StringVar(var_frame, self.id_columns[0])
		if len(self.id_columns) > 1:
			ttk.Label(var_frame, text="Row ID:").grid(row=0, column=0, sticky=tk.NW, padx=(6,3), pady=(3,3))
			var_frame.rowconfigure(0, weight=0)
			id_col_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.id_col_var,
					values=self.id_columns, width=24)
			id_col_sel.grid(row=1, column=0, sticky=tk.NE, padx=(6,3), pady=(3,3))
			var_frame.rowconfigure(1, weight=0)
			id_col_sel.bind('<<ComboboxSelected>>', self.q_recalc)

		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.numeric_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.numeric_columns)))
		self.column_frame.grid(row=2, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(2, weight=5)
		self.column_table.bind('<ButtonRelease-1>', self.q_recalc)

		# Add a default-disabled selector for the number of end members to use.
		em_frame = tk.Frame(var_frame)
		em_frame.grid(row=3, column=0, sticky=tk.NSEW)
		self.em_var = tk.IntVar(em_frame, 0)
		em_lbl = ttk.Label(em_frame, text="End members:")
		em_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.em_sel = ttk.Combobox(em_frame, textvariable=self.em_var, values=[], width=5)
		self.em_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.em_sel["state"] = tk.DISABLED

		# 'Unmix' button
		self.calc_btn = ttk.Button(var_frame, text="  Unmix  ", command=self.recalc, state=tk.DISABLED)
		self.calc_btn.grid(row=4, column=0, sticky=tk.E, padx=(3,18), pady=(3,3))

		# frame for output.  This contains a tabbed Notebook widget,
		# with separate pages for the end member composition and the sample fractions of eacm EM
		self.output_frame = tk.Frame(frame_panes, borderwidth=3, relief=tk.RIDGE)
		self.output_frame.grid(row=0, column=1, sticky=tk.NSEW)
		self.output_frame.rowconfigure(0, weight=1)
		self.output_frame.columnconfigure(0, weight=1, minsize=600)
		frame_panes.add(self.output_frame, weight=12)
		self.output_pages = ttk.Notebook(self.output_frame)
		self.output_pages.bind("<<NotebookTabChanged>>", self.tab_changed)
		self.output_pages.grid(row=0, column=0, sticky=tk.NSEW)
		# Page for EM composition
		self.em_page = tk.Frame(self.output_pages)
		self.em_page.name = "ems"
		self.em_page.rowconfigure(0, weight=1)
		self.em_page.columnconfigure(0, weight=1)
		self.em_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.output_pages.add(self.em_page, text="EM compositions")
		# Page for EM fractional contributions to each sample
		self.contrib_page = tk.Frame(self.output_pages)
		self.contrib_page.name = "fractions"
		self.contrib_page.rowconfigure(0, weight=1)
		self.contrib_page.columnconfigure(0, weight=1)
		self.contrib_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.output_pages.add(self.contrib_page, text="EM contributions")
		# Page for EM values for each sample
		self.values_page = tk.Frame(self.output_pages)
		self.values_page.name = "values"
		self.values_page.rowconfigure(0, weight=1)
		self.values_page.columnconfigure(0, weight=1)
		self.values_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.output_pages.add(self.values_page, text="EM values")
		# Page for the EM composition plots
		self.emplot_page = tk.Frame(self.output_pages)
		self.emplot_page.columnconfigure(0, weight=1)
		self.emplot_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.output_pages.add(self.emplot_page, text="Profile plots")
		self.emplot = Plot(self.emplot_page)
		# Page for the EM diagnostics table
		self.diag_page = tk.Frame(self.output_pages)
		self.diag_page.rowconfigure(0, weight=1)
		self.diag_page.columnconfigure(0, weight=1)
		self.diag_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.output_pages.add(self.diag_page, text="Diagnostics")
		# Page for the EM diagnostics plot
		self.dplot_page = tk.Frame(self.output_pages)
		self.dplot_page.columnconfigure(0, weight=1)
		self.dplot_page.grid(row=0, column=0, sticky=tk.NSEW)
		self.output_pages.add(self.dplot_page, text="Diagnostic plots")
		dphdr = tk.Frame(self.dplot_page)
		dphdr.grid(row=0, column=0, sticky=tk.NSEW)
		dpfig = tk.Frame(self.dplot_page)
		dpfig.grid(row=1, column=0, sticky=tk.NSEW)
		self.dplot_page.rowconfigure(1, weight=1)
		ttk.Label(dphdr, text="Diagnostic:").grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(2,2))
		self.diag_var = tk.StringVar(dphdr, "")
		self.diag_sel = ttk.Combobox(dphdr, state="readonly", textvariable=self.diag_var,
				values=["Tot. resid.", "RSS", "RMSE", "Frob. norm"], width=13, height=4)
		self.diag_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(2,2))
		self.diag_sel.bind('<<ComboboxSelected>>', self.diagplot_redraw)
		self.diagplot = Plot(dpfig)

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.data_btn = new_data_button(self.dlg, self.btn_frame, 1, self.show_data)
		self.addcols_btn = new_addcol_button(self.dlg, self.btn_frame, 2, self.add_columns)
		self.addcols_btn["text"]="Add Columns"
		new_close_button(self.dlg, self.btn_frame, 3, self.do_close)
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(3, weight=1)

		# initialize content frame with headings, no data
		self.clear_output()

	def q_recalc(self, *args):
		self.clear_output()
		self.em_var.set(0)
		self.em_sel["state"] = tk.DISABLED
		enable_if(self.calc_btn, len(self.column_table.selection()) > 2)

	def get_data(self):
		self.loading_dlg.display("Unmixing data")
		self.dataset = None
		self.n_dataset_columns = 0
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		if len(column_list) > 0:
			self.n_dataset_columns = len(column_list)
			column_list.append(self.id_col_var.get())
			# Get either only the selected data or all data.
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) < 3:
				self.loading_dlg.hide()
				nrows = len(dataset[0])
				nvars = self.n_dataset_columns
				warning(f"There are {nvars} variables and {nrows} rows selected; there must be at least three of each", {"parent": self.dlg})
				self.dataset = None
				self.data_labels = None
			else:
				if self.cleantype_var.get() == "Drop cases":
					# Remove rows with missing values
					dataset = clean_missing(dataset, list(range(self.n_dataset_columns)))
				else:
					# Remove columns with missing data.  This may remove all columns.
					dataset, column_list, n_removed = clean_missing_columns(dataset,
							column_list, list(range(self.n_dataset_columns)))
					self.n_dataset_columns = self.n_dataset_columns - n_removed
				if self.n_dataset_columns < 3 or len(dataset[0]) < 3:
					self.dataset = None
					self.data_labels = None
					self.loading_dlg.hide()
					nrows = len(dataset[0])
					nvars = self.n_dataset_columns
					warning(f"There are {nvars} variables and {nrows} rows after cleaning; there must be at least three of each", {"parent": self.dlg})
				else:
					# Convert to floats
					for i in range(self.n_dataset_columns):
						dataset[i] = [conv_float(v) for v in dataset[i]]
					# Check that all values are non-negative
					any_neg = False
					for var in dataset[:self.n_dataset_columns]:
						if any([x < 0 for x in var]):
							any_neg = True
							self.loading_dlg.hide()
							warning("The data set must contain only positive values.", {"parent": self.dlg})
							break
					if any_neg:
						self.dataset = None
						self.data_labels = None
					else:
						self.dataset = dataset
						self.data_labels = column_list
						self.data_btn["state"] = tk.NORMAL
						self.loading_dlg.hide()

	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels, "Data for NMF unmixing")

	def recalc(self):
		self.get_data()
		self.loading_dlg.display("Unmixing data")
		if self.dataset is not None and len(self.dataset[0]) > 2:
			# If the EM-number dropdown is disabled, do multiple NMF runs to find the optimum EM number
			# else run NMF once for the given EM number.
			ds = np.array(columns_to_rows(self.dataset[:self.n_dataset_columns]))
			n_rows = len(self.dataset[0])
			n = n_rows * self.n_dataset_columns
			if self.em_var.get() == 0:
				max_ems = min(12, self.n_dataset_columns-1, len(self.dataset[0])-1)
				test_ems = list(range(1, max_ems))
				#rmses = []
				frobs = []
				unmixes = []
				diagnostics = []
				for em_no in test_ems:
					W, H, frobnorm = iterate_nmf(ds, em_no)
					fit_ds = np.matmul(W, H)
					fit_rss = array_rss(ds, fit_ds)
					fit_rmse = math.sqrt(mean_squared_error(ds, fit_ds))
					frobs.append(frobnorm)
					unmixes.append((W, H))
					# Diagnostics: Total residual, RSS, RMSE, Frobenius norm of matrix difference, AICc
					diagnostics.append([em_no, fp_display(arraysum(fit_ds)-arraysum(ds), 4), fp_display(fit_rss,5),
						fp_display(fit_rmse,4), fp_display(frobnorm,4)])
				if len(frobs) > 1:
					frob_diffs = [frobs[i] - frobs[i+1] for i in range(len(frobs)-1)]
					sel_em = frob_diffs.index(max(frob_diffs))+2
					if len(frob_diffs) > 2:
						frob_diffs2 = [frob_diffs[i] - frob_diffs[i+1] for i in range(len(frob_diffs)-1)]
						sel_em2 = frob_diffs2.index(max(frob_diffs2))+2
						if sel_em != max_ems-1:
							sel_em = sel_em2
					else:
						sel_em = frob_diffs.index(max(frob_diffs))+2
					self.em_var.set(sel_em)
					sel_em_index = sel_em - 1
				else:
					self.em_var.set(1)
					sel_em_index = 0
				self.em_sel["values"] = test_ems
				self.em_sel["state"] = tk.NORMAL
				W, H = unmixes[sel_em_index]
				self.em_diagnostics = diagnostics
				self.em_diagnostics_labels = ["EMs", "Tot. resid.", "RSS", "RMSE", "Frob. norm"]
				self.out_tables[4] = (self.em_diagnostics_labels, diagnostics, "Diagnostics")
				tframe, tdata = treeview_table(self.diag_page, self.em_diagnostics, self.em_diagnostics_labels)
				tframe.grid(row=0, column=0, sticky=tk.NSEW)
				self.em_diagnostics_cols = rows_to_columns(diagnostics)
			else:
				W, H, frobnorm = iterate_nmf(ds, self.em_var.get())
			# Output H is row-wise EMs by chemicals -- the EM patterns, or compositions.
			# Output W is row-wise samples by EMs -- the EM contribution to each sample.
			fit_ds = np.matmul(W, H)
			em_vals = sample_EMs(W, H)
			# Display the composition matrix (H with EM identifiers and row headers)
			em_labels = ["EM"+str(n+1) for n in range(self.em_var.get())]
			H = H.tolist()
			for i in range(len(em_labels)):
				H[i].insert(0, em_labels[i])
			em_page_labels = ["End member"]+self.data_labels[:self.n_dataset_columns]
			self.out_tables[0] = (em_page_labels, H, "EM compositions")
			tframe, tdata = treeview_table(self.em_page, H, em_page_labels)
			tframe.grid(row=0, column=0, sticky=tk.NSEW)
			# Display the contribution matrix (W with sample identifiers and row headers
			# Prefix every row in W with the row ID
			W = W.tolist()
			for i in range(len(W)):
				W[i].insert(0, self.dataset[self.n_dataset_columns][i])
			W_labels = [self.id_col_var.get()]+em_labels
			self.out_tables[1] = (W_labels, W, "EM contributions")
			ltframe, ltdata = treeview_table(self.contrib_page, W, W_labels)
			ltframe.grid(row=0, column=0, sticky=tk.NSEW)
			# Display the sample EM values, with residuals
			# Prefix every row in W with the row ID
			em_totals = [sum(r) for r in em_vals]
			samp_totals = [sum(r) for r in ds]
			residuals = [samp_totals[i] - em_totals[i] for i in range(len(em_totals))]
			for i in range(len(em_vals)):
				em_vals[i].append(residuals[i])
				em_vals[i].insert(0, self.dataset[self.n_dataset_columns][i])
			em_vals_labels = [self.id_col_var.get()]+em_labels+["Residual"]
			self.out_tables[2] = (em_vals_labels, em_vals, "EM values")
			ltframe, ltdata = treeview_table(self.values_page, em_vals, em_vals_labels)
			ltframe.grid(row=0, column=0, sticky=tk.NSEW)
			# Display the plot of EM compositions (H)
			self.emplot.figure.clear()
			nplots = self.em_var.get()
			if nplots > 1:
				self.emplot.axes = axs = self.emplot.figure.subplots(nplots, sharex=True, sharey=True)
				emx = self.data_labels[:self.n_dataset_columns]
				for em_idx in range(len(H)):
					axs[em_idx].bar(emx, H[em_idx][1:])
					axs[em_idx].set_ylabel(H[em_idx][0])
					axs[em_idx].label_outer()
			else:
				self.emplot.clear()
				self.emplot.axes.bar(self.data_labels[:self.n_dataset_columns], H[0][1:])
				self.emplot.axes.set_ylabel(H[0][0])
			self.emplot.draw()

		self.dlg.bind("<Control-s>", self.save_data)
		self.loading_dlg.hide()

	def clear_output(self):
		for ctl in self.em_page.winfo_children():
			ctl.destroy()
		for ctl in self.contrib_page.winfo_children():
			ctl.destroy()
		for ctl in self.values_page.winfo_children():
			ctl.destroy()
		for ctl in self.diag_page.winfo_children():
			ctl.destroy()
		clear_dlg_hotkeys(self.dlg)
		self.dataset = None
		self.data_btn["state"] = tk.DISABLED
		self.addcols_btn["state"] = tk.DISABLED
		tframe, tdata = treeview_table(self.em_page, [], [])
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		ltframe, ltdata = treeview_table(self.contrib_page, [], [])
		ltframe.grid(row=0, column=0, sticky=tk.NSEW)
		ltframe, ltdata = treeview_table(self.values_page, [], [])
		ltframe.grid(row=0, column=0, sticky=tk.NSEW)
		dframe, ddata = treeview_table(self.diag_page, [], [])
		dframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.emplot.clear()
		self.diagplot.clear()
		self.diag_var.set("")

	def diagplot_redraw(self, *args):
		diagvar = self.diag_var.get()
		if self.em_diagnostics_cols is not None and diagvar != '':
			self.diagplot.clear()
			ems = self.em_diagnostics_cols[0]
			diag_vals = [float(x) for x in self.em_diagnostics_cols[self.em_diagnostics_labels.index(diagvar)]]
			self.diagplot.axes.plot(ems, diag_vals, marker=".")
			self.diagplot.set_axis_labels("End members", diagvar)
			self.diagplot.redraw()

	def add_columns(self, *args):
		dlg = CustomContentDialog(parent=self.dlg, title="Save EM Values", prompt="Save EM values in table columns.")
		self.prefix_var = tk.StringVar(dlg.content_frame, "")
		col_lbl = ttk.Label(dlg.content_frame, justify=tk.RIGHT, text="Prefix for new column names:")
		col_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,6))
		self.prefix_entry = ttk.Entry(dlg.content_frame, textvariable=self.prefix_var, width=6)
		self.prefix_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,12), pady=(3,6))
		def ck_entry(*args):
			enable_if(dlg.ok_btn, len(self.prefix_var.get()) > 0)
		self.prefix_var.trace_add("write", ck_entry)
		self.prefix_entry.focus()
		ok = dlg.show()
		self.dlg.lift()
		if ok:
			self.loading_dlg.display("Saving columns of EM values")
			col_prefix = self.prefix_var.get()
			em_hdrs, em_tbl, nm = self.out_tables[2]
			key_col_id = em_hdrs[0]
			db_key_col_id = db_colnames([key_col_id])[0]
			em_names = em_hdrs[1:]
			em_table = rows_to_columns(em_tbl)
			key_col_vals = em_table[0]
			em_table = em_table[1:]
			cur = data_db.cursor()
			kv_list = ",".join(["'"+kv+"'" for kv in key_col_vals])
			data_rowids = cur.execute(f"select treeviewid from mapdata where {db_colnames([key_col_id])[0]} in ({kv_list});").fetchall()
			data_rowids = [str(item[0]) for item in data_rowids]
			for col_idx in range(len(em_names)):
				col_name = col_prefix + "_" + em_names[col_idx]
				db_col_name = db_colnames([col_name])[0]
				# Update the database
				new_column = col_name not in self.parent.headers
				if new_column:
					cur.execute(f"alter table mapdata add column {db_col_name} REAL;")
				for i in range(len(key_col_vals)):
					cur.execute(f"update mapdata set {db_col_name} = '{em_table[col_idx][i]}' where {db_key_col_id} = '{key_col_vals[i]}';")
				# Update the Treeview
				if new_column:
					add_tv_column(self.parent.tbl, col_name, data_rowids, em_table[col_idx])
					self.parent.headers = self.parent.headers + [col_name]
					viscols = list(self.parent.tbl["displaycolumns"])
					if viscols[0] != '#all':
						viscols.append(col_name)
						self.parent.tbl["displaycolumns"] = viscols
				else:
					for rowid in self.parent.tbl.get_children():
						if rowid in data_rowids:
							em_index = data_rowids.index(rowid)
							self.parent.tbl.set(rowid, column=col_name, value=em_table[col_idx][em_index])
						else:
							self.parent.tbl.set(rowid, column=col_name, value='')
				# Update the table specs
				dt = "float"
				coldata = self.parent.get_all_data([col_name])[0]
				missing = len([v for v in coldata if v is None or v == ''])
				unique = len(set([v for v in coldata if v is not None and v != '']))
				if new_column:
					self.parent.data_types.append([col_name, dt, missing, unique])
				else:
					col_ix = self.parent.headers.index(col_name)
					self.parent.data_types[col_ix] = [col_name, dt, missing, unique]
			self.loading_dlg.hide()

	def tab_changed(self, *args):
		self.tab_no = self.output_pages.index(self.output_pages.select())
		enable_if(self.addcols_btn, self.tab_no == 2)
	def save_data(self, *args):
		if self.tab_no in self.out_tables:
			hdrs, tbl, name = self.out_tables[self.tab_no]
			export_data_table(hdrs, tbl, sheetname=name)
			self.dlg.lift()
	def do_close(self, *args):
		self.parent.remove_unmixing(self)
		super().do_cancel(args)



class CategCorrespDialog(Dialog):
	def __init__(self, parent, column_specs):
		self.parent = parent
		self.column_specs = column_specs
		super().__init__("Categorical Variable Correspondence",
				"Select two categorical variables to see the prevalence of all combinations.",
				help_url="https://mapdata.readthedocs.io/en/latest/categcorresp.html")
		# Data
		self.dataset = None
		self.categ_columns2 = [c[0] for c in self.column_specs if c[1] in ("string", "boolean", "date")]
		self.categ_columns2.sort()
		self.column_headers = []
		self.statdata = []

		# Controls
		self.var1 = tk.StringVar(self.ctrl_frame, "")
		var1_lbl = ttk.Label(self.ctrl_frame, text="Variable 1:")
		var1_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.var1_sel = ttk.Combobox(self.ctrl_frame, state=tk.NORMAL, textvariable=self.var1, width=24)
		self.var1_sel["values"] = self.categ_columns2
		self.var1_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.var1_sel.bind("<<ComboboxSelected>>", self.check_var1)

		self.var2 = tk.StringVar(self.ctrl_frame, "")
		var2_lbl = ttk.Label(self.ctrl_frame, text="Variable 2:")
		var2_lbl.grid(row=0, column=2, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.var2_sel = ttk.Combobox(self.ctrl_frame, state=tk.NORMAL, textvariable=self.var2, width=24)
		self.var2_sel["values"] = self.categ_columns2
		self.var2_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		self.var2_sel.bind("<<ComboboxSelected>>", self.check_var2)

		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 1, 0, self.q_recalc, colspan=2)

		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(3, weight=1)

		# Buttons
		self.data_btn = add_help_src_close_btns(self.dlg, self.btn_frame, self.do_help, self.show_data, self.do_close)

		self.dlg.minsize(width=300, height=300)

	def clear_output(self):
		for ctl in self.content_frame.winfo_children():
			ctl.destroy()
		clear_dlg_hotkeys(self.dlg)
		self.dataset = None

	def show_data(self, *args):
		# Show data that have been collected for co-occurrence display
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Original data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for categorical correspondence")

	def check_var1(self, *args):
		self.clear_output()
		v1 = self.var1.get()
		v2 = self.var2.get()
		if v1 != '':
			if v2 != '':
				if v1 == v2:
					self.var2.set('')
				else:
					self.q_recalc()

	def check_var2(self, *args):
		self.clear_output()
		v1 = self.var1.get()
		v2 = self.var2.get()
		if v2 != '':
			if v1 != '':
				if v2 == v1:
					self.var1.set('')
				else:
					self.q_recalc()

	def q_recalc(self, get_data=True, *args):
		# Conditionally (re)calculate the co-occurrences
		can_recalc = self.var1.get() != '' and self.var2.get() != ''
		if can_recalc:
			if get_data or self.dataset is None:
				self.get_data()
			if self.dataset is not None:
				self.recalc()

	def get_data(self):
		# Get the selected data into 'dataset'
		self.data_btn["state"] = tk.DISABLED
		self.dataset = None
		column_list = [self.var1.get(), self.var2.get()]
		# Get either only the selected data or all data.
		dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
		if dataset is None or len(dataset[0]) == 0:
			self.dataset = None
			self.data_labels = None
			self.data_btn["state"] = tk.DISABLED
		else:
			# DO NOT remove missing data for the categorical co-occurrence summary
			self.dataset = dataset
			self.data_labels = [self.var1.get(), self.var2.get()]
			self.data_btn["state"] = tk.NORMAL

	def recalc(self):
		# Put the data into a SQLite db for summarization
		self.db = sqlite3.connect(":memory:")
		cur = self.db.cursor()
		colnames = db_colnames(self.data_labels)
		cur.execute("create table cdata (%s);" % ",".join(colnames))
		tbldata = []
		for row_no in range(len(self.dataset[0])):
			row_vals = [self.dataset[0][row_no], self.dataset[1][row_no]]
			row_vals = [None if isinstance(x, str) and x.strip() == '' else x for x in row_vals]
			tbldata.append(row_vals)
		cur.executemany("insert into cdata values (?,?);", tbldata)
		# Create the summary
		sqlcmd = """select %s, count(*) as data_rows, round(100 * count(*)/total_rows, 3) as percent
from cdata cross join (select cast(count(*) as double) as total_rows from cdata)
group by 1,2 order by 1;""" % ",".join(colnames)
		result = cur.execute(sqlcmd)
		# Stuff the result into the content frame.
		self.output_columns = self.data_labels + ['Data rows', 'Percent']
		self.statdata = result.fetchall()
		tframe, tdata = treeview_table(self.content_frame, self.statdata, self.output_columns)
		cur.close()
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		self.dlg.bind("<Control-s>", self.save_table)

	def save_table(self, *args):
		export_data_table(self.output_columns, self.statdata, sheetname="Selected map items")
	def do_close(self, *args):
		self.parent.remove_categcorresp(self)
		super().do_cancel(args)



class CategSimDialog(Dialog):
	def __init__(self, parent, column_specs, id_columns):
		self.parent = parent
		self.id_columns = id_columns
		self.column_specs = column_specs
		super().__init__("Categorical Similarity",
				"Select one or more variables from the left to see the similarity matrix.  Use Ctrl-click or Shift-click to select multiple rows.",
				help_url="https://mapdata.readthedocs.io/en/latest/categsimmatrix.html")
		self.loading_dlg = LoadingDialog(self.dlg)
		# Data
		self.auto_update = True
		self.dataset = None
		self.data_labels = None
		self.data_ids = None
		self.sim_table = None
		self.sim_table_labels = None
		self.categorical_columns = sorted([c[0] for c in self.column_specs if c[1] == "string"])
		self.show_labels = True
		self.dlg.bind("<Alt-l>")
		self.dlg.bind("<Control-s>")

		# Controls
		# Top controls are only the 'Selected only' checkbox
		self.sel_only_var, self.sel_only_ck = add_sel_only(self.ctrl_frame, 0, 0, self.q_redraw)
		self.autoupdate_var = add_autoupdate(self.ctrl_frame, 0, 1, self.set_autoupdate)
		self.ctrl_frame.columnconfigure(0, weight=0)
		self.ctrl_frame.columnconfigure(1, weight=1)

		# The data_frame encompasses the two panes of the variable frame and the content frame
		data_frame = tk.Frame(self.content_frame, borderwidth=5)
		data_frame.rowconfigure(0, weight=1)
		data_frame.columnconfigure(0, weight=1)
		data_frame.grid(row=0, column=0, sticky=tk.NSEW)
		frame_panes = ttk.PanedWindow(data_frame, orient=tk.HORIZONTAL)
		frame_panes.grid(row=0, column=0, sticky=tk.NSEW)

		# Variable frame for list of categorical columns/variables
		var_frame = tk.Frame(frame_panes, borderwidth=2, relief=tk.RIDGE)
		var_frame.grid(row=0, column=0, sticky=tk.NSEW)
		var_frame.rowconfigure(2, weight=1)
		var_frame.columnconfigure(0, weight=1)
		frame_panes.add(var_frame, weight=1)
		# If there's more than one ID column, allow the user to select.
		self.id_col_var = tk.StringVar(var_frame, self.id_columns[0])
		if len(self.id_columns) > 1:
			ttk.Label(var_frame, text="Row ID:").grid(row=0, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
			id_col_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.id_col_var,
					values=self.id_columns, width=24)
			id_col_sel.grid(row=1, column=0, sticky=tk.NW, stick=tk.E, padx=(6,3), pady=(3,3))
			id_col_sel.bind('<<ComboboxSelected>>', self.q_redraw)
		# Add multi-select list of variables to the leftmost pane
		self.column_frame, self.column_table = treeview_table(var_frame, rowset=[[v] for v in self.categorical_columns], \
				column_headers=['Variable'], select_mode=tk.EXTENDED, nrows=min(10, len(self.id_columns)))
		self.column_frame.grid(row=2, column=0, sticky=tk.NSEW)
		self.column_table.bind('<ButtonRelease-1>', self.q_redraw)
		self.sim_type_var = tk.StringVar(var_frame, "Lin")
		ttk.Label(var_frame, text="Similarity type:").grid(row=3, column=0, sticky=tk.W, padx=(6,3), pady=(3,3))
		sim_type_sel = ttk.Combobox(var_frame, state="readonly", textvariable=self.sim_type_var, width=24)
		sim_type_sel["values"] = ["Lin", "Goodall3", "OF", "IOF", "Overlap"]
		sim_type_sel.grid(row=4, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		sim_type_sel.bind('<<ComboboxSelected>>', self.q_redraw)

		# Output frame for similarity matrix figure
		self.output_frame = tk.Frame(frame_panes, borderwidth=3, relief=tk.RIDGE)
		self.output_frame.grid(row=0, column=1, sticky=tk.NSEW)
		frame_panes.add(self.output_frame, weight=12)
		self.plot = Plot(self.output_frame, 3.5, 3.5)

		# Buttons
		new_help_button(self.dlg, self.btn_frame, self.do_help)
		self.data_btn = new_data_button(self.dlg, self.btn_frame, 1, self.show_data)
		self.sim_data_btn = new_button(self.btn_frame, "Similarities", 0, 2, self.show_sim_data, tk.W, (3,3), state=tk.DISABLED)
		new_close_button(self.dlg, self.btn_frame, 3, self.do_close)
		self.btn_frame.columnconfigure(0, weight=0)
		self.btn_frame.columnconfigure(3, weight=1)

		# initialize content frame with an empty plot
		self.clear_output()


	def show_data(self, *args):
		if self.dataset is not None:
			show_columnar_table(self.dlg, "Source Data", "Selected data:", self.dataset, self.data_labels[0:len(self.dataset)], \
					"Data for similarity matrix")

	def show_sim_data(self, *args):
		if self.sim_table is not None:
			show_table(self.dlg, "Similarities", "Categorical similarities:", self.sim_table, self.sim_table_labels, \
					"Categorical similarities")

	def clear_output(self):
		self.plot.clear()
		self.data_btn["state"] = tk.DISABLED
		self.sim_data_btn["state"] = tk.DISABLED

	def q_redraw(self, *args, get_data=True):
		self.clear_output()
		if self.dataset is None or get_data:
			self.get_data()
		if self.dataset is not None and len(self.dataset) > 1 and len(self.dataset[0]) > 1:
			self.data_btn["state"] = tk.NORMAL
			self.redraw()
		else:
			self.data_btn["state"] = tk.DISABLED
			self.sim_data_btn["state"] = tk.DISABLED

	def get_data(self):
		# Get the selected data into 'dataset'
		self.clear_output()
		self.dataset = None
		column_list = []
		for sel_row in self.column_table.selection():
			datarow = self.column_table.item(sel_row)["values"]
			column_list.append(datarow[0])
		self.n_data_columns = len(column_list)
		column_list.append(self.id_col_var.get())
		if self.n_data_columns > 0:
			# Get either only the selected data or all data.
			dataset = chosen_dataset(self.parent, column_list, self.sel_only_var.get() == "1")
			if dataset is None or len(dataset[0]) == 0:
				self.dataset = None
				self.data_labels = None
			else:
				# Remove missing data casewise
				clean_data = clean_missing(dataset, list(range(self.n_data_columns)))
				if len(clean_data[0]) == 0:
					self.dataset = None
					self.data_labels = None
					self.data_ids = None
				else:
					self.dataset = sort_columns(clean_data, sortby=self.n_data_columns)
					self.data_labels = column_list
					self.data_ids = clean_data[self.n_data_columns]
					self.data_btn["state"] = tk.NORMAL

	def redraw(self):
		# Recalculate and redraw the similarity matrix
		if self.data_labels is not None and len(self.data_labels) > 1:
			self.loading_dlg.display("Calculating categorical\nsimilarities")
			self.dlg.bind("<Alt-l>", self.set_labeling)
			ncases = len(self.data_ids)
			simmat = np.eye(ncases)
			simtype = self.sim_type_var.get()
			if simtype == "Lin":
				simfn = cat_sim_Lin
			elif simtype == "Goodall3":
				simfn = cat_sim_Goodall3
			elif simtype == "IOF":
				simfn = cat_sim_IOF
			elif simtype == "OF":
				simfn = cat_sim_OF
			else:	# Overlap
				simfn = cat_sim_Overlap
			sim_results = simfn(self.dataset[:self.n_data_columns], self.data_ids)
			for row in sim_results:
				row_i = self.data_ids.index(row[0])
				col_i = self.data_ids.index(row[1])
				simmat[row_i][col_i] = row[2]
				simmat[col_i][row_i] = row[2]
			caxes = self.plot.axes.matshow(simmat, cmap="BuGn", vmin=0.0, vmax=1.0)
			self.plot.figure.colorbar(caxes)
			self.plot.axes.set_xticks(range(ncases), labels=self.data_ids, rotation=25)
			self.plot.axes.set_yticks(range(ncases), labels=self.data_ids)
			if self.show_labels:
				for i in range(ncases):
					for j in range(ncases):
						v = simmat[i,j]
						c = "white" if abs(v) > 0.40 else "black"
						self.plot.axes.text(j, i, f"{simmat[i,j]:.2f}", ha="center", va="center", color=c)
			self.plot.axes.set_title(simtype)
			self.plot.draw()
			# Save the similarities for display
			self.sim_table_labels = [f"{self.id_col_var.get()} 1", f"{self.id_col_var.get()} 2", f"{simtype} similarity"]
			self.sim_table = sim_results
			self.sim_data_btn["state"] = tk.NORMAL
			self.loading_dlg.hide()

	def set_autoupdate(self, *args):
		self.auto_update = self.autoupdate_var.get() == "1"
		if self.auto_update:
			self.q_redraw(None)
	def set_labeling(self, *args):
		self.show_labels = not self.show_labels
		self.q_redraw([], get_data=False)
	def do_close(self, *args):
		self.parent.remove_simmat(self)
		super().do_cancel(args)



class RandomSelDialog(object):
	def __init__(self, n_data_rows, n_selected_rows, last_n, last_usesel):
		self.n_data_rows = n_data_rows
		self.n_selected_rows = n_selected_rows
		self.dlg = tk.Toplevel()
		self.dlg.title("Random Selection")
		self.dlg.rowconfigure(0, weight=1)
		self.dlg.columnconfigure(0, weight=2)
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(6,6))
		prompt_frame.rowconfigure(0, weight=1)
		prompt_frame.columnconfigure(0, weight=2)
		msg_lbl = ttk.Label(prompt_frame, wraplength=100, text="Randomly select data. These selections will replace any previous selections.")
		msg_lbl.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)

		spec_frame = tk.Frame(self.dlg)
		spec_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=(6,6), pady=(6,6))
		entry_lbl = ttk.Label(spec_frame, text="Randomly select:")
		entry_lbl.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		if last_usesel and n_selected_rows > 1:
			self.init_n = min(last_n, n_selected_rows - 1)
			init_sel = "1"
			init_to = n_selected_rows - 1
		else:
			self.init_n = min(last_n, n_data_rows - 1)
			init_sel = "0"
			init_to = n_data_rows - 1
		self.entry_var = tk.IntVar(self.dlg, self.init_n)
		self.val_entry = ttk.Spinbox(spec_frame, textvariable=self.entry_var, from_=1, to=init_to, width=6)
		self.val_entry.grid(row=0, column=1, sticky=tk.W, padx=(3,3), pady=(3,3))
		self.sel_only_var, self.sel_only_ck = add_sel_only(spec_frame, 1, 1, self.check_count)
		self.sel_only_var.set(init_sel)
		if n_selected_rows < 2:
			self.sel_only_ck["state"] = tk.DISABLED

		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=2)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(0,0))
		btn_frame.columnconfigure(0, weight=1)
		# Buttons
		self.canceled = False
		self.ok_btn = new_ok_button(self.dlg, btn_frame, 0, self.do_select, ok_enabled=True)
		cancel_btn = new_cancel_button(self.dlg, btn_frame, 1, self.do_cancel)
		self.val_entry.focus()
	def check_count(self, *args):
		if self.sel_only_var.get() == '1':
			if self.entry_var.get() > self.n_selected_rows - 1:
				self.entry_var.set(min(self.n_selected_rows - 1, self.init_n))
			self.val_entry["to"] = self.n_selected_rows - 1
		else:
			self.val_entry["to"] = self.n_data_rows - 1
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def do_cancel(self, *args):
		self.dlg.destroy()
		self.canceled = True
	def show(self):
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.attributes('-topmost', 'false')
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return None, None
		else:
			return self.entry_var.get(), self.sel_only_var.get() == '1'



class MsgDialog(Dialog):
	def __init__(self, title, message, parent=None, bgcolor=None, can_resize=True, content=False):
		super().__init__(title, message, parent, modal=True)
		if bgcolor is not None:
			self.dlg.configure(bg=bgcolor)
		self.ctrl_frame.destroy()
		if not content:
			self.content_frame.destroy()
		# Buttons
		ok_btn = new_close_button(self.dlg, self.btn_frame, 0, self.do_select)
		self.dlg.resizable(can_resize, can_resize)
		self.dlg.minsize(width=325, height=50)
		ok_btn.focus()


class MsgDialog2(MsgDialog):
	# With an extra content frame.
	def __init__(self, title, message, parent=None, bgcolor=None, can_resize=True):
		super().__init__(title, message, parent, bgcolor, can_resize, content=True)


def show_table(parent, wintitle, msg, tabledata, tablehdrs, sheetname, dlg_args={}, tv_args={}):
	dlg = MsgDialog2(wintitle, msg, parent, **dlg_args)
	tframe, tdata = treeview_table(dlg.content_frame, tabledata, tablehdrs, **tv_args)
	tframe.grid(row=0, column=0, sticky=tk.NSEW)
	def save_data(*args):
		export_data_table(tablehdrs, tabledata, sheetname=sheetname)
		dlg.dlg.lift()
	dlg.dlg.bind("<Control-s>", save_data)
	dlg.show()

def show_columnar_table(parent, wintitle, msg, tabledata, tablehdrs, sheetname, dlg_args={}, tv_args={}):
	rowtable = columns_to_rows(tabledata)
	show_table(parent, wintitle, msg, rowtable, tablehdrs, sheetname, dlg_args, tv_args)


class CustomContentDialog(Dialog):
	def __init__(self, parent, title, prompt):
		super().__init__(title, prompt, parent)
		self.rv = False		# i.e., canceled
		self.ctrl_frame.destroy()
		self.ok_btn = new_ok_button(self.dlg, self.btn_frame, 0, self.do_select, ok_enabled=True)
		cancel_btn = new_cancel_button(self.dlg, self.btn_frame, 1, self.do_cancel)
	def do_select(self, *args):
		self.rv = True
		super().do_select(args)


class OneEntryDialog(Dialog):
	def __init__(self, parent, title, prompt, msgwraplength=80, init_value=None, nullable=False, obscure=False):
		self.nullable = nullable
		super().__init__(title, prompt, parent, msgwraplength)
		self.ctrl_frame.destroy()
		self.entry_var = tk.StringVar(self.dlg, init_value or "")
		self.entry_var.trace('w', self.check_enable)
		self.val_entry = ttk.Entry(self.content_frame, width=50, textvariable=self.entry_var)
		if obscure:
			self.val_entry["show"] = "*"
		self.val_entry.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		self.ok_btn = new_ok_button(self.dlg, self.btn_frame, 4, self.do_select, ok_enabled=True)
		cancel_btn = new_cancel_button(self.dlg, self.btn_frame, 5, self.do_cancel)
		self.dlg.resizable(True, False)
		self.val_entry.focus()
	def check_enable(self, *args):
		if self.nullable:
			self.ok_btn["state"] = tk.NORMAL
		else:
			enable_if(self.ok_btn, self.entry_var.get() != '')
	def do_select(self, *args):
		if self.entry_var.get() != '':
			self.canceled = False
			self.rv = self.entry_var.get()
			self.dlg.destroy()
		elif self.nullable:
			self.canceled = False
			self.rv = None
			self.dlg.destroy()


class OneCheckboxDialog(Dialog):
	def __init__(self, title, prompt, checkbox_value):
		super().__init__(title, prompt, parent)
		self.ctrl_frame.destroy()
		self.check_var = tk.StringVar(self.dlg, "0")
		if checkbox_value:
			self.check_var.set("1")
		self.check_ck = ttk.Checkbutton(self.content_frame, text=prompt, state=tk.NORMAL, variable=self.check_var,
				onvalue="1", offvalue="0")
		self.check_ck.grid(row=0, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		self.ok_btn = new_ok_button(self.dlg, self.btn_frame, 0, self.do_select, ok_enabled=True)
		cancel_btn = new_cancel_button(self.dlg, self.btn_frame, 1, self.do_cancel)
		self.dlg.resizable(False, False)
		self.check_ck.focus()
	def do_select(self, *args):
		if self.check_var.get() != '':
			self.rv = self.check_var.get()
			super().do_select(args)


class OneFloatDialog(Dialog):
	def __init__(self, parent, title, prompt, min_value, max_value, initial):
		super().__init__(title, prompt, parent)
		self.ctrl_frame.destroy()
		self.rv = initial
		self.entry_var = tk.DoubleVar(self.dlg, initial)
		self.val_entry = ttk.Spinbox(self.content_frame, textvariable=self.entry_var, from_=min_value, to=max_value)
		self.val_entry.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		self.ok_btn = new_ok_button(self.dlg, self.btn_frame, 0, self.do_select, ok_enabled=True)
		cancel_btn = new_cancel_button(self.dlg, self.btn_frame, 1, self.do_cancel)
		self.dlg.resizable(True, False)
		self.val_entry.focus()
	def do_select(self, *args):
		self.rv = self.entry_var.get()
		super().do_select(args)


class OneIntDialog(OneFloatDialog):
	def __init__(self, parent, title, prompt, min_value, max_value, initial):
		super().__init__(parent, title, prompt, min_value, max_value, initial)
		self.entry_var = tk.IntVar(self.dlg, initial)
		self.val_entry = ttk.Spinbox(self.content_frame, textvariable=self.entry_var, from_=min_value, to=max_value)
		self.val_entry.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))


class PlotConfigDialog(object):
	# Prompts for configuration settings for plot dialogs
	def __init__(self, show_regr_stats, wrap_width, wrap_underscores):
		self.show_regr_stats = show_regr_stats
		self.wrap_width = wrap_width
		self.wrap_underscores = wrap_underscores
		self.dlg = tk.Toplevel()
		self.dlg.title("Plot Configuration")
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, pady=(3,3))
		prompt_frame.columnconfigure(0, weight=1)

		self.regr_check_var = tk.StringVar(self.dlg, "0" if not show_regr_stats else "1")
		self.regr_check_ck = ttk.Checkbutton(prompt_frame,
				text="Display regression statistics when drawing scatter plots", state=tk.NORMAL, variable=self.regr_check_var,
				onvalue="1", offvalue="0")
		self.regr_check_ck.grid(row=0, column=0, columnspan=2, sticky=tk.W, padx=(6,6), pady=(3,3))

		wrap_msg_lbl = ttk.Label(prompt_frame, text="Text wrapping width for plot tick labels:")
		wrap_msg_lbl.grid(row=1, column=0, sticky=tk.W, padx=(6,6), pady=(3,3))
		self.wrap_var = tk.IntVar(self.dlg, self.wrap_width)
		self.wrap_entry = ttk.Spinbox(prompt_frame, textvariable=self.wrap_var, width=5, from_=5, to=50)
		self.wrap_entry.grid(row=1, column=1, sticky=tk.W, padx=(6,6), pady=(3,3))

		self.wundr_check_var = tk.StringVar(self.dlg, "0" if not wrap_underscores else "1")
		self.wundr_check_ck = ttk.Checkbutton(prompt_frame,
				text="Wrap labels at underscores", state=tk.NORMAL, variable=self.wundr_check_var,
				onvalue="1", offvalue="0")
		self.wundr_check_ck.grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=(6,6), pady=(3,3))

		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.EW, pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		self.regr_check_ck.focus()
		# Buttons
		self.canceled = False
		self.ok_btn = new_ok_button(self.dlg, btn_frame, 0, self.do_select, ok_enabled=True)
		cancel_btn = new_cancel_button(self.dlg, btn_frame, 1, self.do_cancel)
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_select(self, *args):
		self.canceled = False
		self.dlg.destroy()
	def show(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(True, False)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return {"show_regr_stats": self.show_regr_stats, "wrapwidth": self.wrap_width,
					"wrap_underscores": self.wrap_underscores}
		else:
			return {"show_regr_stats": self.regr_check_var.get() == '1', "wrapwidth": self.wrap_var.get(),
					"wrap_underscores": self.wundr_check_var.get() == '1'}



class GetEditorDialog(OneEntryDialog):
	def __init__(self, parent, current_editor):
		super().__init__(parent, "Set Text Editor",
				"Choose the text editor to be used to edit SQL commands.  The editor is set from the EDITOR environment variable on startup, and may also be changed using a configuration file.",
				msgwraplength=600)
		self.rv = current_editor
		if current_editor is not None:
			self.rv = current_editor
			self.entry_var.set(current_editor)
		fn_btn = ttk.Button(self.btn_frame, text="Browse", command=self.get_fn, underline=0)
		fn_btn.grid(row=0, column=0, sticky=tk.W, padx=(3,3))
		self.dlg.bind("<Alt-b>", self.get_fn)
		self.val_entry.focus()
		self.check_enable()
	def get_fn(self, *args):
		fn = tkfiledialog.askopenfilename(parent=self.dlg)
		if fn is not None and fn != '' and fn != ():
			self.entry_var.set(fn)
	def check_enable(self, *args):
		enable_if(self.ok_btn, self.entry_var.get() != '')
	def do_select(self, *args):
		if self.entry_var.get() != '':
			self.rv = self.entry_var.get()
			super().do_select(args)


class HelpHotkeysDialog(object):
	def __init__(self):
		self.dlg = tk.Toplevel()
		self.dlg.title("Hotkeys")
		self.dlg.rowconfigure(0, weight=1)
		self.dlg.columnconfigure(0, weight=2)
		keyframe = tk.Frame(self.dlg)
		keyframe.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(6,6))
		keyframe.rowconfigure(0, weight=1)

		ttk.Label(keyframe, width=10, text="Ctrl-S").grid(row=0, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Save the data tables shown by the plot dialog and the statistics dialogs.").grid(row=0, column=1)
		ttk.Label(keyframe, width=10, text="Ctrl-Z").grid(row=1, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Save the table of univariate statistics for log-transformed data.").grid(row=1, column=1)
		ttk.Label(keyframe, width=10, text="Alt-A").grid(row=2, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Change the opacity (alpha value) of symbols on scatter plots, bubble plots, line plots, stripcharts, Q-Q plots, and KD plots.").grid(row=2, column=1)
		ttk.Label(keyframe, width=10, text="Alt-B").grid(row=3, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="For histograms and binned min-max plots: change the number of bins used. For scatter plots: toggle on or off the display of vertical and horizontal lines delineating the Jenks Natural Breaks in X and Y variables, respectively. For line plots: toggle on or off the display of vertical lines delineating the Jenks Natural Breaks in the X variable.").grid(row=3, column=1)
		ttk.Label(keyframe, width=10, text="Alt-C").grid(row=4, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Toggles display of the count of data rows at each location on the map.").grid(row=4, column=1)
		ttk.Label(keyframe, width=10, text="Alt-F").grid(row=5, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Flip (invert) the Y axis on min-max plots, bar plots, bubble plots, scatter plots, and Y-range plots.").grid(row=5, column=1)
		ttk.Label(keyframe, width=10, text="Alt-G").grid(row=6, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="From the main application (map) window, open the plot dialog.  On a Normal Q-Q plot, toggle the coloring of points to correspond to groups defined by Jenks Natural Breaks.").grid(row=6, column=1)
		ttk.Label(keyframe, width=10, text="Alt-L").grid(row=7, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Toggle the display of a LOESS smoothing line on line plots and scatter plots.  Show or hide correlation coefficients in a correlation matrix.").grid(row=7, column=1)
		ttk.Label(keyframe, width=10, text="Alt-P").grid(row=8, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Open the Plot Data dialog from the map view.").grid(row=8, column=1)
		ttk.Label(keyframe, width=10, text="Alt-Q").grid(row=9, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Open the Query Data dialog from the map view.").grid(row=9, column=1)
		ttk.Label(keyframe, width=10, text="Alt-R").grid(row=10, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="For line plots and scatter plots: toggle the display of an ordinary least-squares linear regression line. For box plots, bar plots, min-max plots, stripcharts and violin plots: rotate the X and Y axes.").grid(row=10, column=1)
		ttk.Label(keyframe, width=10, text="Alt-S").grid(row=11, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Toggle the display of a Theil-Sen line on line and scatter plots, and the plot for the bivariate statistsics summary.").grid(row=11, column=1)
		ttk.Label(keyframe, width=10, text="Alt-T").grid(row=12, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Add or change the plot title.").grid(row=12, column=1)
		ttk.Label(keyframe, width=10, text="Alt-X").grid(row=13, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Change the label on the plot's X axis.").grid(row=13, column=1)
		ttk.Label(keyframe, width=10, text="Alt-Y").grid(row=14, column=0, sticky=tk.N+tk.W)
		ttk.Label(keyframe, width=50, justify=tk.LEFT, wraplength=350, text="Change the label on the plot's Y axis.  For pair plots, rotate the Y-axis label").grid(row=14, column=1)

		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=2)
		btn_frame.grid(row=1, column=0, sticky=tk.EW, pady=(0,0))
		btn_frame.columnconfigure(0, weight=1)
		# Buttons
		self.canceled = False
		ok_btn = new_close_button(self.dlg, btn_frame, 0, self.do_select)
		self.dlg.bind("<Return>", self.do_select)
		self.dlg.resizable(False, False)
		ok_btn.focus()
	def do_select(self, *args):
		self.dlg.destroy()
	def show(self):
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.attributes('-topmost', 'true')
		self.dlg.attributes('-topmost', 'false')
		self.dlg.wait_window(self.dlg)


class SelDataSrcDialog(object):
	def __init__(self, parent, mapui):
		self.parent = parent
		self.mapui = mapui
		self.canceled = False
		self.dlg = tk.Toplevel(parent)
		self.dlg.title("Select Mapping Data")
		self.dlg.protocol("WM_DELETE_WINDOW", self.do_cancel)
		self.dlg.columnconfigure(0, weight=1)
		self.rv = (None, None, None, None, None, None, None, None, None, None)
		# Prompt
		prompt_frame = tk.Frame(self.dlg)
		prompt_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(6,6), pady=(6,3))
		msg_lbl = ttk.Label(prompt_frame, width=30, wraplength=100, anchor=tk.W, justify=tk.LEFT, text="Select the type of data source to use.  You will then be prompted for details about the selected data source.")
		msg_lbl.grid(row=0, column=0, padx=(6,6), pady=(3,3))
		def wrap_msg(event):
			msg_lbl.configure(wraplength=event.width - 5)
		msg_lbl.bind("<Configure>", wrap_msg)
		# Select buttons
		sel_frame = tk.Frame(self.dlg)
		sel_frame.grid(row=1, column=0, sticky=tk.NSEW, pady=(6,9))
		csv_btn = ttk.Button(sel_frame, text=" CSV file  ", command=self.sel_csv)
		csv_btn.grid(row=0, column=0, sticky=tk.EW, padx=(3,3), pady=(3,3))
		ss_btn = ttk.Button(sel_frame,  text="Spreadsheet", command=self.sel_spreadsheet)
		ss_btn.grid(row=0, column=1, sticky=tk.EW, padx=(3,3), pady=(3,3))
		db_btn = ttk.Button(sel_frame,  text=" Database  ", command=self.sel_database)
		db_btn.grid(row=0, column=2, sticky=tk.EW, padx=(3,3), pady=(3,3))
		# Help and Cancel buttons
		btn_frame = tk.Frame(self.dlg, borderwidth=3, relief=tk.RIDGE)
		btn_frame.columnconfigure(0, weight=1)
		btn_frame.grid(row=2, column=0, sticky=tk.S+tk.EW, padx=(3,3), pady=(3,3))
		btn_frame.columnconfigure(0, weight=1)
		self.canceled = False
		new_help_button(self.dlg, btn_frame, self.do_help)
		cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.do_cancel)
		cancel_btn.grid(row=0, column=0, sticky=tk.E, padx=(3,6))
		self.dlg.bind("<Escape>", self.do_cancel)
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/", new=2, autoraise=True)
	def sel_csv(self):
		dfd = DataFileDialog()
		self.rv = dfd.get_datafile()
		if self.rv[0] is not None:
			self.dlg.destroy()
	def sel_spreadsheet(self):
		dfd = ImportSpreadsheetDialog(self.dlg, self.mapui)
		self.rv = dfd.get_datafile()
		if self.rv[0] is not None:
			self.dlg.destroy()
	def sel_database(self):
		dbd = DbConnectDialog(self.dlg, self.mapui)
		self.rv = dbd.get_data()
		if self.rv[0] is not None:
			self.dlg.destroy()
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def select(self):
		self.dlg.grab_set()
		center_window(self.dlg)
		raise_window(self.dlg)
		self.dlg.resizable(False, False)
		self.dlg.focus()
		self.dlg.wait_window(self.dlg)
		return self.rv


class EncodedFile(object):
	# A class providing an open method for an encoded file, allowing reading
	# and writing using unicode, without explicit decoding or encoding.
	def __repr__(self):
		return u"EncodedFile(%r, %r)" % (self.filename, self.encoding)
	def __init__(self, filename, file_encoding):
		self.filename = filename
		self.encoding = file_encoding
		self.bom_length = 0
		def detect_by_bom(path, default_enc):
			with io.open(path, 'rb') as f:
				raw = f.read(4)
			for enc, boms, bom_len in (
							('utf-8-sig', (codecs.BOM_UTF8,), 3),
							('utf_16', (codecs.BOM_UTF16_LE, codecs.BOM_UTF16_BE), 2),
							('utf_32', (codecs.BOM_UTF32_LE, codecs.BOM_UTF32_BE), 4)):
				if any(raw.startswith(bom) for bom in boms):
					return enc, bom_len
			return default_enc, 0
		if os.path.exists(filename):
			self.encoding, self.bom_length = detect_by_bom(filename, file_encoding)
		self.fo = None
	def open(self, mode='r'):
		self.fo = io.open(file=self.filename, mode=mode, encoding="UTF8", newline=None)
		return self.fo
	def close(self):
		if self.fo is not None:
			self.fo.close()


class LineDelimiter(object):
	def __init__(self, delim, quote, escchar):
		self.delimiter = delim
		self.joinchar = delim if delim else u""
		self.quotechar = quote
		if quote:
			if escchar:
				self.quotedquote = escchar+quote
			else:
				self.quotedquote = quote+quote
		else:
			self.quotedquote = None
	def delimited(self, datarow, add_newline=True):
		global conf
		if self.quotechar:
			d_row = []
			for e in datarow:
				if isinstance(e, str):
					if (self.quotechar in e) or (self.delimiter is not None and self.delimiter in e) or (u'\n' in e) or (u'\r' in e):
						d_row.append(u"%s%s%s" % (self.quotechar, e.replace(self.quotechar, self.quotedquote), self.quotechar))
					else:
						d_row.append(e)
				else:
					if e is None:
						d_row.append('')
					else:
						d_row.append(e)
			text = self.joinchar.join([type(u"")(d) for d in d_row])
		else:
			d_row = []
			for e in datarow:
				if e is None:
					d_row.append('')
				else:
					d_row.append(e)
			text = self.joinchar.join([type(u"")(d) for d in d_row])
		if add_newline:
			text = text + u"\n"
		return text


def write_delimited_file(outfile, filefmt, column_headers, rowsource, file_encoding='utf8', append=False):
	delim = None
	quote = None
	escchar = None
	if filefmt.lower() == 'csv':
		delim = ","
		quote = '"'
		escchar = None
	elif filefmt.lower() in ('tab', 'tsv'):
		delim = "\t"
		quote = None
		escchar = None
	elif filefmt.lower() in ('tabq', 'tsvq'):
		delim = "\t"
		quote = '"'
		escchar = None
	elif filefmt.lower() in ('unitsep', 'us'):
		delim = chr(31)
		quote = None
		escchar = None
	elif filefmt.lower() == 'plain':
		delim = " "
		quote = ''
		escchar = None
	elif filefmt.lower() == 'tex':
		delim = "&"
		quote = ''
		escchar = None
	line_delimiter = LineDelimiter(delim, quote, escchar)
	fmode = "w" if not append else "a"
	ofile = EncodedFile(outfile, file_encoding).open(mode=fmode)
	fdesc = outfile
	if not (filefmt.lower() == 'plain' or append):
		datarow = line_delimiter.delimited(column_headers)
		ofile.write(datarow)
	for rec in rowsource:
		datarow = line_delimiter.delimited(rec)
		ofile.write(datarow)
	ofile.close()



class OdsFile(object):
	def __repr__(self):
		return u"OdsFile()"
	def __init__(self):
		self.filename = None
		self.wbk = None
		self.cell_style_names = []
	def open(self, filename):
		self.filename = filename
		if os.path.isfile(filename):
			self.wbk = odf.opendocument.load(filename)
			# Get a list of all cell style names used, so as not to re-define them.
			# Adapted from http://www.pbertrand.eu/reading-an-odf-document-with-odfpy/
			for sty in self.wbk.automaticstyles.childNodes:
				try:
					fam = sty.getAttribute("family")
					if fam == "table-cell":
						name = sty.getAttribute("name")
						if not name in self.cell_style_names:
							self.cell_style_names.append(name)
				except:
					pass
		else:
			self.wbk = odf.opendocument.OpenDocumentSpreadsheet()
	def define_body_style(self):
		st_name = "body"
		if not st_name in self.cell_style_names:
			body_style = odf.style.Style(name=st_name, family="table-cell")
			body_style.addElement(odf.style.TableCellProperties(attributes={"verticalalign":"top"}))
			self.wbk.styles.addElement(body_style)
			self.cell_style_names.append(st_name)
	def define_header_style(self):
		st_name = "header"
		if not st_name in self.cell_style_names:
			header_style = odf.style.Style(name=st_name, family="table-cell")
			header_style.addElement(odf.style.TableCellProperties(attributes={"borderbottom":"1pt solid #000000",
				"verticalalign":"bottom"}))
			self.wbk.styles.addElement(header_style)
			self.cell_style_names.append(st_name)
	def define_iso_datetime_style(self):
		st_name = "iso_datetime"
		if not st_name in self.cell_style_names:
			dt_style = odf.number.DateStyle(name="iso-datetime")
			dt_style.addElement(odf.number.Year(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Month(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Day(style="long"))
			# odfpy collapses text elements that have only spaces, so trying to insert just a space between the date
			# and time actually results in no space between them.  Other Unicode invisible characters
			# are also trimmed.  The delimiter "T" is used instead, and conforms to ISO-8601 specifications.
			dt_style.addElement(odf.number.Text(text=u"T"))
			dt_style.addElement(odf.number.Hours(style="long"))
			dt_style.addElement(odf.number.Text(text=u":"))
			dt_style.addElement(odf.number.Minutes(style="long"))
			dt_style.addElement(odf.number.Text(text=u":"))
			dt_style.addElement(odf.number.Seconds(style="long", decimalplaces="3"))
			self.wbk.styles.addElement(dt_style)
			self.define_body_style()
			dts = odf.style.Style(name=st_name, datastylename="iso-datetime", parentstylename="body", family="table-cell")
			self.wbk.automaticstyles.addElement(dts)
			self.cell_style_names.append(st_name)
	def define_iso_date_style(self):
		st_name = "iso_date"
		if st_name not in self.cell_style_names:
			dt_style = odf.number.DateStyle(name="iso-date")
			dt_style.addElement(odf.number.Year(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Month(style="long"))
			dt_style.addElement(odf.number.Text(text=u"-"))
			dt_style.addElement(odf.number.Day(style="long"))
			self.wbk.styles.addElement(dt_style)
			self.define_body_style()
			dts = odf.style.Style(name=st_name, datastylename="iso-date", parentstylename="body", family="table-cell")
			self.wbk.automaticstyles.addElement(dts)
			self.cell_style_names.append(st_name)
	def sheetnames(self):
		# Returns a list of the worksheet names in the specified ODS spreadsheet.
		return [sheet.getAttribute("name") for sheet in self.wbk.spreadsheet.getElementsByType(odf.table.Table)]
	def sheet_named(self, sheetname):
		# Return the sheet with the matching name.  If the name is actually an integer,
		# return that sheet number.
		if isinstance(sheetname, int):
			sheet_no = sheetname
		else:
			try:
				sheet_no = int(sheetname)
				if sheet_no < 1:
					sheet_no = None
			except:
				sheet_no = None
		if sheet_no is not None:
			for i, sheet in enumerate(self.wbk.spreadsheet.getElementsByType(odf.table.Table)):
				if i+1 == sheet_no:
					return sheet
			else:
				sheet_no = None
		if sheet_no is None:
			for sheet in self.wbk.spreadsheet.getElementsByType(odf.table.Table):
				if sheet.getAttribute("name").lower() == sheetname.lower():
					return sheet
		return None
	def sheet_data(self, sheetname, junk_header_rows=0):
		sheet = self.sheet_named(sheetname)
		if not sheet:
			warning("There is no sheet named %s" % sheetname, kwargs={})
			raise
		def row_data(sheetrow):
			# Adapted from http://www.marco83.com/work/wp-content/uploads/2011/11/odf-to-array.py
			cells = sheetrow.getElementsByType(odf.table.TableCell)
			rowdata = []
			for cell in cells:
				p_content = []
				repeat = cell.getAttribute("numbercolumnsrepeated")
				if not repeat:
					repeat = 1
					spanned = int(cell.getAttribute("numbercolumnsspanned") or 0)
					if spanned > 1:
						repeat = spanned
				ps = cell.getElementsByType(odf.text.P)
				if len(ps) == 0:
					for rr in range(int(repeat)):
						p_content.append(None)
				else:
					for p in ps:
						pval = type(u"")(p)
						if len(pval) == 0:
							for rr in range(int(repeat)):
								p_content.append(None)
						else:
							for rr in range(int(repeat)):
								p_content.append(pval)
				if len(p_content) == 0:
					for rr in range(int(repeat)):
						rowdata.append(None)
				elif p_content[0] != u'#':
					rowdata.extend(p_content)
			return rowdata
		rows = sheet.getElementsByType(odf.table.TableRow)
		if junk_header_rows > 0:
			rows = rows[junk_header_rows: ]
		return [row_data(r) for r in rows]
	def new_sheet(self, sheetname):
		# Returns a sheet (a named Table) that has not yet been added to the workbook
		return odf.table.Table(name=sheetname)
	def add_row_to_sheet(self, datarow, odf_table, header=False):
		if header:
			self.define_header_style()
			style_name = "header"
		else:
			self.define_body_style()
			style_name = "body"
		tr = odf.table.TableRow()
		odf_table.addElement(tr)
		for item in datarow:
			if isinstance(item, bool):
				# Booleans must be evaluated before numbers.
				# Neither of the first two commented-out lines actually work (a bug in odfpy?).
				# Booleans *can* be written as either integers or strings; integers are chosen below.
				#tc = odf.table.TableCell(booleanvalue='true' if item else 'false')
				#tc = odf.table.TableCell(valuetype="boolean", value='true' if item else 'false')
				tc = odf.table.TableCell(valuetype="boolean", value=1 if item else 0, stylename=style_name)
				#tc = odf.table.TableCell(valuetype="string", stringvalue='True' if item else 'False')
			elif isinstance(item, float) or isinstance(item, int):
				tc = odf.table.TableCell(valuetype="float", value=item, stylename=style_name)
			elif isinstance(item, datetime.datetime):
				self.define_iso_datetime_style()
				tc = odf.table.TableCell(valuetype="date", datevalue=item.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3], stylename="iso_datetime")
			elif isinstance(item, datetime.date):
				self.define_iso_date_style()
				tc = odf.table.TableCell(valuetype="date", datevalue=item.strftime("%Y-%m-%d"), stylename="iso_date")
			elif isinstance(item, datetime.time):
				self.define_iso_datetime_style()
				timeval = datetime.datetime(1899, 12, 30, item.hour, item.minute, item.second, item.microsecond, item.tzinfo)
				tc = odf.table.TableCell(timevalue=timeval.strftime("PT%HH%MM%S.%fS"), stylename="iso_datetime")
				tc.addElement(odf.text.P(text=timeval.strftime("%H:%M:%S.%f")))
			elif isinstance(item, str):
				item = item.replace(u'\n', u' ').replace(u'\r', u' ')
				tc = odf.table.TableCell(valuetype="string", stringvalue=item, stylename=style_name)
			else:
				tc = odf.table.TableCell(value=item, stylename=style_name)
			if item is not None:
				tc.addElement(odf.text.P(text=item))
			tr.addElement(tc)
	def add_sheet(self, odf_table):
		self.wbk.spreadsheet.addElement(odf_table)
	def save_close(self):
		ofile = io.open(self.filename, "wb")
		self.wbk.write(ofile)
		ofile.close()
		self.filename = None
		self.wbk = None
	def close(self):
		self.filename = None
		self.wbk = None


def ods_data(filename, sheetname, junk_header_rows=0):
	# Returns the data from the specified worksheet as a list of headers and a list of lists of rows.
	wbk = OdsFile()
	try:
		wbk.open(filename)
	except:
		warning("%s is not a valid OpenDocument spreadsheet." % filename, kwargs={})
		raise
	try:
		alldata = wbk.sheet_data(sheetname, junk_header_rows)
	except:
		warning("%s is not a worksheet in %s." % (sheetname, filename), kwargs={})
		raise
	colhdrs = alldata[0]
	if any([x is None or len(x.strip())==0 for x in colhdrs]):
		if conf.del_empty_cols:
			blanks = [i for i in range(len(colhdrs)) if colhdrs[i] is None or len(colhdrs[i].strip())==0]
			while len(blanks) > 0:
				b = blanks.pop()
				for r in range(len(alldata)):
					del(alldata[r][b])
			colhdrs = alldata[0]
		else:
			if conf.create_col_hdrs:
				for i in range(len(colhdrs)):
					if colhdrs[i] is None or len(colhdrs[i]) == 0:
						colhdrs[i] = "Col%s" % str(i+1)
			else:
				warning("The input file %s, sheet %s has missing column headers." % (filename, sheetname), kwargs={})
				raise
	return colhdrs, alldata[1:]


def export_ods(outfile, hdrs, rows, append=False, querytext=None, sheetname=None, desc=None):
	# If not given, determine the worksheet name to use.  The pattern is "Sheetx", where x is
	# the first integer for which there is not already a sheet name.
	if append and os.path.isfile(outfile):
		wbk = OdsFile()
		wbk.open(outfile)
		sheet_names = wbk.sheetnames()
		name = sheetname or u"Sheet"
		sheet_name = name
		sheet_no = 1
		while True:
			if sheet_name not in sheet_names:
				break
			sheet_no += 1
			sheet_name = u"%s%d" % (name, sheet_no)
		wbk.close()
	else:
		sheet_name = sheetname or u"Sheet1"
		if os.path.isfile(outfile):
			os.unlink(outfile)
	wbk = OdsFile()
	wbk.open(outfile)
	# Add the data to a new sheet.
	tbl = wbk.new_sheet(sheet_name)
	wbk.add_row_to_sheet(hdrs, tbl, header=True)
	for row in rows:
		wbk.add_row_to_sheet(row, tbl)
	# Add sheet to workbook
	wbk.add_sheet(tbl)
	# Save and close the workbook.
	wbk.save_close()


class XlsFile(object):
	def __repr__(self):
		return u"XlsFile()"
	def __init__(self):
		self.filename = None
		self.encoding = None
		self.wbk = None
		self.datemode = 0
	def open(self, filename, encoding=None, read_only=False):
		self.filename = filename
		self.encoding = encoding
		self.read_only = read_only
		self.wbk = xlrd.open_workbook(filename, encoding_override=self.encoding)
		self.datemode = self.wbk.datemode
	def sheetnames(self):
		return self.wbk.sheet_names()
	def sheet_named(self, sheetname):
		# Return the sheet with the matching name.  If the name is actually an integer,
		# return that sheet number.
		if isinstance(sheetname, int):
			sheet_no = sheetname
		else:
			try:
				sheet_no = int(sheetname)
				if sheet_no < 1:
					sheet_no = None
			except:
				sheet_no = None
		if sheet_no is None:
			sheet = self.wbk.sheet_by_name(sheetname)
		else:
			# User-specified sheet numbers should be 1-based; xlrd sheet indexes are 0-based
			sheet = self.wbk.sheet_by_index(max(0, sheet_no-1))
		return sheet
	def sheet_data(self, sheetname, junk_header_rows=0):
		sheet = self.sheet_named(sheetname)
		# Don't rely on sheet.ncols and sheet.nrows, because Excel will count columns
		# and rows that have ever been filled, even if they are now empty.  Base the column count
		# on the number of contiguous non-empty cells in the first row, and process the data up to nrows until
		# a row is entirely empty.
		def row_data(sheetrow, columns=None):
			cells = sheet.row_slice(sheetrow)
			if columns:
				d = [cells[c] for c in range(columns)]
			else:
				d = [cell for cell in cells]
			datarow = []
			for c in d:
				if c.ctype == 0:
					# empty
					datarow.append(None)
				elif c.ctype == 1:
					datarow.append(c.value)
				elif c.ctype == 2:
					# float, but maybe should be int
					if c.value - int(c.value) == 0:
						datarow.append(int(c.value))
					else:
						datarow.append(c.value)
				elif c.ctype == 3:
					# date
					dt = xlrd.xldate_as_tuple(c.value, self.datemode)
					# Convert to time or datetime
					if not any(dt[:3]):
						# No date values
						datarow.append(datetime.time(*dt[3:]))
					else:
						datarow.append(datetime.datetime(*dt))
				elif c.ctype == 4:
					# Boolean
					datarow.append(bool(c.value))
				elif c.ctype == 5:
					# Error code
					datarow.append(xlrd.error_text_from_code(c.value))
				elif c.ctype == 6:
					# blank
					datarow.append(None)
				else:
					datarow.append(c.value)
			return datarow
		hdr_row = row_data(junk_header_rows)
		ncols = 0
		for c in range(len(hdr_row)):
			if not hdr_row[c]:
				break
			ncols += 1
		sheet_data = []
		for r in range(junk_header_rows, sheet.nrows - junk_header_rows):
			datarow = row_data(r, ncols)
			if datarow.count(None) == len(datarow):
				break
			sheet_data.append(datarow)
		return sheet_data


class XlsxFile(object):
	def __repr__(self):
		return u"XlsxFile()"
	def __init__(self):
		self.filename = None
		self.encoding = None
		self.wbk = None
		self.read_only = False
	def open(self, filename, encoding=None, read_only=False):
		self.filename = filename
		self.encoding = encoding
		self.read_only = read_only
		if os.path.isfile(filename):
			if read_only:
				self.wbk = openpyxl.load_workbook(filename, read_only=True)
			else:
				self.wbk = openpyxl.load_workbook(filename)
	def close(self):
		if self.wbk is not None:
			self.wbk.close()
			self.wbk = None
			self.filename = None
			self.encoding = None
	def sheetnames(self):
		return self.wbk.sheetnames
	def sheet_named(self, sheetname):
		# Return the sheet with the matching name.  If the name is actually an integer,
		# return that sheet number.
		if isinstance(sheetname, int):
			sheet_no = sheetname
		else:
			try:
				sheet_no = int(sheetname)
				if sheet_no < 1:
					sheet_no = None
			except:
				sheet_no = None
		if sheet_no is not None:
			# User-specified sheet numbers should be 1-based
			sheet = self.wbk[self.wbk.sheetnames[sheet_no - 1]]
		else:
			sheet = self.wbk[sheetname]
		return sheet
	def sheet_data(self, sheetname, junk_header_rows=0):
		sheet = self.sheet_named(sheetname)
		# Don't rely on sheet.max_column and sheet.max_row, because Excel will count columns
		# and rows that have ever been filled, even if they are now empty.  Base the column count
		# on the number of contiguous non-empty cells in the first row, and process the data up to nrows until
		# a row is entirely empty.
		# Get the header row, skipping junk rows
		rowsrc = sheet.iter_rows(max_row = junk_header_rows + 1, values_only = True)
		for hdr_row in rowsrc:
			pass
		# Get the number of columns
		ncols = 0
		for c in range(len(hdr_row)):
			if not hdr_row[c]:
				break
			ncols += 1
		# Get all the data rows
		sheet_data = []
		rowsrc = sheet.iter_rows(min_row = junk_header_rows + 1, values_only = True)
		for r in rowsrc:
			if not any(r):
				break
			sheet_data.append(list(r))
		for r in range(len(sheet_data)):
			rd = sheet_data[r]
			for c in range(len(rd)):
				if isinstance(rd[c], str):
					if rd[c] == '=FALSE()':
						rd[c] = False
					elif rd[c] == '=TRUE()':
						rd[c] = True
		return sheet_data


def xls_data(filename, sheetname, junk_header_rows, encoding=None):
	# Returns the data from the specified worksheet as a list of headers and a list of lists of rows.
	root, ext = os.path.splitext(filename)
	ext = ext.lower()
	if ext == ".xls":
		wbk = XlsFile()
	else:
		wbk = XlsxFile()
	try:
		wbk.open(filename, encoding, read_only=True)
	except:
		warning("%s is not a valid Excel spreadsheet." % filename, kwargs={})
		raise
	alldata = wbk.sheet_data(sheetname, junk_header_rows)
	if len(alldata) == 0:
		raise ErrInfo(type="cmd", other_msg="There are no data on worksheet %s of file %s." % (sheetname, filename))
	if ext == 'xlsx':
		wbk.close()
	if len(alldata) == 1:
		return alldata[0], []
	colhdrs = alldata[0]
	# Delete columns with missing headers
	if any([x is None or (isinstance(x, str) and len(x.strip())==0) for x in colhdrs]):
		blanks = [i for i in range(len(colhdrs)) if colhdrs[i] is None or len(colhdrs[i].strip())==0]
		while len(blanks) > 0:
			b = blanks.pop()
			for r in range(len(alldata)):
				del(alldata[r][b])
		colhdrs = alldata[0]
	return colhdrs, alldata[1:]



def file_data(filename, junk_headers=0):
	# Get headers and rows from the specified CSV file
	csvreader = CsvFile(filename, junk_header_lines=junk_headers)
	headers = csvreader.next()
	rows = []
	for line in csvreader:
		if len(line) > 0:
			rows.append(line)
	return headers, rows


#***************************************************************************************************
#***************************  SQL Scripting Extensions  ********************************************
#***************************************************************************************************

# Support for SQL scripts used to obtain a data table from a database.
# These are a subset of features in execsql.py.

#===============================================================================================
#-----  GLOBAL VARIABLES FOR SQL INTERPRETYER

# Other variables are defined in the context of further class and function definitions.

# A list of errors found while processing the SQL script.  Each item in this list is
# a two-element list consiting of a) a description of the error, and b) the line number of the error.
script_errors = []

# The last command run.  This should be a ScriptCmd object.
last_command = None

# A compiled regex to match prefixed regular expressions, used to check
# for unsubstituted variables.  This is global rather than local to SqlStmt and
# MetacommandStmt objects because Python 2 can't deepcopy a compiled regex.
varlike = re.compile(r'!![$@&~#]?\w+!!', re.I)

# A ScriptExecSpec object for a script to be executed when the program halts due to an error.
# This is intially None, but may be set and re-set by metacommands.
err_halt_exec = None

# A ScriptExecSpec object for a script to be executed when the program halts due
# user cancellation.
# This is intially None, but may be set and re-set by metacommands.
cancel_halt_exec = None

# A stack of the CommandList objects currently in the queue to be executed.
# The list on the top of the stack is the currently executing script.
commandliststack = []

# A dictionary of CommandList objects (ordinarily created by
# BEGIN/END SCRIPT metacommands) that may be inserted into the
# commandliststack.
savedscripts = {}

# A stack of CommandList objects that are used when compiling the
# statements within a loop (between LOOP and END LOOP metacommands).
loopcommandstack = []
# A global flag to indicate that commands should be compiled into
# the topmost entry in the loopcommandstack rather than executed.
compiling_loop = False
# Compiled regex for END LOOP metacommand, which is immediate.
endloop_rx = re.compile(r'^\s*END\s+LOOP\s*$', re.I)
# Compiled regex for *start of* LOOP metacommand, for testing
# while compiling commands within a loop.
loop_rx = re.compile(r'\s*LOOP\s+', re.I)
# Nesting counter, to ensure loops are only ended when nesting
# level is zero.
loop_nest_level = 0

# A count of all of the commands run.
cmds_run = 0

# Pattern for deferred substitution, e.g.: "!{somevar}!"
defer_rx = re.compile(r'(!{([$@&~#]?[a-z0-9_]+)}!)', re.I)

#	End of global variables (1) for execsql interpreter
#===============================================================================================


#===============================================================================================
#-----  CONFIGURATION

class ConfigData(object):
	def __init__(self):
		self.db_encoding = None
		self.script_encoding = 'utf8'
		self.output_encoding = 'utf8'
		self.import_encoding = 'utf8'
		self.empty_rows = True
		self.del_empty_cols = False
		self.create_col_hdrs = False
		self.trim_col_hdrs = 'none'
		self.clean_col_hdrs = False
		self.fold_col_hdrs = 'no'
		self.dedup_col_hdrs = False
		self.trim_strings = False
		self.replace_newlines = False
		self.export_row_buffer = 1000

#	End of configuration for execsql interpreter
#===============================================================================================
	
#===============================================================================================
#-----  SUPPORT FUNCTIONS AND CLASSES (1)

def ins_rxs(rx_list, fragment1, fragment2):
	# Returns a tuple of all strings consisting of elements of the 'rx_list' tuple
	# inserted between 'fragment1' and 'fragment2'.  The fragments may themselves
	# be tuples.
	if type(fragment1) != tuple:
		fragment1 = (fragment1, )
	if fragment2 is None:
		fragment2 = ('', )
	if type(fragment2) != tuple:
		fragment2 = (fragment2, )
	rv = []
	for te in rx_list:
		for f1 in fragment1:
			for f2 in fragment2:
				rv.append(f1 + te + f2)
	return tuple(rv)

def ins_quoted_rx(fragment1, fragment2, rx):
	return ins_rxs((rx, r'"%s"' % rx), fragment1, fragment2)

def ins_schema_rxs(fragment1, fragment2, suffix=None):
	schema_exprs = (r'"(?P<schema>[A-Za-z0-9_\- ]+)"',
					r'(?P<schema>[A-Za-z0-9_\-]+)',
					r'\[(?P<schema>[A-Za-z0-9_\- ]+)\]'
					)
	if suffix:
		schema_exprs = tuple([s.replace("schema", "schema"+suffix) for s in schema_exprs])
	return ins_rxs(schema_exprs, fragment1, fragment2)

def ins_table_rxs(fragment1, fragment2, suffix=None):
	tbl_exprs = (r'(?:"(?P<schema>[A-Za-z0-9_\- ]+)"\.)?"(?P<table>[A-Za-z0-9_\-\# ]+)"',
					r'(?:(?P<schema>[A-Za-z0-9_\-]+)\.)?(?P<table>[A-Za-z0-9_\-\#]+)',
					r'(?:"(?P<schema>[A-Za-z0-9_\- ]+)"\.)?(?P<table>[A-Za-z0-9_\-\#]+)',
					r'(?:(?P<schema>[A-Za-z0-9_\-]+)\.)?"(?P<table>[A-Za-z0-9_\-\# ]+)"',
					r'(?:\[(?P<schema>[A-Za-z0-9_\- ]+)\]\.)?\[(?P<table>[A-Za-z0-9_\-\# ]+)\]',
					r'(?:(?P<schema>[A-Za-z0-9_\-]+)\.)?(?P<table>[A-Za-z0-9_\-\#]+)',
					r'(?:\[(?P<schema>[A-Za-z0-9_\- ]+)\]\.)?(?P<table>[A-Za-z0-9_\-\#]+)',
					r'(?:(?P<schema>[A-Za-z0-9_\-]+)\.)?\[(?P<table>[A-Za-z0-9_\-\# ]+)\]'
					)
	if suffix:
		tbl_exprs = tuple([s.replace("schema", "schema"+suffix).replace("table", "table"+suffix) for s in tbl_exprs])
	return ins_rxs(tbl_exprs, fragment1, fragment2)

def ins_table_list_rxs(fragment1, fragment2):
	tbl_exprs = (r'(?:(?P<tables>(?:"[A-Za-z0-9_\- ]+"\.)?"[A-Za-z0-9_\-\# ]+"(?:\s*,\s*(?:"[A-Za-z0-9_\- ]+"\.)?"[A-Za-z0-9_\-\# ]+")*))',
				r'(?:(?P<tables>(?:[A-Za-z0-9_\-]+\.)?[A-Za-z0-9_\-\#]+(?:\s*,\s*(?:[A-Za-z0-9_\-]+\.)?[A-Za-z0-9_\-\#]+)*))'
				)
	return ins_rxs(tbl_exprs, fragment1, fragment2)


def ins_fn_rxs(fragment1, fragment2, symbolicname="filename"):
	if os.name == 'posix':
		fns = (r'(?P<%s>[\w\.\-\\\/\'~`!@#$^&()+={}\[\]:;,]*[\w\.\-\\\/\'~`!@#$^&(+={}\[\]:;,])' % symbolicname, r'"(?P<%s>[\w\s\.\-\\\/\'~`!@#$^&()+={}\[\]:;,]+)"' % symbolicname)
	else:
		fns = (r'(?P<%s>([A-Z]\:)?[\w+\,()!@#$^&\+=;\'{}\[\]~`\.\-\\\/]*[\w+\,(!@#$^&\+=;\'{}\[\]~`\.\-\\\/])' % symbolicname, r'"(?P<%s>([A-Z]\:)?[\w+\,()!@#$^&\+=;\'{}\[\]~`\s\.\-\\\/]+)"' % symbolicname)
	return ins_rxs(fns, fragment1, fragment2)


dt_fmts = collections.deque((
			"%x %X",
			"%m/%d/%Y %H:%M",
			"%m/%d/%Y %H%M",
			"%m/%d/%Y %H:%M:%S",
			"%Y-%m-%d %H:%M:%S",
			"%Y-%m-%dT%H:%M:%S",
			"%Y-%m-%d %H%M",
			"%Y-%m-%d %H:%M",
			"%Y-%m-%d %I:%M%p",
			"%Y-%m-%d %I:%M %p",
			"%Y-%m-%d %I:%M:%S%p",
			"%Y-%m-%d %I:%M:%S %p",
			"%m/%d/%Y %I:%M%p",
			"%m/%d/%Y %I:%M %p",
			"%m/%d/%Y %I:%M:%S%p",
			"%m/%d/%Y %I:%M:%S %p",
			"%Y/%m/%d %H%M",
			"%Y/%m/%d %H:%M",
			"%Y/%m/%d %H:%M:%S",
			"%Y/%m/%d %I:%M%p",
			"%Y/%m/%d %I:%M %p",
			"%Y/%m/%d %I:%M:%S%p",
			"%Y/%m/%d %I:%M:%S %p",
			"%Y/%m/%d %X",
			"%c",
			"%b %d, %Y %X",
			"%b %d, %Y %I:%M %p",
			"%b %d %Y %X",
			"%b %d %Y %I:%M %p",
			"%d %b, %Y %X",
			"%d %b, %Y %I:%M %p",
			"%d %b %Y %X",
			"%d %b %Y %I:%M %p",
			"%b. %d, %Y %X",
			"%b. %d, %Y %I:%M %p",
			"%b. %d %Y %X",
			"%b. %d %Y %I:%M %p",
			"%d %b., %Y %X",
			"%d %b., %Y %I:%M %p",
			"%d %b. %Y %X",
			"%d %b. %Y %I:%M %p",
			"%B %d, %Y %X",
			"%B %d, %Y %I:%M %p",
			"%B %d %Y %X",
			"%B %d %Y %I:%M %p",
			"%d %B, %Y %X",
			"%d %B, %Y %I:%M %p",
			"%d %B %Y %X",
			"%d %B %Y %I:%M %p"
			))
def parse_datetime(datestr):
	if type(datestr) == datetime.datetime:
		return datestr
	if not isinstance(datestr, str):
		try:
			datestr = str(datestr)
		except:
			return None
	dt = None
	for i, f in enumerate(dt_fmts):
		try:
			dt = datetime.datetime.strptime(datestr, f)
		except:
			continue
		break
	if i:
		del dt_fmts[i]
		dt_fmts.appendleft(f)
	return dt

dtzrx = re.compile(r"(.+)\s*([+-])(\d{1,2}):?(\d{2})$")
timestamptz_fmts = collections.deque((
	"%Y-%m-%d %H%M%Z", "%Y-%m-%d %H%M %Z",
	"%m/%d/%Y%Z", "%m/%d/%Y %Z",
	"%m/%d/%y%Z", "%m/%d/%y %Z",
	"%m/%d/%Y %H%M%Z", "%m/%d/%Y %H%M %Z",
	"%m/%d/%Y %H:%M%Z", "%m/%d/%Y %H:%M %Z",
	"%Y-%m-%dT%H%M%Z", "%Y-%m-%dT%H%M %Z",
	"%Y-%m-%d %H:%M%Z", "%Y-%m-%d %H:%M %Z",
	"%Y-%m-%dT%H:%M%Z", "%Y-%m-%dT%H:%M %Z",
	"%Y-%m-%d %H:%M:%S%Z", "%Y-%m-%d %H:%M:%S %Z",
	"%Y-%m-%dT%H:%M:%S%Z", "%Y-%m-%dT%H:%M:%S %Z",
	"%Y-%m-%d %I:%M%p%Z", "%Y-%m-%d %I:%M%p %Z",
	"%Y-%m-%dT%I:%M%p%Z", "%Y-%m-%dT%I:%M%p %Z",
	"%Y-%m-%d %I:%M %p%Z", "%Y-%m-%d %I:%M %p %Z",
	"%Y-%m-%dT%I:%M %p%Z", "%Y-%m-%dT%I:%M %p %Z",
	"%Y-%m-%d %I:%M:%S%p%Z", "%Y-%m-%d %I:%M:%S%p %Z",
	"%Y-%m-%dT%I:%M:%S%p%Z", "%Y-%m-%dT%I:%M:%S%p %Z",
	"%Y-%m-%d %I:%M:%S %p%Z", "%Y-%m-%d %I:%M:%S %p %Z",
	"%Y-%m-%dT%I:%M:%S %p%Z", "%Y-%m-%dT%I:%M:%S %p %Z",
	"%c%Z", "%c %Z",
	"%x %X%Z", "%x %X %Z",
	"%m/%d/%Y %H:%M:%S%Z", "%m/%d/%Y %H:%M:%S %Z",
	"%m/%d/%Y %I:%M%p%Z", "%m/%d/%Y %I:%M%p %Z",
	"%m/%d/%Y %I:%M %p%Z", "%m/%d/%Y %I:%M %p %Z",
	"%m/%d/%Y %I:%M:%S%p%Z", "%m/%d/%Y %I:%M:%S%p %Z",
	"%m/%d/%Y %I:%M:%S %p%Z", "%m/%d/%Y %I:%M:%S %p %Z",
	"%Y/%m/%d %H%M%Z", "%Y/%m/%d %H%M %Z",
	"%Y/%m/%d %H:%M%Z", "%Y/%m/%d %H:%M %Z",
	"%Y/%m/%d %H:%M:%S%Z", "%Y/%m/%d %H:%M:%S %Z",
	"%Y/%m/%d %I:%M%p%Z", "%Y/%m/%d %I:%M%p %Z",
	"%Y/%m/%d %I:%M %p%Z", "%Y/%m/%d %I:%M %p %Z",
	"%Y/%m/%d %I:%M:%S%p%Z", "%Y/%m/%d %I:%M:%S%p %Z",
	"%Y/%m/%d %I:%M:%S %p%Z", "%Y/%m/%d %I:%M:%S %p %Z",
	"%Y/%m/%d %X%Z", "%Y/%m/%d %X %Z",
	"%b %d, %Y %X%Z", "%b %d, %Y %X %Z",
	"%b %d, %Y %I:%M %p%Z", "%b %d, %Y %I:%M %p %Z",
	"%b %d %Y %X%Z", "%b %d %Y %X %Z",
	"%b %d %Y %I:%M %p%Z", "%b %d %Y %I:%M %p %Z",
	"%d %b, %Y %X%Z", "%d %b, %Y %X %Z",
	"%d %b, %Y %I:%M %p%Z", "%d %b, %Y %I:%M %p %Z",
	"%d %b %Y %X%Z", "%d %b %Y %X %Z",
	"%d %b %Y %I:%M %p%Z", "%d %b %Y %I:%M %p %Z",
	"%b. %d, %Y %X%Z", "%b. %d, %Y %X %Z",
	"%b. %d, %Y %I:%M %%Z", "%b. %d, %Y %I:%M %p %Z",
	"%b. %d %Y %X%Z", "%b. %d %Y %X %Z",
	"%b. %d %Y %I:%M %p%Z", "%b. %d %Y %I:%M %p %Z",
	"%d %b., %Y %X%Z", "%d %b., %Y %X %Z",
	"%d %b., %Y %I:%M %p%Z", "%d %b., %Y %I:%M %p %Z",
	"%d %b. %Y %X%Z", "%d %b. %Y %X %Z",
	"%d %b. %Y %I:%M %p%Z", "%d %b. %Y %I:%M %p %Z",
	"%B %d, %Y %X%Z", "%B %d, %Y %X %Z",
	"%B %d, %Y %I:%M %p%Z", "%B %d, %Y %I:%M %p %Z",
	"%B %d %Y %X%Z", "%B %d %Y %X %Z",
	"%B %d %Y %I:%M %p%Z", "%B %d %Y %I:%M %p %Z",
	"%d %B, %Y %X%Z", "%d %B, %Y %X %Z",
	"%d %B, %Y %I:%M %p%Z", "%d %B, %Y %I:%M %p %Z",
	"%d %B %Y %X%Z", "%d %B %Y %X %Z",
	"%d %B %Y %I:%M %p%Z", "%d %B %Y %I:%M %p %Z"
	))
def parse_datetimetz(data):
	if type(data) == type(datetime.datetime.now()):
		if data.tzinfo is None or data.tzinfo.utcoffset(data) is None:
			return None
		return data
	if not isinstance(data, str):
		return None
	dt = None
	# Check for numeric timezone
	try:
		datestr, sign, hr, min = dtzrx.match(data).groups()
		dt = parse_datetime(datestr)
		if not dt:
			return None
		sign = -1 if sign=='-' else 1
		return datetime.datetime(dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second, tzinfo=Tz(sign, int(hr), int(min)))
	except:
		# Check for alphabetic timezone
		for i,f in enumerate(timestamptz_fmts):
			try:
				dt = datetime.datetime.strptime(data, f)
			except:
				continue
			break
		if i:
			del timestamptz_fmts[i]
			timestamptz_fmts.appendleft(f)
		return dt

date_fmts = collections.deque(("%x",
		"%Y-%m-%d",
		"%Y/%m/%d",
		"%m/%d/%Y",
		"%d/%m/%Y",
		"%b %d, %Y",
		"%b %d %Y",
		"%d %b, %Y",
		"%d %b %Y",
		"%b. %d, %Y",
		"%b. %d %Y",
		"%d %b., %Y",
		"%d %b. %Y",
		"%B %d, %Y",
		"%B %d %Y",
		"%d %B, %Y",
		"%d %B %Y"
		))
def parse_date(data):
	if data is None:
		return None
	if isinstance(data, datetime.date):
		return data
	if not isinstance(data, str):
		return None
	for i,f in enumerate(date_fmts):
		try:
			dt = datetime.datetime.strptime(data, f)
			dtt = datetime.date(dt.year, dt.month, dt.day)
		except:
			continue
		break
	else:
		return None
	if i:
		del date_fmts[i]
		date_fmts.appendleft(f)
	return dtt

def parse_boolean(data):
	if data is None:
		return None
	true_strings = ('yes', 'true', '1')
	bool_strings = ('yes', 'no', 'true', 'false', '1', '0')
	if type(data) == bool:
		return data
	elif isinstance(data, int) and data in (0, 1):
		return data == 1
	elif isinstance(data, str) and data.lower() in bool_strings:
		return data.lower() in true_strings
	else:
		return None



#	End of support functions (1)
#===============================================================================================


#===============================================================================================
#-----  STATUS RECORDING

class StatObj(object):
	# A generic object to maintain status indicators.  These status
	# indicators are primarily those used in the metacommand
	# environment rather than for the program as a whole.
	def __init__(self):
		self.halt_on_err = True
		self.sql_error = False
		self.halt_on_metacommand_err = True
		self.metacommand_error = False
		self.cancel_halt = True
		self.batch = BatchLevels()

# End of status recording class.
#===============================================================================================


#===============================================================================================
#-----  ERROR HANDLING

class ErrInfo(Exception):
	def __repr__(self):
		return u"ErrInfo(%r, %r, %r, %r)" % (self.type, self.command, self.exception, self.other)
	def __init__(self, type, command_text=None, exception_msg=None, other_msg=None):
		# Argument 'type' should be "db", "cmd", "log", "error", or "exception".
		# Arguments for each type are as follows:
		# 	"db"		: command_text, exception_msg
		# 	"cmd"	: command_text, <exception_msg | other_msg>
		# 	"log"	: other_msg [, exception_msg]
		# 	"error"	: other_msg [, exception_msg]
		#	"systemexit" : other_msg
		# 	"exception"	: exception_msg [, other_msg]
		self.type = type
		self.command = command_text
		self.exception = None if not exception_msg else exception_msg.replace(u'\n', u'\n     ')
		self.other = None if not other_msg else other_msg.replace(u'\n', u'\n     ')
		if last_command is not None:
			self.script_line_no = current_script_line()
			self.cmd = last_command.command.statement
			self.cmdtype = last_command.command_type
		else:
			self.script_file = None
			self.script_line_no = None
			self.cmd = None
			self.cmdtype = None
		self.error_message = None
		subvars.add_substitution("$ERROR_MESSAGE", self.errmsg())
	def script_info(self):
		if self.script_line_no:
			return u"Line %d of script" % self.script_line_no
		else:
			return None
	def cmd_info(self):
		if self.cmdtype:
			if self.cmdtype == "cmd":
				em = u"Metacommand: %s" % self.cmd
			else:
				em = u"SQL statement: \n         %s" % self.cmd.replace(u'\n', u'\n         ')
			return em
		else:
			return None
	def eval_err(self):
		if self.type == 'db':
			self.error_message = u"**** Error in SQL statement."
		elif self.type == 'cmd':
			self.error_message = u"**** Error in metacommand."
		elif self.type == 'log':
			self.error_message = u"**** Error in logging."
		elif self.type == 'error':
			self.error_message = u"**** General error."
		elif self.type == 'systemexit':
			self.error_message = u"**** Exit."
		elif self.type == 'exception':
			self.error_message = u"**** Exception."
		else:
			self.error_message = u"**** Error of unknown type: %s" % self.type
		sinfo = self.script_info()
		cinfo = self.cmd_info()
		if sinfo:
			self.error_message += u"\n     %s" % sinfo
		if self.exception:
			self.error_message += u"\n     %s" % self.exception
		if self.other:
			self.error_message += u"\n     %s" % self.other
		if self.command:
			self.error_message += u"\n     %s" % self.command
		if cinfo:
			self.error_message += u"\n     %s" % cinfo
		self.error_message += u"\n     Error occurred at %s UTC." % time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())
		return self.error_message
	def write(self):
		errmsg = self.eval_err()
		output.write_err(errmsg)
		return errmsg
	def errmsg(self):
		return self.eval_err()


def exception_info():
	# Returns the exception type, value, source file name, source line number, and source line text.
	strace = traceback.extract_tb(sys.exc_info()[2])[-1:]
	traces = traceback.extract_tb(sys.exc_info()[2])
	xline = 0
	for trace in traces:
		if u"mapdata" in trace[0]:
			xline = trace[1]
	exc_message = u''
	exc_param = sys.exc_info()[1]
	if isinstance(exc_param, str):
		exc_message = exc_param
	else:
		if hasattr(exc_param, 'message') and isinstance(exc_param.message, str) and len(exc_param.message) > 0:
			exc_message = exc_param.message
		elif hasattr(exc_param, 'value') and isinstance(exc_param.value, str) and len(exc_param.value) > 0:
			exc_message = exc_param.value
		else:
			exc_message = type(u"")(exc_param)
	try:
		exc_message = type(u"")(exc_message)
	except:
		exc_message = repr(exc_message)
	xinfo = sys.exc_info()[0]
	xname = getattr(xinfo, "__name__", "")
	return xname, exc_message, strace[0][0], xline, strace[0][3]


def exception_desc():
	exc_type, exc_strval, exc_filename, exc_lineno, exc_linetext = exception_info()
	return u"%s: %s in %s on line %s of program." % (exc_type, exc_strval, exc_filename, exc_lineno)

#-----  End of ERROR HANDLING
#===============================================================================================



#===============================================================================================
#-----  DATABASE TYPES


class DbTypeError(Exception):
	def __init__(self, dbms_id, data_type, error_msg):
		self.dbms_id = dbms_id
		self.data_type = data_type
		self.error_msg = error_msg or "Unspecified error"
	def __repr__(self):
		return u"DbTypeError(%r, %r)" % (self.dbms_id, self.data_type, self.error_msg)
	def __str__(self):
		if self.data_type:
			return "%s DBMS type error with data type %s: %s" % (self.dbms_id, self.data_type.data_type_name, self.error_msg)
		else:
			return "%s DBMS type error: %s" % (self.dbms_id, self.error_msg)


class DbType(object):
	def __init__(self, DBMS_id, db_obj_quotes=u'""'):
		# The DBMS_id is the name by which this DBMS is identified.
		# db_obj_quotechars is a string of two characters that are the opening and closing quotes
		# for identifiers (schema, table, and column names) that need to be quoted.
		self.dbms_id = DBMS_id
		self.quotechars = db_obj_quotes
		# The dialect is a dictionary of DBMS-specific names for each column type.
		# Dialect keys are DataType classes.
		# Dialect objects are 4-tuples consisting of:
		#	0. a data type name (str)--non-null
		#	1. a Boolean indicating whether or not the length is part of the data type definition
		#		(e.g., for varchar)--non-null
		#	2. a name to use with the 'cast' operator as an alternative to the data type name--nullable.
		#	3. a function to perform a dbms-specific modification of the type conversion result produced
		#		by the 'from_data()' method of the data type.
		#	4. the precision for numeric data types.
		#	5. the scale for numeric data types.
		self.dialect = None
		# The dt_xlate dictionary translates one data type to another.
		# This is specifically needed for Access pre v. 4.0, which has no numeric type, and which
		# therefore requires the numeric data type to be treated as a float data type.
		self.dt_xlate = {}
	def __repr__(self):
		return u"DbType(%r, %r)" % (self.dbms_id, self.quotechars)
	def name_datatype(self, data_type, dbms_name, length_required=False, casting_name=None, conv_mod_fn=None, precision=None, scale=None):
		# data_type is a DataType class object.
		# dbms_name is the DBMS-specific name for this data type.
		# length_required indicates whether length information is required.
		# casting_name is an alternate to the data type name to use in SQL "cast(x as <casting_name>)" expressions.
		# conv_mod_fn is a function that modifies the result of data_type().from_data(x).
		if self.dialect is None:
			self.dialect = {}
		self.dialect[data_type] = (dbms_name, length_required, casting_name, conv_mod_fn, precision, scale)
	def datatype_name(self, data_type):
		# A convenience function to simplify access to data type namess.
		#if not isinstance(data_type, DataType):
		#	raise DbTypeError(self.dbms_id, None, "Unrecognized data type: %s" % data_type)
		try:
			return self.dialect[data_type][0]
		except:
			raise DbTypeError(self.dbms_id, data_type, "%s DBMS type has no specification for data type %s" % (self.dbms_id, data_type.data_type_name))
	def quoted(self, dbms_object):
		if re.search(r'\W', dbms_object):
			if self.quotechars[0] == self.quotechars[1] and self.quotechars[0] in dbms_object:
				dbms_object = dbms_object.replace(self.quotechars[0], self.quotechars[0]+self.quotechars[0])
			return self.quotechars[0] + dbms_object + self.quotechars[1]
		return dbms_object
	def spec_type(self, data_type):
		# Returns a translated data type or the original if there is no translation.
		if data_type in self.dt_xlate:
			return self.dt_xlate[data_type]
		return data_type
	def column_spec(self, column_name, data_type, max_len=None, is_nullable=False, precision=None, scale=None):
		# Returns a column specification as it would be used in a CREATE TABLE statement.
		# The arguments conform to those returned by Column().column_type
		#if not isinstance(data_type, DataType):
		#	raise DbTypeError(self.dbms_id, None, "Unrecognized data type: %s" % data_type)
		data_type = self.spec_type(data_type)
		try:
			dts = self.dialect[data_type]
		except:
			raise DbTypeError(self.dbms_id, data_type, "%s DBMS type has no specification for data type %s" % (self.dbms_id, data_type.data_type_name))
		if max_len and max_len > 0 and dts[1]:
			spec = "%s %s(%d)" % (self.quoted(column_name), dts[0], max_len)
		elif data_type.precspec and precision and scale:
			# numeric
			spec = "%s %s(%s,%s)" % (self.quoted(column_name), dts[0], precision, scale)
		else:
			spec = "%s %s" % (self.quoted(column_name), dts[0])
		if not is_nullable:
			spec += " NOT NULL"
		return spec

# Create a DbType object for each DBMS supported by execsql.

dbt_postgres = DbType("PostgreSQL")
dbt_sqlite = DbType("SQLite")
dbt_duckdb = DbType("DuckDB")
dbt_sqlserver = DbType("SQL Server")
dbt_mysql = DbType("MySQL")
dbt_firebird = DbType("Firebird")
dbt_oracle = DbType("Oracle")


#-----  End of DATABASE TYPES
#===============================================================================================



#===============================================================================================
#-----  DATABASE CONNECTIONS

class DatabaseNotImplementedError(Exception):
	def __init__(self, db_name, method):
		self.db_name = db_name
		self.method = method
	def __repr__(self):
		return u"DatabaseNotImplementedError(%r, %r)" % (self.db_name, self.method)
	def __str__(self):
		return "Method %s is not implemented for database %s" % (self.method, self.db_name)

class Database(object):
	def __init__(self, server_name, db_name, user_name=None, need_passwd=None, port=None, encoding=None):
		self.type = None
		self.server_name = server_name
		self.db_name = db_name
		self.user = user_name
		self.need_passwd = need_passwd
		self.password = None
		self.port = port
		self.encoding = encoding
		self.encode_commands = True
		self.paramstr = '?'
		self.conn = None
		self.autocommit = True
	def __repr__(self):
		return u"Database(%r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.encoding)
	def name(self):
		if self.server_name:
			return "%s(server %s; database %s)" % (self.type.dbms_id, self.server_name, self.db_name)
		else:
			return "%s(file %s)" % (self.type.dbms_id, self.db_name)
	def open_db(self):
		raise DatabaseNotImplementedError(self.name(), 'open_db')
	def cursor(self):
		if self.conn is None:
			self.open_db()
		return self.conn.cursor()
	def close(self):
		if self.conn:
			self.conn.close()
			self.conn = None
	def paramsubs(self, paramcount):
		return ",".join((self.paramstr,) * paramcount)
	def execute(self, sql, paramlist=None):
		# A shortcut to self.cursor().execute() that handles encoding.
		# Whether or not encoding is needed depends on the DBMS.
		global subvars
		if type(sql) in (tuple, list):
			sql = u" ".join(sql)
		try:
			curs = self.cursor()
			if self.encoding and self.encode_commands and sys.version_info < (3,):
				curs.execute(sql.encode(self.encoding))
			else:
				if paramlist is None:
					curs.execute(sql)
				else:
					curs.execute(sql, paramlist)
			try:
				# DuckDB does not support the 'rowcount' attribute
				subvars.add_substitution("$LAST_ROWCOUNT", curs.rowcount)
			except:
				pass
		except Exception as e:
			try:
				self.rollback()
			except:
				pass
			raise e
	def autocommit_on(self):
		self.autocommit = True
	def autocommit_off(self):
		self.autocommit = False
	def commit(self):
		if self.conn and self.autocommit:
			self.conn.commit()
	def rollback(self):
		if self.conn:
			try:
				self.conn.rollback()
			except:
				pass
	def schema_qualified_table_name(self, schema_name, table_name):
		table_name = self.type.quoted(table_name)
		if schema_name:
			schema_name = self.type.quoted(schema_name)
			return u'%s.%s' % (schema_name, table_name)
		return table_name
	def select_data(self, sql):
		# Returns the results of the sql select statement.
		curs = self.cursor()
		try:
			curs.execute(sql)
		except:
			self.rollback()
			raise
		try:
			subvars.add_substitution("$LAST_ROWCOUNT", curs.rowcount)
		except:
			pass
		rows = curs.fetchall()
		return [d[0] for d in curs.description], rows
	def select_rowsource(self, sql):
		# Return 1) a list of column names, and 2) an iterable that yields rows.
		curs = self.cursor()
		try:
			# DuckDB cursors have no 'arraysize' attribute
			curs.arraysize = conf.export_row_buffer
		except:
			pass
		try:
			curs.execute(sql)
		except:
			self.rollback()
			raise
		try:
			subvars.add_substitution("$LAST_ROWCOUNT", curs.rowcount)
		except:
			pass
		def decode_row():
			while True:
				rows = curs.fetchmany()
				if not rows:
					break
				else:
					for row in rows:
						if self.encoding:
							if sys.version_info < (3,):
								yield [c.decode(self.encoding, "backslashreplace") if type(c) == type("") else c for c in row]
							else:
								yield [c.decode(self.encoding, "backslashreplace") if type(c) == type(b'') else c for c in row]
						else:
							yield row
		return [d[0] for d in curs.description], decode_row()
	def schema_exists(self, schema_name):
		curs = self.cursor()
		curs.execute(u"SELECT schema_name FROM information_schema.schemata WHERE schema_name = '%s';" % schema_name)
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def table_exists(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select table_name from information_schema.tables where table_name = '%s'%s;" % (table_name, "" if not schema_name else " and table_schema='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
								other_msg=u"Failed test for existence of table %s in %s" % (table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def column_exists(self, table_name, column_name, schema_name=None):
		curs = self.cursor()
		sql = "select column_name from information_schema.columns where table_name='%s'%s and column_name='%s';" % (table_name, "" if not schema_name else " and table_schema='%s'" % schema_name, column_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed test for existence of column %s in table %s of %s" % (column_name, table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def table_columns(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select column_name from information_schema.columns where table_name='%s'%s order by ordinal_position;" % (table_name, "" if not schema_name else " and table_schema='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed to get column names for table %s of %s" % (table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return [row[0] for row in rows]
	def view_exists(self, view_name, schema_name=None):
		curs = self.cursor()
		sql = "select table_name from information_schema.views where table_name = '%s'%s;" % (view_name, "" if not schema_name else " and table_schema='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed test for existence of view %s in %s" % (view_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def role_exists(self, rolename):
		raise DatabaseNotImplementedError(self.name(), 'role_exists')
	def drop_table(self, tablename):
		# The 'tablename' argument should be schema-qualified and quoted as necessary.
		self.execute(u"drop table if exists %s cascade;" % tablename)
		self.commit()


class SqlServerDatabase(Database):
	def __init__(self, server_name, db_name, user_name, need_passwd=False, port=1433, encoding='latin1', password=None):
		global pyodbc
		try:
			import pyodbc
		except:
			fatal_error(u"The pyodbc module is required.  See http://github.com/mkleehammer/pyodbc", kwargs={})
		self.type = dbt_sqlserver
		self.server_name = server_name
		self.db_name = db_name
		self.user = user_name
		self.need_passwd = need_passwd
		self.password = password
		self.port = port if port else 1433
		self.encoding = encoding or 'latin1'    # Default on installation of SQL Server
		self.encode_commands = True
		self.paramstr = '?'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"SqlServerDatabase(%r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.encoding)
	def open_db(self):
		if self.conn is None:
			if self.user and self.need_passwd and not self.password:
				raise ErrInfo("error", other_msg="Password required but not provided")
			# Use pyodbc to connect.  Try different driver versions from newest to oldest.
			ssdrivers = ('ODBC Driver 17 for SQL Server', 'ODBC Driver 13.1 for SQL Server',
					'ODBC Driver 13 for SQL Server', 'ODBC Driver 11 for SQL Server',
					'SQL Server Native Client 11.0', 'SQL Server Native Client 10.0',
					'SQL Native Client', 'SQL Server')
			for drv in ssdrivers:
				if self.user:
					if self.password:
						connstr = "DRIVER={%s};SERVER=%s;MARS_Connection=Yes; DATABASE=%s;Uid=%s;Pwd=%s" % (drv, self.server_name, self.db_name, self.user, self.password)
					else:
						connstr = "DRIVER={%s};SERVER=%s;MARS_Connection=Yes; DATABASE=%s;Uid=%s" % (drv, self.server_name, self.db_name, self.user)
				else:
					connstr = "DRIVER={%s};SERVER=%s;MARS_Connection=Yes; DATABASE=%s;Trusted_Connection=yes" % (drv, self.server_name, self.db_name)
				try:
					self.conn = pyodbc.connect(connstr)
				except:
					pass
				else:
					break
			if not self.conn:
				raise ErrInfo(type="error", other_msg=u"Can't open SQL Server database %s on %s" % (self.db_name, self.server_name))
			curs = self.conn.cursor()
			curs.execute("SET IMPLICIT_TRANSACTIONS OFF;")
			curs.execute("SET ANSI_NULLS ON;")
			curs.execute("SET ANSI_PADDING ON;")
			curs.execute("SET ANSI_WARNINGS ON;")
			curs.execute("SET QUOTED_IDENTIFIER ON;")
			self.conn.commit()
	def schema_exists(self, schema_name):
		curs = self.cursor()
		curs.execute(u"select * from sys.schemas where name = '%s';" % schema_name)
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def role_exists(self, rolename):
		curs = self.cursor()
		curs.execute(u"select name from sys.database_principals where type in ('R', 'S') and name = '%s';" % rolename)
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def drop_table(self, tablename):
		# SQL Server and Firebird will throw an error if there are foreign keys to the table.
		tablename = self.type.quoted(tablename)
		self.execute(u"drop table %s;" % tablename)


class PostgresDatabase(Database):
	def __init__(self, server_name, db_name, user_name, need_passwd=False, port=5432, new_db=False, encoding='UTF8', password=None):
		global psycopg2
		try:
			import psycopg2
		except:
			fatal_error(u"The psycopg2 module is required to connect to PostgreSQL.", kwargs={})
		self.type = dbt_postgres
		self.server_name = server_name
		self.db_name = db_name
		self.user = user_name
		self.need_passwd = need_passwd
		self.password = password
		self.port = port if port else 5432
		self.new_db = new_db
		self.encoding = encoding or 'UTF8'
		self.encode_commands = False
		self.paramstr = '%s'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"PostgresDatabase(%r, %r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.new_db, self.encoding)
	def open_db(self):
		def db_conn(db, db_name):
			if db.user and db.password:
				return psycopg2.connect(host=str(db.server_name), database=str(db_name), port=db.port, user=db.user, password=db.password)
			else:
				return psycopg2.connect(host=str(db.server_name), database=db_name, port=db.port)
		def create_db(db):
			conn = db_conn(db, 'postgres')
			conn.autocommit = True
			curs = conn.cursor()
			curs.execute("create database %s encoding '%s';" % (db.db_name, db.encoding))
			conn.close()
		if self.conn is None:
			try:
				if self.user and self.need_passwd and not self.password:
					raise ErrInfo("error", "Password required but not provided")
				if self.new_db:
					create_db(self)
				self.conn = db_conn(self, self.db_name)
			except ErrInfo:
				raise
			except:
				msg = u"Failed to open PostgreSQL database %s on %s" % (self.db_name, self.server_name)
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=msg)
			# (Re)set the encoding to match the database.
			self.encoding = self.conn.encoding
	def role_exists(self, rolename):
		curs = self.cursor()
		curs.execute(u"select rolname from pg_roles where rolname = '%s';" % rolename)
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def table_exists(self, table_name, schema_name=None):
		curs = self.cursor()
		if schema_name is not None:
			sql = "select table_name from information_schema.tables where table_name = '%s'%s;" % (table_name, "" if not schema_name else " and table_schema='%s'" % schema_name)
		else:
			sql = """select table_name from information_schema.tables where table_name = '%s' and
			         table_schema in (select nspname from pg_namespace where oid = pg_my_temp_schema()
                     union
                     select trim(unnest(string_to_array(replace(setting, '"$user"', CURRENT_USER), ',')))
                     from pg_settings where name = 'search_path');""" % table_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
								other_msg=u"Failed test for existence of table %s in %s" % (table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def view_exists(self, view_name, schema_name=None):
		curs = self.cursor()
		if schema_name is not None:
			sql = "select table_name from information_schema.views where table_name = '%s'%s;" % (view_name, "" if not schema_name else " and table_schema='%s'" % schema_name)
		else:
			sql = """select table_name from information_schema.views where table_name = '%s' and
			         table_schema in (select nspname from pg_namespace where oid = pg_my_temp_schema()
                     union
                     select trim(unnest(string_to_array(replace(setting, '"$user"', CURRENT_USER), ',')))
                     from pg_settings where name = 'search_path');""" % view_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed test for existence of view %s in %s" % (view_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def vacuum(self, argstring):
		self.commit()
		self.conn.set_session(autocommit=True)
		self.conn.cursor().execute("VACUUM %s;" % argstring)
		self.conn.set_session(autocommit=False)


class OracleDatabase(Database):
	def __init__(self, server_name, db_name, user_name, need_passwd=False, port=5432, encoding='UTF8', password=None):
		global cx_Oracle
		try:
			import cx_Oracle
		except:
			fatal_error(u"The cx-Oracle module is required to connect to Oracle.   See https://pypi.org/project/cx-Oracle/", kwargs={})
		self.type = dbt_oracle
		self.server_name = server_name
		self.db_name = db_name
		self.user = user_name
		self.need_passwd = need_passwd
		self.password = password
		self.port = port if port else 1521
		self.encoding = encoding or 'UTF8'
		self.encode_commands = False
		self.paramstr = ':1'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"OracleDatabase(%r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.encoding)
	def open_db(self):
		def db_conn(db, db_name):
			dsn = cx_Oracle.makedsn(db.server_name, db.port, service_name=db_name)
			if db.user and db.password:
				return cx_Oracle.connect(user=db.user, password=db.password, dsn=dsn)
			else:
				return cx_Oracle.connect(dsn=dsn)
		if self.conn is None:
			try:
				if self.user and self.need_passwd and not self.password:
					raise ErrInfo("error", other_msg="Password required but not provided")
				self.conn = db_conn(self, self.db_name)
			except ErrInfo:
				raise
			except:
				msg = u"Failed to open Oracle database %s on %s" % (self.db_name, self.server_name)
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=msg)
	def execute(self, sql, paramlist=None):
		# Strip any semicolon off the end and pass to the parent method.
		if sql[-1:] == ";":
			super(OracleDatabase, self).execute(sql[:-1], paramlist)
		else:
			super(OracleDatabase, self).execute(sql, paramlist)
	def select_data(self, sql):
		if sql[-1:] == ";":
			return super(OracleDatabase, self).select_data(sql[:-1])
		else:
			return super(OracleDatabase, self).select_data(sql)
	def select_rowsource(self, sql):
		if sql[-1:] == ";":
			return super(OracleDatabase, self).select_rowsource(sql[:-1])
		else:
			return super(OracleDatabase, self).select_rowsource(sql)
	def schema_exists(self, schema_name):
		raise DatabaseNotImplementedError(self.name(), 'schema_exists')
	def table_exists(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select table_name from sys.all_tables where table_name = '%s'%s" % (table_name, "" if not schema_name else " and owner ='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
								other_msg=u"Failed test for existence of table %s in %s" % (table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def column_exists(self, table_name, column_name, schema_name=None):
		curs = self.cursor()
		sql = "select column_name from all_tab_columns where table_name='%s'%s and column_name='%s'" % (table_name, "" if not schema_name else " and owner ='%s'" % schema_name, column_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed test for existence of column %s in table %s of %s" % (column_name, table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def table_columns(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select column_name from all_tab_columns where table_name='%s'%s order by column_id" % (table_name, "" if not schema_name else " and owner='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed to get column names for table %s of %s" % (table_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return [row[0] for row in rows]
	def view_exists(self, view_name, schema_name=None):
		curs = self.cursor()
		sql = "select view_name from sys.all_views where view_name = '%s'%s" % (view_name, "" if not schema_name else " and owner ='%s'" % schema_name)
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed test for existence of view %s in %s" % (view_name, self.name()))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def role_exists(self, rolename):
		curs = self.cursor()
		curs.execute(u"select role from dba_roles where role = '%s' union " \
				" select username from all_users where username = '%s';" % (rolename, rolename))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def drop_table(self, tablename):
		tablename = self.type.quoted(tablename)
		self.execute(u"drop table %s cascade constraints" % tablename)
	def paramsubs(self, paramcount):
		return ",".join(":"+str(d) for d in range(1, paramcount+1))


class SQLiteDatabase(Database):
	def __init__(self, SQLite_fn):
		global sqlite3
		self.type = dbt_sqlite
		self.server_name = None
		self.db_name = SQLite_fn
		self.user = None
		self.need_passwd = False
		self.encoding = 'UTF-8'
		self.encode_commands = False
		self.paramstr = '?'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"SQLiteDabase(%r)" % self.db_name
	def open_db(self):
		if self.conn is None:
			try:
				self.conn = sqlite3.connect(self.db_name)
			except ErrInfo:
				raise
			except:
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=u"Can't open SQLite database %s" % self.db_name)
		pragma_cols, pragma_data = self.select_data("pragma encoding;")
		self.encoding = pragma_data[0][0]
	def table_exists(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select name from sqlite_master where type='table' and name='%s';" % table_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(), other_msg=u'Failed test for existence of SQLite table "%s";' % table_name)
		rows = curs.fetchall()
		return len(rows) > 0
	def column_exists(self, table_name, column_name, schema_name=None):
		curs = self.cursor()
		sql = "select %s from %s limit 1;" % (column_name, table_name)
		try:
			curs.execute(sql)
		except:
			return False
		return True
	def table_columns(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select * from %s where 1=0;" % table_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed to get column names for table %s of %s" % (table_name, self.name()))
		return [d[0] for d in curs.description]
	def view_exists(self, view_name):
		curs = self.cursor()
		sql = "select name from sqlite_master where type='view' and name='%s';" % view_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(), other_msg=u'Failed test for existence of SQLite view "%s";' % view_name)
		rows = curs.fetchall()
		return len(rows) > 0
	def schema_exists(self, schema_name):
		return False
	def drop_table(self, tablename):
		tablename = self.type.quoted(tablename)
		self.execute(u"drop table if exists %s;" % tablename)


class DuckDBDatabase(Database):
	def __init__(self, DuckDB_fn):
		global duckdb
		try:
			import duckdb
		except:
			fatal_error(u"The duckdb module is required.", kwargs={})
		self.type = dbt_duckdb
		self.server_name = None
		self.db_name = DuckDB_fn
		self.catalog_name = os.path.splitext(DuckDB_fn)[0]
		self.user = None
		self.need_passwd = False
		self.encoding = 'UTF-8'
		self.encode_commands = False
		self.paramstr = '?'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"DuckDBDabase(%r)" % self.db_name
	def open_db(self):
		if self.conn is None:
			try:
				self.conn = duckdb.connect(self.db_name, read_only=False)
			except ErrInfo:
				raise
			except:
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=u"Can't open DuckDB database %s" % self.db_name)
	def view_exists(self, view_name):
		# DuckDB information_schema has no 'views' table; views are listed in 'tables'
		return self.table_exists(view_name)
	def schema_exists(self, schema_name):
		# In DuckDB, the 'schemata' view is not limited to the current database.
		curs = self.cursor()
		curs.execute(u"SELECT schema_name FROM information_schema.schemata WHERE schema_name = '%s' and catalog_name = '%s';" % (schema_name, self.catalog_name))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def execute(self, sql, paramlist=None):
		if type(sql) in (tuple, list):
			sql = u" ".join(sql)
		try:
			curs = self.cursor()
			if paramlist is None:
				curs.execute(sql)
			else:
				curs.execute(sql, paramlist)
			# DuckDB does not support the 'rowcount' attribute, so $LAST_ROWCOUNT is not set
		except Exception as e:
			try:
				self.rollback()
			except:
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=u"Can't open DuckDB database %s" % self.db_name)


class MySQLDatabase(Database):
	def __init__(self, server_name, db_name, user_name, need_passwd=False, port=3306, encoding='latin1', password=None):
		global mysql_lib
		try:
			import pymysql as mysql_lib
		except:
			fatal_error(u"The pymysql module is required to connect to MySQL.   See https://pypi.python.org/pypi/PyMySQL", kwargs={})
		self.type = dbt_mysql
		self.server_name = str(server_name)
		self.db_name = str(db_name)
		self.user = str(user_name)
		self.need_passwd = need_passwd
		self.password = password
		self.port = 3306 if not port else port
		self.encoding = encoding or 'latin1'
		self.encode_commands = True
		self.paramstr = '%s'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"MySQLDatabase(%r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.encoding)
	def open_db(self):
		def db_conn():
			if self.user and self.password:
				return mysql_lib.connect(host=self.server_name, database=self.db_name, port=self.port, user=self.user, password=self.password, charset=self.encoding, local_infile=True)
			else:
				return mysql_lib.connect(host=self.server_name, database=self.db_name, port=self.port, charset=self.encoding, local_infile=True)
		if self.conn is None:
			try:
				if self.user and self.need_passwd and not self.password:
					raise ErrInfo("error", other_msg="Password required but not provided")
				self.conn = db_conn()
				self.execute("set session sql_mode='ANSI';")
			except ErrInfo:
				raise
			except:
				msg = u"Failed to open MySQL database %s on %s" % (self.db_name, self.server_name)
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=msg)
	def schema_exists(self, schema_name):
		return False
	def role_exists(self, rolename):
		curs = self.cursor()
		curs.execute(u"select distinct user as role from mysql.user where user = '%s'" \
				" union select distinct role_name as role from information_schema.applicable_roles" \
				" where role_name = '%s'" % (rolename, rolename))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0


class FirebirdDatabase(Database):
	def __init__(self, server_name, db_name, user_name, need_passwd=False, port=3050, encoding='latin1', password=None):
		global firebird_lib
		try:
			import fdb as firebird_lib
		except:
			fatal_error(u"The fdb module is required to connect to MySQL.   See https://pypi.python.org/pypi/fdb/", kwargs={})
		self.type = dbt_firebird
		self.server_name = str(server_name)
		self.db_name = str(db_name)
		self.user = str(user_name)
		self.need_passwd = need_passwd
		self.password = password
		self.port = 3050 if not port else port
		self.encoding = encoding or 'latin1'
		self.encode_commands = True
		self.paramstr = '?'
		self.conn = None
		self.autocommit = True
		self.open_db()
	def __repr__(self):
		return u"FirebirdDatabase(%r, %r, %r, %r, %r, %r)" % (self.server_name, self.db_name, self.user,
				self.need_passwd, self.port, self.encoding)
	def open_db(self):
		def db_conn():
			if self.user and self.password:
				return firebird_lib.connect(host=self.server_name, database=self.db_name, port=self.port, user=self.user, password=self.password, charset=self.encoding)
			else:
				return firebird_lib.connect(host=self.server_name, database=self.db_name, port=self.port, charset=self.encoding)
		if self.conn is None:
			try:
				if self.user and self.need_passwd and not self.password:
					raise ErrInfo("error", other_msg="Password required but not provided")
				self.conn = db_conn()
				#self.execute('set autoddl off;')
			except ErrInfo:
				raise
			except:
				msg = u"Failed to open Firebird database %s on %s" % (self.db_name, self.server_name)
				raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=msg)
	def table_exists(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "SELECT RDB$RELATION_NAME FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG=0 AND RDB$VIEW_BLR IS NULL AND RDB$RELATION_NAME='%s';" % table_name.upper()
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			e = ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(), other_msg=u"Failed test for existence of Firebird table %s" % table_name)
			try:
				self.rollback()
			except:
				pass
			raise e
		rows = curs.fetchall()
		self.conn.commit()
		curs.close()
		return len(rows) > 0
	def column_exists(self, table_name, column_name, schema_name=None):
		curs = self.cursor()
		sql = "select first 1 %s from %s;" % (column_name, table_name)
		try:
			curs.execute(sql)
		except:
			return False
		return True
	def table_columns(self, table_name, schema_name=None):
		curs = self.cursor()
		sql = "select first 1 * from %s;" % table_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(),
							other_msg=u"Failed to get column names for table %s of %s" % (table_name, self.name()))
		return [d[0] for d in curs.description]
	def view_exists(self, view_name, schema_name=None):
		curs = self.cursor()
		sql = "select distinct rdb$view_name from rdb$view_relations where rdb$view_name = '%s';" % view_name
		try:
			curs.execute(sql)
		except ErrInfo:
			raise
		except:
			self.rollback()
			raise ErrInfo(type="db", command_text=sql, exception_msg=exception_desc(), other_msg=u"Failed test for existence of Firebird view %s" % view_name)
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def schema_exists(self, schema_name):
		return False
	def role_exists(self, rolename):
		curs = self.cursor()
		curs.execute(u"SELECT DISTINCT USER FROM RDB$USER_PRIVILEGES WHERE USER = '%s' union " \
				" SELECT DISTINCT RDB$ROLE_NAME FROM RDB$ROLES WHERE RDB$ROLE_NAME = '%s';" % (rolename, rolename))
		rows = curs.fetchall()
		curs.close()
		return len(rows) > 0
	def drop_table(self, tablename):
		# Firebird will thrown an error if there are foreign keys into the table.
		tablename = self.type.quoted(tablename)
		self.execute(u"DROP TABLE %s;" % tablename)
		self.conn.commit()


class DatabasePool(object):
	# Define an object that maintains a set of database connection objects, each with
	# a name (alias), and with the current and initial databases identified.
	def __init__(self):
		self.pool = {}
		self.initial_db = None
		self.current_db = None
		self.do_rollback = True
	def __repr__(self):
		return u"DatabasePool()"
	def add(self, db_alias, db_obj):
		db_alias = db_alias.lower()
		if db_alias == 'initial' and len(self.pool) > 0:
			raise ErrInfo(type="error", other_msg="You may not use the name 'INITIAL' as a database alias.")
		if len(self.pool) == 0:
			self.initial_db = db_alias
			self.current_db = db_alias
		if db_alias in self.pool:
			# Don't allow reassignment of a database that is used in any batch.
			if status.batch.uses_db(self.pool[db_alias]):
				raise ErrInfo(type="error", other_msg="You may not reassign the alias of a database that is currently used in a batch.")
			self.pool[db_alias].close()
		self.pool[db_alias] = db_obj
	def aliases(self):
		# Return a list of the currently defined aliases
		return list(self.pool)
	def current(self):
		# Return the current db object.
		return self.pool[self.current_db]
	def current_alias(self):
		# Return the alias of the current db object.
		return self.current_db
	def initial(self):
		return self.pool[self.initial_db]
	def aliased_as(self, db_alias):
		return self.pool[db_alias]
	def make_current(self, db_alias):
		# Change the current database in use.
		db_alias = db_alias.lower()
		if not db_alias in self.pool:
			raise ErrInfo(type="error", other_msg=u"Database alias '%s' is unrecognized; cannnot use it." % db_alias)
		self.current_db = db_alias
	def disconnect(self, alias):
		if alias == self.current_db or (alias == 'initial' and 'initial' in self.pool):
			raise ErrInfo(type="error", other_msg=u"Database alias %s can't be removed or redefined while it is in use." % alias)
		if alias in self.pool:
			self.pool[alias].close()
			del self.pool[alias]
	def closeall(self):
		for alias, db in self.pool.items():
			nm = db.name()
			try:
				if self.do_rollback:
					db.rollback()
				db.close()
			except:
				pass
		self.__init__()

# End of database connections
#===============================================================================================



#===============================================================================================
#-----  SCRIPTING

class BatchLevels(object):
	# A stack to keep a record of the databases used in nested batches.
	class Batch(object):
		def __init__(self):
			self.dbs_used = []
	def __init__(self):
		self.batchlevels = []
	def in_batch(self):
		return len(self.batchlevels) > 0
	def new_batch(self):
		self.batchlevels.append(self.Batch())
	def using_db(self, db):
		if len(self.batchlevels) > 0 and not db in self.batchlevels[-1].dbs_used:
			self.batchlevels[-1].dbs_used.append(db)
	def uses_db(self, db):
		if len(self.batchlevels) == 0:
			return False
		for batch in self.batchlevels:
			if db in batch.dbs_used:
				return True
	def rollback_batch(self):
		if len(self.batchlevels) > 0:
			b = self.batchlevels[-1]
			for db in b.dbs_used:
				db.rollback()
	def end_batch(self):
		b = self.batchlevels.pop()
		for db in b.dbs_used:
			db.commit()

class IfItem(object):
	# An object representing an 'if' level, with context data.
	def __init__(self, tf_value):
		self.tf_value = tf_value
		self.scriptline = current_script_line()
	def value(self):
		return self.tf_value
	def invert(self):
		self.tf_value = not self.tf_value
	def change_to(self, tf_value):
		self.tf_value = tf_value
	def script_line(self):
		return self.scriptline

class IfLevels(object):
	# A stack of True/False values corresponding to a nested set of conditionals,
	# with methods to manipulate and query the set of conditional states.
	# This stack is used by the IF metacommand and related metacommands.
	def __init__(self):
		self.if_levels = []
	def nest(self, tf_value):
		self.if_levels.append(IfItem(tf_value))
	def unnest(self):
		if len(self.if_levels) == 0:
			raise ErrInfo(type="error", other_msg="Can't exit an IF block; no IF block is active.")
		else:
			self.if_levels.pop()
	def invert(self):
		if len(self.if_levels) == 0:
			raise ErrInfo(type="error", other_msg="Can't change the IF state; no IF block is active.")
		else:
			self.if_levels[-1].invert()
	def replace(self, tf_value):
		if len(self.if_levels) == 0:
			raise ErrInfo(type="error", other_msg="Can't change the IF state; no IF block is active.")
		else:
			self.if_levels[-1].change_to(tf_value)
	def current(self):
		if len(self.if_levels) == 0:
			raise ErrInfo(type="error", other_msg="No IF block is active.")
		else:
			return self.if_levels[-1].value()
	def all_true(self):
		if self.if_levels == []:
			return True
		return all([tf.value() for tf in self.if_levels])
	def only_current_false(self):
		# Returns True if the current if level is false and all higher levels are True.
		# Metacommands such as ELSE and ENDIF are executed in this state.
		if len(self.if_levels) == 0:
			return False
		elif len(self.if_levels) == 1:
			return not self.if_levels[-1].value()
		else:
			return not self.if_levels[-1].value() and all([tf.value() for tf in self.if_levels[:-1]])
	def script_lines(self, top_n):
		# Returns a list of tuples containing the script name and line number
		# for the topmost 'top_n' if levels, in bottom-up order.
		if len(self.if_levels) < top_n:
			raise ErrInfo(type="error", other_msg="Invalid IF stack depth reference.")
		levels = self.if_levels[len(self.if_levels) - top_n:]
		return [lvl.script_line() for lvl in levels]

class CounterVars(object):
	# A dictionary of dynamically created named counter variables.  Counter
	# variables are created when first referenced, and automatically increment
	# the integer value returned with each reference.
	_COUNTER_RX = re.compile(r'!!\$(COUNTER_\d+)!!', re.I)
	def __init__(self):
		self.counters = {}
	def _ctrid(self, ctr_no):
		return u'counter_%d' % ctr_no
	def set_counter(self, ctr_no, ctr_val):
		self.counters[self._ctrid(ctr_no)] = ctr_val
	def remove_counter(self, ctr_no):
		ctr_id = self._ctrid(ctr_no)
		if ctr_id in self.counters:
			del self.counters[ctr_id]
	def remove_all_counters(self):
		self.counters = {}
	def substitute(self, command_str):
		# Substitutes any counter variable references with the counter
		# value and returns the modified command string and a flag
		# indicating whether any replacements were made.
		match_found = False
		m = self._COUNTER_RX.search(command_str, re.I)
		if m:
			ctr_id = m.group(1).lower()
			if not ctr_id in self.counters:
				self.counters[ctr_id] = 0
			new_count = self.counters[ctr_id] + 1
			self.counters[ctr_id] = new_count
			return command_str.replace(u'!!$'+m.group(1)+u'!!', str(new_count)), True
		return command_str, False
	def substitute_all(self, any_text):
		subbed = True
		any_subbed = False
		while subbed:
			any_text, subbed = self.substitute(any_text)
			if subbed:
				any_subbed = True
		return any_text, any_subbed

class SubVarSet(object):
	# A pool of substitution variables.  Each variable consists of a name and
	# a (string) value.  All variable names are stored as lowercase text.
	# This is implemented as a list of tuples rather than a dictionary to enforce
	# ordered substitution.
	def __init__(self):
		self.substitutions = []
		#List of acceptable single-character variable name prefixes
		self.prefix_list = ['$','&','@']
		# Regex for matching
		# Don't construct/compile on init because deepcopy() can't handle compiled regexes.
		# 'Regular' variables, dereferenced with "!!"
		self.var_rx = None
	def compile_var_rx(self):
		# Compile regex to validate variable name, using the prefix list
		# This is: any character from the prefix (optionally), followed by one or more word chars
		self.var_rx_str = r'^[' +  "".join(self.prefix_list) + r']?\w+$'
		self.var_rx = re.compile(self.var_rx_str, re.I)
	def var_name_ok(self, varname):
		if self.var_rx is None:
			self.compile_var_rx()
		return self.var_rx.match(varname) is not None
	def check_var_name(self, varname):
		if not self.var_name_ok(varname.lower()):
			raise ErrInfo("error", other_msg="Invalid variable name (%s) in this context." % varname)
	def remove_substitution(self, template_str):
		self.check_var_name(template_str)
		old_sub = template_str.lower()
		self.substitutions = [sub for sub in self.substitutions if sub[0] != old_sub]
	def add_substitution(self, varname, repl_str):
		self.check_var_name(varname)
		varname = varname.lower()
		self.remove_substitution(varname)
		self.substitutions.append((varname, repl_str))
	def append_substitution(self, varname, repl_str):
		self.check_var_name(varname)
		varname = varname.lower()
		oldsub = [x for x in self.substitutions if x[0] == varname]
		if len(oldsub) == 0:
			self.add_substitution(varname, repl_str)
		else:
			self.add_substitution(varname, "%s\n%s" % (oldsub[0][1], repl_str))
	def varvalue(self, varname):
		self.check_var_name(varname)
		vname = varname.lower()
		for vardef in self.substitutions:
			if vardef[0] == vname:
				return vardef[1]
		return None
	def increment_by(self, varname, numeric_increment):
		self.check_var_name(varname)
		varvalue = self.varvalue(varname)
		if varvalue is None:
			varvalue = "0"
			self.add_substitution(varname, varvalue)
		numvalue = as_numeric(varvalue)
		numinc = as_numeric(numeric_increment)
		if numvalue is None or numinc is None:
			newval = "%s+%s" % (varvalue, numeric_increment)
		else:
			newval = str(numvalue + numinc)
		self.add_substitution(varname, newval)
	def sub_exists(self, template_str):
		self.check_var_name(template_str)
		test_str = template_str.lower()
		return test_str in [s[0] for s in self.substitutions]
	def merge(self, other_subvars):
		# Return a new SubVarSet object with this object's variables
		# merged with the 'other_subvars' substitutions; the latter
		# takes precedence.
		# Also merges the prefix lists
		if other_subvars is not None:
			newsubs = SubVarSet()
			newsubs.substitutions = self.substitutions
			newsubs.prefix_list = list(set(self.prefix_list + other_subvars.prefix_list))
			newsubs.compile_var_rx()
			for vardef in other_subvars.substitutions:
				newsubs.add_substitution(vardef[0], vardef[1])
			return newsubs
		return self
	def substitute(self, command_str):
		# Replace any substitution variables in the command string.
		# This does only one round of replacements: if the first round of replacements
		# produces more substitution variables that could be replaced, those derived
		# matching strings are not replaced.  The second value returned by this
		# function indicates whether any substitutions were made, so that this
		# method can be called repeatedly.
		match_found = False
		if isinstance(command_str, str):
			for match, sub in self.substitutions:
				if sub is None:
					sub = ''
				sub = str(sub)
				if match[0] == "$":
					match = "\\"+match
				if os.name != 'posix':
					sub = sub.replace("\\", "\\\\")
				pat = "!!%s!!" % match
				patq = "!'!%s!'!" % match
				patdq = '!"!%s!"!' % match
				if re.search(pat, command_str, re.I):
					return re.sub(pat, sub, command_str, flags=re.I), True
				if re.search(patq, command_str, re.I):
					sub = sub.replace("'", "''")
					return re.sub(patq, sub, command_str, flags=re.I), True
				if re.search(patdq, command_str, re.I):
					sub = '"' + sub + '"'
					return re.sub(patdq, sub, command_str, flags=re.I), True
		return command_str, False
	def substitute_all(self, any_text):
		subbed = True
		any_subbed = False
		while subbed:
			any_text, subbed = self.substitute(any_text)
			if subbed:
				any_subbed = True
		return any_text, any_subbed


class LocalSubVarSet(SubVarSet):
	# A pool of local substitution variables.
	# Inherits everything from the base class except the allowed prefix list.
	# For local variables, only '~' is allowed as a prefix and MUST be present
	def __init__(self):
		SubVarSet.__init__(self)
		self.prefix_list = ['~']
	def compile_var_rx(self):
		# This is different from the base class because the prefix is required, not optional
		self.var_rx_str = r'^[' +  "".join(self.prefix_list) + r']\w+$'
		self.var_rx = re.compile(self.var_rx_str, re.I)


class ScriptArgSubVarSet(SubVarSet):
	# A pool of script argument names.
	# Inherits everything from the base class except the allowed prefix list.
	# For script arguments, only '#' is allowed as a prefix and MUST be present
	def __init__(self):
		SubVarSet.__init__(self)
		self.prefix_list = ['#']
	def compile_var_rx(self):
		# This is different from the base class because the prefix is required, not optional
		self.var_rx_str = r'^[' +  "".join(self.prefix_list) + r']\w+$'
		self.var_rx = re.compile(self.var_rx_str, re.I)


class MetaCommand(object):
	# A compiled metacommand that can be run if it matches a metacommand command string in the input.
	def __init__(self, rx, exec_func, description=None, run_in_batch=False, run_when_false=False, set_error_flag=True):
		# rx: a compiled regular expression
		# exec_func: a function object that carries out the work of the metacommand.
		#			This function must take keyword arguments corresponding to those named
		#			in the regex, and must return a value (which is used only for conditional
		#			metacommands) or None.
		# run_in_batch: determines whether a metacommand should be run inside a batch.  Only 'END BATCH'
		#			should be run inside a batch.
		# run_when_false: determines whether a metacommand should be run when the exec state is False.
		#			only 'ELSE', 'ELSEIF', 'ORIF', and 'ENDIF' should be run when False, and only when
		#			all higher levels are True.  This condition is evaluated by the script processor.
		# set_error_flag: When run, sets or clears status.metacommand_error.
		self.next_node = None
		self.rx = rx
		self.exec_fn = exec_func
		self.description = description
		self.run_in_batch = run_in_batch
		self.run_when_false = run_when_false
		self.set_error_flag = set_error_flag
		self.hitcount = 0
	def __repr__(self):
		return u"MetaCommand(%r, %r, %r, %r, %r)" % (self.rx.pattern, self.exec_fn, self.description,
				self.run_in_batch, self.run_when_false)
	def run(self, cmd_str):
		# Runs the metacommand if the command string matches the regex.
		# Returns a 2-tuple consisting of:
		#	0. True or False indicating whether the metacommand applies.  If False, the
		#		remaining return value is None and has no meaning.
		#	1. The return value of the metacommand function.
		#		Exceptions are caught and converted to ErrInfo exceptions.
		m = self.rx.match(cmd_str.strip())
		if m:
			cmdargs = m.groupdict()
			cmdargs['metacommandline'] = cmd_str
			er = None
			try:
				rv = self.exec_fn(**cmdargs)
			except ErrInfo as errinf:
				# This variable reassignment is required by Python 3;
				# if the line is "except ErrInfo as er:" then an
				# UnboundLocalError occurs at the "if er:" statement.
				er = errinf
			except:
				er = ErrInfo("cmd", command_text=cmd_str, exception_msg=exception_desc())
			if er:
				if status.halt_on_metacommand_err:
					raise er
				if self.set_error_flag:
					status.metacommand_error = True
					return True, None
			else:
				if self.set_error_flag:
					status.metacommand_error = False
				self.hitcount += 1
				return True, rv
		return False, None

class MetaCommandList(object):
	# The head node for a linked list of MetaCommand objects.
	def __init__(self):
		self.next_node = None
	def __iter__(self):
		n1 = self.next_node
		while n1 is not None:
			yield n1
			n1 = n1.next_node
	def insert_node(self, new_node):
		new_node.next_node = self.next_node
		self.next_node = new_node
	def add(self, matching_regexes, exec_func, description=None, run_in_batch=False, run_when_false=False, set_error_flag=True):
		# Creates a new Metacomman and adds it at the head of the linked list.
		if type(matching_regexes) in (tuple, list):
			self.regexes = [re.compile(rx, re.I) for rx in tuple(matching_regexes)]
		else:
			self.regexes = [re.compile(matching_regexes, re.I)]
		for rx in self.regexes:
			self.insert_node(MetaCommand(rx, exec_func, description, run_in_batch, run_when_false, set_error_flag))
	def eval(self, cmd_str):
		# Evaluates the given metacommand string (line from the SQL script).
		# Searches the linked list of MetaCommand objects.  If a match is found, the metacommand
		# is run, and that MetaCommand is moved to the head of the list.
		# Returns a 2-tuple consisting of:
		#	0. True or False indicating whether the metacommand applies.  If False, the
		#		remaining return value is None and has no meaning.
		#	1. The return value of the metacommand function.
		#		Exceptions are caught and converted to ErrInfo exceptions.
		n1 = self
		node_no = 0
		while n1 is not None:
			n2 = n1.next_node
			if n2 is not None:
				node_no += 1
				if if_stack.all_true() or n2.run_when_false:
					success, value = n2.run(cmd_str)
					if success:
						# Move n2 to the head of the list.
						n1.next_node = n2.next_node
						n2.next_node = self.next_node
						self.next_node = n2
						return True, value
			n1 = n2
		return False, None
	def get_match(self, cmd):
		# Tries to match the command 'cmd' to any MetaCommand.  If a match
		# is found, returns a tuple containing the MetaCommand object and
		# the match object; if not, returns None.
		n1 = self.next_node
		while n1 is not None:
			m = n1.rx.match(cmd.strip())
			if m is not None:
				return (n1, m)
			n1 = n1.next_node
		return None


# Global linked lists of MetaCommand objects (commands and conditional tests).
# These are filled in the 'MetaCommand Functions' and 'Conditional Tests for Metacommands' sections.
metacommandlist = MetaCommandList()
conditionallist = MetaCommandList()


class SqlStmt(object):
	# A SQL statement to be passed to a database to execute.
	# The purpose of storing a SQL statement as a SqlStmt object rather
	# than as a simple string is to allow the definition of a 'run()'
	# method that is different from the 'run()' method of a MetacommandStmt.
	# In effect, the SqlStmt and MetacommandStmt classes are both
	# subclasses of a Stmt class, but the Stmt class, and subclassing,
	# are not implemented because the Stmt class would be trivial: just
	# an assignment in the init method.
	def __init__(self, sql_statement):
		self.statement = re.sub(r'\s*;(\s*;\s*)+$', ';', sql_statement)
	def __repr__(self):
		return u"SqlStmt(%s)" % self.statement
	def run(self, localvars=None, commit=True):
		# Run the SQL statement on the current database.  The current database
		# is obtained from the global database pool "dbs".
		# 'localvars' must be a SubVarSet object.
		if if_stack.all_true():
			e = None
			status.sql_error = False
			cmd = substitute_vars(self.statement, localvars)
			if varlike.search(cmd):
				lno = current_script_line()
				script_errors.append(["There is a potential un-substituted variable in the command %s" % cmd, lno])
			try:
				db = dbs.current()
				db.execute(cmd)
				if commit:
					db.commit()
			except ErrInfo as errinfo:
				# This variable reassignment is required by Python 3;
				# if the line is "except ErrInfo as e:" then an
				# UnboundLocalError occurs at the "if e:" statement.
				e = errinfo
			except:
				e = ErrInfo(type="exception", exception_msg=exception_desc())
			if e:
				subvars.add_substitution("$LAST_ERROR", cmd)
				status.sql_error = True
				if status.halt_on_err:
					raise e
				return
			subvars.add_substitution("$LAST_SQL", cmd)
	def commandline(self):
		return self.statement


class MetacommandStmt(object):
	# A metacommand to be handled by execsql.
	def __init__(self, metacommand_statement):
		self.statement = metacommand_statement
	def __repr__(self):
		return u"MetacommandStmt(%s)" % self.statement
	def run(self, localvars=None, commit=False):
		# Tries all metacommands in "metacommandlist" until one runs.
		# Returns the result of the metacommand that was run, or None.
		# Arguments:
		#	localvars: a SubVarSet object.
		#	commit   : not used; included to allow an isomorphic interface with SqlStmt.run().
		errmsg = "Unknown metacommand"
		cmd = substitute_vars(self.statement, localvars)
		if if_stack.all_true() and varlike.search(cmd):
			lno = current_script_line()
			script_errors.append(["There is a potential un-substituted variable in the command %s" % cmd, lno])
		e = None
		try:
			applies, result = metacommandlist.eval(cmd)
			if applies:
				return result
		except ErrInfo as errinfo:
			# This variable reassignment is required by Python 3;
			# if the line is "except ErrInfo as e:" then an
			# UnboundLocalError occurs at the "if e:" statement.
			e = errinfo
		except:
			e = ErrInfo(type="exception", exception_msg=exception_desc())
		if e:
			status.metacommand_error = True
			subvars.add_substitution("$LAST_ERROR", cmd)
			if status.halt_on_metacommand_err:
				raise e
				#raise ErrInfo(type="cmd", command_text=cmd, other_msg=errmsg)
		if if_stack.all_true():
			# but nothing applies, because we got here.
			status.metacommand_error = True
			lno = current_script_line()
			script_errors.append(["%s:  %s" % (errmsg, cmd), lno])
			#raise ErrInfo(type="cmd", command_text=cmd, other_msg=errmsg)
		return None
	def commandline(self):
		# Returns the SQL or metacommand as in a script
		return  u"-- !x! " + self.statement


class ScriptCmd(object):
	# A SQL script object that is either a SQL statement (SqlStmt object)
	# or an execsql metacommand (MetacommandStmt object).
	# This is the basic uniform internal representation of a single
	# command or statement from an execsql script file.
	# The object attributes include source file information.
	# 'command_type' is "sql" or "cmd".
	# 'script_command' is either a SqlStmt or a MetacommandStmt object.
	def __init__(self, command_source_name, command_line_no, command_type, script_command):
		self.source = command_source_name
		self.line_no = command_line_no
		self.command_type = command_type
		self.command = script_command
	def __repr__(self):
		return u"ScriptCmd(%r, %r, %r, %r)" % (self.source, self.line_no, self.command_type, repr(self.command))
	def current_script_line(self):
		return self.line_no
	def commandline(self):
		# Returns the SQL or metacommand as in a script
		return self.command.statement if self.command_type == "sql" else u"-- !x! " + self.command.statement


class CommandList(object):
	# A list of ScriptCmd objects, including an index into the list, an
	# optional list of parameter names, and an optional set of parameter
	# values (SubvarSet).  This is the basic internal representation of
	# a list of interspersed SQL commands and metacommands.
	def __init__(self, cmdlist, listname, paramnames=None):
		# Arguments:
		#    cmdlist    : A Python list of ScriptCmd objects.  May be an empty list.
		#    listname   : A string to identify the list (e.g., a source file name or SCRIPT name).
		#    paramnames : A list of strings identifying parameters the script expects.
		# Parameter names will be used to check the names of actual arguments
		# if they are specified, but are optional: a sub-script may take
		# arguments even if parameter names have not been specified.
		if cmdlist is None:
			raise ErrInfo("error", other_msg="Initiating a command list without any commands.")
		self.listname = listname
		self.cmdlist = cmdlist
		self.cmdptr = 0
		self.paramnames = paramnames
		self.paramvals = None
		# Local variables must start with a tilde.  Other types are not allowed.
		self.localvars = LocalSubVarSet()
		self.init_if_level = None
	def add(self, script_command):
		# Adds the given ScriptCmd object to the end of the command list.
		self.cmdlist.append(script_command)
	def set_paramvals(self, paramvals):
		# Parameter values should ordinarily set immediately before the script
		# (command list) is run.
		# Arguments:
		#    paramvals : A SubVarSet object.
		self.paramvals = paramvals
		if self.paramnames is not None:
			# Check that all named parameters are provided.
			# Strip '#' off passed parameter names
			passed_paramnames = [p[0][1:] if p[0][0]=='#' else p[0][1:] for p in paramvals.substitutions]
			if not all([p in passed_paramnames for p in self.paramnames]):
				raise ErrInfo("error", other_msg="Formal and actual parameter name mismatch in call to %s." % self.listname)
	def current_command(self):
		if self.cmdptr > len(self.cmdlist) - 1:
			return None
		return self.cmdlist[self.cmdptr]
	def check_iflevels(self):
		if_excess = len(if_stack.if_levels) - self.init_if_level
		if if_excess > 0:
			sources = if_stack.script_lines(if_excess)
			src_msg = ", ".join(["input line %s" % src for src in sources])
			raise ErrInfo(type="error", other_msg="IF level mismatch at beginning and end of script; origin at or after: %s." % src_msg)
	def run_and_increment(self):
		global last_command
		global loop_nest_level
		cmditem = self.cmdlist[self.cmdptr]
		if compiling_loop:
			# Don't run this command, but save it or complete the loop and add the loop's set of commands to the stack.
			if cmditem.command_type == 'cmd' and loop_rx.match(cmditem.command.statement):
				loop_nest_level += 1
    			# Substitute any deferred substitution variables with regular substition var flags, e.g.: "!!somevar!!"
				m = defer_rx.findall(cmditem.command.statement)
				if m is not None:
					for dv in m:
						rep = "!!" +  dv[1] + "!!"
						cmditem.command.statement = cmditem.command.statement.replace(dv[0], rep)
				loopcommandstack[-1].add(cmditem)
			elif cmditem.command_type == 'cmd' and endloop_rx.match(cmditem.command.statement):
				if loop_nest_level == 0:
					endloop()
				else:
					loop_nest_level -= 1
					loopcommandstack[-1].add(cmditem)
			else:
				loopcommandstack[-1].add(cmditem)
		else:
			last_command = cmditem
			if cmditem.command_type == "sql" and status.batch.in_batch():
				status.batch.using_db(dbs.current())
			subvars.add_substitution("$CURRENT_SCRIPT", cmditem.source)
			subvars.add_substitution("$CURRENT_SCRIPT_PATH", os.path.dirname(os.path.abspath(cmditem.source)) + os.sep)
			subvars.add_substitution("$CURRENT_SCRIPT_NAME", os.path.basename(cmditem.source))
			subvars.add_substitution("$CURRENT_SCRIPT_LINE", str(cmditem.line_no))
			subvars.add_substitution("$SCRIPT_LINE", str(cmditem.line_no))
			cmditem.command.run(self.localvars.merge(self.paramvals), not status.batch.in_batch())
		self.cmdptr += 1
	def run_next(self):
		global last_command
		if self.cmdptr == 0:
			self.init_if_level = len(if_stack.if_levels)
		if self.cmdptr > len(self.cmdlist) - 1:
			self.check_iflevels()
			raise StopIteration
		self.run_and_increment()
	def __iter__(self):
		return self
	def __next__(self):
		if self.cmdptr > len(self.cmdlist) - 1:
			raise StopIteration
		scriptcmd = self.cmdlist[self.cmdptr]
		self.cmdptr += 1
		return scriptcmd


class CommandListWhileLoop(CommandList):
	# Subclass of CommandList() that will loop WHILE a condition is met.
	# Additional argument:
	#	loopcondition : A string containing the conditional for continuing the WHILE loop.
	def __init__(self, cmdlist, listname, paramnames, loopcondition):
		super(CommandListWhileLoop, self).__init__(cmdlist, listname, paramnames)
		self.loopcondition = loopcondition
	def run_next(self):
		global last_command
		if self.cmdptr == 0:
			self.init_if_level = len(if_stack.if_levels)
			if not CondParser(substitute_vars(self.loopcondition)).parse().eval():
				raise StopIteration
		if self.cmdptr > len(self.cmdlist) - 1:
			self.check_iflevels()
			self.cmdptr = 0
		else:
			self.run_and_increment()


class CommandListUntilLoop(CommandList):
	# Subclass of CommandList() that will loop UNTIL a condition is met.
	# Additional argument:
	#    loopcondition : A string containing the conditional for terminating the UNTIL loop.
	def __init__(self, cmdlist, listname, paramnames, loopcondition):
		super(CommandListUntilLoop, self).__init__(cmdlist, listname, paramnames)
		self.loopcondition = loopcondition
	def run_next(self):
		global last_command
		if self.cmdptr == 0:
			self.init_if_level = len(if_stack.if_levels)
		if self.cmdptr > len(self.cmdlist) - 1:
			self.check_iflevels()
			if CondParser(substitute_vars(self.loopcondition)).parse().eval():
				raise StopIteration
			self.cmdptr = 0
		else:
			self.run_and_increment()

class ScriptFile(EncodedFile):
	# A file reader that returns lines and records the line number.
	def __init__(self, scriptfname, file_encoding):
		super(ScriptFile, self).__init__(scriptfname, file_encoding)
		self.lno = 0
		self.f = self.open("r")
	def __repr__(self):
		return u"ScriptFile(%r, %r)" % (super(ScriptFile, self).filename, super(ScriptFile, self).encoding)
	def __iter__(self):
		return self
	def __next__(self):
		l = next(self.f)
		self.lno += 1
		return l


class ScriptExecSpec(object):
	# An object that stores the specifications for executing a SCRIPT,
	# for later use.  This is specifically intended to be used by
	# ON ERROR_HALT EXECUTE SCRIPT and ON CANCEL_HALT EXECUTE SCRIPT.
	args_rx = re.compile(r'(?P<param>#?\w+)\s*=\s*(?P<arg>(?:(?:[^"\'\[][^,\)]*)|(?:"[^"]*")|(?:\'[^\']*\')|(?:\[[^\]]*\])))', re.I)
	def __init__(self, **kwargs):
		self.script_id = kwargs["script_id"].lower()
		if self.script_id not in savedscripts.keys():
			raise ErrInfo("cmd", other_msg="There is no SCRIPT named %s." % self.script_id)
		self.arg_exp = kwargs["argexp"]
		self.looptype = kwargs["looptype"].upper() if "looptype" in kwargs and kwargs["looptype"] is not None else None
		self.loopcond = kwargs["loopcond"] if "loopcond" in kwargs else None
	def execute(self):
		# Copy the saved script because otherwise the memory-recovery nullification
		# of completed commands will erase the saved script commands.
		cl = copy.deepcopy(savedscripts[self.script_id])
		# If looping is specified, redirect to appropriate CommandList() subclass 
		if self.looptype is not None:
			cl = CommandListWhileLoop(cl.cmdlist, cl.listname, cl.paramnames, self.loopcond) if self.looptype == 'WHILE' else CommandListUntilLoop(cl.cmdlist, cl.listname, cl.paramnames, self.loopcond)
		# If there are any argument expressions, parse the arguments
		if self.arg_exp is not None:
			# Clean arg_exp
			all_args = re.findall(self.args_rx, self.arg_exp)
			all_cleaned_args = [(ae[0], wo_quotes(ae[1])) for ae in all_args]
			# Prepend '#' on each param name if the user did not include one
			all_prepared_args = [(ae[0] if ae[0][0]=='#' else '#' + ae[0], ae[1]) for ae in all_cleaned_args]
			scriptvarset = ScriptArgSubVarSet()
			for param, arg in all_prepared_args:
				scriptvarset.add_substitution(param, arg)
			cl.set_paramvals(scriptvarset)
		# If argument expressions were NOT found, confirm that the command list is not expecting named params
		else:
			# because if it IS, there's a problem.
			if cl.paramnames is not None:
				raise ErrInfo("error", other_msg="Missing expected parameters (%s) in call to %s." % (", ".join(cl.paramnames), cl.listname))
		commandliststack.append(cl)


# End of scripting classes.
#===============================================================================================



#===============================================================================================
#----- Parsers
#
# Parsers for conditional and numeric expressions.

#-------------------------------------------------------------------------------------
# Source string objects.  These are strings (metacommands arguments) with
# a pointer into the string.
#-------------------------------------------------------------------------------------
class SourceString(object):
	def __init__(self, source_string):
		self.str = source_string
		self.currpos = 0
	def eoi(self):
		# Returns True or False indicating whether or not there is any of
		# the source string left to be consumed.
		return self.currpos >= len(self.str)
	def eat_whitespace(self):
		while not self.eoi() and self.str[self.currpos] in [' ', '\t', '\n']:
			self.currpos += 1
	def match_str(self, str):
		# Tries to match the 'str' argument at the current position in the
		# source string.  Matching is case-insensitive.  If matching succeeds,
		# the matched string is returned and the internal pointer is incremented.
		# If matching fails, None is returned and the internal pointer is unchanged.
		self.eat_whitespace()
		if self.eoi():
			return None
		else:
			found = self.str.lower().startswith(str.lower(), self.currpos)
			if found:
				matched = self.str[self.currpos:self.currpos+len(str)]
				self.currpos += len(str)
				return matched
			else:
				return None
	def match_regex(self, regex):
		# Tries to match the 'regex' argument at the current position in the
		# source string.  If it succeeds, a dictionary of all of the named
		# groups is returned, and the internal pointer is incremented.
		self.eat_whitespace()
		if self.eoi():
			return None
		else:
			m = regex.match(self.str[self.currpos:])
			if m:
				self.currpos += m.end(0)
				return m.groupdict() or {}
			else:
				return None
	def match_metacommand(self, commandlist):
		# Tries to match text at the current position to any metacommand
		# in the specified commandlist. 
		# If it succeeds, the return value is a tuple of the MetaCommand object
		# and a dictionary of all of the named groups.  The internal pointer is
		# incremented past the match.
		self.eat_whitespace()
		if self.eoi():
			return None
		else:
			m = commandlist.get_match(self.str[self.currpos:])
			if m is not None:
				self.currpos += m[1].end(0)
				return (m[0], m[1].groupdict() or {})
			else:
				return None
	def remainder(self):
		return self.str[self.currpos:]
 
#-------------------------------------------------------------------------------------
#	Classes for AST operator types.
#-------------------------------------------------------------------------------------
class CondTokens(object):
	AND, OR, NOT, CONDITIONAL = range(4)

class NumTokens(object):
	MUL, DIV, ADD, SUB, NUMBER = range(5)

#-------------------------------------------------------------------------------------
#	AST for conditional expressions
#-------------------------------------------------------------------------------------
class CondAstNode(CondTokens, object):
	def __init__(self, type, cond1, cond2):
		# 'type' should be one of the constants AND, OR, NOT, CONDITIONAL.
		# For AND and OR types, 'cond1' and 'cond2' should be a subtree (a CondAstNode)
		# For NOT type, 'cond1' should be a CondAstNOde and 'cond2' should be None
		# For CONDITIONAL type, cond1' should be a tuple consisting of metacommand object and
		# its dictionary of named groups (mcmd, groupdict) and 'cond2' should be None.
		self.type = type
		self.left = cond1
		if type not in (self.CONDITIONAL, self.NOT):
			self.right = cond2
		else:
			self.right = None
	def eval(self):
		# Evaluates the subtrees and/or conditional value for this node,
		# returning True or False.
		if self.type == self.CONDITIONAL:
			exec_fn = self.left[0].exec_fn
			cmdargs = self.left[1]
			return exec_fn(**cmdargs)
		if self.type == self.NOT:
			return not self.left.eval()
		lcond = self.left.eval()
		if self.type == self.AND:
			if not lcond: return False
			return self.right.eval()
		if self.type == self.OR:
			if lcond: return True
			return self.right.eval()

#-------------------------------------------------------------------------------------
#	AST for numeric expressions
#-------------------------------------------------------------------------------------
class NumericAstNode(NumTokens, object):
	def __init__(self, type, value1, value2):
		# 'type' should be one of the constants MUL, DIV, ADD, SUB, OR NUMBER.
		# 'value1' and 'value2' should each be either a subtree (a
		# NumericAstNode) or (only 'value1' should be) a number.
		self.type = type
		self.left = value1
		if type != self.NUMBER:
			self.right = value2
		else:
			self.right = None
	def eval(self):
		# Evaluates the subtrees and/or numeric value for this node,
		# returning a numeric value.
		if self.type == self.NUMBER:
			return self.left
		else:
			lnum = self.left.eval()
			rnum = self.right.eval()
			if self.type == self.MUL:
				return lnum * rnum
			elif self.type == self.DIV:
				return lnum / rnum
			elif self.type == self.ADD:
				return lnum + rnum
			else:
				return lnum - rnum

#-------------------------------------------------------------------------------------
#	Conditional Parser
#-------------------------------------------------------------------------------------
class CondParserError(Exception):
	def __init__(self, msg):
		self.value = msg
	def __repr__(self):
		return "ConditionalParserError(%r)" % self.value


class CondParser(CondTokens, object):
	# Takes a conditional expression string.
	def __init__(self, condexpr):
		self.condexpr = condexpr
		self.cond_expr = SourceString(condexpr)
	def match_not(self):
		# Try to match 'NOT' operator. If not found, return None
		m1 = self.cond_expr.match_str('NOT')
		if m1 is not None:
			return self.NOT
		return None
	def match_andop(self):
		# Try to match 'AND' operator. If not found, return None
		m1 = self.cond_expr.match_str('AND')
		if m1 is not None:
			return self.AND
		return None
	def match_orop(self):
		# Try to match 'OR' operator. If not found, return None
		m1 = self.cond_expr.match_str('OR')
		if m1 is not None:
			return self.OR
		return None
	def factor(self): 
		m1 = self.match_not()
		if m1 is not None:
			m1 = self.factor()
			return CondAstNode(self.NOT, m1, None)
		# Find the matching metacommand -- get a tuple consisting of (metacommand, groupdict)
		m1 = self.cond_expr.match_metacommand(conditionallist)
		if m1 is not None:
			m1[1]["metacommandline"] = self.condexpr
			return CondAstNode(self.CONDITIONAL, m1, None)
		else:
			if self.cond_expr.match_str("(") is not None:
				m1 = self.expression()
				rp = self.cond_expr.match_str(")")
				if rp is None:
					raise CondParserError("Expected closing parenthesis at position %s of %s." % (self.cond_expr.currpos, self.cond_expr.str))
				return m1
			else:
				raise CondParserError("Can't parse a factor at position %s of %s." % (self.cond_expr.currpos, self.cond_expr.str))
	def term(self):
		m1 = self.factor()
		andop = self.match_andop()
		if andop is not None:
			m2 = self.term()
			return CondAstNode(andop, m1, m2)
		else:
			return m1
	def expression(self):
		e1 = self.term()
		orop = self.match_orop()
		if orop is not None:
			e2 = self.expression()
			return CondAstNode(orop, e1, e2)
		else:
			return e1
	def parse(self):
		exp = self.expression()
		if not self.cond_expr.eoi():
			raise CondParserError("Conditional expression parser did not consume entire string; remainder = %s." % self.cond_expr.remainder())
		return exp

#-------------------------------------------------------------------------------------
#		Numeric Parser
#-------------------------------------------------------------------------------------
class NumericParserError(Exception):
	def __init__(self, msg):
		self.value = msg
	def __repr__(self):
		return "NumericParserError(%r)" % self.value

class NumericParser(NumTokens, object):
	# Takes a numeric expression string
	def __init__(self, numexpr):
		self.num_expr = SourceString(numexpr)
		self.rxint = re.compile(r'(?P<int_num>[+-]?[0-9]+)')
		self.rxfloat = re.compile(r'(?P<float_num>[+-]?(?:(?:[0-9]*\.[0-9]+)|(?:[0-9]+\.[0-9]*)))')
	def match_number(self):
		# Try to match a number in the source string.
		# Return it if matched, return None if unmatched.
		m1 = self.num_expr.match_regex(self.rxfloat)
		if m1 is not None:
			return float(m1['float_num'])
		else:
			m2 = self.num_expr.match_regex(self.rxint)
			if m2 is not None:
				return int(m2['int_num'])
		return None
	def match_mulop(self):
		# Try to match a multiplication or division operator in the source string.
		# if found, return the matching operator type.  If not found, return None.
		m1 = self.num_expr.match_str("*")
		if m1 is not None:
			return self.MUL
		else:
			m2 = self.num_expr.match_str("/")
			if m2 is not None:
				return self.DIV
		return None
	def match_addop(self):
		# Try to match an addition or division operator in the source string.
		# if found, return the matching operator type.  If not found, return None.
		m1 = self.num_expr.match_str("+")
		if m1 is not None:
			return self.ADD
		else:
			m2 = self.num_expr.match_str("-")
			if m2 is not None:
				return self.SUB
		return None
	def factor(self):
		# Parses a factor out of the source string and returns the
		# AST node that is created.
		m1 = self.match_number()
		if m1 is not None:
			return NumericAstNode(self.NUMBER, m1, None)
		else:
			if self.num_expr.match_str("(") is not None:
				m1 = self.expression()
				rp = self.num_expr.match_str(")")
				if rp is None:
					raise NumericParserError("Expected closing parenthesis at position %s of %s." % (self.num_expr.currpos, self.num_expr.str))
				else:
					return m1
			else:
				raise NumericParserError("Can't parse a factor at position %s of %s." % (self.num_expr.currpos, self.num_expr.str))
	def term(self):
		# Parses a term out of the source string and returns the
		# AST node that is created.
		m1 = self.factor()
		mulop = self.match_mulop()
		if mulop is not None:
			m2 = self.term()
			return NumericAstNode(mulop, m1, m2)
		else:
			return m1
	def expression(self):
		# Parses an expression out of the source string and returns the
		# AST node that is created.
		e1 = self.term()
		if e1 is None:
			return
		addop = self.match_addop()
		if addop is not None:
			e2 = self.expression()
			return NumericAstNode(addop, e1, e2)
		else:
			return e1
	def parse(self):
		exp = self.expression()
		if not self.num_expr.eoi():
			raise NumericParserError("Numeric expression parser did not consume entire string; remainder = %s." % self.num_expr.remainder())
		return exp


# End of Parser classes
#===============================================================================================



#===============================================================================================
#-----  METACOMMAND FUNCTIONS


#****	DEBUG WRITE METACOMMANDLIST
# Undocumented; used to acquire data used to set the ordering of metacommands.
def x_debug_write_metacommands(**kwargs):
	output_dest = kwargs['filename']
	if output_dest is None or output_dest == 'stdout':
		ofile = output
	else:
		ofile = EncodedFile(output_dest, conf.output_encoding).open("w")
	for m in metacommandlist:
		ofile.write(u"(%s)  %s\n" % (m.hitcount, m.rx.pattern))

metacommandlist.add(ins_fn_rxs(r'^\s*DEBUG\s+WRITE\s+METACOMMANDLIST\s+TO\s+', r'\s*$'), x_debug_write_metacommands)


#****	BREAK
def x_break(**kwargs):
	global commandlistack
	global loopcommandstack
	if len(commandliststack) == 1:
		line_no = current_script_line()
		script_errors.append(["BREAK metacommand with no command nesting", line_no])
	else:
		if_stack.if_levels = if_stack.if_levels[:commandliststack[-1].init_if_level]
		commandliststack.pop()
	return None

metacommandlist.add(r'^\s*BREAK\s*$', x_break)


#****	CD
def x_cd(**kwargs):
	new_dir = unquoted(kwargs['dir'])
	if not os.path.isdir(new_dir):
		raise ErrInfo(type="cmd", command_text=kwargs['metacommandline'], other_msg='Directory does not exist')
	os.chdir(new_dir)
	lno = current_script_line()
	return None

metacommandlist.add(r'^\s*CD\s+(?P<dir>.+)\s*$', x_cd)


#****	SUB_LOCAL
# Define a local variable.  Local variables must start with a tilde.  As a convenience, one
# will be added if missing.
def x_sub_local(**kwargs):
	varname = kwargs['match']
	if varname[0] != '~':
		varname = '~' + varname
	global commandliststack
	commandliststack[-1].localvars.add_substitution(varname, kwargs['repl'])
	return None

metacommandlist.add(r'^\s*SUB_LOCAL\s+(?P<match>~?\w+)\s+(?P<repl>.+)$', x_sub_local, "SUB", "Define a local variable consisting of a string to match and a replacement for it.")


#****	WAIT_UNTIL
def x_wait_until(**kwargs):
	countdown = int(kwargs['seconds'])
	while countdown > 0:
		if xcmd_test(kwargs['condition']):
			return
		time.sleep(1)
		countdown -= 1
	if kwargs['end'].lower() == 'halt':
		return None
	return None

metacommandlist.add(r'^\s*WAIT_UNTIL\s+(?P<condition>.+)\s+(?P<end>HALT|CONTINUE)\s+AFTER\s+(?P<seconds>\d+)\s+SECONDS\s*$', x_wait_until)


#****	ON ERROR_HALT EXECUTE SCRIPT CLEAR
def x_error_halt_exec_clear(**kwargs):
	global err_halt_exec
	err_halt_exec = None

metacommandlist.add(r'^\s*ON\s+ERROR_HALT\s+EXEC\s+CLEAR\s*$', x_error_halt_exec_clear)


#****	RESET COUNTER
def x_reset_counter(**kwargs):
	ctr_no = int(kwargs["counter_no"])
	counters.remove_counter(ctr_no)

metacommandlist.add(r'^\s*RESET\s+COUNTER\s+(?P<counter_no>\d+)\s*$', x_reset_counter)


#****	RESET COUNTERS
def x_reset_counters(**kwargs):
	counters.remove_all_counters()

metacommandlist.add(r'^\s*RESET\s+COUNTERS\s*$', x_reset_counters)


#****	SET COUNTER
def x_set_counter(**kwargs):
	ctr_no = int(kwargs["counter_no"])
	ctr_expr = kwargs["value"]
	counters.set_counter(ctr_no, int(math.floor(NumericParser(ctr_expr).parse().eval())))

metacommandlist.add(r'^\s*SET\s+COUNTER\s+(?P<counter_no>\d+)\s+TO\s+(?P<value>[0-9+\-*/() ]+)\s*$', x_set_counter)


#****	TIMER
def x_timer(**kwargs):
	onoff = kwargs["onoff"].lower()
	if onoff == 'on':
		timer.start()
	else:
		timer.stop()

metacommandlist.add(r'^\s*TIMER\s+(?P<onoff>ON|OFF)\s*$', x_timer)


#****	BEGIN BATCH
def x_begin_batch(**kwargs):
	status.batch.new_batch()
	return None

metacommandlist.add(r'^\s*BEGIN\s+BATCH\s*$', x_begin_batch)


#****	END BATCH
def x_end_batch(**kwargs):
	status.batch.end_batch()
	return None

# Set a name so this can be found and evaluated during processing, when all other metacommands are ignored.
metacommandlist.add(r'^\s*END\s+BATCH\s*$', x_end_batch, "END BATCH", run_in_batch=True)


#****	ROLLBACK BATCH
def x_rollback(**kwargs):
	status.batch.rollback_batch()

metacommandlist.add(r'^\s*ROLLBACK(:?\s+BATCH)?\s*$', x_rollback, "ROLLBACK BATCH", run_in_batch=True)


#****	ERROR_HALT
def x_error_halt(**kwargs):
	flag = kwargs['onoff'].lower()
	if not flag in ('on', 'off', 'yes', 'no', 'true', 'false'):
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg=u"Unrecognized flag for error handling: %s" % flag)
	status.halt_on_err = flag in ('on', 'yes', 'true')
	return None

metacommandlist.add(r'\s*ERROR_HALT\s+(?P<onoff>ON|OFF|YES|NO|TRUE|FALSE)\s*$', x_error_halt)


#****	METACOMMAND_ERROR_HALT
def x_metacommand_error_halt(**kwargs):
	flag = kwargs['onoff'].lower()
	if not flag in ('on', 'off', 'yes', 'no', 'true', 'false'):
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg=u"Unrecognized flag for metacommand error handling: %s" % flag)
	status.halt_on_metacommand_err = flag in ('on', 'yes', 'true')
	return None

metacommandlist.add(r'\s*METACOMMAND_ERROR_HALT\s+(?P<onoff>ON|OFF|YES|NO|TRUE|FALSE)\s*$', x_metacommand_error_halt, set_error_flag=False)


#****	LOOP
def x_loop(**kwargs):
	global compiling_loop
	compiling_loop = True
	looptype = kwargs["looptype"].upper()
	loopcond = kwargs["loopcond"]
	listname = 'loop'+str(len(loopcommandstack)+1)
	if looptype == 'WHILE':
		loopcommandstack.append(CommandListWhileLoop([], listname, paramnames=None, loopcondition=loopcond))
	else:
		loopcommandstack.append(CommandListUntilLoop([], listname, paramnames=None, loopcondition=loopcond))

metacommandlist.add(r'^\s*LOOP\s+(?P<looptype>WHILE|UNTIL)\s*\(\s*(?P<loopcond>.+)\s*\)\s*$', x_loop)


#****	END LOOP
def endloop():
	if len(loopcommandstack) == 0:
		raise ErrInfo("error", other_msg="END LOOP metacommand without a matching preceding LOOP metacommand.")
	global compiling_loop
	compiling_loop = False
	commandliststack.append(loopcommandstack[-1])
	loopcommandstack.pop()


#****	SUB_EMPTY
def x_sub_empty(**kwargs):
	varname = kwargs['match']
	# Get subvarset assignment and cleansed variable name
	subvarset, varname = get_subvarset(varname, kwargs['metacommandline'])
	subvarset.add_substitution(varname, u'')
	return None

metacommandlist.add(r'^\s*SUB_EMPTY\s+(?P<match>[+~]?\w+)\s*$', x_sub_empty)


#****	SUB_ADD
def x_sub_add(**kwargs):
	varname = kwargs["match"]
	increment_expr = kwargs["increment"]
	# Get subvarset assignment and cleansed variable name
	subvarset, varname = get_subvarset(varname, kwargs['metacommandline'])
	subvarset.increment_by(varname, NumericParser(increment_expr).parse().eval())
	return None

metacommandlist.add(r'^\s*SUB_ADD\s+(?P<match>[+~]?\w+)\s+(?P<increment>[+\-0-9\.*/() ]+)\s*$', x_sub_add)


#****	SUB_APPEND
def x_sub_append(**kwargs):
	varname = kwargs["match"]
	# Get subvarset assignment and cleansed variable name
	subvarset, varname = get_subvarset(varname, kwargs['metacommandline'])
	subvarset.append_substitution(varname, kwargs['repl'])
	return None

metacommandlist.add(r'^\s*SUB_APPEND\s+(?P<match>[+~]?\w+)\s(?P<repl>(.|\n)*)$', x_sub_append)


#****	BLOCK ORIF
def x_if_orif(**kwargs):
	if if_stack.all_true():
		return None		# Short-circuit evaluation
	if if_stack.only_current_false():
		if_stack.replace(xcmd_test(kwargs['condtest']))
	return None

metacommandlist.add(r'^\s*ORIF\s*\(\s*(?P<condtest>.+)\s*\)\s*$', x_if_orif, run_when_false=True)


#****	EXTEND SCRIPT WITH SCRIPT
#****	APPEND SCRIPT
def x_extendscript(**kwargs):
	script1 = kwargs["script1"].lower()
	if script1 not in savedscripts:
		raise ErrInfo("cmd", other_msg="There is no SCRIPT named %s." % script1)
	script2 = kwargs["script2"].lower()
	if script1 not in savedscripts:
		raise ErrInfo("cmd", other_msg="There is no SCRIPT named %s." % script2)
	s1 = savedscripts[script1]
	s2 = savedscripts[script2]
	for cmd in s1.cmdlist:
		s2.add(cmd)
	if s1.paramnames is not None:
		if s2.paramnames is None:
			s2.paramnames = []
		for param in s1.paramnames:
			if param not in s2.paramnames:
				s2.paramnames.append(param)

metacommandlist.add(r'\s*EXTEND\s+SCRIPT\s+(?P<script2>\w+)\s+WITH\s+SCRIPT\s+(?P<script1>\w+)\s*$', x_extendscript)
metacommandlist.add(r'\s*APPEND\s+SCRIPT\s+(?P<script1>\w+)\s+TO\s+(?P<script2>\w+)\s*$', x_extendscript)


#****	EXTEND SCRIPT WITH METACOMMAND
def x_extendscript_metacommand(**kwargs):
	script = kwargs["script"].lower()
	if script not in savedscripts:
		raise ErrInfo("cmd", other_msg="There is no SCRIPT named %s." % script)
	script_line_no = current_script_line()
	savedscripts[script].add(ScriptCmd(script_file, script_line_no, 'cmd', MetacommandStmt(kwargs["cmd"])))

metacommandlist.add(r'\s*EXTEND\s+SCRIPT\s+(?P<script>\w+)\s+WITH\s+METACOMMAND\s+(?P<cmd>.+)\s*$', x_extendscript_metacommand)


#****	EXTEND SCRIPT WITH SQL
def x_extendscript_sql(**kwargs):
	script = kwargs["script"].lower()
	if script not in savedscripts:
		raise ErrInfo("cmd", other_msg="There is no SCRIPT named %s." % script)
	sql = kwargs["sql"]
	script_line_no = current_script_line()
	savedscripts[script].add(ScriptCmd(script_file, script_line_no , 'sql', SqlStmt(kwargs["sql"])))

metacommandlist.add(r'\s*EXTEND\s+SCRIPT\s+(?P<script>\w+)\s+WITH\s+SQL\s+(?P<sql>.+;)\s*$', x_extendscript_sql)


#****	ON ERROR_HALT EXECUTE SCRIPT
def x_error_halt_exec(**kwargs):
	global err_halt_exec
	err_halt_exec = ScriptExecSpec(**kwargs)

metacommandlist.add(r'^\s*ON\s+ERROR_HALT\s+EXEC(?:UTE)?\s+SCRIPT\s+(?P<script_id>\w+)(?:(?:\s+WITH)?(?:\s+ARG(?:UMENT)?S?)?\s*\(\s*(?P<argexp>#?\w+\s*=\s*(?:(?:[^"\'\[][^,\)]*)|(?:"[^"]*")|(?:\'[^\']*\')|(?:\[[^\]]*\]))(?:\s*,\s*#?\w+\s*=\s*(?:(?:[^"\'\[][^,\)]*)|(?:"[^"]*")|(?:\'[^\']*\')|(?:\[[^\]]*\])))*)\s*\))?\s*$', x_error_halt_exec)


#****	DEBUG_WRITE
def x_debug_write(**kwargs):
	msg = u'%s\n' % kwargs['text']
	print("%s\n" % msg)
	return None

metacommandlist.add(r'^\s*WRITE\s+\~(?P<text>([^\~]|\n)*)\~\s*$', x_debug_write)
metacommandlist.add(r'^\s*WRITE\s+\#(?P<text>([^\#]|\n)*)\#\s*$', x_debug_write)
metacommandlist.add(r'^\s*WRITE\s+\`(?P<text>([^\`]|\n)*)\`\s*$', x_debug_write)
metacommandlist.add(r'^\s*WRITE\s+\[(?P<text>([^\]]|\n)*)\]\s*$', x_debug_write)
metacommandlist.add(r'^\s*WRITE\s+\'(?P<text>([^\']|\n)*)\'\s*$', x_debug_write)
metacommandlist.add(r'^\s*WRITE\s+"(?P<text>([^"]|\n)*)"\s*$', x_debug_write)



#****	INCLUDE
def x_include(**kwargs):
	filename = kwargs['filename']
	if len(filename) > 1 and filename[0] == "~" and filename[1] == os.sep:
		filename = os.path.join(os.path.expanduser(r'~'), filename[2:])
	exists = kwargs['exists']
	if exists is not None:
		if os.path.isfile(filename):
			read_sqlfile(filename)
	else:
		if not os.path.isfile(filename):
			raise ErrInfo(type="error", other_msg="File %s does not exist." % filename)
		read_sqlfile(filename)
	return None

metacommandlist.add(ins_fn_rxs(r'^\s*INCLUDE(?P<exists>\s+IF\s+EXISTS?)?\s+', r'\s*$'), x_include)


#****	RM_SUB
def x_rm_sub(**kwargs):
	varname = kwargs["match"]
	subvarset = subvars if varname[0] != '~' else commandliststack[-1].localvars
	subvarset.remove_substitution(varname)
	return None

metacommandlist.add(r'^\s*RM_SUB\s+(?P<match>~?\w+)\s*$', x_rm_sub)


#****	BLOCK ELSEIF
def x_if_elseif(**kwargs):
	if if_stack.only_current_false():
		if_stack.replace(xcmd_test(kwargs['condtest']))
	else:
		if_stack.replace(False)
	return None

metacommandlist.add(r'^\s*ELSEIF\s*\(\s*(?P<condtest>.+)\s*\)\s*$', x_if_elseif, run_when_false=True)


#****	AUTOCOMMIT OFF
def x_autocommit_off(**kwargs):
	db = dbs.current()
	db.autocommit_off()

metacommandlist.add(r'^\s*AUTOCOMMIT\s+OFF\s*$', x_autocommit_off)


#****	AUTOCOMMIT ON
def x_autocommit_on(**kwargs):
	action = kwargs['action']
	if action is not None:
		action = action.lower()
	db = dbs.current()
	db.autocommit_on()
	if action is not None:
		if action == 'commit':
			db.commit()
		else:
			db.rollback()

metacommandlist.add(r'^\s*AUTOCOMMIT\s+ON(?:\s+WITH\s+(?P<action>COMMIT|ROLLBACK))?\s*$', x_autocommit_on)


#****	BLOCK ANDIF
def x_if_andif(**kwargs):
	if if_stack.all_true():
		if_stack.replace(if_stack.current() and xcmd_test(kwargs['condtest']))
	return None

metacommandlist.add(r'^\s*ANDIF\s*\(\s*(?P<condtest>.+)\s*\)\s*$', x_if_andif)


#****	SELECT_SUB
def x_selectsub(**kwargs):
	sql = u"select * from %s;" % kwargs["datasource"]
	db = dbs.current()
	line_no = current_script_line()
	nodatamsg = "There are no data in %s to use with the SELECT_SUB metacommand (line %d)." % (kwargs["datasource"], line_no)
	try:
		hdrs, rec = db.select_rowsource(sql)
	except ErrInfo:
		raise
	except:
		raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg="Can't get headers and rows from %s." % sql)
	# Remove any existing variables with these names
	for subvar in hdrs:
		subvar = u'@'+subvar
		if subvars.sub_exists(subvar):
			subvars.remove_substitution(subvar)
	try:
		row1 = next(rec)
	except StopIteration:
		row1 = None
	except:
		raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg=nodatamsg)
	if row1:
		for i, item in enumerate(row1):
			if item is None:
				item = u''
			if sys.version_info < (3,):
				item = unicode(item)
			else:
				item = str(item)
			match_str = u"@" + hdrs[i]
			subvars.add_substitution(match_str, item)
	return None

metacommandlist.add(r'^\s*SELECT_SUB\s+(?P<datasource>.+)\s*$', x_selectsub)


#****	SUBDATA
def x_subdata(**kwargs):
	varname = kwargs["match"]
	sql = u"select * from %s;" % kwargs["datasource"]
	db = dbs.current()
	line_no = current_script_line()
	errmsg = "There are no data in %s to use with the SUBDATA metacommand (line %d)." % (kwargs["datasource"], line_no)
	# Get subvarset assignment and cleansed variable name
	subvarset, varname = get_subvarset(varname, kwargs['metacommandline'])
	subvarset.remove_substitution(varname)
	# Exceptions should be trapped by the caller, so are re-raised here after settting status
	try:
		hdrs, rec = db.select_rowsource(sql)
	except ErrInfo:
		raise
	except:
		raise ErrInfo(type="exception", exception_msg=exception_desc(), other_msg="Can't get headers and rows from %s." % sql)
	try:
		row1 = next(rec)
	except:
		row1 = None
	if row1:
		dataval = row1[0]
		if dataval is None:
			dataval = u''
		if not isinstance(dataval, str):
			if sys.version_info < (3,):
				dataval = unicode(dataval)
			else:
				dataval = str(dataval)
		subvarset.add_substitution(varname, dataval)
	return None

metacommandlist.add(r'^\s*SUBDATA\s+(?P<match>[+~]?\w+)\s+(?P<datasource>.+)\s*$', x_subdata)


#****	IF
def x_if(**kwargs):
	tf_value = xcmd_test(kwargs['condtest'])
	if tf_value:
		line_no = current_script_line()
		metacmd = MetacommandStmt(kwargs['condcmd'])
		script_cmd = ScriptCmd(src, line_no, "cmd", metacmd)
		cmdlist = CommandList([script_cmd], "%s_%d" % (src, line_no))
		commandliststack.append(cmdlist)
	return None

metacommandlist.add(r'^\s*IF\s*\(\s*(?P<condtest>.+)\s*\)\s*{\s*(?P<condcmd>.+)\s*}\s*$', x_if)


#****	EXECUTE SCRIPT
def x_executescript(**kwargs):
	exists = kwargs["exists"]
	script_id = kwargs["script_id"].lower()
	if exists is None or (exists is not None and script_id in savedscripts):
		ScriptExecSpec(**kwargs).execute()

metacommandlist.add(r'^\s*EXEC(?:UTE)?\s+SCRIPT(?:\s+(?P<exists>IF\s+EXISTS))?\s+(?P<script_id>\w+)(?:(?:\s+WITH)?(?:\s+ARG(?:UMENT)?S?)?\s*\(\s*(?P<argexp>#?\w+\s*=\s*(?:(?:[^"\'\[][^,\)]*)|(?:"[^"]*")|(?:\'[^\']*\')|(?:\[[^\]]*\]))(?:\s*,\s*#?\w+\s*=\s*(?:(?:[^"\'\[][^,\)]*)|(?:"[^"]*")|(?:\'[^\']*\')|(?:\[[^\]]*\])))*)\s*\))?(?:\s+(?P<looptype>WHILE|UNTIL)\s*\(\s*(?P<loopcond>.+)\s*\))?\s*$', x_executescript)


#****	BLOCK ELSE
def x_if_else(**kwargs):
	if if_stack.all_true() or if_stack.only_current_false():
		if_stack.invert()
	return None

metacommandlist.add(r'^\s*ELSE\s*$', x_if_else, run_when_false=True)


#****	SUB
def x_sub(**kwargs):
	varname = kwargs['match']
	# Get subvarset assignment and cleansed variable name
	subvarset, varname = get_subvarset(varname, kwargs['metacommandline'])
	subvarset.add_substitution(varname, kwargs['repl'])
	return None

metacommandlist.add(r'^\s*SUB\s+(?P<match>[+~]?\w+)\s+(?P<repl>.+)$', x_sub, "SUB", "Define a string to match and a replacement for it.")


#****	BLOCK IF
def x_if_block(**kwargs):
	if if_stack.all_true():
		test_result = xcmd_test(kwargs['condtest'])
		if_stack.nest(test_result)
	else:
		if_stack.nest(False)
	return None

metacommandlist.add(r'^\s*IF\s*\(\s*(?P<condtest>.+)\s*\)\s*$', x_if_block, run_when_false=True)


#****	BLOCK ENDIF
def x_if_end(**kwargs):
	if_stack.unnest()
	return None

metacommandlist.add(r'^\s*ENDIF\s*$', x_if_end, run_when_false=True)


#****	DEBUG WRITE METACOMMANDLIST
# Undocumented; used to acquire data used to set the ordering of metacommands.
def x_debug_write_metacommands(**kwargs):
	output_dest = kwargs['filename']
	if output_dest is None or output_dest == 'stdout':
		ofile = output
	else:
		ofile = EncodedFile(output_dest, conf.output_encoding).open("w")
	for m in metacommandlist:
		ofile.write(u"(%s)  %s\n" % (m.hitcount, m.rx.pattern))

metacommandlist.add(ins_fn_rxs(r'^\s*DEBUG\s+WRITE\s+METACOMMANDLIST\s+TO\s+', r'\s*$'), x_debug_write_metacommands)


#	End of metacommand definitions.
#===============================================================================================


#===============================================================================================
#-----  CONDITIONAL TESTS FOR METACOMMANDS


def xf_contains(**kwargs):
	s1 = kwargs["string1"]
	s2 = kwargs["string2"]
	if kwargs["ignorecase"] and kwargs["ignorecase"].lower() == "i":
		s1 = s1.lower()
		s2 = s2.lower()
	return s2 in s1

conditionallist.add(r'^\s*CONTAINS\s*\(\s*(?P<string1>[^ )]+)\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r'^\s*CONTAINS\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r'^\s*CONTAINS\s*\(\s*(?P<string1>[^ )]+)\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*(?P<string1>[^ )]+)\s*,\s*'(?P<string2>[^']+)'\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*(?P<string1>[^ )]+)\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*'(?P<string1>[^']+)'\s*,\s*(?P<string2>[^ )]+)\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_contains)
conditionallist.add(r'^\s*CONTAINS\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*'(?P<string1>[^']+)'\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*'(?P<string1>[^']+)'\s*,\s*\"(?P<string2>[^\"]+)\"(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*\"(?P<string1>[^\"]+)\"\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r'^\s*CONTAINS\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r'^\s*CONTAINS\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)
conditionallist.add(r"^\s*CONTAINS\s*\(\s*'(?P<string1>[^']+)'\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_contains)


def xf_startswith(**kwargs):
	s1 = kwargs["string1"]
	s2 = kwargs["string2"]
	if kwargs["ignorecase"] and kwargs["ignorecase"].lower() == "i":
		s1 = s1.lower()
		s2 = s2.lower()
	return s1[:len(s2)] == s2

conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*'(?P<string2>[^']+)'\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*(?P<string2>[^ )]+)\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_startswith)
conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*\"(?P<string2>[^\"]+)\"(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*\"(?P<string1>[^\"]+)\"\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r'^\s*STARTS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)
conditionallist.add(r"^\s*STARTS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_startswith)



def xf_endswith(**kwargs):
	s1 = kwargs["string1"]
	s2 = kwargs["string2"]
	if kwargs["ignorecase"] and kwargs["ignorecase"].lower() == "i":
		s1 = s1.lower()
		s2 = s2.lower()
	return s1[-len(s2):] == s2

conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*'(?P<string2>[^']+)'\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*(?P<string1>[^ )]+)\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*(?P<string2>[^ )]+)(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*(?P<string2>[^ )]+)\s*(?:\s*,\s*(?P<ignorecase>I))?\)", xf_endswith)
conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*\"(?P<string2>[^\"]+)\"(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*\"(?P<string1>[^\"]+)\"\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*"(?P<string2>[^"]+)"(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r'^\s*ENDS_WITH\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)', xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*`(?P<string1>[^`]+)`\s*,\s*'(?P<string2>[^']+)'(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)
conditionallist.add(r"^\s*ENDS_WITH\s*\(\s*'(?P<string1>[^']+)'\s*,\s*`(?P<string2>[^`]+)`(?:\s*,\s*(?P<ignorecase>I))?\s*\)", xf_endswith)


def xf_hasrows(**kwargs):
	queryname = kwargs["queryname"]
	sql = u"select count(*) from %s;" % queryname
	# Exceptions should be trapped by the caller, so are re-raised here after settting status
	try:
		hdrs, rec = dbs.current().select_data(sql)
	except ErrInfo:
		raise
	except:
		raise ErrInfo("db", sql, exception_msg=exception_desc())
	nrows = rec[0][0]
	return nrows > 0

conditionallist.add(r'^\s*HASROWS\((?P<queryname>[^)]+)\)', xf_hasrows)
conditionallist.add(r'^\s*HAS_ROWS\((?P<queryname>[^)]+)\)', xf_hasrows)

def xf_sqlerror(**kwargs):
	return status.sql_error

conditionallist.add(r'^\s*sql_error\(\s*\)', xf_sqlerror)

def xf_fileexists(**kwargs):
	filename = kwargs["filename"]
	return os.path.isfile(filename.strip())

conditionallist.add(ins_fn_rxs(r'^FILE_EXISTS\(\s*', r'\)'), xf_fileexists)

def xf_direxists(**kwargs):
	dirname = kwargs["dirname"]
	return os.path.isdir(dirname.strip())

conditionallist.add(r'^DIRECTORY_EXISTS\(\s*("?)(?P<dirname>[^")]+)\1\)', xf_direxists)

def xf_schemaexists(**kwargs):
	schemaname = kwargs["schema"]
	return dbs.current().schema_exists(schemaname)

conditionallist.add((
	r'^SCHEMA_EXISTS\(\s*(?P<schema>[A-Za-z0-9_\-\: ]+)\s*\)',
	r'^SCHEMA_EXISTS\(\s*"(?P<schema>[A-Za-z0-9_\-\: ]+)"\s*\)'
	), xf_schemaexists)


def xf_tableexists(**kwargs):
	schemaname = kwargs["schema"]
	tablename = kwargs["tablename"]
	return  dbs.current().table_exists(tablename.strip(), schemaname)

conditionallist.add((
	r'^TABLE_EXISTS\(\s*(?:(?P<schema>[A-Za-z0-9_\-\/\: ]+)\.)?(?P<tablename>[A-Za-z0-9_\-\/\: ]+)\)',
	r'^TABLE_EXISTS\(\s*(?:\[(?P<schema>[A-Za-z0-9_\-\/\: ]+)\]\.)?\[(?P<tablename>[A-Za-z0-9_\-\/\: ]+)\]\)',
	r'^TABLE_EXISTS\(\s*(?:"(?P<schema>[A-Za-z0-9_\-\/\: ]+)"\.)?"(?P<tablename>[A-Za-z0-9_\-\/\: ]+)"\)',
	r'^TABLE_EXISTS\(\s*(?:(?P<schema>[A-Za-z0-9_\-\/]+)\.)?(?P<tablename>[A-Za-z0-9_\-\/]+)\)'
	), xf_tableexists)

def xf_roleexists(**kwargs):
	rolename = kwargs["role"]
	return dbs.current().role_exists(rolename)

conditionallist.add((
	r'^ROLE_EXISTS\(\s*(?P<role>[A-Za-z0-9_\-\:\$ ]+)\s*\)',
	r'^ROLE_EXISTS\(\s*"(?P<role>[A-Za-z0-9_\-\:\$ ]+)"\s*\)'
	), xf_roleexists)


def xf_sub_defined(**kwargs):
	varname = kwargs["match_str"]
	subvarset = subvars if varname[0] not in ('~','#') else commandliststack[-1].localvars if varname[0] == '~' else commandliststack[-1].paramvals
	return subvarset.sub_exists(varname) if subvarset else False

conditionallist.add(r'^SUB_DEFINED\s*\(\s*(?P<match_str>[\$&@~#]?\w+)\s*\)', xf_sub_defined)


def xf_sub_empty(**kwargs):
	varname = kwargs["match_str"]
	subvarset = subvars if varname[0] not in ('~','#') else commandliststack[-1].localvars if varname[0] == '~' else commandliststack[-1].paramvals
	if not subvarset.sub_exists(varname):
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg=u"Unrecognized substitution variable name: %s" % varname)
	return subvarset.varvalue(varname) == u''

conditionallist.add(r'^SUB_EMPTY\s*\(\s*(?P<match_str>[\$&@~#]?\w+)\s*\)', xf_sub_empty)

def xf_script_exists(**kwargs):
	script_id = kwargs["script_id"].lower()
	return script_id in savedscripts

conditionallist.add(r'^\s*SCRIPT_EXISTS\s*\(\s*(?P<script_id>\w+)\s*\)', xf_script_exists)


def xf_equals(**kwargs):
	import unicodedata
	s1 = unicodedata.normalize('NFC', kwargs["string1"]).lower().strip('"')
	s2 = unicodedata.normalize('NFC', kwargs["string2"]).lower().strip('"')
	converters = (int, float, parse_datetime, parse_datetimetz, parse_date, parse_boolean)
	converted = False
	for convf in converters:
		try:
			v1 = convf(s1)
			v2 = convf(s2)
			if not(v1 is None and v2 is None):
				converted = True
				break
		except:
			continue
	if converted:
		return v1 == v2
	else:
		return s1 == s2

conditionallist.add(r'^\s*EQUAL(S)?\s*\(\s*(?P<string1>[^ )]+)\s*,\s*(?P<string2>[^ )]+)\s*\)', xf_equals)
conditionallist.add(r'^\s*EQUAL(S)?\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*(?P<string2>[^ )]+)\s*\)', xf_equals)
conditionallist.add(r'^\s*EQUAL(S)?\s*\(\s*(?P<string1>[^ )]+)\s*,\s*"(?P<string2>[^"]+)"\s*\)', xf_equals)
conditionallist.add(r'^\s*EQUAL(S)?\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*"(?P<string2>[^"]+)"\s*\)', xf_equals)

def xf_identical(**kwargs):
	s1 = kwargs["string1"].strip('"')
	s2 = kwargs["string2"].strip('"')
	return s1 == s2

conditionallist.add(r'^\s*IDENTICAL\s*\(\s*(?P<string1>[^ ,)]+)\s*,\s*(?P<string2>[^ )]+)\s*\)', xf_identical)
conditionallist.add(r'^\s*IDENTICAL\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*(?P<string2>[^ )]+)\s*\)', xf_identical)
conditionallist.add(r'^\s*IDENTICAL\s*\(\s*(?P<string1>[^ ,]+)\s*,\s*"(?P<string2>[^"]+)"\s*\)', xf_identical)
conditionallist.add(r'^\s*IDENTICAL\s*\(\s*"(?P<string1>[^"]+)"\s*,\s*"(?P<string2>[^"]+)"\s*\)', xf_identical)

def xf_isnull(**kwargs):
	item = kwargs["item"].strip().strip(u'"')
	return item == u""

conditionallist.add(r'^\s*IS_NULL\(\s*(?P<item>"[^"]*")\s*\)', xf_isnull)

def xf_iszero(**kwargs):
	val = kwargs["value"].strip()
	try:
		v = float(val)
	except:
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg="The value {%s} is not numeric." % val)
	return v == 0

conditionallist.add(r'^\s*IS_ZERO\(\s*(?P<value>[^)]*)\s*\)', xf_iszero)

def xf_isgt(**kwargs):
	val1 = kwargs["value1"].strip()
	val2 = kwargs["value2"].strip()
	try:
		v1 = float(val1)
		v2 = float(val2)
	except:
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg="Values {%s} and {%s} are not both numeric." % (val1, val2))
	return v1 > v2

conditionallist.add(r'^\s*IS_GT\(\s*(?P<value1>[^)]*)\s*,\s*(?P<value2>[^)]*)\s*\)', xf_isgt)


def xf_isgte(**kwargs):
	val1 = kwargs["value1"].strip()
	val2 = kwargs["value2"].strip()
	try:
		v1 = float(val1)
		v2 = float(val2)
	except:
		raise ErrInfo(type="cmd", command_text=kwargs["metacommandline"], other_msg="Values {%s} and {%s} are not both numeric." % (val1, val2))
	return v1 >= v2

conditionallist.add(r'^\s*IS_GTE\(\s*(?P<value1>[^)]*)\s*,\s*(?P<value2>[^)]*)\s*\)', xf_isgte)


def xf_boolliteral(**kwargs):
	return unquoted(kwargs["value"].strip()).lower() in ('true', 'yes', '1')

conditionallist.add((
	r'^\s*(?P<value>1)\s*',
	r'^\s*(?P<value>"1")\s*',
	r'^\s*(?P<value>0)\s*',
	r'^\s*(?P<value>"0")\s*',
	r'^\s*(?P<value>Yes)\s*',
	r'^\s*(?P<value>"Yes")\s*',
	r'^\s*(?P<value>No)\s*',
	r'^\s*(?P<value>"No")\s*',
	r'^\s*(?P<value>"False")\s*',
	r'^\s*(?P<value>False)\s*',
	r'^\s*(?P<value>"True")\s*',
	r'^\s*(?P<value>True)\s*'
	), xf_boolliteral)


def xf_istrue(**kwargs):
	return unquoted(kwargs["value"].strip()).lower() in ('yes', 'y', 'true', 't', '1')

conditionallist.add(r'^\s*IS_TRUE\(\s*(?P<value>[^)]*)\s*\)', xf_istrue)

def xf_dbms(**kwargs):
	dbms = kwargs["dbms"]
	return dbs.current().type.dbms_id.lower() == dbms.strip().lower()

conditionallist.add((
	r'^\s*DBMS\(\s*(?P<dbms>[A-Z0-9_\-\(\/\\\. ]+)\s*\)',
	r'^\s*DBMS\(\s*"(?P<dbms>[A-Z0-9_\-\(\)\/\\\. ]+)"\s*\)'
	), xf_dbms)


def xf_dbname(**kwargs):
	dbname = kwargs["dbname"]
	return dbs.current().name().lower() == dbname.strip().lower()

                           
conditionallist.add((
	r'^\s*DATABASE_NAME\(\s*(?P<dbname>[A-Z0-9_;\-\(\/\\\. ]+)\s*\)', 
	r'^\s*DATABASE_NAME\(\s*"(?P<dbname>[A-Z0-9_;\-\(\)\/\\\. ]+)"\s*\)'), xf_dbname)

def xf_viewexists(**kwargs):
	viewname = kwargs["viewname"]
	return dbs.current().view_exists(viewname.strip())

conditionallist.add(r'^\s*VIEW_EXISTS\(\s*("?)(?P<viewname>[^")]+)\1\)', xf_viewexists)


def xf_columnexists(**kwargs):
	tablename = kwargs["tablename"]
	schemaname = kwargs["schema"]
	columnname = kwargs["columnname"]
	return dbs.current().column_exists(tablename.strip(), columnname.strip(), schemaname)

conditionallist.add((
	r'^COLUMN_EXISTS\(\s*(?P<columnname>[A-Za-z0-9_\-\:]+)\s+IN\s+(?:(?P<schema>[A-Za-z0-9_\-\: ]+)\.)?(?P<tablename>[A-Za-z0-9_\-\: ]+)\)',
	r'^COLUMN_EXISTS\(\s*(?P<columnname>[A-Za-z0-9_\-\:]+)\s+IN\s+(?:\[(?P<schema>[A-Za-z0-9_\-\: ]+)\]\.)?\[(?P<tablename>[A-Za-z0-9_\-\: ]+)\]\)', 
	r'^COLUMN_EXISTS\(\s*(?P<columnname>[A-Za-z0-9_\-\:]+)\s+IN\s+(?:"(?P<schema>[A-Za-z0-9_\-\: ]+)"\.)?"(?P<tablename>[A-Za-z0-9_\-\: ]+)"\)',
	r'^COLUMN_EXISTS\(\s*"(?P<columnname>[A-Za-z0-9_\-\: ]+)"\s+IN\s+(?:(?P<schema>[A-Za-z0-9_\-\: ]+)\.)?(?P<tablename>[A-Za-z0-9_\-\: ]+)\)',
	r'^COLUMN_EXISTS\(\s*"(?P<columnname>[A-Za-z0-9_\-\: ]+)"\s+IN\s+(?:\[(?P<schema>[A-Za-z0-9_\-\: ]+)\]\.)?\[(?P<tablename>[A-Za-z0-9_\-\: ]+)\]\)',
	r'^COLUMN_EXISTS\(\s*"(?P<columnname>[A-Za-z0-9_\-\: ]+)"\s+IN\s+(?:"(?P<schema>[A-Za-z0-9_\-\: ]+)"\.)?"(?P<tablename>[A-Za-z0-9_\-\: ]+)"\)'
	), xf_columnexists)

def xf_aliasdefined(**kwargs):
	alias = kwargs["alias"]
	return alias in dbs.aliases()

conditionallist.add(r'^\s*ALIAS_DEFINED\s*\(\s*(?P<alias>\w+)\s*\)', xf_aliasdefined)


def xf_metacommanderror(**kwargs):
	return status.metacommand_error

conditionallist.add(r'^\s*metacommand_error\(\s*\)', xf_metacommanderror)


def xcmd_test(teststr):
	result = CondParser(teststr).parse().eval()
	if result is not None:
		return result
	else:
		raise ErrInfo(type="cmd", command_text=teststr, other_msg="Unrecognized conditional")


#	End of conditional tests for metacommands.
#===============================================================================================




#===============================================================================================
#-----  SUPPORT FUNCTIONS (2)

def set_system_vars():
	# (Re)define the system substitution variables that are not script-specific.
	global subvars
	subvars.add_substitution("$ERROR_HALT_STATE", "ON" if status.halt_on_err else "OFF")
	subvars.add_substitution("$METACOMMAND_ERROR_HALT_STATE", "ON" if status.halt_on_metacommand_err else "OFF")
	subvars.add_substitution("$CURRENT_TIME", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
	subvars.add_substitution("$CURRENT_DIR", os.path.abspath(os.path.curdir))
	subvars.add_substitution("$CURRENT_PATH", os.path.abspath(os.path.curdir) + os.sep)
	subvars.add_substitution("$AUTOCOMMIT_STATE", "ON" if dbs.current().autocommit else "OFF")
	subvars.add_substitution("$RANDOM", str(random.random()))
	subvars.add_substitution("$UUID", str(uuid.uuid4()))


def substitute_vars(command_str, localvars = None):
	# Substitutes global variables, global counters, and local variables
	# into the command string until no more substitutions can be made.
	# Returns the modified command_str.
	global subvars
	global counters
	if localvars is not None:
		subs = subvars.merge(localvars)
	else:
		subs = subvars
	cmdstr = copy.copy(command_str)
	# Substitute variables and counters until no more substitutions are made.
	subs_made = True
	while subs_made:
		subs_made = False
		cmdstr, subs_made = subs.substitute_all(cmdstr)
		cmdstr, any_subbed = counters.substitute_all(cmdstr)
		subs_made = subs_made or any_subbed
	m = defer_rx.findall(cmdstr)
    # Substitute any deferred substitution variables with regular substition var flags, e.g.: "!!somevar!!"
	if m is not None:
		for dv in m:
			rep = "!!" +  dv[1] + "!!"
			cmdstr = cmdstr.replace(dv[0], rep)
	return cmdstr


def current_script_line():
	if len(commandliststack) > 0:
		current_cmds = commandliststack[-1]
		if current_cmds.current_command() is not None:
			return current_cmds.current_command().current_script_line()
		else:
			return len(current_cmds.cmdlist)
	else:
		return 0


def wo_quotes(argstr):
	# Strip first and last quotes off an argument.
	argstr = argstr.strip()
	if argstr[0]=='"' and argstr[-1]=='"' or argstr[0]=="'" and argstr[-1]=="'" or argstr[0]=="[" and argstr[-1]=="]":
		return argstr[1:-1]
	return argstr


def get_subvarset(varname, metacommandline):
	# Supports the exec functions for the substitution metacommands that allow
	# substitution variables with a "+" prefix, to reference outer scope local
	# variables
	subvarset = None
	# Outer scope variable
	if varname[0] == '+':
		varname = re.sub('^[+]', '~', varname)
		for cl in reversed(commandliststack[0:-1]):
			if cl.localvars.sub_exists(varname):
				subvarset = cl.localvars
				break
		# Raise error if local variable not found anywhere down in commandliststack
		if not subvarset:
			raise ErrInfo(type="cmd", command_text=metacommandline, other_msg="Outer-scope referent variable (%s) has no matching local variable (%s)." % (re.sub('^[~]', '+', varname), varname)) # Global or local variable else:
	# Global or local variable
	else:
		subvarset = subvars if varname[0] != '~' else commandliststack[-1].localvars
	return subvarset, varname

# End of support functions (2).
#===============================================================================================



#===============================================================================================
#-----  GLOBAL OBJECT INITIALIZATION FOR EXECSQL INTERPRETER

# Status object with status-related attributes.
status = StatObj()

# Stack of conditional levels to support IF metacommands.
if_stack = IfLevels()

# Global counter variables.
counters = CounterVars()

# Global substitution variables.  (There may also be SCRIPT-specific
# substitution variables used as parameters.)
subvars = SubVarSet()
for k in os.environ.keys():
	try:
		subvars.add_substitution(u"&"+k, os.environ[k])
	except:
		# Ignore "ProgramFiles(x86)" on Windows and any others with invalid characters.
		pass
subvars.add_substitution("$LAST_ROWCOUNT", None)

dt_now = datetime.datetime.now()
subvars.add_substitution("$SCRIPT_START_TIME", dt_now.strftime("%Y-%m-%d %H:%M"))
subvars.add_substitution("$DATE_TAG", dt_now.strftime("%Y%m%d"))
subvars.add_substitution("$DATETIME_TAG", dt_now.strftime("%Y%m%d_%H%M"))
subvars.add_substitution("$LAST_SQL", "")
subvars.add_substitution("$LAST_ERROR", "")
subvars.add_substitution("$ERROR_MESSAGE", "")
subvars.add_substitution("$PATHSEP", os.sep)
osys = sys.platform
if osys.startswith('linux'):
	osys = 'linux'
elif osys.startswith('win'):
	osys = 'windows'
subvars.add_substitution("$OS", osys)

conf = ConfigData()

# Storage for all the (named) databases that are opened.  Databases are added in 'main()'
# and by the CONNECT metacommand.
dbs = DatabasePool()

#	End of global object initialization for execsql interpreter.
#===============================================================================================



def process_sql(sql_commands):
	# Read lines from the list of SQL commands, create a list of ScriptCmd objects,
	# and append the list to the top of the stack of script commands.
	# The filename (fn) and line number are stored with each command.
	# Arguments:
	#    sql_file_name:  The name of the execql script to read and store.
	# Return value:
	#    No return value.
	# Side effects:
	#    1. The script that is read is appended to the global 'commandliststack'.
	#    2. Items may be added to the global 'savedscripts' if there are any
	#       BEGIN/END SCRIPT commands in the file.
	#
	# Lines containing execsql command statements must begin with "-- !x!"
	# Currently this routine knows only three things about SQL:
	#	1. Lines that start with "--" are comments.
	#	2. Lines that end with ";" terminate a SQL statement.'
	#	3. Lines that start with "/*" begin a block comment, and lines that
	#		end with "*/" end a block comment.
	# The following metacommands are executed IMMEDIATELY during this process:
	#	* BEGIN SCRIPT <scriptname>
	#	* END SCRIPT
	#	* BEGIN SQL
	#	* END SQL
	#
	# May update the script_errors global list.
	#
	# Returns True if there are no errors or only warnings.  Returns False
	# if there are any fatal errors.
	beginscript = re.compile(r'^\s*--\s*!x!\s*(?:BEGIN|CREATE)\s+SCRIPT\s+(?P<scriptname>\w+)(?:(?P<paramexpr>\s*\S+.*))?$', re.I)
	endscript = re.compile(r'^\s*--\s*!x!\s*END\s+SCRIPT(?:\s+(?P<scriptname>\w+))?\s*$', re.I)
	beginsql = re.compile(r'^\s*--\s*!x!\s*BEGIN\s+SQL\s*$', re.I)
	endsql = re.compile(r'^\s*--\s*!x!\s*END\s+SQL\s*$', re.I)
	execline = re.compile(r'^\s*--\s*!x!\s*(?P<cmd>.+)$', re.I)
	cmtline = re.compile(r'^\s*--')
	in_block_cmt = False
	in_block_sql = False
	sqllist = []
	sqlline = 0
	subscript_stack = []
	sql_file_name = ""
	scriptfilename = ""
	file_lineno = 0
	currcmd = ''
	for line in sql_commands:
		file_lineno += 1
		# Remove trailing whitespace but not leading whitespace; this may be a plpythonu command in Postgres, where leading whitespace is significant.
		line = line.rstrip()
		is_comment_line = False
		comment_match = cmtline.match(line)
		metacommand_match = execline.match(line)
		if len(line) > 0:
			if in_block_cmt:
				is_comment_line = True
				if len(line) > 1 and line[-2:] == u"*/":
					in_block_cmt = False
			else:
				# Not in block comment
				if len(line.strip()) > 1 and line.strip()[0:2] == u"/*":
					in_block_cmt = True
					is_comment_line = True
					if line.strip()[-2:] == u"*/":
						in_block_cmt = False
				else:
					if comment_match:
						is_comment_line = not metacommand_match
			if not is_comment_line:
				if metacommand_match:
					if beginsql.match(line):
						in_block_sql = True
					if in_block_sql:
						if endsql.match(line):
							in_block_sql = False
							if len(currcmd) > 0:
								cmd = ScriptCmd(sql_file_name, sqlline, 'sql', SqlStmt(currcmd))
								if len(subscript_stack) == 0:
									sqllist.append(cmd)
								else:
									subscript_stack[-1].add(cmd)
								currcmd = ''
					else:
						if len(currcmd) > 0:
							script_errors.append(["Incomplete SQL statement", sqlline])
						begs = beginscript.match(line)
						if not begs:
							ends = endscript.match(line)
						if begs:
							# This is a BEGIN SCRIPT metacommand.
							scriptname = begs.group('scriptname').lower()
							paramnames = None
							paramexpr = begs.group('paramexpr')
							if paramexpr:
								withparams = re.compile(r'(?:\s+WITH)?(?:\s+PARAM(?:ETER)?S)?\s*\(\s*(?P<params>\w+(?:\s*,\s*\w+)*)\s*\)\s*$', re.I)
								wp = withparams.match(paramexpr)
								if not wp:
									raise ErrInfo(type="cmd", command_text=line, other_msg="Invalid BEGIN SCRIPT metacommand on line %s of file %s." % (file_lineno, sql_file_name))
								else:
									param_rx = re.compile(r'\w+', re.I)
									paramnames = re.findall(param_rx, wp.group('params'))
							# If there are no parameter names to pass, paramnames will be None
							subscript_stack.append(CommandList([], scriptname, paramnames))
						elif ends:
							# This is an END SCRIPT metacommand.
							endscriptname = ends.group('scriptname')
							if endscriptname is not None:
								endscriptname = endscriptname.lower()
							if len(subscript_stack) == 0:
								raise ErrInfo(type="cmd", command_text=line, other_msg="Unmatched END SCRIPT metacommand on line %s of file %s." % (file_lineno, sql_file_name))
							if len(currcmd) > 0:
								raise ErrInfo(type="cmd", command_text=line, other_msg="Incomplete SQL statement\n  (%s)\nat END SCRIPT metacommand on line %s of file %s." % (currcmd, file_lineno, sql_file_name))
							if endscriptname is not None and endscriptname != scriptname:
								raise ErrInfo(type="cmd", command_text=line, other_msg="Mismatched script name in the END SCRIPT metacommand on line %s of file %s." % (file_lineno, sql_file_name))
							sub_script = subscript_stack.pop()
							savedscripts[sub_script.listname] = sub_script
						else:
							# This is a non-IMMEDIATE metacommand.
							cmd = ScriptCmd(sql_file_name, file_lineno, 'cmd', MetacommandStmt(metacommand_match.group('cmd').strip()))
							if len(subscript_stack) == 0:
								sqllist.append(cmd)
							else:
								subscript_stack[-1].add(cmd)
				else:
					# This line is not a comment and not a metacommand, therefore should be
					# part of a SQL statement.
					cmd_end = True if line[-1] == ';' else False
					if line[-1] == '\\':
						line = line[:-1].strip()
					if currcmd == '':
						sqlline = file_lineno
						currcmd = line
					else:
						currcmd = u"%s \n%s" % (currcmd, line)
					if cmd_end and not in_block_sql:
						cmd = ScriptCmd(sql_file_name, sqlline, 'sql', SqlStmt(currcmd.strip()))
						if len(subscript_stack) == 0:
							sqllist.append(cmd)
						else:
							subscript_stack[-1].add(cmd)
						currcmd = ''
	if len(subscript_stack) > 0:
		raise ErrInfo(type="error", other_msg="Unmatched BEGIN SCRIPT metacommand at end of file %s." % sql_file_name)
	if len(currcmd) > 0:
		raise ErrInfo(type="error", other_msg="Incomplete SQL statement starting on line %s at end of file %s." % (sqlline, sql_file_name))
	if len(sqllist) > 0:
		# The file might be all comments or just a BEGIN/END SCRIPT metacommand.
		commandliststack.append(CommandList(sqllist, scriptfilename))

def read_sqlfile(sql_file_name):
	with io.open(sql_file_name, "r") as f:
		script_cmds = f.read()
	process_sql(script_cmds.splitlines())


def runscripts(db, sql_text):
	# Repeatedly run the next statement from the script at the top of the
	# command list stack until there are no more statements.
	# Metacommands may modify the stack or the commands in a stack entry.
	# This is the central script processing routine for execsql script extensions.
	global commandliststack
	global cmds_run
	global script_errors
	dbs.add('mapdata_connection', db)
	subvars.add_substitution("$CURRENT_DBMS", db.type.dbms_id)
	subvars.add_substitution("$CURRENT_DATABASE", db.name())
	process_sql(sql_text.splitlines())
	if len(script_errors) > 0:
		dlg = MsgDialog2("Warnings", "The following unexpected conditions were encountered while parsing the SQL script.", can_resize=True)
		hdrs = ["Warning message", "Line no"]
		tframe, tdata = treeview_table(dlg.content_frame, script_errors, hdrs)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		dlg.show()
	script_errors = []
	while len(commandliststack) > 0:
		current_cmds = commandliststack[-1]
		set_system_vars()
		try:
			current_cmds.run_next()
		except StopIteration:
			commandliststack.pop()
		except ErrInfo:
			commandliststack.pop()
			raise
		except:
			commandliststack.pop()
			raise ErrInfo(type="exception", exception_msg=exception_desc())
		cmds_run += 1
	if len(script_errors) > 0:
		dlg = MsgDialog2("Warnings", "The following unexpected conditions were encountered while running the SQL script.", can_resize=True)
		hdrs = ["Warning message", "Line no"]
		tframe, tdata = treeview_table(dlg.content_frame, script_errors, hdrs)
		tframe.grid(row=0, column=0, sticky=tk.NSEW)
		dlg.show()


def runscriptfile(db, sql_file):
	with open(sql_file, 'r') as scriptfile:
		script_text = scriptfile.read()
	runscripts(db, script_text)


#***************************************************************************************************
#***************************  End of SQL Scripting Extensions  *************************************
#***************************************************************************************************


def open_dbms(dbms_type, server, db_name, user, need_pw=True, pw=None, port=None, db_file=None):
	# Returns a db connection object or None if the connection can't be made.
	db = None
	if dbms_type == 'PostgreSQL':
		port = 5432 if port is None else port
		try:
			db = PostgresDatabase(server, db_name, user, need_pw, port, password=pw)
		except ErrInfo as e:
			warning(e.eval_err(), kwargs={})
			db = None
		except:
			warning("Cannot open the Postgres database.", kwargs={})
			db = None

	elif dbms_type == 'SQLite':
		try:
			db = SQLiteDatabase(db_file)
		except ErrInfo as e:
			warning(e.eval_err(), kwargs={})
			db = None
		except:
			warning("Cannot open the file %s as a SQLite database." % db_file, kwargs={})
			db = None

	elif dbms_type == 'DuckDB':
		try:
			db = DuckDBDatabase(db_file)
		except ErrInfo as e:
			warning(e.eval_err(), kwargs={})
			db = None
		except:
			warning("Cannot open the file %s as a DuckDB database." % db_file, kwargs={})
			db = None

	elif dbms_type == 'MariaDB' or dbms_type == 'MySQL':
		port = 3306 if port is None else port
		try:
			db = MySQLDatabase(server, db_name, user, need_pw, port, password=pw)
		except ErrInfo as e:
			warning(e.eval_err(), kwargs={})
			db = None
		except:
			warning("Cannot open the MariaDB/MySQL database.", kwargs={})
			db = None

	elif dbms_type == 'SQL Server':
		port = 1433 if port is None else port
		try:
			db = SqlServerDatabase(server, db_name, user, need_pw, port, password=pw)
		except ErrInfo as e:
			warning(e.eval_err(), kwargs={})
			db = None
		except:
			warning("Cannot open the SQL Server database.", kwargs={'parent': self.dlg})
			db = None

	elif dbms_type == 'Oracle':
		port = 1521 if port is None else port
		try:
			db = SqlServerDatabase(server, db_name, user, need_pw, port, password=pw)
		except ErrInfo as e:
			warning(e.eval_err(), kwargs={})
			db = None
		except:
			warning("Cannot open the Oracle database.", kwargs={})
			db = None

	elif dbms_type == 'Firebird':
		port = 3050 if port is None else port
		try:
			db = FirebirdDatabase(server, db_name, user, need_pw, port, password=pw)
		except ErrInfo as e:
			warning(e.eval_err(), kwargs={})
			db = None
		except:
			warning("Cannot open the Firebird database.", kwargs={})
			db = None
	return db



class DbConnectDialog(object):
	FILE, SERVER, FILE_PW = range(3)
	def __init__(self, parent, mapui):
		self.parent = parent
		self.mapui = mapui
		self.exit_status = 0	# Canceled
		self.exit_svr = None	# For caller
		self.exit_db = None	# For caller
		self.xpos = None
		self.ypos = None
		self.scriptfilepath = None
		# Values of db_params indicate whether server information is needed.
		self.db_params = {u"PostgreSQL": self.SERVER, u"SQLite": self.FILE, u"DuckDB": self.FILE,
							u"SQL Server": self.SERVER, u"MySQL": self.SERVER, u"Firebird": self.SERVER,
							u"MariaDB": self.SERVER, u"Oracle": self.SERVER}
		self.dlg = tk.Toplevel(parent)
		self.title = "Database Table for Map Display"
		self.dlg.title(self.title)
		self.dlg.protocol("WM_DELETE_WINDOW", self.do_cancel)
		self.headers = None
		self.header_list = None
		self.datarows = None

		# Main frames
		msgframe = tk.Frame(self.dlg)
		msgframe.grid(column=0, row=0, padx=6, pady=2, sticky=tk.EW)
		# Database selection is in one wizard pane, table and script selection are in a second, and column selection is in a third wizard pane
		wiz1_frame = tk.Frame(self.dlg)
		wiz1_frame.grid(column=0, row=1, sticky=tk.NSEW)
		wiz2_frame = tk.Frame(self.dlg)
		wiz2_frame.grid(column=0, row=1, sticky=tk.NSEW)
		wiz3_frame = tk.Frame(self.dlg)
		wiz3_frame.grid(column=0, row=1, sticky=tk.NSEW)
		self.dlg.rowconfigure(0, weight=0)
		self.dlg.rowconfigure(1, weight=1)
		self.dlg.columnconfigure(0, weight=1)
		wiz1_frame.rowconfigure(0, weight=1)
		wiz1_frame.columnconfigure(0, weight=1)
		wiz2_frame.rowconfigure(0, weight=1)
		wiz2_frame.columnconfigure(0, weight=1)
		wiz3_frame.rowconfigure(0, weight=1)
		wiz3_frame.columnconfigure(0, weight=1)

		# Populate message frame
		msg_label = ttk.Label(msgframe, text="The database, table, and columns to be used for mapping must be specified.", anchor=tk.W, justify=tk.LEFT, wraplength=500)
		msg_label.grid(column=0, row=0, sticky=tk.EW)


		# Wizard page 1
		# Database selector
		# On the left side will be a combobox to choose the database type.
		# On the right side will be a prompt for the server, db, user name, and pw,
		# or for the filename (and possibly user name and pw).  Each of these alternative
		# types of prompts will be in its own frame, which will be in the same place.
		# Only one will be shown, controlled by the item in the self.db_params dictionary.
		# A separate frame for the table name is below the database parameters frame.
		dbframe = tk.Frame(wiz1_frame)
		dbtypeframe = tk.Frame(dbframe)
		rightframe = tk.Frame(dbframe)
		paramframe = tk.Frame(rightframe)
		self.serverparamframe = tk.Frame(paramframe)
		self.fileparamframe = tk.Frame(paramframe)
		self.filepwparamframe = tk.Frame(paramframe)
		w1btnframe = tk.Frame(wiz1_frame, borderwidth=3, relief=tk.RIDGE)

		# Grid wiz1 frame widgets
		def param_choices(*args, **kwargs):
			svr_params = self.db_params[self.db_type_var.get()]
			if svr_params == self.SERVER:
				self.fileparamframe.grid_remove()
				self.filepwparamframe.grid_remove()
				self.serverparamframe.grid()
			elif svr_params == self.FILE_PW:
				self.serverparamframe.grid_remove()
				self.fileparamframe.grid_remove()
				self.filepwparamframe.grid()
			else:
				self.serverparamframe.grid_remove()
				self.filepwparamframe.grid_remove()
				self.fileparamframe.grid()
			check_w1enable()

		def check_w1enable(*args):
			dbms = self.db_type_var.get()
			if dbms != '':
				dbtype = self.db_params[dbms]
				if dbtype == self.SERVER:
					if self.server.get() != '' and self.db.get != '':
						w1next_btn["state"] = tk.NORMAL
					else:
						w1next_btn["state"] = tk.DISABLED
				elif dbtype == self.FILE_PW:
					if self.db_file.get() != '' and self.user.get() != '' and self.pw.get() != '':
						w1next_btn["state"] = tk.NORMAL
					else:
						w1next_btn["state"] = tk.DISABLED
				else:	# self.FILE
					if self.db_file.get() != '':
						w1next_btn["state"] = tk.NORMAL
					else:
						w1next_btn["state"] = tk.DISABLED

		dbframe.grid(column=0, row=0, sticky=tk.NSEW)
		dbtypeframe.grid(column=0, row=0, padx=5, sticky=tk.NW)
		rightframe.grid(column=1, row=0, padx=5, sticky=tk.N + tk.EW)
		paramframe.grid(column=0, row=0, padx=5, sticky=tk.N + tk.EW)
		# Put serverparamframe, fileparamframe, and filepwparamframe in the same place in paramframe.
		# Leave only serverparamframe visible.
		self.fileparamframe.grid(row=0, column=0, sticky=tk.N + tk.EW)
		self.fileparamframe.grid_remove()
		self.filepwparamframe.grid(row=0, column=0, sticky=tk.N + tk.EW)
		self.filepwparamframe.grid_remove()
		self.serverparamframe.grid(row=0, column=0, sticky=tk.N + tk.EW)
		w1btnframe.grid(column=0, row=2, sticky=tk.S+tk.EW)
		w1btnframe.columnconfigure(0, weight=1)

		# Populate dbframe
		self.db_type_var = tk.StringVar()
		self.encoding = tk.StringVar()
		self.table_var = tk.StringVar()
		self.table_var.trace('w', check_w1enable)
		# Database type selection
		ttk.Label(dbtypeframe, text="DBMS:").grid(column=0, row=0, padx=3, pady=3, sticky=tk.NE)
		dbmss = [k for k in self.db_params.keys()]
		dbmss.sort()
		self.db_choices = ttk.Combobox(dbtypeframe, textvariable=self.db_type_var, width=12,
						values=dbmss)
		self.db_choices.bind("<<ComboboxSelected>>", param_choices)
		self.db_choices.config(state='readonly')
		self.db_choices.grid(column=1, row=0, padx=3, pady=3, sticky=tk.NW)
		self.db_choices.focus()
		ttk.Label(dbtypeframe, text="Encoding:").grid(column=0, row=1, padx=3, pady=3, sticky=tk.NE)
		self.db_choices.set('PostgreSQL')
		enc_choices = ttk.Combobox(dbtypeframe, textvariable=self.encoding, width=12,
						values=('UTF8', 'Latin1', 'Win1252'))
		enc_choices.set('UTF8')
		enc_choices.grid(column=1, row=1, padx=3, pady=3, sticky=tk.NW)
		# Database parameter entry frames
		self.server = tk.StringVar()
		self.server.trace('w', check_w1enable)
		self.port = tk.StringVar()
		self.db = tk.StringVar()
		self.db.trace('w', check_w1enable)
		self.user = tk.StringVar()
		self.user.trace('w', check_w1enable)
		self.pw = tk.StringVar()
		self.pw.trace('w', check_w1enable)
		self.db_file = tk.StringVar()
		self.db_file.trace('w', check_w1enable)

		# Server databases
		ttk.Label(self.serverparamframe, text="Server:").grid(column=0, row=0, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.server).grid(column=1, row=0, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="Database:").grid(column=0, row=1, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.db).grid(column=1, row=1, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="User:").grid(column=0, row=2, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.user).grid(column=1, row=2, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="Password:").grid(column=0, row=3, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=30, textvariable=self.pw, show="*").grid(column=1, row=3, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.serverparamframe, text="Port:").grid(column=0, row=4, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.serverparamframe, width=4, textvariable=self.port).grid(column=1, row=4, padx=3, pady=3, sticky=tk.W)

		# File databases
		ttk.Label(self.fileparamframe, text="Database file:").grid(column=0, row=0, padx=3, pady=3, sticky=tk.NW)
		ttk.Entry(self.fileparamframe, width=40, textvariable=self.db_file).grid(column=0, row=1, padx=3, pady=3, sticky=tk.NW)
		ttk.Button(self.fileparamframe, text="Browse...", command=self.set_sel_fn).grid(column=1, row=1)

		# File databases with user name and password
		ttk.Label(self.filepwparamframe, text="Database file:").grid(column=0, row=0, columnspan=2, padx=3, pady=3, sticky=tk.NW)
		ttk.Entry(self.filepwparamframe, width=40, textvariable=self.db_file).grid(column=0, row=1, columnspan=2, padx=3, pady=3, sticky=tk.NW)
		ttk.Button(self.filepwparamframe, text="Browse...", command=self.set_sel_fn).grid(column=2, row=1)
		ttk.Label(self.filepwparamframe, text="User:").grid(column=0, row=2, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.filepwparamframe, width=30, textvariable=self.user).grid(column=1, row=2, padx=3, pady=3, sticky=tk.W)
		ttk.Label(self.filepwparamframe, text="Password:").grid(column=0, row=3, padx=3, pady=3, sticky=tk.E)
		ttk.Entry(self.filepwparamframe, width=30, textvariable=self.pw, show="*").grid(column=1, row=3, padx=3, pady=3, sticky=tk.W)

		# Put serverparamframe, fileparamframe, and filepwparamframe in the same place in paramframe
		self.fileparamframe.grid(row=0, column=0, sticky=tk.NW)
		self.fileparamframe.grid_remove()
		self.filepwparamframe.grid(row=0, column=0, sticky=tk.NW)
		self.filepwparamframe.grid_remove()
		self.serverparamframe.grid(row=0, column=0, sticky=tk.NW)
		self.db_type_var.set(u"PostgreSQL")

		def w1_next(*args):
			self.dlg.bind("<Alt-p>", load_script)
			self.dlg.bind("<Alt-s>", save_script)
			self.dlg.bind("<Alt-e>", edit_sql)
			wiz2_frame.lift()
			# The following conditional fails
			#if w1next_btn["state"] == tk.NORMAL:
			#	wiz2_frame.lift()
			self.dlg.bind("<Alt-n>")
			self.dlg.bind("<Alt-n>", w2_next)
			self.dlg.bind("<Alt-b>")
			self.dlg.bind("<Alt-b>", w2_back)
			self.table_entry.focus()

		# Populate w1btnframe
		w1help_btn = ttk.Button(w1btnframe, text="Help", command=self.do_help, underline=0)
		w1help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w1next_btn = ttk.Button(w1btnframe, text="Next", command=w1_next, underline=0)
		w1next_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w1cancel_btn = ttk.Button(w1btnframe, text="Cancel", command=self.do_cancel, underline=0)
		w1cancel_btn.grid(row=0, column=2, sticky=tk.E, padx=(3,6))
		w1next_btn["state"] = tk.DISABLED
		self.dlg.bind("<Alt-n>", w1_next)
		self.dlg.bind("<Alt-c>", self.do_cancel)
		self.dlg.bind("<Escape>", self.do_cancel)
		

		# Wizard page 2
		# Database table and optional query
		def check_w2enable(*args):
			enable_if(w2next_btn, self.table_var.get() != '')
		def w2_back(*args):
			self.dlg.unbind("<Alt-p>")
			self.dlg.unbind("<Alt-s>")
			self.dlg.unbind("<Alt-e>")
			wiz1_frame.lift()
			self.dlg.bind("<Alt-n>", w1_next)
			self.dlg.bind("<Alt-b>")
			self.db_choices.focus()
		def w2_next(*args):
			if self.table_var.get() != '':
				self.mapui.loading_dlg.display("Querying database")
				sql = "select * from %s;" % self.table_var.get()
				#conn = None
				db = None
				# Open database, get table data and column headers, populate wiz2 comboboxes
				dbms = self.db_type_var.get()
				server = self.server.get() if self.server.get() != '' else None
				dbname = self.db.get() if self.db.get() != '' else None
				user = self.user.get() if self.user.get() != '' else None
				need_pw = self.pw.get() != ''
				pw = self.pw.get() if self.pw.get() != '' else None
				port = int(self.port.get()) if self.port.get() != '' else None
				db_file = self.db_file.get()
				db = open_dbms(dbms, server, dbname, user, need_pw, pw, port, db_file)

				if db is not None:
					script_error = False
					script_text = self.script_text.get("1.0", "end")
					if len(script_text) > 0:
						try:
							runscripts(db, script_text)
						except ErrInfo as e:
							warning(e.eval_err(), kwargs={'parent':self.dlg})
							script_error = True
						except:
							script_error = True
							raise
					if not script_error:
						try:
							self.headers, self.rows = db.select_data(sql)
						except ErrInfo as e:
							warning(e.eval_err(), kwargs={'parent': self.dlg})
						except:
							warning("Cannot select data from table %s." % self.table_var.get(), kwargs={'parent':self.dlg})
						else:
							self.header_list = list(self.headers)
							# Set list box values
							self.id_sel["values"] = self.header_list
							self.lat_sel["values"] = self.header_list
							self.lon_sel["values"] = self.header_list
							self.sym_sel["values"] = self.header_list
							self.col_sel["values"] = self.header_list
							self.dlg.unbind("<Alt-p>")
							self.dlg.unbind("<Alt-s>")
							self.dlg.unbind("Alt-e>")
							wiz3_frame.lift()
							self.dlg.bind("<Alt-n>")
							self.dlg.bind("<Alt-b>")
							self.dlg.bind("<Alt-b>", w3_back)
							self.dlg.bind('<Alt-o>', self.do_select)
					db.close()
				self.mapui.loading_dlg.hide()

		def load_script(*args):
			fn = tkfiledialog.askopenfilename(parent=self.dlg, title="SQL script file to open", filetypes=([('SQL script files', '.sql')]))
			if not (fn is None or fn == '' or fn == ()):
				path, filename = os.path.split(os.path.abspath(fn))
				self.scriptfilepath = path
				with open(fn, "r") as f:
					sql = f.read()
				self.script_text.insert("end", sql)
		def save_script(*args):
			outfile = tkfiledialog.asksaveasfilename(initialdir=self.scriptfilepath, parent=self.dlg, title="SQL script file to save", filetypes=[('SQL script files', '.sql')])
			if not (outfile is None or outfile == ''):
				sql = self.script_text.get("1.0", "end")
				with open(outfile, "w") as f:
					f.write(sql)
		def edit_sql(*args):
			td = tempfile.TemporaryDirectory()
			edit_fn = os.path.join(td.name, "mapfile_temp.sql")
			with open(edit_fn, "w") as f:
				f.write(self.script_text.get("1.0", "end"))
			returncode = subprocess.call([editor, edit_fn])
			if returncode == 0:
				with open(edit_fn, "r") as f:
					sql = f.read()
				self.script_text.delete("1.0", "end")
				self.script_text.insert("end", sql)
			else:
				warning("Failure attempting to edit the SQL with %s" % editor, kwargs={'parent':self.dlg})

		w2req_frame = ttk.LabelFrame(wiz2_frame, text="Required")
		w2req_frame.grid(row=0, column=0, sticky=tk.EW+tk.N, padx=(6,6), pady=(3,3))
		w2req_frame.columnconfigure(1, weight=1)
		#
		tbl_label = ttk.Label(w2req_frame, text="Table:")
		tbl_label.grid(row=0, column=0, sticky=tk.E, padx=(6,3), pady=(3,3))
		self.table_var = tk.StringVar(w2req_frame, '')
		self.table_var.trace('w', check_w2enable)
		self.table_entry = ttk.Entry(w2req_frame, width=30, textvariable=self.table_var)
		self.table_entry.grid(row=0, column=1, padx=(3,6), pady=3, sticky=tk.W)
		#
		w2opt_frame = ttk.LabelFrame(wiz2_frame, text="Optional")
		w2opt_frame.grid(row=1, column=0, sticky=tk.EW+tk.N, padx=(6,6), pady=(3,3))
		w2opt_frame.columnconfigure(1, weight=1)
		#
		self.script_text = tk.Text(w2opt_frame, width=40, height=4)
		self.script_text.grid(row=0, column=1, columnspan=2, rowspan=4, sticky=tk.NSEW, padx=(3,0), pady=(3,3))
		scr_label = ttk.Label(w2opt_frame, text="Script:")
		scr_label.grid(row=0, column=0, sticky=tk.NE, padx=(6,3), pady=(2,2))
		load_btn = ttk.Button(w2opt_frame, text="Open", command=load_script, underline=1)
		load_btn.grid(row=1, column=0, sticky=tk.E, padx=(6,3))
		save_btn = ttk.Button(w2opt_frame, text="Save", command=save_script, underline=0)
		save_btn.grid(row=2, column=0, sticky=tk.E, padx=(3,3))
		edit_btn = ttk.Button(w2opt_frame, text="Edit", command=edit_sql, underline=0)
		edit_btn.grid(row=3, column=0, sticky=tk.E, padx=(3,3), pady=(0,2))
		enable_if(edit_btn, editor is not None)
		sbar = tk.Scrollbar(w2opt_frame)
		sbar.grid(row=0, column=2, rowspan=4, sticky=tk.NS, padx=(0,3), pady=(3,3))
		sbar.config(command=self.script_text.yview)
		self.script_text.config(yscrollcommand = sbar.set)
		#
		w2btn_frame = tk.Frame(wiz2_frame, borderwidth=3, relief=tk.RIDGE)
		w2btn_frame.grid(row=2, column=0, sticky=tk.S+tk.EW, padx=(3,3), pady=(2,2))
		w2btn_frame.columnconfigure(0, weight=1)
		#
		w2help_btn = ttk.Button(w2btn_frame, text="Help", command=self.do_help, underline=0)
		w2help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		w2prev_btn = ttk.Button(w2btn_frame, text="Back", command=w2_back, underline=0)
		w2prev_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w2next_btn = ttk.Button(w2btn_frame, text="Next", command=w2_next, underline=0)
		w2next_btn.grid(row=0, column=2, sticky=tk.E, padx=3)
		w2cancel_btn = ttk.Button(w2btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w2cancel_btn.grid(row=0, column=3, sticky=tk.E, padx=(3,6))
		w2next_btn["state"] = tk.DISABLED


		# Wizard page 3
		# Column selectors
		def check_w3enable(*args):
			enable_if(w3ok_btn, self.lat_var.get() != '' and self.lon_var.get() != '')
		w3req_frame = ttk.LabelFrame(wiz3_frame, text="Required")
		w3req_frame.grid(row=0, column=0, sticky=tk.EW, padx=(6,6), pady=(3,3))
		w3req_frame.columnconfigure(0, weight=1)
		#
		lat_label = ttk.Label(w3req_frame, text="Latitude column:")
		lat_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lat_var = tk.StringVar(w3req_frame, '')
		self.lat_var.trace('w', check_w3enable)
		self.lat_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lat_var, values=self.header_list, width=24)
		self.lat_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		lon_label = ttk.Label(w3req_frame, text="Longitude column:")
		lon_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.lon_var = tk.StringVar(w3req_frame, '')
		self.lon_var.trace('w', check_w3enable)
		self.lon_sel = ttk.Combobox(w3req_frame, state="readonly", textvariable=self.lon_var, values=self.header_list, width=24)
		self.lon_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))

		w3opt_frame = ttk.LabelFrame(wiz3_frame, text="Optional")
		w3opt_frame.grid(row=1, column=0, sticky=tk.EW, padx=(6,6), pady=(6,3))
		w3opt_frame.columnconfigure(0, weight=1)
		#
		id_label = ttk.Label(w3opt_frame, text="Label column:")
		id_label.grid(row=0, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.id_var = tk.StringVar(w3opt_frame, '')
		self.id_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.id_var, values=self.header_list, width=24)
		self.id_sel.grid(row=0, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		crs_label = ttk.Label(w3opt_frame, text="CRS:")
		crs_label.grid(row=0, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.crs_var = tk.IntVar(w3opt_frame, 4326)
		self.crs_var.trace('w', check_w3enable)
		self.crs_sel = ttk.Entry(w3opt_frame, width=8, textvariable=self.crs_var)
		self.crs_sel.grid(row=0, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		sym_label = ttk.Label(w3opt_frame, text="Symbol column:")
		sym_label.grid(row=1, column=0, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.sym_var = tk.StringVar(w3opt_frame, '')
		self.sym_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.sym_var, values=self.header_list, width=24)
		self.sym_sel.grid(row=1, column=1, sticky=tk.W, padx=(3,30), pady=(3,3))
		#
		col_label = ttk.Label(w3opt_frame, text="Color column:")
		col_label.grid(row=1, column=2, sticky=tk.E, padx=(3,3), pady=(3,3))
		self.col_var = tk.StringVar(w3opt_frame, '')
		self.col_sel = ttk.Combobox(w3opt_frame, state="readonly", textvariable=self.col_var, values=self.header_list, width=24)
		self.col_sel.grid(row=1, column=3, sticky=tk.W, padx=(3,6), pady=(3,3))
		#
		desc_label = ttk.Label(w3opt_frame, text="Description:")
		desc_label.grid(row=2, column=0, sticky=tk.E, padx=(3,3), pady=(3,6))
		self.desc_var = tk.StringVar(w3opt_frame, '')
		desc_entry = ttk.Entry(w3opt_frame, width=60, textvariable=self.desc_var)
		desc_entry.grid(row=2, column=1, columnspan=3, sticky=tk.W, padx=(3,3), pady=(3,6))

		def w3_back(*args):
			self.dlg.bind("<Alt-p>", load_script)
			self.dlg.bind("<Alt-s>", save_script)
			self.dlg.bind("<Alt-e>", edit_sql)
			self.dlg.bind("<Alt-o>")
			wiz2_frame.lift()
			self.dlg.bind("<Alt-n>", w2_next)
			self.dlg.bind("<Alt-b>")
			self.dlg.bind("<Alt-b>", w2_back)
			self.table_entry.focus()

		w3btn_frame = tk.Frame(wiz3_frame, borderwidth=3, relief=tk.RIDGE)
		w3btn_frame.grid(row=2, column=0, sticky=tk.S+tk.EW, padx=(3,3), pady=(3,3))
		w3btn_frame.columnconfigure(0, weight=1)
		#
		w3help_btn = ttk.Button(w3btn_frame, text="Help", command=self.do_help, underline=0)
		w3help_btn.grid(row=0, column=0, sticky=tk.W, padx=(6,3))
		self.dlg.bind("<Alt-h>", self.do_help)
		w3prev_btn = ttk.Button(w3btn_frame, text="Back", command=w3_back, underline=0)
		w3prev_btn.grid(row=0, column=1, sticky=tk.E, padx=3)
		w3ok_btn = ttk.Button(w3btn_frame, text="OK", command=self.do_select, underline=0)
		w3ok_btn.grid(row=0, column=2, sticky=tk.E, padx=3)
		w3cancel_btn = ttk.Button(w3btn_frame, text="Cancel", command=self.do_cancel, underline=0)
		w3cancel_btn.grid(row=0, column=3, sticky=tk.E, padx=(3,6))
		w3ok_btn["state"] = tk.DISABLED

		wiz1_frame.lift()

		self.canceled = True
		# Limit resizing
		self.dlg.resizable(False, False)
		center_window(self.dlg)
	def set_sel_fn(self):
		fn = tkfiledialog.askopenfilename(parent=self.fileparamframe, title=self.title)
		if fn is not None and fn != '' and fn != ():
			self.db_file.set(fn)
			#self.clearstatus()
	def do_select(self, *args):
		if self.table_var.get() != '' and self.lat_var.get() != '' and self.lon_var.get() != '':
			self.canceled = False
			self.dlg.destroy()
	def do_cancel(self, *args):
		self.canceled = True
		self.dlg.destroy()
	def do_help(self, *args):
		webbrowser.open("https://mapdata.readthedocs.io/en/latest/importingdata.html#importing-from-a-file-based-database", new=2, autoraise=True)
	def get_data(self):
		self.dlg.grab_set()
		self.dlg.focus_force()
		self.db_choices.focus()
		self.dlg.wait_window(self.dlg)
		if self.canceled:
			return (None, None, None, None, None, None, None, None, None, None)
		else:
			return self.table_var.get(), self.id_var.get(), self.lat_var.get(), self.lon_var.get(), \
					self.crs_var.get(), self.sym_var.get(), self.col_var.get(), self.desc_var.get(), \
					self.headers, self.rows





def read_all_config(argobj):
	# 'argobj' has attributes for command-line settings.  This is used in two ways:
	#  1. If there is a non-null 'file' setting, it is used to look for a configuration
	#     file in that location.
	#  2. Values that are read from configuration files that correspond to command-line
	#     arguments will replace those values in 'argobj', but only after all configuration
	#     files are read, and the corresponding values in the 'argobj' are null.
	class ClArgs(object):
		pass
	cl_arg_defaults = ClArgs()

	global config_files
	config_files = []
	if os.name == 'posix':
		sys_config_file = os.path.join("/etc", config_file_name)
	else:
		sys_config_file = os.path.join(os.path.expandvars(r'%APPDIR%'), config_file_name)
	if os.path.isfile(sys_config_file):
		config_files.append(sys_config_file)
	program_dir_config = os.path.join(os.path.abspath(sys.argv[0]), config_file_name)
	if os.path.isfile(program_dir_config) and not program_dir_config in config_files:
		config_files.append(program_dir_config)
	user_config_file = os.path.join(os.path.expanduser(r'~/.config'), config_file_name)
	if os.path.isfile(user_config_file) and not user_config_file in config_files:
		config_files.append(user_config_file)
	datafile = argobj.file
	if datafile is not None:
		data_config_file = os.path.join(os.path.abspath(datafile), config_file_name)
		if os.path.isfile(data_config_file) and not data_config_file in config_files:
			config_files.append(data_config_file)
	startdir_config_file = os.path.join(os.path.abspath(os.path.curdir), config_file_name)
	if os.path.isfile(startdir_config_file) and not startdir_config_file in config_files:
		config_files.append(startdir_config_file)
	files_read = []
	while len(config_files) > 0:
		cf = config_files.pop(0)
		if not cf in files_read and os.path.isfile(cf):
			files_read.append(cf)
			next_cf = read_config(cf, cl_arg_defaults)
			if next_cf is not None:
				if not (next_cf in files_read):
					config_files.insert(0, next_cf)
	
	# Update 'argobj'
	for att in ("db_type", "server", "port", "database", "user", "table", "script", \
			"lon", "lat", "id", "projection", "message", "file", "sheet", "symbol", \
			"color", "message", "imagefile", "imagewait"):
		if hasattr(cl_arg_defaults, att) and getattr(argobj, att) is None:
			setattr(argobj, att, getattr(cl_arg_defaults, att))


def read_config(configfile, arg_settings):
	# Reads a single configuration file, assigning values to global variables
	# (including attributes of map_settings).
	_BASEMAP_SECTION = "basemap_tile_servers"
	_APIKEYS_SECTION = "api_keys"
	_ATTRIBUTIONS_SECTION = "map_attributions"
	_COLORADJ_SECTION = "map_color_adj"
	_SYMBOL_SECTION = "symbols"
	_CONNECT_SECTION = "connect"
	_DEFAULTS_SECTION = "defaults"
	_MISC_SECTION = "misc"
	cp = ConfigParser()
	cp.optionxform = str
	cp.read(configfile)
	next_config_file = None

	#==== Tile servers
	if cp.has_section(_BASEMAP_SECTION):
		basemap_sources = cp.items(_BASEMAP_SECTION)
		for name, url in basemap_sources:
			if url is None:
				if name in bm_servers and len(bm_servers) > 1:
					del(bm_servers[name])
			else:
				bm_servers[name] = url

	#==== API keys
	if cp.has_section(_APIKEYS_SECTION):
		apikeys = cp.items(_APIKEYS_SECTION)
		for name, apikey in apikeys:
			if apikey is None:
				if name in api_keys and len(api_keys) > 1:
					del(api_keys[name])
			else:
				api_keys[name.capitalize()] = apikey

	#==== Map attributions
	if cp.has_section(_ATTRIBUTIONS_SECTION):
		atts = cp.items(_ATTRIBUTIONS_SECTION)
		for name, citation in atts:
			if citation is not None:
				map_attributions[name] = citation

	#==== Map color adjustments
	if cp.has_section(_COLORADJ_SECTION):
		adjustments = cp.items(_COLORADJ_SECTION)
		for name, coloradj in adjustments:
			if coloradj is not None:
				try:
					adjs = [float(x) for x in re.split(r"[ ,]+", coloradj)]
				except:
					warning(f"Invalid color adjustment values for {name}.", {})
				else:
					if len(adjs) != 3:
						warning(f"Wrong number of color adjustment values for {name}.", {})
					else:
						map_color_adj[name] = adjs

	#==== Symbols
	if cp.has_section(_SYMBOL_SECTION):
		symbols = cp.items(_SYMBOL_SECTION)
		for name, filename in symbols:
			if not os.path.isfile(filename):
				config_warning_messages.append(f"The symbol file {filename} does not exist.")
			else:
				import_symbol(name, filename)

	#==== Connection settings
	# All of these update the 'arg_settings' object, even if the keyword has no value.
	if cp.has_option(_CONNECT_SECTION, "db_type"):
		t = cp.get(_CONNECT_SECTION, "db_type").lower()
		if len(t) != 1 or t not in ('l', 'p', 'f', 'm', 'o', 's'):
			config_warning_messages.append(f"Invalid database type: {t}.")
		arg_settings.db_type = t
	if cp.has_option(_CONNECT_SECTION, "server"):
		arg_settings.server = cp.get(_CONNECT_SECTION, "server")
	if cp.has_option(_CONNECT_SECTION, "port"):
		try:
			arg_settings.port = cp.getint(self._CONNECT_SECTION, "port")
		except:
			config_warning_messages.append("Invalid port number.")
	# Accept 'db' for 'database', for compatibility with execsql
	if cp.has_option(_CONNECT_SECTION, "db"):
		arg_settings.database = cp.get(_CONNECT_SECTION, "db")
	if cp.has_option(_CONNECT_SECTION, "database"):
		arg_settings.database = cp.get(_CONNECT_SECTION, "database")
	# 'file' also appears in the defaults section
	if cp.has_option(_CONNECT_SECTION, "file"):
		arg_settings.file = cp.get(_CONNECT_SECTION, "file")
	# 'username' is accepted as well as 'user', for execsql compatibility
	if cp.has_option(_CONNECT_SECTION, "username"):
		arg_settings.user = cp.get(_CONNECT_SECTION, "username")
	if cp.has_option(_CONNECT_SECTION, "user"):
		arg_settings.user = cp.get(_CONNECT_SECTION, "user")
	if cp.has_option(_CONNECT_SECTION, "no_password"):
		try:
			self.no_passwd = cp.getboolean(_CONNECT_SECTION, "no_password")
		except:
			config_warning_messages.append("Invalid argument for 'no_password' in configuration file.")
	if cp.has_option(_CONNECT_SECTION, "table"):
		arg_settings.table = cp.get(_CONNECT_SECTION, "table")
	if cp.has_option(_CONNECT_SECTION, "script"):
		arg_settings.script = cp.get(_CONNECT_SECTION, "script")

	#==== Defaults
	#-- Equivalent to command-line arguments
	if cp.has_option(_DEFAULTS_SECTION, "x_column"):
		arg_settings.lon = cp.get(_DEFAULTS_SECTION, "x_column")
	if cp.has_option(_DEFAULTS_SECTION, "y_column"):
		arg_settings.lat = cp.get(_DEFAULTS_SECTION, "y_column")
	if cp.has_option(_DEFAULTS_SECTION, "id_column"):
		arg_settings.id = cp.get(_DEFAULTS_SECTION, "id_column")
	if cp.has_option(_DEFAULTS_SECTION, "symbol_column"):
		arg_settings.symbol = cp.get(_DEFAULTS_SECTION, "symbol_column")
	if cp.has_option(_DEFAULTS_SECTION, "color_column"):
		arg_settings.color = cp.get(_DEFAULTS_SECTION, "color_column")
	if cp.has_option(_DEFAULTS_SECTION, "projection"):
		proj = cp.get(_DEFAULTS_SECTION, "projection")
		try:
			proj = int(proj)
		except:
			config_fatal_messages.append(f"Invalid projection in configuration file: {proj}.")
		else:
			arg_settings.projection = proj
	if cp.has_option(_DEFAULTS_SECTION, "message"):
		arg_settings.message = cp.get(_DEFAULTS_SECTION, "message")
	if cp.has_option(_DEFAULTS_SECTION, "datafile"):
		arg_settings.file = cp.get(_DEFAULTS_SECTION, "datafile")
	if cp.has_option(_DEFAULTS_SECTION, "sheet"):
		arg_settings.sheet = cp.get(_DEFAULTS_SECTION, "sheet")
	if cp.has_option(_DEFAULTS_SECTION, "imagefile"):
		arg_settings.imagefile = cp.get(_DEFAULTS_SECTION, "imagefile")
	if cp.has_option(_DEFAULTS_SECTION, "imagewait"):
		try:
			arg_settings.imagewait = cp.getint(_DEFAULTS_SECTION, "imagewait")
		except:
			config_warning_messages.append("Invalid 'imagewait' setting in configuration file.")
	#-- Other defaults
	if cp.has_option(_DEFAULTS_SECTION, "multiselect"):
		global multiselect
		err = False
		try:
			multi = cp.getboolean(_DEFAULTS_SECTION, "multiselect")
		except:
			err = True
			config_warning_messages.append("Invalid argument to the 'multiselect' configuration option.")
		if not err:
			multiselect = "1" if multi else "0"
	if cp.has_option(_DEFAULTS_SECTION, "basemap"):
		global initial_basemap
		bm = cp.get(_DEFAULTS_SECTION, "basemap")
		if bm is None or bm not in bm_servers:
			config_warning_messages.append(f"Invalid argument to the 'basemap' configuration option: {bm}.")
		else:
			initial_basemap = bm
	if cp.has_option(_DEFAULTS_SECTION, "location_marker"):
		loc_mkr = cp.get(_DEFAULTS_SECTION, "location_marker")
		if loc_mkr is not None:
			map_settings.location_marker = loc_mkr
	if cp.has_option(_DEFAULTS_SECTION, "location_color"):
		loc_color = cp.get(_DEFAULTS_SECTION, "location_color")
		if loc_color is not None:
			if loc_color not in color_names:
				config_warning_messages.append(f"Invalid argument to the 'location_color' configuration option: {loc_color}.")
			else:
				map_settings.location_color = loc_color
	if cp.has_option(_DEFAULTS_SECTION, "use_data_marker"):
		global use_data_marker
		loc_mkr = cp.getboolean(_DEFAULTS_SECTION, "use_data_marker")
		if loc_mkr is not None:
			map_settings.use_data_marker = loc_mkr
	if cp.has_option(_DEFAULTS_SECTION, "use_data_color"):
		global use_data_color
		loc_clr = cp.getboolean(_DEFAULTS_SECTION, "use_data_color")
		if loc_clr is not None:
			map_settings.use_data_color = loc_clr
	if cp.has_option(_DEFAULTS_SECTION, "select_symbol"):
		global select_symbol
		default_symbol = cp.get(_DEFAULTS_SECTION, "select_symbol")
		if default_symbol is not None:
			if default_symbol not in icon_xbm:
				config_warning_messages.append(f"Unrecognized symbol name for the 'select_symbol' configuration option: {default_symbol}.")
			else:
				map_settings.select_symbol = default_symbol
	if cp.has_option(_DEFAULTS_SECTION, "select_color"):
		global select_color
		sel_color = cp.get(_DEFAULTS_SECTION, "select_color")
		if sel_color is not None:
			if sel_color not in color_names:
				config_warning_messages.append("Invalid argument to the 'multiselect' configuration option.")
			else:
				map_settings.select_color = sel_color
	if cp.has_option(_DEFAULTS_SECTION, "label_color"):
		global label_color
		lbl_color = cp.get(_DEFAULTS_SECTION, "label_color")
		if lbl_color is not None:
			if lbl_color not in color_names:
				config_warning_messages.append(f"Invalid argument to the 'label_color' configuration option: {lbl_color}.")
			else:
				map_settings.label_color = lbl_color
	if cp.has_option(_DEFAULTS_SECTION, "label_font"):
		global label_font
		lbl_font = cp.get(_DEFAULTS_SECTION, "label_font")
		if lbl_font is not None:
			if lbl_font not in list(tk.font.families()):
				config_warning_messages.append(f"Invalid argument to the 'label_font' configuration option: {lbl_font}.")
			else:
				map_settings.label_font = lbl_font
	if cp.has_option(_DEFAULTS_SECTION, "label_size"):
		global label_size
		err = False
		try:
			lbl_size = cp.getint(_DEFAULTS_SECTION, "label_size")
		except:
			err = True
			config_warning_messages.append(f"Invalid argument to the 'label_size' configuration option: {lbl_size}.")
		if not err:
			if lbl_size is not None and lbl_size > 6:
				map_settings.label_size = lbl_size
	if cp.has_option(_DEFAULTS_SECTION, "label_bold"):
		global label_bold
		err = False
		try:
			lbl_bold = cp.getboolean(_DEFAULTS_SECTION, "label_bold")
		except:
			err = True
			config_warning_messages.append(f"Invalid argument to the 'label_bold' configuration option: {lbl_bold}.")
		if not err:
			if lbl_bold is not None:
				map_settings.label_bold = lbl_bold
	if cp.has_option(_DEFAULTS_SECTION, "label_position"):
		global label_position
		lbl_position = cp.get(_DEFAULTS_SECTION, "label_position")
		if lbl_position is not None:
			lbl_position = lbl_position.lower()
			if lbl_position not in ("above", "below"):
				config_warning_messages.append(f"Invalid argument to the 'label_position' configuration option: {lbl_position}.")
			else:
				map_settings.label_position = lbl_position
	if cp.has_option(_DEFAULTS_SECTION, "show_regression_stats"):
		global show_regression_stats
		err = False
		try:
			srs = cp.getboolean(_DEFAULTS_SECTION, "show_regression_stats")
		except:
			err = True
			config_warning_messages.append(f"Invalid argument to the 'show_regression_stats' configuration option: {srs}.")
		if not err:
			show_regression_stats = srs
	if cp.has_option(_DEFAULTS_SECTION, "wrapwidth"):
		global wrapwidth
		err = False
		try:
			wr = cp.getint(_DEFAULTS_SECTION, "wrapwidth")
		except:
			err = True
			config_warning_messages.append(f"Invalid argument to the 'wrapwidth' configuration option: {wr}.")
		if not err:
			wrapwidth = wr
	if cp.has_option(_MISC_SECTION, "temp_dbfile"):
		global temp_dbfile
		err = False
		try:
			dbfile = cp.getboolean(_MISC_SECTION, "temp_dbfile")
		except:
			err = True
			config_warning_messages.append(f"Invalid argument to the 'temp_dbfile' configuration option: {dbfile}.")
		if not err:
			temp_dbfile = dbfile
	if cp.has_option(_MISC_SECTION, "editor"):
		global editor
		err = False
		try:
			ed = cp.get(_MISC_SECTION, "editor")
		except:
			err = True
			config_warning_messages.append(f"Invalid argument to the 'editor' configuration option: {ed}.")
		if not err:
			editor = ed
	if cp.has_option(_MISC_SECTION, "read_config"):
		err = False
		try:
			config_file = cp.get(_MISC_SECTION, "read_config")
		except:
			err = True
			config_warning_messages.append(f"Invalid argument to the 'config_file' configuration option: {config_file}.")
		if not err:
			if os.path.isdir(config_file):
				config_file = os.path.join(config_file, "mapdata.conf")
			if not os.path.isfile(config_file):
				config_warning_messages.append(f"The file specified with the 'config_file' configuration option ({config_file}) does not exist.")
			else:
				next_config_file = config_file
	return next_config_file
			


def import_symbol(symbol_name, filename):
	with open(filename, mode='r') as f:
		symbol_def = f.read()
	icon_xbm[symbol_name] = symbol_def



def clparser():
	# Every argument must be assigned a default value.
	desc_msg = "Display an interactive map with points read from a data file or database. Version %s, %s" % (version, vdate)
	parser = argparse.ArgumentParser(description=desc_msg)
	parser.add_argument('-a', '--table', dest='table', default=None,
			help="The name of a database table to be imported.")
	parser.add_argument('-c', '--color', default=None, dest='color',
			help="The name of the column in the data file or table containing color names.")
	parser.add_argument('-d', '--database', dest='database', default=None,
			help="The name of a client-server database from which to import dat.")
	parser.add_argument('-e', '--server', default=None,
			help="The name of the database server.")
	parser.add_argument('-f', '--file', default=None,
			help="The name of a CSV file, spreadsheet file, or database file containing data with latitude and longitude coordinates.")
	parser.add_argument('-g', '--image', dest='imagefile', default=None,
			help="The name of an image file to which the map will be exported--no UI will be created.")
	parser.add_argument('-i', '--identifier', dest='id',
			help="The name of the column in the data file containing location identifiers or labels.")
	parser.add_argument("-k", "--db_type", type=str, choices=['p', 's', 'l', 'm', 'k', 'o', 'f'], dest="db_type", default=None,
			help="Database type: 'p'-PostgreSQL; 's'-SQL Server; 'l'-SQLite, 'm'-MySQL/MariaDB, 'k'-DuckDB, 'o'-Oracle, 'f'-Firebird.")
	parser.add_argument('-m', '--message', dest='message',
			help='A message to display above the map.')
	parser.add_argument("-n", "--no_passwd", action="store_true", dest="no_passwd", default=False,
			help="Do not prompt for the password when the user is specified (default is to prompt).")
	parser.add_argument("-o", "--port", action="store", type=int, dest="port", default=None,
			help="Database server port.")
	parser.add_argument('-p', '--projection', default=None, type=int, dest="projection",
			help="The coordinate reference system (CRS) if the data are projected")
	parser.add_argument('-r', '--script', default=None, dest='script',
			help="The name of a SQL script file to run before importing a database table.")
	parser.add_argument('-s', '--symbol', default=None, dest='symbol',
			help="The name of the column in the data file or table containing symbol names.")
	parser.add_argument('-t', '--sheet', default=None,
			help="The name of the worksheet to import when the data source is a spreadsheet.")
	parser.add_argument("-u", "--user", action="store", type=str, dest="user", default=None,
			help="Database user name.")
	parser.add_argument('-w', '--imagewait', default=12, type=int,
			help="The time in seconds to wait before exporting the map to an image file.")
	parser.add_argument('-x', '--lon', dest='lon',
			help="The name of the column in the data file containing longitude values (default: x_coord).")
	parser.add_argument('-y', '--lat', dest='lat',
			help="The name of the column in the data file containg latitude values (default: y_coord).")
	return parser



def main():
	args = clparser().parse_args()
	read_all_config(args)
	make_data_db()
	app = MapUI(args)
	app.win.mainloop()


main()


