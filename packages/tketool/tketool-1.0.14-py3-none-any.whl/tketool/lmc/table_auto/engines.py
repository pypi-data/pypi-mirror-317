import abc
import pandas as pd
import os.path
import openpyxl
import re, math, json
from tketool.lmc.prompts.prompt_controller import get_prompt, get_prompt_by_path
from tketool.utils.progressbar import process_status_bar
from tketool.utils.MultiTask import do_multitask
from tketool.lmc.lmc_linked import *
from tketool.lmc.lmc_linked_flow import lmc_linked_flow_model
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class excel_pointer:
    def __init__(self, sheet_name, col_name, col_excel_name, row_index):
        self.sheet_name = sheet_name
        self.col_name = col_name
        self.col_excel_name = col_excel_name
        self.row_index = row_index


# Excel代理基类,定义了Excel处理的基本接口
# 使用abc.ABCMeta作为元类来定义抽象基类
class excel_agent(metaclass=abc.ABCMeta):
    @property
    @abc.abstractmethod
    def match_str(self):
        pass

    @property
    @abc.abstractmethod
    def params_list(self):
        return

    @property
    def agent_des(self):
        return ""

    @abc.abstractmethod
    def call(self, llm, row_dict, cur_col_name, sheet_obj, params, content, logs_list):
        pass

    def init_task(self, llm, sheet_obj, params, content):
        pass

    def is_nan(self, value):
        # Check for None
        if value is None:
            return True

        # Check for empty string
        if value == '':
            return True

        # Check for NaN
        if isinstance(value, float) and math.isnan(value):
            return True

        return False

    def get_datas_by(self, sheet_obj, index_key, cur_col=None):
        def split_string(s):
            # 使用正则表达式匹配字符和数字
            letters = re.findall(r'[A-Za-z]+', s)
            digits = re.findall(r'\d+', s)

            # 确保返回两个元素
            if letters:
                letters_part = letters[0]
            else:
                letters_part = ''

            if digits:
                digits_part = digits[0]
            else:
                digits_part = ''

            return [letters_part, digits_part]

        sheet_name = cur_col.sheet_name if cur_col is not None else ""

        sheet_split_list = index_key.split("::")
        if len(sheet_split_list) > 1:
            sheet_name = sheet_split_list[0]
            r_c_key = sheet_split_list[1]
        else:
            r_c_key = sheet_split_list[0]

        rc = split_string(r_c_key)

        if rc[0] == '':
            row_index = int(rc[1])
            row_data = {c: v
                        for v, c in
                        zip(sheet_obj[sheet_name].rows[row_index], sheet_obj[sheet_name].excel_column_names)}
            return row_data
        elif rc[1] == '':
            col_index = sheet_obj[sheet_name].excel_column_names.index(rc[0])
            col_list = [s[col_index] for s in sheet_obj[sheet_name].rows]
            return col_list
        else:
            col_index = sheet_obj[sheet_name].excel_column_names.index(rc[0])
            row_index = int(rc[1])
            return sheet_obj[sheet_name].rows[row_index][col_index]


class double_shape(excel_agent):

    @property
    def params_list(self):
        return []

    @property
    def match_str(self):
        return "##"

    def call(self, llm, row_dict, cur_name, sheet_obj, params, content, logs_list):
        llm_invoker = lmc_linked_model(llm).set_prompt_template(content)
        result = llm_invoker(**row_dict)
        if len(result.results) > 0:
            return str(result.result)
        else:
            return ""


class prompt_file_shape(excel_agent):

    def __init__(self):
        self.invoker_dict = {}

    @property
    def params_list(self):
        return [("prompt_file_path", "提示词文件路径"),
                ("prompt_key_mapping", "提示词的mapping key替换, A=wordA#B=wordB"),
                ("prompt_model_output", "返回的结果结构选择器(可选)"), ]

    @property
    def match_str(self):
        return "#promptfile"

    def init_task(self, llm, sheet_obj, params, content):
        invoker_key = params["prompt_file_path"]
        if invoker_key not in self.invoker_dict:
            path = params["prompt_file_path"]
            parse_field = get_prompt_by_path(path)
            self.invoker_dict[invoker_key] = lmc_linked_flow_model(parse_field, retry_time=2)

    def call(self, llm, row_dict, cur_name, sheet_obj, params, content, logs_list):
        invoker_key = params["prompt_file_path"]
        mapping_dict = {}
        for k, v in row_dict.items():
            if isinstance(v, str):
                mapping_dict[k] = v
            else:
                if not math.isnan(v):
                    mapping_dict[k] = v
        if len(params) > 1:
            mapping_change_dict = {}
            mapping_change = params["prompt_key_mapping"].split("#")
            for spp in mapping_change:
                s = spp.split("=")
                if len(s) == 2:
                    mapping_change_dict[s[0]] = s[1]

            for k, v in mapping_change_dict.items():
                if k in mapping_dict:
                    mapping_dict[v] = mapping_dict[k]
                if v in mapping_dict:
                    mapping_dict[k] = mapping_dict[v]

        llmresult = self.invoker_dict[invoker_key](llm, **mapping_dict)
        if not llmresult.passed:
            logs_list.append("invoke llm error")
        if len(params) > 2:
            output_str = params["prompt_model_output"]
            result_str = eval("llmresult.result." + output_str)
            return result_str
        else:
            return llmresult.result.json()


class sheet_data:
    def __init__(self, sheet_obj, all_agents):
        self.sheet_obj = sheet_obj
        self.all_agents = all_agents
        self.column_names = []
        self.column_name_index = {}
        self.excel_column_names = []
        self.column_agent = []
        self.rows = []
        self.log_col = None

        for col_name in sheet_obj.columns.tolist():
            self.column_names.append(col_name)
            excel_col_name = openpyxl.utils.get_column_letter(sheet_obj.columns.get_loc(col_name) + 1)
            self.excel_column_names.append(excel_col_name)
            self.column_agent.append(self.parse_colstr(col_name))
            if col_name == '#log':
                self.log_col = excel_col_name

        self.mapping_agent = []  # all_agents[cmd] if cmd in all_agents else None for cmd, par, con in self.column_agent]

        for cmd, par, con in self.column_agent:
            if cmd in all_agents:
                self.mapping_agent.append(all_agents[cmd])
            else:
                self.mapping_agent.append(None)

        for idx, row in sheet_obj.iterrows():
            cur_row = []
            for c_idx in self.column_names:
                cur_row.append(row[c_idx])
            self.rows.append(cur_row)

        self.column_name_index = {colname: idx for idx, colname in enumerate(self.column_names)}

    def parse_colstr(self, col_str):
        if not isinstance(col_str, str):
            return (None, None, None)
        if col_str.startswith("##"):
            return ("##", [], col_str[2:])
        else:
            pattern = r'(#\w+)(?:\(([^)]+)\))?(?::\s*(.*))?'
            match = re.match(pattern, col_str)
            if match:
                command = match.group(1).strip()  # 提取指令
                params = match.group(2)  # 提取参数，可能为None
                content = match.group(3)

                if params:
                    params_list = []
                    for p in params.split(','):
                        p_s = p.strip()
                        if (p_s.startswith('"') and p_s.endswith('"')) or (p_s.startswith("'") and p_s.endswith("'")):
                            params_list.append(p_s[1:-1])
                        else:
                            params_list.append(p_s)
                    # param.strip() for param in params.split(',')]
                else:
                    params_list = []

                if command in self.all_agents:
                    zip_params = {k[0]: v for k, v in zip(self.all_agents[command].params_list, params_list)}
                    return (command, zip_params, content)

        return (None, None, None)

    def set(self, row_index, col_name, value):
        target_col_name = ""
        if col_name in self.column_name_index:
            target_col_name = col_name
        else:
            target_col_name = self.column_names[self.excel_column_names.index(col_name)]
        self.sheet_obj.at[row_index, target_col_name] = value
        col_index = self.column_name_index[target_col_name]
        self.rows[row_index][col_index] = value
        pass


class excel_work_process:
    def __init__(self):
        self.max_sheet_count = 0
        self.done_sheet_count = 0


class excel_engine:
    def __init__(self, llm, *args, thread=1):
        self.all_agents = {}
        self.llm = llm
        self.thread_count = thread
        for arg in args:
            if isinstance(arg, excel_agent):
                self.all_agents[arg.match_str] = arg

    def parse_sheet(self, sheet_obj_dict, sheet_key, pass_row_count, pb, progress_callback=None):
        def do_task(idx):
            if progress_callback:
                progress_callback(idx)
            
            sheet_obj = sheet_obj_dict[sheet_key]
            logs_list = []
            for op, row_tile, coln, coln2 in zip(sheet_obj.mapping_agent, sheet_obj.column_agent,
                                                 sheet_obj.column_names,
                                                 sheet_obj.excel_column_names):
                if op is None:
                    continue

                row_dict = {k: v for k, v in
                            zip(sheet_obj.excel_column_names, sheet_obj.rows[idx])}

                if not isinstance(row_dict[coln2], str) and math.isnan(row_dict[coln2]):
                    params = row_tile[1]
                    content = row_tile[2]
                    pointer = excel_pointer(sheet_key, coln, coln2, idx)
                    call_result = op.call(self.llm, row_dict, pointer, sheet_obj_dict, params, content, logs_list)
                    sheet_obj_dict[sheet_key].set(idx, coln, call_result)
            if sheet_obj.log_col:
                sheet_obj_dict[sheet_key].set(idx, sheet_obj.log_col, "\n".join(logs_list))
            pass

        rows_index = list(range(pass_row_count, len(sheet_obj_dict[sheet_key].rows)))
        for row, c_row in pb.iter_bar(
                do_multitask(rows_index, do_task, self.thread_count, self.thread_count * 2),
                key="row", max=len(rows_index)):
            pass

    def call_file(self, excel_file_path: str, start_row_index=0, progress_callback=None):
        xls = pd.ExcelFile(excel_file_path)
        pb = process_status_bar()

        all_sheets = {}
        column_widths = {}  # 用于存储每个工作表的列宽信息
        
        total_sheets = len(xls.sheet_names)
        current_sheet = 0

        for sheet in xls.sheet_names:
            sheet_content = sheet_data(xls.parse(sheet), all_agents=self.all_agents)
            all_sheets[sheet] = sheet_content
            
            # 计算总行数用于进度计算
            total_rows = len(sheet_content.rows) - start_row_index
            
            # 加载现有的Excel文件，以获取列宽信息
            wb = load_workbook(excel_file_path)
            ws = wb[sheet]

            column_widths[sheet] = {}

            # 获取工作表的列维度
            for col in ws.column_dimensions:
                # 在openpyxl中，列宽默认是None，如果未设置
                # 所以我们需要检查是否已经设置���列宽
                if ws.column_dimensions[col].width is not None:
                    column_widths[sheet][col] = ws.column_dimensions[col].width
                else:
                    # 如果列宽未设置，可以设置一个默认值或根据内容计算宽度
                    column_widths[sheet][col] = 10  # 假设默认列宽为10

            if progress_callback:
                progress_callback({
                    'stage': 'init',
                    'current_sheet': sheet,
                    'sheet_progress': current_sheet / total_sheets,
                    'total_sheets': total_sheets,
                    'current_sheet_index': current_sheet,
                    'total_rows': total_rows
                })
            current_sheet += 1

        for k, sheet in all_sheets.items():
            for cmd, par, con in sheet.column_agent:
                if cmd in self.all_agents:
                    self.all_agents[cmd].init_task(self.llm, all_sheets, par, con)

        current_sheet = 0
        for s_name, op_sheet in pb.iter_bar(all_sheets.items(), key="sheet"):
            try:
                total_rows = len(op_sheet.rows) - start_row_index
                current_row = 0
                
                def row_progress_callback(row_index):
                    nonlocal current_row
                    current_row = row_index - start_row_index
                    if progress_callback:
                        progress_callback({
                            'stage': 'processing',
                            'current_sheet': s_name,
                            'sheet_progress': current_sheet / total_sheets,
                            'row_progress': current_row / total_rows,
                            'current_row': current_row,
                            'total_rows': total_rows,
                            'total_sheets': total_sheets,
                            'current_sheet_index': current_sheet
                        })
                
                self.parse_sheet(all_sheets, s_name, start_row_index, pb=pb, progress_callback=row_progress_callback)
            finally:
                current_sheet += 1

        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            for s_name, op_sheet in pb.iter_bar(all_sheets.items(), key="sheet"):
                all_sheets[s_name].sheet_obj.to_excel(writer, sheet_name=s_name, index=False)

                # 设置列宽
                worksheet = writer.sheets[s_name]
                for col, width in column_widths[s_name].items():
                    worksheet.column_dimensions[col].width = width

                # 设置自动换行
                for col in worksheet.columns:
                    for cell in col:
                        cell.alignment = Alignment(wrapText=True)


from tketool.lmc.llms.openai import OpenAI_Complete_Model
from typing import Optional, List, Dict, Mapping, Any


class local_model(OpenAI_Complete_Model):
    def __init__(self, **kwargs: Any):
        super().__init__("qwen2-72b-gptq", (0.01, 0.03), apitoken="EMPTY",
                         base_url="http://10.193.3.200:8000/v1", **kwargs)


ee = excel_engine(local_model(), double_shape(), prompt_file_shape())
ee.call_file("/Users/kejiang/Downloads/test.xlsx", start_row_index=1)
