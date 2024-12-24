'''
Created: Oct 22, 2023
Author: @kumykov

License tagging and grouping workflow

Copyright © 2024 Synopsys, Inc. All Rights Reserved.
This Synopsys «product» and all associated documentation are proprietary to Synopsys, Inc. and may only be used
pursuant to the terms and conditions of a written license agreement with Synopsys, Inc. All other use, reproduction,
modification, or distribution of the Synopsys «product» or the associated documentation is strictly prohibited

'''

import argparse
import logging
import sys
from pprint import pprint

import openpyxl
from blackduck import Client


class License:
    def __init__(self, values) -> None:
        self.name = values[0] if len(values) > 0 else None
        self.url = values[1] if len(values) > 1 else None
        self.family = values[2] if len(values) > 2 else None
        self.family_url = values[3] if len(values) > 3 else None
        self.spdx_id = values[4] if len(values) > 4 else None
        self.tags = values[5:] if len(values) > 5 else []

    #def __str__(self):
    #    return f"{self.name} {self.family}"

    def __repr__(self):
        return f"[{self.name}, {self.url}, {self.family}, {self.family_url}, {self.spdx_id}, {self.tags}"

    def __eq__(self, __value: object) -> bool:
        return (self.name == __value.name
                and self.url == __value.url
                and self.family == __value.family
                and self.family_url == __value.family_url)

class LicenseStorage:

    def __init__(self, **kwargs) -> None:
        self.licenses = []
        self.tags = []
        keys = kwargs.keys()
        if 'filepath' in keys:
            self.filepath = kwargs['filepath']
        try:
            self.open(self.filepath)
        except(FileNotFoundError):
            self.initialize()
        self.read_data()

    def initialize(self):
        self.workbook = openpyxl.Workbook()
        license_sheet = self.workbook.active
        license_sheet.title = 'Licenses'
        tag_sheet = self.workbook.create_sheet(title='LicenseTags', index=2)
        self.save()

    def open(self,filepath):
        self.workbook = openpyxl.load_workbook(filepath)

    def read_data(self):
        ws = self.workbook['Licenses']
        self.licenses = [License(l) for l in ws.values]
        ws = self.workbook['LicenseTags']
        self.tags = [License(l) for l in ws.values]

    def update_license(self,license, save = True):
        ws = self.workbook['Licenses']
        found = False
        for row in ws.iter_rows():
            if row[0].value == license.name:
                found = True
                ws.cell(row[0].row, column = 2).value = license.url
                ws.cell(row[0].row, column = 3).value = license.family
                ws.cell(row[0].row, column = 4).value = license.family_url
                ws.cell(row[0].row, column = 5).value = license.spdx_id
                break
        if not found:
            data  = [license.name,
                     license.url,
                     license.family,
                     license.family_url,
                     license.spdx_id]
            ws.append(data)
        if save:
            self.save()

    def update_tag(self,tag: "License"):
        ws = self.workbook['LicenseTags']
        for row in ws.iter_rows():
            if row[0].value == tag.name:
                ws.cell(row[0].row, column = 2).value = tag.url
                ws.cell(row[0].row, column = 3).value = tag.family
                ws.cell(row[0].row, column = 4).value = tag.family_url
        self.save()

    def get_tags(self):
        return self.tags

    def get_tag(self, tagname):
        return [x for x in self.tags if tagname == x.name][0]

    def get_license_group(self, tag):
        return [x for x in self.licenses if tag.name in x.tags]

    def get_expanded_parameters(self, matches):
        parameters = dict()
        parameters['data'] = []
        parameters['values'] = []
        for tagname in matches:
            tag = self.get_tag(tagname)
            parameters['data'].append({"licenseName": tagname, "licenseUrl": tag.url})
            parameters['values'].append(tag.url)
            license_group = self.get_license_group(tag)
            for license in license_group:
                parameters['data'].append({"licenseName": license.name, "licenseUrl": license.url})
                parameters['values'].append(license.url)
        return parameters

    def save(self):
        self.workbook.save(self.filepath)

logging.basicConfig(format='%(asctime)s:%(levelname)s:%(message)s', stream=sys.stderr, level=logging.DEBUG)
logging.getLogger("requests").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("blackduck").setLevel(logging.DEBUG)

tag_family = 'TagPlaceholder'
tag_family_number = None

def update_licenses():
    json = bd.get_json('/api/licenses?offset=0&limit=0')
    pprint(f"Found {json['totalCount']} licenses.")
    licenses = bd.get_items('/api/licenses')
    for license in licenses:
        license_name = license['name']
        license_url = license['_meta']['href']
        license_family = license['licenseFamily']
        license_spdx_id = license.get('spdxId',None)
        print (f"{license_name} {license_family['name']} ::: {license_spdx_id}")

def validate_placeholder_license_family():
    params = {"filter": ["licenseFamilySource:custom"]}
    custom_families = bd.get_items('/api/license-families', params = params)
    dataset = [x for x in custom_families if x['name'] == tag_family]
    if len(dataset) == 0:
        data = {"name": tag_family,
                "description": "Placeholder family for license tagging"}
        result = bd.session.post('/api/license-families', json=data)
        if result.ok:
            return validate_placeholder_license_family()
    elif len(dataset) == 1:
        return dataset[0]['_meta']['href'].split("/")[5]
    return None

def validate_placeholder_license(license, family_number):
    params = {"filter": [f"licenseFamily:{family_number}"]}
    placeholder_licenses = bd.get_items('/api/licenses', params = params)
    dataset = [x for x in placeholder_licenses if x['name'] == license.name]
    if len(dataset) == 0:
        data = {"name" : license.name,
                "licenseFamily" : f"/api/license-families/{family_number}",
                "text" : "This is a placeholder license",
                "licenseStatus" : "APPROVED",
                }
        result = bd.session.post('/api/licenses', json=data)
        if result.ok:
            license_data = result.json()
            values = [license_data['name'],
                      license_data['_meta']['href'],
                      license_data['licenseFamily']['name'],
                      license_data['licenseFamily']['href']]
            return License(values)
    elif len(dataset) == 1:
        license_data = dataset[0]
        values = [license_data['name'],
                    license_data['_meta']['href'],
                    license_data['licenseFamily']['name'],
                    license_data['licenseFamily']['href']]
        return License(values)
    return None

def validate_local_placeholder_licenses(tag_family_number, license_store):
    local_placeholder_licenses = license_store.get_tags()
    for placeholder_license in local_placeholder_licenses:
        bd_license = validate_placeholder_license(placeholder_license, tag_family_number)
        if (bd_license == placeholder_license):
            logging.info(f"License {placeholder_license.name} matches server definitions")
        else:
            logging.info(f"License {placeholder_license.name} does not match server definitions, updating")
            license_store.update_tag(bd_license)

def validate_remote_placeholder_licenses(family_number, license_store):
    local_placeholder_licenses = license_store.get_tags()
    params = {"filter": [f"licenseFamily:{family_number}"]}
    remote_placeholder_licenses = bd.get_items('/api/licenses', params = params)
    local_placeholder_names = [l.name for l in local_placeholder_licenses]
    dataset = [l for l in remote_placeholder_licenses if l['name'] not in local_placeholder_names]
    logging.info(f"Found {len(dataset)} placeholder licenses not defiled in the local store.")
    if len(dataset) > 0:
        for l in dataset:
            logging.info(f"Removing {l['name']} from server")
            response = bd.session.delete(l['_meta']['href'])
            logging.info(f"Response {response}")

def validate_placeholder_licenses(family_number, license_store):
    validate_local_placeholder_licenses(family_number, license_store)
    validate_remote_placeholder_licenses(family_number, license_store)

def retrieve_licenses():
    json = bd.get_json('/api/licenses?offset=0&limit=10')
    pprint(f"Found {json['totalCount']} licenses.")
    licenses = bd.get_items('/api/licenses')
    dataset = []
    for license in licenses:
        if license['licenseFamily']['name'] == tag_family:
            continue
        values = [license['name'],
                license['_meta']['href'],
                license['licenseFamily']['name'],
                license['licenseFamily']['href'],
                license.get('spdxId',None)]
        dataset.append(License(values))
    return dataset

def tag_matches(expression, tags):
    if expression['name'] != 'SINGLE_LICENSE':
        return []
    license_names = [item['licenseName'] for item in expression['parameters']['data']]
    tag_names = [x.name for x in tags]
    return list(set(license_names).intersection(set(tag_names)))

def update_policy(policy, matches, license_store):
    logging.info(f"Submitting policy {policy['name']} update to the server")
    policy_url = policy['_meta']['href']
    response = bd.session.put(policy_url, json=policy)
    logging.info(f"{response}")

def process_policies(license_store):
    params = {"filter": "policyRuleEnabled:true"}
    policies = bd.get_items('/api/policy-rules', params = params)
    tags = license_store.get_tags()
    for policy in policies:
        affected = False
        for expression in policy['expression']['expressions']:
            matches = tag_matches(expression, tags)
            if len(matches) > 0:
                affected = True
                parameters = license_store.get_expanded_parameters(matches)
                expression.update({"parameters": parameters})
                expression.update({"operation": 'IN'})
                logging.info(f"Expression for policy {policy['name']} updated with {matches}")

        if affected:
            update_policy(policy, matches, license_store)

def execute_process(blackduck, license_store_file, license_check):
    global bd
    bd = blackduck

    tag_family_number = validate_placeholder_license_family()
    if tag_family_number:
        logging.info(f"Placeholder family {tag_family} with id {tag_family_number} present")
    license_store = LicenseStorage(filepath=license_store_file)
    validate_placeholder_licenses(tag_family_number, license_store)
    if license_check:
        license_data = retrieve_licenses()
        for license in license_data:
            logging.info(f"Updating information for {license.name}")
            license_store.update_license(license, save = False)
        license_store.save()

    process_policies(license_store)


def parse_command_args():

    parser = argparse.ArgumentParser("license_tagging.py")
    parser.add_argument("-u", "--base-url",     required=True, help="Hub server URL e.g. https://your.blackduck.url")
    parser.add_argument("-t", "--token-file",   required=True, help="File containing access token")
    parser.add_argument("-nv", "--no-verify",   action='store_false', help="Disable TLS certificate verification")
    parser.add_argument("-lsf", "--license-store-file", required=False, default='LicenseStorage.xlsx', help="Local license information storage")
    parser.add_argument("-nlc", "--no-license-check", action='store_false', help="Skip scanning trough all licenses")
    return parser.parse_args()

def main():
    args = parse_command_args()
    with open(args.token_file) as tf:
        access_token = tf.readline().strip()

    bd = Client(base_url=args.base_url, token=access_token, verify=args.no_verify, timeout=460.0, retries=4)

    execute_process(bd, args.license_store_file, args.no_license_check)

if __name__ == "__main__":
    sys.exit(main())
