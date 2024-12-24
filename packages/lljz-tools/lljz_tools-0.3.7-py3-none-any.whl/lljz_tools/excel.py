# coding=utf-8

"""
@fileName       :   excel.py
@data           :   2024/2/22
@author         :   jiangmenggui@hosonsoft.com
"""
import csv
import os.path
from typing import Iterator

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:

    def __init__(self, excel_filename):
        self._file = excel_filename
        self._is_csv = False
        if isinstance(self._file, str):
            if not os.path.exists(self._file):
                raise FileNotFoundError(f"{self._file} not found!")
            self._is_csv = self._file.endswith('.csv')
        self._csv = self._excel = None  # type: ignore
        if self._is_csv:
            self._csv = open(self._file)
        else:
            self._excel = load_workbook(self._file, read_only=True, data_only=True)

    def _get_excel(self):
        if self._is_csv:
            return open(self._file)
        else:
            return load_workbook(self._file, read_only=True, data_only=True)

    def _get_sheet(self, sheet: int | None | str) -> Worksheet:
        if not self._excel:
            raise ValueError('excel not open')
        if isinstance(sheet, int):
            return self._excel.worksheets[sheet]
        elif isinstance(sheet, str):
            return self._excel[sheet]
        return self._excel.active  # type: ignore

    def _csv_read(self, /, *, has_title: bool = True, min_row=None, max_row=None,
                  min_col=None, max_col=None):

        def read(min_row, max_row, min_col, max_col):
            reader = csv.reader(self._csv)  # type: ignore
            for i in range(min_row - 1):
                next(reader)
            for row, i in zip(reader, range(max_row)):
                yield row[min_col - 1:max_col]

        reader = read(min_row or 1, max_row or 10000000, min_col or 1, max_col or 10000000)

        if has_title:
            title = next(reader)
            yield from (dict(zip(title, r)) for r in reader)
        else:
            yield from reader

    def read(self, __sheet: int | None | str = None, /, *, has_title: bool = True, min_row=None, max_row=None,
             min_col=None, max_col=None):
        """
        读取Excel的Sheet页中的数据

        如果没有指定索引，则范围从A1开始。如果工作表中没有单元格，则返回一个空元组。

        :param __sheet: sheet页的顺序或名字或默认
        :type sheet: int | None | str

        :param has_title: 是否包含标题，如果包含标题则每行数据都会以字典的形式返回，否则以元组的形式返回
        :type has_title: bool

        :param min_col: 最小列索引(从1开始)
        :type min_col: int

        :param min_row: 最小行索引(从1开始)
        :type min_row: int

        :param max_col: 最大列索引(从1开始)
        :type max_col: int

        :param max_row: 最大行索引(从1开始)
        :type max_row: int

        :rtype: generator
        """
        if self._is_csv:
            yield from self._csv_read(has_title=has_title, min_row=min_row, max_row=max_row,
                                      min_col=min_col, max_col=max_col)
            return
        sheet = self._get_sheet(__sheet)
        iter_rows = sheet.iter_rows(min_row, max_row, min_col, max_col, values_only=True)
        if has_title:
            title = next(iter_rows)
            yield from (dict(zip(title, r)) for r in iter_rows)
        else:
            yield from iter_rows

    @property
    def sheet_names(self):
        if self._is_csv:
            raise ValueError('csv no sheet names')
        if not self._excel:
            raise ValueError('excel not open')
        return self._excel.sheetnames

    def close(self):
        if self._csv:
            self._csv.close()
        if self._excel:
            self._excel.close()


class ExcelWriter:

    def __init__(self, excel_filename):
        self._file = excel_filename
        self._sheets: dict[str, Worksheet] = {}
        self._excel = self._get_excel()

    def _get_excel(self) -> Workbook:  # noqa
        return Workbook(write_only=True)

    def write(self, data: Iterator[list | dict], /, *, sheet_name=None) -> Worksheet:
        if not sheet_name:
            sheet_name = f'Sheet{len(self._sheets) + 1}'
        if sheet_name not in self._sheets:
            self._sheets[sheet_name] = self._excel.create_sheet(title=sheet_name)
        if not data:
            return self._sheets[sheet_name]

        data = iter(data)
        first = next(data)
        if isinstance(first, dict):
            self._sheets[sheet_name].append(tuple(first.keys()))  # write_title
            self._sheets[sheet_name].append(tuple(first.values()))
        else:
            self._sheets[sheet_name].append(first)
        for r in data:
            if isinstance(r, dict):
                self._sheets[sheet_name].append(tuple(r.values()))
            else:
                self._sheets[sheet_name].append(r)
        return self._sheets[sheet_name]

    def save(self):
        self._excel.save(self._file)

    def close(self):
        self._excel.close()


class Excel(ExcelReader, ExcelWriter):

    def __init__(self, excel_filename: str):
        super().__init__(excel_filename=excel_filename)
        if not self._excel:
            raise ValueError('excel not open')
        self._sheets = {name: self._excel[name] for name in self._excel.sheetnames}  

    def _get_excel(self):  # noqa
        return load_workbook(self._file, data_only=True)

    def write(self, data, /, *, sheet_name=None):
        if not sheet_name:
            sheet_name = self._excel.active.title  # type: ignore
        return super().write(data, sheet_name=sheet_name)

    def __getitem__(self, item) -> Worksheet:
        if not self._excel:
            raise ValueError('excel not open')
        return self._excel[item]  # type: ignore


if __name__ == '__main__':
    pass
