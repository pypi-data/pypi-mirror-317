import os
import openpyxl
from openpyxl import load_workbook
import pandas as pd

class Excel:

    def __init__(self):
        pass

    def read_from_excel(self, filename: str, sheet_name: str='Sheet1', header: int=0, index_col: bool=False, keep_default_na: bool=True, skip_rows=None) -> pd.DataFrame:
        return pd.read_excel(filename, sheet_name=sheet_name, header=header, index_col=None, keep_default_na=keep_default_na, skiprows=skip_rows, engine='openpyxl', parse_dates=[0])

    def append_to_excel(self, filename, data, sheet_name='Sheet1', startrow=None, truncate_sheet=False, header=False, **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        @param filename: File path or existing ExcelWriter
                        (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                        (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                        Per default (startrow=None) calculate the last row
                        in the existing DF and write to the next row...
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                            before writing DataFrame to Excel file
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None

        Usage examples:

        >>> append_df_to_excel('d:/temp/test.xlsx', df)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                            index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                            index=False, startrow=25)

        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """

        df = pd.DataFrame(data)

        # Excel file doesn't exist - saving and exiting
        if not os.path.isfile(filename):
            
            df.to_excel(
                filename,
                sheet_name=sheet_name, 
                startrow=startrow if startrow is not None else 0, 
                **to_excel_kwargs)
            return
        
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        if_sheet_exists = 'replace' if truncate_sheet is True else 'overlay'
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists=if_sheet_exists)

        # get the last row in the existing Excel sheet if it was not specified explicitly
        if startrow is None or truncate_sheet is True:
            startrow = 0
        elif startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy new sheet from template sheet
        if 'template' in writer.book.sheetnames and sheet_name not in writer.book.sheetnames:
            sheet_to_copy = writer.book.get_sheet_by_name('template')
            sheet_copied = writer.book.copy_worksheet(sheet_to_copy)
            sheet_copied.title = sheet_name

        # copy existing sheets
        # writer.sheets = { ws.title:ws for ws in writer.book.worksheets }

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, header=header, **to_excel_kwargs)

        # make columns filterable
        if sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.worksheets[idx].auto_filter.ref = writer.book.worksheets[idx].dimensions
        else:
            writer.book.active.auto_filter.ref = writer.book.active.dimensions

        # make columns autofit
        if sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            worksheet = writer.book.worksheets[idx]

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = adjusted_width


        # save the workbook
        writer.save()

    def write_to_excel(self, filename: str, data: any, sheet_name: str='Sheet1') -> None:
        # create excel writer object
        writer = pd.ExcelWriter(filename, datetime_format='hh:mm:ss')

        if type(data) is dict:     
            for sheet_name, dataframe in data.items():
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # write dataframe to excel
            data = pd.DataFrame(data)
            data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # save the excel
        writer.save()

    def get_sheets(self, filename: str) -> 'list[str]':
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        # try to open an existing workbook
        writer.book = load_workbook(filename)

        return writer.book.sheetnames

    def insert_image(self, filename: str, imagename: str, sheet_name: str, cell: str='A1') -> None:
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        sheet = None
        for i, sheetname in enumerate(writer.book.sheetnames):
            if sheet_name == sheetname:
                sheet = writer.book.worksheets[i]
                break
        if sheet is None:
            sheet = writer.book.create_sheet(sheet_name, len(writer.book.sheetnames))
        else:
            del writer.book[sheet_name]
            sheet = writer.book.create_sheet(sheet_name, len(writer.book.sheetnames))
        
        picture = openpyxl.drawing.image.Image(imagename)
        sheet.add_image(picture, cell)

        # save the excel
        writer.save()