import logging
from pathlib import Path
from re import Pattern
from typing import List, Any

import openpyxl

from sage_templater.plugin_manager import hookimpl
from sage_templater.schemas import PetitCashRecordSchema

logger = logging.getLogger(__name__)


def get_wb_and_sheets(file_path: Path, only_visible: bool = True) -> (openpyxl.Workbook, List[str]):
    """Get workbook and sheets from an Excel file."""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    if only_visible:
        return wb, [sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == "visible"]
    return wb, wb.sheetnames


def get_raw_rows(wb: openpyxl.Workbook, sheet_name: str, start_row: int, end_row: int) -> List[List[str]]:
    """Get raw rows from a sheet with the small box format."""
    sheet = wb[sheet_name]
    raw_rows = []
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        raw_row = []
        for cell in row:
            raw_row.append(str(cell.value))
        raw_rows.append(raw_row)
    return raw_rows


def is_empty_row(raw_row: List[str]) -> bool:
    """Check if a row is empty."""
    return all(cell is None or cell == "None" for cell in raw_row)


def check_regular_expression(regexp: Pattern, cell_value: Any) -> bool:
    """Check if a cell value matches a regular expression."""
    if cell_value is None or not isinstance(cell_value, str):
        return False
    match = regexp.match(cell_value)
    return match is not None


@hookimpl
def parse_file(file_path: str) -> List[PetitCashRecordSchema]:
    if not file_path.endswith(".xlsx"):
        return []
    return None
