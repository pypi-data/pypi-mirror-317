# Fecha
# No. Ck.
# A nombre de:
# RUC
# DV
# Monto
# ITBMS
# Total
# Concepto
# Proyecto
import re
from decimal import Decimal
from pathlib import Path
from typing import List

import openpyxl

from sage_templater.plugins.parsers.excel_parser import check_regular_expression, get_wb_and_sheets
from sage_templater.schemas import CheckRecordSchema


def get_start_and_end_row_numbers(wb: openpyxl.Workbook, sheet_name: str) -> tuple[int, int]:
    """Get start and end row numbers from a sheet with checks and transfer format."""
    date_regexp = re.compile(r"\s*([Ff][Ee][cC][hH][aA])\s*")
    amount_regexp = re.compile(r"\s*(Monto)\s*")
    sheet = wb[sheet_name]
    start_row = -1
    if sheet.max_row is None:
        return start_row, -1
    end_row = sheet.max_row
    i = 0
    for row in sheet.iter_rows():
        i += 1
        date_cell_value = row[0].value
        amount_value = row[5].value
        if date_cell_value is None or not isinstance(date_cell_value, str):
            continue
        date_match = check_regular_expression(date_regexp, date_cell_value)
        amount_match = check_regular_expression(amount_regexp, amount_value)
        if date_match and amount_match:
            start_row = i
            break
    return start_row, end_row


def is_check_template(excel_file: Path) -> bool:
    if excel_file.suffix != ".xlsx":
        return False
    wb, sheets = get_wb_and_sheets(excel_file)
    start, end = get_start_and_end_row_numbers(wb, sheets[0])
    return start != -1


def is_valid_row(raw_row: List[str]) -> bool:
    try:
        amount = convert_to_decimal(raw_row, 5)
        tax = convert_to_decimal(raw_row, 6)
        total = convert_to_decimal(raw_row, 7)
        CheckRecordSchema(
            date=raw_row[0],
            check_number=raw_row[1],
            name=raw_row[2],
            national_id=raw_row[3],
            verification_digit=raw_row[4],
            amount=amount,
            tax=tax,
            total=total,
            description=raw_row[8],
            project=raw_row[9],
        )
        return True
    except Exception:
        return False


def convert_to_decimal(raw_row, i):
    if raw_row[i] is None or len(raw_row[i]) == 0:
        amount = Decimal("0.0")
    else:
        amount = Decimal(raw_row[i])
    return amount
