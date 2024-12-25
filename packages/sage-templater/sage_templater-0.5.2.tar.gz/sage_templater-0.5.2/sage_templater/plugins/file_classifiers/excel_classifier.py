import warnings
from time import time
from typing import List

import click

from sage_templater.plugins.parsers.check_excel_parsers import is_check_template
from sage_templater.plugins.parsers.petit_cash_excel_parsers import is_petit_cash_template

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

import xlrd
from openpyxl import Workbook
from pathlib import Path


def convert_xls_to_xlsx(xls_file: Path, xlsx_file: Path):
    # Open the xls file
    workbook_xls = xlrd.open_workbook(xls_file)
    sheet_xls = workbook_xls.sheet_by_index(0)

    # Create a new xlsx file
    workbook_xlsx = Workbook()
    sheet_xlsx = workbook_xlsx.active

    # Copy data from xls to xlsx
    for row in range(sheet_xls.nrows):
        for col in range(sheet_xls.ncols):
            sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

    # Save the xlsx file
    workbook_xlsx.save(xlsx_file)


def copy_xls_to_xlsx(input_xls: Path, output_xlsx: Path):
    # Open the xls file using xlrd
    workbook_xls = xlrd.open_workbook(input_xls, formatting_info=False)

    # Create a new xlsx workbook
    workbook_xlsx = Workbook()

    # Remove the default empty sheet created by openpyxl
    if len(workbook_xlsx.sheetnames) > 0:
        default_sheet = workbook_xlsx.active
        workbook_xlsx.remove(default_sheet)

    # Iterate through all sheets in the xls file
    for sheet_idx in range(workbook_xls.nsheets):
        sheet_xls = workbook_xls.sheet_by_index(sheet_idx)
        sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)

        # Iterate through rows and columns to copy values
        for row_idx in range(sheet_xls.nrows):
            for col_idx in range(sheet_xls.ncols):
                cell_value = sheet_xls.cell_value(row_idx, col_idx)
                sheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

    # Save the xlsx workbook
    workbook_xlsx.save(output_xlsx)
    print(f"Successfully copied '{input_xls}' to '{output_xlsx}'")


def get_excel_files(folder: Path, pattern: str = "**/*.xls*") -> List[Path]:
    """Get Excel files from a folder in xls and xlsx format."""
    excel_suffixes = (".xls", ".xlsx")
    excel_files = [file for file in folder.glob(pattern) if file.suffix in excel_suffixes]

    return excel_files


def main(pattern: str = "menuda"):
    folder = Path.home() / "Downloads" / "sage"  # / "data_ls"
    excel_files = get_excel_files(folder, pattern="**/*.xls")
    for i, excel_file in enumerate(excel_files):
        try:
            start = time()
            is_petit_cash = is_petit_cash_template(excel_file)
            is_check = is_check_template(excel_file)
            elapsed = time() - start
            if is_petit_cash:
                file_type = "PETIT_CASH"
                click.secho(f"{i + 1} {excel_file.relative_to(folder)} {file_type} time: {elapsed:.2f}", fg="green")
            elif is_check:
                file_type = "CHECK"
                click.secho(f"{i + 1} {excel_file.relative_to(folder)} {file_type} time: {elapsed:.2f}", fg="blue")
            else:
                file_type = "UNKNOWN"
                click.secho(f"{i + 1} {excel_file.relative_to(folder)} {file_type} time: {elapsed:.2f}",
                            fg="yellow")

        except Exception as e:
            click.secho(f"{i + 1} {excel_file.relative_to(folder)} {e}", fg="red")


if __name__ == '__main__':
    main()
