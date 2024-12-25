import re
from decimal import Decimal
from pathlib import Path
from typing import List

import openpyxl

from sage_templater.exceptions import SageParseRawError
from sage_templater.plugins.parsers.excel_parser import get_wb_and_sheets, is_empty_row, logger
from sage_templater.schemas import PetitCashRecordSchema


def get_start_and_end_row_numbers(wb: openpyxl.Workbook, sheet_name: str) -> tuple[int, int]:
    """Get start and end row numbers from a sheet with the small box format."""
    regexp = re.compile(r"\s*([Cc][oOóÓ][Dd][Ii][Gg][Oo])\s*")
    sheet = wb[sheet_name]
    start_row = -1
    if sheet.max_row is None:
        return start_row, -1
    end_row = sheet.max_row
    i = 0
    for row in sheet.iter_rows():
        i += 1
        cell_value = row[0].value
        if cell_value is None or not isinstance(cell_value, str):
            continue
        match = regexp.match(cell_value)
        if match:
            start_row = i
            break
    return start_row, end_row


def is_petit_cash_template(excel_file: Path) -> bool:
    if excel_file.suffix != ".xlsx":
        return False
    wb, sheets = get_wb_and_sheets(excel_file)
    start, end = get_start_and_end_row_numbers(wb, sheets[0])
    return start != -1


def is_valid_row(raw_row: List[str]) -> bool:
    try:
        PetitCashRecordSchema(
            code=raw_row[0],
            national_id=raw_row[1],
            verification_digit=raw_row[2],
            name=raw_row[3],
            invoice=raw_row[4],
            date=raw_row[5],
            amount=Decimal(raw_row[6]),
            tax=raw_row[7],
            total=Decimal(raw_row[8]),
            description=raw_row[9],
            # source_file=str(source_file),
            # source_sheet=source_sheet,
        )
        if raw_row[9] == "None" or raw_row[9] == "" or raw_row[9] is None:
            return False
        return True
    except Exception:
        return False


def clean_raw_rows(raw_rows: List[List[str]]) -> List[List[str]]:
    """Clean raw rows from a sheet with the small box format."""
    cleaned_raw_rows = []
    for raw_row in raw_rows:
        if not is_empty_row(raw_row) and is_valid_row(raw_row):
            cleaned_raw_rows.append(raw_row)
    return cleaned_raw_rows


def parse_raw_rows(
    raw_rows: List[List[str]], source_file: Path, source_sheet: str, has_headers: bool = False
) -> List[PetitCashRecordSchema]:
    """Parse raw rows from a sheet with the small box format."""
    records = []
    for i, raw_row in enumerate(raw_rows, 1):
        try:
            if len(raw_row) < 10 or (i == 1 and has_headers):
                logger.debug("Skipping row %s. Row: %s", i, raw_row)
                continue
            if raw_row[6] is None or raw_row[6] == "None":
                logger.debug("Stopping row %s. Row: %s", i, raw_row)
                break
            logger.debug("Parsing row %s. Row: %s", i, raw_row)
            record = PetitCashRecordSchema(
                code=raw_row[0],
                national_id=raw_row[1],
                verification_digit=raw_row[2],
                name=raw_row[3],
                invoice=raw_row[4],
                date=raw_row[5],
                amount=raw_row[6],
                tax=raw_row[7],
                total=raw_row[8],
                description=raw_row[9],
                source_file=str(source_file),
                source_sheet=source_sheet,
            )
            records.append(record)
        except Exception as e:
            logger.error("Error parsing row %s from %s - %s. Row: %s", i, source_file, source_sheet, raw_row)
            error_message = (
                f"Error parsing row {i} from {source_file} - {source_sheet}."
                f" Error type: {e.__class__.__name__} Error: {e}"
            )
            raise SageParseRawError(error_message) from e
    return records
